import base64
import io
import re
import zipfile
from datetime import datetime
from pathlib import Path
from uuid import uuid4

from flask import send_file

from backend.config.settings import DATABASE_PATH, UPLOAD_FOLDER, get_db_connection
from backend.image.processor import decode_image_base64, guess_image_extension
from backend.repositories.material_repository import (
    bulk_update_aluminum_images,
    create_aluminum_record,
    create_change_request,
    delete_aluminum_record,
    get_aluminum_record,
    get_change_request,
    list_all_aluminum_codes,
    list_all_aluminum_image_records,
    list_all_aluminum_records_raw,
    list_all_change_requests_raw,
    list_aluminum_records,
    list_change_requests,
    list_table_columns,
    update_aluminum_record,
    update_change_request_status,
)
from backend.utils.converters import normalize_lookup_code
from backend.utils.constants import (
    API_SPEC_COLUMN,
    CHANGE_ACTIONS,
    CHANGE_ACTION_CREATE,
    CHANGE_ACTION_DELETE,
    CHANGE_ACTION_IMPORT_IMAGES,
    CHANGE_ACTION_BATCH_UPDATE_PRICES,
    CHANGE_ACTION_BATCH_UPDATE_DATA,
    CHANGE_ACTION_UPDATE,
    CHANGE_STATUSES,
    CHANGE_STATUS_APPROVED,
    CHANGE_STATUS_PENDING,
    CHANGE_STATUS_REJECTED,
    CHANGE_STATUS_WITHDRAWN,
    DB_CODE_COLUMN,
    DB_CODE_ATTRIBUTE_COLUMN,
    DB_IMAGE_BASE64_COLUMN,
    DB_IMAGE_COLUMN,
    DB_NAME_COLUMN,
    DB_NAME_KO_COLUMN,
    DB_PRICE_COLUMN,
    DB_SPEC_COLUMN,
    DB_TABLE_NAME,
    DB_UNIT_COLUMN,
    DB_WEIGHT_COLUMN,
)


EXPORT_PREVIEW_COLUMN = '图片预览'
EXPORT_IMAGE_STATUS_COLUMN = '图片状态'
EXPORT_MAIN_COLUMNS = (
    DB_CODE_COLUMN,
    DB_SPEC_COLUMN,
    DB_NAME_COLUMN,
    DB_NAME_KO_COLUMN,
    DB_UNIT_COLUMN,
    DB_PRICE_COLUMN,
    DB_CODE_ATTRIBUTE_COLUMN,
    DB_WEIGHT_COLUMN,
)
CHANGE_REQUEST_EXPORT_COLUMNS = (
    'id',
    'action',
    'target_code',
    'requester',
    'requester_role',
    'status',
    'submitted_at',
    'reviewed_at',
    'reviewed_by',
    'review_note',
    'payload_json',
    'snapshot_json',
)


def get_now_iso():
    return datetime.now().isoformat(timespec='seconds')


def ensure_admin_role(actor_role):
    if actor_role != 'admin':
        raise PermissionError('只有 admin 可以直接写入数据库或执行审核')


PENDING_CHANGE_UPLOAD_DIR = UPLOAD_FOLDER / 'pending_change_requests'


def ensure_pending_change_upload_dir():
    PENDING_CHANGE_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    return PENDING_CHANGE_UPLOAD_DIR


def create_pending_upload_path(original_filename, fallback_suffix=''):
    ensure_pending_change_upload_dir()
    source_name = extract_filename_only(original_filename) or f'upload{fallback_suffix}'
    source_path = Path(source_name)
    safe_name = sanitize_archive_stem(source_path.stem or source_name)
    suffix = source_path.suffix or fallback_suffix
    return PENDING_CHANGE_UPLOAD_DIR / f'{datetime.now().strftime("%Y%m%d_%H%M%S")}_{uuid4().hex}_{safe_name}{suffix}'


def save_request_file(file_storage, fallback_suffix=''):
    original_filename = extract_filename_only(getattr(file_storage, 'filename', '')) or f'upload{fallback_suffix}'
    target_path = create_pending_upload_path(original_filename, fallback_suffix=fallback_suffix)
    file_storage.save(target_path)
    return {
        'original_filename': original_filename,
        'stored_path': str(target_path),
    }


def save_image_request_bundle(files):
    file_list = [file for file in (files or []) if getattr(file, 'filename', '')]
    if not file_list:
        raise ValueError('请先选择要上传的图片或压缩包')

    if len(file_list) == 1:
        filename = str(getattr(file_list[0], 'filename', '') or '').lower()
        if filename.endswith(('.xlsx', '.xls', '.zip')):
            saved = save_request_file(file_list[0], fallback_suffix=Path(filename).suffix or '.bin')
            saved['submission_mode'] = 'excel' if filename.endswith(('.xlsx', '.xls')) else 'archive'
            saved['file_count'] = 1
            return saved

    target_path = create_pending_upload_path('image_request_bundle.zip', fallback_suffix='.zip')
    with zipfile.ZipFile(target_path, mode='w', compression=zipfile.ZIP_STORED) as archive:
        for file_storage in file_list:
            filename = extract_filename_only(getattr(file_storage, 'filename', ''))
            if not filename:
                continue
            file_bytes = file_storage.read()
            archive.writestr(filename, file_bytes)
            try:
                file_storage.stream.seek(0)
            except Exception:
                pass

    return {
        'original_filename': target_path.name,
        'stored_path': str(target_path),
        'submission_mode': 'archive',
        'file_count': len(file_list),
    }


def build_file_change_request_payload(saved_file, extra=None):
    payload = dict(extra or {})
    payload.update({
        'stored_path': str(saved_file.get('stored_path') or '').strip(),
        'original_filename': str(saved_file.get('original_filename') or '').strip(),
    })
    if saved_file.get('submission_mode'):
        payload['submission_mode'] = saved_file['submission_mode']
    if saved_file.get('file_count') is not None:
        payload['file_count'] = saved_file['file_count']
    return payload


def open_saved_file_storage(stored_path, filename=''):
    from werkzeug.datastructures import FileStorage

    path = Path(str(stored_path or '').strip())
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f'待审核文件不存在: {path}')
    stream = path.open('rb')
    return FileStorage(stream=stream, filename=filename or path.name)


def normalize_image_payload(data):
    normalized = dict(data or {})
    image_base64 = str(normalized.get(DB_IMAGE_BASE64_COLUMN, '') or '').strip()
    image_value = str(normalized.get(DB_IMAGE_COLUMN, '') or '').strip()

    if not image_base64 and image_value.startswith('data:image'):
        image_base64 = image_value
        normalized[DB_IMAGE_COLUMN] = ''

    if image_base64:
        _, _, image_status = decode_image_base64(image_base64)
        if image_status != 'ready':
            raise ValueError('图片 base64 数据无效，请重新上传图片')
        normalized[DB_IMAGE_BASE64_COLUMN] = image_base64

    return normalized


INVALID_ARCHIVE_NAME_PATTERN = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
SUPPORTED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'}


def sanitize_archive_stem(value):
    text = str(value or '').strip().rstrip('.')
    if not text:
        return 'unnamed'
    sanitized = INVALID_ARCHIVE_NAME_PATTERN.sub('_', text)
    return sanitized or 'unnamed'


def build_archive_filename(code, image_ext, used_names):
    safe_stem = sanitize_archive_stem(code)
    normalized_ext = str(image_ext or 'png').lower().replace('jpeg', 'jpg')
    candidate = f'{safe_stem}.{normalized_ext}'
    suffix = 2

    while candidate.lower() in used_names:
        candidate = f'{safe_stem}_{suffix}.{normalized_ext}'
        suffix += 1

    used_names.add(candidate.lower())
    return candidate


def extract_filename_only(filename):
    normalized = str(filename or '').replace('\\', '/').strip()
    if not normalized:
        return ''
    return normalized.rsplit('/', 1)[-1].strip()


def extract_code_from_filename(filename):
    basename = extract_filename_only(filename)
    if not basename:
        return ''

    parts = basename.rsplit('.', 1)
    if len(parts) == 2 and parts[0]:
        return parts[0].strip()
    return basename.strip()


def normalize_upload_extension(filename, image_bytes):
    basename = extract_filename_only(filename)
    suffix = basename.rsplit('.', 1)[-1].lower() if '.' in basename else ''
    guessed = guess_image_extension(image_bytes, '').lower()
    if guessed:
        return guessed.replace('jpeg', 'jpg')
    if suffix in SUPPORTED_IMAGE_EXTENSIONS:
        return suffix.replace('jpeg', 'jpg')
    return ''


def normalize_image_bytes_for_storage(image_bytes, filename=''):
    image_ext = normalize_upload_extension(filename, image_bytes)
    if image_ext:
        return image_bytes, image_ext

    from PIL import Image as PILImage

    source = io.BytesIO(image_bytes)
    with PILImage.open(source) as img:
        if img.mode in ('RGBA', 'LA', 'P'):
            rgba_img = img.convert('RGBA')
            background = PILImage.new('RGB', rgba_img.size, (255, 255, 255))
            background.paste(rgba_img, mask=rgba_img.split()[-1])
            img = background
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        else:
            img = img.copy()

        output = io.BytesIO()
        img.save(output, format='PNG')
        return output.getvalue(), 'png'


def encode_image_bytes_to_data_url(image_bytes, filename=''):
    if not image_bytes:
        raise ValueError('图片文件为空')

    normalized_bytes, image_ext = normalize_image_bytes_for_storage(image_bytes, filename)
    encoded = base64.b64encode(normalized_bytes).decode('ascii')
    return f'data:image/{image_ext};base64,{encoded}'


def safe_excel_value(value):
    if value is None:
        return ''
    if isinstance(value, (int, float)):
        return value
    return str(value)


def build_excel_thumbnail(image_bytes, max_size=(96, 72)):
    from PIL import Image as PILImage

    source = io.BytesIO(image_bytes)
    with PILImage.open(source) as img:
        if img.mode in ('RGBA', 'LA', 'P'):
            rgba_img = img.convert('RGBA')
            background = PILImage.new('RGB', rgba_img.size, (255, 255, 255))
            background.paste(rgba_img, mask=rgba_img.split()[-1])
            img = background
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        else:
            img = img.copy()

        img.thumbnail(max_size, PILImage.Resampling.LANCZOS)
        output = io.BytesIO()
        img.save(output, format='PNG')
        output.seek(0)
        return output, img.width, img.height


def style_export_worksheet(ws, header_row=1):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_fill = PatternFill(fill_type='solid', start_color='DDEBF7', end_color='DDEBF7')
    header_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='D9E2F3')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    ws.freeze_panes = f'A{header_row + 1}'
    ws.auto_filter.ref = ws.dimensions


def autosize_columns(ws, width_overrides=None, max_width=40):
    from openpyxl.utils import get_column_letter

    width_overrides = width_overrides or {}
    for idx, column_cells in enumerate(ws.columns, 1):
        letter = get_column_letter(idx)
        if letter in width_overrides:
            ws.column_dimensions[letter].width = width_overrides[letter]
            continue

        max_length = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            if len(text) > max_length:
                max_length = len(text)
        ws.column_dimensions[letter].width = min(max(max_length + 2, 10), max_width)


def create_aluminum_export_workbook(actor_role):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    aluminum_rows = list_all_aluminum_records_raw()
    change_request_rows = list_all_change_requests_raw()
    aluminum_columns = [
        column for column in list_table_columns(DB_TABLE_NAME)
        if column not in {DB_IMAGE_COLUMN, DB_IMAGE_BASE64_COLUMN}
    ]

    workbook = Workbook()
    main_ws = workbook.active
    main_ws.title = DB_TABLE_NAME

    main_headers = list(aluminum_columns)
    main_ws.append(main_headers)

    for record in aluminum_rows:
        row_values = [safe_excel_value(record.get(column)) for column in aluminum_columns]
        main_ws.append(row_values)

    style_export_worksheet(main_ws)
    autosize_columns(main_ws, max_width=48)

    requests_ws = workbook.create_sheet(title='change_requests')
    requests_ws.append(list(CHANGE_REQUEST_EXPORT_COLUMNS))
    for row in change_request_rows:
        requests_ws.append([safe_excel_value(row.get(column)) for column in CHANGE_REQUEST_EXPORT_COLUMNS])
    style_export_worksheet(requests_ws)
    autosize_columns(requests_ws, max_width=48)

    meta_ws = workbook.create_sheet(title='meta')
    meta_ws.append(['key', 'value'])
    meta_rows = [
        ('exported_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('exported_by_role', safe_excel_value(actor_role or '')),
        ('database_path', safe_excel_value(DATABASE_PATH)),
        ('main_table', DB_TABLE_NAME),
        ('main_record_count', len(aluminum_rows)),
        ('main_export_column_count', len(aluminum_columns)),
        ('excluded_columns', f'{DB_IMAGE_COLUMN}, {DB_IMAGE_BASE64_COLUMN}'),
        ('change_request_count', len(change_request_rows)),
    ]
    for key, value in meta_rows:
        meta_ws.append([key, value])
    meta_ws['A1'].font = Font(bold=True)
    meta_ws['B1'].font = Font(bold=True)
    meta_ws.column_dimensions['A'].width = 24
    meta_ws.column_dimensions['B'].width = 42

    return workbook


def iter_uploaded_image_entries(files):
    for storage in files or []:
        filename = extract_filename_only(getattr(storage, 'filename', ''))
        if not filename:
            continue

        if filename.lower().endswith('.zip'):
            try:
                with zipfile.ZipFile(storage.stream) as archive:
                    for info in archive.infolist():
                        if info.is_dir():
                            continue
                        entry_name = extract_filename_only(info.filename)
                        if not entry_name:
                            continue
                        yield entry_name, archive.read(info)
            except zipfile.BadZipFile as exc:
                raise ValueError(f'压缩包 {filename} 不是有效的 ZIP 文件') from exc
            finally:
                try:
                    storage.stream.seek(0)
                except Exception:
                    pass
            continue

        image_bytes = storage.read()
        try:
            storage.stream.seek(0)
        except Exception:
            pass
        yield filename, image_bytes


def normalize_request_payload(data):
    if data is None:
        raise ValueError('请求体不能为空')
    if not isinstance(data, dict):
        raise ValueError('请求体格式不正确')

    normalized = dict(data)
    if '\n(mm)/(米)' in ''.join(normalized.keys()) and API_SPEC_COLUMN not in normalized:
        normalized[API_SPEC_COLUMN] = normalized.pop('规格说明\n(mm)/(米)')
    return normalized


def validate_aluminum_payload(data, *, require_code=True):
    normalized = normalize_request_payload(data)
    required_fields = [API_SPEC_COLUMN, DB_NAME_COLUMN, DB_UNIT_COLUMN]
    if require_code:
        required_fields.insert(0, DB_CODE_COLUMN)

    for field in required_fields:
        if not str(normalized.get(field, '') or '').strip():
            raise ValueError(f'{field} 不能为空')

    return normalize_image_payload(normalized)


def extract_search_filters(args):
    return {
        'keyword': str(args.get('keyword', '') or args.get('search', '') or '').strip(),
        'code': str(args.get('code', '')).strip(),
        'name': str(args.get('name', '')).strip(),
        'name_ko': str(args.get('name_ko', '')).strip(),
        'spec': str(args.get('spec', '')).strip(),
    }


def get_aluminum_list(args):
    try:
        page = max(int(args.get('page', 1)), 1)
        page_size = max(int(args.get('page_size', 10)), 1)
        filters = extract_search_filters(args)
        result = list_aluminum_records(page, page_size, search=filters.get('keyword', ''), filters=filters)
        return {
            'success': True,
            **result,
            'filters': filters,
            'message': '查询成功',
        }, 200
    except Exception as exc:
        return {'success': False, 'message': f'查询失败: {exc}'}, 500


def get_aluminum_by_id(record_id):
    try:
        row = get_aluminum_record(record_id)
        if not row:
            return {'success': False, 'message': '记录不存在'}, 404

        return {
            'success': True,
            'data': row,
            'message': '查询成功',
        }, 200
    except Exception as exc:
        return {'success': False, 'message': f'查询失败: {exc}'}, 500


def create_aluminum(data, actor_role):
    try:
        ensure_admin_role(actor_role)
    except PermissionError as exc:
        return {'success': False, 'message': str(exc)}, 403

    try:
        normalized_data = validate_aluminum_payload(data, require_code=True)
        create_aluminum_record(normalized_data)
        return {'success': True, 'message': '新增成功'}, 200
    except ValueError as exc:
        return {'success': False, 'message': str(exc)}, 400
    except Exception as exc:
        return {'success': False, 'message': f'新增失败: {exc}'}, 500


def update_aluminum(record_id, data, actor_role):
    try:
        ensure_admin_role(actor_role)
    except PermissionError as exc:
        return {'success': False, 'message': str(exc)}, 403

    try:
        normalized_data = validate_aluminum_payload(data, require_code=False)
        payload_code = str(normalized_data.get(DB_CODE_COLUMN, '') or '').strip()
        if payload_code and payload_code != record_id:
            return {'success': False, 'message': '暂不支持修改工程编码'}, 400

        normalized_data.pop(DB_CODE_COLUMN, None)
        update_aluminum_record(record_id, normalized_data)
        return {'success': True, 'message': '更新成功'}, 200
    except ValueError as exc:
        return {'success': False, 'message': str(exc)}, 400
    except LookupError as exc:
        return {'success': False, 'message': str(exc)}, 404
    except Exception as exc:
        return {'success': False, 'message': f'更新失败: {exc}'}, 500


def delete_aluminum(record_id, actor_role):
    try:
        ensure_admin_role(actor_role)
    except PermissionError as exc:
        return {'success': False, 'message': str(exc)}, 403

    try:
        delete_aluminum_record(record_id)
        return {'success': True, 'message': '删除成功'}, 200
    except LookupError as exc:
        return {'success': False, 'message': str(exc)}, 404
    except Exception as exc:
        return {'success': False, 'message': f'删除失败: {exc}'}, 500


def export_aluminum_images(actor_role):
    try:
        ensure_admin_role(actor_role)
    except PermissionError as exc:
        return {'success': False, 'message': str(exc)}, 403

    try:
        from backend.config.settings import DATABASE_PATH
        from backend.repositories.material_repository import get_db_connection, row_to_dict

        connection = get_db_connection()
        cursor = connection.cursor()
        cursor.execute(
            f'''
            SELECT
                "{DB_CODE_COLUMN}" AS "{DB_CODE_COLUMN}",
                "{DB_IMAGE_BASE64_COLUMN}" AS "{DB_IMAGE_BASE64_COLUMN}"
            FROM {DB_TABLE_NAME}
            WHERE COALESCE(TRIM("{DB_IMAGE_BASE64_COLUMN}"), '') != ''
            '''
        )
        rows = [row_to_dict(row) for row in cursor.fetchall()]
        connection.close()
        archive_buffer = io.BytesIO()
        used_names = set()
        exported_count = 0

        with zipfile.ZipFile(archive_buffer, mode='w', compression=zipfile.ZIP_STORED) as archive:
            for row in rows:
                code = str(row.get(DB_CODE_COLUMN) or '').strip()
                if not code:
                    continue

                image_value = row.get(DB_IMAGE_BASE64_COLUMN)
                image_bytes, image_ext, image_status = decode_image_base64(image_value)
                if image_status != 'ready':
                    continue

                archive_name = build_archive_filename(code, image_ext, used_names)
                archive.writestr(archive_name, image_bytes)
                exported_count += 1

        if exported_count <= 0:
            return {'success': False, 'message': '数据库中没有可导出的图片'}, 400

        archive_buffer.seek(0)
        download_name = f'aluminum_images_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
        return send_file(
            archive_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=download_name,
        )
    except Exception as exc:
        return {'success': False, 'message': f'批量下载图片失败: {exc}'}, 500


def create_batch_change_request(action, actor_user, actor_role, payload):
    request_id = create_change_request(
        action=action,
        target_code='',
        requester=str(actor_user or '').strip() or 'unknown',
        requester_role=str(actor_role or '').strip(),
        submitted_at=get_now_iso(),
        payload=payload,
        snapshot={},
    )
    return {
        'success': True,
        'request_id': request_id,
        'status': CHANGE_STATUS_PENDING,
        'message': '已提交给 admin 审核',
    }, 200


def can_bypass_image_review(actor_role, actor_user=''):
    if actor_role == 'admin':
        return True
    username = str(actor_user or '').strip().lower()
    return username == 'adjust'


def import_aluminum_images(files, actor_role, actor_user=''):

    file_list = [file for file in (files or []) if getattr(file, 'filename', '')]
    if not file_list:
        return {'success': False, 'message': '请先选择要上传的图片或 ZIP 压缩包'}, 400

    if not can_bypass_image_review(actor_role, actor_user):
        try:
            saved_file = save_image_request_bundle(file_list)
            return create_batch_change_request(
                CHANGE_ACTION_IMPORT_IMAGES,
                actor_user,
                actor_role,
                build_file_change_request_payload(saved_file),
            )
        except ValueError as exc:
            return {'success': False, 'message': str(exc)}, 400
        except Exception as exc:
            return {'success': False, 'message': f'提交图片导入审核失败: {exc}'}, 500

    try:
        code_lookup = {}
        for code in list_all_aluminum_codes():
            normalized_code = normalize_lookup_code(code)
            if normalized_code and normalized_code not in code_lookup:
                code_lookup[normalized_code] = code

        seen_updates = {}
        missing_codes = []
        skipped_files = []
        duplicate_codes = []
        total_entries = 0

        for entry_name, image_bytes in iter_uploaded_image_entries(file_list):
            total_entries += 1
            code_from_name = extract_code_from_filename(entry_name)
            normalized_code = normalize_lookup_code(code_from_name)

            if not normalized_code:
                skipped_files.append(f'{entry_name}: 文件名无法识别工程编码')
                continue

            target_code = code_lookup.get(normalized_code)
            if not target_code:
                missing_codes.append(code_from_name or entry_name)
                continue

            try:
                data_url = encode_image_bytes_to_data_url(image_bytes, entry_name)
            except ValueError as exc:
                skipped_files.append(f'{entry_name}: {exc}')
                continue

            if target_code in seen_updates:
                duplicate_codes.append(target_code)
            seen_updates[target_code] = data_url

        update_payload = [
            {'code': code, 'image_base64': image_base64}
            for code, image_base64 in seen_updates.items()
        ]
        updated_count = bulk_update_aluminum_images(update_payload)

        missing_codes = sorted(dict.fromkeys(missing_codes))
        duplicate_codes = sorted(dict.fromkeys(duplicate_codes))

        if updated_count <= 0:
            return {
                'success': False,
                'message': '没有匹配到可更新的图片，请检查图片名称是否与工程编码一致',
                'total_files': total_entries,
                'updated_count': 0,
                'missing_codes': missing_codes,
                'duplicate_codes': duplicate_codes,
                'skipped_files': skipped_files,
            }, 400

        return {
            'success': True,
            'message': f'批量上传完成，已更新 {updated_count} 条图片',
            'total_files': total_entries,
            'updated_count': updated_count,
            'updated_codes': list(seen_updates.keys()),
            'missing_codes': missing_codes,
            'duplicate_codes': duplicate_codes,
            'skipped_files': skipped_files,
        }, 200
    except ValueError as exc:
        return {'success': False, 'message': str(exc)}, 400
    except Exception as exc:
        return {'success': False, 'message': f'批量上传图片失败: {exc}'}, 500


def download_aluminum_database(actor_role):
    try:
        if not DATABASE_PATH.exists():
            return {'success': False, 'message': '数据库文件不存在'}, 404

        workbook = create_aluminum_export_workbook(actor_role)
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        download_name = f'aluminum_pricing_database_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            buffer,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as exc:
        return {'success': False, 'message': f'下载完整数据库失败: {exc}'}, 500


def submit_aluminum_change_request(data, actor_user='', actor_role=''):
    try:
        normalized = normalize_request_payload(data)
        action = str(normalized.get('action', '')).strip()
        if action not in {CHANGE_ACTION_CREATE, CHANGE_ACTION_UPDATE, CHANGE_ACTION_DELETE}:
            return {'success': False, 'message': '变更类型不支持'}, 400

        requester = str(actor_user or normalized.get('requester', '') or '').strip() or 'unknown'
        requester_role = str(actor_role or normalized.get('requester_role', '') or '').strip()
        submitted_at = get_now_iso()
        request_data = normalized.get('data') or {}

        target_code = ''
        payload = {}
        snapshot = {}

        if action == CHANGE_ACTION_CREATE:
            payload = validate_aluminum_payload(request_data, require_code=True)
            target_code = str(payload.get(DB_CODE_COLUMN, '')).strip()
        elif action == CHANGE_ACTION_UPDATE:
            target_code = str(normalized.get('record_id', '')).strip()
            if not target_code:
                return {'success': False, 'message': '缺少待修改的工程编码'}, 400
            snapshot = get_aluminum_record(target_code)
            if not snapshot:
                return {'success': False, 'message': '待修改记录不存在'}, 404
            payload = validate_aluminum_payload(request_data, require_code=False)
            payload_code = str(payload.get(DB_CODE_COLUMN, '') or '').strip()
            if payload_code and payload_code != target_code:
                return {'success': False, 'message': '暂不支持修改工程编码'}, 400
            payload[DB_CODE_COLUMN] = target_code
        elif action == CHANGE_ACTION_DELETE:
            target_code = str(normalized.get('record_id', '')).strip()
            if not target_code:
                return {'success': False, 'message': '缺少待删除的工程编码'}, 400
            snapshot = get_aluminum_record(target_code)
            if not snapshot:
                return {'success': False, 'message': '待删除记录不存在'}, 404

        request_id = create_change_request(
            action=action,
            target_code=target_code,
            requester=requester,
            requester_role=requester_role,
            submitted_at=submitted_at,
            payload=payload,
            snapshot=snapshot,
        )
        return {
            'success': True,
            'request_id': request_id,
            'status': CHANGE_STATUS_PENDING,
            'message': '已提交给 admin 审核',
        }, 200
    except ValueError as exc:
        return {'success': False, 'message': str(exc)}, 400
    except Exception as exc:
        return {'success': False, 'message': f'提交失败: {exc}'}, 500


def list_aluminum_change_requests(args, actor_role, actor_user=''):
    try:
        page = max(int(args.get('page', 1)), 1)
        page_size = max(int(args.get('page_size', 10)), 1)
        status = str(args.get('status', '')).strip()
        requester = str(args.get('requester', '')).strip()
        if status and status not in CHANGE_STATUSES:
            return {'success': False, 'message': '审核状态不支持'}, 400

        if actor_role != 'admin':
            requester = requester or str(actor_user or '').strip()
            if not requester:
                return {'success': False, 'message': '缺少申请人标识，无法查询个人申请记录'}, 400

        result = list_change_requests(page, page_size, status=status, requester=requester)
        return {
            'success': True,
            **result,
            'requester': requester,
            'message': '查询成功',
        }, 200
    except Exception as exc:
        return {'success': False, 'message': f'查询失败: {exc}'}, 500


def execute_batch_change_request(request_row):
    payload = dict(request_row.get('payload') or {})
    action = str(request_row.get('action') or '').strip()
    stored_path = str(payload.get('stored_path') or '').strip()
    original_filename = str(payload.get('original_filename') or '').strip()
    submission_mode = str(payload.get('submission_mode') or '').strip()

    if not stored_path:
        raise ValueError('缺少待审核文件路径')

    file_storage = open_saved_file_storage(stored_path, original_filename)
    try:
        if action == CHANGE_ACTION_IMPORT_IMAGES:
            is_excel = submission_mode == 'excel' or original_filename.lower().endswith(('.xlsx', '.xls'))
            result = (
                import_aluminum_images_from_excel(file_storage, 'admin', 'admin')
                if is_excel
                else import_aluminum_images([file_storage], 'admin', 'admin')
            )
        elif action in (CHANGE_ACTION_BATCH_UPDATE_PRICES, CHANGE_ACTION_BATCH_UPDATE_DATA):
            result = batch_update_data_from_excel(file_storage, 'admin', 'admin')
        else:
            raise ValueError('变更类型不支持')
    finally:
        file_storage.close()

    if isinstance(result, tuple):
        payload, status = result
        if status >= 400 or payload.get('success') is False:
            raise ValueError(payload.get('message') or '执行待审核批量任务失败')
        return payload
    return result


def approve_aluminum_change_request(request_id, data, actor_role, actor_user):
    try:
        ensure_admin_role(actor_role)
    except PermissionError as exc:
        return {'success': False, 'message': str(exc)}, 403

    review_note = str((data or {}).get('review_note', '')).strip()
    reviewer = str(actor_user or '').strip() or 'admin'

    try:
        request_row = get_change_request(request_id)
        if not request_row:
            return {'success': False, 'message': '审核记录不存在'}, 404
        if request_row.get('status') != CHANGE_STATUS_PENDING:
            return {'success': False, 'message': '该申请已处理，请刷新后重试'}, 400

        action = request_row.get('action')
        target_code = str(request_row.get('target_code') or '').strip()
        payload = dict(request_row.get('payload') or {})

        if action == CHANGE_ACTION_CREATE:
            create_aluminum_record(payload)
        elif action == CHANGE_ACTION_UPDATE:
            payload.pop(DB_CODE_COLUMN, None)
            update_aluminum_record(target_code, payload)
        elif action == CHANGE_ACTION_DELETE:
            delete_aluminum_record(target_code)
        elif action in {CHANGE_ACTION_IMPORT_IMAGES, CHANGE_ACTION_BATCH_UPDATE_PRICES, CHANGE_ACTION_BATCH_UPDATE_DATA}:
            execute_batch_change_request(request_row)
        else:
            return {'success': False, 'message': '变更类型不支持'}, 400

        update_change_request_status(
            request_id=request_id,
            status=CHANGE_STATUS_APPROVED,
            reviewed_by=reviewer,
            review_note=review_note,
            reviewed_at=get_now_iso(),
        )
        return {'success': True, 'message': '审核通过并已写入数据库'}, 200
    except ValueError as exc:
        return {'success': False, 'message': str(exc)}, 400
    except LookupError as exc:
        return {'success': False, 'message': str(exc)}, 404
    except Exception as exc:
        return {'success': False, 'message': f'审核失败: {exc}'}, 500


def reject_aluminum_change_request(request_id, data, actor_role, actor_user):
    try:
        ensure_admin_role(actor_role)
    except PermissionError as exc:
        return {'success': False, 'message': str(exc)}, 403

    review_note = str((data or {}).get('review_note', '')).strip()
    reviewer = str(actor_user or '').strip() or 'admin'

    try:
        request_row = get_change_request(request_id)
        if not request_row:
            return {'success': False, 'message': '审核记录不存在'}, 404
        if request_row.get('status') != CHANGE_STATUS_PENDING:
            return {'success': False, 'message': '该申请已处理，请刷新后重试'}, 400

        update_change_request_status(
            request_id=request_id,
            status=CHANGE_STATUS_REJECTED,
            reviewed_by=reviewer,
            review_note=review_note,
            reviewed_at=get_now_iso(),
        )
        return {'success': True, 'message': '已驳回该申请'}, 200
    except LookupError as exc:
        return {'success': False, 'message': str(exc)}, 404
    except Exception as exc:
        return {'success': False, 'message': f'驳回失败: {exc}'}, 500


def withdraw_aluminum_change_request(request_id, data, actor_role, actor_user):
    requester = str(actor_user or '').strip() or str((data or {}).get('requester', '')).strip()
    if not requester:
        return {'success': False, 'message': '缺少申请人标识，无法撤回申请'}, 400

    try:
        request_row = get_change_request(request_id)
        if not request_row:
            return {'success': False, 'message': '审核记录不存在'}, 404
        if request_row.get('status') != CHANGE_STATUS_PENDING:
            return {'success': False, 'message': '只能撤回待审核状态的申请'}, 400

        owner = str(request_row.get('requester') or '').strip()
        if owner != requester:
            return {'success': False, 'message': '只能撤回自己提交的申请'}, 403

        update_change_request_status(
            request_id=request_id,
            status=CHANGE_STATUS_WITHDRAWN,
            reviewed_by=requester,
            review_note='申请人已撤回',
            reviewed_at=get_now_iso(),
        )
        return {'success': True, 'message': '申请已撤回'}, 200
    except LookupError as exc:
        return {'success': False, 'message': str(exc)}, 404
    except Exception as exc:
        return {'success': False, 'message': f'撤回失败: {exc}'}, 500


def import_aluminum_images_from_excel(file_storage, actor_role, actor_user=''):
    if not file_storage or not getattr(file_storage, 'filename', ''):
        return {'success': False, 'message': '请上传 Excel 文件'}, 400

    filename = file_storage.filename.lower()
    if not filename.endswith(('.xlsx', '.xls')):
        return {'success': False, 'message': '仅支持 .xlsx / .xls 文件'}, 400

    if not can_bypass_image_review(actor_role, actor_user):
        try:
            saved_file = save_request_file(file_storage, fallback_suffix=Path(filename).suffix or '.xlsx')
            return create_batch_change_request(
                CHANGE_ACTION_IMPORT_IMAGES,
                actor_user,
                actor_role,
                build_file_change_request_payload(
                    {
                        **saved_file,
                        'submission_mode': 'excel',
                        'file_count': 1,
                    }
                ),
            )
        except Exception as exc:
            return {'success': False, 'message': f'提交图片导入审核失败: {exc}'}, 500

    try:
        import tempfile, os
        import xml.etree.ElementTree as ET
        import openpyxl

        tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
        try:
            file_storage.save(tmp_path)
            os.close(tmp_fd)
        except Exception:
            os.close(tmp_fd)

        image_map = {}
        with zipfile.ZipFile(tmp_path, 'r') as zf:
            rid_to_target = {}
            try:
                rels_xml = zf.read('xl/_rels/cellimages.xml.rels').decode('utf-8')
                for rel in ET.fromstring(rels_xml):
                    rid = rel.get('Id')
                    target = rel.get('Target')
                    if rid and target:
                        rid_to_target[rid] = target
            except KeyError:
                rels_xml = None

            try:
                ci_xml = zf.read('xl/cellimages.xml').decode('utf-8')
            except KeyError:
                ci_xml = None

            if ci_xml:
                ns = {
                    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
                    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                    'etc': 'http://www.wps.cn/officeDocument/2017/etCustomData',
                }
                root = ET.fromstring(ci_xml)
                for ci in root.findall('.//etc:cellImage', ns):
                    cnv = ci.find('.//xdr:cNvPr', ns)
                    if cnv is None:
                        continue
                    img_name = cnv.get('name', '')
                    blip = ci.find('.//a:blip', ns)
                    rid = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed', '') if blip is not None else ''
                    if img_name and rid:
                        target = rid_to_target.get(rid, '')
                        if target:
                            media_path = 'xl/' + target
                            try:
                                img_bytes = zf.read(media_path)
                                image_map[img_name] = {
                                    'bytes': img_bytes,
                                    'filename': target,
                                }
                            except KeyError:
                                pass

            if not image_map:
                wb_tmp = openpyxl.load_workbook(tmp_path)
                ws_tmp = wb_tmp.active
                for idx, img_obj in enumerate(ws_tmp._images):
                    from io import BytesIO
                    buf = BytesIO()
                    raw_image = img_obj._data()
                    if hasattr(raw_image, 'save'):
                        raw_image.save(buf)
                    else:
                        buf.write(raw_image)
                    image_map[f'_img_{idx}'] = {
                        'bytes': buf.getvalue(),
                        'filename': getattr(getattr(img_obj, 'ref', None), 'filename', '') or f'image_{idx}.png',
                    }
                wb_tmp.close()

        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active

        code_lookup = {}
        for code in list_all_aluminum_codes():
            normalized_code = normalize_lookup_code(code)
            if normalized_code and normalized_code not in code_lookup:
                code_lookup[normalized_code] = code

        seen_updates = {}
        missing_codes = []
        skipped_rows = []

        img_col = 1
        code_col = 2
        for probe_r in range(2, min(ws.max_row + 1, 10)):
            v1 = str(ws.cell(row=probe_r, column=1).value or '')
            v2 = str(ws.cell(row=probe_r, column=2).value or '')
            if 'DISPIMG' in v1.upper():
                img_col = 1
                code_col = 2
                break
            if 'DISPIMG' in v2.upper():
                img_col = 2
                code_col = 1
                break

        for r in range(2, ws.max_row + 1):
            codes_text = ws.cell(row=r, column=code_col).value
            if not codes_text or not str(codes_text).strip():
                continue

            cell_img = ws.cell(row=r, column=img_col).value
            img_id = None
            cell_img_text = str(cell_img or '')
            if cell_img_text and 'DISPIMG(' in cell_img_text.upper():
                match = re.search(r'DISPIMG\("([^"]+)"', cell_img_text, re.IGNORECASE)
                if match:
                    img_id = match.group(1)

            img_entry = image_map.get(img_id) if img_id else None
            if not img_entry:
                if img_id:
                    skipped_rows.append(f'Row {r}: 未找到图片资源 {img_id}')
                else:
                    skipped_rows.append(f'Row {r}: 未识别到图片公式')
                continue

            codes = [c.strip() for c in str(codes_text).replace('\\n', '\n').split('\n') if c.strip()]
            if not codes:
                continue

            try:
                data_url = encode_image_bytes_to_data_url(
                    img_entry['bytes'],
                    img_entry.get('filename', ''),
                )
            except ValueError as exc:
                skipped_rows.append(f'Row {r}: 图片处理失败 - {exc}')
                continue

            for code_str in codes:
                normalized_code = normalize_lookup_code(code_str)
                if not normalized_code:
                    skipped_rows.append(f'Row {r}: 编码格式无效 "{code_str}"')
                    continue
                target_code = code_lookup.get(normalized_code)
                if not target_code:
                    missing_codes.append(code_str)
                    continue
                seen_updates[target_code] = data_url

        wb.close()
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

        update_payload = [
            {'code': code, 'image_base64': img_b64}
            for code, img_b64 in seen_updates.items()
        ]
        updated_count = bulk_update_aluminum_images(update_payload)

        missing_codes = sorted(dict.fromkeys(missing_codes))

        if updated_count <= 0:
            return {
                'success': False,
                'message': '没有匹配到可更新的编码，请检查Excel中编码列是否与数据库工程编码一致',
                'updated_count': 0,
                'missing_codes': missing_codes,
                'skipped_rows': skipped_rows,
            }, 400

        return {
            'success': True,
            'message': f'Excel图片导入完成，已更新 {updated_count} 条图片',
            'updated_count': updated_count,
            'updated_codes': list(seen_updates.keys()),
            'missing_codes': missing_codes,
            'skipped_rows': skipped_rows,
        }, 200
    except ValueError as exc:
        return {'success': False, 'message': str(exc)}, 400
    except Exception as exc:
        return {'success': False, 'message': f'Excel图片导入失败: {exc}'}, 500


def batch_update_prices_from_excel(file_storage, actor_role, actor_user=''):
    return batch_update_data_from_excel(file_storage, actor_role, actor_user)


def batch_update_data_from_excel(file_storage, actor_role, actor_user=''):
    if not file_storage or not getattr(file_storage, 'filename', ''):
        return {'success': False, 'message': '请上传 Excel 文件'}, 400

    if actor_role != 'admin':
        try:
            saved_file = save_request_file(file_storage, fallback_suffix='.xlsx')
            return create_batch_change_request(
                CHANGE_ACTION_BATCH_UPDATE_DATA,
                actor_user,
                actor_role,
                build_file_change_request_payload(saved_file),
            )
        except Exception as exc:
            return {'success': False, 'message': f'提交数据更新审核失败: {exc}'}, 500

    try:
        import openpyxl
        from backend.config.settings import DATABASE_PATH

        wb = openpyxl.load_workbook(file_storage, data_only=True)
        ws = wb.active

        header_row = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=c).value
            if val:
                header_row[c] = str(val).strip()

        code_col = None
        for c, name in header_row.items():
            if name == DB_CODE_COLUMN:
                code_col = c
                break

        if code_col is None:
            return {'success': False, 'message': 'Excel 第一行未找到"工程编码"列'}, 400

        conn = get_db_connection()
        cursor = conn.cursor()

        col_info = conn.execute(f'PRAGMA table_info({DB_TABLE_NAME})').fetchall()
        db_columns = {row[1] for row in col_info}
        skip_columns = {DB_IMAGE_COLUMN, DB_IMAGE_BASE64_COLUMN}

        EXCEL_TO_DB_COLUMN_MAP = {
            '单位': DB_UNIT_COLUMN,
        }
        for c in list(header_row.keys()):
            name = header_row[c]
            if name in EXCEL_TO_DB_COLUMN_MAP:
                header_row[c] = EXCEL_TO_DB_COLUMN_MAP[name]

        new_columns = []
        for c, name in header_row.items():
            if c == code_col or name in skip_columns or name in db_columns:
                continue
            new_columns.append(name)
            conn.execute(f'ALTER TABLE {DB_TABLE_NAME} ADD COLUMN "{name}" TEXT')
        if new_columns:
            conn.commit()
            col_info = conn.execute(f'PRAGMA table_info({DB_TABLE_NAME})').fetchall()
            db_columns = {row[1] for row in col_info}

        valid_cols = {}
        for c, name in header_row.items():
            if c == code_col or name in skip_columns:
                continue
            if name in db_columns:
                valid_cols[c] = name

        if not valid_cols:
            conn.close()
            return {'success': False, 'message': 'Excel 中没有可更新的数据列'}, 400

        updated_count = 0
        inserted_count = 0
        skipped_count = 0
        missing_codes = []

        for r in range(2, ws.max_row + 1):
            code_val = ws.cell(row=r, column=code_col).value
            if not code_val or not str(code_val).strip():
                continue
            code = str(code_val).strip()

            updates = {}
            for c, col_name in valid_cols.items():
                cell_val = ws.cell(row=r, column=c).value
                if cell_val is not None:
                    updates[col_name] = cell_val

            if not updates:
                continue

            cursor.execute(
                f'SELECT COUNT(*) FROM {DB_TABLE_NAME} WHERE "{DB_CODE_COLUMN}" = ?',
                (code,)
            )
            if cursor.fetchone()[0] == 0:
                insert_cols = [DB_CODE_COLUMN] + list(updates.keys())
                insert_vals = [code] + list(updates.values())
                col_names = ', '.join([f'"{col}"' for col in insert_cols])
                placeholders = ', '.join(['?'] * len(insert_cols))
                cursor.execute(
                    f'INSERT INTO {DB_TABLE_NAME} ({col_names}) VALUES ({placeholders})',
                    insert_vals,
                )
                inserted_count += 1
                continue

            set_clause = ', '.join([f'"{col}" = ?' for col in updates.keys()])
            sql = f'UPDATE {DB_TABLE_NAME} SET {set_clause} WHERE "{DB_CODE_COLUMN}" = ?'
            cursor.execute(sql, list(updates.values()) + [code])
            if cursor.rowcount > 0:
                updated_count += 1

        conn.commit()
        conn.close()

        msg = f'批量更新完成，共更新 {updated_count} 条记录'
        if inserted_count:
            msg += f'，新增 {inserted_count} 条'
        if skipped_count:
            msg += f'，跳过 {skipped_count} 条'
        if new_columns:
            msg += f'，新增 {len(new_columns)} 列：{", ".join(new_columns)}'

        return {
            'success': True,
            'message': msg,
            'updated_count': updated_count,
            'inserted_count': inserted_count,
            'skipped_count': skipped_count,
            'missing_codes_sample': missing_codes[:20],
            'updated_columns': list(valid_cols.values()),
            'new_columns': new_columns,
        }, 200
    except ValueError as exc:
        return {'success': False, 'message': str(exc)}, 400
    except Exception as exc:
        return {'success': False, 'message': f'批量更新数据失败: {exc}'}, 500


def add_aluminum_column(data, actor_role):
    if actor_role != 'admin':
        return {'success': False, 'message': '仅管理员可新增列'}, 403

    column_name = str((data or {}).get('column_name', '')).strip()
    if not column_name:
        return {'success': False, 'message': '列名不能为空'}, 400

    try:
        conn = get_db_connection()
        col_info = conn.execute(f'PRAGMA table_info({DB_TABLE_NAME})').fetchall()
        existing = {row[1] for row in col_info}
        if column_name in existing:
            conn.close()
            return {'success': False, 'message': f'列 "{column_name}" 已存在'}, 400

        conn.execute(f'ALTER TABLE {DB_TABLE_NAME} ADD COLUMN "{column_name}" TEXT')
        conn.commit()

        new_col_info = conn.execute(f'PRAGMA table_info({DB_TABLE_NAME})').fetchall()
        conn.close()
        return {
            'success': True,
            'message': f'已新增列 "{column_name}"',
            'column_name': column_name,
            'total_columns': len(new_col_info),
        }, 200
    except Exception as exc:
        return {'success': False, 'message': f'新增列失败: {exc}'}, 500
