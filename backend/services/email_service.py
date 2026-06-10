import smtplib
import os
import mimetypes
import html as _html_module
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime

from backend.config.settings import (
    INQUIRY_SMTP_HOST,
    INQUIRY_SMTP_PORT,
    INQUIRY_SMTP_USER,
    INQUIRY_SMTP_PASSWORD,
    INQUIRY_SMTP_FROM,
    INQUIRY_SMTP_TO,
    INQUIRY_SMTP_CC,
)


def _esc(value):
    return _html_module.escape(str(value))


def send_inquiry_email(
    file_path,
    project_name='',
    material_count=0,
    unmatched_products=None,
    sender_name='',
    bom_filename='',
    remark='',
    extra_attachments=None,
):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f'询价表文件不存在: {file_path}')

    if not INQUIRY_SMTP_USER or not INQUIRY_SMTP_PASSWORD or not INQUIRY_SMTP_FROM or not INQUIRY_SMTP_TO:
        raise RuntimeError('询价邮件未配置，请检查 KS_INQUIRY_SMTP_* 环境变量')

    now = datetime.now()
    date_str = now.strftime('%Y-%m-%d %H:%M')
    date_short = now.strftime('%Y%m%d')

    safe_project = _esc(project_name)
    if project_name:
        subject = f'【inquiry询价】{project_name} - {material_count}项物料待询价 ({date_str})'
    else:
        subject = f'【inquiry询价】新询价单 - {material_count}项物料待询价 ({date_str})'

    body_lines = [
        '<html><body style="font-family: Arial, sans-serif; font-size: 14px; color: #333;">',
        '<div style="max-width: 700px; margin: 0 auto;">',
        '<h2 style="color: #1a56db; border-bottom: 2px solid #1a56db; padding-bottom: 8px;">询价通知</h2>',
        f'<p>发送时间: <strong>{date_str}</strong></p>',
    ]

    if project_name:
        body_lines.append(f'<p>项目名称: <strong>{safe_project}</strong></p>')

    if sender_name:
        body_lines.append(f'<p>询价人: <strong>{_esc(sender_name)}</strong></p>')

    if bom_filename:
        body_lines.append(f'<p>源文件: {_esc(bom_filename)}</p>')

    body_lines.append(f'<p>待询价物料数量: <strong style="color: #dc2626;">{material_count} 项</strong></p>')

    if unmatched_products:
        body_lines.append('<hr style="border: 1px solid #e5e7eb; margin: 16px 0;">')
        body_lines.append('<h3 style="color: #374151;">物料清单</h3>')
        body_lines.append(
            '<table style="border-collapse: collapse; width: 100%; font-size: 13px;">'
            '<tr style="background: #f3f4f6;">'
            '<th style="border: 1px solid #d1d5db; padding: 6px 10px; text-align: left;">序号</th>'
            '<th style="border: 1px solid #d1d5db; padding: 6px 10px; text-align: left;">产品编码</th>'
            '<th style="border: 1px solid #d1d5db; padding: 6px 10px; text-align: left;">产品名称</th>'
            '<th style="border: 1px solid #d1d5db; padding: 6px 10px; text-align: left;">规格</th>'
            '<th style="border: 1px solid #d1d5db; padding: 6px 10px; text-align: right;">数量</th>'
            '</tr>'
        )
        for idx, p in enumerate(unmatched_products[:50], 1):
            bg = '#ffffff' if idx % 2 == 1 else '#f9fafb'
            body_lines.append(
                f'<tr style="background: {bg};">'
                f'<td style="border: 1px solid #d1d5db; padding: 4px 10px;">{idx}</td>'
                f'<td style="border: 1px solid #d1d5db; padding: 4px 10px;">{_esc(p.get("code", ""))}</td>'
                f'<td style="border: 1px solid #d1d5db; padding: 4px 10px;">{_esc(p.get("name", ""))}</td>'
                f'<td style="border: 1px solid #d1d5db; padding: 4px 10px;">{_esc(p.get("spec", ""))}</td>'
                f'<td style="border: 1px solid #d1d5db; padding: 4px 10px; text-align: right;">{_esc(p.get("quantity", ""))}</td>'
                f'</tr>'
            )
        if len(unmatched_products) > 50:
            body_lines.append(
                f'<tr><td colspan="5" style="border: 1px solid #d1d5db; padding: 6px 10px; text-align: center; color: #6b7280;">'
                f'... 还有 {len(unmatched_products) - 50} 项，详见附件</td></tr>'
            )
        body_lines.append('</table>')

    if remark:
        body_lines.append('<hr style="border: 1px solid #e5e7eb; margin: 16px 0;">')
        body_lines.append('<h3 style="color: #374151;">备注</h3>')
        body_lines.append(
            f'<div style="background: #fefce8; border: 1px solid #fde68a; border-radius: 6px; padding: 10px 14px; '
            f'font-size: 13px; color: #92400e; white-space: pre-wrap; word-break: break-word;">'
            f'{_esc(remark)}</div>'
        )

    body_lines.append('<hr style="border: 1px solid #e5e7eb; margin: 16px 0;">')
    body_lines.append('<p style="color: #6b7280; font-size: 12px;">此邮件由BOM智能报价系统自动发送，请勿直接回复。</p>')
    body_lines.append('</div></body></html>')

    msg = MIMEMultipart()
    msg['From'] = INQUIRY_SMTP_FROM
    msg['To'] = INQUIRY_SMTP_TO
    cc_list = [addr.strip() for addr in INQUIRY_SMTP_CC.split(',') if addr.strip()] if INQUIRY_SMTP_CC else []
    if cc_list:
        msg['Cc'] = ', '.join(cc_list)
    msg['Subject'] = subject

    msg.attach(MIMEText(''.join(body_lines), 'html', 'utf-8'))

    filename = os.path.basename(file_path)
    if not project_name:
        attach_name = filename
    else:
        ext = os.path.splitext(filename)[1]
        safe_name = project_name.replace(' ', '_').replace('/', '_')
        attach_name = f'{safe_name}_询价表_{date_short}{ext}'

    with open(file_path, 'rb') as f:
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(f.read())

    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment', filename=attach_name)
    msg.attach(part)

    if extra_attachments:
        for att in extra_attachments:
            att_bytes = att.get('data')
            att_name = att.get('filename', 'attachment')
            if not att_bytes:
                continue
            mime_type, _ = mimetypes.guess_type(att_name)
            if not mime_type:
                mime_type = 'application/octet-stream'
            maintype, subtype = mime_type.split('/', 1)
            att_part = MIMEBase(maintype, subtype)
            att_part.set_payload(att_bytes)
            encoders.encode_base64(att_part)
            att_part.add_header('Content-Disposition', 'attachment', filename=att_name)
            msg.attach(att_part)

    all_recipients = [INQUIRY_SMTP_TO] + cc_list
    server = None
    try:
        print(f'[EMAIL] Connecting to {INQUIRY_SMTP_HOST}:{INQUIRY_SMTP_PORT} ...')
        if INQUIRY_SMTP_PORT == 465:
            server = smtplib.SMTP_SSL(INQUIRY_SMTP_HOST, INQUIRY_SMTP_PORT, timeout=30)
        else:
            server = smtplib.SMTP(INQUIRY_SMTP_HOST, INQUIRY_SMTP_PORT, timeout=30)
            server.ehlo()
            server.starttls()
            server.ehlo()

        server.login(INQUIRY_SMTP_USER, INQUIRY_SMTP_PASSWORD)
        server.sendmail(INQUIRY_SMTP_FROM, all_recipients, msg.as_string())
        print(f'[EMAIL] Sent inquiry email to {all_recipients}, subject: {subject}')
    except smtplib.SMTPAuthenticationError as exc:
        print(f'[EMAIL] SMTP auth failed: {exc}')
        raise RuntimeError(
            '邮箱认证失败（SMTP 535）。'
            '请在腾讯企业邮管理后台为 inquiry@xmkseng.com 开启SMTP服务，'
            '并生成客户端专用密码（授权码），将其填入 .env.local 的 KS_INQUIRY_SMTP_PASSWORD。'
        ) from exc
    except smtplib.SMTPException as exc:
        print(f'[EMAIL] SMTP error: {exc}')
        raise RuntimeError(f'邮件发送失败（SMTP错误）: {exc}') from exc
    finally:
        if server:
            try:
                server.quit()
            except Exception:
                pass

    return True


def send_image_inquiry_email(
    codes,
    project_name='',
    sender_name='',
    designer_email='',
    remark='',
    items=None,
    cc_emails=None,
):
    if not INQUIRY_SMTP_USER or not INQUIRY_SMTP_PASSWORD or not INQUIRY_SMTP_FROM:
        raise RuntimeError('询价邮件未配置，请检查 KS_INQUIRY_SMTP_* 环境变量')

    if not codes and not items:
        raise ValueError('缺失图片编码列表为空')

    if not codes:
        codes = [item.get('code', '') for item in (items or [])]

    now = datetime.now()
    date_str = now.strftime('%Y-%m-%d %H:%M')
    date_short = now.strftime('%Y%m%d')

    safe_project = _esc(project_name)
    if project_name:
        subject = f'[询图] {project_name} - 缺失图片编码列表 ({date_str})'
    else:
        subject = f'[询图] 缺失图片编码列表 ({date_str})'

    body_lines = [
        '<html><body style="font-family: Arial, sans-serif; font-size: 14px; color: #333;">',
        '<div style="max-width: 700px; margin: 0 auto;">',
        '<h2 style="color: #7c3aed; border-bottom: 2px solid #7c3aed; padding-bottom: 8px;">询图通知</h2>',
        f'<p>发送时间: <strong>{date_str}</strong></p>',
    ]

    if project_name:
        body_lines.append(f'<p>项目名称: <strong>{safe_project}</strong></p>')

    if sender_name:
        body_lines.append(f'<p>发送人: <strong>{_esc(sender_name)}</strong></p>')

    body_lines.append(f'<p>缺失图片编码数量: <strong style="color: #7c3aed;">{len(codes)} 项</strong></p>')

    body_lines.append('<hr style="border: 1px solid #e5e7eb; margin: 16px 0;">')
    body_lines.append('<h3 style="color: #374151;">缺失图片编码清单</h3>')
    body_lines.append(
        '<table style="border-collapse: collapse; width: 100%; font-size: 13px;">'
        '<tr style="background: #f3f4f6;">'
        '<th style="border: 1px solid #d1d5db; padding: 6px 10px; text-align: left;">序号</th>'
        '<th style="border: 1px solid #d1d5db; padding: 6px 10px; text-align: left;">工程编码</th>'
        '<th style="border: 1px solid #d1d5db; padding: 6px 10px; text-align: left;">工程品名</th>'
        '</tr>'
    )
    name_map = {}
    if items:
        for it in items:
            name_map[str(it.get('code', ''))] = str(it.get('name', ''))
    for idx, code in enumerate(codes[:200], 1):
        bg = '#ffffff' if idx % 2 == 1 else '#f9fafb'
        item_name = _esc(name_map.get(str(code), ''))
        body_lines.append(
            f'<tr style="background: {bg};">'
            f'<td style="border: 1px solid #d1d5db; padding: 4px 10px;">{idx}</td>'
            f'<td style="border: 1px solid #d1d5db; padding: 4px 10px;">{_esc(str(code))}</td>'
            f'<td style="border: 1px solid #d1d5db; padding: 4px 10px;">{item_name}</td>'
            f'</tr>'
        )
    if len(codes) > 200:
        body_lines.append(
            f'<tr><td colspan="3" style="border: 1px solid #d1d5db; padding: 6px 10px; text-align: center; color: #6b7280;">'
            f'... 还有 {len(codes) - 200} 项</td></tr>'
        )
    body_lines.append('</table>')

    body_lines.append('<hr style="border: 1px solid #e5e7eb; margin: 16px 0;">')
    body_lines.append('<h3 style="color: #374151;">回复说明</h3>')
    body_lines.append(
        '<div style="background: #f0f9ff; border: 1px solid #bae6fd; border-radius: 6px; padding: 10px 14px; '
        'font-size: 13px; color: #0c4a6e;">'
        '请在回复邮件时，以附件形式发送对应编码的产品图片（支持 png/jpg/jpeg/bmp/webp 格式）。<br>'
        '每张图片的文件名必须包含对应的工程编码，例如：<code>WTP-ABC123.jpg</code><br>'
        '系统将自动解析回复邮件中的图片并写入数据库。</div>'
    )

    if remark:
        body_lines.append('<hr style="border: 1px solid #e5e7eb; margin: 16px 0;">')
        body_lines.append('<h3 style="color: #374151;">备注</h3>')
        body_lines.append(
            f'<div style="background: #fefce8; border: 1px solid #fde68a; border-radius: 6px; padding: 10px 14px; '
            f'font-size: 13px; color: #92400e; white-space: pre-wrap; word-break: break-word;">'
            f'{_esc(remark)}</div>'
        )

    import openpyxl
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws['A1'] = '图片'
    ws['B1'] = '对应编码'
    ws['C1'] = '工程品名'
    for i, code in enumerate(codes, start=2):
        ws.cell(row=i, column=2, value=str(code))
        item_name = name_map.get(str(code), '')
        if item_name:
            ws.cell(row=i, column=3, value=item_name)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    body_lines.append('<hr style="border: 1px solid #e5e7eb; margin: 16px 0;">')
    body_lines.append('<p style="color: #6b7280; font-size: 12px;">此邮件由BOM智能报价系统自动发送，请将产品图片填入附件模板的"图片"列并回复。</p>')
    body_lines.append('</div></body></html>')

    if not designer_email:
        raise ValueError('询图邮件必须指定组长邮箱（designer_email），不能为空')

    msg = MIMEMultipart()
    msg['From'] = INQUIRY_SMTP_FROM
    recipient = designer_email
    msg['To'] = recipient
    msg['Subject'] = subject

    cc_list = [addr.strip() for addr in (cc_emails or []) if addr.strip()]
    if cc_list:
        msg['Cc'] = ', '.join(cc_list)

    msg.attach(MIMEText(''.join(body_lines), 'html', 'utf-8'))

    safe_name = project_name.replace(' ', '_').replace('/', '_') if project_name else '缺失图片'
    attach_name = f'{safe_name}_编码模板_{date_short}.xlsx'
    part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    part.set_payload(buf.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment', filename=attach_name)
    msg.attach(part)

    server = None
    try:
        print(f'[EMAIL] Connecting to {INQUIRY_SMTP_HOST}:{INQUIRY_SMTP_PORT} ...')
        if INQUIRY_SMTP_PORT == 465:
            server = smtplib.SMTP_SSL(INQUIRY_SMTP_HOST, INQUIRY_SMTP_PORT, timeout=30)
        else:
            server = smtplib.SMTP(INQUIRY_SMTP_HOST, INQUIRY_SMTP_PORT, timeout=30)
            server.ehlo()
            server.starttls()
            server.ehlo()

        server.login(INQUIRY_SMTP_USER, INQUIRY_SMTP_PASSWORD)
        all_recipients = [recipient] + cc_list
        server.sendmail(INQUIRY_SMTP_FROM, all_recipients, msg.as_string())
        print(f'[EMAIL] Sent image inquiry email to {recipient} cc={cc_list}, subject: {subject}')
    except smtplib.SMTPAuthenticationError as exc:
        print(f'[EMAIL] SMTP auth failed: {exc}')
        raise RuntimeError(
            '邮箱认证失败（SMTP 535）。'
            '请在腾讯企业邮管理后台为 inquiry@xmkseng.com 开启SMTP服务，'
            '并生成客户端专用密码（授权码），将其填入 .env.local 的 KS_INQUIRY_SMTP_PASSWORD。'
        ) from exc
    except smtplib.SMTPException as exc:
        print(f'[EMAIL] SMTP error: {exc}')
        raise RuntimeError(f'邮件发送失败（SMTP错误）: {exc}') from exc
    finally:
        if server:
            try:
                server.quit()
            except Exception:
                pass

    return True


def forward_inquiry_reply(
    original_subject: str,
    to_email: str,
    text_parts: list,
    image_parts: list,
    parsed_items: list,
):
    if not INQUIRY_SMTP_USER or not INQUIRY_SMTP_PASSWORD or not INQUIRY_SMTP_FROM:
        raise RuntimeError('SMTP 未配置')

    now = datetime.now()
    date_str = now.strftime('%Y-%m-%d %H:%M')
    forward_subject = f'[报价回复] {original_subject}'

    body_lines = [
        '<html><body style="font-family: Arial, sans-serif; font-size: 14px; color: #333;">',
        '<div style="max-width: 700px; margin: 0 auto;">',
        '<h2 style="color: #059669; border-bottom: 2px solid #059669; padding-bottom: 8px;">报价回复通知</h2>',
        f'<p>转发时间: <strong>{date_str}</strong></p>',
        f'<p>原始询价主题: {_esc(original_subject)}</p>',
    ]

    if parsed_items:
        body_lines.append('<hr style="border: 1px solid #e5e7eb; margin: 16px 0;">')
        body_lines.append('<h3 style="color: #374151;">报价解析结果</h3>')
        body_lines.append(
            '<table style="border-collapse: collapse; width: 100%; font-size: 13px;">'
            '<tr style="background: #f3f4f6;">'
            '<th style="border: 1px solid #d1d5db; padding: 6px 8px; text-align: left;">物料编码</th>'
            '<th style="border: 1px solid #d1d5db; padding: 6px 8px; text-align: left;">名称</th>'
            '<th style="border: 1px solid #d1d5db; padding: 6px 8px; text-align: left;">规格</th>'
            '<th style="border: 1px solid #d1d5db; padding: 6px 8px; text-align: right;">数量</th>'
            '<th style="border: 1px solid #d1d5db; padding: 6px 8px; text-align: right;">售价(美元)</th>'
            '<th style="border: 1px solid #d1d5db; padding: 6px 8px; text-align: right;">售价(人民币)</th>'
            '<th style="border: 1px solid #d1d5db; padding: 6px 8px; text-align: right;">售价(欧元)</th>'
            '<th style="border: 1px solid #d1f5db; padding: 6px 8px; text-align: left;">单位</th>'
            '<th style="border: 1px solid #d1d5db; padding: 6px 8px; text-align: left;">有效期</th>'
            '</tr>'
        )
        for idx, item in enumerate(parsed_items[:100], 1):
            bg = '#ffffff' if idx % 2 == 1 else '#f9fafb'
            usd = item.get('unit_price_usd') or ''
            cny = item.get('unit_price_cny') or ''
            eur = item.get('unit_price_eur') or ''
            body_lines.append(
                f'<tr style="background: {bg};">'
                f'<td style="border: 1px solid #d1d5db; padding: 4px 8px;">{_esc(item.get("material_code", ""))}</td>'
                f'<td style="border: 1px solid #d1d5db; padding: 4px 8px;">{_esc(item.get("name", ""))}</td>'
                f'<td style="border: 1px solid #d1d5db; padding: 4px 8px;">{_esc(item.get("spec", ""))}</td>'
                f'<td style="border: 1px solid #d1d5db; padding: 4px 8px; text-align: right;">{_esc(str(item.get("quantity", "")))}</td>'
                f'<td style="border: 1px solid #d1d5db; padding: 4px 8px; text-align: right;">{_esc(str(usd))}</td>'
                f'<td style="border: 1px solid #d1d5db; padding: 4px 8px; text-align: right;">{_esc(str(cny))}</td>'
                f'<td style="border: 1px solid #d1d5db; padding: 4px 8px; text-align: right;">{_esc(str(eur))}</td>'
                f'<td style="border: 1px solid #d1d5db; padding: 4px 8px;">{_esc(item.get("unit", ""))}</td>'
                f'<td style="border: 1px solid #d1d5db; padding: 4px 8px;">{_esc(item.get("valid_until", ""))}</td>'
                f'</tr>'
            )
        body_lines.append('</table>')

    if text_parts:
        body_lines.append('<hr style="border: 1px solid #e5e7eb; margin: 16px 0;">')
        body_lines.append('<h3 style="color: #374151;">原始邮件内容</h3>')
        for tp in text_parts[:3]:
            safe_text = _esc(tp[:2000])
            body_lines.append(
                f'<pre style="background: #f9fafb; padding: 10px; border: 1px solid #e5e7eb; '
                f'border-radius: 4px; font-size: 12px; white-space: pre-wrap; word-break: break-all;">'
                f'{safe_text}</pre>'
            )

    body_lines.append('<hr style="border: 1px solid #e5e7eb; margin: 16px 0;">')
    body_lines.append('<p style="color: #6b7280; font-size: 12px;">此邮件由BOM智能报价系统自动转发，供应商报价已解析并缓存。</p>')
    body_lines.append('</div></body></html>')

    msg = MIMEMultipart()
    msg['From'] = INQUIRY_SMTP_FROM
    msg['To'] = to_email
    msg['Subject'] = forward_subject

    msg.attach(MIMEText(''.join(body_lines), 'html', 'utf-8'))

    import base64 as _b64
    img_idx = 0
    for mime_type, b64_data in image_parts:
        img_idx += 1
        ext = mime_type.split('/')[-1]
        if ext == 'jpeg':
            ext = 'jpg'
        filename = f'报价截图_{img_idx}.{ext}'
        try:
            img_bytes = _b64.b64decode(b64_data)
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(img_bytes)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment', filename=filename)
            msg.attach(part)
        except Exception:
            pass

    server = None
    try:
        print(f'[EMAIL] Forwarding reply to {to_email}')
        if INQUIRY_SMTP_PORT == 465:
            server = smtplib.SMTP_SSL(INQUIRY_SMTP_HOST, INQUIRY_SMTP_PORT, timeout=30)
        else:
            server = smtplib.SMTP(INQUIRY_SMTP_HOST, INQUIRY_SMTP_PORT, timeout=30)
            server.ehlo()
            server.starttls()
            server.ehlo()

        server.login(INQUIRY_SMTP_USER, INQUIRY_SMTP_PASSWORD)
        server.sendmail(INQUIRY_SMTP_FROM, [to_email], msg.as_string())
        print(f'[EMAIL] Forwarded reply to {to_email}, subject: {forward_subject}')
    except smtplib.SMTPException as exc:
        print(f'[EMAIL] Forward SMTP error: {exc}')
        raise RuntimeError(f'转发邮件失败: {exc}') from exc
    finally:
        if server:
            try:
                server.quit()
            except Exception:
                pass

    return True
