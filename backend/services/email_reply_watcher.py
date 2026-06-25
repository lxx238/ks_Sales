import base64
import imaplib
import email as email_lib
import json
import os
import re
import tempfile
import threading
from datetime import datetime, timedelta
from email.header import decode_header
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

from backend.config.settings import (
    INQUIRY_IMAP_HOST,
    INQUIRY_IMAP_PASSWORD,
    INQUIRY_IMAP_PORT,
    INQUIRY_IMAP_USER,
)
from backend.repositories.inquiry_repository import (
    backfill_orphan_cache,
    ensure_price_cache_schema,
    find_record_by_subject,
    upsert_price_cache,
    update_inquiry_reply,
    _strip_reply_prefix,
    _get_conn as _get_db_conn,
)
from backend.services.llm_service import (
    extract_quotation_from_images,
    extract_quotation_from_text,
    extract_quotation_mixed,
)
from backend.repositories.user_repository import list_unified_contacts


_last_scan_time = None
_last_scan_dt = None

_check_lock = threading.Lock()

_PROCESSED_TABLE = 'ks_processed_emails'


def get_last_scan_time():
    return _last_scan_time


def _parse_imap_internaldate(raw: str) -> Optional[datetime]:
    try:
        import email.utils as eutils
        for part in raw.split('INTERNALDATE'):
            if '"' in part:
                date_str = part.split('"')[1]
                parsed = eutils.parsedate_to_datetime(date_str)
                return parsed
    except Exception:
        pass
    return None


def _ensure_processed_table():
    conn = _get_db_conn()
    try:
        conn.execute(f'''
            CREATE TABLE IF NOT EXISTS {_PROCESSED_TABLE} (
                message_id TEXT PRIMARY KEY,
                subject TEXT NOT NULL DEFAULT '',
                processed_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
            )
        ''')
        conn.commit()
    finally:
        conn.close()


def _is_already_processed(message_id: str) -> bool:
    if not message_id:
        return False
    conn = _get_db_conn()
    try:
        row = conn.execute(
            f'SELECT 1 FROM {_PROCESSED_TABLE} WHERE message_id = ?',
            (message_id.strip(),),
        ).fetchone()
        return row is not None
    finally:
        conn.close()


def _mark_processed(message_id: str, subject: str):
    if not message_id:
        return
    conn = _get_db_conn()
    try:
        conn.execute(
            f'INSERT OR IGNORE INTO {_PROCESSED_TABLE} (message_id, subject) VALUES (?, ?)',
            (message_id.strip(), subject[:200]),
        )
        conn.commit()
    except Exception as exc:
        _log(f'Failed to mark email processed: {exc}')
    finally:
        conn.close()


_REPLY_PREFIXES = ('Re:', 'RE:', '回复:', '回复：')
_FORWARD_PREFIXES = ('Fwd:', 'FW:', '转发:', '转发：')
_QUOTATION_KEYWORDS = ('【inquiry询价】', '[询价]', '询价', '报价', ' quotation', 'quote', 'price', '价格', '单价')
_IMAGE_KEYWORDS = ('[询图]', '询图', '缺失图片', '图片缺失', 'product image')


def _classify_email(subject: str) -> str:
    s = str(subject or '').strip()
    raw_lower = s.lower()
    stripped = s
    changed = True
    while changed:
        changed = False
        for p in (_REPLY_PREFIXES + _FORWARD_PREFIXES):
            if stripped.startswith(p) or f' {p}' in stripped:
                for prefix in (_REPLY_PREFIXES + _FORWARD_PREFIXES):
                    if stripped.startswith(prefix):
                        stripped = stripped[len(prefix):].strip()
                        changed = True
                        break
    stripped_lower = stripped.lower()
    for kw in _IMAGE_KEYWORDS:
        if kw.lower() in stripped_lower:
            return 'image'
    for kw in _QUOTATION_KEYWORDS:
        if kw.lower() in stripped_lower:
            return 'quotation'
    for p in (_REPLY_PREFIXES + _FORWARD_PREFIXES):
        if s.startswith(p) or f' {p}' in s:
            return 'quotation'
    return 'unknown'


def _is_quotation_email(subject: str) -> bool:
    return _classify_email(subject) == 'quotation'


def _is_image_inquiry_email(subject: str) -> bool:
    return _classify_email(subject) == 'image'


def _log(msg: str):
    print(f'[EMAIL-WATCHER] {msg}', flush=True)


def _decode_header_value(value: str) -> str:
    if not value:
        return ''
    decoded_parts = decode_header(value)
    result = []
    for part, charset in decoded_parts:
        if isinstance(part, bytes):
            result.append(part.decode(charset or 'utf-8', errors='replace'))
        else:
            result.append(part)
    return ''.join(result).strip()


def _connect_imap() -> imaplib.IMAP4_SSL:
    if not INQUIRY_IMAP_USER or not INQUIRY_IMAP_PASSWORD:
        raise RuntimeError('IMAP 未配置，请检查 KS_INQUIRY_IMAP_* 环境变量')
    imap = imaplib.IMAP4_SSL(INQUIRY_IMAP_HOST, INQUIRY_IMAP_PORT, timeout=30)
    imap.login(INQUIRY_IMAP_USER, INQUIRY_IMAP_PASSWORD)
    return imap


_PRICE_COL_KEYWORDS = ['单价', '售价', '价格', 'price', '含税价', '未税价', '税込', '単価']
_QTY_COL_KEYWORDS = ['数量', 'qty', 'quantity', '数']
_NAME_COL_KEYWORDS = ['名称', '品名', '物料名称', '品名/规格', 'product', '品名/型式', '物品名', '品名\n规格']
_SPEC_COL_KEYWORDS = ['规格', '型号', '规格型号', 'spec', '型式']
_CODE_COL_KEYWORDS = ['物料编码', '编码', '物料号', '品番', 'code', '部品番号', '品番/规格']
_UNIT_COL_KEYWORDS = ['单位', 'unit', '単位']
_CATEGORY_COL_KEYWORDS = ['类别', '分类', 'category', '種類']
_DATE_COL_KEYWORDS = ['报价日期', '日期', 'date', '納期', '出荷']
_VALID_COL_KEYWORDS = ['有效期', '有效期限', 'valid']
_DISCOUNT_COL_KEYWORDS = ['折扣', 'discount', '割引']

_CURRENCY_SYMBOL_MAP = [('€', 'EUR'), ('$', 'USD'), ('￥', 'CNY'), ('¥', 'CNY')]
_CURRENCY_TEXT_MAP = [
    ('人民币', 'CNY'), ('rmb', 'CNY'), ('cny', 'CNY'),
    ('美元', 'USD'), ('usd', 'USD'), ('dollar', 'USD'),
    ('欧元', 'EUR'), ('eur', 'EUR'), ('euro', 'EUR'),
]

def _detect_currency(*texts) -> Optional[str]:
    for text in texts:
        t = str(text or '').strip()
        if not t:
            continue
        for sym, cur in _CURRENCY_SYMBOL_MAP:
            if sym in t:
                return cur
        tl = t.lower()
        for kw, cur in _CURRENCY_TEXT_MAP:
            if kw.lower() in tl:
                return cur
        if t == '元' or t.startswith('元/') or t.startswith('元\\') or t.startswith('元 '):
            return 'CNY'
    return None


def _match_col_index(columns, keywords):
    for i, col in enumerate(columns):
        c = str(col).strip().lower()
        for kw in keywords:
            if kw.lower() in c:
                return i
    return -1


def _extract_items_from_excel(data: bytes, filename: str) -> List[Dict[str, Any]]:
    try:
        ext = os.path.splitext(filename)[1].lower()
        engine = 'xlrd' if ext == '.xls' else None
        xls = pd.ExcelFile(BytesIO(data), engine=engine) if engine else pd.ExcelFile(BytesIO(data))
        all_items = []
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
            df = df.fillna('')
            cols = [str(c).strip() for c in df.columns]

            price_idx = _match_col_index(cols, _PRICE_COL_KEYWORDS)
            if price_idx < 0:
                continue

            qty_idx = _match_col_index(cols, _QTY_COL_KEYWORDS)
            name_idx = _match_col_index(cols, _NAME_COL_KEYWORDS)
            spec_idx = _match_col_index(cols, _SPEC_COL_KEYWORDS)
            code_idx = _match_col_index(cols, _CODE_COL_KEYWORDS)
            unit_idx = _match_col_index(cols, _UNIT_COL_KEYWORDS)
            cat_idx = _match_col_index(cols, _CATEGORY_COL_KEYWORDS)
            date_idx = _match_col_index(cols, _DATE_COL_KEYWORDS)
            valid_idx = _match_col_index(cols, _VALID_COL_KEYWORDS)
            disc_idx = _match_col_index(cols, _DISCOUNT_COL_KEYWORDS)

            price_col_header = cols[price_idx] if price_idx >= 0 else ''
            prev_col_idx = price_idx - 1 if price_idx > 0 else -1

            for _, row in df.iterrows():
                raw_price = str(row.iloc[price_idx]).strip() if price_idx >= 0 else ''
                if not raw_price or raw_price in ('', '-', '/', 'nan', 'None', '0', '0.0'):
                    continue
                raw_unit = str(row.iloc[unit_idx]).strip() if unit_idx >= 0 else ''
                raw_prev = str(row.iloc[prev_col_idx]).strip() if prev_col_idx >= 0 else ''
                cell_currency = _detect_currency(raw_price, raw_prev, raw_unit, price_col_header)
                try:
                    price = float(raw_price.replace(',', '').replace('￥', '').replace('¥', '').replace('$', '').replace('€', '').strip())
                except (TypeError, ValueError):
                    continue
                if price <= 0:
                    continue

                raw_qty = str(row.iloc[qty_idx]).strip() if qty_idx >= 0 else '0'
                try:
                    qty = float(raw_qty.replace(',', '').strip())
                except (TypeError, ValueError):
                    qty = 0.0

                item = {
                    'material_code': str(row.iloc[code_idx]).strip() if code_idx >= 0 else '',
                    'name': str(row.iloc[name_idx]).strip() if name_idx >= 0 else '',
                    'spec': str(row.iloc[spec_idx]).strip() if spec_idx >= 0 else '',
                    'category': str(row.iloc[cat_idx]).strip() if cat_idx >= 0 else '',
                    'quantity': qty,
                    'unit_price': price,
                    'unit_price_usd': price if cell_currency == 'USD' else None,
                    'unit_price_cny': price if cell_currency == 'CNY' else None,
                    'unit_price_eur': price if cell_currency == 'EUR' else None,
                    'unit': str(row.iloc[unit_idx]).strip() if unit_idx >= 0 else '',
                    'quotation_date': str(row.iloc[date_idx]).strip() if date_idx >= 0 else '',
                    'valid_until': str(row.iloc[valid_idx]).strip() if valid_idx >= 0 else '',
                    'discount': str(row.iloc[disc_idx]).strip() if disc_idx >= 0 else '',
                }
                if item['material_code'] in ('nan', 'None'):
                    item['material_code'] = ''
                if item['name'] in ('nan', 'None'):
                    item['name'] = ''
                if item['spec'] in ('nan', 'None'):
                    item['spec'] = ''
                all_items.append(item)

        return all_items
    except Exception as exc:
        _log(f'Direct Excel extraction failed for {filename}: {exc}')
        return []


def _extract_text_from_excel(data: bytes, filename: str) -> str:
    try:
        ext = os.path.splitext(filename)[1].lower()
        engine = 'xlrd' if ext == '.xls' else None
        all_lines = []
        xls = pd.ExcelFile(BytesIO(data), engine=engine) if engine else pd.ExcelFile(BytesIO(data))
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
            df = df.fillna('')
            lines = []
            for _, row in df.iterrows():
                parts = [str(v).strip() for v in row.values if str(v).strip()]
                if parts:
                    lines.append(' | '.join(parts))
            if lines:
                all_lines.append(f'=== Sheet: {sheet_name} ===')
                all_lines.extend(lines)
        return '\n'.join(all_lines)
    except Exception as exc:
        _log(f'Excel parse failed for {filename}: {exc}')
        return ''


def _extract_attachments(msg: email_lib.message.Message) -> Tuple[List[str], List[Tuple[str, str]], List[Dict[str, Any]]]:
    text_parts = []
    image_parts = []
    pre_parsed_items = []

    for part in msg.walk():
        content_type = part.get_content_type() or ''
        content_disposition = str(part.get('Content-Disposition') or '')
        filename = part.get_filename() or ''

        if filename:
            filename = _decode_header_value(filename)

        if 'attachment' in content_disposition or filename:
            if not filename:
                continue
            payload = part.get_payload(decode=True)
            if not payload:
                continue

            ext = os.path.splitext(filename)[1].lower()
            if ext in ('.xlsx', '.xls', '.csv'):
                items = _extract_items_from_excel(payload, filename)
                if items:
                    pre_parsed_items.extend(items)
                    _log(f'Directly extracted {len(items)} items from {filename}')
                else:
                    text = _extract_text_from_excel(payload, filename)
                    if text:
                        text_parts.append(text)
            elif ext in ('.png', '.jpg', '.jpeg', '.bmp', '.webp'):
                mime_map = {
                    '.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
                    '.bmp': 'image/bmp', '.webp': 'image/webp',
                }
                mime_type = mime_map.get(ext, 'image/png')
                b64 = base64.b64encode(payload).decode('ascii')
                image_parts.append((mime_type, b64))

        elif content_type.startswith('image/'):
            payload = part.get_payload(decode=True)
            if payload:
                cid = str(part.get('Content-ID') or '').strip('<>')
                if cid or 'inline' in content_disposition:
                    b64 = base64.b64encode(payload).decode('ascii')
                    image_parts.append((content_type, b64))

    body_text = ''
    if msg.is_multipart():
        for part in msg.walk():
            ct = part.get_content_type()
            if ct == 'text/plain':
                payload = part.get_payload(decode=True)
                if payload:
                    charset = part.get_content_charset() or 'utf-8'
                    body_text = payload.decode(charset, errors='replace')
                    break
            elif ct == 'text/html' and not body_text:
                payload = part.get_payload(decode=True)
                if payload:
                    charset = part.get_content_charset() or 'utf-8'
                    html = payload.decode(charset, errors='replace')
                    body_text = re.sub(r'<[^>]+>', ' ', html)
                    body_text = re.sub(r'\s+', ' ', body_text).strip()
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            charset = msg.get_content_charset() or 'utf-8'
            body_text = payload.decode(charset, errors='replace')

    if body_text and len(body_text) > 20:
        text_parts.insert(0, f'邮件正文:\n{body_text}')

    return text_parts, image_parts, pre_parsed_items


def _extract_email_address(from_header: str) -> str:
    if not from_header:
        return ''
    m = re.search(r'<([^>]+@[^>]+)>', from_header)
    if m:
        return m.group(1).strip()
    m = re.search(r'[\w.+-]+@[\w.-]+\.\w+', from_header)
    if m:
        return m.group(0).strip()
    return ''


def _find_contact_email(requester_name: str) -> Optional[str]:
    if not requester_name:
        return None
    raw = requester_name.strip()

    name_only = re.sub(r'\s*<[^>]+>\s*$', '', raw).strip()

    email_in_brackets = ''
    m = re.search(r'<([^>]+@[^>]+)>', raw)
    if m:
        email_in_brackets = m.group(1).strip().lower()

    try:
        contacts = list_unified_contacts()
        for c in contacts:
            email_addr = (c.get('email') or '').strip()
            if not email_addr:
                continue
            if c.get('name_china') and c['name_china'].strip() in (raw, name_only):
                return email_addr
            if email_in_brackets and email_addr.lower() == email_in_brackets:
                return email_addr
            if c.get('username') and c['username'].strip().lower() == raw.lower():
                return email_addr
            if c.get('nickname') and c['nickname'].strip() in (raw, name_only):
                return email_addr
            if email_in_brackets and c.get('name_china') and name_only == c['name_china'].strip():
                return email_addr
    except Exception as exc:
        _log(f'Failed to lookup contact for "{requester_name}": {exc}')
    return None


def _loose_match_subject(subject: str) -> Optional[int]:
    from backend.repositories.inquiry_repository import _get_conn, INQUIRY_TABLE
    clean = _strip_reply_prefix(subject)
    clean = re.sub(r'\s*\(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}(:\d{2})?\)\s*$', '', clean).strip()
    short = re.sub(r'\s*[-–—]\s*\d+项物料待询价.*$', '', clean).strip()
    conn = _get_conn()
    try:
        for candidate in [short, clean]:
            if not candidate:
                continue
            rows = conn.execute(
                f"SELECT * FROM {INQUIRY_TABLE} WHERE status='sent' AND email_subject LIKE ? ORDER BY created_at DESC LIMIT 5",
                (f'%{candidate}%',),
            ).fetchall()
            for row in rows:
                return row['id']
        rows = conn.execute(
            f"SELECT * FROM {INQUIRY_TABLE} WHERE status='sent' ORDER BY created_at DESC"
        ).fetchall()
        for row in rows:
            db_subj = str(row['email_subject'] or '').strip()
            if db_subj and (db_subj in clean or clean in db_subj):
                return row['id']
        return None
    finally:
        conn.close()


def _forward_reply_email(
    original_subject: str,
    to_email: str,
    text_parts: List[str],
    image_parts: List[Tuple[str, str]],
    parsed_items: List[Dict[str, Any]],
):
    from backend.services.email_service import forward_inquiry_reply
    try:
        forward_inquiry_reply(
            original_subject=original_subject,
            to_email=to_email,
            text_parts=text_parts,
            image_parts=image_parts,
            parsed_items=parsed_items,
        )
    except Exception as exc:
        _log(f'Forward email failed to {to_email}: {exc}')


def _process_one_email(
    imap: imaplib.IMAP4_SSL,
    msg_id: bytes,
    message_id: str,
    scanned: int,
    parsed_count: int,
    forwarded: int,
) -> Tuple[int, int, int]:
    _, msg_data = imap.fetch(msg_id, '(BODY.PEEK[])')
    if not msg_data or not msg_data[0]:
        return scanned, parsed_count, forwarded

    raw_email = msg_data[0][1] if isinstance(msg_data[0], tuple) else msg_data[0]
    msg = email_lib.message_from_bytes(raw_email)

    subject = _decode_header_value(msg.get('Subject', ''))
    _log(f'Processing email: {subject}')

    record = find_record_by_subject(subject)
    scanned += 1
    is_external = not record
    clean_subject = _strip_reply_prefix(subject)
    is_our_inquiry = '【inquiry询价】' in clean_subject

    from_email = _decode_header_value(msg.get('From', ''))
    _log(f'Email from: {from_email}')

    if is_external:
        _log(f'No matching inquiry record for: {subject}, creating external record')

    requester = ''
    record_id = None
    if record:
        requester = record.get('inquiry_requester', '')
        record_id = record.get('id')

    text_parts, image_parts, pre_parsed_items = _extract_attachments(msg)

    if not text_parts and not image_parts and not pre_parsed_items:
        _log(f'No extractable content in email: {subject}')
        return scanned, parsed_count, forwarded

    # ── Step 1: Forward FIRST (before parsing) ──
    # Only forward if the subject contains 【inquiry询价】 (our outgoing inquiry).
    # CC'd emails are just for record-keeping / price caching, not forwarding.
    # Forward to the original requester (业务), never to the email sender.
    forwarded_to = ''
    if is_our_inquiry and requester:
        contact_email = _find_contact_email(requester)
        if contact_email:
            try:
                _forward_reply_email(subject, contact_email, text_parts, image_parts, [])
                forwarded_to = contact_email
                forwarded += 1
                _log(f'Forwarded reply to requester {contact_email}')
            except Exception as exc:
                _log(f'Forward failed: {exc}')
        else:
            _log(f'No contact email found for requester: {requester}')
    elif is_our_inquiry and not requester:
        _log(f'Our inquiry but no requester on record, cannot forward: {subject}')
    else:
        _log(f'Not our inquiry email, skipping forward (price cache only): {subject}')

    # ── Step 2: Parse / extract ──
    items = []
    parse_error = False
    if pre_parsed_items:
        items = pre_parsed_items
        _log(f'Directly extracted {len(items)} items from Excel in: {subject}')
    else:
        try:
            if text_parts and image_parts:
                items = extract_quotation_mixed('\n\n'.join(text_parts), image_parts)
            elif image_parts:
                items = extract_quotation_from_images(image_parts)
            else:
                text_input = '\n\n'.join(text_parts)
                _log(f'Text extraction input length: {len(text_input)} chars')
                items = extract_quotation_from_text(text_input)
                if not items and len(text_input) > 6000:
                    truncated = text_input[:6000]
                    _log(f'Retrying with truncated input ({len(truncated)} chars)')
                    items = extract_quotation_from_text(truncated + '\n\n(内容已截断)')
            _log(f'LLM extracted {len(items)} items from: {subject}')
        except Exception as exc:
            _log(f'KIMI2 extraction failed: {exc}')
            parse_error = True
            items = []

    if not items and not parse_error:
        _log(f'KIMI2 returned empty items for: {subject}')
        parse_error = True

    # ── Step 3: Cache prices (only if items parsed) ──
    import json as _json
    reply_json_str = ''
    if items:
        reply_json_str = _json.dumps(items, ensure_ascii=False)

        for item in items:
            item['source_email'] = subject
            item['source_record_id'] = record_id
            if not item.get('inquirer'):
                item['inquirer'] = requester

        if not record_id:
            try:
                from backend.repositories.inquiry_repository import insert_inquiry_record
                record_id = insert_inquiry_record(
                    project_name=f'[外部邮件] {subject[:60]}',
                    bom_filename='',
                    inquiry_requester=from_email[:80],
                    material_count=len(items),
                    materials=[],
                    email_subject=subject,
                    remark='auto_external',
                )
                for item in items:
                    item['source_record_id'] = record_id
                _log(f'Created external record #{record_id} for: {subject}')
            except Exception as exc:
                _log(f'Failed to create external record: {exc}')
                record_id = None

        try:
            upsert_price_cache(items)
            _log(f'Cached {len(items)} price items from email: {subject}')
        except Exception as exc:
            _log(f'Cache write failed: {exc}')

        if is_external and not record_id:
            try:
                loose_id = _loose_match_subject(subject)
                if loose_id:
                    backfill_orphan_cache(subject, items, loose_id)
                    _log(f'Backfilled orphan cache -> record #{loose_id} for: {subject}')
            except Exception as exc:
                _log(f'Backfill failed: {exc}')

        parsed_count += 1

    # ── Step 4: Update record status ──
    if record_id:
        if forwarded_to and parse_error:
            status = 'forwarded_parse_failed'
        elif is_external and not forwarded_to:
            status = 'parsed_external'
        elif forwarded_to:
            status = 'parsed_forwarded'
        elif items:
            status = 'parsed'
        else:
            status = 'parse_failed'
        update_inquiry_reply(
            record_id,
            reply_json=reply_json_str,
            status=status,
            forwarded_to=forwarded_to,
        )

    _mark_processed(message_id, subject)
    try:
        imap.store(msg_id, '+FLAGS', '\\Seen')
    except Exception:
        pass

    return scanned, parsed_count, forwarded


def _extract_images_from_xlsx_attachment(payload: bytes, filename: str) -> List[Dict[str, str]]:
    images = []
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ('.xlsx',):
        if ext == '.xls':
            _log(f'Skipping .xls embedded image extraction (not supported): {filename}')
        return images
    try:
        import openpyxl
        import io as _io
        buf = _io.BytesIO(payload)
        wb = openpyxl.load_workbook(buf)
        for ws in wb.worksheets:
            code_col = _detect_code_column(ws)

            row_image_map = {}
            for img in ws._images:
                img_anchor = img.anchor
                if hasattr(img_anchor, '_from'):
                    img_row = img_anchor._from.row + 1
                elif hasattr(img_anchor, 'row'):
                    img_row = img_anchor.row + 1
                else:
                    continue
                row_image_map.setdefault(img_row, []).append(img)

            _log(f'Excel sheet "{ws.title}": {len(row_image_map)} rows with images, code_col={code_col}')

            for row_idx in range(2, ws.max_row + 1):
                code_val = ''
                if code_col:
                    code_val = str(ws.cell(row=row_idx, column=code_col).value or '').strip()

                row_imgs = row_image_map.get(row_idx, [])
                for img in row_imgs:
                    try:
                        img_data = img._data()
                        if hasattr(img_data, 'read'):
                            img_bytes = img_data.read()
                        elif isinstance(img_data, bytes):
                            img_bytes = img_data
                        else:
                            continue
                        if not img_bytes:
                            continue

                        from backend.image.processor import guess_image_extension
                        img_ext = guess_image_extension(img_bytes, 'png')
                        b64 = base64.b64encode(img_bytes).decode('ascii')
                        images.append({
                            'filename': f'row{row_idx}_{code_val}.{img_ext}' if code_val else f'row{row_idx}.{img_ext}',
                            'material_code': code_val,
                            'base64': b64,
                            'ext': img_ext,
                        })
                    except Exception:
                        continue

        wb.close()
        _log(f'Extracted {len(images)} images from xlsx attachment: {filename}')
    except Exception as exc:
        _log(f'Failed to extract images from xlsx {filename}: {exc}')
    return images


def _extract_image_attachments_from_email(msg: email_lib.message.Message) -> List[Dict[str, str]]:
    images = []
    for part in msg.walk():
        content_type = part.get_content_type() or ''
        content_disposition = str(part.get('Content-Disposition') or '')
        filename = part.get_filename() or ''

        if filename:
            filename = _decode_header_value(filename)

        if 'attachment' in content_disposition or filename:
            if not filename:
                continue
            payload = part.get_payload(decode=True)
            if not payload:
                continue

            ext = os.path.splitext(filename)[1].lower()
            if ext in ('.png', '.jpg', '.jpeg', '.bmp', '.webp'):
                b64 = base64.b64encode(payload).decode('ascii')
                code_from_filename = _extract_code_from_filename(filename)
                images.append({
                    'filename': filename,
                    'material_code': code_from_filename,
                    'base64': b64,
                    'ext': ext.lstrip('.'),
                })
            elif ext in ('.xlsx', '.xls'):
                xlsx_images = _extract_images_from_xlsx_attachment(payload, filename)
                images.extend(xlsx_images)
        elif content_type.startswith('image/'):
            payload = part.get_payload(decode=True)
            if payload:
                cid = str(part.get('Content-ID') or '').strip('<>')
                if cid or 'inline' in content_disposition:
                    ext = content_type.split('/')[-1]
                    if ext == 'jpeg':
                        ext = 'jpg'
                    b64 = base64.b64encode(payload).decode('ascii')
                    images.append({
                        'filename': cid or f'inline.{ext}',
                        'material_code': '',
                        'base64': b64,
                        'ext': ext,
                    })
    return images


def _extract_code_from_filename(filename: str) -> str:
    from backend.core.shared.image_utils import extract_code_from_filename as _rich_extract
    code = _rich_extract(filename)
    if code:
        return code
    name = os.path.splitext(filename)[0].strip()
    if not name:
        return ''
    patterns = [
        r'^([A-Za-z]{1,4}[-][A-Za-z0-9][-A-Za-z0-9]*)',
        r'^([A-Za-z]{1,4}\d+[-][A-Za-z0-9][-A-Za-z0-9]*)',
    ]
    for pattern in patterns:
        m = re.match(pattern, name)
        if m:
            return m.group(1).strip()
    parts = re.split(r'[_\s\-]+', name)
    for part in parts:
        if re.match(r'^[A-Za-z]{1,4}[-]', part):
            return part.strip()
    return ''


def _extract_image_code_pairs_from_excel(msg: email_lib.message.Message) -> dict:
    result = {}
    for part in msg.walk():
        content_disposition = str(part.get('Content-Disposition') or '')
        filename = part.get_filename() or ''
        if filename:
            filename = _decode_header_value(filename)
        ext = os.path.splitext(filename)[1].lower()
        if ext not in ('.xlsx',):
            if ext == '.xls':
                _log(f'Skipping .xls image extraction (not supported by openpyxl): {filename}')
            continue
        payload = part.get_payload(decode=True)
        if not payload:
            continue
        try:
            import openpyxl
            buf = BytesIO(payload)
            wb = openpyxl.load_workbook(buf)
            for ws in wb.worksheets:
                code_col = _detect_code_column(ws)
                if code_col is None:
                    _log(f'No code column found in sheet "{ws.title}" of {filename}')
                    continue

                row_image_map = {}
                for img in ws._images:
                    img_anchor = img.anchor
                    if hasattr(img_anchor, '_from'):
                        img_row = img_anchor._from.row + 1
                    elif hasattr(img_anchor, 'row'):
                        img_row = img_anchor.row + 1
                    else:
                        continue
                    row_image_map.setdefault(img_row, []).append(img)

                for row_idx in range(2, ws.max_row + 1):
                    code_val = str(ws.cell(row=row_idx, column=code_col).value or '').strip()
                    if not code_val or code_val in ('nan', 'None'):
                        continue

                    row_imgs = row_image_map.get(row_idx, [])
                    for img in row_imgs:
                        try:
                            img_data = img._data()
                            if hasattr(img_data, 'read'):
                                img_bytes = img_data.read()
                            elif isinstance(img_data, bytes):
                                img_bytes = img_data
                            else:
                                continue
                            if not img_bytes:
                                continue

                            from backend.image.processor import guess_image_extension
                            img_ext = guess_image_extension(img_bytes, 'png')
                            b64 = base64.b64encode(img_bytes).decode('ascii')
                            mime_type = f'image/{img_ext}' if img_ext != 'jpg' else 'image/jpeg'
                            data_url = f'data:{mime_type};base64,{b64}'
                            if code_val not in result:
                                result[code_val] = data_url
                        except Exception:
                            continue

            wb.close()
            _log(f'Extracted {len(result)} code-image pairs from Excel: {filename}')
        except Exception as exc:
            _log(f'Failed to extract images from Excel attachment {filename}: {exc}')
    return result


def _detect_code_column(ws) -> Optional[int]:
    _CODE_COL_KEYWORDS = ['对应编码', '物料编码', '编码', '品番', 'code', '部品番号']
    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col_idx).value or '').strip().lower()
        if not header:
            continue
        for kw in _CODE_COL_KEYWORDS:
            if kw.lower() in header:
                return col_idx
    for col_idx in range(1, min(ws.max_column + 1, 10)):
        for row_idx in range(2, min(ws.max_row + 1, 6)):
            val = str(ws.cell(row=row_idx, column=col_idx).value or '').strip()
            if val and re.match(r'^[A-Za-z]{1,4}[-][A-Za-z0-9]', val):
                return col_idx
    return None


def _decode_image_base64_safe(value):
    from backend.image.processor import decode_image_base64
    return decode_image_base64(value)


def _match_images_to_codes(images: List[Dict], codes: List[str]) -> Dict[str, str]:
    from backend.utils.converters import normalize_lookup_code as _norm

    code_image_map = {}

    if not images:
        return code_image_map

    unmatched_images = list(images)

    if codes:
        normalized_codes = [(code, _norm(code)) for code in codes if code]
        normalized_codes = [(c, n) for c, n in normalized_codes if n]

        still_unmatched = []
        for img in unmatched_images:
            matched_code = ''
            img_fn_upper = img.get('filename', '').upper()

            if img.get('material_code'):
                img_code_norm = _norm(img['material_code'])
                for code, code_norm in normalized_codes:
                    if img_code_norm == code_norm:
                        matched_code = code
                        break
                if not matched_code:
                    for code, code_norm in normalized_codes:
                        if code_norm and code_norm in img_fn_upper:
                            matched_code = code
                            break

            if not matched_code:
                for code, code_norm in normalized_codes:
                    if code_norm and code_norm in img_fn_upper:
                        matched_code = code
                        break

            if not matched_code:
                fn_code = _extract_code_from_filename(img.get('filename', ''))
                if fn_code:
                    fn_code_norm = _norm(fn_code)
                    for code, code_norm in normalized_codes:
                        if fn_code_norm == code_norm:
                            matched_code = code
                            break

            if matched_code:
                b64_data = img['base64']
                img_ext = img.get('ext', 'png')
                mime_type = f'image/{img_ext}' if img_ext != 'jpg' else 'image/jpeg'
                data_url = f'data:{mime_type};base64,{b64_data}'
                code_image_map[matched_code] = data_url
            else:
                still_unmatched.append(img)

        unmatched_images = still_unmatched

    else:
        for img in unmatched_images:
            code = img.get('material_code') or _extract_code_from_filename(img.get('filename', ''))
            if code:
                b64_data = img['base64']
                img_ext = img.get('ext', 'png')
                mime_type = f'image/{img_ext}' if img_ext != 'jpg' else 'image/jpeg'
                data_url = f'data:{mime_type};base64,{b64_data}'
                code_image_map[code] = data_url
        unmatched_images = []

    if unmatched_images and codes:
        used_codes = set(_norm(c) for c in code_image_map.keys())
        remaining_codes = [(c, _norm(c)) for c in codes if _norm(c) not in used_codes]
        for i, img in enumerate(unmatched_images):
            if i < len(remaining_codes):
                code, _ = remaining_codes[i]
                b64_data = img['base64']
                img_ext = img.get('ext', 'png')
                mime_type = f'image/{img_ext}' if img_ext != 'jpg' else 'image/jpeg'
                data_url = f'data:{mime_type};base64,{b64_data}'
                code_image_map[code] = data_url
                _log(f'Positional fallback: image {i+1} -> code {code}')

    return code_image_map


def _process_image_reply(
    imap: imaplib.IMAP4_SSL,
    msg_id: bytes,
    message_id: str,
    scanned: int,
) -> Tuple[int, int]:
    from backend.repositories.image_inquiry_repository import (
        ensure_image_inquiry_schema,
        find_image_inquiry_by_subject,
        insert_image_inquiry_record,
        update_image_inquiry_status,
        update_image_inquiry_reply,
    )
    from backend.repositories.material_repository import bulk_update_aluminum_images

    _, msg_data = imap.fetch(msg_id, '(BODY.PEEK[])')
    if not msg_data or not msg_data[0]:
        return scanned, 0

    raw_email = msg_data[0][1] if isinstance(msg_data[0], tuple) else msg_data[0]
    msg = email_lib.message_from_bytes(raw_email)

    subject = _decode_header_value(msg.get('Subject', ''))
    from_email = _decode_header_value(msg.get('From', ''))
    _log(f'Processing image inquiry reply: {subject} from {from_email}')

    scanned += 1

    record = find_image_inquiry_by_subject(subject)
    record_id = None
    codes = []
    if record:
        record_id = record.get('id')
        try:
            codes = json.loads(record.get('codes_json', '[]'))
        except (json.JSONDecodeError, TypeError):
            codes = []

    if not record_id:
        try:
            record_id = insert_image_inquiry_record(
                project_name=f'[外部邮件] {subject[:60]}',
                sender_name=from_email[:80],
                designer_email=from_email[:80],
                code_count=0,
                codes=[],
                email_subject=subject,
                remark='auto_external',
            )
            _log(f'Created external image inquiry record #{record_id} for: {subject}')
        except Exception as exc:
            _log(f'Failed to create external image inquiry record: {exc}')

    if record_id:
        update_image_inquiry_status(record_id, status='received')
    _log(f'Image inquiry #{record_id} marked as received')

    images = _extract_image_attachments_from_email(msg)
    _log(f'Extracted {len(images)} image attachments from email')

    code_image_map = _match_images_to_codes(images, codes)
    _log(f'Filename-based matching: {len(code_image_map)} codes matched from {len(images)} images')

    if not code_image_map:
        excel_items = _extract_image_code_pairs_from_excel(msg)
        if excel_items:
            code_image_map = excel_items
            _log(f'Excel-based matching: {len(code_image_map)} codes matched')

    if not code_image_map:
        error_detail = json.dumps({
            'error': 'no_matching_images',
            'images_found': len(images),
            'codes_expected': len(codes),
            'image_filenames': [img.get('filename', '') for img in images[:20]],
            'codes_sample': [str(c) for c in codes[:20]],
        }, ensure_ascii=False)
        _log(f'No matching images found in image inquiry reply: {subject}')
        _log(f'  images_found={len(images)}, codes_expected={len(codes)}')
        if record_id:
            update_image_inquiry_status(record_id, status='parse_failed')
            try:
                conn = _get_db_conn()
                conn.execute(
                    f"UPDATE ks_image_inquiry_records SET reply_json = ? WHERE id = ?",
                    (error_detail, record_id),
                )
                conn.commit()
                conn.close()
            except Exception:
                pass
        try:
            from backend.repositories.image_inquiry_repository import update_item_by_code
            for code in codes:
                update_item_by_code(code, status='failed')
        except Exception:
            pass
        return scanned, 0

    images_json = json.dumps(images, ensure_ascii=False)
    if record_id:
        update_image_inquiry_reply(record_id, images_json=images_json, status='parsed')
    _log(f'Saved {len(images)} images to image inquiry record #{record_id}')
    _log(f'Image inquiry #{record_id} matched {len(code_image_map)} codes')

    image_updates = [
        {'code': code, 'image_base64': data_url}
        for code, data_url in code_image_map.items()
    ]

    updated_count = 0
    db_error = ''
    try:
        updated_count = bulk_update_aluminum_images(image_updates)
        _log(f'Updated {updated_count} images in database from image inquiry reply: {subject}')
    except Exception as exc:
        db_error = str(exc)
        _log(f'Failed to update images in database: {exc}')

    reply_data = [
        {'code': code, 'updated': code_image_map[code] is not None}
        for code in code_image_map
    ]

    if record_id:
        update_image_inquiry_status(
            record_id,
            status='db_updated',
            reply_json=json.dumps(reply_data, ensure_ascii=False),
            images_updated=updated_count,
        )

    try:
        from backend.repositories.image_inquiry_repository import update_item_by_code
        for code in code_image_map:
            data_url = code_image_map[code]
            update_item_by_code(code, status='db_updated', image_base64=data_url)
        _log(f'Updated {len(code_image_map)} items in ks_image_inquiry_items')
    except Exception as exc:
        _log(f'Failed to update image inquiry items: {exc}')

    _mark_processed(message_id, subject)
    try:
        imap.store(msg_id, '+FLAGS', '\\Seen')
    except Exception:
        pass

    return scanned, 1


def _repair_orphan_cache():
    from backend.repositories.inquiry_repository import (
        _get_conn, INQUIRY_TABLE, PRICE_CACHE_TABLE,
        insert_inquiry_record, update_inquiry_reply,
    )
    conn = _get_conn()
    try:
        orphans = conn.execute(
            f"SELECT DISTINCT source_email FROM {PRICE_CACHE_TABLE} WHERE source_record_id IS NULL"
        ).fetchall()
        if not orphans:
            return
        _log(f'Found {len(orphans)} orphan cache groups, repairing...')
        for row in orphans:
            source_email = row['source_email']
            cache_rows = conn.execute(
                f"SELECT * FROM {PRICE_CACHE_TABLE} WHERE source_email = ? AND source_record_id IS NULL",
                (source_email,),
            ).fetchall()
            if not cache_rows:
                continue
            try:
                reply_data = []
                for ci in cache_rows:
                    reply_data.append({
                        'material_code': ci['material_code'],
                        'name': ci['name'],
                        'spec': ci['spec'],
                        'quantity': ci['quantity'],
                        'unit_price': ci['unit_price'],
                        'unit': ci['unit'],
                        'valid_until': ci['valid_until'],
                        'discount': ci['discount'],
                        'inquirer': ci['inquirer'] or '',
                        'source_email': source_email,
                    })
                record_id = insert_inquiry_record(
                    project_name=f'[外部邮件] {source_email[:60]}',
                    bom_filename='',
                    inquiry_requester=str(cache_rows[0]['inquirer'] or '')[:80],
                    material_count=len(reply_data),
                    materials=[],
                    email_subject=source_email,
                    remark='auto_external',
                )
                conn.execute(
                    f"UPDATE {PRICE_CACHE_TABLE} SET source_record_id = ? WHERE source_email = ? AND source_record_id IS NULL",
                    (record_id, source_email),
                )
                now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                conn.execute(
                    f"UPDATE {INQUIRY_TABLE} SET status='parsed_external', reply_json=?, reply_received_at=? WHERE id=?",
                    (json.dumps(reply_data, ensure_ascii=False), now_str, record_id),
                )
                conn.commit()
                _log(f'Repaired orphan cache: {source_email[:50]} -> record #{record_id} ({len(reply_data)} items)')
            except Exception as exc:
                _log(f'Repair failed for {source_email[:50]}: {exc}')
    finally:
        conn.close()


def check_email_replies(include_seen: bool = False) -> Dict[str, Any]:
    global _last_scan_time, _last_scan_dt

    acquired = _check_lock.acquire(blocking=False)
    if not acquired:
        _log('Another email check is already running, skipping this call')
        return {'scanned': 0, 'parsed': 0, 'forwarded': 0, 'image_parsed': 0, 'skipped': True}

    try:
        return _check_email_replies_inner(include_seen=include_seen)
    finally:
        _check_lock.release()


def _check_email_replies_inner(include_seen: bool = True) -> Dict[str, Any]:
    global _last_scan_time, _last_scan_dt
    _log(f'Starting email reply check (include_seen={include_seen})...')
    scanned = 0
    parsed_count = 0
    forwarded = 0
    image_parsed = 0

    ensure_price_cache_schema()
    _ensure_processed_table()
    _repair_orphan_cache()

    from backend.repositories.image_inquiry_repository import ensure_image_inquiry_schema
    ensure_image_inquiry_schema()

    try:
        imap = _connect_imap()
    except Exception as exc:
        _log(f'IMAP connection failed: {exc}')
        return {'scanned': 0, 'parsed': 0, 'forwarded': 0, 'image_parsed': 0, 'error': str(exc)}

    try:
        imap.select('INBOX')

        scan_lookback_days = 3
        since_date = (datetime.now() - timedelta(days=scan_lookback_days)).strftime('%d-%b-%Y')
        if include_seen:
            search_criteria = f'(SINCE {since_date})'
        else:
            search_criteria = f'(UNSEEN SINCE {since_date})'
        status, data = imap.search(None, search_criteria)
        if status != 'OK' or not data or not data[0]:
            _log(f'No emails found (criteria: {search_criteria})')
            _last_scan_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            _last_scan_dt = datetime.now()
            return {'scanned': 0, 'parsed': 0, 'forwarded': 0, 'image_parsed': 0}

        msg_ids = data[0].split()
        _log(f'Found {len(msg_ids)} emails (SINCE {since_date})')

        quotation_tasks = []
        image_tasks = []
        skipped_dupes = 0
        skipped_old = 0

        for msg_id in msg_ids:
            _, subj_data = imap.fetch(msg_id, '(BODY.PEEK[HEADER.FIELDS (SUBJECT FROM MESSAGE-ID)])')
            if subj_data and isinstance(subj_data[0], tuple):
                header_text = subj_data[0][1].decode('utf-8', errors='replace')
                decoded_subject = _decode_header_value(
                    header_text.split('Subject:', 1)[-1].strip()
                    if 'Subject:' in header_text else ''
                )
                raw_mid = ''
                if 'Message-ID:' in header_text or 'Message-Id:' in header_text:
                    for line in header_text.splitlines():
                        if line.lower().startswith('message-id:'):
                            raw_mid = line.split(':', 1)[1].strip().strip('<>')
                            break
                if _is_already_processed(raw_mid):
                    skipped_dupes += 1
                    try:
                        imap.store(msg_id, '+FLAGS', '\\Seen')
                    except Exception:
                        pass
                    continue
                email_type = _classify_email(decoded_subject)
                if email_type == 'image':
                    image_tasks.append((msg_id, raw_mid))
                elif email_type == 'quotation':
                    quotation_tasks.append((msg_id, raw_mid))

        if skipped_dupes:
            _log(f'Skipped {skipped_dupes} already-processed emails')

        _log(f'Found {len(quotation_tasks)} quotation emails, {len(image_tasks)} image inquiry emails')

        # 询价已改为网页填价，不再处理询价回复邮件；仅处理询图回复。
        if quotation_tasks:
            _log(f'Skipping {len(quotation_tasks)} quotation reply email(s) — inquiry is now web-form based')

        for msg_id, mid in image_tasks:
            try:
                scanned, img_count = _process_image_reply(
                    imap, msg_id, mid, scanned
                )
                image_parsed += img_count
            except Exception as exc:
                _log(f'Error processing image reply {msg_id}: {exc}')
                import traceback
                traceback.print_exc()

        _last_scan_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        _last_scan_dt = datetime.now()
        return {
            'scanned': scanned,
            'parsed': parsed_count,
            'forwarded': forwarded,
            'image_parsed': image_parsed,
        }
    except Exception as exc:
        _log(f'Email scan error: {exc}')
        import traceback
        traceback.print_exc()
        _last_scan_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        _last_scan_dt = datetime.now()
        return {'scanned': scanned, 'parsed': parsed_count, 'forwarded': forwarded, 'image_parsed': image_parsed, 'error': str(exc)}
    finally:
        try:
            imap.close()
            imap.logout()
        except Exception:
            pass
