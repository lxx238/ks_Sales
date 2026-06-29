import os
import re
import json
import uuid
import tempfile
from datetime import datetime

from openpyxl import Workbook

from backend.repositories.inquiry_repository import (
    insert_inquiry_items,
    list_inquiry_items,
    list_inquiry_projects,
    get_inquiry_item,
    update_inquiry_item,
    delete_inquiry_item as repo_delete_inquiry_item,
    delete_inquiry_case,
    count_inquiry_items_by_status,
    upsert_price_cache,
    cleanup_expired_records,
    CASE_LOCK_MINUTES,
    get_case_lock,
    list_case_locks,
    upsert_case_lock,
    insert_ton_price_history,
    list_ton_price_history,
    cleanup_expired_ton_history,
    get_case_meta,
    list_case_metas,
    upsert_case_meta,
    insert_attachment,
    list_attachments_batch,
    get_attachment,
    delete_attachment,
)
from backend.utils.converters import normalize_lookup_code
from backend.utils.constants import DB_UNIT_COLUMN
from backend.core.shared.weight_utils import (
    extract_length_from_spec,
    WEIGHT_BY_LENGTH_ATTRIBUTES,
)


def _current_user_name():
    try:
        from backend.services.auth_service import get_current_account as _get_acct
        _acct = _get_acct(optional=True)
        if _acct:
            _china_name = str(_acct.get('name') or '').strip()
            if _china_name:
                return _china_name
            _u = str(_acct.get('username') or '').strip()
            if _u:
                return _u
    except Exception:
        pass
    return ''


def _is_case_locked(project_name):
    """案件金额保存后 7 天内可改；超过则锁定。返回 (locked: bool, lock: dict|None, minutes_left: int)。"""
    lock = get_case_lock(project_name)
    if not lock or not lock.get('last_priced_at'):
        return False, lock, None
    try:
        priced_at = datetime.strptime(lock['last_priced_at'], '%Y-%m-%d %H:%M:%S')
    except (ValueError, TypeError):
        return False, lock, None
    elapsed_min = (datetime.now() - priced_at).total_seconds() / 60.0
    if elapsed_min > CASE_LOCK_MINUTES:
        return True, lock, 0
    return False, lock, int(CASE_LOCK_MINUTES - elapsed_min)

def _parse_float(value, default=0.0):
    try:
        return float(str(value).replace(',', '').strip())
    except (TypeError, ValueError):
        return default


# ----------------------------------------------------------------------
# 吨价换算 —— 完全对齐「临时价格设置」(temp_price_settings) 的模型
#   * 吨价按「吨价类型 + 长度档 + 吨位档」录入（基准/内部吨价，币种无关）
#   * 单价 = 单重 × 吨价
#       - 人民币(无汇率): 单重 × 吨价
#       - 美元/欧元/人民币外汇: 单重 × 吨价 ÷ 汇率 ÷ 点数
#   * 吨价与汇率/点数均存入 temp_price_settings（与临时价格设置同一库）
# ----------------------------------------------------------------------
_TEMP_SETTINGS_TABLE = 'temp_price_settings'
_LEN_TIER_KEY = {'0-1': '01', '1-3': '13', '3+': '3'}        # 长度档 → 设置键
_WEIGHT_TIER_DISPLAY = {'05': '0-5', '550': '5-50', '50999': '50+'}  # 吨位档
_RATE_CURS = ['usd', 'eur', 'rmb_fx']                          # 带汇率的币种
_RATE_KEYS = (
    [f'exchange_rate_{c}' for c in _RATE_CURS]
    + [f'points_{c}' for c in _RATE_CURS]
)

# 汇率/点数分类（不同定价属性使用不同类别的汇率/点数）
RATE_CATEGORIES = [
    {'key': 'dizhuang', 'label': '地桩', 'attrs': ('D',)},
    {'key': 'lv', 'label': '铝', 'attrs': ('M',)},
    {'key': 'lvpj', 'label': '铝配件', 'attrs': ('F', 'Q')},
    {'key': 'tie', 'label': '铁', 'attrs': ('WTX', 'WTP')},
    {'key': 'waigou', 'label': '外购件', 'attrs': ('W',)},
]
# 定价属性 → 类别键（无映射返回 None → 需手动填价）
_ATTR_TO_CAT = {a: c['key'] for c in RATE_CATEGORIES for a in c['attrs']}
# 全部类别键集合（用于校验前端手动选择的 price_category）
_CAT_KEYS = {c['key'] for c in RATE_CATEGORIES}
# 各类别的汇率/点数键（exchange_rate_{cat}_{cur} / points_{cat}_{cur}）
_RATE_CAT_KEYS = []
for _c in RATE_CATEGORIES:
    for _cur in _RATE_CURS:
        _RATE_CAT_KEYS.append(f'exchange_rate_{_c["key"]}_{_cur}')
        _RATE_CAT_KEYS.append(f'points_{_c["key"]}_{_cur}')

# 铝配件额外系数（计算时 base × 系数，默认 1.03，前端可修改）
_LVPJ_COEFF_KEY = 'lvpj_coefficient'
_LVPJ_COEFF_DEFAULT = 1.03


def _category_for_attr(pricing_attr):
    """定价属性 → 汇率类别键；无映射返回 None（需手动填价）。"""
    return _ATTR_TO_CAT.get(str(pricing_attr or '').strip().upper())

# 计价单位为「米」的标识（同前端 _isMeterUnit）
_METER_UNITS = {'米', 'm', 'M'}


def _spec_factor(unit, spec):
    """每件价格/重量系数：米单位物料单重为 kg/m，每件 = 单重 × 规格(mm)/1000 → 系数 = 规格(mm)/1000；
    非米单位（配件等）单重即每件 → 系数 = 1。规格解析失败时按 1 计。
    与报价引擎 weight_utils / 临时价格设置口径一致。"""
    if str(unit or '').strip() not in _METER_UNITS:
        return 1.0
    length_mm = extract_length_from_spec(str(spec or ''))
    if length_mm and length_mm > 0:
        return length_mm / 1000.0
    return 1.0


def _ton_type_from_code(code, length_tier=None):
    code = str(code or '').strip()
    if not code:
        return ''
    if code.startswith('FEPJ'):
        return code.replace('-', '_')  # 配件：下划线以对齐临时价格设置
    if not str(length_tier or '').strip():
        return code  # 非米单位：完整编码
    idx = code.find('-')  # 米单位：取 '-' 前缀
    return code[:idx] if idx > 0 else code


def _weight_tier_key(total_kg):
    ton = (_parse_float(total_kg, 0)) / 1000.0
    if ton <= 5:
        return '05'
    if ton <= 50:
        return '550'
    return '50999'


def _weight_tier_display(total_kg):
    return _WEIGHT_TIER_DISPLAY.get(_weight_tier_key(total_kg), '50+')


# 包装类型（吨价按包装区分：简易包装/铁托），与临时价格设置一致
PACK_TYPES = [{'key': 'jybz', 'label': '简易包装'}, {'key': 'tietuo', 'label': '铁托'}]
PACK_KEYS = [p['key'] for p in PACK_TYPES]
DEFAULT_PACK = 'jybz'
_VALID_PACKS = set(PACK_KEYS)


def _normalize_pack(pack):
    pk = str(pack or '').strip().lower()
    return pk if pk in _VALID_PACKS else DEFAULT_PACK


def _ton_setting_key(ton_type, length_tier, total_kg, pack=None):
    """内部吨价设置键（与临时价格设置 _build_ton_keys 一致，含包装后缀）。

    FEPJ 配件无长度档: ton_{type}_int_{吨位档}_{pack}
    其余: ton_{type}_int_{长度档}_{吨位档}_{pack}
    total_kg 为该组总重量(kg)，内部折算为吨位档。
    pack 默认简易包装(jybz)，可选铁托(tietuo)。
    """
    if not ton_type:
        return ''
    pk = _normalize_pack(pack)
    wk = _weight_tier_key(total_kg)
    lt = str(length_tier or '').strip()
    if not lt:
        return f'ton_{ton_type}_int_{wk}_{pk}'  # 无长度档（非米单位/配件）
    len_key = _LEN_TIER_KEY.get(lt, '3')
    return f'ton_{ton_type}_int_{len_key}_{wk}_{pk}'


def _ensure_setting_cols(connection, keys):
    cols = {r[1] for r in connection.execute(f'PRAGMA table_info({_TEMP_SETTINGS_TABLE})').fetchall()}
    if not cols:
        cols_def = ', '.join(f'"{k}" REAL' for k in keys)
        connection.execute(
            f'CREATE TABLE IF NOT EXISTS {_TEMP_SETTINGS_TABLE} '
            f'(id INTEGER PRIMARY KEY DEFAULT 1, {cols_def}, updated_at TEXT)'
        )
    else:
        for k in keys:
            if k and k not in cols:
                connection.execute(f'ALTER TABLE {_TEMP_SETTINGS_TABLE} ADD COLUMN "{k}" REAL')
    connection.execute(f'INSERT OR IGNORE INTO {_TEMP_SETTINGS_TABLE} (id) VALUES (1)')
    connection.commit()


def _read_temp_settings(keys):
    if not keys:
        return {}
    from backend.config.settings import get_db_connection
    conn = get_db_connection()
    try:
        _ensure_setting_cols(conn, list(keys))
        row = conn.execute(f'SELECT * FROM {_TEMP_SETTINGS_TABLE} WHERE id=1').fetchone()
        d = dict(row) if row else {}
    finally:
        conn.close()
    result = {}
    for k in keys:
        try:
            result[k] = float(d.get(k)) if d.get(k) not in (None, '') else None
        except (TypeError, ValueError):
            result[k] = None
    return result


def _write_temp_settings(updates):
    if not updates:
        return
    from backend.config.settings import get_db_connection
    keys = list(updates.keys())
    conn = get_db_connection()
    try:
        _ensure_setting_cols(conn, keys)
        set_parts = [f'"{k}"=?' for k in keys]
        vals = []
        for k in keys:
            v = updates[k]
            try:
                vals.append(float(v) if str(v).strip() not in ('', 'None') else 0.0)
            except (TypeError, ValueError):
                vals.append(0.0)
        vals.append(1)
        conn.execute(
            f'UPDATE {_TEMP_SETTINGS_TABLE} SET {", ".join(set_parts)}, '
            f"updated_at=datetime('now','localtime') WHERE id=?",
            vals,
        )
        conn.commit()
    finally:
        conn.close()


def get_case_ton_settings(groups, pack=None):
    """读取各组对应的内部吨价 + 当前汇率/点数（来自 temp_price_settings）。

    groups: [{'material_code', 'length_tier', 'total_kg', 'pack'(可选)}, ...]
    pack: 案件级包装类型（jybz/tietuo），覆盖各组；默认简易包装。
    返回: {'ton_prices': [{material_code,length_tier,weight_tier,setting_key,ton_price}, ...],
           'rates': {cur: {'rate':, 'points':}, ...}}
    """
    groups = groups or []
    case_pack = _normalize_pack(pack)
    ton_keys = []
    meta = []
    for g in groups:
        code = str(g.get('material_code') or '').strip()
        length_tier = str(g.get('length_tier') or '').strip()
        total_kg = _parse_float(g.get('total_kg'), 0)
        g_pack = g.get('pack') or case_pack
        ton_type = _ton_type_from_code(code, length_tier)
        wkey = _weight_tier_key(total_kg)
        skey = _ton_setting_key(ton_type, length_tier, total_kg, g_pack)
        meta.append({
            'material_code': code, 'length_tier': length_tier,
            'weight_tier': _WEIGHT_TIER_DISPLAY.get(wkey, '50+'),
            'setting_key': skey,
        })
        if skey and skey not in ton_keys:
            ton_keys.append(skey)

    settings = _read_temp_settings(ton_keys + _RATE_CAT_KEYS + [_LVPJ_COEFF_KEY])
    ton_prices = []
    for m in meta:
        ton_prices.append({**m, 'ton_price': settings.get(m['setting_key'])})
    # 返回各类别汇率/点数（{cat_key: {cur: {rate, points}}}
    rates = {}
    for cat in RATE_CATEGORIES:
        ck = cat['key']
        rates[ck] = {}
        for cur in _RATE_CURS:
            rates[ck][cur] = {
                'rate': settings.get(f'exchange_rate_{ck}_{cur}'),
                'points': settings.get(f'points_{ck}_{cur}'),
            }
    _coeff = settings.get(_LVPJ_COEFF_KEY)
    lvpj_coeff = float(_coeff) if _coeff not in (None, '') and float(_coeff) > 0 else _LVPJ_COEFF_DEFAULT
    return {
        'ton_prices': ton_prices, 'rates': rates,
        'rate_categories': RATE_CATEGORIES, 'lvpj_coefficient': lvpj_coeff,
        'pack': case_pack, 'pack_types': PACK_TYPES,
    }


def _fetch_material_attrs(material_codes):
    """查询常规报价物料库(aluminum_pricing)，返回 {norm: {'unit','unit_weight','attr'}}。

    单重取值顺序与「临时价格设置」一致：单重 → 米重/km → 参考重量 → 重量
    （aluminum_pricing 的 重量 列常为"暂无数据"，真正的单重在 单重/参考重量 列）。
    """
    norms = [normalize_lookup_code(str(c or '')) for c in (material_codes or [])]
    norms = list(dict.fromkeys(n for n in norms if n))
    if not norms:
        return {}
    try:
        from backend.config.settings import get_db_connection
        from backend.utils.constants import DB_TABLE_NAME
    except Exception as exc:
        print(f'[INQUIRY-ITEMS] material import failed: {exc}')
        return {}
    try:
        conn = get_db_connection()
    except Exception as exc:
        print(f'[INQUIRY-ITEMS] material db failed: {exc}')
        return {}
    try:
        existing = {r[1] for r in conn.execute(f'PRAGMA table_info({DB_TABLE_NAME})').fetchall()}
        weight_cols = [c for c in ('单重', '米重/km', '参考重量', '重量') if c in existing]
        sel_cols = [c for c in ('工程编码', DB_UNIT_COLUMN, '编码属性', '定价属性') if c in existing] + weight_cols
        if '工程编码' not in sel_cols or not weight_cols:
            return {}
        col_sql = ', '.join(f'"{c}"' for c in dict.fromkeys(sel_cols))
        placeholders = ', '.join(['?'] * len(norms))
        rows = conn.execute(
            f'SELECT {col_sql} FROM {DB_TABLE_NAME} '
            f"WHERE UPPER(REPLACE(TRIM(\"工程编码\"), ' ', '')) IN ({placeholders})",
            norms,
        ).fetchall()
    except Exception as exc:
        print(f'[INQUIRY-ITEMS] material lookup failed: {exc}')
        return {}
    finally:
        conn.close()

    attrs = {}
    for row in rows:
        d = dict(row)
        n = normalize_lookup_code(str(d.get('工程编码') or ''))
        if not n:
            continue
        uw = 0.0
        for col in ('单重', '米重/km', '参考重量', '重量'):
            raw = d.get(col)
            if raw is None or str(raw).strip() in ('', '暂无数据'):
                continue
            try:
                v = float(str(raw).replace(',', '').strip())
                if v > 0:
                    uw = v
                    break
            except (TypeError, ValueError):
                continue
        attrs[n] = {
            'unit': str(d.get(DB_UNIT_COLUMN) or '').strip(),
            'unit_weight': uw,
            'attr': str(d.get('编码属性') or '').strip().upper(),
            'pricing_attr': str(d.get('定价属性') or '').strip(),
        }
    return attrs


def _compute_item_weights(products):
    """按 (编码, 规格) 分组聚合，从铝价数据库查询单重并计算总重量。

    - 相同编码 + 相同规格 → 合并数量
    - 相同编码 + 不同规格 → 各自独立

    返回 [(code, spec, name, unit, quantity, unit_weight, total_weight, preinstall), ...]
    """
    if not products:
        return []

    grouped = {}
    for p in products:
        code = str(p.get('code') or '').strip()
        if not code:
            continue
        spec = str(p.get('spec') or '').strip()
        norm = normalize_lookup_code(code)
        qty = _parse_float(p.get('quantity'), 0.0)
        name = str(p.get('name') or '').strip()
        unit = str(p.get('unit') or '').strip()
        key = (code, spec)
        entry = grouped.setdefault(key, {
            'code': code, 'spec': spec, 'norm': norm,
            'name': name, 'unit': unit, 'quantity': 0.0,
            'bom_weight': 0.0,
            'preinstall': '',
        })
        entry['quantity'] += qty
        bw = _parse_float(p.get('weight'), 0.0)
        if bw and not entry['bom_weight']:
            entry['bom_weight'] = bw
        _pi = str(p.get('preinstall') or '').strip()
        if _pi and (not entry['preinstall'] or _pi == '非预装'):
            entry['preinstall'] = _pi

    if not grouped:
        return []

    norms = list({e['norm'] for e in grouped.values() if e['norm']})
    attrs = _fetch_material_attrs(norms)

    result = []
    for (code, spec), entry in grouped.items():
        norm = entry['norm']
        info = attrs.get(norm, {})
        db_weight = info.get('unit_weight') or 0.0
        # DB 查不到单重时，回退使用 BOM 自带单重
        if not db_weight:
            db_weight = entry.get('bom_weight') or 0.0
        attr = info.get('attr') or ''
        # 单重存原始值(kg/m)；长度属性物料(型材 A/F/TX)按规格折算每件实际重量
        # 每件 = 单重 × 规格(mm)/1000（与报价引擎 weight_utils 口径一致）
        unit_weight = db_weight
        eff_weight = db_weight
        if db_weight and attr in WEIGHT_BY_LENGTH_ATTRIBUTES:
            length_mm = extract_length_from_spec(spec)
            if length_mm and length_mm > 0:
                eff_weight = db_weight * length_mm / 1000.0

        unit_weight_r = round(unit_weight, 4) if unit_weight else None
        total_weight = round(eff_weight * entry['quantity'], 2) if eff_weight else None

        # 计价单位以常规报价物料库为准；BOM 自带单位为空时回退到库内单位
        resolved_unit = entry['unit'] or info.get('unit') or ''

        result.append((
            code, spec, entry['name'], resolved_unit,
            round(entry['quantity'], 4),
            unit_weight_r,
            total_weight,
            entry['preinstall'] or '预装',
        ))
    return result


def enrich_inquiry_items(items):
    """从常规报价物料库补全计价单位/单重，并按规格重算总重量并持久化。

    - 计价单位/单重：仅填充当前为空的字段（单重存原始 kg/m 值，不覆盖）
    - 总重量：长度属性物料(型材 A/F/TX)按 每件=单重×规格(mm)/1000 × 数量 重算，值变化才写入
    """
    if not items:
        return items

    def _empty(v):
        return v in (None, '', 0)

    targets = [it for it in items if it.get('id') is not None and str(it.get('material_code') or '').strip()]
    if not targets:
        return items

    attrs = _fetch_material_attrs([it.get('material_code') for it in targets])

    for it in targets:
        norm = normalize_lookup_code(str(it.get('material_code') or ''))
        info = attrs.get(norm)
        if not info:
            continue
        it['pricing_attr'] = info.get('pricing_attr') or ''
        fields = {}
        if _empty(it.get('unit')) and info.get('unit'):
            fields['unit'] = info['unit']

        db_w = info.get('unit_weight') or 0.0
        if db_w:
            attr = info.get('attr') or ''
            raw_r = round(db_w, 4)
            # 单重存原始值(kg/m)：与库内一致（纠偏历史可能折算过的值），值变化才写入
            try:
                stored_uw = float(it.get('unit_weight')) if it.get('unit_weight') not in (None, '') else None
            except (TypeError, ValueError):
                stored_uw = None
            if stored_uw is None or abs(stored_uw - raw_r) > 0.001:
                fields['unit_weight'] = raw_r
            # 每件实际重量：长度属性物料按规格折算 单重 × 规格(mm)/1000
            eff = db_w
            if attr in WEIGHT_BY_LENGTH_ATTRIBUTES:
                length_mm = extract_length_from_spec(str(it.get('spec') or ''))
                if length_mm and length_mm > 0:
                    eff = db_w * length_mm / 1000.0
            try:
                qty = float(it.get('quantity') or 0)
            except (TypeError, ValueError):
                qty = 0.0
            if qty:
                new_total = round(eff * qty, 2)
                try:
                    old_f = float(it.get('total_weight')) if it.get('total_weight') not in (None, '') else None
                except (TypeError, ValueError):
                    old_f = None
                if old_f is None or abs(old_f - new_total) > 0.01:
                    fields['total_weight'] = new_total
        if fields:
            try:
                update_inquiry_item(it['id'], **fields)
            except Exception as exc:
                print(f'[INQUIRY-ITEMS] enrich persist failed for id={it.get("id")}: {exc}')
            it.update(fields)
    return items


def submit_inquiry_items(data):
    """BOM 生成后将未匹配物料提交为询价项（不再发送邮件）。

    相同编码 + 相同规格合并数量；不同规格各自独立。
    """
    if not data:
        return {'success': False, 'message': '请求体不能为空'}, 400

    project_name = str(data.get('project_name', '')).strip()
    unmatched_products = data.get('unmatched_products') or []

    inquirer = str(data.get('inquiry_requester', '')).strip() or _current_user_name()
    business_name = str(data.get('business_name', '')).strip() or inquirer

    computed = _compute_item_weights(unmatched_products)

    items = []
    for code, spec, name, unit, quantity, unit_weight, total_weight, preinstall in computed:
        items.append({
            'project_name': project_name,
            'business_name': business_name,
            'material_code': code,
            'name': name,
            'spec': spec,
            'quantity': quantity,
            'unit_weight': unit_weight,
            'total_weight': total_weight,
            'unit': unit,
            'inquirer': inquirer,
            'preinstall': preinstall,
        })

    inserted = insert_inquiry_items(items, source='bom')

    # 保存案件备注（提交询价项时附带，供 admin 填价参考）
    remark = str(data.get('remark') or '').strip()
    if project_name and remark:
        try:
            upsert_case_meta(project_name, remark)
        except Exception as exc:
            print(f'[INQUIRY-ITEMS] save case remark failed: {exc}')

    # 异步发送钉钉提醒（不阻塞 BOM 报表响应）
    # 只要本次有询价项就提醒（即使因重复未新增，也要通知报价人员）
    if items:
        import threading

        def _run_notify():
            try:
                print(f'[INQUIRY-NOTIFY] 开始发送提交通知: items={len(items)} inserted={inserted} project={project_name!r}')
                result = notify_inquiry_submitted(project_name, inserted, items)
                print(f'[INQUIRY-NOTIFY] 提交通知结果: {result}')
            except Exception as exc:
                import traceback
                print(f'[INQUIRY-NOTIFY] 提交通知异常(被线程捕获): {exc}')
                traceback.print_exc()

        threading.Thread(target=_run_notify, daemon=True).start()
    else:
        print('[INQUIRY-NOTIFY] 本次无询价项，未发送通知')

    return {
        'success': True,
        'message': f'已提交 {inserted} 条询价项（共 {len(items)} 项，相同规格已合并）到询价填价页面',
        'submitted_count': len(items),
        'inserted_count': inserted,
        'material_count': len(items),
    }, 200


def _project_name_from_filename(filename):
    """从文件名推导案件名：去掉扩展名和常见前缀（物料询价表/询价表）。
    '物料询价表-广东世必达物流.xlsx' → '广东世必达物流'
    """
    import os
    name = os.path.splitext(os.path.basename(filename or ''))[0]
    for prefix in ('物料询价表-', '物料询价表', '询价表-', '询价表'):
        if name.startswith(prefix):
            name = name[len(prefix):]
            break
    return name.strip() or '手动导入'


# 标准询价表列索引（第3行起数据，0-based）
_INQ_COL = {
    'inquirer': 0,       # A 询价人
    'date': 1,           # B 询价日期
    'code': 2,           # C 产品编码
    'name': 3,           # D 产品名称
    'spec': 4,           # E 规格(mm)
    'quantity': 5,       # F 数量
    'unit': 6,           # G 销售单位
    'material': 7,       # H 备注/材质
    'unit_weight': 14,   # O 重量属性(kg)
    'total_weight': 15,  # P 物料总重(kg)
    'project': 18,       # S 产品归属项目
}


def import_inquiry_excel(filepath, original_filename, project_name=''):
    """手动导入标准询价表 Excel，解析后写入 ks_inquiry_items 并通知报价人员。

    Excel 格式：第1行分组标题，第2列表头，第3行起数据（与 create_inquiry_sheet 输出一致）。
    """
    from openpyxl import load_workbook

    project = str(project_name or '').strip()
    if not project:
        project = _project_name_from_filename(original_filename)

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as exc:
        return {'success': False, 'message': f'无法读取 Excel 文件: {exc}'}, 400

    ws = wb.active or wb[wb.sheetnames[0]]

    inquirer_default = _current_user_name()
    items = []
    skipped = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 3:
            continue
        code = str(row[_INQ_COL['code']] or '').strip() if len(row) > _INQ_COL['code'] else ''
        if not code:
            skipped += 1
            continue
        # 跳过合计/空行
        if '合计' in code or '小计' in code:
            skipped += 1
            continue

        qty = _parse_float(row[_INQ_COL['quantity']] if len(row) > _INQ_COL['quantity'] else 0, 0)
        uw = _parse_float(row[_INQ_COL['unit_weight']] if len(row) > _INQ_COL['unit_weight'] else 0, 0)
        tw = _parse_float(row[_INQ_COL['total_weight']] if len(row) > _INQ_COL['total_weight'] else 0, 0)
        # 如果总重量为空但单重和数量有值，则计算
        if (not tw or tw <= 0) and uw > 0 and qty > 0:
            tw = round(uw * qty, 2)

        proj_override = str(row[_INQ_COL['project']] or '').strip() if len(row) > _INQ_COL['project'] else ''

        items.append({
            'project_name': proj_override or project,
            'material_code': code,
            'name': str(row[_INQ_COL['name']] or '').strip() if len(row) > _INQ_COL['name'] else '',
            'spec': str(row[_INQ_COL['spec']] or '').strip() if len(row) > _INQ_COL['spec'] else '',
            'quantity': qty,
            'unit_weight': uw if uw > 0 else None,
            'total_weight': tw if tw > 0 else None,
            'unit': str(row[_INQ_COL['unit']] or '').strip() if len(row) > _INQ_COL['unit'] else '',
            'inquirer': str(row[_INQ_COL['inquirer']] or '').strip() if len(row) > _INQ_COL['inquirer'] else inquirer_default,
            'preinstall': '预装',
        })

    if not items:
        return {'success': False, 'message': '未解析到有效数据行（请确认格式：第3行起为数据，C列为物料编码）'}, 400

    inserted = insert_inquiry_items(items, source='manual_import')

    # 异步发送钉钉提醒
    notify_result = {'sent': False, 'reason': '未触发'}
    if items:
        import threading

        def _run_notify():
            try:
                result = notify_inquiry_submitted(project, inserted, items)
                print(f'[INQUIRY-IMPORT] 通知结果: {result}')
            except Exception as exc:
                print(f'[INQUIRY-IMPORT] 通知异常: {exc}')

        threading.Thread(target=_run_notify, daemon=True).start()

    return {
        'success': True,
        'message': f'已导入 {inserted} 条询价项（共 {len(items)} 项，跳过 {skipped} 行）到案件「{project}」',
        'inserted_count': inserted,
        'total_count': len(items),
        'skipped': skipped,
        'project_name': project,
    }, 200


def add_inquiry_item_manual(data):
    if not data:
        return {'success': False, 'message': '请求体不能为空'}, 400
    code = str(data.get('material_code') or data.get('code') or '').strip()
    if not code:
        return {'success': False, 'message': '物料编码不能为空'}, 400

    inquirer = str(data.get('inquirer', '')).strip() or _current_user_name()

    raw_uw = data.get('unit_weight')
    raw_tw = data.get('total_weight')

    item = {
        'project_name': str(data.get('project_name', '')).strip(),
        'business_name': str(data.get('business_name', '')).strip() or inquirer,
        'material_code': code,
        'name': str(data.get('name', '')).strip(),
        'spec': str(data.get('spec', '')).strip(),
        'quantity': _parse_float(data.get('quantity'), 0.0),
        'unit_weight': _parse_float(raw_uw, None) if raw_uw not in (None, '') else None,
        'total_weight': _parse_float(raw_tw, None) if raw_tw not in (None, '') else None,
        'unit': str(data.get('unit', '')).strip(),
        'inquirer': inquirer,
        'preinstall': str(data.get('preinstall') or '预装').strip(),
    }
    inserted = insert_inquiry_items([item], source='manual')
    if not inserted:
        return {'success': False, 'message': '该编码在相同项目下已存在询价项，已忽略'}, 200
    return {'success': True, 'message': '已新增询价项'}, 200


def save_inquiry_item_price(item_id, data):
    if not data:
        return {'success': False, 'message': '请求体不能为空'}, 400
    item = get_inquiry_item(item_id)
    if not item:
        return {'success': False, 'message': '询价项不存在'}, 404

    # 案件金额锁定：保存后 7 天内可改，超时禁止再次修改
    project_name = str(item.get('project_name') or '').strip()
    locked, lock, _ = _is_case_locked(project_name)
    if locked:
        locked_at = (lock or {}).get('last_priced_at') or ''
        return {
            'success': False,
            'locked': True,
            'message': f'该案件已于 {locked_at} 保存金额，超过 7 天，无法修改',
        }, 403

    def _num(key):
        raw = data.get(key)
        if raw is None or str(raw).strip() == '':
            return None
        return _parse_float(raw, None)

    usd = _num('unit_price_usd')
    cny = _num('unit_price_cny')
    eur = _num('unit_price_eur')
    rmb = _num('unit_price_rmb')
    unit = str(data.get('unit') or item.get('unit') or '').strip()
    quotation_date = str(data.get('quotation_date') or item.get('quotation_date') or '').strip()
    valid_until = str(data.get('valid_until') or item.get('valid_until') or '').strip()
    discount = str(data.get('discount') or item.get('discount') or '').strip()
    remark = str(data.get('remark') or item.get('remark') or '').strip()
    inquirer = str(data.get('inquirer') or item.get('inquirer') or _current_user_name()).strip()

    has_price = any(v not in (None, '') for v in (usd, cny, eur, rmb))
    status = 'priced' if has_price else 'pending'

    update_ok = update_inquiry_item(
        item_id,
        unit_price_usd=usd,
        unit_price_cny=cny,
        unit_price_eur=eur,
        unit_price_rmb=rmb,
        unit=unit,
        quotation_date=quotation_date,
        valid_until=valid_until,
        discount=discount,
        remark=remark,
        inquirer=inquirer,
        status=status,
    )

    # 同步到价格缓存，使后续报价自动匹配
    if has_price:
        price_val = usd or cny or eur or rmb
        try:
            upsert_price_cache([{
                'material_code': item.get('material_code', ''),
                'spec': item.get('spec', ''),
                'quantity': item.get('quantity', 0),
                'name': item.get('name', ''),
                'unit_price': price_val,
                'unit_price_usd': usd,
                'unit_price_cny': cny,
                'unit_price_eur': eur,
                'unit_price_rmb': rmb,
                'unit': unit,
                'quotation_date': quotation_date or datetime.now().strftime('%Y-%m-%d'),
                'valid_until': valid_until,
                'discount': discount,
                'inquirer': inquirer,
                'source_email': 'web_form',
                'preinstall': item.get('preinstall', '预装'),
            }])
        except Exception as exc:
            print(f'[INQUIRY-ITEMS] mirror to price cache failed: {exc}')

    # 记录案件金额锁定时间点（7 天可改窗口）
    if update_ok and has_price:
        try:
            upsert_case_lock(project_name, _current_user_name())
        except Exception as exc:
            print(f'[INQUIRY-ITEMS] upsert case lock failed: {exc}')

    return {'success': bool(update_ok), 'message': '已保存' if update_ok else '保存失败'}, 200 if update_ok else 500


def save_case_ton_prices(data):
    """按案件批量保存吨价（对齐「临时价格设置」模型）并换算每项单价、同步价格库。

    entries: [
        { material_code, length_tier, total_kg, ton_price, valid_until, item_ids: [id,...] }, ...
    ]
    * 吨价(基准/内部)写入 temp_price_settings: ton_{type}_int_{长度档}_{吨位档}
    * 单价 = 单重 × 吨价；美元/欧元/人民币外汇 再 ÷ 汇率 ÷ 点数
    """
    if not data:
        return {'success': False, 'message': '请求体不能为空'}, 400

    entries = data.get('entries') or []
    if not entries:
        return {'success': False, 'message': '没有需要保存的吨价'}, 400

    project_name = str(data.get('project_name') or '').strip()

    # 案件金额锁定：保存后 7 天内可改，超时禁止再次修改
    locked, lock, _minutes_left = _is_case_locked(project_name)
    if locked:
        locked_at = (lock or {}).get('last_priced_at') or ''
        return {
            'success': False,
            'locked': True,
            'message': f'该案件已于 {locked_at} 保存金额，超过 7 天，无法修改',
        }, 403

    case_valid = str(data.get('valid_until') or '').strip()
    quotation_date = datetime.now().strftime('%Y-%m-%d')
    case_pack = _normalize_pack(data.get('pack'))

    # 保存报价人员备注（手写备注 + 价格公式备注，公式部分不随钉钉通知外发）+ 案件级包装
    if project_name:
        try:
            upsert_case_meta(project_name, pricer_remark=str(data.get('pricer_remark') or '') if 'pricer_remark' in data else None, pack=case_pack if 'pack' in data else None)
        except Exception as exc:
            print(f'[INQUIRY-ITEMS] save case meta failed: {exc}')

    # 1) 汇总各组吨价 → 写入 temp_price_settings（与临时价格设置共用同一库）
    ton_updates = {}
    for entry in entries:
        code = str(entry.get('material_code') or '').strip()
        length_tier = str(entry.get('length_tier') or '').strip()
        ton_price = _parse_float(entry.get('ton_price'), None)
        if ton_price in (None, 0) or not code:
            continue
        total_kg = _parse_float(entry.get('total_kg'), 0)
        ton_type = _ton_type_from_code(code, length_tier)
        e_pack = entry.get('pack') or case_pack
        skey = _ton_setting_key(ton_type, length_tier, total_kg, e_pack)
        if skey:
            ton_updates[skey] = ton_price
    if ton_updates:
        try:
            _write_temp_settings(ton_updates)
        except Exception as exc:
            print(f'[INQUIRY-ITEMS] write ton settings failed: {exc}')

    # 2) 读取吨价 + 各类别汇率/点数 + 铝配件系数
    settings = _read_temp_settings(list(ton_updates.keys()) + _RATE_CAT_KEYS + [_LVPJ_COEFF_KEY])

    def _sval(key):
        v = settings.get(key)
        return float(v) if v not in (None, '') and v > 0 else 0.0

    lvpj_coeff = _sval(_LVPJ_COEFF_KEY) or _LVPJ_COEFF_DEFAULT

    # 批量查询各组定价属性（决定汇率类别）
    all_codes = list({str(e.get('material_code') or '').strip() for e in entries if e.get('material_code')})
    mat_attrs = _fetch_material_attrs(all_codes) if all_codes else {}

    def _cat_of(code):
        norm = normalize_lookup_code(str(code or ''))
        info = mat_attrs.get(norm) or {}
        return _category_for_attr(info.get('pricing_attr'))

    updated = 0
    cache_items = []
    history_records = []  # 每组吨价快照（用于历史回溯）

    for entry in entries:
        code = str(entry.get('material_code') or '').strip()
        length_tier = str(entry.get('length_tier') or '').strip()
        ton_price = _parse_float(entry.get('ton_price'), None)
        total_kg = _parse_float(entry.get('total_kg'), 0)
        entry_valid = str(entry.get('valid_until') or case_valid or '').strip()
        item_ids = entry.get('item_ids') or []
        if not code:
            continue

        # 报价方式（吨价/单价）：吨价模式按 单重×吨价 计算；单价模式直接使用输入值
        pricing_method = str(entry.get('pricing_method') or '').strip().lower()
        # 当定价属性查不到时，前端可手动选择类别（地桩/铝/铝配件/铁/外购件）
        price_category = str(entry.get('price_category') or '').strip().lower()
        # 组级附加信息（折扣/模具费/起订量/备注）
        g_discount = str(entry.get('discount') or '').strip()
        g_mold_fee = str(entry.get('mold_fee') or '').strip()
        g_moq = str(entry.get('moq') or '').strip()
        g_remark = str(entry.get('remark') or '').strip()

        cat = _cat_of(code) or (price_category if price_category in _CAT_KEYS else None)
        # 手动填价：无汇率类别，或有类别但无单重（前端发送 unit_price_* 字段）
        is_manual = cat is None or 'unit_price_usd' in entry

        ton_type = _ton_type_from_code(code, length_tier)
        e_pack = entry.get('pack') or case_pack
        skey = _ton_setting_key(ton_type, length_tier, total_kg, e_pack)

        if is_manual:
            # 手动填价模式：直接使用前端传入的各币种单价
            manual_prices = {
                'usd': _parse_float(entry.get('unit_price_usd'), 0.0),
                'eur': _parse_float(entry.get('unit_price_eur'), 0.0),
                'rmb_fx': _parse_float(entry.get('unit_price_cny'), 0.0),
                'rmb': _parse_float(entry.get('unit_price_rmb'), 0.0),
            }
            if not any(v > 0 for v in manual_prices.values()):
                continue
            ton_val = 0.0
        else:
            if ton_price in (None, 0):
                continue
            ton_val = float(ton_price)

        # 本组使用的汇率/点数快照（写入历史，回溯用）
        if is_manual:
            rate_snap = {k: 0.0 for k in _RATE_KEYS}
            rate_category = 'manual'
        else:
            rate_snap = {}
            for cur in _RATE_CURS:
                rate_snap[f'exchange_rate_{cur}'] = _sval(f'exchange_rate_{cat}_{cur}')
                rate_snap[f'points_{cur}'] = _sval(f'points_{cat}_{cur}')
            rate_category = cat

        group_prices = []  # 本组各明细的换算价格（历史快照）

        for item_id in item_ids:
            try:
                item_id_int = int(item_id)
            except (TypeError, ValueError):
                continue
            item = get_inquiry_item(item_id_int)
            if not item:
                continue

            factor = _spec_factor(item.get('unit'), item.get('spec'))
            if is_manual:
                # 手动填价：直接使用传入的单价（不按单重换算）
                w = 0.0
                prices = {
                    'rmb': manual_prices.get('rmb', 0.0),
                    'usd': manual_prices.get('usd', 0.0),
                    'eur': manual_prices.get('eur', 0.0),
                    'rmb_fx': manual_prices.get('rmb_fx', 0.0),
                }
            else:
                # 报价方式：
                #   采购吨价 = 单重 × 采购吨价 × 规格/1000 × 系数（人民币 base，再 ÷汇率÷点数）
                #   采购单价 = 采购单价 × 规格/1000（输入人民币成本，人民币 base，再 ÷汇率÷点数）
                #   售价 = 输入美元 × 规格/1000，反算人民币 base 后换算欧元/外汇
                w = _parse_float(item.get('unit_weight'), 0) or 0
                if pricing_method == 'usd':
                    # 售价模式：输入即美元 × 规格/1000，反算人民币后换算其他币种
                    usd_val = ton_val * factor
                    ex_usd = rate_snap['exchange_rate_usd']
                    pt_usd = rate_snap['points_usd']
                    if cat == 'dizhuang':
                        base = usd_val * ex_usd if ex_usd > 0 else 0
                    else:
                        base = usd_val * ex_usd * pt_usd if (ex_usd > 0 and pt_usd > 0) else 0
                    prices = {'rmb': round(base, 6), 'usd': round(usd_val, 6)}
                    for cur in _RATE_CURS:
                        if cur == 'usd':
                            continue
                        ex = rate_snap[f'exchange_rate_{cur}']
                        pt = rate_snap[f'points_{cur}']
                        if cat == 'dizhuang':
                            prices[cur] = round(base / ex, 6) if ex > 0 else 0.0
                        else:
                            prices[cur] = round(base / ex / pt, 6) if ex > 0 and pt > 0 else 0.0
                else:
                    # 采购吨价 / 采购单价：先算出人民币 base，再换算其他币种
                    if pricing_method == 'ton' and w > 0:
                        # 采购吨价：base = 单重 × 采购吨价 × 规格/1000
                        base = w * ton_val * factor
                    else:
                        # 采购单价：base = 采购单价 × 规格/1000（输入人民币成本）
                        base = ton_val * factor
                    if cat == 'lvpj':
                        base *= lvpj_coeff
                    prices = {'rmb': round(base, 6)}
                    for cur in _RATE_CURS:
                        ex = rate_snap[f'exchange_rate_{cur}']
                        pt = rate_snap[f'points_{cur}']
                        if cat == 'dizhuang':
                            prices[cur] = round(base / ex, 6) if ex > 0 else 0.0
                        else:
                            prices[cur] = round(base / ex / pt, 6) if ex > 0 and pt > 0 else 0.0

            # 保存价格时一并写入组级附加信息（折扣/模具费/起订量/备注/报价方式/类别）
            update_inquiry_item(
                item_id_int,
                unit_price_usd=prices.get('usd'),
                unit_price_eur=prices.get('eur'),
                unit_price_cny=prices.get('rmb_fx'),
                unit_price_rmb=prices.get('rmb'),
                quotation_date=quotation_date,
                valid_until=entry_valid,
                discount=g_discount,
                mold_fee=g_mold_fee,
                moq=g_moq,
                remark=g_remark,
                pricing_method=pricing_method,
                price_category=price_category if not _cat_of(code) else '',
                status='priced',
            )
            updated += 1
            group_prices.append({
                'item_id': item_id_int,
                'spec': item.get('spec', ''),
                'unit_weight': w,
                'factor': factor,
                'usd': prices.get('usd'),
                'eur': prices.get('eur'),
                'rmb_fx': prices.get('rmb_fx'),
                'rmb': prices.get('rmb'),
            })
            cache_items.append({
                'material_code': item.get('material_code', ''),
                'spec': item.get('spec', ''),
                'quantity': item.get('quantity', 0),
                'name': item.get('name', ''),
                'unit_price': prices.get('usd') or prices.get('rmb_fx')
                or prices.get('eur') or prices.get('rmb') or 0,
                'unit_price_usd': prices.get('usd'),
                'unit_price_cny': prices.get('rmb_fx'),
                'unit_price_eur': prices.get('eur'),
                'unit_price_rmb': prices.get('rmb'),
                'unit': item.get('unit', '') or '米',
                'quotation_date': quotation_date,
                'valid_until': entry_valid,
                'discount': g_discount,
                'mold_fee': g_mold_fee,
                'moq': g_moq,
                'remark': g_remark,
                'inquirer': item.get('inquirer', ''),
                'source_email': 'web_form',
                'preinstall': item.get('preinstall', '预装'),
            })

        # 本组吨价快照写入历史（保存 1 年，可回溯）
        history_records.append({
            'project_name': project_name,
            'material_code': code,
            'ton_type': ton_type,
            'length_tier': length_tier,
            'weight_tier': _weight_tier_display(total_kg),
            'ton_price': ton_val if not is_manual else None,
            'prices': group_prices,
            'valid_until': entry_valid,
            'saved_by': _current_user_name(),
            'rate_category': rate_category,
            **rate_snap,
        })

    if cache_items:
        try:
            upsert_price_cache(cache_items)
        except Exception as exc:
            print(f'[INQUIRY-ITEMS] case ton price mirror to cache failed: {exc}')

    # 记录案件金额锁定时间点（7 天可改窗口）
    if updated > 0:
        try:
            upsert_case_lock(project_name, _current_user_name())
        except Exception as exc:
            print(f'[INQUIRY-ITEMS] upsert case lock failed: {exc}')

    # 吨价历史保存（1 年，可回溯）
    if history_records:
        try:
            insert_ton_price_history(history_records)
        except Exception as exc:
            print(f'[INQUIRY-ITEMS] insert ton history failed: {exc}')

    # 钉钉通知当前登入用户：价格已填写完毕
    notify_result = {'sent': False}
    if updated > 0:
        notify_result = _notify_case_priced(project_name, updated)

    return {
        'success': True, 'updated': updated,
        'saved_ton_prices': len(ton_updates),
        'notify': notify_result,
        'message': f'已保存 {updated} 条，吨价 {len(ton_updates)} 项',
    }, 200


def update_inquiry_item_fields(item_id, data):
    if not data:
        return {'success': False, 'message': '请求体不能为空'}, 400
    item = get_inquiry_item(item_id)
    if not item:
        return {'success': False, 'message': '询价项不存在'}, 404

    fields = {}
    for key in ('project_name', 'material_code', 'name', 'spec', 'unit', 'remark', 'inquirer', 'status'):
        if key in data:
            fields[key] = str(data.get(key) or '').strip()
    for key in ('quantity', 'unit_weight', 'total_weight'):
        if key in data and data.get(key) not in (None, ''):
            fields[key] = _parse_float(data.get(key), 0.0)

    if not fields:
        return {'success': False, 'message': '没有可更新的字段'}, 400

    ok = update_inquiry_item(item_id, **fields)
    return {'success': ok, 'message': '已更新' if ok else '更新失败'}, 200 if ok else 500


def remove_inquiry_item(item_id):
    ok = repo_delete_inquiry_item(item_id)
    return {'success': ok, 'message': '已删除' if ok else '记录不存在'}, 200 if ok else 404


def remove_inquiry_case(project_name):
    """删除整个案件（询价项 + 锁定 + 备注 + 附件记录与磁盘文件）。仅 admin 调用。"""
    project_name = str(project_name or '').strip()
    if not project_name:
        return {'success': False, 'message': '案件名称不能为空'}, 400
    existing = list_inquiry_items(project=project_name, page=1, page_size=1)
    if not (existing.get('items') or []) and not (existing.get('total') or 0):
        return {'success': False, 'message': '案件不存在或无询价项'}, 404

    result = delete_inquiry_case(project_name)
    removed_files = 0
    for stored in result.get('attachments') or []:
        try:
            path = _attachment_disk_path(stored)
            if os.path.exists(path):
                os.remove(path)
                removed_files += 1
        except Exception as exc:
            print(f'[INQUIRY-ITEMS] remove case attachment file failed: {exc}')

    item_count = result.get('item_count', 0)
    return {
        'success': True,
        'deleted_count': item_count,
        'removed_files': removed_files,
        'message': f'已删除案件「{project_name}」（{item_count} 条询价项）',
    }, 200


def run_inquiry_cleanup():
    deleted = cleanup_expired_records()
    deleted_history = 0
    try:
        deleted_history = cleanup_expired_ton_history()
    except Exception as exc:
        print(f'[INQUIRY-ITEMS] cleanup ton history failed: {exc}')
    total = deleted + deleted_history
    return {
        'success': True,
        'deleted_count': total,
        'message': f'已清理 {deleted} 条过期询价记录，{deleted_history} 条过期吨价历史',
    }


# ====================================================================
# 案件附件 / 案件备注 / 吨价历史 / 下载（#2 #3 #7）
# ====================================================================

_ATTACHMENT_ALLOWED_EXT = {
    '.xlsx', '.xls', '.csv', '.pdf', '.png', '.jpg', '.jpeg', '.gif',
    '.bmp', '.webp', '.zip', '.rar', '.7z', '.doc', '.docx', '.txt',
}
_ATTACHMENT_MAX_SIZE = 30 * 1024 * 1024  # 30MB


def save_inquiry_attachments(project_name, files, uploaded_by=''):
    """保存提交的附件到本地服务器文件夹，并记录元数据。返回 (saved_count, errors)。"""
    if not files:
        return 0, []
    try:
        from backend.config.settings import UPLOAD_FOLDER
        save_dir = UPLOAD_FOLDER / 'inquiry'
    except Exception:
        save_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'uploads', 'inquiry')
    os.makedirs(save_dir, exist_ok=True)

    project_name = str(project_name or '').strip()
    uploaded_by = str(uploaded_by or '').strip()
    saved = 0
    errors = []
    for f in files:
        fname = ''
        try:
            fname = getattr(f, 'filename', '') or ''
        except Exception:
            fname = ''
        fname = os.path.basename(str(fname))
        if not fname:
            continue
        ext = os.path.splitext(fname)[1].lower()
        if ext not in _ATTACHMENT_ALLOWED_EXT:
            errors.append(f'{fname}: 不支持的文件类型({ext})')
            continue
        try:
            content = f.read()
        except Exception as exc:
            errors.append(f'{fname}: 读取失败({exc})')
            continue
        if len(content) > _ATTACHMENT_MAX_SIZE:
            errors.append(f'{fname}: 超过 30MB 限制')
            continue
        stored_name = f'{datetime.now().strftime("%Y%m%d%H%M%S")}_{uuid.uuid4().hex[:8]}{ext}'
        try:
            with open(os.path.join(save_dir, stored_name), 'wb') as fp:
                fp.write(content)
        except Exception as exc:
            errors.append(f'{fname}: 保存失败({exc})')
            continue
        try:
            insert_attachment(project_name, fname, stored_name, len(content), uploaded_by)
            saved += 1
        except Exception as exc:
            errors.append(f'{fname}: 记录失败({exc})')
    return saved, errors


def _attachment_disk_path(stored_name):
    try:
        from backend.config.settings import UPLOAD_FOLDER
        return str(UPLOAD_FOLDER / 'inquiry' / stored_name)
    except Exception:
        return os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'uploads', 'inquiry', stored_name)


def get_inquiry_attachment_path(attachment_id):
    att = get_attachment(attachment_id)
    if not att:
        return None
    path = _attachment_disk_path(att.get('stored_name') or '')
    if not os.path.exists(path):
        return None
    return {
        'path': path,
        'original_name': att.get('original_name') or att.get('stored_name') or 'attachment',
        'stored_name': att.get('stored_name'),
        'size': att.get('file_size'),
    }


def update_case_remark(project_name, remark=None, pricer_remark=None):
    """更新案件备注。remark=业务备注, pricer_remark=报价人员备注，None=不修改。"""
    upsert_case_meta(project_name, remark=remark, pricer_remark=pricer_remark)
    return {'success': True, 'message': '备注已保存'}


def get_ton_history(project=None, page=1, page_size=50):
    page = max(1, int(page or 1))
    page_size = max(1, min(int(page_size or 50), 500))
    project = str(project or '').strip() or None
    return list_ton_price_history(project=project, page=page, page_size=page_size)


def build_case_price_workbook(project_name):
    """构建案件 Excel（下载吨价/价格），3个Sheet：吨价、明细生成价格、案件信息。"""
    from openpyxl.styles import Font, PatternFill, Alignment

    project_name = str(project_name or '').strip()
    result = list_inquiry_items(project=project_name or None, page=1, page_size=2000)
    items = result.get('items') or []
    meta = get_case_meta(project_name)

    save_time = ''
    for it in items:
        if it.get('quotation_date'):
            save_time = it['quotation_date']
            break
    valid_until = ''
    for it in items:
        if it.get('valid_until'):
            valid_until = it['valid_until']
            break
    business_remark = (meta or {}).get('remark') or ''
    pricer_remark = (meta or {}).get('pricer_remark') or ''
    case_pack = _normalize_pack((meta or {}).get('pack'))
    business_name = (items[0].get('business_name') or items[0].get('inquirer') or '') if items else ''

    # 吨价组（与前端 _buildTonGroups 口径一致）
    grouped = {}
    code_totals = {}
    for it in items:
        uw = _parse_float(it.get('unit_weight'), 0)
        if not uw or uw <= 0:
            continue
        unit = str(it.get('unit') or '').strip()
        has_len = unit in _METER_UNITS
        len_tier = ''
        if has_len:
            mm = extract_length_from_spec(str(it.get('spec') or ''))
            if mm and mm > 0:
                m = mm / 1000.0
                len_tier = '0-1' if m < 1 else ('1-3' if m < 3 else '3+')
        code = str(it.get('material_code') or '').strip()
        ton_type = _ton_type_from_code(code, len_tier)
        key = f'{code}@@{len_tier}' if has_len else ton_type
        entry = grouped.setdefault(key, {
            'material_code': code, 'name': it.get('name') or code or '',
            'length_tier': len_tier, 'ton_type': ton_type,
            'items': [], 'total_kg': 0.0,
        })
        entry['items'].append(it)
        tw = _parse_float(it.get('total_weight'), 0)
        entry['total_kg'] += tw
        code_totals[code] = code_totals.get(code, 0.0) + tw
    for g in grouped.values():
        g['total_kg'] = code_totals.get(g['material_code'], g['total_kg'])

    # 读取吨价设置
    settings_keys = []
    for g in grouped.values():
        skey = _ton_setting_key(g['ton_type'], g['length_tier'], g['total_kg'], case_pack)
        g['setting_key'] = skey
        if skey and skey not in settings_keys:
            settings_keys.append(skey)
    settings = _read_temp_settings(settings_keys + _RATE_CAT_KEYS + [_LVPJ_COEFF_KEY])

    def _sval(key):
        v = settings.get(key)
        return float(v) if v not in (None, '') and v > 0 else 0.0

    # 确定每组使用的汇率类别（定价属性 → 类别；查不到则用手动保存的 price_category）
    _cat_label_map = {c['key']: c['label'] for c in RATE_CATEGORIES}
    for g in grouped.values():
        first = (g['items'][0] if g.get('items') else {}) or {}
        cat = _category_for_attr(first.get('pricing_attr'))
        if not cat:
            pc = str(first.get('price_category') or '').strip().lower()
            if pc in _CAT_KEYS:
                cat = pc
        g['category'] = cat

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F4E78')

    def _style_header(ws):
        for c in ws[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center')

    # ── Sheet 1: 吨价 ──
    ws1 = wb.active
    ws1.title = '吨价'
    headers1 = ['案件名称', '序号', '保存时间', '物料编码', '名称', '吨价类型',
                '长度范围', '吨位范围', '总重量(kg)', '吨价(基准)', '有效期',
                '汇率类别', '美元汇率', '欧元汇率', '人民币外汇汇率',
                '美元点数', '欧元点数', '人民币外汇点数']
    ws1.append(headers1)
    _style_header(ws1)
    for idx, g in enumerate(grouped.values(), start=1):
        ton_val = _sval(g.get('setting_key', ''))
        cat = g.get('category') or ''
        cat_label = _cat_label_map.get(cat, '手动填价' if not cat else cat)
        if cat:
            rate_cells = [
                _sval(f'exchange_rate_{cat}_usd') or '',
                _sval(f'exchange_rate_{cat}_eur') or '',
                _sval(f'exchange_rate_{cat}_rmb_fx') or '',
                _sval(f'points_{cat}_usd') or '',
                _sval(f'points_{cat}_eur') or '',
                _sval(f'points_{cat}_rmb_fx') or '',
            ]
        else:
            rate_cells = ['', '', '', '', '', '']
        ws1.append([
            project_name, idx, save_time,
            g['material_code'], str(g['name'])[:20], g['ton_type'],
            (g['length_tier'] + ' 米') if g['length_tier'] else '',
            _weight_tier_display(g['total_kg']), round(g['total_kg'], 2),
            round(ton_val, 6) if ton_val else '',
            valid_until,
            cat_label, *rate_cells,
        ])
    for i, width in enumerate([16, 6, 14, 14, 18, 14, 12, 12, 12, 14, 12,
                               10, 12, 12, 16, 12, 12, 16], start=1):
        ws1.column_dimensions[ws1.cell(row=1, column=i).column_letter].width = width

    # ── Sheet 2: 明细生成价格 ──
    ws2 = wb.create_sheet('明细生成价格')
    headers2 = ['案件名称', '序号', '保存时间', '物料编码', '名称', '规格', '数量',
                '单重(kg)', '总重量(kg)', '单位',
                '价格(美元)', '价格(欧元)', '价格(人民币外汇)', '价格(人民币)', '有效期']
    ws2.append(headers2)
    _style_header(ws2)
    for idx, it in enumerate(items, start=1):
        ws2.append([
            project_name, idx, save_time,
            it.get('material_code', ''), str(it.get('name') or '')[:20],
            it.get('spec', ''), it.get('quantity', 0),
            it.get('unit_weight') if it.get('unit_weight') is not None else '',
            it.get('total_weight') if it.get('total_weight') is not None else '',
            it.get('unit', ''),
            it.get('unit_price_usd') if it.get('unit_price_usd') is not None else '',
            it.get('unit_price_eur') if it.get('unit_price_eur') is not None else '',
            it.get('unit_price_cny') if it.get('unit_price_cny') is not None else '',
            it.get('unit_price_rmb') if it.get('unit_price_rmb') is not None else '',
            it.get('valid_until', ''),
        ])
    for i, width in enumerate([16, 6, 14, 14, 18, 18, 8, 10, 12, 8, 14, 14, 16, 14, 12], start=1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = width

    # ── Sheet 3: 案件信息 ──
    ws3 = wb.create_sheet('案件信息')
    headers3 = ['案件名称', '保存时间', '有效期', '业务备注', '报价人员备注', '业务员',
                '折扣', '模具费', '起订量', '备注']
    ws3.append(headers3)
    _style_header(ws3)
    # 按物料编码去重，每组一行
    seen = {}
    for it in items:
        code = str(it.get('material_code') or '').strip()
        if not code or code in seen:
            continue
        seen[code] = {
            'discount': it.get('discount') or '',
            'mold_fee': it.get('mold_fee') or '',
            'moq': it.get('moq') or '',
            'remark': it.get('remark') or '',
        }
    if seen:
        for info in seen.values():
            ws3.append([project_name, save_time, valid_until, business_remark, pricer_remark, business_name,
                        info['discount'], info['mold_fee'], info['moq'], info['remark']])
    else:
        ws3.append([project_name, save_time, valid_until, business_remark, pricer_remark, business_name, '', '', '', ''])
    for i, width in enumerate([16, 14, 12, 24, 24, 12, 8, 10, 10, 16], start=1):
        ws3.column_dimensions[ws3.cell(row=1, column=i).column_letter].width = width

    return wb


def build_ton_history_workbook(records):
    """构建吨价历史 Excel（吨价历史下载），3个Sheet：吨价、明细生成价格、案件信息。"""
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F4E78')

    def _style_header(ws):
        for c in ws[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center')

    # ── Sheet 1: 吨价 ──
    ws1 = wb.active
    ws1.title = '吨价'
    headers1 = ['案件名称', '序号', '保存时间', '物料编码', '吨价类型',
                '长度范围', '吨位范围', '吨价(基准)', '有效期',
                '汇率类别', '美元汇率', '欧元汇率', '人民币外汇汇率',
                '美元点数', '欧元点数', '人民币外汇点数']
    ws1.append(headers1)
    _style_header(ws1)
    _cat_labels = {c['key']: c['label'] for c in RATE_CATEGORIES}
    for idx, r in enumerate(records or [], start=1):
        cat_label = _cat_labels.get(r.get('rate_category'), '手动填价' if r.get('rate_category') == 'manual' else '')
        ws1.append([
            r.get('project_name', ''), idx, r.get('saved_at', ''),
            r.get('material_code', ''), r.get('ton_type', ''),
            r.get('length_tier', ''), r.get('weight_tier', ''),
            r.get('ton_price') if r.get('ton_price') is not None else '',
            r.get('valid_until', ''),
            cat_label,
            r.get('exchange_rate_usd') if r.get('exchange_rate_usd') is not None else '',
            r.get('exchange_rate_eur') if r.get('exchange_rate_eur') is not None else '',
            r.get('exchange_rate_rmb_fx') if r.get('exchange_rate_rmb_fx') is not None else '',
            r.get('points_usd') if r.get('points_usd') is not None else '',
            r.get('points_eur') if r.get('points_eur') is not None else '',
            r.get('points_rmb_fx') if r.get('points_rmb_fx') is not None else '',
        ])
    for i, width in enumerate([16, 6, 18, 14, 14, 12, 12, 14, 12,
                               10, 12, 12, 16, 12, 12, 16], start=1):
        ws1.column_dimensions[ws1.cell(row=1, column=i).column_letter].width = width

    # ── Sheet 2: 明细生成价格 ──
    ws2 = wb.create_sheet('明细生成价格')
    headers2 = ['案件名称', '序号', '保存时间', '物料编码', '规格', '单重(kg)',
                '价格(美元)', '价格(欧元)', '价格(人民币外汇)', '价格(人民币)', '有效期']
    ws2.append(headers2)
    _style_header(ws2)
    det_idx = 0
    for r in records or []:
        for p in (r.get('prices') or []):
            det_idx += 1
            ws2.append([
                r.get('project_name', ''), det_idx, r.get('saved_at', ''),
                r.get('material_code', ''), p.get('spec', ''),
                p.get('unit_weight') if p.get('unit_weight') is not None else '',
                p.get('usd') if p.get('usd') is not None else '',
                p.get('eur') if p.get('eur') is not None else '',
                p.get('rmb_fx') if p.get('rmb_fx') is not None else '',
                p.get('rmb') if p.get('rmb') is not None else '',
                r.get('valid_until', ''),
            ])
    for i, width in enumerate([16, 6, 18, 14, 18, 10, 14, 14, 16, 14, 12], start=1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = width

    # ── Sheet 3: 案件信息 ──
    ws3 = wb.create_sheet('案件信息')
    headers3 = ['案件名称', '保存时间', '有效期', '业务备注', '报价人员备注', '业务员',
                '汇率类别', '美元汇率', '欧元汇率', '人民币外汇汇率',
                '美元点数', '欧元点数', '人民币外汇点数']
    ws3.append(headers3)
    _style_header(ws3)
    _cat_labels = {c['key']: c['label'] for c in RATE_CATEGORIES}
    for r in records or []:
        cat_label = _cat_labels.get(r.get('rate_category'), '手动填价' if r.get('rate_category') == 'manual' else '')
        ws3.append([
            r.get('project_name', ''), r.get('saved_at', ''),
            r.get('valid_until', ''),
            '', '', r.get('saved_by', ''),
            cat_label,
            r.get('exchange_rate_usd') if r.get('exchange_rate_usd') is not None else '',
            r.get('exchange_rate_eur') if r.get('exchange_rate_eur') is not None else '',
            r.get('exchange_rate_rmb_fx') if r.get('exchange_rate_rmb_fx') is not None else '',
            r.get('points_usd') if r.get('points_usd') is not None else '',
            r.get('points_eur') if r.get('points_eur') is not None else '',
            r.get('points_rmb_fx') if r.get('points_rmb_fx') is not None else '',
        ])
    for i, width in enumerate([16, 18, 12, 24, 24, 12, 10, 12, 12, 16, 12, 12, 16], start=1):
        ws3.column_dimensions[ws3.cell(row=1, column=i).column_letter].width = width

    return wb


# ====================================================================
# 钉钉询价提醒
# ====================================================================

_INQUIRY_PAGE_LINK = '/#email-mgmt'


def _inquiry_page_url():
    try:
        from backend.config import settings as _s
        base = (_s.PUBLIC_BASE_URL or '').strip().rstrip('/')
    except Exception:
        base = ''
    if not base:
        base = 'http://192.168.0.235:8080'
    return f'{base}{_INQUIRY_PAGE_LINK}'


def _first_inquirer(items):
    for it in items:
        name = str(it.get('inquirer') or '').strip()
        if name:
            return name
    return ''


def _build_inquiry_excel(items, project_name=''):
    """根据询价项生成标准询价表（与报价生成时产出的询价表一致），返回 (文件路径, 文件名)。"""
    from backend.core.quotation_engine import create_inquiry_sheet

    wb = Workbook()
    wb.remove(wb.active)

    products = []
    for it in items:
        products.append({
            'code': it.get('material_code') or it.get('code') or '',
            'name': it.get('name') or '',
            'spec': it.get('spec') or '',
            'quantity': it.get('quantity', 0),
            'unit': it.get('unit') or '',
            'weight': it.get('unit_weight') or 0,
        })

    requester = _first_inquirer(items)
    create_inquiry_sheet(wb, products, '汇总', inquiry_requester=requester)

    safe_project = re.sub(r'[\\/:*?"<>|]', '_', project_name or '询价')[:30]
    date_str = datetime.now().strftime('%Y%m%d_%H%M')
    fname = f'【询价表】{safe_project}_{date_str}.xlsx'
    path = os.path.join(tempfile.gettempdir(), fname)
    wb.save(path)
    return path, fname


def _build_submission_message(project_name, inserted_count, total_count, pending_total, salesperson, excel_name):
    if inserted_count > 0:
        summary = f'本次提交 **{total_count}** 项（新增 {inserted_count}），待填价共 **{pending_total}** 项'
        title = f'询价填价提醒：{project_name or "新业务"} 新增 {inserted_count} 项'
    else:
        summary = f'本次提交 **{total_count}** 项（均已存在，无新增），待填价共 **{pending_total}** 项'
        title = f'询价填价提醒：{project_name or "新业务"} {total_count} 项'
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    text = (
        f'## 【询价提醒】\n\n'
        f'1、业务员：{salesperson or "-"}\n\n'
        f'2、项目：{project_name or "-"}\n\n'
        f'3、日期：{now_str}\n\n'
        f'{summary}\n\n'
        f'excel：{excel_name}'
    )
    return title, text


def _build_reminder_message(pending_total, salesperson, project_name, excel_name):
    title = f'询价待办提醒：{pending_total} 项待填价'
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    text = (
        f'## 【询价提醒】\n\n'
        f'1、业务员：{salesperson or "-"}\n\n'
        f'2、项目：{project_name or "-"}\n\n'
        f'3、日期：{now_str}\n\n'
        f'待填价共 **{pending_total}** 项\n\n'
        f'excel：{excel_name}'
    )
    return title, text


def _send_via_bot(title, text, excel_path=None, excel_name=None):
    try:
        from backend.services.dingtalk_service import get_inquiry_bot_client, DingTalkError
        client = get_inquiry_bot_client()
        if not client:
            print('[INQUIRY-NOTIFY] 询价提醒机器人未配置（KS_INQUIRY_DT_*），跳过')
            return {'sent': False, 'reason': '未配置'}
        task_id, _ = client.send_markdown(title, text)
        file_task = ''
        if excel_path and os.path.exists(excel_path):
            try:
                file_task, _ = client.send_file(excel_path, display_name=excel_name)
            except Exception as fexc:
                print(f'[INQUIRY-NOTIFY] 文件发送失败(文本已发): {fexc}')
        print(f'[INQUIRY-NOTIFY] 钉钉消息已发送: {title} (task={task_id}, file={file_task})')
        return {'sent': True, 'task_id': task_id}
    except Exception as exc:
        print(f'[INQUIRY-NOTIFY] 发送失败: {exc}')
        return {'sent': False, 'reason': str(exc)}


def _send_to_user(userid, title, text):
    """向指定钉钉 userid 发送机器人单聊消息。"""
    try:
        from backend.services.dingtalk_service import get_inquiry_bot_client, DingTalkError
        client = get_inquiry_bot_client()
        if not client:
            print('[INQUIRY-NOTIFY] 询价提醒机器人未配置（KS_INQUIRY_DT_*），跳过')
            return {'sent': False, 'reason': '未配置'}
        if not userid:
            return {'sent': False, 'reason': '无钉钉 userid'}
        task_id, _ = client.send_markdown(title, text, userid=userid)
        print(f'[INQUIRY-NOTIFY] 钉钉消息已发送给 {userid}: {title} (task={task_id})')
        return {'sent': True, 'task_id': task_id}
    except Exception as exc:
        print(f'[INQUIRY-NOTIFY] 发送失败: {exc}')
        return {'sent': False, 'reason': str(exc)}


def _build_material_notify_lines(project_name):
    """汇总某案件各物料的备注 / 起订量 / 模具费，用于钉钉通知（同一编码合并）。"""
    try:
        result = list_inquiry_items(project=str(project_name or '').strip() or None,
                                    page=1, page_size=2000)
        items = result.get('items') or []
    except Exception as exc:
        print(f'[INQUIRY-NOTIFY] build material lines failed: {exc}')
        return ''
    seen = {}
    for it in items:
        code = str(it.get('material_code') or '').strip()
        if not code or code in seen:
            continue
        moq = str(it.get('moq') or '').strip()
        mold = str(it.get('mold_fee') or '').strip()
        remark = str(it.get('remark') or '').strip()
        if not (moq or mold or remark):
            continue
        parts = [f'- {code}']
        if moq:
            parts.append(f'起订量 {moq}')
        if mold:
            parts.append(f'模具费 {mold}')
        if remark:
            parts.append(f'备注 {remark}')
        seen[code] = '：'.join([parts[0], ' / '.join(parts[1:])]) if len(parts) > 1 else parts[0]
    return '\n\n'.join(seen.values())


def _notify_case_priced(project_name, updated_count):
    """案件价格保存后，向当前登入用户的钉钉发送通知（含案件备注）。"""
    try:
        from backend.services.auth_service import get_current_account as _get_acct
        acct = _get_acct(optional=True) or {}
    except Exception:
        acct = {}
    user_name = str(acct.get('name') or acct.get('username') or '').strip()
    userid = str(acct.get('dingtalkId') or '').strip()

    pricer_remark = ''
    try:
        meta = get_case_meta(project_name) or {}
        pricer_remark = str(meta.get('pricer_remark') or '').split('\u3010\u4ef7\u683c\u516c\u5f0f\u5907\u6ce8\u3011', 1)[0].strip()
    except Exception:
        pass

    detail_lines = _build_material_notify_lines(project_name)

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    _default_name = '\u65b0\u4e1a\u52a1'
    title = f'\u6848\u4ef6\u4ef7\u683c\u5df2\u586b\u5199\uff1a{project_name or _default_name}'
    remark_line = f'4\u3001\u5907\u6ce8\uff1a{pricer_remark}\n\n' if pricer_remark else ''
    detail_block = f'5\u3001\u7269\u6599\u660e\u7ec6\uff1a\n\n{detail_lines}\n\n' if detail_lines else ''
    text = (
        f'## \u3010\u4ef7\u683c\u5df2\u586b\u597d\u3011\n\n'
        f'1\u3001\u62a5\u4ef7\u5458\uff1a{user_name or "-"}\n\n'
        f'2\u3001\u9879\u76ee\uff1a{project_name or "-"}\n\n'
        f'3\u3001\u65e5\u671f\uff1a{now_str}\n\n'
        f'{remark_line}'
        f'{detail_block}'
        f'\u5df2\u586b\u5199 **{updated_count}** \u6761\u4ef7\u683c\uff0c\u8bf7\u53ca\u65f6\u67e5\u770b\u3002\n\n'
        f'[\u70b9\u6b64\u67e5\u770b]({_inquiry_page_url()})'
    )
    # 优先发给当前登录用户的钉钉 userid，失败时回退到机器人默认目标
    if userid:
        result = _send_to_user(userid, title, text)
        if result.get('sent'):
            return result
        print(f'[INQUIRY-NOTIFY] 发给 {userid} 失败({result.get("reason")})，回退到机器人默认目标')
    return _send_via_bot(title, text)


def notify_inquiry_submitted(project_name, inserted_count, items):
    """BOM 提交询价项后，向报价人员发送钉钉提醒（精简消息 + 询价 Excel）。

    只要本次有询价项（即使因重复未新增），都发送提醒。
    """
    if not items:
        return {'sent': False, 'reason': '无询价项'}
    counts = count_inquiry_items_by_status()
    pending_total = counts.get('pending', 0)
    salesperson = _first_inquirer(items)
    excel_path, excel_name = _build_inquiry_excel(items, project_name)
    title, text = _build_submission_message(project_name, inserted_count, len(items), pending_total, salesperson, excel_name)
    return _send_via_bot(title, text, excel_path, excel_name)


def send_pending_inquiry_reminder():
    """定时任务：检查待填价项，发送提醒（无待填价项则跳过）。"""
    counts = count_inquiry_items_by_status()
    pending_total = counts.get('pending', 0)
    if pending_total <= 0:
        return {'sent': False, 'reason': '无待填价项'}
    result = list_inquiry_items(status='pending', page=1, page_size=50)
    items = result.get('items') or []
    salesperson = _first_inquirer(items)
    projects = sorted({it.get('project_name') for it in items if it.get('project_name')})
    project_name = '、'.join(projects) if projects else ''
    excel_path, excel_name = _build_inquiry_excel(items, project_name)
    title, text = _build_reminder_message(pending_total, salesperson, project_name, excel_name)
    return _send_via_bot(title, text, excel_path, excel_name)


def test_inquiry_notification():
    """发送测试消息，验证询价提醒机器人是否连通。"""
    title = '询价提醒机器人测试'
    text = (
        f'## ✅ 测试消息\n\n'
        f'询价提醒机器人已连通。\n\n'
        f'⏰ {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    )
    return _send_via_bot(title, text)
