import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.worksheet.page import PageMargins
import json
import re
import os
import sqlite3

# ==================== 加载JSON数据 ====================
def load_json_data(file_path):
    if not os.path.exists(file_path):
        print(f"警告: 文件 {file_path} 不存在")
        return []
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def find_chinese_packaging(chinese_list, packaging_name):
    for item in chinese_list:
        if item.get('包装清单名称') == packaging_name:
            return item
    return None

def find_english_kit(english_list, kit_code):
    for item in english_list:
        if item.get('套装编码') == kit_code:
            return item
    suffix_match = re.search(r'-(\d+)$', kit_code)
    if suffix_match:
        suffix = suffix_match.group(1)
        for item in english_list:
            eng_code = item.get('套装编码', '')
            if eng_code.startswith('ASM-') and eng_code.endswith('-' + suffix):
                return item
    return None

# ==================== 物料展开函数 ====================
def expand_materials(material, quantity, chinese_list, english_list, visited=None, from_english_kit=False):
    if visited is None:
        visited = set()
    
    material_code = material.get('物料编码', '')
    material_type = material.get('类型', '')
    material_qty_str = material.get('数量', '0')
    material_qty = extract_number(material_qty_str) * quantity
    
    result = []
    
    result.append({
        '编码': material_code,
        '类型': material_type,
        '数量': material_qty,
        '单套数量': extract_number(material_qty_str),
        'from_english_kit': from_english_kit
    })
    
    if material_code in visited:
        return result
    visited.add(material_code)
    
    if material_type == '中文名称':
        chn_item = find_chinese_packaging(chinese_list, material_code)
        if chn_item:
            sub_materials = chn_item.get('物料明细', [])
            for sub in sub_materials:
                sub_code = sub.get('物料编码', '')
                sub_qty_str = sub.get('数量', '0')
                sub_qty = extract_number(sub_qty_str) * material_qty
                
                sub_type = '中文名称' if any(item.get('包装清单名称') == sub_code for item in chinese_list) else '英文物料'
                if find_english_kit(english_list, sub_code):
                    sub_type = '套装编码'
                
                sub_material = {
                    '物料编码': sub_code,
                    '类型': sub_type,
                    '数量': f"{sub_qty}{get_unit_suffix(sub_qty_str)}"
                }
                
                sub_result = expand_materials(sub_material, 1, chinese_list, english_list, visited.copy(), from_english_kit=False)
                result.extend(sub_result)
    
    elif material_type == '套装编码':
        eng_item = find_english_kit(english_list, material_code)
        if eng_item:
            sub_materials = eng_item.get('物料明细', [])
            for sub in sub_materials:
                sub_code = sub.get('物料编码', '')
                sub_qty_str = sub.get('数量', '0')
                sub_qty = extract_number(sub_qty_str) * material_qty
                
                sub_type = '英文物料'
                if find_english_kit(english_list, sub_code):
                    sub_type = '套装编码'
                elif any(item.get('包装清单名称') == sub_code for item in chinese_list):
                    sub_type = '中文名称'
                
                sub_material = {
                    '物料编码': sub_code,
                    '类型': sub_type,
                    '数量': f"{sub_qty}{get_unit_suffix(sub_qty_str)}"
                }
                
                sub_result = expand_materials(sub_material, 1, chinese_list, english_list, visited.copy(), from_english_kit=True)
                result.extend(sub_result)
    
    elif material_type == '英文物料':
        pass
    
    return result

def get_unit_suffix(quantity_str):
    match = re.search(r'[^\d\.]+$', str(quantity_str))
    return match.group(0) if match else ''

# ==================== 从JSON文件读取托盘数据 ====================
def load_pallet_data(json_file_path):
    with open(json_file_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def parse_volume(volume_str):
    match = re.search(r'<([\d\.]+)\*([\d\.]+)\*([\d\.]+)>', volume_str)
    if match:
        return float(match.group(1)), float(match.group(2)), float(match.group(3))
    return 0, 0, 0

def extract_number(quantity_str):
    if isinstance(quantity_str, (int, float)):
        return float(quantity_str)
    match = re.search(r'(\d+(?:\.\d+)?)', str(quantity_str))
    return float(match.group(1)) if match else 0

def extract_pallet_number(pallet_no):
    match = re.search(r'(\d+)', str(pallet_no))
    return int(match.group(1)) if match else 0

def expand_pallet_range(pallets):
    expanded = []
    for pallet in pallets:
        pallet_no = str(pallet.get('托盘序号', ''))
        range_match = re.match(r'^(\d+)#?\s*[-~]\s*(\d+)#?$', pallet_no)
        if range_match:
            start_num = int(range_match.group(1))
            end_num = int(range_match.group(2))
            for i in range(start_num, end_num + 1):
                new_pallet = pallet.copy()
                new_pallet['托盘序号'] = f'{i:03d}#'
                expanded.append(new_pallet)
        else:
            expanded.append(pallet)
    return expanded

def load_material_db(db_path):
    if not os.path.exists(db_path):
        print(f"警告: 数据库文件 {db_path} 不存在")
        return {}
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("SELECT 工程编码, 工程品名_日语, 工程品名_英语, 计价单位, 规格说明, 重量 FROM 名称对应")
    result = {}
    for row in cur.fetchall():
        code = row[0]
        name_ja = row[1] if row[1] else ''
        name_en = row[2] if row[2] else ''
        unit = row[3] if row[3] else ''
        spec = row[4] if row[4] else ''
        weight = row[5] if row[5] else ''
        result[code] = {'工程品名_日语': name_ja, '工程品名_英语': name_en, '计价单位': unit, '规格说明': spec, '重量': weight}
    conn.close()
    return result

def _lookup_db(material_db, code):
    if code in material_db:
        return material_db[code]
    parts = code.split('-')
    if len(parts) >= 3:
        short_code = parts[0] + '-' + parts[-1]
        if short_code in material_db:
            return material_db[short_code]
    if len(parts) >= 2:
        first_two = parts[0] + '-' + parts[1]
        if first_two in material_db:
            return material_db[first_two]
    if parts[0] in material_db:
        return material_db[parts[0]]
    for key in material_db:
        if key.startswith(code[:3]) and (material_db[key]['工程品名_日语'] or material_db[key]['工程品名_英语']):
            return material_db[key]
    return None

def strip_middle_b10(code):
    parts = code.split('-')
    if len(parts) >= 3 and parts[1] in ('B10', '10'):
        return [parts[0]] + parts[2:]
    return parts

def get_display_code(code, db_unit=''):
    stripped = strip_middle_b10(code)
    if db_unit == '米' and len(stripped) >= 2:
        result = stripped[0]
    else:
        result = '-'.join(stripped)
    if result.endswith('-PKG'):
        result = result[:-4]
    return result

def resolve_material_info(material_db, code):
    spec = ''
    stripped_parts = strip_middle_b10(code)
    display_code = '-'.join(stripped_parts)

    db_info = _lookup_db(material_db, display_code)
    if not db_info:
        db_info = _lookup_db(material_db, stripped_parts[0])

    if not db_info:
        db_info = _lookup_db(material_db, code)

    if not db_info:
        return '', '', 0, ''

    description = db_info.get('工程品名_日语', '') or db_info.get('工程品名_英语', '')
    db_unit = db_info['计价单位']
    db_weight_str = db_info.get('重量', '')
    try:
        db_weight = float(db_weight_str) if db_weight_str and db_weight_str not in ('暂无数据', '') else 0
    except ValueError:
        db_weight = 0

    if db_unit == '米' and len(stripped_parts) >= 2:
        last = stripped_parts[-1]
        num_match = re.search(r'(\d+)', last)
        if num_match:
            spec = num_match.group(1)

    if not spec:
        db_spec = db_info.get('规格说明', '')
        if db_spec and db_spec not in ('补充长度(mm)', '待增', '0'):
            spec = db_spec

    return description, spec, db_weight, db_unit

PALLET_TYPE_MAP_JA = {
    "铁托": "鉄パレット", "木托": "木パレット", "木框": "木フレーム",
    "木箱": "木箱", "纸箱": "ダンボール", "裸装": "包装無し"
}

def translate_pallet_type_ja(val):
    if not val:
        return val
    return PALLET_TYPE_MAP_JA.get(str(val).strip(), val)

def format_spec_display(val):
    if val is None or val == '':
        return ''
    s = str(val).strip()
    if s != '' and re.match(r'^\d+(\.\d+)?$', s):
        return s + 'mm'
    return s

def create_pallet_sheet(wb, sheet_name, data_row_count, main_mark, summary_row,
                        summary_name="汇总", logo_path=None):
    ws = wb.create_sheet(title=sheet_name)
    sn = f"'{summary_name}'"

    border = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
    )
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
    fn = 'MS UI Gothic'

    ws.merge_cells('A1:D1')
    ws['A1'] = 'SHIPPING MARK'
    ws['A1'].font = Font(name=fn, size=48)
    ws['A1'].alignment = align_center
    for c in range(1, 5):
        ws.cell(row=1, column=c).fill = gray_fill
        ws.cell(row=1, column=c).border = border

    ws.merge_cells('E1:F6')
    for r in range(1, 7):
        for c in range(5, 7):
            ws.cell(row=r, column=c).border = border

    if logo_path:
        from openpyxl.drawing.image import Image as XlImage
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU
        img = XlImage(logo_path)
        img.width = 336
        img.height = 386
        marker = AnchorMarker(col=4, colOff=pixels_to_EMU(50), row=0, rowOff=pixels_to_EMU(50))
        img.anchor = OneCellAnchor(
            _from=marker,
            ext=XDRPositiveSize2D(pixels_to_EMU(336), pixels_to_EMU(386))
        )
        ws.add_image(img)

    ws.merge_cells('A2:B2')
    ws.cell(row=2, column=1, value='メインマーク').font = Font(name=fn, size=20)
    ws.cell(row=2, column=1).alignment = align_center
    ws.cell(row=2, column=2).alignment = align_center
    ws.merge_cells('C2:D2')
    ws.cell(row=2, column=3, value=f'={sn}!E2').font = Font(name=fn, size=36)
    ws.cell(row=2, column=3).alignment = align_center
    ws.cell(row=2, column=4).alignment = align_center
    for c in range(1, 5):
        ws.cell(row=2, column=c).border = border

    ws.merge_cells('A3:B3')
    ws.cell(row=3, column=1, value='工場番号').font = Font(name=fn, size=20)
    ws.cell(row=3, column=1).alignment = align_center
    ws.cell(row=3, column=2).alignment = align_center
    ws.merge_cells('C3:D3')
    ws.cell(row=3, column=3, value=f'={sn}!N{summary_row}').font = Font(name=fn, size=20)
    ws.cell(row=3, column=3).alignment = align_center
    ws.cell(row=3, column=4).alignment = align_center
    for c in range(1, 5):
        ws.cell(row=3, column=c).border = border

    ws.merge_cells('A4:B4')
    ws.cell(row=4, column=1, value='案件名').font = Font(name=fn, size=20)
    ws.cell(row=4, column=1).alignment = align_center
    ws.cell(row=4, column=2).alignment = align_center
    ws.merge_cells('C4:D4')
    ws.cell(row=4, column=3, value=f'={sn}!O{summary_row}').font = Font(name=fn, size=20)
    ws.cell(row=4, column=3).alignment = align_center
    ws.cell(row=4, column=4).alignment = align_center
    for c in range(1, 5):
        ws.cell(row=4, column=c).border = border

    ws.merge_cells('A5:B5')
    ws.cell(row=5, column=1, value='パレット番号').font = Font(name=fn, size=20)
    ws.cell(row=5, column=1).alignment = align_center
    ws.cell(row=5, column=2).alignment = align_center
    ws.cell(row=5, column=3, value=f'={sn}!L{summary_row}').font = Font(name=fn, size=36)
    ws.cell(row=5, column=3).alignment = align_center
    pt_ref = f'INDEX({sn}!B:B,MATCH(C5,{sn}!L:L,0))'
    ws.cell(row=5, column=4, value=f'=IF({pt_ref}="铁托","鉄パレット",IF({pt_ref}="木托","木パレット",IF({pt_ref}="木框","木フレーム",IF({pt_ref}="木箱","木箱",IF({pt_ref}="纸箱","ダンボール",IF({pt_ref}="裸装","包装無し",{pt_ref}))))))').font = Font(name=fn, size=20)
    ws.cell(row=5, column=4).alignment = align_center
    for c in range(1, 5):
        ws.cell(row=5, column=c).border = border

    ws.merge_cells('A6:B6')
    ws.cell(row=6, column=1, value='サイズ').font = Font(name=fn, size=20)
    ws.cell(row=6, column=1).alignment = align_center
    ws.cell(row=6, column=2).alignment = align_center
    ws.merge_cells('C6:D6')
    ws.cell(row=6, column=3, value=f'={sn}!J{summary_row}').font = Font(name=fn, size=20)
    ws.cell(row=6, column=3).alignment = align_center
    ws.cell(row=6, column=4).alignment = align_center
    for c in range(1, 5):
        ws.cell(row=6, column=c).border = border

    ws.merge_cells('A7:B7')
    ws.cell(row=7, column=1, value='容量（㎥）').font = Font(name=fn, size=20)
    ws.cell(row=7, column=1).alignment = align_center
    ws.cell(row=7, column=2).alignment = align_center
    ws.cell(row=7, column=3, value=f'=ROUND({sn}!K{summary_row},2)').font = Font(name=fn, size=20)
    ws.cell(row=7, column=3).alignment = align_center
    ws.cell(row=7, column=4, value='重量(㎏s)').font = Font(name=fn, size=20)
    ws.cell(row=7, column=4).alignment = align_center
    ws.merge_cells('E7:F7')
    ws.cell(row=7, column=5, value=f'=ROUND({sn}!I{summary_row},2)').font = Font(name=fn, size=20)
    ws.cell(row=7, column=5).alignment = align_center
    for c in range(1, 7):
        ws.cell(row=7, column=c).border = border

    headers = ['パレット番号', '品番', '品名', '規格', '数量（PCS）', '備考']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=8, column=i, value=h)
        cell.font = Font(name=fn, size=20, bold=True)
        cell.fill = gray_fill
        cell.border = border
        cell.alignment = align_center

    vlookup_range = f'{sn}!P:T'
    for seq in range(1, data_row_count + 1):
        r = 8 + seq
        ws.cell(row=r, column=1, value=f'No.{seq}').font = Font(name=fn, size=20)
        ws.cell(row=r, column=2, value=f'=VLOOKUP($C$5&"-No."&{seq},{vlookup_range},2,0)').font = Font(name=fn, size=20)
        ws.cell(row=r, column=3, value=f'=VLOOKUP($C$5&"-No."&{seq},{vlookup_range},3,0)').font = Font(name=fn, size=20)
        ws.cell(row=r, column=4, value=f'=VLOOKUP($C$5&"-No."&{seq},{vlookup_range},4,0)').font = Font(name=fn, size=20)
        ws.cell(row=r, column=5, value=f'=VLOOKUP($C$5&"-No."&{seq},{vlookup_range},5,0)').font = Font(name=fn, size=20)
        ws.cell(row=r, column=6, value='').font = Font(name=fn, size=20)
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).alignment = align_center

    mic_row = 9 + data_row_count
    ws.merge_cells(f'A{mic_row}:F{mic_row}')
    ws.cell(row=mic_row, column=1, value='Made    in    China')
    ws.cell(row=mic_row, column=1).font = Font(name=fn, size=30, bold=True)
    ws.cell(row=mic_row, column=1).alignment = align_right
    for c in range(1, 7):
        ws.cell(row=mic_row, column=c).border = border

    ws.row_dimensions[1].height = 100
    ws.row_dimensions[2].height = 100
    for r in range(3, mic_row + 1):
        ws.row_dimensions[r].height = 50

    for i, w in enumerate([30, 45, 35, 60, 30, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait'
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = '1:8'
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5, header=0.1, footer=0.1)

    ws.sheet_view.view = 'pageBreakPreview'
    ws.sheet_view.zoomScale = 55

    return ws

def convert_pallet_to_rows(pallet, chinese_list, english_list, material_db):
    rows_data = []
    
    pallet_no = pallet['托盘序号']
    packaging = pallet.get('包装', '铁托')
    order_no = pallet.get('订单号', '')
    volume_str = pallet.get('体积', '<0*0*0>')
    length_cbm, width_cbm, height_cbm = parse_volume(volume_str)
    volume_value = round(length_cbm * width_cbm * height_cbm, 2)
    pallet_number = extract_pallet_number(pallet_no)
    
    project_name = ""
    pallet_num = extract_pallet_number(pallet_no)
    for item in chinese_list:
        source_pallet = item.get('来源托盘', '')
        range_match = re.match(r'(\d+)#~(\d+)#', source_pallet)
        if range_match:
            start_num = int(range_match.group(1))
            end_num = int(range_match.group(2))
            if start_num <= pallet_num <= end_num:
                project_name = item.get('包装清单名称', '')
                break
        elif pallet_no in source_pallet:
            project_name = item.get('包装清单名称', '')
            break
    
    materials = pallet.get('物料明细', [])
    
    print(f"\n{'='*60}")
    print(f"处理托盘: {pallet_no}")
    print(f"{'='*60}")
    
    all_materials = []
    
    for material in materials:
        material_code = material.get('物料编码', '')
        material_type = material.get('类型', '')
        quantity_str = material.get('数量', '0')
        parent_qty = extract_number(quantity_str)
        
        print(f"\n[+] 原始物料: {material_code} ({material_type}) x {parent_qty}")
        
        mat_obj = {
            '物料编码': material_code,
            '类型': material_type,
            '数量': quantity_str
        }
        
        expanded = expand_materials(mat_obj, 1, chinese_list, english_list)
        
        for exp in expanded:
            if exp['类型'] == '中文名称':
                continue
            total_qty = exp['数量']
            per_unit_qty = total_qty / parent_qty if parent_qty > 0 else total_qty
            exp['parent_qty'] = parent_qty
            exp['per_unit_qty'] = per_unit_qty
            tag = " [ENGLISH KIT]" if exp.get('from_english_kit', False) else ""
            all_materials.append(exp)
            print(f"   -> {exp['编码']} ({exp['类型']}) x {parent_qty}*{per_unit_qty:.0f}{tag}")
    
    print(f"\n[-] 共 {len(all_materials)} 行物料")
    
    kit_groups = []
    i = 0
    while i < len(all_materials):
        if all_materials[i].get('from_english_kit', False):
            start = i
            if start > 0 and all_materials[start - 1]['类型'] == '套装编码':
                start = start - 1
            while i < len(all_materials) and all_materials[i].get('from_english_kit', False):
                i += 1
            kit_groups.append((start, i - 1))
        else:
            i += 1
    
    kit_group_set = set()
    for (gs, ge) in kit_groups:
        for gi in range(gs, ge + 1):
            kit_group_set.add(gi)
    
    for idx, mat in enumerate(all_materials):
        code = mat['编码']
        mat_type = mat['类型']
        parent_qty = mat['parent_qty']
        per_unit_qty = mat['per_unit_qty']
        is_kit_header = (mat_type == '套装编码' and idx in kit_group_set)
        
        description, spec, db_weight, db_unit = resolve_material_info(material_db, code)
        
        if db_weight > 0 and db_unit == '米' and spec:
            has_unit_weight = True
        elif db_weight > 0:
            has_unit_weight = True
        else:
            has_unit_weight = False
        
        row_data = {
            'row_num': idx,
            'pallet_no': pallet_no,
            'packaging': packaging,
            'item_no': get_display_code(code, db_unit),
            'description': description,
            'spec': spec,
            'parent_qty': parent_qty,
            'per_unit_qty': per_unit_qty,
            'is_kit_header': is_kit_header,
            'has_unit_weight': has_unit_weight,
            'db_weight': db_weight,
            'db_unit': db_unit,
            'lwh': f'{length_cbm}*{width_cbm}*{height_cbm}',
            'cbm': volume_value,
            'pallet_number': pallet_number,
            'pallets': 1,
            'order_no': order_no,
            'project_name': project_name
        }
        rows_data.append(row_data)
    
    return rows_data, kit_groups

# ==================== 创建Excel工作簿 ====================
def create_excel_from_pallets(pallet_json_path, output_file_path, 
                              chinese_json_path="中文包装清单参考.json", 
                              english_json_path="英文单套预装明细.json",
                              db_path="物料数据库.db",
                              main_mark="KSN-SN20260305-MG-01S"):
    
    pallets = load_pallet_data(pallet_json_path)
    pallets = expand_pallet_range(pallets)
    chinese_list = load_json_data(chinese_json_path)
    english_list = load_json_data(english_json_path)
    material_db = load_material_db(db_path)
    
    print(f"\n[*] 已加载中文包装清单: {len(chinese_list)} 条")
    print(f"[*] 已加载英文套装明细: {len(english_list)} 条")
    print(f"[*] 已加载物料数据库: {len(material_db)} 条")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "汇总"
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='medium', color='000000'), 
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'), 
        bottom=Side(style='medium', color='000000')
    )
    
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    ws.merge_cells('C1:F1')
    ws['C1'] = '转存PDF 后再发客人'
    ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('H1:O1')
    ws['H1'] = '不能>26T,超过要加柜子'
    ws['H1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('C2:D2')
    ws.merge_cells('E2:O2')
    ws['C2'] = 'MARKING'
    ws['E2'] = main_mark
    ws['C2'].fill = header_fill
    ws['E2'].fill = header_fill
    ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['E2'].alignment = Alignment(horizontal='center', vertical='center')
    
    merge_ranges = ['A3:A4', 'B3:B4', 'C3:C4', 'D3:D4', 'E3:E4', 'F3:F4',
                    'G3:I3', 'J3:K3', 'L3:L4', 'M3:M4', 'N3:N4', 'O3:O4']
    for rng in merge_ranges:
        ws.merge_cells(rng)
    
    ws['A3'] = 'パレット番号'
    ws['B3'] = '包装'
    ws['C3'] = '品番'
    ws['D3'] = '品名'
    ws['E3'] = '規格'
    ws['F3'] = '数量\n（PCS）'
    ws['G3'] = '重量（KG）'
    ws['J3'] = '体積（m³）'
    ws['L3'] = 'パレット番号'
    ws['M3'] = '件数'
    ws['N3'] = '工場番号'
    ws['O3'] = '案件名'
    
    ws['G4'] = '単品重'
    ws['H4'] = '正味重量\n(NW)'
    ws['I4'] = '総重量\n(GW)'
    ws['J4'] = 'サイズ'
    ws['K4'] = '容量\n（m³）'
    
    for row in [3, 4]:
        for col in range(1, 16):
            cell = ws.cell(row=row, column=col)
            if col not in [1, 2]:
                cell.border = thin_border
                cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    all_rows = []
    pallet_ranges = []
    all_kit_groups = []
    
    for pallet in pallets:
        rows_data, kit_groups = convert_pallet_to_rows(pallet, chinese_list, english_list, material_db)
        if not rows_data:
            continue
        start_idx = len(all_rows)
        all_rows.extend(rows_data)
        end_idx = len(all_rows) - 1
        pallet_ranges.append({
            'start': start_idx,
            'end': end_idx,
            'pallet_no': rows_data[0]['pallet_no'] if rows_data else ''
        })
        for (kg_start, kg_end) in kit_groups:
            all_kit_groups.append((start_idx + kg_start, start_idx + kg_end))
    
    current_row = 5
    row_positions = []
    
    for row_data in all_rows:
        row_positions.append(current_row)
        
        ws.cell(row=current_row, column=1, value=row_data['pallet_no'])
        ws.cell(row=current_row, column=2, value=row_data['packaging'])
        ws.cell(row=current_row, column=3, value=row_data['item_no'])
        ws.cell(row=current_row, column=4, value=row_data['description'])
        ws.cell(row=current_row, column=5, value=row_data['spec'])
        if row_data.get('is_kit_header', False):
            ws.cell(row=current_row, column=6, value='')
        else:
            pq = row_data['parent_qty']
            uq = row_data['per_unit_qty']
            if uq == int(uq):
                ws.cell(row=current_row, column=6, value=f'={pq:.0f}*{int(uq)}')
            else:
                ws.cell(row=current_row, column=6, value=f'={pq:.0f}*{uq}')
        if row_data['has_unit_weight']:
            db_w = row_data['db_weight']
            db_u = row_data['db_unit']
            sp = row_data['spec']
            if db_u == '米' and sp:
                spec_len = float(sp)
                ws.cell(row=current_row, column=7,
                        value=f'=F{current_row}*{db_w}/1000*{spec_len}')
            else:
                ws.cell(row=current_row, column=7,
                        value=f'=F{current_row}*{db_w}')
        else:
            ws.cell(row=current_row, column=7, value='')
        ws.cell(row=current_row, column=10, value=row_data['lwh'])
        ws.cell(row=current_row, column=11, value=row_data['cbm'])
        ws.cell(row=current_row, column=12, value=row_data['pallet_number'])
        ws.cell(row=current_row, column=13, value=row_data['pallets'])
        ws.cell(row=current_row, column=14, value=row_data['order_no'])
        ws.cell(row=current_row, column=15, value=row_data['project_name'])
        
        for col in range(1, 16):
            cell = ws.cell(row=current_row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if col not in [1, 2]:
                cell.border = thin_border
        
        current_row += 1
    
    for pr in pallet_ranges:
        start_row = row_positions[pr['start']]
        end_row = row_positions[pr['end']]
        
        ws.cell(row=start_row, column=8, value=f'=SUM(G{start_row}:G{end_row})')
        ws.cell(row=start_row, column=9, value=f'=H{start_row}+22')
    
    for pr in pallet_ranges:
        start_row = row_positions[pr['start']]
        end_row = row_positions[pr['end']]
        
        if start_row != end_row:
            for col in ['A', 'B', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
                ws.merge_cells(f'{col}{start_row}:{col}{end_row}')
    
    red_side = Side(style='medium', color='FF0000')
    for (kg_start, kg_end) in all_kit_groups:
        row_start = row_positions[kg_start]
        row_end = row_positions[kg_end]
        for row_idx in range(row_start, row_end + 1):
            for col in range(3, 8):
                cell = ws.cell(row=row_idx, column=col)
                old_border = cell.border
                new_left = red_side if col == 3 else old_border.left
                new_right = red_side if col == 7 else old_border.right
                new_top = red_side if row_idx == row_start else old_border.top
                new_bottom = red_side if row_idx == row_end else old_border.bottom
                cell.border = Border(left=new_left, right=new_right, top=new_top, bottom=new_bottom)
    
    total_qty = sum(row['parent_qty'] * row['per_unit_qty'] for row in all_rows)
    total_rows_count = len(all_rows)
    first_data_row = 5
    
    total_row = current_row
    
    for col in ['A', 'B']:
        cell = ws.cell(row=total_row, column=column_index_from_string(col), value="")
        cell.border = Border()
    
    ws.merge_cells(f'C{total_row}:E{total_row}')
    total_cell = ws.cell(row=total_row, column=3, value="Total")
    total_cell.alignment = Alignment(horizontal='center', vertical='center')
    total_cell.border = thin_border
    total_cell.fill = total_fill
    
    summary_formulas = {
        'F': f'=SUM(F{first_data_row}:F{total_row-1})',
        'G': f'=SUM(G{first_data_row}:G{total_row-1})',
        'H': f'=SUM(H{first_data_row}:H{total_row-1})',
        'I': f'=SUM(I{first_data_row}:I{total_row-1})',
        'K': f'=SUM(K{first_data_row}:K{total_row-1})',
        'M': f'=SUM(M{first_data_row}:M{total_row-1})'
    }
    
    for col_letter, value in summary_formulas.items():
        col_idx = column_index_from_string(col_letter)
        cell = ws.cell(row=total_row, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.fill = total_fill
    
    for col_idx in range(1, 16):
        col_letter = get_column_letter(col_idx)
        if col_letter in ['A', 'B', 'C', 'D', 'E']:
            continue
        if col_letter in summary_formulas:
            continue
        cell = ws.cell(row=total_row, column=col_idx, value="")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.fill = total_fill
    
    last_row = total_row
    
    for row in range(2, last_row + 1):
        for col in range(3, 16):
            cell = ws.cell(row=row, column=col)
            new_border = Border(
                left=cell.border.left if cell.border else None,
                right=cell.border.right if cell.border else None,
                top=cell.border.top if cell.border else None,
                bottom=cell.border.bottom if cell.border else None
            )
            if row == 2:
                new_border.top = Side(style='medium', color='000000')
            if row == last_row:
                new_border.bottom = Side(style='medium', color='000000')
            if col == 3:
                new_border.left = Side(style='medium', color='000000')
            if col == 15:
                new_border.right = Side(style='medium', color='000000')
            cell.border = new_border
    
    ws['C2'].border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    ws['D2'].border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    ws['E2'].border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    for row in [3, 4]:
        ws.cell(row=row, column=3).border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        ws.cell(row=row, column=15).border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    
    for col in range(3, 16):
        cell = ws.cell(row=last_row, column=col)
        current_border = cell.border
        if current_border:
            cell.border = Border(
                left=current_border.left,
                right=current_border.right,
                top=current_border.top,
                bottom=Side(style='medium', color='000000')
            )
        else:
            cell.border = Border(bottom=Side(style='medium', color='000000'))
    
    col_widths = [8, 8, 25, 18, 18, 12, 10, 10, 10, 18, 10, 12, 10, 12, 14]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 30
    
    ws.print_area = f'C2:O{last_row}'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.3, footer=0.3)
    ws.sheet_view.view = 'pageBreakPreview'
    ws.freeze_panes = 'C5'
    
    # ========== 添加辅助列（用于拆分页VLOOKUP） ==========
    helper_col = 16
    for pr in pallet_ranges:
        pallet_rows = all_rows[pr['start']:pr['end'] + 1]
        seq = 0
        for i, rd in enumerate(pallet_rows):
            seq += 1
            r = row_positions[pr['start'] + i]
            ws.cell(row=r, column=helper_col, value=f"{rd['pallet_number']}-No.{seq}")
            ws.cell(row=r, column=helper_col + 1, value=f'=C{r}')
            ws.cell(row=r, column=helper_col + 2, value=f'=D{r}')
            ws.cell(row=r, column=helper_col + 3, value=f'=E{r}')
            ws.cell(row=r, column=helper_col + 4, value=f'=F{r}')

    for c in range(helper_col, helper_col + 5):
        ws.column_dimensions[get_column_letter(c)].hidden = True

    # ========== 拆分：为每个托盘生成独立工作表 ==========
    logo_path = os.path.join(os.path.dirname(output_file_path), 'logo_ja.png')
    logo_exists = os.path.exists(logo_path)

    for pr in pallet_ranges:
        pallet_rows = all_rows[pr['start']:pr['end'] + 1]
        if not pallet_rows:
            continue

        summary_start_row = row_positions[pr['start']]
        data_row_count = len(pallet_rows)
        sheet_name = str(pallet_rows[0]['pallet_number'])
        create_pallet_sheet(wb, sheet_name, data_row_count, main_mark,
                            summary_row=summary_start_row,
                            summary_name="汇总",
                            logo_path=logo_path if logo_exists else None)

    wb.save(output_file_path)
    print(f"\n{'='*60}")
    print(f"[OK] Excel ファイル '{output_file_path}' が生成されました!")
    print(f"[STAT] {len(pallets)} パレット、{total_rows_count} 行の明細データを処理しました。")
    print(f"\n[SUMMARY]")
    print(f"  数量 (PCS): {total_qty}")
    print(f"  パレット数: {total_rows_count}")
    print(f"{'='*60}")

# ==================== 主程序入口 ====================
if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(os.path.dirname(script_dir))
    session_output = os.environ.get('SESSION_OUTPUT_DIR', os.path.join(script_dir, "output"))
    output_filename = os.environ.get('OUTPUT_FILENAME', '物流汇总表v10_ja.xlsx')
    create_excel_from_pallets(
        output_file_path=os.path.join(session_output, output_filename),
        pallet_json_path=os.path.join(session_output, "托盘清单.json"),
        chinese_json_path=os.path.join(session_output, "中文包装清单参考.json"),
        english_json_path=os.path.join(session_output, "英文单套预装明细.json"),
        db_path=os.path.join(project_dir, "物料数据库.db")
    )
