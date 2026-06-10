import io
import json

from flask import Blueprint, jsonify, request, send_file

from backend.services.auth_service import ensure_admin_account
from backend.repositories.quotation_log_repository import (
    get_overview_stats,
    get_user_stats,
    get_group_stats,
    get_trend,
    query_logs,
    get_user_detail,
    query_all_logs,
)
from backend.repositories.user_repository import list_accounts


def _build_china_name_map():
    accounts = list_accounts()
    return {a['username']: a.get('name', '') for a in accounts if a.get('username')}


def _get_admin_usernames():
    accounts = list_accounts()
    return [a['username'] for a in accounts if a.get('role') == 'admin' and a.get('username')]

usage_bp = Blueprint('usage', __name__, url_prefix='/api/admin/usage')


@usage_bp.get('/overview')
def overview_route():
    ensure_admin_account()
    return jsonify({'success': True, 'data': get_overview_stats(exclude_usernames=_get_admin_usernames())})


@usage_bp.get('/by-user')
def by_user_route():
    ensure_admin_account()
    start = request.args.get('start')
    end = request.args.get('end')
    rows = get_user_stats(start=start, end=end, exclude_usernames=_get_admin_usernames())
    china_name_map = _build_china_name_map()
    for r in rows:
        r['china_name'] = china_name_map.get(r.get('username', ''), '')
    return jsonify({'success': True, 'data': rows})


@usage_bp.get('/by-group')
def by_group_route():
    ensure_admin_account()
    start = request.args.get('start')
    end = request.args.get('end')
    return jsonify({'success': True, 'data': get_group_stats(start=start, end=end, exclude_usernames=_get_admin_usernames())})


@usage_bp.get('/trend')
def trend_route():
    ensure_admin_account()
    granularity = request.args.get('granularity', 'day')
    start = request.args.get('start')
    end = request.args.get('end')
    return jsonify({'success': True, 'data': get_trend(granularity=granularity, start=start, end=end, exclude_usernames=_get_admin_usernames())})


@usage_bp.get('/details')
def details_route():
    ensure_admin_account()
    start = request.args.get('start')
    end = request.args.get('end')
    username = request.args.get('username')
    group_name = request.args.get('group')
    status = request.args.get('status')
    limit = request.args.get('limit', 100, type=int)
    offset = request.args.get('offset', 0, type=int)
    rows, total = query_logs(
        start=start, end=end, username=username,
        group_name=group_name, status=status,
        limit=limit, offset=offset,
        exclude_usernames=_get_admin_usernames() if not username else None,
    )
    china_name_map = _build_china_name_map()
    for r in rows:
        r['china_name'] = china_name_map.get(r.get('username', ''), '')
    return jsonify({'success': True, 'data': rows, 'total': total})


@usage_bp.get('/user/<username>')
def user_detail_route(username):
    ensure_admin_account()
    detail = get_user_detail(username)
    if not detail:
        return jsonify({'success': False, 'message': '该用户暂无使用记录'}), 404
    china_name_map = _build_china_name_map()
    detail['china_name'] = china_name_map.get(username, '')
    return jsonify({'success': True, 'data': detail})


@usage_bp.get('/export')
def export_route():
    ensure_admin_account()
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    start = request.args.get('start')
    end = request.args.get('end')

    user_rows = get_user_stats(start=start, end=end, exclude_usernames=_get_admin_usernames())
    all_logs = query_all_logs(start=start, end=end, exclude_usernames=_get_admin_usernames())
    china_name_map = _build_china_name_map()

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    group_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    group_font = Font(bold=True, size=11)
    subtotal_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    subtotal_font = Font(bold=True, size=11, color='475569')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    ws1 = wb.active
    ws1.title = '每人总计'
    s1_headers = ['序号', '用户名', '中文名称', '语言组', '生成次数', '成功率', '平均耗时(秒)', '最近活跃时间']
    ws1.append(s1_headers)
    for col_idx, _ in enumerate(s1_headers, 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for idx, r in enumerate(user_rows, 1):
        total = r.get('total_count') or 0
        success = r.get('success_count') or 0
        rate = f"{round((success / total) * 100)}%" if total > 0 else '0%'
        avg_s = round((r.get('avg_duration_ms') or 0) / 1000, 1)
        uname = r.get('username', '')
        ws1.append([
            idx,
            uname,
            china_name_map.get(uname, ''),
            r.get('group_name', ''),
            total,
            rate,
            avg_s,
            r.get('last_active', ''),
        ])
        for col_idx in range(1, len(s1_headers) + 1):
            cell = ws1.cell(row=ws1.max_row, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    for col_idx in range(1, len(s1_headers) + 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = 18

    ws2 = wb.create_sheet('按组员明细')
    s2_headers = ['序号', '时间', '用户名', '中文名称', '语言组', '项目名称', 'BOM文件', '业务类型', '匹配(成功/总数)', 'Sheet数', '耗时(秒)', '状态']
    ws2.append(s2_headers)
    for col_idx, _ in enumerate(s2_headers, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    grouped = {}
    for log in all_logs:
        g = log.get('group_name') or '未知'
        grouped.setdefault(g, []).append(log)

    group_order = ['韩语组', '日语组', '英语组']
    other_groups = sorted(k for k in grouped if k not in group_order)
    ordered_groups = [g for g in group_order if g in grouped] + other_groups

    row_num = 2
    for g in ordered_groups:
        logs = grouped[g]
        ws2.cell(row=row_num, column=1, value=f'【{g}】')
        ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(s2_headers))
        for col_idx in range(1, len(s2_headers) + 1):
            cell = ws2.cell(row=row_num, column=col_idx)
            cell.fill = group_fill
            cell.font = group_font
            cell.border = thin_border
        row_num += 1

        user_grouped = {}
        for log in logs:
            u = log.get('username') or '未知'
            user_grouped.setdefault(u, []).append(log)

        seq = 1
        for u, u_logs in user_grouped.items():
            for log in u_logs:
                match_str = ''
                try:
                    ms = json.loads(log.get('match_stats') or '{}') if isinstance(log.get('match_stats'), str) else (log.get('match_stats') or {})
                    match_str = f"{ms.get('matched_count', 0)}/{ms.get('total_products', 0)}"
                except Exception:
                    pass
                dur_s = round((log.get('duration_ms') or 0) / 1000, 1)
                status_cn = '成功' if log.get('status') == 'success' else '失败'
                ws2.append([
                    seq,
                    log.get('created_at', ''),
                    log.get('username', ''),
                    china_name_map.get(log.get('username', ''), ''),
                    log.get('group_name', ''),
                    log.get('project_name', ''),
                    log.get('bom_filename', ''),
                    log.get('case_type', ''),
                    match_str,
                    log.get('sheet_count', 0),
                    dur_s,
                    status_cn,
                ])
                for col_idx in range(1, len(s2_headers) + 1):
                    cell = ws2.cell(row=row_num, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                row_num += 1
                seq += 1

            ws2.cell(row=row_num, column=1, value=f'{u} 小计')
            ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            ws2.cell(row=row_num, column=3, value=len(u_logs))
            ws2.cell(row=row_num, column=3).font = subtotal_font
            for col_idx in range(1, len(s2_headers) + 1):
                cell = ws2.cell(row=row_num, column=col_idx)
                cell.fill = subtotal_fill
                cell.font = subtotal_font
                cell.border = thin_border
            row_num += 1

    for col_idx in range(1, len(s2_headers) + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = '报价使用统计.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
