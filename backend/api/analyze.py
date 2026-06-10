import io

from flask import Blueprint, jsonify, request, send_file

from backend.services.auth_service import ensure_permission
from backend.services.analyze_service import analyze_bom_db_only


analyze_bp = Blueprint('analyze', __name__, url_prefix='/api')


@analyze_bp.post('/analyze')
def analyze_bom_route():
    ensure_permission('quotation')
    payload, status = analyze_bom_db_only(request.get_json(silent=True))
    return jsonify(payload), status


@analyze_bp.post('/download-missing-image-template')
def download_missing_image_template():
    ensure_permission('quotation')
    data = request.get_json(silent=True) or {}
    codes = data.get('codes') or []
    items = data.get('items') or []
    if not codes and not items:
        return jsonify({'success': False, 'message': '编码列表为空'}), 400

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws['A1'] = '图片'
    ws['B1'] = '对应编码'
    ws['C1'] = '工程品名'

    if items:
        for i, item in enumerate(items, start=2):
            ws.cell(row=i, column=2, value=str(item.get('code', '')))
            ws.cell(row=i, column=3, value=str(item.get('name', '')))
    else:
        for i, code in enumerate(codes, start=2):
            ws.cell(row=i, column=2, value=str(code))

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name='需维护编码模板.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@analyze_bp.post('/download-missing-ja-template')
def download_missing_ja_template():
    ensure_permission('quotation')
    data = request.get_json(silent=True) or {}
    items = data.get('items') or []
    if not items:
        return jsonify({'success': False, 'message': '列表为空'}), 400

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws['A1'] = '工程编码'
    ws['B1'] = '工程品名--日语'

    for i, item in enumerate(items, start=2):
        ws.cell(row=i, column=1, value=str(item.get('code', '')))
        ws.cell(row=i, column=2, value=str(item.get('name', '')))

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name='缺失日语名清单.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
