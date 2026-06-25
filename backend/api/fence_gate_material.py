import base64
import io
import os
import re
import tempfile
import zipfile
from datetime import datetime

from flask import Blueprint, jsonify, request, send_file

from backend.config.settings import DATABASE_PATH, get_db_connection
from backend.image.processor import decode_image_base64, guess_image_extension
from backend.repositories.fence_gate_material_repository import (
    add_column,
    bulk_upsert,
    bulk_update_images,
    create_material,
    delete_material,
    get_material,
    list_all_codes,
    list_all_prices,
    list_all_records_raw,
    list_image_records,
    list_materials,
    list_table_columns,
    update_material,
    VALID_CATEGORIES,
    TABLE,
)
from backend.services.auth_service import ensure_permission, get_current_account


fence_material_bp = Blueprint('fence_material', __name__, url_prefix='/api/fence-materials')


def _get_actor_role():
    from backend.services.auth_service import get_current_account
    account = get_current_account(optional=True)
    if account:
        return str(account.get('role') or '').strip()
    return request.headers.get('X-KS-Role', '').strip()


def _get_actor_user():
    from backend.services.auth_service import get_current_account
    account = get_current_account(optional=True)
    if account:
        return str(account.get('username') or '').strip()
    return request.headers.get('X-KS-User', '').strip()


def _extract_filename_only(filename):
    normalized = str(filename or '').replace('\\', '/').strip()
    if not normalized:
        return ''
    return normalized.rsplit('/', 1)[-1].strip()


def _extract_code_from_filename(filename):
    basename = _extract_filename_only(filename)
    if not basename:
        return ''
    parts = basename.rsplit('.', 1)
    if len(parts) == 2 and parts[0]:
        return parts[0].strip()
    return basename.strip()


def _encode_image_bytes_to_data_url(image_bytes, filename=''):
    if not image_bytes:
        raise ValueError('图片文件为空')
    image_ext = guess_image_extension(image_bytes, '').lower().replace('jpeg', 'jpg')
    if not image_ext:
        suffix = filename.rsplit('.', 1)[-1].lower() if '.' in (filename or '') else ''
        if suffix in {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'}:
            image_ext = suffix.replace('jpeg', 'jpg')
        else:
            image_ext = 'png'
    encoded = base64.b64encode(image_bytes).decode('ascii')
    return f'data:image/{image_ext};base64,{encoded}'


SUPPORTED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'}


def _sanitize_archive_stem(value):
    text = str(value or '').strip().rstrip('.')
    if not text:
        return 'unnamed'
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', text) or 'unnamed'


def _build_archive_filename(code, image_ext, used_names):
    safe_stem = _sanitize_archive_stem(code)
    normalized_ext = str(image_ext or 'png').lower().replace('jpeg', 'jpg')
    candidate = f'{safe_stem}.{normalized_ext}'
    suffix = 2
    while candidate.lower() in used_names:
        candidate = f'{safe_stem}_{suffix}.{normalized_ext}'
        suffix += 1
    used_names.add(candidate.lower())
    return candidate


def _iter_uploaded_image_entries(files):
    for storage in files or []:
        filename = _extract_filename_only(getattr(storage, 'filename', ''))
        if not filename:
            continue
        if filename.lower().endswith('.zip'):
            try:
                with zipfile.ZipFile(storage.stream) as archive:
                    for info in archive.infolist():
                        if info.is_dir():
                            continue
                        entry_name = _extract_filename_only(info.filename)
                        if not entry_name:
                            continue
                        ext = entry_name.rsplit('.', 1)[-1].lower() if '.' in entry_name else ''
                        if ext not in SUPPORTED_IMAGE_EXTENSIONS:
                            continue
                        yield entry_name, archive.read(info)
            except zipfile.BadZipFile:
                pass
            finally:
                try:
                    storage.stream.seek(0)
                except Exception:
                    pass
            continue

        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
        if ext not in SUPPORTED_IMAGE_EXTENSIONS:
            continue
        image_bytes = storage.read()
        try:
            storage.stream.seek(0)
        except Exception:
            pass
        yield filename, image_bytes


def _normalize_lookup_code(code):
    return re.sub(r'\s+', '', str(code or '').strip()).upper()


@fence_material_bp.get('/columns')
def columns_route():
    ensure_permission('database')
    columns = list_table_columns()
    return jsonify({'success': True, 'columns': columns})


@fence_material_bp.get('/prices')
def prices_route():
    get_current_account()
    return jsonify({'success': True, 'data': list_all_prices()})


@fence_material_bp.get('')
def list_route():
    ensure_permission('database')
    category = request.args.get('category', '')
    keyword = request.args.get('keyword', '')
    page = max(1, int(request.args.get('page', 1)))
    page_size = min(200, max(1, int(request.args.get('page_size', 50))))
    return jsonify({'success': True, 'data': list_materials(category, keyword, page, page_size)})


@fence_material_bp.get('/<path:code>')
def get_route(code):
    ensure_permission('database')
    item = get_material(code)
    if not item:
        return jsonify({'success': False, 'message': '未找到'}), 404
    return jsonify({'success': True, 'data': item})


@fence_material_bp.post('')
def create_route():
    ensure_permission('database_submit')
    body = request.get_json(silent=True) or {}
    code = str(body.get('code') or '').strip()
    if not code:
        return jsonify({'success': False, 'message': '编码不能为空'}), 400
    try:
        def _safe_float(val, default=0.0):
            try:
                return float(val)
            except (ValueError, TypeError):
                return default

        item = create_material(
            code=code,
            category=str(body.get('category') or '').strip(),
            name=str(body.get('name') or '').strip(),
            spec=str(body.get('spec') or '').strip(),
            price_usd=_safe_float(body.get('price_usd')),
            price_rmb=_safe_float(body.get('price_rmb')),
            remark=str(body.get('remark') or '').strip(),
            image_base64=str(body.get('image_base64') or '').strip(),
            price_3_5_usd=_safe_float(body.get('price_3_5_usd'), None) if body.get('price_3_5_usd') is not None else None,
            ja_name=str(body.get('日语名称') or '').strip(),
            material_dip=str(body.get('材質表面処理_浸塑') or '').strip(),
            material_hd=str(body.get('材質表面処理_热镀锌') or '').strip(),
        )
        _update_dynamic_fence_columns(code, body)
        return jsonify({'success': True, 'data': get_material(code)}), 201
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 409
    except Exception as e:
        return jsonify({'success': False, 'message': f'创建失败: {e}'}), 500


@fence_material_bp.put('/<path:code>')
def update_route(code):
    ensure_permission('database_submit')
    body = request.get_json(silent=True) or {}
    try:
        def _safe_float(val, default=0.0):
            if val is None:
                return None
            try:
                return float(val)
            except (ValueError, TypeError):
                return default

        item = update_material(
            code=code,
            category=body.get('category'),
            name=body.get('name'),
            spec=body.get('spec'),
            price_usd=_safe_float(body.get('price_usd')),
            price_rmb=_safe_float(body.get('price_rmb')),
            remark=body.get('remark'),
            image_base64=body.get('image_base64'),
            price_3_5_usd=_safe_float(body.get('price_3_5_usd'), None) if body.get('price_3_5_usd') is not None else None,
            ja_name=body.get('日语名称'),
            material_dip=body.get('材質表面処理_浸塑'),
            material_hd=body.get('材質表面処理_热镀锌'),
        )
        _update_dynamic_fence_columns(code, body)
        return jsonify({'success': True, 'data': get_material(code)})
    except LookupError as e:
        return jsonify({'success': False, 'message': str(e)}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': f'更新失败: {e}'}), 500


def _update_dynamic_fence_columns(code, body):
    known_keys = {
        'code', 'category', 'name', 'spec', 'price_usd', 'price_rmb',
        'remark', 'image_base64', 'price_3_5_usd', '日语名称',
        '材質表面処理_浸塑', '材質表面処理_热镀锌',
    }
    extra = {k: v for k, v in body.items() if k not in known_keys}
    if not extra:
        return
    conn = get_db_connection()
    try:
        db_columns = {row[1] for row in conn.execute(f'PRAGMA table_info({TABLE})').fetchall()}
        new_cols = [k for k in extra if k not in db_columns]
        for col in new_cols:
            conn.execute(f'ALTER TABLE {TABLE} ADD COLUMN "{col}" TEXT')
        if new_cols:
            conn.commit()
        set_parts = [f'"{k}" = ?' for k in extra]
        vals = list(extra.values()) + [code]
        conn.execute(f'UPDATE {TABLE} SET {", ".join(set_parts)} WHERE code = ?', vals)
        conn.commit()
    finally:
        conn.close()


@fence_material_bp.delete('/<path:code>')
def delete_route(code):
    ensure_permission('database_submit')
    ok = delete_material(code)
    if not ok:
        return jsonify({'success': False, 'message': '未找到'}), 404
    return jsonify({'success': True})


@fence_material_bp.post('/bulk')
def bulk_route():
    ensure_permission('database_submit')
    body = request.get_json(silent=True) or {}
    items = body.get('items') or []
    result = bulk_upsert(items)
    return jsonify({'success': True, 'data': result})


@fence_material_bp.get('/download')
def download_route():
    ensure_permission('database_download')
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        records = list_all_records_raw()
        columns = [c for c in list_table_columns() if c != 'image_base64']

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = TABLE
        ws.append(columns)

        for record in records:
            row_values = [record.get(column, '') if record.get(column) is not None else '' for column in columns]
            ws.append(row_values)

        header_fill = PatternFill(fill_type='solid', start_color='DDEBF7', end_color='DDEBF7')
        header_font = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='D9E2F3')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        for idx, column_cells in enumerate(ws.columns, 1):
            from openpyxl.utils import get_column_letter
            max_length = 0
            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(idx)].width = min(max(max_length + 2, 10), 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        download_name = f'fence_gate_materials_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            buffer,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as exc:
        return jsonify({'success': False, 'message': f'下载失败: {exc}'}), 500


@fence_material_bp.post('/batch-update')
def batch_update_route():
    ensure_permission('database_submit')
    actor_role = _get_actor_role()
    if actor_role != 'admin':
        return jsonify({'success': False, 'message': '仅管理员可批量更新'}), 403

    file_storage = request.files.get('file')
    if not file_storage or not getattr(file_storage, 'filename', ''):
        return jsonify({'success': False, 'message': '请上传 Excel 文件'}), 400

    filename = file_storage.filename.lower()
    if not filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '仅支持 .xlsx / .xls 文件'}), 400

    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_storage, data_only=True)
        ws = wb.active

        header_row = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=c).value
            if val:
                header_row[c] = str(val).strip()

        code_col = None
        for c, name in header_row.items():
            if name == 'code':
                code_col = c
                break

        if code_col is None:
            return jsonify({'success': False, 'message': 'Excel 第一行未找到"code"列'}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        col_info = conn.execute(f'PRAGMA table_info({TABLE})').fetchall()
        db_columns = {row[1] for row in col_info}
        skip_columns = {'image_base64'}

        new_columns = []
        for c, name in header_row.items():
            if c == code_col or name in skip_columns or name in db_columns:
                continue
            new_columns.append(name)
            conn.execute(f'ALTER TABLE {TABLE} ADD COLUMN "{name}" TEXT')
        if new_columns:
            conn.commit()

        valid_cols = {}
        for c, name in header_row.items():
            if c == code_col or name in skip_columns:
                continue
            if name in db_columns or name in new_columns:
                valid_cols[c] = name

        if not valid_cols:
            conn.close()
            return jsonify({'success': False, 'message': 'Excel 中没有可更新的数据列'}), 400

        updated_count = 0
        skipped_count = 0
        missing_codes = []

        for r in range(2, ws.max_row + 1):
            code_val = ws.cell(row=r, column=code_col).value
            if not code_val or not str(code_val).strip():
                continue
            code = str(code_val).strip()

            updates = {}
            for c, col_name in valid_cols.items():
                cell_val = ws.cell(row=r, column=c).value
                if cell_val is not None:
                    updates[col_name] = cell_val

            if not updates:
                continue

            cursor.execute(f'SELECT COUNT(*) FROM {TABLE} WHERE code = ?', (code,))
            if cursor.fetchone()[0] == 0:
                if len(missing_codes) < 50:
                    missing_codes.append(code)
                skipped_count += 1
                continue

            set_clause = ', '.join([f'"{col}" = ?' for col in updates.keys()])
            sql = f'UPDATE {TABLE} SET {set_clause} WHERE code = ?'
            cursor.execute(sql, list(updates.values()) + [code])
            if cursor.rowcount > 0:
                updated_count += 1

        conn.commit()
        conn.close()

        msg = f'批量更新完成，共更新 {updated_count} 条记录'
        if skipped_count:
            msg += f'，跳过 {skipped_count} 条（数据库中无对应编码）'
        if new_columns:
            msg += f'，新增 {len(new_columns)} 列：{", ".join(new_columns)}'

        return jsonify({
            'success': True,
            'message': msg,
            'updated_count': updated_count,
            'skipped_count': skipped_count,
            'missing_codes_sample': missing_codes[:20],
            'new_columns': new_columns,
        }), 200
    except Exception as exc:
        return jsonify({'success': False, 'message': f'批量更新失败: {exc}'}), 500


def _import_images_from_excel(file_storage, code_lookup):
    import xml.etree.ElementTree as ET
    import openpyxl

    filename = (getattr(file_storage, 'filename', '') or '').lower()
    if not filename.endswith(('.xlsx', '.xls')):
        return None

    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
    try:
        file_storage.save(tmp_path)
        os.close(tmp_fd)
    except Exception:
        os.close(tmp_fd)

    try:
        image_map = {}
        with zipfile.ZipFile(tmp_path, 'r') as zf:
            rid_to_target = {}
            try:
                rels_xml = zf.read('xl/_rels/cellimages.xml.rels').decode('utf-8')
                for rel in ET.fromstring(rels_xml):
                    rid = rel.get('Id')
                    target = rel.get('Target')
                    if rid and target:
                        rid_to_target[rid] = target
            except KeyError:
                pass

            ci_xml = None
            try:
                ci_xml = zf.read('xl/cellimages.xml').decode('utf-8')
            except KeyError:
                pass

            if ci_xml:
                ns = {
                    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
                    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                    'etc': 'http://www.wps.cn/officeDocument/2017/etCustomData',
                }
                root = ET.fromstring(ci_xml)
                for ci in root.findall('.//etc:cellImage', ns):
                    cnv = ci.find('.//xdr:cNvPr', ns)
                    if cnv is None:
                        continue
                    img_name = cnv.get('name', '')
                    blip = ci.find('.//a:blip', ns)
                    rid = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed', '') if blip is not None else ''
                    if img_name and rid:
                        target = rid_to_target.get(rid, '')
                        if target:
                            media_path = 'xl/' + target
                            try:
                                img_bytes = zf.read(media_path)
                                image_map[img_name] = {
                                    'bytes': img_bytes,
                                    'filename': target,
                                }
                            except KeyError:
                                pass

            if not image_map:
                wb_tmp = openpyxl.load_workbook(tmp_path)
                ws_tmp = wb_tmp.active
                for idx, img_obj in enumerate(ws_tmp._images):
                    from io import BytesIO
                    buf = BytesIO()
                    raw_image = img_obj._data()
                    if hasattr(raw_image, 'save'):
                        raw_image.save(buf)
                    else:
                        buf.write(raw_image)
                    image_map[f'_img_{idx}'] = {
                        'bytes': buf.getvalue(),
                        'filename': getattr(getattr(img_obj, 'ref', None), 'filename', '') or f'image_{idx}.png',
                    }
                wb_tmp.close()

        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active

        img_col = 1
        code_col = 2
        for probe_r in range(2, min(ws.max_row + 1, 10)):
            v1 = str(ws.cell(row=probe_r, column=1).value or '')
            v2 = str(ws.cell(row=probe_r, column=2).value or '')
            if 'DISPIMG' in v1.upper():
                img_col = 1
                code_col = 2
                break
            if 'DISPIMG' in v2.upper():
                img_col = 2
                code_col = 1
                break

        header = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=c).value
            if val:
                header[c] = str(val).strip().lower()
        for c, name in header.items():
            if name == 'code' and c != code_col:
                code_col = c
                break

        seen_updates = {}
        missing_codes = []
        skipped_rows = []
        total_entries = 0

        for r in range(2, ws.max_row + 1):
            codes_text = ws.cell(row=r, column=code_col).value
            if not codes_text or not str(codes_text).strip():
                continue

            cell_img = ws.cell(row=r, column=img_col).value
            img_id = None
            cell_img_text = str(cell_img or '')
            if cell_img_text and 'DISPIMG(' in cell_img_text.upper():
                match = re.search(r'DISPIMG\("([^"]+)"', cell_img_text, re.IGNORECASE)
                if match:
                    img_id = match.group(1)

            img_entry = image_map.get(img_id) if img_id else None
            if not img_entry:
                if img_id:
                    skipped_rows.append(f'Row {r}: 未找到图片资源 {img_id}')
                else:
                    skipped_rows.append(f'Row {r}: 未识别到图片公式')
                continue

            codes = [c.strip() for c in str(codes_text).replace('\\n', '\n').split('\n') if c.strip()]
            if not codes:
                continue

            total_entries += 1

            try:
                data_url = _encode_image_bytes_to_data_url(
                    img_entry['bytes'],
                    img_entry.get('filename', ''),
                )
            except ValueError as exc:
                skipped_rows.append(f'Row {r}: 图片处理失败 - {exc}')
                continue

            for code_str in codes:
                normalized = _normalize_lookup_code(code_str)
                if not normalized:
                    skipped_rows.append(f'Row {r}: 编码格式无效 "{code_str}"')
                    continue
                target_code = code_lookup.get(normalized)
                if not target_code:
                    missing_codes.append(code_str)
                    continue
                seen_updates[target_code] = data_url

        wb.close()

        return {
            'seen_updates': seen_updates,
            'missing_codes': sorted(dict.fromkeys(missing_codes)),
            'skipped_rows': skipped_rows,
            'total_entries': total_entries,
        }
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


@fence_material_bp.post('/images/import')
def import_images_route():
    ensure_permission('database_submit')
    actor_role = _get_actor_role()
    if actor_role != 'admin':
        return jsonify({'success': False, 'message': '仅管理员可批量上传图片'}), 403

    files = request.files.getlist('files')
    if not files:
        single_file = request.files.get('file')
        if single_file:
            files = [single_file]

    file_list = [f for f in files if getattr(f, 'filename', '')]
    if not file_list:
        return jsonify({'success': False, 'message': '请先选择要上传的图片或 ZIP 压缩包'}), 400

    try:
        code_lookup = {}
        for code in list_all_codes():
            normalized = _normalize_lookup_code(code)
            if normalized and normalized not in code_lookup:
                code_lookup[normalized] = code

        seen_updates = {}
        missing_codes = []
        skipped_files = []
        total_entries = 0

        excel_file = file_list[0] if len(file_list) == 1 else None
        excel_filename = (getattr(excel_file, 'filename', '') or '').lower()
        if excel_file and excel_filename.endswith(('.xlsx', '.xls')):
            result = _import_images_from_excel(excel_file, code_lookup)
            if result is not None:
                seen_updates = result['seen_updates']
                missing_codes = result['missing_codes']
                skipped_files = result['skipped_rows']
                total_entries = result['total_entries']
            else:
                return jsonify({'success': False, 'message': 'Excel 图片导入失败'}), 400
        else:
            for entry_name, image_bytes in _iter_uploaded_image_entries(file_list):
                total_entries += 1
                code_from_name = _extract_code_from_filename(entry_name)
                normalized = _normalize_lookup_code(code_from_name)

                if not normalized:
                    skipped_files.append(f'{entry_name}: 文件名无法识别编码')
                    continue

                target_code = code_lookup.get(normalized)
                if not target_code:
                    missing_codes.append(code_from_name or entry_name)
                    continue

                try:
                    data_url = _encode_image_bytes_to_data_url(image_bytes, entry_name)
                except ValueError as exc:
                    skipped_files.append(f'{entry_name}: {exc}')
                    continue

                seen_updates[target_code] = data_url

        update_payload = [
            {'code': code, 'image_base64': img_b64}
            for code, img_b64 in seen_updates.items()
        ]
        updated_count = bulk_update_images(update_payload)

        missing_codes = sorted(dict.fromkeys(missing_codes))

        if updated_count <= 0:
            return jsonify({
                'success': False,
                'message': '没有匹配到可更新的图片，请检查图片名称是否与物料编码一致',
                'total_files': total_entries,
                'updated_count': 0,
                'missing_codes': missing_codes,
                'skipped_files': skipped_files,
            }), 400

        return jsonify({
            'success': True,
            'message': f'批量上传完成，已更新 {updated_count} 条图片',
            'total_files': total_entries,
            'updated_count': updated_count,
            'updated_codes': list(seen_updates.keys()),
            'missing_codes': missing_codes,
            'skipped_files': skipped_files,
        }), 200
    except Exception as exc:
        return jsonify({'success': False, 'message': f'批量上传图片失败: {exc}'}), 500


@fence_material_bp.get('/images/export')
def export_images_route():
    ensure_permission('database')
    actor_role = _get_actor_role()
    if actor_role != 'admin':
        return jsonify({'success': False, 'message': '仅管理员可批量下载图片'}), 403

    try:
        records = list_image_records()
        archive_buffer = io.BytesIO()
        used_names = set()
        exported_count = 0

        with zipfile.ZipFile(archive_buffer, mode='w', compression=zipfile.ZIP_STORED) as archive:
            for row in records:
                code = str(row.get('code') or '').strip()
                if not code:
                    continue

                image_value = row.get('image_base64')
                image_bytes, image_ext, image_status = decode_image_base64(image_value)
                if image_status != 'ready':
                    continue

                archive_name = _build_archive_filename(code, image_ext, used_names)
                archive.writestr(archive_name, image_bytes)
                exported_count += 1

        if exported_count <= 0:
            return jsonify({'success': False, 'message': '数据库中没有可导出的图片'}), 400

        archive_buffer.seek(0)
        download_name = f'fence_materials_images_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
        return send_file(
            archive_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=download_name,
        )
    except Exception as exc:
        return jsonify({'success': False, 'message': f'批量下载图片失败: {exc}'}), 500


@fence_material_bp.post('/add-column')
def add_column_route():
    ensure_permission('database_submit')
    actor_role = _get_actor_role()
    if actor_role != 'admin':
        return jsonify({'success': False, 'message': '仅管理员可新增列'}), 403

    body = request.get_json(silent=True) or {}
    column_name = str(body.get('column_name', '')).strip()
    if not column_name:
        return jsonify({'success': False, 'message': '列名不能为空'}), 400

    try:
        add_column(column_name)
        columns = list_table_columns()
        return jsonify({
            'success': True,
            'message': f'已新增列 "{column_name}"',
            'column_name': column_name,
            'total_columns': len(columns),
        }), 200
    except ValueError as exc:
        return jsonify({'success': False, 'message': str(exc)}), 400
    except Exception as exc:
        return jsonify({'success': False, 'message': f'新增列失败: {exc}'}), 500
