from flask import Blueprint, jsonify, request

from backend.config.settings import get_db_connection
from backend.services.auth_service import ensure_permission, get_current_account
from backend.utils.constants import (
    DB_PILE_15_18UM_CODE_COLUMN,
    DB_PILE_15_18UM_PRICE_EUR,
    DB_PILE_15_18UM_PRICE_RMB,
    DB_PILE_15_18UM_PRICE_USD,
    DB_PILE_15_18UM_TABLE,
)

pile_price_bp = Blueprint('pile_price', __name__, url_prefix='/api/pile-prices')


def _get_actor_role():
    account = get_current_account(optional=True)
    return (account or {}).get('role', '')


@pile_price_bp.get('/list')
def list_pile_prices():
    ensure_permission('database_submit')
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            f'''
            SELECT "{DB_PILE_15_18UM_CODE_COLUMN}",
                   "{DB_PILE_15_18UM_PRICE_USD}",
                   "{DB_PILE_15_18UM_PRICE_EUR}",
                   "{DB_PILE_15_18UM_PRICE_RMB}"
            FROM "{DB_PILE_15_18UM_TABLE}"
            ORDER BY "{DB_PILE_15_18UM_CODE_COLUMN}"
            '''
        )
        rows = cursor.fetchall()
        conn.close()
        result = []
        for row in rows:
            result.append({
                'code': row[DB_PILE_15_18UM_CODE_COLUMN],
                'price_usd': row[DB_PILE_15_18UM_PRICE_USD],
                'price_eur': row[DB_PILE_15_18UM_PRICE_EUR],
                'price_rmb': row[DB_PILE_15_18UM_PRICE_RMB],
            })
        return jsonify({'success': True, 'data': result}), 200
    except Exception as exc:
        return jsonify({'success': False, 'message': f'查询失败: {exc}'}), 500


@pile_price_bp.post('/batch-update')
def batch_update_pile_prices():
    ensure_permission('database_submit')
    actor_role = _get_actor_role()
    if actor_role != 'admin':
        return jsonify({'success': False, 'message': '仅管理员可批量更新'}), 403

    file_storage = request.files.get('file')
    if not file_storage or not getattr(file_storage, 'filename', ''):
        return jsonify({'success': False, 'message': '请上传 Excel 文件'}), 400

    filename = file_storage.filename.lower()
    if not filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '仅支持 .xlsx / .xls 文件'}), 400

    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_storage, data_only=True)
        ws = wb.active

        header_row = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=c).value
            if val:
                header_row[c] = str(val).strip()

        code_col = None
        for c, name in header_row.items():
            if name == DB_PILE_15_18UM_CODE_COLUMN:
                code_col = c
                break

        if code_col is None:
            return jsonify({'success': False, 'message': f'Excel 第一行未找到"{DB_PILE_15_18UM_CODE_COLUMN}"列'}), 400

        expected_columns = {
            DB_PILE_15_18UM_PRICE_USD,
            DB_PILE_15_18UM_PRICE_EUR,
            DB_PILE_15_18UM_PRICE_RMB,
        }

        valid_cols = {}
        for c, name in header_row.items():
            if c == code_col:
                continue
            if name in expected_columns:
                valid_cols[c] = name

        if not valid_cols:
            return jsonify({'success': False, 'message': 'Excel 中没有可更新的价格列'}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        updated_count = 0
        skipped_count = 0
        missing_codes = []

        for r in range(2, ws.max_row + 1):
            code_val = ws.cell(row=r, column=code_col).value
            if not code_val or not str(code_val).strip():
                continue
            code = str(code_val).strip()

            updates = {}
            for c, col_name in valid_cols.items():
                cell_val = ws.cell(row=r, column=c).value
                if cell_val is not None:
                    try:
                        updates[col_name] = float(cell_val)
                    except (TypeError, ValueError):
                        updates[col_name] = cell_val

            if not updates:
                continue

            cursor.execute(
                f'SELECT COUNT(*) FROM "{DB_PILE_15_18UM_TABLE}" WHERE "{DB_PILE_15_18UM_CODE_COLUMN}" = ?',
                (code,),
            )
            if cursor.fetchone()[0] == 0:
                if len(missing_codes) < 50:
                    missing_codes.append(code)
                skipped_count += 1
                continue

            set_clause = ', '.join([f'"{col}" = ?' for col in updates.keys()])
            sql = f'UPDATE "{DB_PILE_15_18UM_TABLE}" SET {set_clause} WHERE "{DB_PILE_15_18UM_CODE_COLUMN}" = ?'
            cursor.execute(sql, list(updates.values()) + [code])
            if cursor.rowcount > 0:
                updated_count += 1

        conn.commit()
        conn.close()

        msg = f'批量更新完成，共更新 {updated_count} 条记录'
        if skipped_count:
            msg += f'，跳过 {skipped_count} 条（数据库中无对应编码）'

        return jsonify({
            'success': True,
            'message': msg,
            'updated_count': updated_count,
            'skipped_count': skipped_count,
            'missing_codes_sample': missing_codes[:20],
        }), 200
    except Exception as exc:
        return jsonify({'success': False, 'message': f'批量更新失败: {exc}'}), 500
