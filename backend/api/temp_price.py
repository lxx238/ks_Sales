import io
from datetime import datetime

from flask import Blueprint, jsonify, request, send_file

from backend.config.settings import get_db_connection
from backend.services.auth_service import ensure_temp_price_access

temp_price_bp = Blueprint('temp_price', __name__, url_prefix='/api/temp-price')

TBL_S = 'temp_price_settings'
TBL_M = 'temp_material_pricing'

TIERS = ['01', '13', '3']
TIER_LABELS = {'01': '0-1m', '13': '1-3m', '3': '3+m'}

TON_TIERS = ['05', '550', '50999']
TON_TIER_LABELS = {'05': '0-5\u5428', '550': '5-50\u5428', '50999': '50-999\u5428'}

STANDARD_TON_TYPES = [
    {'key': 'FEC004', 'label': 'FEC004'},
    {'key': 'FEC006', 'label': 'FEC006'},
    {'key': 'FEC020', 'label': 'FEC020'},
    {'key': 'FEC027', 'label': 'FEC027'},
    {'key': 'FEC035', 'label': 'FEC035'},
    {'key': 'FEPJ_0103', 'label': 'FEPJ-0103'},
    {'key': 'FEPJ_0173', 'label': 'FEPJ-0173'},
    {'key': 'FEPJ_0178', 'label': 'FEPJ-0178'},
    {'key': 'FEPJ_2201', 'label': 'FEPJ-2201'},
    {'key': 'FEPJ_2804', 'label': 'FEPJ-2804'},
    {'key': 'FEPJ_2808', 'label': 'FEPJ-2808'},
]

STANDARD_TP_KEYS = {t['key'] for t in STANDARD_TON_TYPES}

# 包装类型维度：吨价按「包装」区分（简易包装 / 铁托）
# 吨价设置键末尾追加 _{pack} 后缀，默认简易包装
PACK_TYPES = [
    {'key': 'jybz', 'label': '简易包装'},
    {'key': 'tietuo', 'label': '铁托'},
]
PACK_KEYS = [p['key'] for p in PACK_TYPES]
PACK_LABELS = {p['key']: p['label'] for p in PACK_TYPES}
DEFAULT_PACK = 'jybz'


def _get_all_ton_types_with_settings(c):
    type_map = {}
    for tp in STANDARD_TON_TYPES:
        type_map[tp['key']] = tp
    for tp in _get_all_ton_types(c):
        if tp['key'] not in type_map:
            type_map[tp['key']] = tp
    cols = {r[1] for r in c.execute(f'PRAGMA table_info({TBL_S})').fetchall()}
    for col in cols:
        if not col.startswith('ton_'):
            continue
        rest = col[4:]
        for marker in ('_int_', '_ext_'):
            idx = rest.find(marker)
            if idx > 0:
                k = rest[:idx]
                if k not in type_map and _is_valid_ton_key(k):
                    label = k.replace('_', '-') if k.startswith('FEPJ') else k
                    type_map[k] = {'key': k, 'label': label}
                break
    return list(type_map.values())


def _is_valid_ton_key(k):
    return bool(k) and all(ch.isalnum() or ch in ('_', '-') for ch in k)


def _get_all_ton_types(c):
    cols = {r[1] for r in c.execute(f'PRAGMA table_info({TBL_M})').fetchall()}
    if '\u5428\u4ef7\u7c7b\u578b' not in cols:
        return list(STANDARD_TON_TYPES)
    rows = c.execute(
        f'SELECT DISTINCT "\u5428\u4ef7\u7c7b\u578b" FROM {TBL_M} '
        f'WHERE "\u5428\u4ef7\u7c7b\u578b" IS NOT NULL AND "\u5428\u4ef7\u7c7b\u578b" != \'\''
    ).fetchall()
    extra = []
    for r in rows:
        k = r[0]
        if k not in STANDARD_TP_KEYS and _is_valid_ton_key(k):
            extra.append({'key': k, 'label': k})
    return list(STANDARD_TON_TYPES) + extra


def _build_ton_keys(types, packs=None):
    keys = []
    packs = packs or PACK_KEYS
    for pk in packs:
        for tp in types:
            k = tp['key']
            if k.startswith('FEPJ_'):
                for tt in TON_TIERS:
                    keys.append(f'ton_{k}_int_{tt}_{pk}')
                keys.append(f'ton_{k}_ext_50999_{pk}')
            else:
                for tier in TIERS:
                    for tt in TON_TIERS:
                        keys.append(f'ton_{k}_int_{tier}_{tt}_{pk}')
                for tier in TIERS:
                    keys.append(f'ton_{k}_ext_{tier}_50999_{pk}')
    return keys


CURRENCIES = [
    {'key': 'usd', 'label': '美元', 'has_rate': True},
    {'key': 'eur', 'label': '欧元', 'has_rate': True},
    {'key': 'rmb_fx', 'label': '人民币外汇', 'has_rate': True},
    {'key': 'rmb_int', 'label': '人民币（无汇率）', 'has_rate': False},
]

CUR_LABELS = {c['key']: c['label'] for c in CURRENCIES}
SIDE_LABELS = {'ext': '外部', 'int': '内部'}
SIDE_KEYS = {'ext': 'external', 'int': 'internal'}
LEN_LABELS = {'01': '0-1', '13': '1-3', '3': '3+'}
MAT_TON_TIERS = ['0-5', '5-50', '50-999']
MAT_LEN_TIERS = ['0-1', '1-3', '3+']

VALID_SIDE_TON = [('int', '0-5'), ('int', '5-50'), ('int', '50-999'), ('ext', '50-999')]

SHARED_KEYS = []
for cur in CURRENCIES:
    if cur['has_rate']:
        SHARED_KEYS.append(f'exchange_rate_{cur["key"]}')
        SHARED_KEYS.append(f'points_{cur["key"]}')

# 汇率/点数分类（不同定价属性使用不同类别）
RATE_CATEGORIES = [
    {'key': 'dizhuang', 'label': '地桩', 'attrs': ['D']},
    {'key': 'lv', 'label': '铝', 'attrs': ['M']},
    {'key': 'lvpj', 'label': '铝配件', 'attrs': ['F', 'Q']},
    {'key': 'tie', 'label': '铁', 'attrs': ['WTX', 'WTP']},
    {'key': 'waigou', 'label': '外购件', 'attrs': ['W']},
]
RATE_CATEGORY_KEYS = []
for _cat in RATE_CATEGORIES:
    for cur in CURRENCIES:
        if cur['has_rate']:
            RATE_CATEGORY_KEYS.append(f'exchange_rate_{_cat["key"]}_{cur["key"]}')
            RATE_CATEGORY_KEYS.append(f'points_{_cat["key"]}_{cur["key"]}')

# 铝配件额外系数（计算时 base × 系数，默认 1.03，前端可修改）
RATE_COEFFICIENT_KEYS = ['lvpj_coefficient']
ALL_RATE_KEYS = RATE_CATEGORY_KEYS + RATE_COEFFICIENT_KEYS


def _all_keys_for_types(types):
    return _build_ton_keys(types) + SHARED_KEYS


TON_KEYS = _build_ton_keys(STANDARD_TON_TYPES)
ALL_SETTINGS_KEYS = TON_KEYS + SHARED_KEYS

PRICE_COLS = []
PRICE_COL_MAP = {}
for pk in PACK_KEYS:
    for side, ton in VALID_SIDE_TON:
        for length in MAT_LEN_TIERS:
            for cur in CURRENCIES:
                col = f'{SIDE_KEYS[side]}_{ton}_{length}_{cur["key"]}_{pk}'
                PRICE_COLS.append(col)
                PRICE_COL_MAP[(side, ton, length, cur['key'], pk)] = col


def _conn():
    c = get_db_connection()
    return c


def _ensure_settings(c, keys=None):
    if keys is None:
        keys = ALL_SETTINGS_KEYS
    existing = {r[1] for r in c.execute(f'PRAGMA table_info({TBL_S})').fetchall()}
    if not existing:
        cols_def = ', '.join(f'"{k}" REAL' for k in keys)
        c.execute(f'''
            CREATE TABLE {TBL_S} (
                id INTEGER PRIMARY KEY DEFAULT 1,
                {cols_def},
                updated_at TEXT
            )
        ''')
        c.execute(f'INSERT INTO {TBL_S} (id) VALUES (1)')
        c.commit()
    else:
        for k in keys:
            if k not in existing:
                c.execute(f'ALTER TABLE {TBL_S} ADD COLUMN "{k}" REAL')
        if not c.execute(f'SELECT 1 FROM {TBL_S} WHERE id=1').fetchone():
            c.execute(f'INSERT INTO {TBL_S} (id) VALUES (1)')
        c.commit()


def _ensure_material_cols(c):
    existing = {r[1] for r in c.execute(f'PRAGMA table_info({TBL_M})').fetchall()}
    if not existing:
        base = '"工程编码" TEXT PRIMARY KEY,"规格说明" TEXT,"工程品名" TEXT,"计价单位" TEXT,"单重" REAL,"定价属性" TEXT,"吨价类型" TEXT'
        price = ', '.join(f'"{col}" REAL' for col in PRICE_COLS)
        c.execute(f'CREATE TABLE {TBL_M} ({base},{price})')
        c.commit()
        return
    for col in PRICE_COLS:
        if col not in existing:
            c.execute(f'ALTER TABLE {TBL_M} ADD COLUMN "{col}" REAL')
    c.commit()


@temp_price_bp.get('/settings')
def get_settings():
    ensure_temp_price_access()
    c = _conn()
    try:
        all_types = _get_all_ton_types(c)
        dynamic_keys = _all_keys_for_types(all_types)
        _ensure_settings(c, dynamic_keys)
        r = c.execute(f'SELECT * FROM {TBL_S} WHERE id=1').fetchone()
        result = {}
        if r:
            cols = r.keys()
            for k in dynamic_keys:
                result[k] = r[k] if k in cols else None
            result['updated_at'] = r['updated_at'] if 'updated_at' in cols else None
        return jsonify({
            'success': True, 'settings': result,
            'currencies': CURRENCIES,
            'tonPriceTypes': all_types,
            'standardTonTypes': STANDARD_TON_TYPES,
            'packTypes': PACK_TYPES,
            'defaultPack': DEFAULT_PACK,
            'tiers': TIERS, 'tierLabels': TIER_LABELS,
            'tonTiers': TON_TIERS, 'tonTierLabels': TON_TIER_LABELS,
            'sharedKeys': SHARED_KEYS,
            'matTonTiers': MAT_TON_TIERS,
            'matLenTiers': MAT_LEN_TIERS,
        })
    finally:
        c.close()


@temp_price_bp.get('/price-cache')
def list_price_cache_route():
    ensure_temp_price_access()
    from backend.repositories.inquiry_repository import list_price_cache
    keyword = (request.args.get('keyword') or '').strip()
    limit = request.args.get('limit', type=int) or 2000
    limit = max(1, min(limit, 5000))
    items = list_price_cache(keyword=keyword or None, limit=limit)
    return jsonify({'success': True, 'items': items, 'total': len(items)})


@temp_price_bp.post('/price-cache/batch-delete')
def batch_delete_price_cache_route():
    ensure_temp_price_access(write=True)
    from backend.repositories.inquiry_repository import delete_price_cache_items
    data = request.get_json(silent=True) or {}
    raw_ids = data.get('ids') or []
    ids = []
    for v in raw_ids:
        try:
            ids.append(int(v))
        except (TypeError, ValueError):
            continue
    deleted = delete_price_cache_items(ids)
    return jsonify({'success': True, 'deleted': deleted})


@temp_price_bp.post('/price-cache/cleanup-expired')
def cleanup_expired_price_cache_route():
    ensure_temp_price_access(write=True)
    from backend.repositories.inquiry_repository import delete_expired_price_cache
    data = request.get_json(silent=True) or {}
    days = data.get('days', 7)
    try:
        days = int(days)
    except (TypeError, ValueError):
        days = 7
    deleted = delete_expired_price_cache(days)
    return jsonify({'success': True, 'deleted': deleted})


@temp_price_bp.get('/rate-settings')
def get_rate_settings():
    ensure_temp_price_access()
    c = _conn()
    try:
        _ensure_settings(c, ALL_RATE_KEYS)
        r = c.execute(f'SELECT * FROM {TBL_S} WHERE id=1').fetchone()
        result = {}
        if r:
            cols = r.keys()
            for k in ALL_RATE_KEYS:
                result[k] = r[k] if k in cols else None
        return jsonify({
            'success': True,
            'settings': result,
            'currencies': [cur for cur in CURRENCIES if cur['has_rate']],
            'rate_categories': RATE_CATEGORIES,
        })
    finally:
        c.close()


@temp_price_bp.post('/rate-settings')
def update_rate_settings():
    ensure_temp_price_access(write=True)
    body = request.get_json(silent=True) or {}
    d = body.get('settings') or {}
    c = _conn()
    try:
        _ensure_settings(c, ALL_RATE_KEYS)
        set_parts = []
        vals = []
        for k in ALL_RATE_KEYS:
            if k in d:
                set_parts.append(f'{k}=?')
                vals.append(_f(d.get(k)))
        if set_parts:
            set_parts.append("updated_at=datetime('now','localtime')")
            vals.append(1)
            c.execute(f'UPDATE {TBL_S} SET {",".join(set_parts)} WHERE id=?', vals)
            c.commit()
        r = c.execute(f'SELECT * FROM {TBL_S} WHERE id=1').fetchone()
        result = {}
        if r:
            cols = r.keys()
            for k in ALL_RATE_KEYS:
                result[k] = r[k] if k in cols else None
        return jsonify({'success': True, 'settings': result})
    finally:
        c.close()


@temp_price_bp.get('/export')
def export_excel():
    ensure_temp_price_access()
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    ton_tier = request.args.get('tonTier', '50-999')
    if ton_tier not in MAT_TON_TIERS:
        ton_tier = '50-999'
    len_tier = request.args.get('lengthTier', '3+')
    if len_tier not in MAT_LEN_TIERS:
        len_tier = '3+'

    show_ext = (ton_tier == '50-999')

    export_cols = []
    export_labels = []
    for cur in CURRENCIES:
        col_key = PRICE_COL_MAP.get(('int', ton_tier, len_tier, cur['key'], DEFAULT_PACK))
        if col_key:
            export_cols.append(col_key)
            export_labels.append(f'{cur["label"]}-内部')
    if show_ext:
        for cur in CURRENCIES:
            col_key = PRICE_COL_MAP.get(('ext', ton_tier, len_tier, cur['key'], DEFAULT_PACK))
            if col_key:
                export_cols.append(col_key)
                export_labels.append(f'{cur["label"]}-外部')

    c = _conn()
    try:
        _ensure_settings(c)
        _ensure_material_cols(c)

        s = c.execute(f'SELECT * FROM {TBL_S} WHERE id=1').fetchone()
        s_cols = s.keys()

        pc = ', '.join(f'"{x}"' for x in export_cols)
        rows = c.execute(
            f'SELECT "\u5de5\u7a0b\u7f16\u7801","\u89c4\u683c\u8bf4\u660e","\u5de5\u7a0b\u54c1\u540d","\u8ba1\u4ef7\u5355\u4f4d","\u5355\u91cd","\u5b9a\u4ef7\u5c5e\u6027",{pc} FROM {TBL_M} ORDER BY "\u5de5\u7a0b\u7f16\u7801"'
        ).fetchall()

        wb = openpyxl.Workbook()

        hdr_fill = PatternFill(fill_type='solid', start_color='DDEBF7', end_color='DDEBF7')
        hdr_font = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')

        ws_s = wb.active
        ws_s.title = '\u8bbe\u7f6e'

        ws_s['A1'] = '\u5428\u91cd\u68af\u5ea6'
        ws_s['B1'] = ton_tier + '\u5428'
        ws_s['A2'] = '\u957f\u5ea6\u68af\u5ea6'
        ws_s['B2'] = len_tier + '\u7c73'
        for cell in ws_s[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
        for cell in ws_s[2]:
            cell.font = hdr_font
            cell.fill = hdr_fill

        ws_s['A4'] = '\u6c47\u7387\u4e0e\u70b9\u6570'
        ws_s['A4'].font = hdr_font
        ws_s['A5'] = '\u5e01\u79cd'
        ws_s['B5'] = '\u6c47\u7387'
        ws_s['C5'] = '\u70b9\u6570'
        for cell in ws_s[5]:
            cell.font = hdr_font
            cell.fill = hdr_fill

        rate_rows = {}
        rate_curs = [cur for cur in CURRENCIES if cur['has_rate']]
        for i, cur in enumerate(rate_curs):
            row = 6 + i
            ws_s[f'A{row}'] = CUR_LABELS[cur['key']]
            ex_key = f'exchange_rate_{cur["key"]}'
            pt_key = f'points_{cur["key"]}'
            ws_s[f'B{row}'] = _f(s[ex_key] if ex_key in s_cols else None)
            ws_s[f'C{row}'] = _f(s[pt_key] if pt_key in s_cols else None)
            rate_rows[cur['key']] = row

        ton_settings_key = MAT_TON_TIERS.index(ton_tier)
        len_settings_key = MAT_LEN_TIERS.index(len_tier)
        settings_ton = TON_TIERS[ton_settings_key]
        settings_len = TIERS[len_settings_key]

        tp_hdr = 6 + len(rate_curs) + 1
        ws_s[f'A{tp_hdr - 1}'] = '\u5428\u4ef7\u5e95\u8868'
        ws_s[f'A{tp_hdr - 1}'].font = hdr_font
        ws_s[f'A{tp_hdr}'] = '\u7c7b\u578b'
        ws_s[f'B{tp_hdr}'] = f'\u5185\u90e8({len_tier}\u7c73)'
        ws_s[f'C{tp_hdr}'] = f'\u5916\u90e8({len_tier}\u7c73)'
        for cell in ws_s[tp_hdr]:
            cell.font = hdr_font
            cell.fill = hdr_fill

        all_types = STANDARD_TON_TYPES
        for i, tp in enumerate(all_types):
            row = tp_hdr + 1 + i
            k = tp['key']
            is_pj = k.startswith('FEPJ_')
            ws_s[f'A{row}'] = tp['label']
            if is_pj:
                key_int = f'ton_{k}_int_{settings_ton}_{DEFAULT_PACK}'
                ws_s[f'B{row}'] = _f(s[key_int] if key_int in s_cols else None)
                key_ext = f'ton_{k}_ext_50999_{DEFAULT_PACK}'
                ws_s[f'C{row}'] = _f(s[key_ext] if key_ext in s_cols else None)
            else:
                key_int = f'ton_{k}_int_{settings_len}_{settings_ton}_{DEFAULT_PACK}'
                ws_s[f'B{row}'] = _f(s[key_int] if key_int in s_cols else None)
                key_ext = f'ton_{k}_ext_{settings_len}_50999_{DEFAULT_PACK}'
                ws_s[f'C{row}'] = _f(s[key_ext] if key_ext in s_cols else None)

        ws_s.column_dimensions['A'].width = 16
        for col_idx in range(2, 4):
            ws_s.column_dimensions[get_column_letter(col_idx)].width = 14

        ws_m = wb.create_sheet('\u7269\u6599\u62a5\u4ef7')

        headers = ['\u5de5\u7a0b\u7f16\u7801', '\u5428\u4ef7\u7c7b\u578b', '\u5355\u91cd']
        headers.extend(export_labels)
        headers.extend(['\u89c4\u683c\u8bf4\u660e', '\u5de5\u7a0b\u54c1\u540d', '\u5355\u4f4d', '\u5c5e\u6027'])
        ws_m.append(headers)

        for cell in ws_m[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center
        ws_m.freeze_panes = 'A2'

        for r_idx, mat in enumerate(rows):
            dr = r_idx + 2
            code = mat['\u5de5\u7a0b\u7f16\u7801'] or ''
            if code.startswith('FEPJ'):
                tp_type = code.replace('-', '_')
            else:
                tp_type = code.split('-')[0] if '-' in code else code
            weight = _f(mat['\u5355\u91cd'])

            ws_m[f'A{dr}'] = code
            ws_m[f'B{dr}'] = tp_type
            ws_m[f'C{dr}'] = weight

            ci = 4
            for col_name in export_cols:
                ws_m.cell(row=dr, column=ci, value=_f(mat[col_name] if col_name in mat.keys() else None))
                ci += 1

            ws_m.cell(row=dr, column=ci, value=mat['\u89c4\u683c\u8bf4\u660e'] or '')
            ci += 1
            ws_m.cell(row=dr, column=ci, value=mat['\u5de5\u7a0b\u54c1\u540d'] or '')
            ci += 1
            ws_m.cell(row=dr, column=ci, value=mat['\u8ba1\u4ef7\u5355\u4f4d'] or '')
            ci += 1
            ws_m.cell(row=dr, column=ci, value=mat['\u5b9a\u4ef7\u5c5e\u6027'] or '')

        for idx, column_cells in enumerate(ws_m.columns, 1):
            max_len = 0
            for cell in column_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws_m.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 8), 25)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        download_name = f'\u4e34\u65f6\u4ef7\u683c\u8868_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            buffer,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        c.close()


@temp_price_bp.get('/materials')
def get_materials():
    ensure_temp_price_access()
    page = max(1, request.args.get('page', 1, type=int))
    ps = min(200, max(1, request.args.get('pageSize', 50, type=int)))
    c = _conn()
    try:
        _ensure_material_cols(c)
        total = c.execute(f'SELECT COUNT(*) FROM {TBL_M}').fetchone()[0]
        pc = ', '.join(f'"{x}"' for x in PRICE_COLS)
        rows = c.execute(
            f'SELECT "\u5de5\u7a0b\u7f16\u7801","\u89c4\u683c\u8bf4\u660e","\u5de5\u7a0b\u54c1\u540d","\u8ba1\u4ef7\u5355\u4f4d","\u5355\u91cd","\u5b9a\u4ef7\u5c5e\u6027","\u5428\u4ef7\u7c7b\u578b",{pc} FROM {TBL_M} ORDER BY "\u5de5\u7a0b\u7f16\u7801" LIMIT ? OFFSET ?',
            [ps, (page - 1) * ps]
        ).fetchall()
        tp = (total + ps - 1) // ps if total else 0
        return jsonify({
            'success': True,
            'materials': [{k: r[k] for k in r.keys()} for r in rows],
            'total': total, 'page': page, 'pageSize': ps, 'totalPages': tp,
        })
    finally:
        c.close()



def _f(v):
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


@temp_price_bp.post('/update')
def update_prices():
    ensure_temp_price_access(write=True)
    body = request.get_json(silent=True) or {}
    d = body.get('settings') or {}

    c = _conn()
    try:
        all_types = _get_all_ton_types(c)
        dynamic_keys = _all_keys_for_types(all_types)
        _ensure_settings(c, dynamic_keys)
        _ensure_material_cols(c)

        vals = []
        set_parts = []
        for k in dynamic_keys:
            set_parts.append(f'{k}=?')
            vals.append(_f(d.get(k)))
        set_parts.append("updated_at=datetime('now','localtime')")
        vals.append(1)
        c.execute(f'UPDATE {TBL_S} SET {",".join(set_parts)} WHERE id=?', vals)

        s = c.execute(f'SELECT * FROM {TBL_S} WHERE id=1').fetchone()
        s_cols = s.keys()

        rows = c.execute(f'SELECT "\u5de5\u7a0b\u7f16\u7801","\u5355\u91cd" FROM {TBL_M}').fetchall()
        updated = 0
        for r in rows:
            w = _f(r['\u5355\u91cd'])
            if w <= 0:
                continue

            code = r['\u5de5\u7a0b\u7f16\u7801'] or ''
            if code.startswith('FEPJ'):
                tp_type = code.replace('-', '_')
            else:
                tp_type = code.split('-')[0] if '-' in code else code
            is_pj = tp_type.startswith('FEPJ_')
            price_vals = {}

            for pk in PACK_KEYS:
                for side, ton in VALID_SIDE_TON:
                    ti = MAT_TON_TIERS.index(ton)
                    ton_settings_key = TON_TIERS[ti]
                    for li, length in enumerate(MAT_LEN_TIERS):
                        len_settings_key = TIERS[li]
                        for cur in CURRENCIES:
                            ck = cur['key']
                            has_rate = cur['has_rate']
                            col = PRICE_COL_MAP[(side, ton, length, ck, pk)]

                            if side == 'int':
                                if is_pj:
                                    k_int = f'ton_{tp_type}_int_{ton_settings_key}_{pk}'
                                else:
                                    k_int = f'ton_{tp_type}_int_{len_settings_key}_{ton_settings_key}_{pk}'
                                ton_int = _f(s[k_int] if k_int in s_cols else None)
                                if has_rate:
                                    ex = _f(s[f'exchange_rate_{ck}'] if f'exchange_rate_{ck}' in s_cols else None)
                                    pt = _f(s[f'points_{ck}'] if f'points_{ck}' in s_cols else None)
                                    price_vals[col] = round(w * ton_int / ex / pt, 6) if ex > 0 and pt > 0 else 0
                                else:
                                    price_vals[col] = round(w * ton_int, 6)
                            else:
                                if is_pj:
                                    k_ext = f'ton_{tp_type}_ext_50999_{pk}'
                                else:
                                    k_ext = f'ton_{tp_type}_ext_{len_settings_key}_50999_{pk}'
                                ton_ext = _f(s[k_ext] if k_ext in s_cols else None)
                                if has_rate:
                                    ex = _f(s[f'exchange_rate_{ck}'] if f'exchange_rate_{ck}' in s_cols else None)
                                    pt = _f(s[f'points_{ck}'] if f'points_{ck}' in s_cols else None)
                                    price_vals[col] = round(w * ton_ext / ex / pt, 6) if ex > 0 and pt > 0 else 0
                                else:
                                    price_vals[col] = round(w * ton_ext, 6)

            set_sql = ', '.join(f'"{col}"=?' for col in price_vals)
            c.execute(
                f'UPDATE {TBL_M} SET {set_sql} WHERE "\u5de5\u7a0b\u7f16\u7801"=?',
                list(price_vals.values()) + [r['\u5de5\u7a0b\u7f16\u7801']]
            )
            updated += 1

        c.commit()
        return jsonify({'success': True, 'updated': updated})
    except Exception as e:
        c.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        c.close()


@temp_price_bp.get('/lookup')
def lookup_material():
    ensure_temp_price_access()
    code = str(request.args.get('code') or '').strip()
    if not code:
        return jsonify({'success': False, 'message': '\u5de5\u7a0b\u7f16\u7801\u4e0d\u80fd\u4e3a\u7a7a'}), 400
    c = _conn()
    try:
        row = c.execute('SELECT * FROM aluminum_pricing WHERE "\u5de5\u7a0b\u7f16\u7801"=?', [code]).fetchone()
        if not row:
            return jsonify({'success': False, 'message': '\u7f16\u7801\u4e0d\u5b58\u5728'}), 404
        cols = row.keys()
        spec = ''
        for col in cols:
            if '\u89c4\u683c\u8bf4\u660e' in col:
                spec = row[col] or ''
                break

        def safe_float(val):
            try:
                v = float(val)
                return v if v > 0 else None
            except (ValueError, TypeError):
                return None

        weight = ''
        for wcol in ['\u5355\u91cd', '\u7c73\u91cd/km', '\u53c2\u8003\u91cd\u91cf']:
            val = row[wcol] if wcol in cols else None
            parsed = safe_float(val)
            if parsed is not None:
                weight = str(parsed)
                break

        ton_type = code.split('-')[0] if '-' in code else code
        for tp in STANDARD_TON_TYPES:
            if tp['key'].startswith('FEPJ_'):
                label_clean = tp['label'].replace('_', '').replace('-', '')
                code_clean = code.replace('-', '').replace('_', '')
                if code_clean.startswith(label_clean.replace('\u914d\u4ef6', '')):
                    ton_type = tp['key']
                    break
            else:
                if code.startswith(tp['key']):
                    ton_type = tp['key']
                    break

        return jsonify({
            'success': True,
            'data': {
                '\u89c4\u683c\u8bf4\u660e': spec,
                '\u5de5\u7a0b\u54c1\u540d': row['\u5de5\u7a0b\u54c1\u540d'] if '\u5de5\u7a0b\u54c1\u540d' in cols else '',
                '\u8ba1\u4ef7\u5355\u4f4d': row['\u8ba1\u4ef7\u5355\u4f4d'] if '\u8ba1\u4ef7\u5355\u4f4d' in cols else '\u7c73',
                '\u5355\u91cd': weight,
                '\u5b9a\u4ef7\u5c5e\u6027': row['\u5b9a\u4ef7\u5c5e\u6027'] if '\u5b9a\u4ef7\u5c5e\u6027' in cols else '',
                '\u5428\u4ef7\u7c7b\u578b': ton_type,
            },
        })
    finally:
        c.close()


@temp_price_bp.post('/material')
def add_material():
    ensure_temp_price_access(write=True)
    d = request.get_json(silent=True) or {}
    code = str(d.get('\u5de5\u7a0b\u7f16\u7801') or '').strip()
    if not code:
        return jsonify({'success': False, 'message': '\u5de5\u7a0b\u7f16\u7801\u4e0d\u80fd\u4e3a\u7a7a'}), 400
    c = _conn()
    try:
        if c.execute(f'SELECT COUNT(*) FROM {TBL_M} WHERE "\u5de5\u7a0b\u7f16\u7801"=?', [code]).fetchone()[0] > 0:
            return jsonify({'success': False, 'message': f'\u5de5\u7a0b\u7f16\u7801 {code} \u5df2\u5b58\u5728'}), 400
        c.execute(
            f'INSERT INTO {TBL_M} ("\u5de5\u7a0b\u7f16\u7801","\u89c4\u683c\u8bf4\u660e","\u5de5\u7a0b\u54c1\u540d","\u8ba1\u4ef7\u5355\u4f4d","\u5355\u91cd","\u5b9a\u4ef7\u5c5e\u6027","\u5428\u4ef7\u7c7b\u578b") VALUES (?,?,?,?,?,?,?)',
            (code, str(d.get('\u89c4\u683c\u8bf4\u660e') or '').strip(), str(d.get('\u5de5\u7a0b\u54c1\u540d') or '').strip(),
             str(d.get('\u8ba1\u4ef7\u5355\u4f4d') or '').strip(), str(d.get('\u5355\u91cd') or '').strip(),
             str(d.get('\u5b9a\u4ef7\u5c5e\u6027') or '').strip(), str(d.get('\u5428\u4ef7\u7c7b\u578b') or 'PJ').strip()))
        c.commit()
        return jsonify({'success': True})
    except Exception as e:
        c.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        c.close()


@temp_price_bp.put('/material/<code>')
def edit_material(code):
    ensure_temp_price_access(write=True)
    d = request.get_json(silent=True) or {}
    c = _conn()
    try:
        if c.execute(f'SELECT COUNT(*) FROM {TBL_M} WHERE "\u5de5\u7a0b\u7f16\u7801"=?', [code]).fetchone()[0] == 0:
            return jsonify({'success': False, 'message': '\u7269\u6599\u4e0d\u5b58\u5728'}), 404
        fields = []
        vals = []
        for col in ('\u89c4\u683c\u8bf4\u660e', '\u5de5\u7a0b\u54c1\u540d', '\u8ba1\u4ef7\u5355\u4f4d', '\u5355\u91cd', '\u5b9a\u4ef7\u5c5e\u6027', '\u5428\u4ef7\u7c7b\u578b'):
            if col in d:
                fields.append(f'"{col}"=?')
                vals.append(str(d[col] or '').strip())
        if not fields:
            return jsonify({'success': True})
        vals.append(code)
        c.execute(f'UPDATE {TBL_M} SET {",".join(fields)} WHERE "\u5de5\u7a0b\u7f16\u7801"=?', vals)
        c.commit()
        return jsonify({'success': True})
    except Exception as e:
        c.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        c.close()


@temp_price_bp.post('/materials/delete')
def delete_materials():
    ensure_temp_price_access(write=True)
    d = request.get_json(silent=True) or {}
    codes = d.get('codes') or []
    if not codes:
        return jsonify({'success': True, 'deleted': 0})
    c = _conn()
    try:
        ph = ','.join(['?'] * len(codes))
        cur = c.execute(f'DELETE FROM {TBL_M} WHERE "\u5de5\u7a0b\u7f16\u7801" IN ({ph})', codes)
        c.commit()
        return jsonify({'success': True, 'deleted': cur.rowcount})
    except Exception as e:
        c.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        c.close()


@temp_price_bp.get('/export-ton')
def export_ton_prices():
    ensure_temp_price_access()
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    conn = _conn()
    try:
        _ensure_settings(conn)
        s = conn.execute(f'SELECT * FROM {TBL_S} WHERE id=1').fetchone()
        s_cols = s.keys()

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        hdr_fill = PatternFill(fill_type='solid', start_color='DDEBF7', end_color='DDEBF7')
        hdr_font = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='D9E2F3')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        all_types = _get_all_ton_types_with_settings(conn)

        # pack 查询参数：仅导出指定包装的单 Sheet（与当前录入包装一致，便于导入回写）
        query_pack = str(request.args.get('pack') or '').strip().lower()
        export_packs = [p for p in PACK_TYPES if p['key'] == query_pack] if query_pack in PACK_KEYS else PACK_TYPES

        for pk in export_packs:
            pack_key = pk['key']
            ws = wb.create_sheet(title=pk['label'])

            ws.merge_cells('A1:A2')
            ws['A1'] = '\u7f16\u7801'
            ws['A1'].font = hdr_font
            ws['A1'].fill = hdr_fill
            ws['A1'].alignment = center
            ws['A1'].border = border
            ws['A2'].border = border

            group_headers = [
                ('\u5185\u90e8 0-5\u5428', 3),
                ('\u5185\u90e8 5-50\u5428', 3),
                ('\u5185\u90e8 50-999\u5428', 3),
                ('\u5916\u90e8 50-999\u5428', 3),
            ]
            col = 2
            for label, span in group_headers:
                cell = ws.cell(row=1, column=col)
                cell.value = label
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = center
                cell.border = border
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
                for c2 in range(col, col + span):
                    ws.cell(row=1, column=c2).border = border
                col += span

            sub_headers = ['0-1', '1-3', '3+'] * 4
            for i, sh in enumerate(sub_headers):
                cell = ws.cell(row=2, column=2 + i)
                cell.value = sh
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = center
                cell.border = border

            for ri, tp in enumerate(all_types):
                row = 3 + ri
                k = tp['key']
                is_pj = k.startswith('FEPJ_')
                ws.cell(row=row, column=1, value=tp['label']).border = border
                ws.cell(row=row, column=1).font = Font(bold=True)
                ci = 2
                for gi in range(4):
                    if is_pj:
                        if gi < 3:
                            key = f'ton_{k}_int_{TON_TIERS[gi]}_{pack_key}'
                        else:
                            key = f'ton_{k}_ext_50999_{pack_key}'
                        raw = s[key] if key in s_cols else None
                        val = float(raw) if raw is not None else None
                        ws.merge_cells(start_row=row, start_column=ci, end_row=row, end_column=ci + 2)
                        cell = ws.cell(row=row, column=ci, value=val)
                        cell.border = border
                        cell.alignment = center
                        for c2 in range(ci, ci + 3):
                            ws.cell(row=row, column=c2).border = border
                        ci += 3
                    else:
                        if gi < 3:
                            tt = TON_TIERS[gi]
                            for lt in TIERS:
                                key = f'ton_{k}_int_{lt}_{tt}_{pack_key}'
                                raw = s[key] if key in s_cols else None
                                val = float(raw) if raw is not None else None
                                cell = ws.cell(row=row, column=ci, value=val)
                                cell.border = border
                                cell.alignment = center
                                ci += 1
                        else:
                            for lt in TIERS:
                                key = f'ton_{k}_ext_{lt}_50999_{pack_key}'
                                raw = s[key] if key in s_cols else None
                                val = float(raw) if raw is not None else None
                                cell = ws.cell(row=row, column=ci, value=val)
                                cell.border = border
                                cell.alignment = center
                                ci += 1

            ws.column_dimensions['A'].width = 14
            for c2 in range(2, 14):
                ws.column_dimensions[get_column_letter(c2)].width = 12

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'\u5428\u4ef7\u5e95\u8868_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()


@temp_price_bp.post('/import-ton')
def import_ton_prices():
    ensure_temp_price_access(write=True)
    import openpyxl

    file_storage = request.files.get('file')
    if not file_storage or not getattr(file_storage, 'filename', ''):
        return jsonify({'success': False, 'message': '\u8bf7\u9009\u62e9\u6587\u4ef6'}), 400

    filename = file_storage.filename.lower()
    if not filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '\u4ec5\u652f\u6301xlsx\u683c\u5f0f'}), 400

    conn = _conn()
    try:
        _ensure_settings(conn)

        wb = openpyxl.load_workbook(file_storage, data_only=True)

        updated_keys = {}

        # 包装归属优先级：前端传入的 pack 查询参数（当前选中包装）> Sheet 名（简易包装/铁托）
        # > 旧版单 Sheet 默认简易包装
        query_pack = str(request.args.get('pack') or '').strip().lower()
        forced_pack = query_pack if query_pack in PACK_KEYS else None

        # 每个 Sheet 对应一种包装：Sheet 名称为包装标签（简易包装/铁托），
        # 兼容旧版单 Sheet（无包装标识）→ 按默认简易包装导入
        for ws in wb.worksheets:
            if forced_pack:
                pack_key = forced_pack
            else:
                sheet_name = str(ws.title or '').strip()
                pack_key = None
                for pk in PACK_TYPES:
                    if sheet_name == pk['label'] or sheet_name == pk['key']:
                        pack_key = pk['key']
                        break
                if pack_key is None and wb.index(ws) == 0:
                    pack_key = DEFAULT_PACK  # 旧版单表 → 简易包装

            for row_idx in range(3, ws.max_row + 1):
                code_cell = ws.cell(row=row_idx, column=1).value
                if not code_cell:
                    continue
                code_str = str(code_cell).strip()

                tp_match = None
                for tp in STANDARD_TON_TYPES:
                    if tp['label'] == code_str or tp['key'] == code_str:
                        tp_match = tp
                        break

                if tp_match:
                    k = tp_match['key']
                    is_pj = k.startswith('FEPJ_')
                else:
                    if code_str.startswith('FEPJ'):
                        k = code_str.replace('-', '_')
                        is_pj = True
                    elif _is_valid_ton_key(code_str):
                        k = code_str
                        is_pj = False
                    else:
                        continue
                    if not _is_valid_ton_key(k):
                        continue

                ci = 2

                if is_pj:
                    for tt in TON_TIERS:
                        val = ws.cell(row=row_idx, column=ci).value
                        key = f'ton_{k}_int_{tt}_{pack_key}'
                        if val is not None:
                            updated_keys[key] = _f(val)
                        ci += 3
                    val = ws.cell(row=row_idx, column=ci).value
                    key = f'ton_{k}_ext_50999_{pack_key}'
                    if val is not None:
                        updated_keys[key] = _f(val)
                else:
                    for tt in TON_TIERS:
                        for lt in TIERS:
                            val = ws.cell(row=row_idx, column=ci).value
                            key = f'ton_{k}_int_{lt}_{tt}_{pack_key}'
                            if val is not None:
                                updated_keys[key] = _f(val)
                            ci += 1
                    for lt in TIERS:
                        val = ws.cell(row=row_idx, column=ci).value
                        key = f'ton_{k}_ext_{lt}_50999_{pack_key}'
                        if val is not None:
                            updated_keys[key] = _f(val)
                        ci += 1

        if updated_keys:
            _ensure_settings(conn, list(updated_keys.keys()))
            set_parts = [f'"{sk}"=?' for sk in updated_keys]
            vals = list(updated_keys.values())
            vals.append(1)
            conn.execute(f'UPDATE {TBL_S} SET {",".join(set_parts)},updated_at=datetime(\'now\',\'localtime\') WHERE id=?', vals)
            conn.commit()

        return jsonify({'success': True, 'updated': len(updated_keys)})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()
