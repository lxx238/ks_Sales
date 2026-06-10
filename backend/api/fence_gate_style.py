import io
import re
from datetime import datetime

from flask import Blueprint, jsonify, request, send_file

from backend.repositories.fence_gate_style_repository import (
    batch_delete_fence_styles,
    batch_delete_gate_styles,
    bulk_upsert_fence_styles,
    bulk_upsert_gate_styles,
    create_fence_style,
    create_gate_style,
    delete_fence_style,
    delete_gate_style,
    get_fence_style,
    get_gate_style,
    list_all_fence_styles_raw,
    list_all_gate_styles_raw,
    list_fence_style_codes,
    list_fence_styles,
    list_gate_style_codes,
    list_gate_styles,
    update_fence_style,
    update_gate_style,
    FENCE_TABLE,
    GATE_TABLE,
)
from backend.services.auth_service import ensure_permission


fence_style_bp = Blueprint('fence_style', __name__, url_prefix='/api/fence-styles')
gate_style_bp = Blueprint('gate_style', __name__, url_prefix='/api/gate-styles')


FENCE_STYLE_COLUMNS = [
    'code', 'mesh_type', 'pipe_spec', 'base_type', 'height',
    'mesh_code', 'mesh_thick_code', 'post_code', 'pile_code',
    'end_cap_code', 'rubber_code',
]
GATE_STYLE_COLUMNS = [
    'code', 'gate_type', 'width', 'height', 'base_type',
    'mesh_shape', 'install_type', 'mesh_base_code',
    'buckle_code', 'bolt_code', 'end_cap_code',
    'horizontal_pin_code', 'vertical_pin_code',
    'pile_code', 'pile_bolt_code', 'rubber_code',
    'buckle_qty', 'bolt_qty', 'end_cap_qty',
    'horizontal_pin_qty', 'vertical_pin_qty',
    'pile_qty', 'pile_bolt_qty', 'rubber_qty',
]


@fence_style_bp.route('', methods=['GET'])
def api_list_fence_styles():
    ensure_permission('database')
    mesh_type = request.args.get('mesh_type', '').strip()
    base_type = request.args.get('base_type', '').strip()
    height = request.args.get('height', '').strip()
    height = int(height) if height else None
    rows = list_fence_styles(mesh_type=mesh_type, base_type=base_type, height=height)
    return jsonify({'data': rows, 'total': len(rows)})


@fence_style_bp.route('/<code>', methods=['GET'])
def api_get_fence_style(code):
    ensure_permission('database')
    row = get_fence_style(code)
    if not row:
        return jsonify({'error': f'围栏款式不存在: {code}'}), 404
    return jsonify({'data': row})


@fence_style_bp.route('', methods=['POST'])
def api_create_fence_style():
    ensure_permission('database_submit')
    data = request.get_json(force=True)
    code = (data.get('code') or '').strip()
    if not code:
        return jsonify({'error': '编码不能为空'}), 400
    try:
        row = create_fence_style(
            code=code,
            mesh_type=(data.get('mesh_type') or '').strip(),
            pipe_spec=(data.get('pipe_spec') or '').strip(),
            base_type=(data.get('base_type') or '').strip(),
            height=int(data.get('height') or 0),
            mesh_code_30=(data.get('mesh_code') or '').strip(),
            mesh_code_35=(data.get('mesh_thick_code') or '').strip(),
            mesh_code_40=(data.get('mesh_code_40') or '').strip(),
            post_code=(data.get('post_code') or '').strip(),
            pile_code=data.get('pile_code') or None,
            end_cap_code=(data.get('end_cap_code') or '').strip(),
            rubber_code=data.get('rubber_code') or None,
            image_base64=(data.get('image_base64') or '').strip(),
        )
        return jsonify({'data': row, 'message': '新增成功'}), 201
    except ValueError as e:
        return jsonify({'error': str(e)}), 409
    except Exception as e:
        return jsonify({'error': f'创建失败: {e}'}), 500


@fence_style_bp.route('/<code>', methods=['PUT'])
def api_update_fence_style(code):
    ensure_permission('database_submit')
    data = request.get_json(force=True)
    if 'mesh_code' in data:
        data['mesh_code_30'] = data.pop('mesh_code')
    if 'mesh_thick_code' in data:
        data['mesh_code_35'] = data.pop('mesh_thick_code')
    try:
        new_code = data.pop('code', None) or code
        row = update_fence_style(code, new_code=new_code, **data)
        return jsonify({'data': row, 'message': '更新成功'})
    except LookupError as e:
        return jsonify({'error': str(e)}), 404
    except Exception as e:
        return jsonify({'error': f'更新失败: {e}'}), 500


@fence_style_bp.route('/<code>', methods=['DELETE'])
def api_delete_fence_style(code):
    ensure_permission('database_submit')
    if delete_fence_style(code):
        return jsonify({'message': f'已删除 {code}'})
    return jsonify({'error': f'围栏款式不存在: {code}'}), 404


@fence_style_bp.route('/bulk', methods=['POST'])
def api_bulk_fence_styles():
    ensure_permission('database_submit')
    data = request.get_json(force=True)
    items = data.get('items') if isinstance(data, dict) else data
    if not isinstance(items, list):
        return jsonify({'error': 'items 必须为数组'}), 400
    result = bulk_upsert_fence_styles(items)
    return jsonify({'success': True, 'data': result})


@fence_style_bp.route('/batch-delete', methods=['POST'])
def api_batch_delete_fence_styles():
    ensure_permission('database_submit')
    data = request.get_json(force=True)
    codes = data.get('codes') or []
    if not codes:
        return jsonify({'error': 'codes 不能为空'}), 400
    deleted = batch_delete_fence_styles(codes)
    return jsonify({'success': True, 'message': f'已删除 {deleted} 条围栏款式', 'deleted_count': deleted})


@fence_style_bp.route('/download', methods=['GET'])
def api_download_fence_styles():
    ensure_permission('database_download')
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        records = list_all_fence_styles_raw()
        columns = FENCE_STYLE_COLUMNS

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'fence_styles'
        ws.append(columns)

        for record in records:
            ws.append([record.get(c, '') if record.get(c) is not None else '' for c in columns])

        header_fill = PatternFill(fill_type='solid', start_color='DDEBF7', end_color='DDEBF7')
        header_font = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='D9E2F3')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        for idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(idx)].width = min(max(max_length + 2, 10), 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        download_name = f'fence_styles_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            buffer, as_attachment=True, download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as exc:
        return jsonify({'success': False, 'message': f'下载失败: {exc}'}), 500


@fence_style_bp.route('/batch-update', methods=['POST'])
def api_batch_update_fence_styles():
    ensure_permission('database_submit')
    file_storage = request.files.get('file')
    if not file_storage or not getattr(file_storage, 'filename', ''):
        return jsonify({'success': False, 'message': '请上传 Excel 文件'}), 400

    filename = file_storage.filename.lower()
    if not filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '仅支持 .xlsx / .xls 文件'}), 400

    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_storage, data_only=True)
        ws = wb.active

        header_row = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=c).value
            if val:
                header_row[c] = str(val).strip()

        code_col = None
        for c, name in header_row.items():
            if name == 'code':
                code_col = c
                break
        if code_col is None:
            return jsonify({'success': False, 'message': 'Excel 第一行未找到"code"列'}), 400

        items = []
        for r in range(2, ws.max_row + 1):
            code_val = ws.cell(row=r, column=code_col).value
            if not code_val or not str(code_val).strip():
                continue
            item = {}
            for c, name in header_row.items():
                cell_val = ws.cell(row=r, column=c).value
                if cell_val is not None:
                    item[name] = cell_val
            if item:
                items.append(item)

        if not items:
            return jsonify({'success': False, 'message': 'Excel 中没有可导入的数据'}), 400

        result = bulk_upsert_fence_styles(items)
        return jsonify({
            'success': True,
            'message': f'批量导入完成，成功 {result["success"]} 条，失败 {len(result["failed"])} 条',
            **result,
        })
    except Exception as exc:
        return jsonify({'success': False, 'message': f'批量更新失败: {exc}'}), 500


@gate_style_bp.route('', methods=['GET'])
def api_list_gate_styles():
    ensure_permission('database')
    gate_type = request.args.get('gate_type', '').strip()
    base_type = request.args.get('base_type', '').strip()
    width = request.args.get('width', '').strip()
    height = request.args.get('height', '').strip()
    width = int(width) if width else None
    height = int(height) if height else None
    rows = list_gate_styles(gate_type=gate_type, base_type=base_type, width=width, height=height)
    return jsonify({'data': rows, 'total': len(rows)})


@gate_style_bp.route('/<code>', methods=['GET'])
def api_get_gate_style(code):
    ensure_permission('database')
    row = get_gate_style(code)
    if not row:
        return jsonify({'error': f'门款式不存在: {code}'}), 404
    return jsonify({'data': row})


@gate_style_bp.route('', methods=['POST'])
def api_create_gate_style():
    ensure_permission('database_submit')
    data = request.get_json(force=True)
    code = (data.get('code') or '').strip()
    if not code:
        return jsonify({'error': '编码不能为空'}), 400
    try:
        def _int(val, default=0):
            try:
                return int(val)
            except (ValueError, TypeError):
                return default

        row = create_gate_style(
            code=code,
            gate_type=(data.get('gate_type') or '').strip(),
            width=_int(data.get('width'), 0),
            height=_int(data.get('height'), 0),
            base_type=(data.get('base_type') or '').strip(),
            mesh_shape=(data.get('mesh_shape') or '折半圆').strip(),
            install_type=(data.get('install_type') or '常规').strip(),
            mesh_base_code=(data.get('mesh_base_code') or '').strip(),
            buckle_code=(data.get('buckle_code') or 'FN-PJ-0002').strip(),
            bolt_code=(data.get('bolt_code') or 'FN-PJ-0004').strip(),
            end_cap_code=(data.get('end_cap_code') or 'XJ-0009').strip(),
            horizontal_pin_code=(data.get('horizontal_pin_code') or '').strip(),
            vertical_pin_code=(data.get('vertical_pin_code') or '').strip(),
            pile_code=data.get('pile_code') or None,
            pile_bolt_code=(data.get('pile_bolt_code') or '').strip(),
            rubber_code=data.get('rubber_code') or None,
            buckle_qty=_int(data.get('buckle_qty'), 0),
            bolt_qty=_int(data.get('bolt_qty'), 0),
            end_cap_qty=_int(data.get('end_cap_qty'), 2),
            horizontal_pin_qty=_int(data.get('horizontal_pin_qty'), 0),
            vertical_pin_qty=_int(data.get('vertical_pin_qty'), 0),
            pile_qty=_int(data.get('pile_qty'), 0),
            pile_bolt_qty=_int(data.get('pile_bolt_qty'), 0),
            rubber_qty=_int(data.get('rubber_qty'), 0),
            image_base64=(data.get('image_base64') or '').strip(),
        )
        return jsonify({'data': row, 'message': '新增成功'}), 201
    except ValueError as e:
        return jsonify({'error': str(e)}), 409
    except Exception as e:
        return jsonify({'error': f'创建失败: {e}'}), 500


@gate_style_bp.route('/<code>', methods=['PUT'])
def api_update_gate_style(code):
    ensure_permission('database_submit')
    data = request.get_json(force=True)
    try:
        int_fields = ['width', 'height', 'buckle_qty', 'bolt_qty', 'end_cap_qty',
                       'horizontal_pin_qty', 'vertical_pin_qty', 'pile_qty', 'pile_bolt_qty', 'rubber_qty']
        for f in int_fields:
            if f in data and not isinstance(data[f], int):
                try:
                    data[f] = int(data[f])
                except (ValueError, TypeError):
                    data[f] = 0
        new_code = data.pop('code', None) or code
        row = update_gate_style(code, new_code=new_code, **data)
        return jsonify({'data': row, 'message': '更新成功'})
    except LookupError as e:
        return jsonify({'error': str(e)}), 404
    except Exception as e:
        return jsonify({'error': f'更新失败: {e}'}), 500


@gate_style_bp.route('/<code>', methods=['DELETE'])
def api_delete_gate_style(code):
    ensure_permission('database_submit')
    if delete_gate_style(code):
        return jsonify({'message': f'已删除 {code}'})
    return jsonify({'error': f'门款式不存在: {code}'}), 404


@gate_style_bp.route('/bulk', methods=['POST'])
def api_bulk_gate_styles():
    ensure_permission('database_submit')
    data = request.get_json(force=True)
    items = data.get('items') if isinstance(data, dict) else data
    if not isinstance(items, list):
        return jsonify({'error': 'items 必须为数组'}), 400
    result = bulk_upsert_gate_styles(items)
    return jsonify({'success': True, 'data': result})


@gate_style_bp.route('/batch-delete', methods=['POST'])
def api_batch_delete_gate_styles():
    ensure_permission('database_submit')
    data = request.get_json(force=True)
    codes = data.get('codes') or []
    if not codes:
        return jsonify({'error': 'codes 不能为空'}), 400
    deleted = batch_delete_gate_styles(codes)
    return jsonify({'success': True, 'message': f'已删除 {deleted} 条门款式', 'deleted_count': deleted})


@gate_style_bp.route('/download', methods=['GET'])
def api_download_gate_styles():
    ensure_permission('database_download')
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        records = list_all_gate_styles_raw()
        columns = GATE_STYLE_COLUMNS

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'gate_styles'
        ws.append(columns)

        for record in records:
            ws.append([record.get(c, '') if record.get(c) is not None else '' for c in columns])

        header_fill = PatternFill(fill_type='solid', start_color='DDEBF7', end_color='DDEBF7')
        header_font = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='D9E2F3')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        for idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(idx)].width = min(max(max_length + 2, 10), 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        download_name = f'gate_styles_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            buffer, as_attachment=True, download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as exc:
        return jsonify({'success': False, 'message': f'下载失败: {exc}'}), 500


@gate_style_bp.route('/batch-update', methods=['POST'])
def api_batch_update_gate_styles():
    ensure_permission('database_submit')
    file_storage = request.files.get('file')
    if not file_storage or not getattr(file_storage, 'filename', ''):
        return jsonify({'success': False, 'message': '请上传 Excel 文件'}), 400

    filename = file_storage.filename.lower()
    if not filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '仅支持 .xlsx / .xls 文件'}), 400

    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_storage, data_only=True)
        ws = wb.active

        header_row = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=c).value
            if val:
                header_row[c] = str(val).strip()

        code_col = None
        for c, name in header_row.items():
            if name == 'code':
                code_col = c
                break
        if code_col is None:
            return jsonify({'success': False, 'message': 'Excel 第一行未找到"code"列'}), 400

        items = []
        for r in range(2, ws.max_row + 1):
            code_val = ws.cell(row=r, column=code_col).value
            if not code_val or not str(code_val).strip():
                continue
            item = {}
            for c, name in header_row.items():
                cell_val = ws.cell(row=r, column=c).value
                if cell_val is not None:
                    item[name] = cell_val
            if item:
                items.append(item)

        if not items:
            return jsonify({'success': False, 'message': 'Excel 中没有可导入的数据'}), 400

        result = bulk_upsert_gate_styles(items)
        return jsonify({
            'success': True,
            'message': f'批量导入完成，成功 {result["success"]} 条，失败 {len(result["failed"])} 条',
            **result,
        })
    except Exception as exc:
        return jsonify({'success': False, 'message': f'批量更新失败: {exc}'}), 500
