import openpyxl
from openpyxl.utils import get_column_letter

try:
    from openpyxl.utils import column_width_to_pixels, row_height_to_pixels
    _use_utils = True
except ImportError:
    _use_utils = False


def _col_to_px(ws, col_idx):
    col_letter = get_column_letter(col_idx)
    width = ws.column_dimensions[col_letter].width
    if width is None:
        width = 8.43
    if _use_utils:
        try:
            return column_width_to_pixels(width)
        except Exception:
            pass
    return width * 7.5


def _row_to_px(ws, row_idx):
    height = ws.row_dimensions[row_idx].height
    if height is None:
        height = 15
    if _use_utils:
        try:
            return row_height_to_pixels(height)
        except Exception:
            pass
    return height * 1.33


def center_images_in_column_d(excel_path, output_path=None):
    if output_path is None:
        output_path = excel_path

    workbook = openpyxl.load_workbook(excel_path)
    EMU_PER_PIXEL = 9525
    target_col = 4

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        for image in list(worksheet._images):
            anchor = image.anchor
            if hasattr(anchor, '_from'):
                from_col = anchor._from.col + 1
                from_row = anchor._from.row + 1
            elif hasattr(anchor, 'from_col'):
                from_col = anchor.from_col + 1
                from_row = anchor.from_row + 1
            else:
                continue

            if from_col != target_col:
                continue

            cell_width_px = _col_to_px(worksheet, from_col)
            cell_height_px = _row_to_px(worksheet, from_row)

            img_width = image.width
            img_height = image.height

            offset_x = max(0, (cell_width_px - img_width) / 2)
            offset_y = max(0, (cell_height_px - img_height) / 2)
            offset_x_emu = int(offset_x * EMU_PER_PIXEL)
            offset_y_emu = int(offset_y * EMU_PER_PIXEL)

            if hasattr(anchor, '_from'):
                anchor._from.colOff = offset_x_emu
                anchor._from.rowOff = offset_y_emu
            elif hasattr(anchor, 'from_col_off'):
                anchor.from_col_off = offset_x_emu
                anchor.from_row_off = offset_y_emu

    workbook.save(output_path)
    return workbook


def center_images_advanced(excel_path, output_path=None, scale_to_fit=True, max_scale=1.0):
    if output_path is None:
        output_path = excel_path.replace('.xlsx', '_centered.xlsx')

    workbook = openpyxl.load_workbook(excel_path)
    EMU_PER_PIXEL = 9525
    target_col = 4

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        for image in list(worksheet._images):
            anchor = image.anchor
            if hasattr(anchor, '_from'):
                from_col = anchor._from.col + 1
                from_row = anchor._from.row + 1
            elif hasattr(anchor, 'from_col'):
                from_col = anchor.from_col + 1
                from_row = anchor.from_row + 1
            else:
                continue

            if from_col != target_col:
                continue

            cell_width = _col_to_px(worksheet, from_col)
            cell_height = _row_to_px(worksheet, from_row)
            img_width = image.width
            img_height = image.height

            if scale_to_fit and img_width > 0 and img_height > 0:
                scale = min(cell_width / img_width, cell_height / img_height, max_scale)
                image.width = int(img_width * scale)
                image.height = int(img_height * scale)

            offset_x = max(0, (cell_width - image.width) / 2)
            offset_y = max(0, (cell_height - image.height) / 2)

            if hasattr(anchor, '_from'):
                anchor._from.colOff = int(offset_x * EMU_PER_PIXEL)
                anchor._from.rowOff = int(offset_y * EMU_PER_PIXEL)
            elif hasattr(anchor, 'from_col_off'):
                anchor.from_col_off = int(offset_x * EMU_PER_PIXEL)
                anchor.from_row_off = int(offset_y * EMU_PER_PIXEL)

    workbook.save(output_path)
    return workbook
