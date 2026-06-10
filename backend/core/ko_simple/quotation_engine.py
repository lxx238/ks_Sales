import os
import re
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins

from backend.core.shared.bom_utils import (
    read_bom_from_dataframe,
    parse_array_to_rows_cols,
    resolve_products_and_array,
)
from backend.core.shared.price_utils import (
    resolve_price_info,
    has_valid_price_info,
    round_to_2_decimal,
    _get_discount_category,
    _get_discount_rate,
)
from backend.core.shared.weight_utils import (
    extract_length_from_spec,
    WEIGHT_BY_LENGTH_ATTRIBUTES,
)
from backend.core.shared.text_utils import (
    normalize_lookup_code,
    extract_main_name,
    parse_decimal_number,
    _strip_cjk_spec,
    normalize_angle,
)
from backend.core.shared.product_utils import (
    _is_valid_product_code,
    _match_exclude_group,
)
from backend.core.shared.constants import (
    EXCLUDE_ITEM_GROUPS,
    CARBON_STEEL_PRICING_ATTRS,
    IMAGE_WIDTH,
    IMAGE_HEIGHT,
    IMAGE_COLUMN_INDEX,
    IMAGE_PADDING,
)
from backend.core.shared.image_utils import (
    prepare_image_for_excel,
    add_image_centered_in_cell,
)
from backend.core.material_translate import translate_material, adjust_material_by_coating

CURRENCY_FMT = '"US$" #,##0.00'


def _set_cell(ws, row, col, val, font=None, align=None, border=None, number_format=None, fill=None):
    cell = ws.cell(row=row, column=col, value=val)
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format
    if fill is not None:
        cell.fill = fill
    return cell


def _is_carbon_steel_cached(product, price_info):
    if price_info:
        attr = str(price_info.get('pricing_attribute', '')).strip().upper()
        if attr in CARBON_STEEL_PRICING_ATTRS:
            return True
    return False


def _classify_products_single_pass(all_products, price_mapping, delete_options, always_exclude, ko_exclude_options=None):
    delete_options = delete_options or {}
    ko_exclude_options = ko_exclude_options or {}
    all_exclude_opts = {k: True for k in EXCLUDE_ITEM_GROUPS} if always_exclude else {}
    for k in list(ko_exclude_options.keys()) + list(delete_options.keys()):
        all_exclude_opts[k] = True
    aluminum = []
    carbon_steel = []
    excluded = []
    pile_products = []
    price_info_cache = {}
    for p in all_products:
        code = str(p.get('code', '')).strip()
        spec = str(p.get('spec', '')).strip()
        cache_key = (code, spec) if spec else code
        if cache_key not in price_info_cache:
            price_info_cache[cache_key] = resolve_price_info(price_mapping, code, spec=spec) if price_mapping else None
        pi = price_info_cache[cache_key]
        p['_price_info'] = pi
        attr = (pi.get('attribute', '') if pi else '') or ''
        if attr.strip() in ('地桩', '地盤杭'):
            p['_is_pile'] = True
            pile_products.append(p)
            continue
        if not _is_valid_product_code(code):
            continue
        qty = p.get('quantity', 0)
        if not qty or qty <= 0:
            continue
        matched_group = _match_exclude_group(p, price_mapping, all_exclude_opts)
        is_cs = _is_carbon_steel_cached(p, pi)
        if matched_group:
            if delete_options.get(matched_group):
                continue
            if ko_exclude_options.get(matched_group):
                excluded.append(p)
            else:
                if is_cs:
                    carbon_steel.append(p)
                else:
                    aluminum.append(p)
        elif is_cs:
            carbon_steel.append(p)
        else:
            aluminum.append(p)
    return aluminum, carbon_steel, excluded, pile_products, price_info_cache


def _is_carbon_steel(product, price_mapping):
    price_info = resolve_price_info(price_mapping, product.get('code', ''), spec=product.get('spec', ''))
    if price_info:
        attr = str(price_info.get('pricing_attribute', '')).strip().upper()
        if attr in CARBON_STEEL_PRICING_ATTRS:
            return True
    return False


def _group_products_for_ksd(all_products, price_mapping, delete_options, always_exclude, ko_exclude_options=None):
    delete_options = delete_options or {}
    ko_exclude_options = ko_exclude_options or {}
    all_exclude_opts = {k: True for k in EXCLUDE_ITEM_GROUPS} if always_exclude else {}
    for k in list(ko_exclude_options.keys()) + list(delete_options.keys()):
        all_exclude_opts[k] = True
    aluminum = []
    carbon_steel = []
    excluded = []
    for p in all_products:
        code = str(p.get('code', '')).strip()
        if not _is_valid_product_code(code):
            continue
        qty = p.get('quantity', 0)
        if not qty or qty <= 0:
            continue
        if p.get('_is_pile'):
            aluminum.append(p)
            continue
        matched_group = _match_exclude_group(p, price_mapping, all_exclude_opts)
        if matched_group:
            if delete_options.get(matched_group):
                continue
            if ko_exclude_options.get(matched_group):
                excluded.append(p)
            else:
                if _is_carbon_steel(p, price_mapping):
                    carbon_steel.append(p)
                else:
                    aluminum.append(p)
        elif _is_carbon_steel(p, price_mapping):
            carbon_steel.append(p)
        else:
            aluminum.append(p)
    return aluminum, carbon_steel, excluded


def _calc_display_price(product, price_mapping):
    code = product.get('code', '')
    price_info = resolve_price_info(price_mapping, code, spec=product.get('spec', ''))
    if not price_info or not has_valid_price_info(price_info):
        return 0, '', price_info
    base_price = float(price_info['price'])
    price_unit = price_info.get('unit', '')
    is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
    if is_meter:
        length_mm = Decimal(str(extract_length_from_spec(product.get('spec', '')) or 0))
        if length_mm > 0:
            display_price = float(Decimal(str(base_price)) * length_mm / Decimal('1000'))
        else:
            display_price = base_price
    else:
        display_price = base_price
    return display_price, price_unit, price_info


def _render_subtable(ws, start_row, products, price_mapping, col_count=12, currency_label='US$'):
    if not products:
        return start_row

    normal_font = Font(name='Calibri', size=10)
    bold_font = Font(name='Calibri', size=10, bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    header_labels = [
        'Item No.', 'Product Name', 'Material', 'Picture', 'Spec.',
        f'Unit Price ({currency_label})\nEXW', 'QTY (PCS)\nEXW', f'Total price({currency_label})', 'REMARK',
    ]
    header_row = start_row
    for ci, label in enumerate(header_labels):
        cell = ws.cell(row=header_row, column=ci + 1, value=label)
        cell.font = bold_font
        cell.alignment = center
        cell.border = thin_border
    ws.cell(row=header_row, column=12, value='单价').font = bold_font
    ws.cell(row=header_row, column=12).alignment = center
    ws.cell(row=header_row, column=12).border = thin_border

    sub_total_price = Decimal('0')
    data_start = header_row + 1
    for idx, product in enumerate(products):
        row = data_start + idx
        ws.row_dimensions[row].height = 30
        code = product.get('code', '')
        price_info = resolve_price_info(price_mapping, code, spec=product.get('spec', ''))
        quantity = product.get('quantity', 0)

        en_name = ''
        cn_name = ''
        base_unit_price = 0
        if price_info:
            en_name = price_info.get('name_en') or price_info.get('name_ko') or price_info.get('name') or product.get('name', '')
            cn_name = price_info.get('name') or ''
            base_unit_price = float(price_info['price']) if has_valid_price_info(price_info) else 0
        else:
            _name_info = resolve_price_info(price_mapping, code)
            if _name_info:
                en_name = _name_info.get('name_en') or _name_info.get('name_ko') or _name_info.get('name') or product.get('name', '')
            else:
                en_name = product.get('name', '')

        if product.get('_is_pile'):
            display_price = 0
            is_matched = True
            total_price = Decimal('0')
        else:
            display_price, price_unit, price_info = _calc_display_price(product, price_mapping)
            is_matched = price_info is not None and has_valid_price_info(price_info)
            total_price = Decimal(str(display_price)) * Decimal(str(quantity)) if display_price and quantity else Decimal('0')
            sub_total_price += total_price

        ws.cell(row=row, column=1, value=idx + 1).alignment = center
        ws.cell(row=row, column=2, value=en_name).alignment = center
        _sub_raw_mat = (price_info.get('db_material') if price_info and price_info.get('db_material') else None) or product.get('material', '')
        ws.cell(row=row, column=3, value=_sub_raw_mat).alignment = center
        ws.cell(row=row, column=4, value='').alignment = center
        _spec_val = product.get('spec', '')
        try:
            _spec_num = float(_spec_val)
            ws.cell(row=row, column=5, value=_spec_num).alignment = center
            ws.cell(row=row, column=5).number_format = '"L"!=#"mm"'
        except (ValueError, TypeError):
            ws.cell(row=row, column=5, value=_strip_cjk_spec(_spec_val)).alignment = center
        if display_price > 0:
            ws.cell(row=row, column=6, value=display_price)
        ws.cell(row=row, column=6).number_format = CURRENCY_FMT
        ws.cell(row=row, column=6).alignment = center
        ws.cell(row=row, column=7, value=int(quantity) if quantity and quantity % 1 == 0 else (quantity or '')).alignment = center
        if total_price > 0:
            ws.cell(row=row, column=8, value=f"=F{row}*G{row}")
        ws.cell(row=row, column=8).number_format = CURRENCY_FMT
        ws.cell(row=row, column=8).alignment = center
        ws.cell(row=row, column=9, value=code).alignment = center
        ws.cell(row=row, column=10, value=code).alignment = center
        ws.cell(row=row, column=11, value=cn_name).alignment = left
        if base_unit_price > 0:
            ws.cell(row=row, column=12, value=base_unit_price).number_format = CURRENCY_FMT
        ws.cell(row=row, column=12).alignment = center

        for ci in range(1, col_count + 1):
            ws.cell(row=row, column=ci).font = normal_font
            ws.cell(row=row, column=ci).border = thin_border
        if not is_matched:
            for ci in range(1, col_count + 1):
                ws.cell(row=row, column=ci).fill = yellow_fill

    data_end = data_start + len(products) - 1
    return data_end, header_row, float(sub_total_price)


def _render_totals(ws, row, sub_total_price, total_price_all, output_kw):
    bold_font = Font(name='Calibri', size=10, bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_a = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=row, column=1, value='TOTAL AMOUNT/TABLE').font = bold_font
    ws.cell(row=row, column=1).alignment = left_a
    ws.cell(row=row, column=8, value=sub_total_price).number_format = CURRENCY_FMT
    ws.cell(row=row, column=8).font = bold_font
    ws.cell(row=row, column=8).alignment = center
    for ci in range(1, 13):
        ws.cell(row=row, column=ci).border = thin_border
    ws.row_dimensions[row].height = 30

    row += 1
    ws.cell(row=row, column=1, value='TOTAL AMOUNT OF  ALL TABLES').font = bold_font
    ws.cell(row=row, column=1).alignment = left_a
    ws.cell(row=row, column=8, value=total_price_all).number_format = CURRENCY_FMT
    ws.cell(row=row, column=8).font = bold_font
    ws.cell(row=row, column=8).alignment = center
    for ci in range(1, 13):
        ws.cell(row=row, column=ci).border = thin_border
    ws.row_dimensions[row].height = 30

    row += 1
    usd_per_w = round(total_price_all / (output_kw * 1000), 7) if output_kw and output_kw > 0 else 0
    ws.cell(row=row, column=1, value=f'{currency_label}/W').font = bold_font
    ws.cell(row=row, column=1).alignment = left_a
    ws.cell(row=row, column=8, value=usd_per_w).number_format = '#,##0.000000'
    ws.cell(row=row, column=8).font = bold_font
    ws.cell(row=row, column=8).alignment = center
    for ci in range(1, 13):
        ws.cell(row=row, column=ci).border = thin_border
    ws.row_dimensions[row].height = 30

    return row


def create_ksd_detail_sheet(
        df, workbook, sheet_name_prefix, price_mapping,
        image_path=None, image_folder=None, code_to_images=None,
        image_temp_dir=None, image_cache=None,
        unmatched_products_list=None, contact_info=None,
        config=None, matrix_data=None, sale_type='export',
        ko_discount_rate=100, ko_steel_discount_rate=84,
        ko_purchased_discount_rate=94, coating_thickness=10,
        delete_options=None, always_exclude_extra_items=False,
        ko_exclude_options=None, trade_method='EXW',
        ko_ddp_address='',
        pre_parsed_products=None,
        need_weight_code=False,
        **kwargs
):
    all_products, array_info, span_info, rows, cols = resolve_products_and_array(
        pre_parsed_products, df, matrix_data,
    )

    aluminum, carbon_steel, excluded, pile_products, price_info_cache = _classify_products_single_pass(
        all_products, price_mapping, delete_options or {},
        always_exclude_extra_items, ko_exclude_options or {}
    )

    matrix_data = matrix_data or {}
    matrix_project_name = str(matrix_data.get('project_name') or '').strip()
    matrix_output_wp = matrix_data.get('output_wp') or 0
    matrix_max_wind = matrix_data.get('max_wind_speed') or ''
    matrix_max_snow = matrix_data.get('max_snow_load') or ''
    matrix_module_watt = matrix_data.get('module_wattage') or ''
    matrix_module_size = matrix_data.get('module_size') or ''
    matrix_set_count = matrix_data.get('set_count') or 1
    if not isinstance(matrix_set_count, int) or matrix_set_count <= 0:
        matrix_set_count = 1
    total_table_count = matrix_set_count
    angle_val = (config or {}).get('angle', '') if config else ''
    if not angle_val:
        angle_val = (matrix_data or {}).get('angle', '')

    sheet_name = extract_main_name(sheet_name_prefix)
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    original = sheet_name
    counter = 1
    while sheet_name in workbook.sheetnames:
        sheet_name = f"{original}_{counter}"
        counter += 1

    ws = workbook.create_sheet(title=sheet_name)
    print(f"[SIMPLE] Creating detail sheet: {sheet_name}")

    thin = Side(style='thin', color='000000')
    thick = Side(style='medium', color='000000')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(name='Malgun Gothic', size=36, bold=True, color='000000')
    normal_font = Font(name='Malgun Gothic', size=16, color='000000')
    header_font = Font(name='Malgun Gothic', size=16, bold=True, color='000000')
    small_bold_font = Font(name='Malgun Gothic', size=16, bold=True, color='000000')
    red_small_font = Font(name='Malgun Gothic', size=16, bold=True, color='FF0000')
    if sale_type == 'domestic':
        currency_number_format = '#,##0.00'
        currency_label = 'RMB'
    else:
        currency_number_format = '#,##0.00'
        currency_label = 'US$'
    a7_font = Font(name='Malgun Gothic', size=16, bold=True, color='FFFFFF')
    a7_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)

    if need_weight_code:
        column_widths = {
            'A': 15.77, 'B': 28.67, 'C': 45, 'D': 30.03, 'E': 34.58,
            'F': 24.58, 'G': 25.49, 'H': 24.13, 'I': 22.31, 'J': 25.49,
        }
    else:
        column_widths = {
            'A': 15.77, 'B': 28.67, 'C': 38, 'D': 30.03, 'E': 34.58,
            'F': 24.58, 'G': 25.49, 'H': 24.13, 'I': 25.49,
        }
    max_col = 10 if need_weight_code else 9
    max_col_letter = get_column_letter(max_col)
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    all_render_products = aluminum + carbon_steel + excluded
    combined_products = aluminum + carbon_steel
    total_products = len(all_render_products)
    data_start_row = 9
    row_heights_temp = {1: 18, 2: 66, 3: 30, 4: 30, 5: 30, 6: 30, 7: 30, 8: 20, 9: 70.7}
    for row, height in row_heights_temp.items():
        ws.row_dimensions[row].height = height

    for col in range(1, max_col + 1):
        ws.cell(row=1, column=col)

    for _hdr_row in range(2, data_start_row):
        for _hdr_col in range(1, max_col + 1):
            ws.cell(row=_hdr_row, column=_hdr_col).border = thin_border

    ws.merge_cells(f'A2:{max_col_letter}2')
    ws['A2'] = 'Solar mounting system Quotation'
    ws['A2'].font = title_font
    ws['A2'].alignment = center_align

    ws.merge_cells('A3:B7')
    if image_path and os.path.exists(image_path):
        try:
            img = XLImage(img=image_path)
            width_a = ws.column_dimensions['A'].width
            width_b = ws.column_dimensions['B'].width
            total_width_pixels = (width_a + width_b) * 7
            height_4 = ws.row_dimensions[4].height
            height_5 = ws.row_dimensions[5].height
            height_6 = ws.row_dimensions[6].height
            total_height_pts = height_4 + height_5 + height_6
            total_height_pixels = total_height_pts * 1.33
            target_width = total_width_pixels * 1.05
            target_height = total_height_pixels * 1.8
            original_width = img.width
            original_height = img.height
            original_ratio = original_width / original_height
            target_ratio = target_width / target_height
            if original_ratio > target_ratio:
                img.width = target_width
                img.height = target_width / original_ratio
            else:
                img.height = target_height
                img.width = target_height * original_ratio
            min_size = 50
            if img.width < min_size or img.height < min_size:
                img.width = max(img.width, min_size)
                img.height = max(img.height, min_size)
            ws.add_image(img, 'A4')
        except Exception:
            ws['A4'] = '[Image not found]'
            ws['A4'].font = normal_font
            ws['A4'].alignment = center_align
    else:
        ws['A3'] = ''
        ws['A3'].alignment = center_align

    contact_defaults = {
        'contact_name': '진설정',
        'phone': '0086-18050053693',
        'tel': 'seol@xmkseng.com',
        'fax': ''
    }
    contact_info = contact_info or {}
    contact_name = contact_info.get('contact_name') or contact_defaults['contact_name']
    phone = contact_info.get('phone') or contact_defaults['phone']
    tel = contact_info.get('tel') or contact_defaults['tel']
    fax = contact_info.get('fax') or contact_defaults['fax']

    ws['C3'] = f"Sales: {contact_name}"
    ws['C3'].font = normal_font
    ws['C3'].alignment = center_align
    ws['D3'] = 'Installation Angle'
    ws['D3'].font = normal_font
    ws['D3'].alignment = center_align
    ws['E3'] = normalize_angle(angle_val)
    ws['E3'].font = normal_font
    ws['E3'].alignment = center_align
    ws['F3'] = 'Panel Size'
    ws['F3'].font = normal_font
    ws['F3'].alignment = center_align

    ws['C4'] = f"Phone: {phone}"
    ws['C4'].font = normal_font
    ws['C4'].alignment = center_align
    ws['D4'] = 'Max Wind Load'
    ws['D4'].font = normal_font
    ws['D4'].alignment = center_align
    ws['E4'] = matrix_max_wind or ''
    ws['E4'].font = normal_font
    ws['E4'].alignment = center_align
    ws['F4'] = 'Power/PC'
    ws['F4'].font = normal_font
    ws['F4'].alignment = center_align
    ws['G4'] = f"{matrix_module_watt}w" if matrix_module_watt else ''
    ws['G4'].font = normal_font
    ws['G4'].alignment = center_align

    ws['C5'] = f"Email: {tel}"
    ws['C5'].font = normal_font
    ws['C5'].alignment = center_align
    ws['D5'] = 'Max Snow Load'
    ws['D5'].font = normal_font
    ws['D5'].alignment = center_align
    ws['E5'] = matrix_max_snow or ''
    ws['E5'].font = normal_font
    ws['E5'].alignment = center_align
    ws['F5'] = 'Total Output'
    ws['F5'].font = normal_font
    ws['F5'].alignment = center_align
    ws['G5'] = f"{matrix_output_wp} Wp" if matrix_output_wp else ''
    ws['G5'].font = normal_font
    ws['G5'].alignment = center_align

    missing_boards_val = matrix_data.get('missing_per_table', 0) or 0
    if missing_boards_val == 0:
        _raw_mb = int((config or {}).get('missing_boards', 0) or 0)
        missing_boards_val = -_raw_mb if _raw_mb > 0 else _raw_mb
    ws['C6'] = 'Missing Modules'
    ws['C6'].font = normal_font
    ws['C6'].alignment = center_align
    ws['D6'] = str(missing_boards_val) if missing_boards_val != 0 else '/'
    ws['D6'].font = normal_font
    ws['D6'].alignment = center_align
    ws['F6'] = 'Span(E/W)'
    ws['F6'].font = normal_font
    ws['F6'].alignment = center_align
    ws['G6'] = f"{span_info} mm" if span_info else ''
    ws['G6'].font = normal_font
    ws['G6'].alignment = center_align

    if matrix_module_size:
        panel_spec = re.sub(r'[-×xX]', '*', str(matrix_module_size))
        if not panel_spec.lower().endswith('mm'):
            panel_spec = panel_spec + 'mm'
    else:
        panel_spec = (config or {}).get('panel_spec', '') if config else ''
    ws['G3'] = panel_spec
    ws['G3'].font = normal_font
    ws['G3'].alignment = center_align

    ws['C7'] = 'Array'
    ws['C7'].font = normal_font
    ws['C7'].alignment = center_align
    if rows is not None:
        ws['D7'] = f"{rows} Row"
    else:
        ws['D7'] = ''
    if cols is not None:
        ws['E7'] = f"{cols} Col"
    else:
        ws['E7'] = ''
    ws['D7'].alignment = center_align
    ws['E7'].alignment = center_align
    ws['D7'].font = normal_font
    ws['E7'].font = normal_font
    ws['F7'] = 'Tables'
    ws['F7'].font = normal_font
    ws['F7'].alignment = center_align
    if matrix_set_count:
        ws['G7'] = matrix_set_count
    else:
        ws['G7'] = ''
    ws['G7'].font = normal_font
    ws['G7'].alignment = center_align

    ws.merge_cells(f'H3:{max_col_letter}7')
    ws['H3'] = '10 years warranty\n20 years service life'
    ws['H3'].font = normal_font
    ws['H3'].alignment = center_align

    ws.merge_cells(f'A8:{max_col_letter}8')

    # ========== Write product data in 3 groups ==========
    row_product_map = {}
    row_price_info_map = {}
    matched_count = 0
    unmatched_count = 0
    image_found_count = 0
    image_not_found_count = 0
    local_unmatched = []

    def _col_width_to_px(char_width):
        return int(char_width * 7.5 + 5)

    def _row_height_to_px(pt_height):
        return int(pt_height * 1.33)

    def _fit_image_to_cell(ws, cell_row, cell_col, max_w, max_h, padding=IMAGE_PADDING):
        col_w = ws.column_dimensions[get_column_letter(cell_col)].width or 8.43
        row_h = ws.row_dimensions[cell_row].height or 15
        avail_w = _col_width_to_px(col_w) - padding * 2
        avail_h = _row_height_to_px(row_h) - padding * 2
        avail_w = max(avail_w, 20)
        avail_h = max(avail_h, 20)
        scale = min(avail_w / max_w, avail_h / max_h, 1.0)
        return int(max_w * scale), int(max_h * scale)

    def _write_product_rows(ws, start_row, products, price_mapping):
        nonlocal matched_count, unmatched_count, local_unmatched
        sub_total_price = Decimal('0')
        sub_total_weight = Decimal('0') if need_weight_code else None
        standard_rows = []
        steel_rows = []
        purchased_rows = []
        inquiry_rows = []
        unmatched_rows = []
        standard_price_sum = Decimal('0')
        steel_price_sum = Decimal('0')
        purchased_price_sum = Decimal('0')
        inquiry_price_sum = Decimal('0')
        is_complex = False
        for p in products:
            pi = p.get('_price_info')
            if pi is None:
                pi = resolve_price_info(price_mapping, p.get('code', ''), spec=p.get('spec', ''))
                p['_price_info'] = pi
            if not pi or not has_valid_price_info(pi):
                is_complex = True
                break
            cat = _get_discount_category(pi, p)
            if cat != 'standard':
                is_complex = True
                break
        for idx, product in enumerate(products):
            row = start_row + idx
            ws.row_dimensions[row].height = 61
            product_code = product['code']
            price_info = product.get('_price_info')
            if price_info is None:
                price_info = resolve_price_info(price_mapping, product_code, spec=product.get('spec', ''))
                product['_price_info'] = price_info
            en_name = (
                price_info.get('name_en') or price_info.get('name_ko')
                or price_info.get('name') or product['name']
            ) if price_info else product['name']
            if not price_info:
                _name_info = resolve_price_info(price_mapping, product_code)
                if _name_info:
                    en_name = (
                        _name_info.get('name_en') or _name_info.get('name_ko')
                        or _name_info.get('name') or en_name
                    )

            row_product_map[row] = product
            row_price_info_map[row] = price_info

            _set_cell(ws, row, 1, f'=ROW()-{start_row - 1}', font=normal_font, align=center_align, border=thin_border)
            _set_cell(ws, row, 2, en_name, font=normal_font, align=center_align, border=thin_border)
            _raw_mat = (price_info.get('db_material') if price_info and price_info.get('db_material') else None) or product['material']
            _set_cell(ws, row, 3, adjust_material_by_coating(translate_material(_raw_mat, 'ko'), coating_thickness),
                      font=normal_font, align=center_align, border=thin_border)
            _set_cell(ws, row, 4, "", font=normal_font, align=center_align, border=thin_border)
            try:
                _spec_num = float(product['spec'])
                _set_cell(ws, row, 5, _spec_num,
                          font=normal_font, align=center_align, border=thin_border)
                ws.cell(row=row, column=5).number_format = '"L"!=#"mm"'
            except (ValueError, TypeError):
                _set_cell(ws, row, 5, _strip_cjk_spec(product['spec']),
                          font=normal_font, align=center_align, border=thin_border)

            unit_price = 0
            display_unit_price = 0
            price_unit = ''
            is_matched = False

            if price_info and has_valid_price_info(price_info):
                unit_price = float(price_info['price'])
                price_unit = price_info.get('unit', '')
                matched_count += 1
                is_matched = True
            else:
                unmatched_count += 1
                if unmatched_products_list is not None and _is_valid_product_code(product_code):
                    local_unmatched.append({
                        'code': product_code,
                        'name': (price_info.get('name') if price_info else None) or product.get('name', ''),
                        'spec': product.get('spec', ''),
                        'quantity': product.get('quantity', 0),
                        'unit': price_info.get('unit', '') if price_info else '',
                        'issue_reason': '数据库无法匹配',
                    })

            is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
            length_mm = Decimal('0')
            if is_meter:
                length_mm = Decimal(str(extract_length_from_spec(product['spec']) or 0))
            if is_meter and length_mm > 0:
                display_unit_price = float(Decimal(str(unit_price)) * length_mm / Decimal('1000'))
            else:
                display_unit_price = unit_price

            if display_unit_price > 0:
                if is_complex and is_matched:
                    cat = _get_discount_category(price_info, product)
                    rate = _get_discount_rate(cat, ko_discount_rate, ko_steel_discount_rate, ko_purchased_discount_rate)
                    factor = rate / 100.0
                    _set_cell(ws, row, 6, f"={display_unit_price}*{factor}",
                              font=normal_font, align=center_align, border=thin_border,
                              number_format=currency_number_format)
                else:
                    _set_cell(ws, row, 6, float(display_unit_price),
                              font=normal_font, align=center_align, border=thin_border,
                              number_format=currency_number_format)
            else:
                _set_cell(ws, row, 6, "",
                          font=normal_font, align=center_align, border=thin_border,
                          number_format=currency_number_format)

            quantity = product['quantity']
            if quantity > 0:
                _set_cell(ws, row, 7, int(quantity) if quantity % 1 == 0 else quantity,
                          font=normal_font, align=center_align, border=thin_border)
            else:
                _set_cell(ws, row, 7, "",
                          font=normal_font, align=center_align, border=thin_border)

            total_price = Decimal('0')
            if display_unit_price > 0 and quantity > 0:
                total_price = Decimal(str(display_unit_price)) * Decimal(str(quantity))
                if is_complex and is_matched:
                    cat = _get_discount_category(price_info, product)
                    rate = _get_discount_rate(cat, ko_discount_rate, ko_steel_discount_rate, ko_purchased_discount_rate)
                    total_price = total_price * Decimal(str(rate)) / Decimal('100')
                _set_cell(ws, row, 8, f"=F{row}*G{row}",
                          font=normal_font, align=center_align, border=thin_border,
                          number_format=currency_number_format)
                sub_total_price += total_price
                cat = _get_discount_category(price_info)
                if cat == 'standard':
                    standard_rows.append(row)
                    standard_price_sum += total_price
                elif cat == 'steel':
                    steel_rows.append(row)
                    steel_price_sum += total_price
                elif cat == 'purchased':
                    purchased_rows.append(row)
                    purchased_price_sum += total_price
                else:
                    inquiry_rows.append(row)
                    inquiry_price_sum += total_price
            else:
                _set_cell(ws, row, 8, f"=F{row}*G{row}",
                          font=normal_font, align=center_align, border=thin_border,
                          number_format=currency_number_format)
                if not is_matched:
                    unmatched_rows.append(row)

            remark_col = 10 if need_weight_code else 9
            if need_weight_code:
                weight_cell = ws.cell(row=row, column=9)
                unit_weight = None
                if price_info:
                    db_w = parse_decimal_number(price_info.get('db_weight'))
                    if db_w is not None and db_w > 0:
                        code_attribute = str(price_info.get('code_attribute') or '').strip().upper()
                        unit_weight = db_w
                        if code_attribute in WEIGHT_BY_LENGTH_ATTRIBUTES:
                            length_mm = extract_length_from_spec(product.get('spec'))
                            if length_mm and length_mm > 0:
                                unit_weight = db_w * Decimal(str(length_mm)) / Decimal('1000')
                            else:
                                unit_weight = None
                if unit_weight is None:
                    bom_w = product.get('weight', 0)
                    if bom_w and bom_w > 0:
                        unit_weight = Decimal(str(bom_w))
                if unit_weight is not None and unit_weight > 0:
                    weight_cell.value = float(round_to_2_decimal(unit_weight))
                    weight_cell.alignment = center_align
                    weight_cell.font = normal_font
                    weight_cell.border = thin_border
                    weight_cell.number_format = '#,##0.00'
                    sub_total_weight += round_to_2_decimal(unit_weight)
                else:
                    db_w = price_info.get('db_weight') if price_info else None
                    bom_w = product.get('weight', 0)
                    print(f"   [SIMPLE-WEIGHT] row={row} code={product_code} qty={quantity} "
                          f"db_weight={db_w} bom_weight={bom_w} => empty")
                    weight_cell.value = ""
                    weight_cell.alignment = center_align
                    weight_cell.font = normal_font
                    weight_cell.border = thin_border

            _set_cell(ws, row, remark_col, product_code,
                      font=normal_font, align=center_align, border=thin_border)

            if not is_matched:
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).fill = yellow_fill

        data_end = start_row + len(products) - 1 if products else start_row - 1
        _sw = float(sub_total_weight) if need_weight_code else 0.0
        return (data_end, float(sub_total_price), _sw,
                standard_rows, steel_rows, purchased_rows, inquiry_rows,
                unmatched_rows,
                float(standard_price_sum), float(steel_price_sum),
                float(purchased_price_sum), float(inquiry_price_sum),
                is_complex)

    def _write_part_title(ws, row, title):
        ws.merge_cells(f'A{row}:{max_col_letter}{row}')
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = small_bold_font
        ws[f'A{row}'].alignment = left_align
        ws.row_dimensions[row].height = 40
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = thin_border

    def _write_group_header(ws, row):
        group_headers = [
            ('Item No.', 'A'), ('Product Name', 'B'), ('Material', 'C'),
            ('Picture', 'D'), ('Spec.', 'E'),
            (f'Unit Price ({currency_label})\nEX Works', 'F'),
            (f'QTY (PCS)', 'G'),
            (f'Total Price ({currency_label})\nEX Works', 'H'),
            ('Remark', 'J' if need_weight_code else 'I')
        ]
        if need_weight_code:
            group_headers.insert(-1, ('Weight(KG)', 'I'))
        green_fill = PatternFill(start_color='98FFEE', end_color='98FFEE', fill_type='solid')
        for text, col in group_headers:
            cell = ws[f'{col}{row}']
            cell.value = text
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            if need_weight_code and col == 'I':
                cell.fill = green_fill
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = thin_border

    def _write_group_totals(ws, row, sub_total, total_all, kw, sub_weight=0, total_weight=0, data_first_row=None, data_last_row=None):
        ws.row_dimensions[row].height = 40
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = 'TOTAL AMOUNT/TABLE'
        ws[f'A{row}'].font = small_bold_font
        ws[f'A{row}'].alignment = right_align
        if data_first_row and data_last_row and data_last_row >= data_first_row:
            ws.cell(row=row, column=8, value=f'=SUM(H{data_first_row}:H{data_last_row})')
        else:
            ws.cell(row=row, column=8, value=sub_total)
        ws.cell(row=row, column=8).number_format = currency_number_format
        ws.cell(row=row, column=8).font = small_bold_font
        ws.cell(row=row, column=8).alignment = center_align
        if need_weight_code:
            if data_first_row and data_last_row and data_last_row >= data_first_row:
                ws.cell(row=row, column=9, value=f'=SUMPRODUCT(I{data_first_row}:I{data_last_row},G{data_first_row}:G{data_last_row})')
            else:
                ws.cell(row=row, column=9, value=round(sub_weight, 2))
            ws.cell(row=row, column=9).number_format = '#,##0.00'
            ws.cell(row=row, column=9).font = small_bold_font
            ws.cell(row=row, column=9).alignment = center_align
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = thin_border

        sub_total_row = row
        row += 1
        ws.row_dimensions[row].height = 40
        ws.merge_cells(f'A{row}:G{row}')
        total_label = f'TOTAL AMOUNT OF {total_table_count} TABLES' if total_table_count > 1 else 'TOTAL AMOUNT OF ALL TABLES'
        ws[f'A{row}'] = total_label
        ws[f'A{row}'].font = small_bold_font
        ws[f'A{row}'].alignment = right_align
        if total_table_count > 1:
            ws.cell(row=row, column=8, value=f'=H{sub_total_row}*G7')
        else:
            ws.cell(row=row, column=8, value=f'=H{sub_total_row}')
        ws.cell(row=row, column=8).number_format = currency_number_format
        ws.cell(row=row, column=8).font = small_bold_font
        ws.cell(row=row, column=8).alignment = center_align
        if need_weight_code:
            if total_table_count > 1:
                ws.cell(row=row, column=9, value=f'=I{sub_total_row}*G7')
            else:
                ws.cell(row=row, column=9, value=f'=I{sub_total_row}')
            ws.cell(row=row, column=9).number_format = '#,##0.00'
            ws.cell(row=row, column=9).font = small_bold_font
            ws.cell(row=row, column=9).alignment = center_align
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = thin_border

        all_tables_row = row
        row += 1
        ws.row_dimensions[row].height = 40
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = f'{currency_label}/W'
        ws[f'A{row}'].font = small_bold_font
        ws[f'A{row}'].alignment = right_align
        if kw and kw > 0:
            ws.cell(row=row, column=8, value=f'=H{all_tables_row}/({kw}*1000)')
        else:
            ws.cell(row=row, column=8, value=0)
        ws.cell(row=row, column=8).number_format = '#,##0.000'
        ws.cell(row=row, column=8).font = small_bold_font
        ws.cell(row=row, column=8).alignment = center_align
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = thin_border

        return row

    multiplier = Decimal(str(total_table_count)) if total_table_count > 1 else Decimal('1')
    output_kw = matrix_data.get('output_kw') or 0
    current_row = data_start_row

    # ========== Part 1: All Products (no carbon steel split) ==========
    part1_sub_total = 0.0
    part1_total_all = 0.0
    part1_sub_weight = 0.0
    part1_total_weight = 0.0
    part1_total_table_row = 0
    part1_std_rows = []
    part1_steel_rows = []
    part1_purchased_rows = []
    part1_inq_rows = []
    part1_unmatched_rows = []
    part1_std_price = 0.0
    part1_steel_price = 0.0
    part1_purchased_price = 0.0
    part1_inq_price = 0.0
    is_complex = False
    if combined_products:
        _write_group_header(ws, current_row)
        current_row += 1
        part1_data_start = current_row
        (data_end, part1_sub_total, part1_sub_weight,
         part1_std_rows, part1_steel_rows, part1_purchased_rows, part1_inq_rows,
         part1_unmatched_rows,
         part1_std_price, part1_steel_price, part1_purchased_price, part1_inq_price,
         is_complex) = _write_product_rows(ws, current_row, combined_products, price_mapping)
        part1_data_end = data_end
        current_row = data_end + 1
        part1_total_all = part1_sub_total * total_table_count
        part1_total_weight = part1_sub_weight
        part1_total_table_row = current_row
        current_row = _write_group_totals(ws, current_row, part1_sub_total, part1_total_all, output_kw, part1_sub_weight, part1_total_weight, data_first_row=part1_data_start, data_last_row=part1_data_end) + 1

    # ========== Part 3: Alternative / Excluded ==========
    part3_sub_total = 0.0
    part3_total_all = 0.0
    part3_sub_weight = 0.0
    part3_total_weight = 0.0
    if excluded:
        ws.merge_cells(f'A{current_row}:{max_col_letter}{current_row}')
        ws[f'A{current_row}'] = 'The following products are alternative, please let me know if you need them.'
        ws[f'A{current_row}'].font = red_small_font
        ws[f'A{current_row}'].alignment = left_align
        for c in range(1, max_col + 1):
            ws.cell(row=current_row, column=c).border = thin_border
        current_row += 1
        _write_part_title(ws, current_row, 'Part 3: Optional Accessories (Earth Clip/Earth Lug etc.)')
        current_row += 1
        _write_group_header(ws, current_row)
        current_row += 1
        part3_data_start = current_row
        data_end, part3_sub_total, part3_sub_weight, _, _, _, _, _, _, _, _, _, _ = _write_product_rows(ws, current_row, excluded, price_mapping)
        part3_data_end = data_end
        current_row = data_end + 1
        part3_total_all = part3_sub_total * total_table_count
        part3_total_weight = part3_sub_weight
        current_row = _write_group_totals(ws, current_row, part3_sub_total, part3_total_all, output_kw, part3_sub_weight, part3_total_weight, data_first_row=part3_data_start, data_last_row=part3_data_end) + 1

    last_data_row = current_row - 1

    # ========== Insert images ==========
    def mark_missing_image(row_number):
        image_cell = ws.cell(row=row_number, column=IMAGE_COLUMN_INDEX)
        image_cell.value = "/"
        image_cell.alignment = center_align
        image_cell.font = normal_font

    for row in range(data_start_row, last_data_row + 1):
        if row not in row_product_map:
            continue
        product_code = ws.cell(row=row, column=max_col).value
        if product_code is None or str(product_code).strip() == '':
            continue
        product_code = str(product_code).strip()
        normalized_code = normalize_lookup_code(product_code)

        img_path = None
        if code_to_images and product_code in code_to_images and code_to_images[product_code]:
            img_path = code_to_images[product_code][0]
        elif code_to_images and normalized_code in code_to_images and code_to_images[normalized_code]:
            img_path = code_to_images[normalized_code][0]

        if img_path and image_folder:
            fit_w, fit_h = _fit_image_to_cell(ws, row, IMAGE_COLUMN_INDEX, IMAGE_WIDTH, IMAGE_HEIGHT)
            temp_img_path = prepare_image_for_excel(
                img_path, target_width=fit_w, target_height=fit_h,
                temp_dir=image_temp_dir, cache=image_cache
            )
            final_img_path = temp_img_path if temp_img_path else img_path
            success = add_image_centered_in_cell(
                ws, final_img_path, row, IMAGE_COLUMN_INDEX,
                img_width=fit_w, img_height=fit_h
            )
            if success:
                image_found_count += 1
                continue
            image_not_found_count += 1
            mark_missing_image(row)
        else:
            image_not_found_count += 1
            mark_missing_image(row)

    print(f"   [SIMPLE] Images: found={image_found_count}, not_found={image_not_found_count}")

    grand_total = part1_total_all + part3_total_all

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(top=0.75, bottom=0.75, left=0.7, right=0.7, header=0.3, footer=0.3)

    if unmatched_products_list is not None:
        for _up in local_unmatched:
            _up['quantity'] = float(_up.get('quantity', 0)) * total_table_count
        unmatched_products_list.extend(local_unmatched)

    quotation_product_codes = set()
    for p in combined_products + excluded:
        c = str(p.get('code', '')).strip()
        if c:
            quotation_product_codes.add(c)

    return {
        'sheet_name': sheet_name,
        'quotation_product_codes': quotation_product_codes,
        'valid_products': len(combined_products) + len(excluded),
        'total_weight': 0,
        'total_price': grand_total,
        'part1_price_per_table': part1_sub_total,
        'part1_price_all_tables': part1_total_all,
        'part1_total_table_row': part1_total_table_row,
        'part1_standard_rows': part1_std_rows,
        'part1_steel_rows': part1_steel_rows,
        'part1_purchased_rows': part1_purchased_rows,
        'part1_inquiry_rows': part1_inq_rows,
        'part1_unmatched_rows': part1_unmatched_rows,
        'part1_standard_total': part1_std_price,
        'part1_steel_total': part1_steel_price,
        'part1_purchased_total': part1_purchased_price,
        'part1_inquiry_total': part1_inq_price,
        'part2_price_per_table': 0,
        'part2_price_all_tables': 0,
        'part2_total_table_row': 0,
        'part2_standard_rows': [],
        'part2_steel_rows': [],
        'part2_purchased_rows': [],
        'part2_inquiry_rows': [],
        'part2_unmatched_rows': [],
        'is_complex': is_complex,
        'part2_standard_total': 0,
        'part2_steel_total': 0,
        'part2_purchased_total': 0,
        'part2_inquiry_total': 0,
        'matched_count': matched_count,
        'unmatched_count': unmatched_count,
        'unmatched_products': local_unmatched,
        'image_found_count': image_found_count,
        'image_not_found_count': image_not_found_count,
        'pile_products': pile_products,
        'sub_total_row': last_data_row,
        'total_row': last_data_row,
        'detail_data_end_row': last_data_row,
        'config': config,
        'matrix_data': matrix_data,
    }


def create_ksd_summary_sheet(
        workbook, all_quotation_results, matrix_data=None,
        ko_discount_rate=100, ko_steel_discount_rate=84,
        ko_purchased_discount_rate=94, sale_type='export',
        contact_info=None, trade_method='EXW',
        image_path=None, ko_ddp_address='',
        pile_summary=None,
        dest_port='BUSAN', container_type='40HQ',
        container_qty=1, ko_cif_freight=0,
        skip_freight=False,
):
    if sale_type == 'domestic':
        ksd_currency_fmt = '#,##0.00'
        ksd_currency_label = 'RMB'
    else:
        ksd_currency_fmt = '#,##0.00'
        ksd_currency_label = 'US$'
    matrix_data = matrix_data or {}
    project_name = str(matrix_data.get('project_name') or '').strip()
    output_kw = matrix_data.get('output_kw') or 0

    ws = workbook.create_sheet(title='Total')

    normal_font = Font(name='Calibri', size=10)
    bold_font = Font(name='Calibri', size=10, bold=True)
    title_font = Font(name='Calibri', size=24, bold=True)
    company_font = Font(name='Calibri', size=12, bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_bottom = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
    right_a = Alignment(horizontal='right', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    thick = Side(style='medium', color='000000')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bottom_border = Border(bottom=thin)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    col_width_map = {'A': 6, 'B': 16, 'C': 16, 'D': 8, 'F': 11}
    for col_letter in 'ABCDEFGHIJKLM':
        ws.column_dimensions[col_letter].width = col_width_map.get(col_letter, 8)

    ws.merge_cells('J1:L1')
    ws['J1'] = f'DATE：\t\t{datetime.now().strftime("%Y/%m/%d")}'
    ws['J1'].font = normal_font
    ws['J1'].alignment = center


    ws.merge_cells('A2:M2')
    ws['A2'] = 'Solar mounting system Quotation'
    ws['A2'].font = title_font
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 30

    if image_path and os.path.exists(image_path):
        try:
            img = XLImage(img=image_path)
            img.width = 280
            img.height = 70
            ws.add_image(img, 'H3')
        except Exception:
            pass

    ws.merge_cells('A5:B5')
    ws['A5'] = 'Project:'
    ws['A5'].font = normal_font
    ws['A5'].alignment = Alignment(horizontal='left', vertical='center')
    project_name_font = Font(name='Calibri', size=9)
    ws.merge_cells('C5:F5')
    ws['C5'] = project_name
    ws['C5'].font = project_name_font
    ws['C5'].alignment = left_bottom
    ws.row_dimensions[5].height =25

    ws.merge_cells('A6:B6')
    ws['A6'] = 'Price Term：'
    ws['A6'].font = normal_font
    ws['A6'].alignment = left_bottom
    ws.merge_cells('C6:F6')
    ws['C6'] = 'EXW'
    ws['C6'].font = Font(name='Calibri', size=10, color='FF0000')
    ws['C6'].alignment = left_bottom

    ws.merge_cells('A7:B7')
    ws['A7'] = 'Delivery time：'
    ws['A7'].font = normal_font
    ws['A7'].alignment = left_bottom
    ws.merge_cells('C7:F7')
    ws['C7'] = '2 Weeks after receiving deposit'
    ws['C7'].font = normal_font
    ws['C7'].alignment = left_bottom
    ws.merge_cells('H7:M7')
    ws['H7'] = 'Xiamen Kseng Metal Tech Co.,Ltd'
    ws['H7'].font = Font(name='Calibri', size=15, bold=True)
    ws['H7'].alignment = left_bottom

    ws.merge_cells('A8:B8')
    ws['A8'] = 'Payment Term:'
    ws['A8'].font = normal_font
    ws['A8'].alignment = left_bottom
    ws.merge_cells('C8:F8')
    ws['C8'] = '30% T/T deposit, 70% balance against B/L Copy'
    ws['C8'].font = normal_font
    ws['C8'].alignment = left_bottom
    ws.merge_cells('H8:M8')
    ws['H8'] = 'Add.: RM 302, Huixin Wealth Centre, No. 891, '
    ws['H8'].font = normal_font
    ws['H8'].alignment = left_bottom

    ws.merge_cells('A9:B9')
    ws['A9'] = 'Validity Date:'
    ws['A9'].font = normal_font
    ws['A9'].alignment = left_bottom
    ws.merge_cells('C9:F9')
    ws['C9'] = 'Quote valid on issue date only. Please reconfirm price when ordering.'
    ws['C9'].font = normal_font
    ws['C9'].alignment = left_bottom
    ws.merge_cells('H9:M9')
    ws['H9'] = 'Fanghu North 2nd Rd, Huli Dist, Xiamen, Fujian, China'
    ws['H9'].font = normal_font
    ws['H9'].alignment = left_bottom

    part1_total_all = Decimal('0')
    part1_per_table = Decimal('0')
    part1_std_total = Decimal('0')
    part1_steel_total = Decimal('0')
    part1_purchased_total = Decimal('0')
    part1_inq_total = Decimal('0')
    total_output_kw = Decimal('0')
    for qr in all_quotation_results:
        part1_total_all += Decimal(str(qr.get('part1_price_all_tables', 0)))
        part1_per_table += Decimal(str(qr.get('part1_price_per_table', 0)))
        md = qr.get('matrix_data') or matrix_data
        _r_set = md.get('set_count') or 1
        try:
            _r_set = int(_r_set)
        except (ValueError, TypeError):
            _r_set = 1
        _set_mult = Decimal(str(_r_set))
        part1_std_total += Decimal(str(qr.get('part1_standard_total', 0))) * _set_mult
        part1_steel_total += Decimal(str(qr.get('part1_steel_total', 0))) * _set_mult
        part1_purchased_total += Decimal(str(qr.get('part1_purchased_total', 0))) * _set_mult
        part1_inq_total += Decimal(str(qr.get('part1_inquiry_total', 0))) * _set_mult
        kw = md.get('output_kw') or 0
        try:
            total_output_kw += Decimal(str(kw))
        except (ArithmeticError, ValueError):
            pass

    any_complex = any(qr.get('is_complex', False) for qr in all_quotation_results)

    if any_complex:
        grand_total = float(part1_std_total) + float(part1_steel_total) + float(part1_purchased_total) + float(part1_inq_total)
    else:
        grand_total = (float(part1_std_total) * ko_discount_rate / 100.0
                       + float(part1_steel_total) * ko_steel_discount_rate / 100.0
                       + float(part1_purchased_total) * ko_purchased_discount_rate / 100.0
                       + float(part1_inq_total))
    unit_per_w = round(grand_total / (float(total_output_kw) * 1000), 7) if total_output_kw > 0 else 0

    _n = len(all_quotation_results)
    _total_row = 15 + _n
    _disc_p1_row = _total_row + 1
    _perw_row = _disc_p1_row + 1

    for r in range(5, 12):
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = bottom_border

    part1_start = 13
    part1_title_font = Font(name='Calibri', size=12, bold=True)
    ws.merge_cells(f'A{part1_start}:L{part1_start}')
    ws.cell(row=part1_start, column=1, value=' Part 1:Solar Mounting System').font = part1_title_font
    ws.row_dimensions[part1_start].height = 25

    hdr_row = part1_start + 1
    part1_headers = ['NO.', 'Row', 'Column', 'Installation Angle', '', 'Table', 'Power（kw）', '', f'Price({ksd_currency_label}/Table)', '', f'Amount({ksd_currency_label})', '']
    for ci, h in enumerate(part1_headers):
        cell = ws.cell(row=hdr_row, column=ci + 1, value=h)
        cell.font = bold_font
        cell.alignment = center
        cell.border = thin_border
    ws.merge_cells(f'D{hdr_row}:E{hdr_row}')
    ws.merge_cells(f'G{hdr_row}:H{hdr_row}')
    ws.merge_cells(f'I{hdr_row}:J{hdr_row}')
    ws.merge_cells(f'K{hdr_row}:L{hdr_row}')
    ws.row_dimensions[hdr_row].height = 30

    data_row = hdr_row + 1

    _arrays_for_sort = matrix_data.get('arrays') or []
    if _arrays_for_sort:
        _arr_pos = {}
        for _ai, _ae in enumerate(_arrays_for_sort):
            _k = (_ae.get('rows'), _ae.get('cols'), _ae.get('table_qty', 1), _ae.get('missing_per_table', 0) or 0)
            if _k not in _arr_pos:
                _arr_pos[_k] = _ai

        def _sort_by_array(item):
            _idx, _d = item
            _explicit = _d.get('_matrix_idx')
            if _explicit is not None:
                return (_explicit, _d.get('accumulated_sub_idx', 0), _idx)
            _ma = _d.get('matched_array') or _d.get('matrix_data') or {}
            _k = (_ma.get('rows'), _ma.get('cols'), _ma.get('table_qty', 1) or 1, _ma.get('missing_per_table', 0) or 0)
            return (_arr_pos.get(_k, len(_arrays_for_sort)), 0, _idx)

        _indexed = list(enumerate(all_quotation_results))
        _indexed.sort(key=_sort_by_array)
        all_quotation_results[:] = [_r for _, _r in _indexed]

    for i, qr in enumerate(all_quotation_results):
        r = data_row + i
        cfg = qr.get('config') or {}
        md = qr.get('matrix_data') or matrix_data
        ma = qr.get('matched_array') or {}
        r_rows = ma.get('rows', md.get('array_rows'))
        r_cols = ma.get('cols', md.get('array_cols'))
        r_set = md.get('set_count') or 1
        try:
            r_set = int(r_set)
        except (ValueError, TypeError):
            pass
        angle_val = cfg.get('angle', '') or md.get('angle', '')
        angle_display = normalize_angle(angle_val)
        r_kw = md.get('output_kw') or 0
        detail_sheet = qr.get('sheet_name', '')
        p1_total_table_row = qr.get('part1_total_table_row', 0)

        _m_idx = qr.get('_matrix_idx')
        _acc_sub = qr.get('accumulated_sub_idx')
        if _m_idx is not None:
            if _acc_sub is not None:
                _no_label = f"({_m_idx + 1}_{_acc_sub + 1})"
            else:
                _no_label = f"({_m_idx + 1})"
        else:
            _no_label = str(i + 1)
        ws.cell(row=r, column=1, value=_no_label).alignment = center
        ws.cell(row=r, column=2, value=r_rows or '').alignment = center
        ws.cell(row=r, column=3, value=r_cols or '').alignment = center
        ws.cell(row=r, column=4, value=angle_display).alignment = center
        ws.merge_cells(f'D{r}:E{r}')
        ws.cell(row=r, column=6, value=r_set or '').alignment = center
        ws.cell(row=r, column=7, value=r_kw).alignment = center
        ws.merge_cells(f'G{r}:H{r}')
        if detail_sheet and p1_total_table_row:
            ws.cell(row=r, column=9, value=f"='{detail_sheet}'!H{p1_total_table_row}").number_format = ksd_currency_fmt
        else:
            ws.cell(row=r, column=9, value=qr.get('part1_price_per_table', 0)).number_format = ksd_currency_fmt
        ws.cell(row=r, column=9).alignment = center
        ws.merge_cells(f'I{r}:J{r}')
        if isinstance(r_set, int) and r_set > 1:
            ws.cell(row=r, column=11, value=f'=I{r}*F{r}').number_format = ksd_currency_fmt
        else:
            ws.cell(row=r, column=11, value=f'=I{r}').number_format = ksd_currency_fmt
        ws.cell(row=r, column=11).alignment = center
        ws.merge_cells(f'K{r}:L{r}')
        for ci in range(1, 13):
            ws.cell(row=r, column=ci).font = normal_font
            ws.cell(row=r, column=ci).border = thin_border
        if not qr.get('part1_price_per_table', 0):
            for ci in range(1, 13):
                ws.cell(row=r, column=ci).fill = yellow_fill
        ws.row_dimensions[r].height = 25

    data_end = data_row + len(all_quotation_results) - 1

    total_row = data_end + 1
    ws.merge_cells(f'A{total_row}:F{total_row}')
    ws.cell(row=total_row, column=1, value='Total').font = bold_font
    ws.cell(row=total_row, column=1).alignment = center
    ws.merge_cells(f'G{total_row}:H{total_row}')
    ws.cell(row=total_row, column=7, value=f'=SUM(G{data_row}:G{data_end})').alignment = center
    ws.merge_cells(f'I{total_row}:J{total_row}')
    ws.merge_cells(f'K{total_row}:L{total_row}')
    ws.cell(row=total_row, column=11, value=f'=SUM(K{data_row}:K{data_end})').number_format = ksd_currency_fmt
    ws.cell(row=total_row, column=11).alignment = center
    for ci in range(1, 13):
        ws.cell(row=total_row, column=ci).border = thin_border
    ws.row_dimensions[total_row].height = 25

    red_bold_font = Font(name='Calibri', size=12, bold=True, color='FF0000')

    disc_p1_row = total_row + 1
    ws.merge_cells(f'A{disc_p1_row}:J{disc_p1_row}')
    ws.cell(row=disc_p1_row, column=1, value='DISCOUNT').font = red_bold_font
    ws.cell(row=disc_p1_row, column=1).alignment = right_a
    ws.merge_cells(f'K{disc_p1_row}:L{disc_p1_row}')
    if any_complex:
        ws.cell(row=disc_p1_row, column=11, value=f'=K{total_row}').number_format = ksd_currency_fmt
    else:
        ws.cell(row=disc_p1_row, column=11, value=f'=K{total_row}*{ko_discount_rate}/100').number_format = ksd_currency_fmt
    ws.cell(row=disc_p1_row, column=11).font = red_bold_font
    ws.cell(row=disc_p1_row, column=11).alignment = center
    for ci in range(1, 13):
        ws.cell(row=disc_p1_row, column=ci).border = thin_border
    ws.row_dimensions[disc_p1_row].height = 25

    perw_row = disc_p1_row + 1
    ws.merge_cells(f'A{perw_row}:J{perw_row}')
    ws.cell(row=perw_row, column=1, value='per/w').font = bold_font
    ws.cell(row=perw_row, column=1).alignment = right_a
    ws.merge_cells(f'K{perw_row}:L{perw_row}')
    ws.cell(row=perw_row, column=11, value=f'=ROUND(IF(G{total_row}>0,K{disc_p1_row}/(G{total_row}*1000),0),3)').number_format = '#,##0.000'
    ws.cell(row=perw_row, column=11).alignment = center
    for ci in range(1, 13):
        ws.cell(row=perw_row, column=ci).border = thin_border
    ws.row_dimensions[perw_row].height = 25

    _port_en_map = {'부산': 'BUSAN', '인천': 'INCHEON'}
    dest_port_en = _port_en_map.get(dest_port, dest_port)

    if skip_freight:
        freight_cif_row = disc_p1_row
        freight_end_row = perw_row
    else:
        freight_title_row = perw_row + 1
        freight_title_font = Font(name='Calibri', size=12, bold=True)
        ws.merge_cells(f'A{freight_title_row}:L{freight_title_row}')
        ws.cell(row=freight_title_row, column=1, value='Part 2:Freight & Insurance').font = freight_title_font
        ws.row_dimensions[freight_title_row].height = 25

        freight_hdr_row = freight_title_row + 1
        freight_headers = ['NO.', 'PORT', '', 'Container', '', '', 'Quantity', '', 'Unit price', '', 'Amount(USD)', '']
        for ci, h in enumerate(freight_headers):
            cell = ws.cell(row=freight_hdr_row, column=ci + 1, value=h)
            cell.font = bold_font
            cell.alignment = center
            cell.border = thin_border
        ws.merge_cells(f'B{freight_hdr_row}:C{freight_hdr_row}')
        ws.merge_cells(f'D{freight_hdr_row}:F{freight_hdr_row}')
        ws.merge_cells(f'G{freight_hdr_row}:H{freight_hdr_row}')
        ws.merge_cells(f'I{freight_hdr_row}:J{freight_hdr_row}')
        ws.merge_cells(f'K{freight_hdr_row}:L{freight_hdr_row}')
        ws.row_dimensions[freight_hdr_row].height = 25

        freight_data_row = freight_hdr_row + 1
        ws.cell(row=freight_data_row, column=1, value=1).alignment = center
        ws.cell(row=freight_data_row, column=2, value=dest_port).alignment = center
        ws.merge_cells(f'B{freight_data_row}:C{freight_data_row}')
        _ct_parts = str(container_type).split('*')
        _ct_type = _ct_parts[0] if _ct_parts else str(container_type)
        _ct_qty = int(_ct_parts[1]) if len(_ct_parts) > 1 else 1
        ws.cell(row=freight_data_row, column=4, value=f'{_ct_qty}X{_ct_type}').alignment = center
        ws.merge_cells(f'D{freight_data_row}:F{freight_data_row}')
        ws.cell(row=freight_data_row, column=7, value=container_qty).alignment = center
        ws.merge_cells(f'G{freight_data_row}:H{freight_data_row}')
        ws.cell(row=freight_data_row, column=9, value=ko_cif_freight).alignment = center
        ws.cell(row=freight_data_row, column=9).number_format = '#,##0.00'
        ws.merge_cells(f'I{freight_data_row}:J{freight_data_row}')
        ws.cell(row=freight_data_row, column=11, value=f'=I{freight_data_row}*G{freight_data_row}').alignment = center
        ws.cell(row=freight_data_row, column=11).number_format = ksd_currency_fmt
        ws.merge_cells(f'K{freight_data_row}:L{freight_data_row}')
        for ci in range(1, 13):
            ws.cell(row=freight_data_row, column=ci).font = normal_font
            ws.cell(row=freight_data_row, column=ci).border = thin_border
        ws.row_dimensions[freight_data_row].height = 25

        freight_cif_row = freight_data_row + 1
        ws.merge_cells(f'A{freight_cif_row}:J{freight_cif_row}')
        ws.cell(row=freight_cif_row, column=1, value=f'{trade_method} {dest_port_en}').font = red_bold_font
        ws.cell(row=freight_cif_row, column=1).alignment = right_a
        ws.merge_cells(f'K{freight_cif_row}:L{freight_cif_row}')
        ws.cell(row=freight_cif_row, column=11, value=f'=K{disc_p1_row}+K{freight_data_row}').number_format = ksd_currency_fmt
        ws.cell(row=freight_cif_row, column=11).font = red_bold_font
        ws.cell(row=freight_cif_row, column=11).alignment = center
        for ci in range(1, 13):
            ws.cell(row=freight_cif_row, column=ci).border = thin_border
        ws.row_dimensions[freight_cif_row].height = 25

        freight_end_row = freight_cif_row

    phone = (contact_info or {}).get('phone', '0086-18050053693')
    email = (contact_info or {}).get('tel', 'using@xmkseng.com')
    ws.merge_cells('A10:B10')
    ws['A10'] = 'Unite Price/W:'
    ws['A10'].font = bold_font
    ws['A10'].alignment = left_bottom
    ws.merge_cells('C10:F10')
    ws.cell(row=10, column=3, value=f'=ROUND(IF(G{total_row}>0,K{freight_cif_row}/(G{total_row}*1000),0),3)').number_format = '#,##0.000'
    ws.cell(row=10, column=3).font = bold_font
    ws.cell(row=10, column=3).alignment = left_bottom
    ws.merge_cells('H10:M10')
    ws['H10'] = f'Tel: {phone}'
    ws['H10'].font = normal_font
    ws['H10'].alignment = left_bottom
    ws.merge_cells('A11:B11')
    ws.merge_cells('H11:M11')
    ws['H11'] = f'Email: {email}'
    ws['H11'].font = normal_font
    ws['H11'].alignment = left_bottom

    # ========== 绘制外圈粗黑边框（修复版） ==========
    outer_start_row = part1_start
    if skip_freight:
        outer_end_row = perw_row
    else:
        outer_end_row = freight_end_row
    outer_left_col = 1
    outer_right_col = 12
    
    thin_side = Side(style='thin', color='000000')
    thick_side = Side(style='medium', color='000000')
    
    # 第一步：给整个区域的所有单元格设置细边框（作为基础）
    for r in range(outer_start_row, outer_end_row + 1):
        for c in range(outer_left_col, outer_right_col + 1):
            cell = ws.cell(row=r, column=c)
            # 跳过合并单元格的非左上角单元格
            is_skip = False
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range and cell.coordinate != merged_range.start_cell.coordinate:
                    is_skip = True
                    break
            if not is_skip:
                # 先设置细边框
                cell.border = Border(
                    left=thin_side, right=thin_side,
                    top=thin_side, bottom=thin_side
                )
    
    # 第二步：加粗外边框 - 左边框
    for r in range(outer_start_row, outer_end_row + 1):
        target_cell = ws.cell(row=r, column=outer_left_col)
        # 找到合并单元格的左上角
        for merged_range in ws.merged_cells.ranges:
            if target_cell.coordinate in merged_range:
                target_cell = merged_range.start_cell
                break
        old_border = target_cell.border
        if old_border is None:
            old_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        target_cell.border = Border(
            left=thick_side, 
            right=old_border.right,
            top=old_border.top, 
            bottom=old_border.bottom
        )
    
    # 第三步：加粗外边框 - 右边框
    for r in range(outer_start_row, outer_end_row + 1):
        target_cell = ws.cell(row=r, column=outer_right_col)
        for merged_range in ws.merged_cells.ranges:
            if target_cell.coordinate in merged_range:
                target_cell = merged_range.start_cell
                break
        old_border = target_cell.border
        if old_border is None:
            old_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        target_cell.border = Border(
            left=old_border.left, 
            right=thick_side,
            top=old_border.top, 
            bottom=old_border.bottom
        )
    
    # 第四步：加粗外边框 - 上边框
    for c in range(outer_left_col, outer_right_col + 1):
        target_cell = ws.cell(row=outer_start_row, column=c)
        for merged_range in ws.merged_cells.ranges:
            if target_cell.coordinate in merged_range:
                target_cell = merged_range.start_cell
                break
        old_border = target_cell.border
        if old_border is None:
            old_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        target_cell.border = Border(
            left=old_border.left, 
            right=old_border.right,
            top=thick_side, 
            bottom=old_border.bottom
        )
    
    # 第五步：加粗外边框 - 下边框
    for c in range(outer_left_col, outer_right_col + 1):
        target_cell = ws.cell(row=outer_end_row, column=c)
        for merged_range in ws.merged_cells.ranges:
            if target_cell.coordinate in merged_range:
                target_cell = merged_range.start_cell
                break
        old_border = target_cell.border
        if old_border is None:
            old_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        target_cell.border = Border(
            left=old_border.left, 
            right=old_border.right,
            top=old_border.top, 
            bottom=thick_side
        )
    # ========== 边框绘制结束 ==========

    hint_row = outer_end_row + 2
    ws.merge_cells(f'A{outer_end_row + 1}:L{outer_end_row + 1}')
    ws.merge_cells(f'A{hint_row}:L{hint_row}')
    ws.cell(row=hint_row, column=1,
            value='Please kindly check the detailed prices of components in the following sheet').font = Font(
        name='Calibri', size=12, color='FF0000')
    ws.cell(row=hint_row, column=1).alignment = left_bottom

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(top=0.75, bottom=0.75, left=0.7, right=0.7, header=0.3, footer=0.3)

    sheet_names = workbook.sheetnames
    if 'Total' in sheet_names:
        idx = sheet_names.index('Total')
        workbook.move_sheet('Total', offset=-idx)

    print(f"   ✅ SIMPLE summary sheet(Total) created")
    return ws.title