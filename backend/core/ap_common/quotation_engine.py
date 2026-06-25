"""亚太组（Asia-Pacific）屋顶分销报价单引擎。

对应模板：【KSENG Quotation】亚太屋顶分销模板.xls
币种：USD（美元）；整份报价单输出为「单页」工作表
（多阵列/多站点合并：Part 1 列出全部站点，Part 2 汇总全部物料）。

版式贴合模板：
    抬头：Quotation / Project / PI No / Date / Validity Date + 公司地址联系方式
    Part 1 - Information for Each Site：站点清单
        No.(B) | Roof Information(C-D) | Row/Table(E) | Column/Table(F) | Tables(G) |
        Total Capacity(kw)(H-J)
    Part 2 - Bill of Material：
        No. | Name | Picture | Material | Specification | EXW Unit Price (USD) | QTY(PCS) |
        Free spare parts(PCS) | Total QTY (PCS) | Amount(USD) | Remark
    合计：{trade_method} Xiamen Total Amount（无底色、无框）
    TERMS & CONDITIONS / BANK INFORMATION

计价说明：单价直接取数据库价格展示（USD，无汇率换算）；
          长度物料(米)按规格长度折算；备品 = QTY × 0.1%；贸易方式 EXW/FOB/CIF 仅影响条款标签（不含运费）。
"""

import os
import sys
from decimal import Decimal

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins
from backend.core.print_settings import apply_print_setup

from backend.core.shared.bom_utils import resolve_products_and_array
from backend.core.shared.price_utils import (
    resolve_price_info,
    has_valid_price_info,
    get_temp_adjusted_base_price,
    get_temp_base_price, apply_temp_preinstall_adjustment,
    round_to_2_decimal,
    _get_discount_category,
)
from backend.core.shared.weight_utils import extract_length_from_spec
from backend.core.shared.text_utils import (
    extract_main_name,
    normalize_lookup_code,
    _strip_cjk_spec,
)
from backend.core.shared.product_utils import (
    _is_valid_product_code,
    _match_exclude_group,
    normalize_preinstall,
)
from backend.core.shared.constants import (
    EXCLUDE_ITEM_GROUPS,
    CARBON_STEEL_PRICING_ATTRS,
    IMAGE_PADDING,
)
from backend.core.shared.image_utils import (
    prepare_image_for_excel,
    add_image_centered_in_cell,
)
from backend.core.material_translate import translate_material, adjust_material_by_coating

if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

# 亚太分销模板币种为 USD（美元）
CURRENCY_LABEL = 'USD'
CURRENCY_FMT = '"$"#,##0.00'

# Part 2 - Bill of Material 表头（11 列，A~K，与模板一致）
BOM_HEADERS = [
    'No.',
    'Name',
    'Picture',
    'Material',
    'Specification',
    f'EXW Unit Price ({CURRENCY_LABEL})',
    'QTY(PCS)',
    'Free spare parts(PCS)',
    'Total QTY (PCS)',
    f'Amount({CURRENCY_LABEL})',
    'Remark',
]
# 图片所在列（Picture = 第 3 列，C）
IMAGE_COL_INDEX = 3
# 亚太组物料图片展示尺寸（像素）
AP_IMAGE_WIDTH = 372
AP_IMAGE_HEIGHT = 92

# Part 1 - Information for Each Site 表头（6 个标签，按列放置，非顺序连续列）
#   No.→B, Roof Information→C-D, Row/Table→E, Column/Table→F, Tables→G, Total Capacity(kw)→H-J
SITE_HEADERS = [
    'No.',
    'Roof Information',
    'Row/Table',
    'Column/Table',
    'Tables',
    'Total Capacity(kw)',
]

# 公司抬头/收款信息（默认值，可由 contact_info 覆盖）
COMPANY = {
    'name_en': 'XIAMEN KSENG METALTECH CO.,LTD',
    'address': 'RM 601, Huixin Wealth Centre, No. 891, Fanghu North 2nd Rd, '
               'Huli Dist, Xiamen, Fujian, China',
    'tel': '0086-180 5002 6891',
    'email': 'hannah@xmkseng.com',
}

BANK_INFO = (
    'BENEFICIARYS NAME：XIAMEN KSENG METALTECH CO.,LTD\n'
    'BENEFICIARY ADD:601, Huixin Financial Bldg, No.891, Fanghu North 2nd Rd, '
    'Huli Dist, Xiamen, 361000, Fujian, China\n'
    'BENEFICIARY BANK:Technology Sub-branch,Xiamen,Industrial Bank Co.,Ltd\n'
    'A/C:129971400100037896\n'
    'SWIFT CODE:FJIBCNBA260\n'
    'BANK ADDRESS:No.103, Building 11-13,XiaDaWestVillage Siming District,Xiamen,Fujian'
)


def _set_cell(ws, row, col, val, font=None, align=None, border=None,
              number_format=None, fill=None):
    cell = ws.cell(row=row, column=col, value=val)
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format
    if fill is not None:
        cell.fill = fill
    return cell


def _is_carbon_steel_cached(product, price_info):
    if price_info:
        attr = str(price_info.get('pricing_attribute', '')).strip().upper()
        if attr in CARBON_STEEL_PRICING_ATTRS:
            return True
    return False


def _classify_products(all_products, price_mapping, delete_options=None,
                       always_exclude=False, ap_exclude_options=None):
    """分类：主物料 / 排除(可选)项 / 地桩（参考 ko_simple._classify_products_single_pass）。"""
    delete_options = delete_options or {}
    ap_exclude_options = ap_exclude_options or {}
    all_exclude_opts = {k: True for k in EXCLUDE_ITEM_GROUPS} if always_exclude else {}
    for k in list(ap_exclude_options.keys()) + list(delete_options.keys()):
        all_exclude_opts[k] = True

    main = []
    excluded = []
    pile_products = []
    price_info_cache = {}
    for p in all_products:
        code = str(p.get('code', '')).strip()
        spec = str(p.get('spec', '')).strip()
        cache_key = (code, spec) if spec else code
        if cache_key not in price_info_cache:
            price_info_cache[cache_key] = (
                resolve_price_info(price_mapping, code, spec=spec) if price_mapping else None
            )
        pi = price_info_cache[cache_key]
        p['_price_info'] = pi
        attr = (pi.get('attribute', '') if pi else '') or ''
        if attr.strip() in ('地桩', '地盤杭'):
            p['_is_pile'] = True
            pile_products.append(p)
            continue
        if not _is_valid_product_code(code):
            continue
        qty = p.get('quantity', 0)
        if not qty or qty <= 0:
            continue
        matched_group = _match_exclude_group(p, price_mapping, all_exclude_opts)
        if matched_group:
            if delete_options.get(matched_group):
                continue
            if ap_exclude_options.get(matched_group):
                excluded.append(p)
            else:
                main.append(p)
        else:
            main.append(p)
    return main, excluded, pile_products, price_info_cache


def _calc_base_display_price(product, price_mapping, group, sale_type):
    """基础单价(长度折算后) → 展示单价（直接取数据库价格，无汇率换算）。"""
    code = product.get('code', '')
    price_info = resolve_price_info(price_mapping, code, spec=product.get('spec', ''))
    if not price_info or not has_valid_price_info(price_info):
        return 0.0, '', price_info
    base_price = float(get_temp_base_price(price_info, product, group, sale_type))
    price_unit = price_info.get('unit', '')
    is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
    if is_meter:
        length_mm = extract_length_from_spec(product.get('spec', '')) or 0
        if length_mm > 0:
            base_price = base_price * length_mm / 1000.0
    base_price = apply_temp_preinstall_adjustment(price_info, base_price, product, group, sale_type)
    return base_price, price_unit, price_info


def collect_ap_array(df, price_mapping, config=None, matrix_data=None,
                     pre_parsed_products=None, sale_type='export',
                     coating_thickness=10, delete_options=None,
                     always_exclude_extra_items=False, ap_exclude_options=None,
                     unmatched_products_list=None, **kwargs):
    """收集单个阵列/站点的产品与站点信息（不渲染）。

    返回：
        {'site': {project/rows/cols/tables/kw},
         'items': [{code,name,material,spec,spec_norm,unit_price,qty_total,is_matched}, ...],
         'pile_products': [...], 'matched': int, 'unmatched': int}
    数量已按 set_count（Tables）放大为「全部台数」的总用量，便于后续跨阵列汇总。
    """
    group = '亚太组'
    all_products, array_info, span_info, rows, cols = resolve_products_and_array(
        pre_parsed_products, df, matrix_data,
    )
    main_products, excluded_products, pile_products, _pic = _classify_products(
        all_products, price_mapping, delete_options or {},
        always_exclude_extra_items, ap_exclude_options or {},
    )

    matrix_data = matrix_data or {}
    set_count = matrix_data.get('set_count') or 1
    if not isinstance(set_count, int) or set_count <= 0:
        set_count = 1
    output_kw = matrix_data.get('output_kw') or 0
    project_name = str(matrix_data.get('project_name') or '').strip()

    render_products = main_products + excluded_products
    items = []
    matched_count = 0
    unmatched_count = 0
    local_unmatched = []
    for product in render_products:
        code = str(product.get('code', '')).strip()
        quantity = product.get('quantity', 0)

        display_price, price_unit, price_info = _calc_base_display_price(
            product, price_mapping, group, sale_type
        )
        is_matched = bool(price_info and has_valid_price_info(price_info))
        if is_matched:
            matched_count += 1
        else:
            unmatched_count += 1
            if unmatched_products_list is not None and _is_valid_product_code(code):
                local_unmatched.append({
                    'code': code,
                    'name': product.get('name', ''),
                    'spec': product.get('spec', ''),
                    'quantity': quantity,
                    'unit': price_info.get('unit', '') if price_info else '',
                    'preinstall': normalize_preinstall(product.get('preinstall')),
                    'issue_reason': '数据库无法匹配',
                })

        # 名称：优先取数据库英文名（无论是否匹配到价格），避免使用 BOM 完整工程品名
        en_name = ''
        _name_src = price_info if price_info else (resolve_price_info(price_mapping, code) if price_mapping else None)
        if _name_src:
            # 仅取工程品名--英语（name_en）/ 韩语（name_ko），不回退到中文工程品名（name / BOM name）
            en_name = (_name_src.get('name_en') or _name_src.get('name_ko') or '')
        raw_mat = ((price_info.get('db_material') if price_info and price_info.get('db_material') else None)
                   or (_name_src.get('db_material') if _name_src and _name_src.get('db_material') else None)
                   or product.get('material', ''))
        material = adjust_material_by_coating(translate_material(raw_mat, 'en'), coating_thickness)
        spec_val = product.get('spec', '')

        qty_total = float(quantity or 0) * set_count
        discount_cat = _get_discount_category(price_info, product)
        items.append({
            'code': code,
            'name': en_name,
            'material': material,
            'spec': spec_val,
            'spec_norm': _strip_cjk_spec(spec_val),
            'unit_price': display_price,
            'qty_total': qty_total,
            'is_matched': is_matched,
            'discount_cat': discount_cat,
        })

    pile_scaled = []
    for pp in pile_products:
        s = dict(pp)
        s['quantity'] = float(pp.get('quantity', 0) or 0) * set_count
        pile_scaled.append(s)

    if unmatched_products_list is not None:
        for _up in local_unmatched:
            _up['quantity'] = float(_up.get('quantity', 0) or 0) * set_count
        unmatched_products_list.extend(local_unmatched)

    return {
        'site': {
            'project': project_name,
            'rows': rows,
            'cols': cols,
            'tables': set_count,
            'kw': output_kw,
        },
        'items': items,
        'pile_products': pile_scaled,
        'matched': matched_count,
        'unmatched': unmatched_count,
    }


def aggregate_ap_items(collected_arrays):
    """跨阵列汇总物料：按 (code, 规格归一) 合并，数量累加。保留首次出现顺序。"""
    agg = {}
    order = []
    for arr in collected_arrays:
        for it in arr.get('items', []):
            key = (it['code'], it['spec_norm'])
            if key not in agg:
                agg[key] = dict(it)
                order.append(key)
            else:
                agg[key]['qty_total'] += it['qty_total']
    return [agg[k] for k in order]


def create_ap_quotation_sheet(
        workbook, sites, aggregated_items, price_mapping=None,
        contact_info=None, matrix_data=None, trade_method='EXW',
        coating_thickness=10, sale_type='export',
        image_path=None, image_folder=None, code_to_images=None,
        image_temp_dir=None, image_cache=None,
        ap_discount_rate=100, ap_steel_discount_rate=100, ap_purchased_discount_rate=100,
):
    """创建单页完整报价工作表（Part1 + Part2 + 合计 + 条款 + 银行信息）。"""
    matrix_data = matrix_data or {}
    project_name = str(matrix_data.get('project_name') or '').strip()
    output_kw = matrix_data.get('output_kw') or 0

    sheet_name = extract_main_name(project_name) if project_name else 'Quotation'
    if not sheet_name:
        sheet_name = 'Quotation'
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    original = sheet_name
    counter = 1
    while sheet_name in workbook.sheetnames:
        sheet_name = f"{original}_{counter}"
        counter += 1
    ws = workbook.create_sheet(title=sheet_name)
    print(f"[AP] Creating quotation sheet: {sheet_name}")

    max_col = len(BOM_HEADERS)  # 11 (A~K)
    max_col_letter = get_column_letter(max_col)

    GREEN_FILL = PatternFill(start_color='FF51B777', end_color='FF51B777', fill_type='solid')
    title_font = Font(name='Calibri', size=26, bold=True)
    section_font = Font(name='Calibri', size=16, bold=True)
    info_font = Font(name='Calibri', size=11)
    header_white_bold = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
    normal_font = Font(name='Calibri', size=12)
    bold_font = Font(name='Calibri', size=12, bold=True)
    total_font = Font(name='Calibri', size=14, bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_a = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_a = Alignment(horizontal='right', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    gray_fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
    bottom_border = Border(bottom=thin)

    def _edge_border(c):
        # A、K 为两侧留白列，整张表格体内不绘制边缘框
        return None if c in (1, max_col) else thin_border

    def _apply_outer_frame(top_row, bottom_row):
        # 在最外层绘制黑色边缘框（保留内部各单元格原有边框）
        left_c, right_c = 1, max_col
        for r in range(top_row, bottom_row + 1):
            for c in range(left_c, right_c + 1):
                cell = ws.cell(row=r, column=c)
                ex = cell.border
                cell.border = Border(
                    left=thin if c == left_c else ex.left,
                    right=thin if c == right_c else ex.right,
                    top=thin if r == top_row else ex.top,
                    bottom=thin if r == bottom_row else ex.bottom,
                )

    column_widths = {
        'A': 13, 'B': 21, 'C': 15, 'D': 24, 'E': 18, 'F': 18,
        'G': 18, 'H': 18, 'I': 18, 'J': 18, 'K': 10,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # ---- 抬头：Quotation（不填充绿色，字体加粗）----
    ws.merge_cells(f'A2:{max_col_letter}2')
    _set_cell(ws, 2, 1, 'Quotation', font=title_font, align=center)
    ws.row_dimensions[2].height = 33.35

    contact_info = contact_info or {}
    for _rr in range(4, 8):
        ws.row_dimensions[_rr].height = 19.5
    ws.merge_cells('C4:G4')
    ws.merge_cells('H4:K4')
    ws.merge_cells('H5:K5')
    ws.merge_cells('H6:K6')
    ws.merge_cells('H7:K7')
    _set_cell(ws, 4, 1, 'Project:', font=info_font, align=left_a)
    _set_cell(ws, 4, 3, project_name, font=info_font, align=right_a)
    _addr_lines = COMPANY['address'].split(', ', 3)
    if len(_addr_lines) >= 4:
        _addr_line1 = ', '.join(_addr_lines[:3]) + ','
        _addr_line2 = _addr_lines[3]
    else:
        _addr_line1 = COMPANY['address']
        _addr_line2 = ''
    _set_cell(ws, 4, 8, 'Add.: ' + _addr_line1, font=info_font, align=left_a)
    _set_cell(ws, 5, 8, _addr_line2, font=info_font, align=left_a)
    _set_cell(ws, 5, 1, 'PI No:', font=info_font, align=left_a)
    _set_cell(ws, 6, 1, 'Date:', font=info_font, align=left_a)
    _set_cell(ws, 6, 2, '=TODAY()', font=info_font, align=left_a, number_format='yyyy-mm-dd')
    _set_cell(ws, 6, 8, 'Tel: ' + (contact_info.get('phone') or COMPANY['tel']), font=info_font, align=left_a)
    _set_cell(ws, 7, 1, 'Validity Date:', font=info_font, align=left_a)
    _set_cell(ws, 7, 2, '2 days', font=info_font, align=left_a)
    _set_cell(ws, 7, 8, 'Email: ' + (contact_info.get('tel') or COMPANY['email']), font=info_font, align=left_a)

    # ---- Part 1 - Information for Each Site（不填充绿色，字体加粗）----
    cur = 9
    ws.merge_cells(f'A{cur}:{max_col_letter}{cur}')
    _set_cell(ws, cur, 1, 'Part 1 -  Information for Each Site', font=section_font, align=center)
    ws.row_dimensions[cur].height = 58
    cur += 1

    # 站点表头行（A~K 填充绿色，字体白色加粗；整行不设边缘框）
    ws.row_dimensions[cur].height = 35
    for c in range(1, max_col + 1):
        _set_cell(ws, cur, c, None, font=header_white_bold, align=center, border=None, fill=GREEN_FILL)
    ws.merge_cells(f'C{cur}:D{cur}')
    ws.merge_cells(f'H{cur}:J{cur}')
    _set_cell(ws, cur, 2, 'No.', font=header_white_bold, align=center, border=None, fill=GREEN_FILL)
    _set_cell(ws, cur, 3, 'Roof Information', font=header_white_bold, align=center, border=None, fill=GREEN_FILL)
    _set_cell(ws, cur, 5, 'Row/Table', font=header_white_bold, align=center, border=None, fill=GREEN_FILL)
    _set_cell(ws, cur, 6, 'Column/Table', font=header_white_bold, align=center, border=None, fill=GREEN_FILL)
    _set_cell(ws, cur, 7, 'Tables', font=header_white_bold, align=center, border=None, fill=GREEN_FILL)
    _set_cell(ws, cur, 8, 'Total Capacity(kw)', font=header_white_bold, align=center, border=None, fill=GREEN_FILL)
    cur += 1

    # 站点数据行（每个阵列一行）
    sites_kw_sum = 0.0
    for idx, site in enumerate(sites):
        ws.row_dimensions[cur].height = 65
        site_kw = site.get('kw') or 0
        try:
            sites_kw_sum += float(site_kw)
        except (TypeError, ValueError):
            pass
        site_name = site.get('project') or site.get('name') or project_name or ''
        for c in range(1, max_col + 1):
            _set_cell(ws, cur, c, None, font=normal_font, align=center, border=_edge_border(c))
        ws.merge_cells(f'C{cur}:D{cur}')
        ws.merge_cells(f'H{cur}:J{cur}')
        _set_cell(ws, cur, 2, idx + 1, font=normal_font, align=center, border=thin_border)
        _set_cell(ws, cur, 3, site_name, font=normal_font, align=center, border=thin_border)
        _set_cell(ws, cur, 5, site.get('rows') or '', font=normal_font, align=center, border=thin_border)
        _set_cell(ws, cur, 6, site.get('cols') or '', font=normal_font, align=center, border=thin_border)
        _set_cell(ws, cur, 7, site.get('tables') or '', font=normal_font, align=center, border=thin_border)
        _set_cell(ws, cur, 8, site_kw or '', font=normal_font, align=center, border=thin_border)
        cur += 1

    # Total Project Capacity(kw)（字体加粗；标签合并 B-G）
    ws.row_dimensions[cur].height = 50
    for c in range(1, max_col + 1):
        _set_cell(ws, cur, c, None, font=bold_font, align=center, border=_edge_border(c))
    ws.merge_cells(f'B{cur}:G{cur}')
    ws.merge_cells(f'H{cur}:J{cur}')
    total_cap = output_kw if output_kw else (sites_kw_sum if sites_kw_sum else '')
    _set_cell(ws, cur, 2, 'Total Project Capacity(kw)', font=bold_font, align=right_a, border=thin_border)
    _set_cell(ws, cur, 8, total_cap, font=bold_font, align=center, border=thin_border)
    cur += 2

    # ---- Part 2 - Bill of Material（不填充绿色，字体加粗）----
    ws.merge_cells(f'A{cur}:{max_col_letter}{cur}')
    _set_cell(ws, cur, 1, 'Part 2 - Bill of Material', font=section_font, align=center)
    ws.row_dimensions[cur].height = 58
    cur += 1

    # BOM 表头行（A~K 填充绿色，字体白色；仅保留下框线）
    ws.row_dimensions[cur].height = 33.5
    for c in range(1, max_col + 1):
        _set_cell(ws, cur, c, None, font=header_white_bold, align=center, border=bottom_border, fill=GREEN_FILL)
    for ci, label in enumerate(BOM_HEADERS):
        _set_cell(ws, cur, ci + 1, label, font=header_white_bold, align=center, border=bottom_border, fill=GREEN_FILL)
    cur += 1

    data_start = cur
    sub_total = Decimal('0')
    matched_count = 0
    unmatched_count = 0
    row_code_map = {}

    for idx, item in enumerate(aggregated_items):
        row = data_start + idx
        ws.row_dimensions[row].height = 65
        code = item['code']
        unit_price = item['unit_price']
        qty_total = item['qty_total']
        is_matched = item['is_matched']
        discount_cat = item.get('discount_cat', 'standard')
        if discount_cat == 'steel':
            discount_rate = ap_steel_discount_rate
        elif discount_cat == 'purchased':
            discount_rate = ap_purchased_discount_rate
        else:
            discount_rate = ap_discount_rate

        if is_matched:
            matched_count += 1
        else:
            unmatched_count += 1

        for c in range(1, max_col + 1):
            _set_cell(ws, row, c, None, font=normal_font, align=center, border=bottom_border)

        _set_cell(ws, row, 1, idx + 1, font=normal_font, align=center, border=bottom_border)
        _set_cell(ws, row, 2, item['name'], font=normal_font, align=center, border=bottom_border)
        _set_cell(ws, row, 3, '', font=normal_font, align=center, border=bottom_border)  # Picture
        _set_cell(ws, row, 4, item['material'], font=normal_font, align=center, border=bottom_border)
        _set_cell(ws, row, 5, _strip_cjk_spec(item['spec']), font=normal_font, align=center, border=bottom_border)
        if is_matched and unit_price > 0:
            # EXW Unit Price = 原价 * 折扣率/100，使用公式显示
            _orig = round(unit_price, 6)
            _set_cell(ws, row, 6, f'={_orig}*{discount_rate}/100',
                      font=normal_font, align=center, border=bottom_border, number_format=CURRENCY_FMT)
            sub_total += (Decimal(str(unit_price)) * Decimal(str(qty_total))
                          * Decimal(str(discount_rate)) / Decimal('100'))
        else:
            _set_cell(ws, row, 6, '',
                      font=normal_font, align=center, border=bottom_border, number_format=CURRENCY_FMT)
        qty_display = int(qty_total) if float(qty_total).is_integer() else qty_total
        _set_cell(ws, row, 7, qty_display, font=normal_font, align=center, border=bottom_border)
        _set_cell(ws, row, 8, f'=G{row}*0.001', font=normal_font, align=center, border=bottom_border, number_format='0')
        _set_cell(ws, row, 9, f'=G{row}+H{row}', font=normal_font, align=center, border=bottom_border, number_format='#,##0.000')
        # Amount(USD) 始终设置公式（即使无价格，F 为空按 0 计算）
        _set_cell(ws, row, 10, f'=F{row}*G{row}', font=normal_font, align=center, border=bottom_border, number_format=CURRENCY_FMT)
        _set_cell(ws, row, 11, code, font=normal_font, align=center, border=bottom_border)
        row_code_map[row] = code
        if not is_matched:
            for c in range(1, max_col + 1):
                ws.cell(row=row, column=c).fill = yellow_fill

    data_end = data_start + len(aggregated_items) - 1 if aggregated_items else data_start - 1

    # ---- 合计：EXW Xiamen Total Amount（不填充颜色、无边缘框）----
    cur = data_end + 1
    ws.merge_cells(f'A{cur}:I{cur}')
    _set_cell(ws, cur, 1, f'{trade_method} Xiamen Total Amount :', font=total_font, align=right_a)
    if data_start <= data_end:
        _set_cell(ws, cur, 10, f'=SUM(J{data_start}:J{data_end})',
                  font=total_font, align=center, number_format=CURRENCY_FMT)
    else:
        _set_cell(ws, cur, 10, 0, font=total_font, align=center, number_format=CURRENCY_FMT)
    total_row = cur
    ws.row_dimensions[cur].height = 21
    sub_total_val = float(sub_total)

    # ---- TERMS & CONDITIONS（每行合并 A~K，左对齐）----
    cur += 2
    ws.merge_cells(f'A{cur}:{max_col_letter}{cur}')
    _set_cell(ws, cur, 1, 'TERMS & CONDITIONS :', font=section_font, align=left_a, fill=gray_fill)
    for _c in range(1, max_col + 1):
        ws.cell(row=cur, column=_c).border = bottom_border
    cur += 1
    terms = [
        f'1.Price Term：{trade_method}',
        '2.Payment Term:30% T/T deposit, 70% balance before shipment',
        '3.Guarantee: 10 years warranty, 20 years service life',
        '4.Delivery time：3-4 Weeks after receiving deposit',
    ]
    for t in terms:
        ws.merge_cells(f'A{cur}:{max_col_letter}{cur}')
        _set_cell(ws, cur, 1, t, font=normal_font, align=left_a)
        ws.row_dimensions[cur].height = 25
        cur += 1

    # ---- BANK INFORMATION（每行合并 A~K，左对齐）----
    cur += 1
    ws.merge_cells(f'A{cur}:{max_col_letter}{cur}')
    _set_cell(ws, cur, 1, 'BANK INFORMATION :', font=section_font, align=left_a, fill=gray_fill)
    for _c in range(1, max_col + 1):
        ws.cell(row=cur, column=_c).border = bottom_border
    cur += 1
    ws.merge_cells(f'A{cur}:{max_col_letter}{cur}')
    _set_cell(ws, cur, 1, BANK_INFO, font=normal_font, align=left_a)
    ws.row_dimensions[cur].height = 115
    last_content_row = cur

    # ---- 图片插入（保持原始比例，不变形）----
    image_found_count = 0
    image_not_found_count = 0

    def _col_width_to_px(char_width):
        return int(char_width * 7.5 + 5)

    def _row_height_to_px(pt_height):
        return int(pt_height * 1.33)

    def _image_src_size(path):
        try:
            from PIL import Image as _PIL
            with _PIL.open(path) as _im:
                return _im.size
        except Exception:
            return None, None

    def _mark_missing_image(row_number):
        cell = ws.cell(row=row_number, column=IMAGE_COL_INDEX)
        cell.value = '/'
        cell.alignment = center
        cell.font = normal_font

    for row in range(data_start, data_end + 1):
        product_code = str(row_code_map.get(row, '') or '').strip()
        if not product_code:
            continue
        normalized_code = normalize_lookup_code(product_code)
        img_path = None
        if code_to_images:
            img_path = (code_to_images.get(product_code) or code_to_images.get(normalized_code) or [None])[0]
        if img_path and image_folder:
            col_w = ws.column_dimensions[get_column_letter(IMAGE_COL_INDEX)].width or 8.43
            row_h = ws.row_dimensions[row].height or 15
            avail_w = max(_col_width_to_px(col_w) - IMAGE_PADDING * 2, 20)
            avail_h = max(_row_height_to_px(row_h) - IMAGE_PADDING * 2, 20)
            src_w, src_h = _image_src_size(img_path)
            if not src_w or not src_h:
                src_w, src_h = AP_IMAGE_WIDTH, AP_IMAGE_HEIGHT
            scale = min(avail_w / src_w, avail_h / src_h, 1.0)
            disp_w = max(int(src_w * scale), 20)
            disp_h = max(int(src_h * scale), 20)
            temp_img_path = prepare_image_for_excel(
                img_path, target_width=disp_w, target_height=disp_h,
                temp_dir=image_temp_dir, cache=image_cache,
            )
            final_img_path = temp_img_path or img_path
            if add_image_centered_in_cell(ws, final_img_path, row, IMAGE_COL_INDEX, img_width=disp_w, img_height=disp_h):
                image_found_count += 1
            else:
                image_not_found_count += 1
                _mark_missing_image(row)
        else:
            image_not_found_count += 1
            _mark_missing_image(row)
    print(f"   [AP] Images: found={image_found_count}, not_found={image_not_found_count}")

    _apply_outer_frame(1, last_content_row)

    apply_print_setup(ws, 'ap_common')

    quotation_product_codes = {str(it['code']).strip() for it in aggregated_items if it.get('code')}

    return {
        'sheet_name': sheet_name,
        'quotation_product_codes': quotation_product_codes,
        'valid_products': len(aggregated_items),
        'total_weight': 0,
        'total_price': sub_total_val,
        'matched_count': matched_count,
        'unmatched_count': unmatched_count,
        'image_found_count': image_found_count,
        'image_not_found_count': image_not_found_count,
        'sub_total_row': total_row,
        'total_row': total_row,
        'detail_data_end_row': data_end,
        'is_complex': False,
        'matrix_data': matrix_data,
    }
