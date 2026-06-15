from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins
from decimal import Decimal
import os
import re

from backend.core.shared.text_utils import _CJK_RE, _strip_cjk_spec


def _lookup_gate_spec_from_db(gate_style_code):
    try:
        from backend.repositories.fence_gate_style_repository import get_gate_style
        gs = get_gate_style(gate_style_code)
        if gs:
            w = int(gs.get('width', 0) or 0)
            h = int(gs.get('height', 0) or 0)
            if w > 0 and h > 0:
                return h, w
    except Exception:
        pass
    return None, None

from backend.core.shared.price_utils import resolve_price_info, has_valid_price_info, round_to_2_decimal
from backend.core.shared.text_utils import normalize_lookup_code
from backend.core.shared.product_utils import _is_valid_product_code
from backend.core.shared.weight_utils import extract_length_from_spec
from backend.core.material_translate import translate_material

SM_FONT = Font(name='Yu Gothic UI', size=10)
SM_FONT_BOLD = Font(name='Yu Gothic UI', size=10, bold=True)
SM_FONT_BOLD_10 = Font(name='Yu Gothic UI', size=10, bold=True)
SM_FONT_TITLE = Font(name='Yu Gothic UI', size=16, bold=True)
SM_FONT_RED = Font(name='Yu Gothic UI', size=10, bold=True, color='FF0000')
SM_FONT_RED_10 = Font(name='Yu Gothic UI', size=10, bold=True, color='FF0000')
SM_FONT_HEADER = Font(name='Yu Gothic UI', size=10, bold=True)
FONT_10 = SM_FONT
FONT_10_BOLD = SM_FONT_BOLD
FONT_11 = Font(name='Yu Gothic UI', size=11)
FONT_11_BOLD = Font(name='Yu Gothic UI', size=11, bold=True)

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT_A = Alignment(horizontal='right', vertical='center', wrap_text=True)
LEFT_A = Alignment(horizontal='left', vertical='center', wrap_text=True)

NUM_FMT = '"US$" #,##0.00'
NUM_FMT = '#,##0.00'

DEFAULT_NV_DISCOUNT_RATE = Decimal('0.72')
DEFAULT_NV_FENCE_DISCOUNT_RATE = Decimal('0.94')

DETAIL_LOGO_WIDTH = 188
DETAIL_LOGO_HEIGHT = 50

DETAIL_COL_WIDTHS = {
    'A': 3, 'B': 10, 'C': 20, 'D': 20, 'E': 15, 'F': 15, 'G': 13, 'H': 15, 'I': 18, 'J': 10, 'K': 10,
}

SUMMARY_COL_WIDTHS = {
    'A': 5, 'B': 8, 'C': 8, 'D': 8, 'E': 8, 'F': 8,
    'G': 10, 'H': 10, 'I': 13, 'J': 13, 'K': 13, 'L': 8, 'M': 8,
}

GRAY_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
LIGHT_BLUE_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

_LOGO_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))), 'input', '集团标1.png')
_FENCE_LOGO_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))), 'input', '集团标2.png')


def _insert_logo(ws, cell_ref, width=120, height=40):
    if not os.path.isfile(_LOGO_PATH):
        return
    try:
        img = XLImage(_LOGO_PATH)
        img.width = width
        img.height = height
        ws.add_image(img, cell_ref)
    except Exception:
        pass


def _set(ws, r, c, val, font=SM_FONT, align=CENTER, border=True):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font
    cell.alignment = align
    if border:
        cell.border = THIN_BORDER
    return cell


def _strip_decimal_zero(val):
    if val is None or val == '':
        return val
    if isinstance(val, (int, float)):
        return int(val) if float(val) == int(val) else val
    return re.sub(r'(\d+)\.0(?!\d)', r'\1', str(val))


def _get_nv_rate_value(nv_params, primary_key, alias_key, default):
    params = nv_params or {}
    value = params.get(primary_key)
    if value in (None, ''):
        value = params.get(alias_key)
    if value in (None, ''):
        value = default
    return value


def _extract_fence_height_mm_from_style(style_code):
    if not style_code:
        return ''
    parts = str(style_code).split('-')
    if len(parts) < 2:
        return ''
    try:
        return str(int(parts[-1]) * 10)
    except (ValueError, TypeError):
        return ''


def _format_mm_to_meter_text(length_mm):
    if length_mm in (None, ''):
        return ''
    try:
        meters = Decimal(str(length_mm)) / Decimal('1000')
    except Exception:
        return ''
    text = format(meters.normalize(), 'f')
    return text.rstrip('0').rstrip('.') or '0'


def _format_pile_spec_display(spec):
    raw_spec = _strip_cjk_spec(spec or '')
    if not raw_spec:
        return ''

    length_mm = extract_length_from_spec(spec)
    if length_mm:
        return _strip_decimal_zero(length_mm)

    match = re.search(r'L\s*[:=]?\s*(\d+(?:\.\d+)?)', raw_spec, re.IGNORECASE)
    if match:
        return match.group(1)

    match_num = re.search(r'(\d+(?:\.\d+)?)', raw_spec)
    if match_num:
        return match_num.group(1)

    return raw_spec[:8]


def _resolve_pile_price(price_mapping, code, spec=None):
    if not price_mapping or not code:
        return None
    pi = resolve_price_info(price_mapping, code, spec=spec)
    if pi and has_valid_price_info(pi):
        return pi
    normalized = normalize_lookup_code(code)
    cleaned = re.sub(r'[-_\s]', '', normalized)
    if cleaned != normalized:
        for key, val in price_mapping.items():
            if re.sub(r'[-_\s]', '', normalize_lookup_code(str(key))) == cleaned:
                if has_valid_price_info(val):
                    print(f"   🔧 杭価格フォールバック一致: bom={code} → db={key}")
                    return val
    if not pi:
        print(f"   ⚠ 杭価格 lookup全失敗: code={code}, normalized={normalized}, cleaned={cleaned}")
    return pi


def create_nv_detail_sheet(workbook, array_info, bom_products, price_mapping,
                           sheet_prefix=None, matrix_data=None,
                           unmatched_products_out=None,
                           coating_thickness=10, nv_params=None,
                           pile_products=None,
                           code_to_images=None, image_temp_dir=None,
                           image_cache=None, angle_override=None,
                           need_weight_code=False, need_weight=False, need_code=False, missing_boards=0,
                           is_inverter=False, span_ew_override=None):
    from backend.core.shared.image_utils import (
        prepare_image_for_excel, add_image_centered_in_cell,
    )

    nv_params = nv_params or {}
    matrix_data = matrix_data or {}

    _add_code = need_code or need_weight_code
    _add_weight = need_weight or need_weight_code
    _extra = (1 if _add_code else 0) + (1 if _add_weight else 0)
    _col_end = 9 + _extra
    _code_col = 10 if _add_code else None
    _weight_col = (10 + (1 if _add_code else 0)) if _add_weight else None
    _merge_end = chr(ord('A') + _col_end - 1)

    img_w, img_h, img_col, img_padding = 65, 50, 5, 2

    def col_width_to_px(cw):
        return int(cw * 7.5 + 5)

    def row_height_to_px(rh):
        return int(rh * 1.33)

    def fit_image(ws, r, c, mw, mh, pad=img_padding):
        cw_val = ws.column_dimensions[get_column_letter(c)].width or 8.43
        rh_val = ws.row_dimensions[r].height or 15
        aw = col_width_to_px(cw_val) - pad * 2
        ah = row_height_to_px(rh_val) - pad * 2
        aw = max(aw, 20)
        ah = max(ah, 20)
        scale = min(aw / mw, ah / mh, 1.0)
        return int(mw * scale), int(mh * scale)

    if sheet_prefix:
        sheet_name = sheet_prefix
    else:
        sheet_name = str(array_info.get('table_qty', 1))
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    original = sheet_name
    counter = 1
    while sheet_name in workbook.sheetnames:
        sheet_name = f"{original}_{counter}"
        counter += 1

    ws = workbook.create_sheet(title=sheet_name)

    for col, width in DETAIL_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 30
    for row in range(2, 6):
        ws.row_dimensions[row].height = 20
    ws.row_dimensions[6].height = 30
    ws.row_dimensions[7].height = 5
    ws.row_dimensions[8].height = 30

    ws.merge_cells('A2:C5')
    for r in range(2, 6):
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN_BORDER
    if os.path.isfile(_LOGO_PATH):
        try:
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils import column_width_to_pixels, row_height_to_pixels

            _logo_img = XLImage(_LOGO_PATH)
            _logo_w, _logo_h = DETAIL_LOGO_WIDTH, DETAIL_LOGO_HEIGHT
            _logo_img.width = _logo_w
            _logo_img.height = _logo_h

            _range_w = sum(
                column_width_to_pixels(ws.column_dimensions[get_column_letter(c)].width or 8.43)
                for c in range(1, 4)
            )
            _range_h = sum(
                row_height_to_pixels(ws.row_dimensions[r].height or 15)
                for r in range(3, 6)
            )

            EMU = 9525
            _logo_img.anchor = OneCellAnchor(
                _from=AnchorMarker(
                    col=1,
                    colOff=int((_range_w - _logo_w) / 2 * EMU),
                    row=2,
                    rowOff=int((_range_h - _logo_h) / 2 * EMU),
                ),
                ext=XDRPositiveSize2D(
                    cx=int(_logo_w * EMU),
                    cy=int(_logo_h * EMU),
                ),
            )
            ws.add_image(_logo_img)
        except Exception:
            _insert_logo(ws, 'B3', width=DETAIL_LOGO_WIDTH, height=DETAIL_LOGO_HEIGHT)

    _title_merge_end = _merge_end
    _title_col_end = _col_end
    ws.merge_cells(f'A1:{_title_merge_end}1')
    if is_inverter:
        _detail_title = str((matrix_data or {}).get('project_name') or '').strip() or 'パワコン部材リスト'
    else:
        _detail_title = str((matrix_data or {}).get('project_name') or '').strip() or '各アレイ部材リスト'
    ws['A1'] = _detail_title
    ws['A1'].font = Font(name='Yu Gothic UI', bold=True, size=16)
    ws['A1'].alignment = CENTER
    for c in range(1, _title_col_end + 1):
        ws.cell(row=1, column=c).border = THIN_BORDER

    rows_val = array_info.get('rows', '')
    cols_val = array_info.get('cols', '')
    qty_val = array_info.get('table_qty', 1)
    panel_count = int(rows_val) * int(cols_val) if rows_val and cols_val else 0
    module_wattage = matrix_data.get('module_wattage') or 0
    panel_size = matrix_data.get('panel_spec') or matrix_data.get('module_size') or ''
    wind_speed = matrix_data.get('max_wind_speed') or ''
    snow_load = matrix_data.get('max_snow_load') or ''
    angle = angle_override if angle_override is not None else (matrix_data.get('angle') or '')
    if isinstance(angle, float) and angle != angle:
        angle = ''
    if str(angle).lower() == 'nan' or str(angle).lower().strip('°') == 'nan':
        angle = ''
    ground_height = nv_params.get('ground_height') or matrix_data.get('ground_height') or ''
    span_ew = span_ew_override if span_ew_override is not None else (nv_params.get('span_ew') or 2700)
    sales_name = nv_params.get('sales_name') or 'Nanami'
    sales_phone = nv_params.get('sales_phone') or '+86-137-7466-5835'
    sales_fax = nv_params.get('sales_fax') or '0086-592-5738212'
    sales_tel = nv_params.get('sales_tel') or '0086-592-5767152'

    ws['A6'] = '〒361-009 中国厦門湖里区枋湖北二路891号 匯鑫財富大厦11-12階'
    ws['A6'].font = SM_FONT
    ws['A6'].alignment = LEFT_A
    ws.merge_cells('A6:C6')
    for c in range(1, 4):
        ws.cell(row=6, column=c).border = THIN_BORDER

    _set(ws, 2, 4, f'Sales man:{sales_name}', align=LEFT_A)
    _set(ws, 3, 4, f'Mob:{sales_phone}', align=LEFT_A)
    _set(ws, 4, 4, f'Tel : {sales_tel}', align=LEFT_A)
    _set(ws, 5, 4, f'Fax: {sales_fax}', align=LEFT_A)

    _set(ws, 2, 5, 'パネル設置角度', align=RIGHT_A)
    if is_inverter:
        _set(ws, 2, 6, '/')
    else:
        angle_display = angle.rstrip('°').strip() if angle else ''
        angle_display = _strip_decimal_zero(angle_display)
        _set(ws, 2, 6, f'横置き{angle_display}°' if angle_display else '', align=LEFT_A)
    _set(ws, 2, 7, 'パネルサイズ', align=RIGHT_A)
    if is_inverter:
        _set(ws, 2, 8, '/')
    else:
        _set(ws, 2, 8, f'{panel_size}mm' if panel_size else '', align=LEFT_A)

    _quality_merge_end = _merge_end
    ws.merge_cells(f'I2:{_quality_merge_end}6')
    quality_text = (
        '品質保証:\n10年間品質保証、それに、20年以上使える\n'
        '参考標準:\nJIS C 8955太陽電池アレイ用支持物設計標準'
    )
    _set(ws, 2, 9, quality_text, font=Font(name='Yu Gothic UI', size=9), align=LEFT_A)
    for r in range(2, 7):
        ws.cell(row=r, column=9).border = THIN_BORDER
    if _add_code:
        for r in range(2, 7):
            ws.cell(row=r, column=_code_col).border = THIN_BORDER
    if _add_weight:
        for r in range(2, 7):
            ws.cell(row=r, column=_weight_col).border = THIN_BORDER

    _set(ws, 3, 5, '最大風速', align=RIGHT_A)
    _set(ws, 3, 6, _strip_decimal_zero(wind_speed), align=LEFT_A)
    _set(ws, 3, 7, '発電量/PC', align=RIGHT_A)
    if is_inverter:
        _set(ws, 3, 8, '/')
    else:
        _cell_mw = _set(ws, 3, 8, module_wattage, align=LEFT_A)
        if module_wattage:
            _cell_mw.number_format = '#,##0 "Wp"'

    _set(ws, 4, 5, '垂直積雪量', align=RIGHT_A)
    _set(ws, 4, 6, _strip_decimal_zero(snow_load), align=LEFT_A)
    _set(ws, 4, 7, '発電量(W)/基', align=RIGHT_A)
    if is_inverter:
        _set(ws, 4, 8, '/')
    elif module_wattage and rows_val and cols_val and span_ew:
        panel_per_set = int(rows_val) * int(cols_val) if rows_val and cols_val else 0
        mb_val = int(missing_boards) if missing_boards else 0
        _cell_gen = _set(ws, 4, 8, f'=H3*({panel_per_set}+{mb_val})', align=LEFT_A)
        _cell_gen.number_format = '#,##0 "Wp"'
    else:
        _set(ws, 4, 8, '')

    _set(ws, 5, 5, '地上高さ', align=RIGHT_A)
    _set(ws, 5, 6, ground_height, align=LEFT_A)
    _set(ws, 5, 7, 'スパン(E/W)', align=RIGHT_A)
    if is_inverter:
        _set(ws, 5, 8, '/')
    else:
        _cell_span = _set(ws, 5, 8, span_ew, align=LEFT_A)
        if span_ew:
            _cell_span.number_format = '#,##0 "mm"'

    _set(ws, 6, 4, 'パネル数')
    if is_inverter:
        _set(ws, 6, 5, '/')
        _set(ws, 6, 6, '')
    else:
        _set(ws, 6, 5, f'{rows_val}段' if rows_val else '')
        _set(ws, 6, 6, f'{cols_val}列' if cols_val else '')
    _set(ws, 6, 7, 'セット数')
    _set(ws, 6, 8, qty_val)

    _hdr_merge_end = _merge_end
    _hdr_col_end = _col_end
    ws.merge_cells(f'A7:{_hdr_merge_end}7')
    for c in range(1, _hdr_col_end + 1):
        ws.cell(row=7, column=c).border = THIN_BORDER

    YELLOW_FILL = PatternFill(start_color='FFDDB3', end_color='FFDDB3', fill_type='solid')
    headers = ['部品名称', '材質', '写真', '規格', '単価(EXW US$)', '数量（PCS)', '総金額(EXW US$)']
    if _add_code:
        headers += ['品番']
    if _add_weight:
        headers += ['重量(KG)']
    ws.merge_cells('A8:B8')
    cell_num = ws.cell(row=8, column=1, value='番号')
    cell_num.font = SM_FONT_BOLD
    cell_num.alignment = CENTER
    cell_num.border = THIN_BORDER
    cell_num.fill = LIGHT_BLUE_FILL
    ws.cell(row=8, column=2).border = THIN_BORDER
    ws.cell(row=8, column=2).fill = LIGHT_BLUE_FILL
    for i, h in enumerate(headers, 3):
        cell = ws.cell(row=8, column=i, value=h)
        cell.font = SM_FONT_BOLD
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.fill = LIGHT_BLUE_FILL

    YELLOW_FILL = PatternFill(start_color='FFDDB3', end_color='FFDDB3', fill_type='solid')
    data_font = SM_FONT
    data_start = 9
    total_price_sum = Decimal('0')
    matched_count = 0
    name_field = 'name_ja'
    _detail_table_qty = int(array_info.get('table_qty', 1)) if array_info else 1
    from backend.core.shared.price_utils import _get_discount_category
    _cat_row_map = {}

    for row_idx in range(data_start, data_start + len(bom_products)):
        ws.row_dimensions[row_idx].height = 40

    for idx, product in enumerate(bom_products):
        row = data_start + idx
        product_code = product.get('code', '')
        price_info = resolve_price_info(price_mapping, product_code, spec=product.get('spec', ''))

        display_name = (
            price_info.get(name_field)
            or price_info.get('name_ko')
            or price_info.get('name')
            or product.get('name', '')
        ) if price_info else product.get('name', '')

        ws.merge_cells(f'A{row}:B{row}')
        _set(ws, row, 1, idx + 1)
        _set(ws, row, 3, display_name)

        raw_material = (price_info.get('db_material') if price_info and price_info.get('db_material') else None) or product.get('material', '')
        display_material = ''
        if raw_material:
            mat = str(raw_material).replace('&', '+')
            if coating_thickness in (15, 18):
                mat = mat.replace('304', '316')
                mat = mat.replace('/316', '')
            else:
                mat = mat.replace('/316', '')
            mat = translate_material(mat, 'ja')
            display_material = mat
            if coating_thickness in (15, 18):
                display_material += f'  {coating_thickness}um'

        _set(ws, row, 4, display_material)
        _set(ws, row, 5, '')
        _set(ws, row, 6, product.get('spec', ''))

        quantity = product.get('quantity', 0)
        unit_price = 0
        display_unit_price = 0
        price_unit = ''
        is_matched = False

        if price_info and has_valid_price_info(price_info):
            unit_price = float(price_info['price'])
            price_unit = price_info.get('unit', '')
            matched_count += 1
            is_matched = True
            _cat_row_map[row] = _get_discount_category(price_info, product)
        else:
            _cat_row_map[row] = _get_discount_category(price_info, product)
            if unmatched_products_out is not None and quantity > 0 and _is_valid_product_code(product_code):
                unmatched_products_out.append({
                    'code': product_code,
                    'name': product.get('name', ''),
                    'spec': product.get('spec', ''),
                    'material': raw_material,
                    'quantity': quantity * _detail_table_qty,
                })

        is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
        length_mm = Decimal('0')
        if is_meter:
            length_mm = Decimal(str(extract_length_from_spec(product.get('spec')) or 0))

        if is_meter and length_mm > 0:
            display_unit_price = float(Decimal(str(unit_price)) * length_mm / Decimal('1000'))
        else:
            display_unit_price = unit_price

        if display_unit_price > 0:
            cell = ws.cell(row=row, column=7, value=display_unit_price)
            cell.border = THIN_BORDER
            cell.number_format = NUM_FMT
            cell.font = data_font
            cell.alignment = CENTER
        else:
            _set(ws, row, 7, '')

        if quantity > 0:
            _set(ws, row, 8, int(quantity) if quantity % 1 == 0 else quantity)
        else:
            _set(ws, row, 8, '')

        total_price = Decimal('0')
        if display_unit_price > 0 and quantity > 0:
            total_price = Decimal(str(display_unit_price)) * Decimal(str(quantity))

        total_price_rounded = round_to_2_decimal(total_price)
        if total_price_rounded > 0:
            cell = ws.cell(row=row, column=9, value=float(total_price))
            cell.border = THIN_BORDER
            cell.number_format = NUM_FMT
            cell.font = data_font
            cell.alignment = CENTER
            total_price_sum += total_price_rounded
        else:
            _set(ws, row, 9, '')

        if _add_code:
            ws.cell(row=row, column=_code_col, value=product_code).border = THIN_BORDER
            ws.cell(row=row, column=_code_col).alignment = CENTER
            ws.cell(row=row, column=_code_col).font = data_font
        if _add_weight:
            weight_val = '/'
            if price_info:
                w = price_info.get('db_weight')
                price_unit = price_info.get('unit', '')
                if w:
                    is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
                    if is_meter:
                        length_mm = extract_length_from_spec(product.get('spec'))
                        if length_mm and length_mm > 0:
                            weight_val = round(float(w) / 1000 * length_mm, 4)
                        else:
                            weight_val = float(w)
                    else:
                        weight_val = float(w)
            ws.cell(row=row, column=_weight_col, value=weight_val).border = THIN_BORDER
            ws.cell(row=row, column=_weight_col).alignment = CENTER
            ws.cell(row=row, column=_weight_col).font = data_font

        if not is_matched:
            _fill_end = _col_end
            for c in range(1, _fill_end + 1):
                ws.cell(row=row, column=c).fill = YELLOW_FILL

    rows_to_delete = []
    for r in range(data_start, data_start + len(bom_products)):
        name_val = ws.cell(row=r, column=3).value
        if name_val is None or name_val == '':
            continue
        qty_val_cell = ws.cell(row=r, column=8).value
        if qty_val_cell is None or qty_val_cell == '' or (isinstance(qty_val_cell, (int, float)) and qty_val_cell == 0):
            rows_to_delete.append(r)

    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r)

    deleted_count = len(rows_to_delete)
    if deleted_count:
        print(f"   🗑️ 数量=0の行を {deleted_count} 行削除")

    new_seq = 1
    for r in range(data_start, ws.max_row + 1):
        name_val = ws.cell(row=r, column=3).value
        if name_val and name_val != '':
            ws.merge_cells(f'A{r}:B{r}')
            ws.cell(row=r, column=1, value=new_seq)
            ws.cell(row=r, column=1).font = SM_FONT
            ws.cell(row=r, column=1).alignment = CENTER
            new_seq += 1

    for r in range(data_start, ws.max_row + 1):
        name_val = ws.cell(row=r, column=3).value
        if name_val and name_val != '':
            ws.cell(row=r, column=9, value=f'=G{r}*H{r}')
            ws.cell(row=r, column=9).font = SM_FONT
            ws.cell(row=r, column=9).alignment = CENTER
            ws.cell(row=r, column=9).number_format = NUM_FMT

    data_end = data_start + len(bom_products) - 1 - deleted_count

    _cat_row_remapped = {}
    for orig_row, cat in _cat_row_map.items():
        if orig_row in set(rows_to_delete):
            continue
        offset = sum(1 for d in rows_to_delete if d < orig_row)
        new_row = orig_row - offset
        _cat_row_remapped[new_row] = cat

    row_product_map = {}
    for idx, product in enumerate(bom_products):
        orig_row = data_start + idx
        pc = product.get('code', '')
        row_product_map[orig_row] = str(pc).strip() if pc else ''

    survived = [r for r in sorted(row_product_map.keys()) if r not in set(rows_to_delete)]
    row_remap = {}
    for orig_row in survived:
        offset = sum(1 for d in rows_to_delete if d < orig_row)
        row_remap[orig_row - offset] = row_product_map[orig_row]

    image_found_count = 0
    image_not_found_count = 0
    for row, product_code in row_remap.items():
        if row < data_start or row > data_end:
            continue
        if not product_code:
            continue
        normalized_code = normalize_lookup_code(product_code)

        img_path = None
        if code_to_images and product_code in code_to_images and code_to_images[product_code]:
            img_path = code_to_images[product_code][0]
        elif code_to_images and normalized_code in code_to_images and code_to_images[normalized_code]:
            img_path = code_to_images[normalized_code][0]

        if not img_path and price_mapping:
            pi = resolve_price_info(price_mapping, product_code)
            if pi and pi.get('image_bytes'):
                img_bytes = pi['image_bytes']
                img_ext = pi.get('image_ext', '.png')
                if image_temp_dir:
                    db_img_name = f"db_{normalized_code or product_code}{img_ext}"
                    db_img_path = os.path.join(image_temp_dir, db_img_name)
                    if db_img_path not in (image_cache or {}):
                        with open(db_img_path, 'wb') as _f:
                            _f.write(img_bytes)
                        if image_cache is not None:
                            image_cache.pop(db_img_path, None)
                    img_path = db_img_path

        if img_path:
            fit_w, fit_h = fit_image(ws, row, img_col, img_w, img_h)
            temp_img_path = prepare_image_for_excel(
                img_path,
                target_width=fit_w,
                target_height=fit_h,
                temp_dir=image_temp_dir,
                cache=image_cache,
            )
            final_img_path = temp_img_path if temp_img_path else img_path
            success = add_image_centered_in_cell(
                ws,
                final_img_path,
                row,
                img_col,
                img_width=fit_w,
                img_height=fit_h,
            )
            if success:
                image_found_count += 1
                continue
        image_not_found_count += 1
        ws.cell(row=row, column=5, value='/').alignment = CENTER

    if image_found_count or image_not_found_count:
        print(f"   🖼️ {sheet_name}: 挿入 {image_found_count} 枚, 未検出 {image_not_found_count} 枚")

    # 修复 create_nv_detail_sheet 函数中的这三行（大约在第 558-597 行附近）

    sub_row = data_end + 1
    _sub_border_end = _col_end
    _set(ws, sub_row, 1, '')
    _set(ws, sub_row, 2, '')
    ws.merge_cells(f'C{sub_row}:H{sub_row}')
    _set(ws, sub_row, 3, 'SUB-TOTAL-（EXW）1基パワコン合計' if is_inverter else 'SUB-TOTAL-（EXW）1基架台合計', font=FONT_10_BOLD, align=Alignment(horizontal='right', vertical='center'))
    ws.cell(row=sub_row, column=1).border = Border(left=THIN_BORDER.left, right=Side(style=None), top=THIN_BORDER.top, bottom=THIN_BORDER.bottom)
    ws.cell(row=sub_row, column=2).border = Border(left=Side(style=None), right=Side(style=None), top=THIN_BORDER.top, bottom=THIN_BORDER.bottom)
    ws.cell(row=sub_row, column=3).border = Border(left=Side(style=None), right=Side(style=None), top=THIN_BORDER.top, bottom=THIN_BORDER.bottom)
    for c in range(4, _sub_border_end + 1):
        ws.cell(row=sub_row, column=c).border = THIN_BORDER
    ws.cell(row=sub_row, column=8).border = Border(left=Side(style=None), right=THIN_BORDER.right, top=THIN_BORDER.top, bottom=THIN_BORDER.bottom)
    cell_h = ws.cell(row=sub_row, column=9, value=f'=SUM(I{data_start}:I{data_end})')
    cell_h.font = FONT_10_BOLD
    cell_h.border = THIN_BORDER
    cell_h.alignment = CENTER
    cell_h.number_format = NUM_FMT
    if _add_code:
        ws.cell(row=sub_row, column=_code_col).border = THIN_BORDER
    if _add_weight:
        ws.cell(row=sub_row, column=_weight_col).border = THIN_BORDER
    ws.row_dimensions[sub_row].height = 25

    total_row = sub_row + 1
    _set(ws, total_row, 1, '')
    _set(ws, total_row, 2, '')
    ws.merge_cells(f'C{total_row}:H{total_row}')
    _set(ws, total_row, 3, f'TOTAL-（EXW）{qty_val}基パワコン合計' if is_inverter else f'TOTAL-（EXW）{qty_val}基架台合計', font=FONT_10_BOLD, align=Alignment(horizontal='right', vertical='center'))
    ws.cell(row=total_row, column=1).border = Border(left=THIN_BORDER.left, right=Side(style=None), top=THIN_BORDER.top, bottom=THIN_BORDER.bottom)
    ws.cell(row=total_row, column=2).border = Border(left=Side(style=None), right=Side(style=None), top=THIN_BORDER.top, bottom=THIN_BORDER.bottom)
    ws.cell(row=total_row, column=3).border = Border(left=Side(style=None), right=Side(style=None), top=THIN_BORDER.top, bottom=THIN_BORDER.bottom)
    for c in range(4, _sub_border_end + 1):
        ws.cell(row=total_row, column=c).border = THIN_BORDER
    ws.cell(row=total_row, column=8).border = Border(left=Side(style=None), right=THIN_BORDER.right, top=THIN_BORDER.top, bottom=THIN_BORDER.bottom)
    cell_total = ws.cell(row=total_row, column=9, value=f'=H6*I{sub_row}')
    cell_total.font = FONT_10_BOLD
    cell_total.border = THIN_BORDER
    cell_total.alignment = CENTER
    cell_total.number_format = NUM_FMT
    if _add_code:
        ws.cell(row=total_row, column=_code_col).border = THIN_BORDER
    if _add_weight:
        ws.cell(row=total_row, column=_weight_col).border = THIN_BORDER
    ws.row_dimensions[total_row].height = 25

    watt_row = total_row + 1

    if is_inverter:
        _set(ws, watt_row, 1, '')
        _set(ws, watt_row, 2, '')
        ws.merge_cells(f'C{watt_row}:H{watt_row}')
        _set(ws, watt_row, 3, '', font=FONT_10_BOLD, align=Alignment(horizontal='right', vertical='center'))
        for c in range(1, _col_end + 1):
            ws.cell(row=watt_row, column=c).border = THIN_BORDER
    else:
        _set(ws, watt_row, 1, '')
        _set(ws, watt_row, 2, '')
        ws.merge_cells(f'C{watt_row}:H{watt_row}')
        _set(ws, watt_row, 3, '1Ｗあたり金額', font=FONT_10_BOLD, align=Alignment(horizontal='right', vertical='center'))
        ws.cell(row=watt_row, column=1).border = Border(left=THIN_BORDER.left, right=Side(style=None), top=THIN_BORDER.top, bottom=THIN_BORDER.bottom)
        ws.cell(row=watt_row, column=2).border = Border(left=Side(style=None), right=Side(style=None), top=THIN_BORDER.top, bottom=THIN_BORDER.bottom)
        ws.cell(row=watt_row, column=3).border = Border(left=Side(style=None), right=THIN_BORDER.right, top=THIN_BORDER.top, bottom=THIN_BORDER.bottom)
        for c in range(4, _sub_border_end + 1):
            ws.cell(row=watt_row, column=c).border = THIN_BORDER
        ws.cell(row=watt_row, column=8).border = Border(left=Side(style=None), right=THIN_BORDER.right, top=THIN_BORDER.top, bottom=THIN_BORDER.bottom)
        cell_watt = ws.cell(row=watt_row, column=9, value=f'=I{sub_row}/H4')
        cell_watt.font = FONT_10_BOLD
        cell_watt.border = THIN_BORDER
        cell_watt.alignment = CENTER
        cell_watt.number_format = '0.00'
        if _add_code:
            ws.cell(row=watt_row, column=_code_col).border = THIN_BORDER
        if _add_weight:
            ws.cell(row=watt_row, column=_weight_col).border = THIN_BORDER
    ws.row_dimensions[watt_row].height = 25

    pile_data_start_row = 0
    pile_data_end_row = 0
    pile_total_per_base = 0.0
    pile_total_qty = 0
    pile_products = pile_products or []
    _pile_img_cache = {}
    pile_display_name = ''
    if pile_products:
        import tempfile as _tf
        _pile_tmp = _tf.mkdtemp(prefix='nv_pile_img_')
        pile_start_row = watt_row + 1

        pile_seq = 1
        pile_data_start_row = pile_start_row
        rendered_pile_count = 0
        for p in pile_products:
            qty_check = int(p.get('quantity', 0) or 0)
            if qty_check <= 0:
                continue
            row = pile_data_start_row + pile_seq - 1
            ws.row_dimensions[row].height = 40
            code = str(p.get('code', '') or '')
            price_info = _resolve_pile_price(price_mapping, code, spec=p.get('spec', ''))
            display_name = (
                price_info.get('name_ja')
                or price_info.get('name')
                or 'スクリュー杭'
            ) if price_info else 'スクリュー杭'
            if not pile_display_name:
                pile_display_name = display_name
            raw_material = (price_info.get('db_material') if price_info and price_info.get('db_material') else None) or p.get('material', '')
            display_material = ''
            if raw_material:
                mat = str(raw_material).replace('&', '+')
                if coating_thickness in (15, 18):
                    mat = mat.replace('304', '316')
                    mat = mat.replace('/316', '')
                else:
                    mat = mat.replace('/316', '')
                mat = translate_material(mat, 'ja')
                display_material = mat
                if coating_thickness in (15, 18):
                    display_material += f'  {coating_thickness}um'
            spec = _format_pile_spec_display(p.get('spec', ''))
            qty = qty_check

            unit_price = 0.0
            display_unit_price = 0.0
            pile_is_matched = False
            YELLOW_FILL = PatternFill(start_color='FFDDB3', end_color='FFDDB3', fill_type='solid')
            if price_info and has_valid_price_info(price_info):
                unit_price = float(price_info['price'])
                pile_is_matched = True
            else:
                print(f"   ⚠ 杭価格未検出: code={code}, spec={p.get('spec', '')}, price_info={'found' if price_info else 'None'}")
            pricing_unit = (price_info.get('unit', '') if price_info else '') or ''
            length_mm = float(p.get('length', 0) or 0)
            if not length_mm:
                length_mm = float(extract_length_from_spec(p.get('spec', '')) or 0)
            if not length_mm:
                _spec_raw = str(p.get('spec', '') or '').strip()
                _m = re.search(r'(\d+(?:\.\d+)?)', _spec_raw)
                if _m:
                    length_mm = float(_m.group(1))
            if not pile_is_matched:
                display_unit_price = ''
            elif pricing_unit == '米' and length_mm > 0:
                display_unit_price = (length_mm / 1000) * unit_price
            elif pricing_unit == '米' and length_mm <= 0:
                display_unit_price = unit_price
                print(f"   ⚠ 杭長さ抽出失敗、メートル単価をそのまま使用: code={code}, spec={p.get('spec', '')}, unit_price={unit_price}")
            else:
                display_unit_price = unit_price
            print(f"   🔍 杭計算: code={code}, unit_price={unit_price}, pricing_unit={pricing_unit}, length_mm={length_mm}, qty={qty}, display_unit_price={display_unit_price}")
            _real_total = display_unit_price * qty if isinstance(display_unit_price, (int, float)) and display_unit_price > 0 and qty > 0 else 0.0
            pile_total_per_base += _real_total
            pile_total_qty += qty

            ws.merge_cells(f'A{row}:B{row}')
            ws.cell(row=row, column=1, value=pile_seq).font = SM_FONT
            ws.cell(row=row, column=1).alignment = CENTER
            ws.cell(row=row, column=3, value=display_name).font = SM_FONT
            ws.cell(row=row, column=3).alignment = CENTER
            ws.cell(row=row, column=4, value=display_material).font = SM_FONT
            ws.cell(row=row, column=4).alignment = CENTER
            _set(ws, row, 5, '')
            try:
                _spec_num = float(spec)
                _spec_display = f'L={int(_spec_num)}mm' if _spec_num == int(_spec_num) else f'L={_spec_num}mm'
            except (ValueError, TypeError):
                _spec_display = str(spec)
            ws.cell(row=row, column=6, value=_spec_display).font = SM_FONT
            ws.cell(row=row, column=6).alignment = CENTER

            ws.cell(row=row, column=7, value=display_unit_price).font = SM_FONT
            ws.cell(row=row, column=7).alignment = CENTER
            ws.cell(row=row, column=7).number_format = NUM_FMT

            ws.cell(row=row, column=8, value=qty).font = SM_FONT
            ws.cell(row=row, column=8).alignment = CENTER

            ws.cell(row=row, column=9, value=f'=G{row}*H{row}').font = SM_FONT
            ws.cell(row=row, column=9).alignment = CENTER
            ws.cell(row=row, column=9).number_format = NUM_FMT

            if _add_code:
                ws.cell(row=row, column=_code_col, value=code).font = SM_FONT
                ws.cell(row=row, column=_code_col).alignment = CENTER
                ws.cell(row=row, column=_code_col).border = THIN_BORDER
            if _add_weight:
                _pile_weight = '/'
                if price_info:
                    _pw = price_info.get('db_weight')
                    _pu = price_info.get('unit', '')
                    if _pw:
                        if _pu in ['米', 'm', 'M', 'meter', 'Meter']:
                            _pl = extract_length_from_spec(p.get('spec', ''))
                            _pile_weight = round(float(_pw) / 1000 * _pl, 4) if _pl and _pl > 0 else float(_pw)
                        else:
                            _pile_weight = float(_pw)
                ws.cell(row=row, column=_weight_col, value=_pile_weight).font = SM_FONT
                ws.cell(row=row, column=_weight_col).alignment = CENTER
                ws.cell(row=row, column=_weight_col).border = THIN_BORDER

            _pile_border_end = _col_end
            for c in range(1, _pile_border_end + 1):
                ws.cell(row=row, column=c).border = THIN_BORDER

            if not pile_is_matched:
                for c in range(1, _pile_border_end + 1):
                    ws.cell(row=row, column=c).fill = YELLOW_FILL

            img_inserted = False
            if price_info and price_info.get('image_bytes'):
                try:
                    import os as _os
                    _img_bytes = price_info['image_bytes']
                    _img_ext = price_info.get('image_ext', '.png') or '.png'
                    _db_img_name = f"pile_{code}{_img_ext}"
                    _db_img_path = _os.path.join(_pile_tmp, _db_img_name)
                    if _db_img_path not in _pile_img_cache:
                        with open(_db_img_path, 'wb') as _f:
                            _f.write(_img_bytes)
                        _pile_img_cache.pop(_db_img_path, None)
                    _fit_w, _fit_h = fit_image(ws, row, img_col, img_w, img_h)
                    _prep = prepare_image_for_excel(
                        _db_img_path,
                        target_width=_fit_w,
                        target_height=_fit_h,
                        temp_dir=_pile_tmp,
                        cache=_pile_img_cache,
                    )
                    _final = _prep if _prep else _db_img_path
                    img_inserted = add_image_centered_in_cell(
                        ws, _final, row, img_col,
                        img_width=_fit_w, img_height=_fit_h,
                    )
                except Exception:
                    pass
            if not img_inserted:
                ws.cell(row=row, column=5, value='/').alignment = CENTER

            pile_seq += 1
            rendered_pile_count += 1

        pile_data_end_row = pile_data_start_row + rendered_pile_count - 1 if rendered_pile_count > 0 else 0

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(top=0.75, bottom=0.75, left=0.25, right=0.25, header=0.3, footer=0.3)

    discount_total_row = None

    return {
        'sheet_name': sheet_name,
        'discount_total_row': discount_total_row,
        'sub_total_row': sub_row,
        'pile_data_start_row': pile_data_start_row,
        'pile_data_end_row': pile_data_end_row,
        'pile_total_per_base': pile_total_per_base,
        'pile_total_qty': pile_total_qty,
        'array_info': array_info,
        'angle': angle,
        'category_rows': _cat_row_remapped,
    }


def _create_fence_detail_sheet(workbook, fence_rows, gate_rows, nv_fgg,
                               matrix_data=None, nv_params=None,
                               image_temp_dir=None, image_cache=None):
    import tempfile
    from backend.image.processor import decode_image_base64
    from backend.core.shared.image_utils import (
        prepare_image_for_excel, add_image_centered_in_cell,
    )

    matrix_data = matrix_data or {}
    nv_params = nv_params or {}
    if image_temp_dir is None:
        image_temp_dir = tempfile.gettempdir()
    if image_cache is None:
        image_cache = {}

    img_w, img_h, img_col, img_padding = 65, 50, 5, 2

    def _col_width_to_px(cw):
        return int(cw * 7.5 + 5)

    def _row_height_to_px(rh):
        return int(rh * 1.33)

    def _fit_image(ws, r, c, mw, mh, pad=img_padding):
        cw_val = ws.column_dimensions[get_column_letter(c)].width or 8.43
        rh_val = ws.row_dimensions[r].height or 15
        aw = _col_width_to_px(cw_val) - pad * 2
        ah = _row_height_to_px(rh_val) - pad * 2
        aw = max(aw, 20)
        ah = max(ah, 20)
        scale = min(aw / mw, ah / mh, 1.0)
        return int(mw * scale), int(mh * scale)

    coating = (nv_fgg or {}).get('coating', '浸塑')
    fence_surface = ((nv_fgg or {}).get('fence') or {}).get('surface', '白色')
    fence_material_map = {}

    def _resolve_fence_material(code, color_prefix):
        normalized_code = str(code or '').strip()
        if not normalized_code:
            return None

        candidate_codes = []

        def _add_candidate(value):
            candidate = str(value or '').strip()
            if candidate and candidate not in candidate_codes:
                candidate_codes.append(candidate)

        _add_candidate(normalized_code)

        if '-' in normalized_code:
            first_part = normalized_code.split('-', 1)[0]
            if first_part.startswith('FN') and len(first_part) == 4:
                _add_candidate(normalized_code.split('-', 1)[1])

        if '-' in normalized_code:
            base_with_dash = normalized_code if normalized_code.startswith('M') else normalized_code.split('-', 1)[1] if normalized_code.startswith('FN') else ''
            if base_with_dash.startswith('M'):
                for prefix in [color_prefix, 'FN01', 'FN11']:
                    _add_candidate(f'{prefix}-{base_with_dash}')

        try:
            from backend.repositories.fence_gate_material_repository import get_material
        except Exception:
            return None

        for candidate_code in candidate_codes:
            try:
                material = get_material(candidate_code)
                if material:
                    return material
            except Exception:
                continue

        return None
    all_codes = list(set(
        str(r.get('code', '')).strip()
        for r in (fence_rows + gate_rows)
        if str(r.get('code', '')).strip()
    ))
    _codes_not_found = []
    _codes_no_image = []
    if all_codes:
        try:
            surface_prefix_map = {
                '白色浸塑': 'FN01', '咖啡色浸塑': 'FN02', '绿色浸塑': 'FN03',
                '灰褐色浸塑': 'FN04', '深茶色浸塑': 'FN05', '银灰色浸塑': 'FN06',
                '黑色浸塑': 'FN07', '深咖色浸塑': 'FN08', '热镀锌': 'FN11',
                '咖啡浸塑': 'FN02', '绿浸塑': 'FN03',
                '灰褐浸塑': 'FN04', '深茶浸塑': 'FN05', '银灰浸塑': 'FN06',
                '黑浸塑': 'FN07', '深咖浸塑': 'FN08',
            }
            color_prefix = surface_prefix_map.get(fence_surface, 'FN01')
            for code in all_codes:
                mat = _resolve_fence_material(code, color_prefix)
                if mat:
                    fence_material_map[code] = mat
                else:
                    _codes_not_found.append(code)
            for code, mat in fence_material_map.items():
                if not mat.get('image_base64'):
                    _codes_no_image.append(code)
        except Exception:
            pass

    if _codes_not_found:
        print(f"   ⚠ フェンス物料DB未検出コード: {_codes_not_found}")
    if _codes_no_image:
        print(f"   ⚠ フェンス物料画像なしコード: {_codes_no_image}")

    fence_section = (nv_fgg or {}).get('fence') or {}
    gate_section = (nv_fgg or {}).get('gate') or {}
    multi_fences = (nv_fgg or {}).get('fences') or []
    multi_gates = (nv_fgg or {}).get('gates') or []
    fence_style = fence_section.get('style', '')
    height_code = '150'
    height_mm = '1500'

    if multi_fences:
        first_fence = multi_fences[0] if multi_fences else {}
        fence_style = first_fence.get('style', fence_style)
        fence_section = first_fence if not fence_section.get('style') else fence_section

    if fence_style:
        parts = fence_style.split('-')
        if len(parts) >= 2:
            height_code = parts[-1]
        height_mm = _extract_fence_height_mm_from_style(fence_style) or height_mm

    if multi_fences and len(multi_fences) > 1:
        sheet_name = 'フェンス明細'
    else:
        sheet_name = f'フェンス H{height_code}'
    if sheet_name in workbook.sheetnames:
        return {'sheet_name': sheet_name, 'discount_total_row': None}
    ws = workbook.create_sheet(title=sheet_name)

    for col, width in DETAIL_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30

    fence_surface = fence_section.get('surface', (nv_fgg or {}).get('surface', '白色'))
    surface_jp = _surface_to_jp_color(fence_surface)
    if multi_gates:
        for g_sec in multi_gates:
            g_color = g_sec.get('gateColor', '')
            if g_color:
                surface_jp = g_color
                break
    elif gate_section:
        g_color = gate_section.get('gateColor', '')
        if g_color:
            surface_jp = g_color
    fence_len = float(fence_section.get('totalLength', 0) or 0)
    if multi_fences:
        fence_len = sum(float(f.get('totalLength', 0) or 0) for f in multi_fences)
    corner_qty = int(fence_section.get('cornerQty', 0) or 0)
    if multi_fences:
        corner_qty = sum(int(f.get('cornerQty', 0) or 0) for f in multi_fences)

    if multi_gates:
        first_gate = multi_gates[0] if multi_gates else {}
        gate_style_code = first_gate.get('gateStyle', '')
        gate_qty_val = sum(int(g.get('gateQty', 0) or 0) for g in multi_gates)
    else:
        gate_style_code = gate_section.get('gateStyle', '')
        gate_qty_val = int(gate_section.get('gateQty', 0) or 0)

    foundation_type = 'コンクリート基礎'
    gate_type_label = ''
    single_gate_qty = 0
    double_gate_qty = 0
    double_gate_width = 4200

    if gate_style_code and len(gate_style_code) >= 6:
        prefix = gate_style_code[:3]
        width_code = gate_style_code[3:6]

        if prefix.endswith('p'):
            foundation_type = '一体式基礎'
        elif prefix.endswith('g'):
            foundation_type = '杭基礎'

        _NV_GATE_TYPE_LABELS = {
            '片開き門扉', '両開き門扉', '引き戸扉', '折りたたみ扉', '伸縮門扉',
        }

        if prefix.lower().startswith('tl'):
            if gate_qty_val > 0:
                gate_type_label = '引き戸扉'
            double_gate_qty = gate_qty_val
            double_gate_width = int(width_code) * 10 if width_code.isdigit() else 2400
        elif prefix.lower().startswith('tf'):
            if gate_qty_val > 0:
                gate_type_label = '折りたたみ扉'
            double_gate_qty = gate_qty_val
            double_gate_width = int(width_code) * 10 if width_code.isdigit() else 2400
        elif width_code == '120':
            single_gate_qty = gate_qty_val
            if gate_qty_val > 0:
                gate_type_label = '片開き門扉'
        else:
            double_gate_qty = gate_qty_val
            db_h_fd, db_w_fd = _lookup_gate_spec_from_db(gate_style_code)
            if db_w_fd is not None:
                double_gate_width = db_w_fd
            else:
                try:
                    double_gate_width = int(width_code) * 10
                except (ValueError, TypeError):
                    double_gate_width = 4200
            if gate_qty_val > 0:
                gate_type_label = '両開き門扉'

    ws.merge_cells('G1:H1')
    ws['G1'] = '見積日：'
    ws['G1'].font = SM_FONT
    ws['G1'].alignment = RIGHT_A
    ws['I1'] = '=TODAY()'
    ws['I1'].number_format = 'yyyy"年"m"月"dd"日"'
    ws['I1'].font = SM_FONT
    ws['I1'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('A2:I2')
    ws['A2'] = 'フェンス御見積書'
    ws['A2'].font = Font(name='Yu Gothic UI', bold=True, size=16)
    ws['A2'].alignment = CENTER

    if os.path.isfile(_LOGO_PATH):
        try:
            _fimg = XLImage(_LOGO_PATH)
            _fimg.width = 160
            _fimg.height = 53
            ws.add_image(_fimg, 'A1')
        except Exception:
            pass

    ws.delete_rows(3)

    ws.merge_cells('H3:I9')
    _height_m = _format_mm_to_meter_text(height_mm) or '1.5'
    desc_text = (
        f'仕様：\n'
        f'1.ディップコーディング　{surface_jp}\n'
        f'2.{foundation_type}\n'
        f'3.H={_height_m}m　L={int(fence_len)}ｍ\n'
        f'4.片開き門扉幅W1200　{single_gate_qty}ヶ所\n'
        f'5.両開き門扉幅W4200　{double_gate_qty}ヶ所\n'
        f'6.柱材コーナー部　{corner_qty}箇所\n'
        f'（0枚予備）'
    )
    _set(ws, 3, 8, desc_text, font=Font(name='Yu Gothic UI', size=9), align=LEFT_A, border=False)

    bottom_border = Border(bottom=Side(style='thin'))

    project_name = str(matrix_data.get('project_name') or '').strip()

    fence_mitsumori_condition = nv_params.get('mitsumori_condition', 'CIF')
    _fence_mc_label = fence_mitsumori_condition
    if fence_mitsumori_condition == 'CIF_DDP':
        _fence_mc_label = 'CIF+DDP'
    elif fence_mitsumori_condition == 'NV':
        _fence_mc_label = 'DDP（船便）'
    fence_torihiki_condition = nv_params.get('torihiki_condition', 'T/Tで発注時30％、B/Lコピー発行後70％支払')
    if fence_mitsumori_condition == 'NV':
        fence_torihiki_condition = '納品月末締め翌月末払い'
    fence_nounyu_period = '発注後2週後工場から出荷'
    fence_yuko_period = '原材料・海上運賃の変動が激しいため、有効期限24時間'

    def _info_row(row_num, label, value):
        ws.merge_cells(f'A{row_num}:B{row_num}')
        ws[f'A{row_num}'] = label
        ws[f'A{row_num}'].font = SM_FONT
        ws.merge_cells(f'C{row_num}:E{row_num}')
        ws[f'C{row_num}'] = value
        ws[f'C{row_num}'].font = SM_FONT

    _info_row(4, '案件名：', f'{project_name} フェンス H{_height_m}m　L={int(fence_len)}ｍ （{surface_jp}）')
    _info_row(5, '見積条件：', _fence_mc_label)
    _info_row(6, '納入期限：', fence_nounyu_period)
    _info_row(7, '取引条件：', fence_torihiki_condition)
    _info_row(8, '有効期限：', fence_yuko_period)

    r = 10

    subtitle_parts = [p for p in [foundation_type, gate_type_label] if p]
    subtitle_suffix = '　'.join(subtitle_parts)
    ws.merge_cells(f'A{r}:I{r}')
    ws[f'A{r}'] = f'フェンス金額----{subtitle_suffix}' if subtitle_suffix else 'フェンス金額'
    ws[f'A{r}'].font = Font(name='Yu Gothic UI', bold=True, size=10)
    ws[f'A{r}'].alignment = LEFT_A

    r += 1
    header_row = r
    headers = ['部品名称', '材質', '写真', '規格', '単価(EXW US$)', '数量（PCS)', '総金額(EXW US$)']
    ws.merge_cells(f'A{header_row}:B{header_row}')
    cell_num = ws.cell(row=header_row, column=1, value='番号')
    cell_num.font = Font(name='Yu Gothic UI', bold=True, size=10)
    cell_num.alignment = CENTER
    cell_num.border = THIN_BORDER
    cell_num.fill = LIGHT_BLUE_FILL
    ws.cell(row=header_row, column=2).border = THIN_BORDER
    for i, h in enumerate(headers, 3):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = Font(name='Yu Gothic UI', bold=True, size=10)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.fill = LIGHT_BLUE_FILL
    ws.row_dimensions[header_row].height = 25

    all_items = []
    seq = 1
    for row_data in fence_rows:
        all_items.append({**row_data, '_seq': seq})
        seq += 1
    for row_data in gate_rows:
        all_items.append({**row_data, '_seq': seq})
        seq += 1

    data_start = header_row + 1
    for idx, item in enumerate(all_items):
        row = data_start + idx
        ws.row_dimensions[row].height = 42
        for c in range(1, 10):
            ws.cell(row=row, column=c).border = THIN_BORDER

        ws.merge_cells(f'A{row}:B{row}')
        _set(ws, row, 1, item.get('_seq', idx + 1))

        item_code = str(item.get('code', '')).strip()
        mat_info = fence_material_map.get(item_code)
        if mat_info:
            display_name = mat_info.get('日语名称', '') or item.get('name', '')
        else:
            display_name = item.get('name', '')
        _set(ws, row, 3, display_name)

        if mat_info:
            if coating == '热镀锌':
                display_material = mat_info.get('材質表面処理_热镀锌', '')
            else:
                display_material = mat_info.get('材質表面処理_浸塑', '')
        else:
            display_material = ''
        _set(ws, row, 4, display_material)

        img_inserted = False
        if mat_info and mat_info.get('image_base64'):
            img_bytes, img_ext, img_status = decode_image_base64(mat_info['image_base64'])
            if img_status == 'ready' and img_bytes:
                db_img_name = f"fence_{item_code}.{img_ext}"
                db_img_path = os.path.join(image_temp_dir, db_img_name)
                if db_img_path not in image_cache:
                    with open(db_img_path, 'wb') as _f:
                        _f.write(img_bytes)
                    image_cache.pop(db_img_path, None)
                fit_w, fit_h = _fit_image(ws, row, img_col, img_w, img_h)
                temp_img_path = prepare_image_for_excel(
                    db_img_path,
                    target_width=fit_w,
                    target_height=fit_h,
                    temp_dir=image_temp_dir,
                    cache=image_cache,
                )
                final_img_path = temp_img_path if temp_img_path else db_img_path
                success = add_image_centered_in_cell(
                    ws,
                    final_img_path,
                    row,
                    img_col,
                    img_width=fit_w,
                    img_height=fit_h,
                )
                if success:
                    img_inserted = True
        if not img_inserted:
            _set(ws, row, 5, '/')

        _set(ws, row, 6, item.get('spec', ''))

        unit_price = item.get('unit_price', 0)
        qty = item.get('qty', 0)

        ws.cell(row=row, column=7, value=unit_price).font = SM_FONT
        ws.cell(row=row, column=7).alignment = CENTER
        ws.cell(row=row, column=7).number_format = NUM_FMT

        _set(ws, row, 8, qty)

        ws.cell(row=row, column=9, value=f'=G{row}*H{row}').border = THIN_BORDER
        ws.cell(row=row, column=9).alignment = CENTER
        ws.cell(row=row, column=9).number_format = NUM_FMT
        ws.cell(row=row, column=9).font = SM_FONT

    data_end = data_start + len(all_items) - 1
    _img_ok = sum(1 for item in all_items
                  if fence_material_map.get(str(item.get('code', '')).strip())
                  and fence_material_map[str(item.get('code', '')).strip()].get('image_base64'))
    _img_ng = len(all_items) - _img_ok
    print(f"   🖼️ フェンス明細: 画像あり {_img_ok}件, 画像なし {_img_ng}件")
    discount_total_row = None
    if all_items:
        sum_row = data_end + 1
        ws.merge_cells(f'A{sum_row}:H{sum_row}')
        _set(ws, sum_row, 1, 'SUB-TOTAL-（EXW）フェンス合計', font=SM_FONT_BOLD, align=Alignment(horizontal='right', vertical='center'))
        for c in range(1, 9):
            ws.cell(row=sum_row, column=c).border = THIN_BORDER
        ws.cell(row=sum_row, column=9, value=f'=SUM(I{data_start}:I{data_end})').border = THIN_BORDER
        ws.cell(row=sum_row, column=9).font = SM_FONT_BOLD
        ws.cell(row=sum_row, column=9).alignment = CENTER
        ws.cell(row=sum_row, column=9).number_format = NUM_FMT
        ws.row_dimensions[sum_row].height = 25

        try:
            fence_discount_rate = Decimal(str(_get_nv_rate_value(nv_params, 'fence_discount_rate', 'nv_fence_discount_rate', 94)))
        except Exception:
            fence_discount_rate = Decimal('94')

        total_row = sum_row + 1
        ws.merge_cells(f'A{total_row}:H{total_row}')
        _set(ws, total_row, 1, '合計金額(USD)', font=SM_FONT_BOLD, align=Alignment(horizontal='right', vertical='center'))
        for c in range(1, 9):
            ws.cell(row=total_row, column=c).border = THIN_BORDER
        ws.cell(row=total_row, column=9, value=f'=I{sum_row}').border = THIN_BORDER
        ws.cell(row=total_row, column=9).font = SM_FONT_BOLD
        ws.cell(row=total_row, column=9).alignment = CENTER
        ws.cell(row=total_row, column=9).number_format = NUM_FMT
        ws.row_dimensions[total_row].height = 25

        discount_total_row = total_row + 1
        ws.merge_cells(f'A{discount_total_row}:H{discount_total_row}')
        _set(ws, discount_total_row, 1, '特別値引き後合計金額(USD)', font=SM_FONT_BOLD, align=Alignment(horizontal='right', vertical='center'))
        for c in range(1, 9):
            ws.cell(row=discount_total_row, column=c).border = THIN_BORDER
        ws.cell(row=discount_total_row, column=9, value=f'=I{total_row}*{fence_discount_rate}/100').border = THIN_BORDER
        ws.cell(row=discount_total_row, column=9).font = SM_FONT_BOLD
        ws.cell(row=discount_total_row, column=9).alignment = CENTER
        ws.cell(row=discount_total_row, column=9).number_format = NUM_FMT
        ws.row_dimensions[discount_total_row].height = 25

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(top=0.75, bottom=0.75, left=0.25, right=0.25, header=0.3, footer=0.3)

    return {'sheet_name': sheet_name, 'discount_total_row': discount_total_row}


BLUE_FILL = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='F3D4A9', end_color='F3D4A9', fill_type='solid')
GREEN_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
THICK_SIDE = Side(style='medium', color='000000')



def _get_cjk_num(n):
    _cjk = ['一', '二', '三', '四', '五', '六', '七', '八', '九', '十']
    return _cjk[n - 1] if 1 <= n <= len(_cjk) else str(n)


def _surface_to_jp_color(surface):
    if not surface:
        return '白色'
    _color_map = [
        ('深茶色', '深茶色'), ('深茶', '深茶色'),
        ('深咖色', 'ダークブラウン'), ('深咖', 'ダークブラウン'),
        ('咖啡色', 'コーヒー色'), ('咖啡', 'コーヒー色'),
        ('灰褐色', '灰褐色'), ('灰褐', '灰褐色'),
        ('银灰色', '銀灰色'), ('银灰', '銀灰色'),
        ('绿色', '緑色'), ('绿', '緑色'),
        ('黑色', '黒色'), ('黑', '黒色'),
        ('白色', '白色'),
        ('热镀锌', 'ガルバ'),
        ('茶', '茶色'),
    ]
    for cn_key, jp_val in _color_map:
        if cn_key in surface:
            return jp_val
    return str(surface)


def create_nv_summary_sheet(workbook, detail_results, matrix_data=None,
                             image_path=None, fence_data=None,
                             nv_params=None, nv_fence_gate_data=None,
                             pile_summary=None,
                             image_temp_dir=None, image_cache=None,
                             sheet_title=None,
                             spare_parts_count=0,
                             dz_spare_count=0,
                             inverter_detail=None):
    from datetime import datetime

    nv_params = nv_params or {}
    matrix_data = matrix_data or {}

    discount_rate_pct = nv_params.get('discount_rate') or nv_params.get('nv_discount_rate', 72)
    fence_discount_rate = _get_nv_rate_value(nv_params, 'fence_discount_rate', 'nv_fence_discount_rate', 94)
    steel_discount_rate = nv_params.get('steel_discount_rate') or nv_params.get('nv_steel_discount_rate', 84)
    purchased_discount_rate = nv_params.get('purchased_discount_rate') or nv_params.get('nv_purchased_discount_rate', 94)
    consumption_tax_pct = nv_params.get('consumption_tax', 10)
    tariff_rate_pct = nv_params.get('tariff_rate', 3)
    mitsumori_condition = nv_params.get('mitsumori_condition', 'CIF')
    _show_cif = mitsumori_condition in ('CIF', 'CIF_DDP', 'NV')
    _show_ddp = mitsumori_condition in ('DDP', 'CIF_DDP', 'NV')
    is_cif = _show_cif
    shipping_fee = nv_params.get('ddp_shipping_total') or nv_params.get('shipping_fee', 0)
    dest_port = nv_params.get('dest_port', '横浜')
    containers = nv_params.get('containers', [])
    if not containers and is_cif:
        container_type_str = nv_params.get('container_type', '混載便')
        if container_type_str and '*' in container_type_str:
            parts = container_type_str.split('+')
            containers = []
            for p in parts:
                if '*' in p:
                    ct, cq = p.split('*', 1)
                    containers.append({'type': ct, 'qty': int(cq), 'freight': 0})
        else:
            containers = [{'type': container_type_str or '混載便', 'qty': nv_params.get('container_qty', 1), 'freight': 0}]
    ocean_freight = nv_params.get('ocean_freight', 0)
    if ocean_freight in (None, ''):
        ocean_freight = 0
    exchange_rate = nv_params.get('exchange_rate', 151)
    nv_truck_fee = nv_params.get('truck_fee')
    if nv_truck_fee in (None, ''):
        nv_truck_fee = 0
    truck_fee = nv_truck_fee if _show_ddp else 0
    truck_desc_val = nv_params.get('truck_desc', '4Tユニック+平車')
    truck_desc_val = re.sub(r'4[TＴ]\+?ユニック', '4Tユニック', str(truck_desc_val))
    truck_desc = truck_desc_val if _show_ddp else ''
    sales_name = nv_params.get('sales_name', 'Nanami')
    sales_phone = nv_params.get('sales_phone', '+86-137-7466-5835')
    customer_name = str(nv_params.get('customer_name', '') or '').strip()
    nounyu_period = '発注後2週後工場から出荷'
    torihiki_condition = nv_params.get('torihiki_condition', 'T/Tで発注時30％、B/Lコピー発行後70％支払')
    if mitsumori_condition == 'NV':
        torihiki_condition = '納品月末締め翌月末払い'
    yuko_period = '原材料・海上運賃の変動が激しいため、有効期限24時間'
    final_price_usd = nv_params.get('final_price', 0)

    project_name = str(matrix_data.get('project_name') or '').strip()
    output_kw = matrix_data.get('output_kw') or 0
    module_wattage = matrix_data.get('module_wattage') or 0

    ws = workbook.create_sheet(title=sheet_title or '合計')

    NV_SUMMARY_WIDTHS = {
        'A': 11, 'B': 11, 'C': 11, 'D': 8, 'E': 13, 'F': 8,
        'G': 8, 'H': 8, 'I': 8, 'J': 13, 'K': 15, 'L': 8, 'M': 8,
    }
    for col, width in NV_SUMMARY_WIDTHS.items():
        ws.column_dimensions[col].width = width

    center = CENTER
    right_a = RIGHT_A
    left_a = LEFT_A
    bottom_only = Border(bottom=Side(style='thin'))
    GRAY_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    def _s(r, c, val, font=FONT_10, align=center, border=True):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = font
        cell.alignment = align
        if border:
            cell.border = THIN_BORDER
        return cell

    # ═══ HEADER ═══

    for _r in range(1, 18):
        ws.row_dimensions[_r].height = 15

    ws.cell(row=1, column=11, value='見積日：').font = FONT_10
    ws.cell(row=1, column=11).alignment = right_a
    ws.cell(row=1, column=12, value='=TODAY()').font = FONT_10
    ws.cell(row=1, column=12).alignment = left_a
    ws.cell(row=1, column=12).number_format = 'yyyy"年"m"月"dd"日"'
    ws.merge_cells('L1:M1')

    ws.cell(row=2, column=1, value='架台御見積書').font = Font(name='Yu Gothic UI', size=21, bold=True)
    ws.cell(row=2, column=1).alignment = center
    ws.merge_cells('A2:M3')

    _insert_logo(ws, 'J9', width=200, height=60)

    ws.merge_cells('A5:E5')
    ws.cell(row=5, column=1, value=customer_name).font = Font(name='Yu Gothic UI', size=12, bold=True)
    ws.cell(row=5, column=1).alignment = center
    for c in range(1, 6):
        ws.cell(row=5, column=c).border = Border(bottom=Side(style='thin'))
    ws.cell(row=5, column=6, value='御中').font = Font(name='Yu Gothic UI', size=10, bold=True)
    ws.cell(row=5, column=6).alignment = left_a
    ws.merge_cells('A6:H6')

    ws.merge_cells('A7:H7')
    ws.cell(row=7, column=1, value='拝啓 貴社益々ご清栄の事とお喜び申し上げます。').font = FONT_10
    ws.cell(row=7, column=1).alignment = left_a
    ws.merge_cells('A8:H8')
    ws.cell(row=8, column=1, value='貴社よりご依頼頂いた案件情報における架台諸費用を下記の通り御見積申し上げます。').font = FONT_10
    ws.cell(row=8, column=1).alignment = left_a
    ws.merge_cells('A9:H9')

    def _info_row(r, label, value, k_value=None):
        ws.cell(row=r, column=1, value=label).font = FONT_10
        ws.cell(row=r, column=1).alignment = left_a
        ws.merge_cells(f'B{r}:F{r}')
        ws.cell(row=r, column=2, value=value).font = FONT_10
        ws.cell(row=r, column=2).alignment = left_a
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = bottom_only
        if k_value is not None:
            ws.merge_cells(f'J{r}:M{r}')
            ws.cell(row=r, column=10, value=k_value).font = FONT_10
            ws.cell(row=r, column=10).alignment = left_a

    _info_row(10, '案件名：', f'{project_name}　{output_kw}KW')
    _mc_label = mitsumori_condition
    if mitsumori_condition == 'CIF_DDP':
        _mc_label = 'CIF+DDP'
    elif mitsumori_condition == 'NV':
        _mc_label = 'DDP（船便）'
    _info_row(11, '見積条件：', _mc_label)
    _info_row(12, '納入期限：', nounyu_period)
    _info_row(13, '取引条件：', torihiki_condition, k_value='厦門カセング金属科技有限公司')
    yuko_font = Font(name='Yu Gothic UI', size=10, color='0000FF')
    _info_row(14, '有効期限：', yuko_period, k_value='〒361-009 中国厦門湖里区枋湖北二路891号')
    ws.cell(row=14, column=2).font = yuko_font

    ws.cell(row=13, column=10).font = Font(name='Yu Gothic UI', size=11, bold=True)
    ws.merge_cells('J15:M15')
    ws.cell(row=15, column=10, value='匯鑫財富大厦11-12階').font = FONT_10
    ws.cell(row=15, column=10).alignment = center

    ws.merge_cells('J16:M16')
    ws.cell(row=16, column=10, value=f'ご担当者：{sales_name}　{sales_phone}').font = FONT_10
    ws.cell(row=16, column=10).alignment = left_a

    _FONT_12 = Font(name='Yu Gothic UI', size=12.5)
    _FONT_12_BOLD = Font(name='Yu Gothic UI', size=12.5, bold=True)
    _bottom_left = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
    _bottom_center = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

    r = 16
    ws.row_dimensions[r].height = 20
    ws.merge_cells(f'A{r}:B{r}')
    ws.cell(row=r, column=1, value='御見積合計金額：').font = _FONT_12_BOLD
    ws.cell(row=r, column=1).alignment = _bottom_left
    ws.merge_cells(f'C{r}:D{r}')
    total_amount_cell_r = r
    ws.merge_cells(f'E{r}:F{r}')
    ws.cell(row=r, column=5, value='（送料、税込）').font = _FONT_12_BOLD
    ws.cell(row=r, column=5).alignment = _bottom_left
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = bottom_only

    r = 17
    ws.row_dimensions[r].height = 20
    ws.merge_cells(f'A{r}:B{r}')
    ws.cell(row=r, column=1, value='部材KW当たり：').font = _FONT_12_BOLD
    ws.cell(row=r, column=1).alignment = _bottom_left
    ws.merge_cells(f'C{r}:D{r}')
    kw_price_cell_r = r
    ws.merge_cells(f'E{r}:F{r}')
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = bottom_only

    # ═══ 一、架台本体金額 ═══

    r = 19
    ws.merge_cells(f'A{r}:M{r}')
    ws.cell(row=r, column=1, value='一、架台本体金額').font = SM_FONT_BOLD_10
    ws.cell(row=r, column=1).alignment = left_a
    ws.row_dimensions[r].height = 22

    r = 20
    hdr1 = ['NO.', 'パネル枚数', '', '', 'その他', '設置角度', '', 'アレイ数', '発電量（kw）', '', '1アレイ単価(USD)', '総金額(USD)', '']
    for i, h in enumerate(hdr1, 1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.font = FONT_10_BOLD
        cell.alignment = center
        cell.border = THIN_BORDER
        if h == '':
            cell.fill = GRAY_FILL
    ws.merge_cells(f'B{r}:D{r}')
    ws.merge_cells(f'F{r}:G{r}')
    ws.merge_cells(f'I{r}:J{r}')
    ws.merge_cells(f'L{r}:M{r}')
    ws.row_dimensions[r].height = 22

    data_row_start = 21
    _accum_seen = set()
    _visible_row = 0

    for i, detail in enumerate(detail_results):
        _acc_gid = detail.get('accumulated_group_id')
        if _acc_gid and _acc_gid in _accum_seen:
            continue
        if _acc_gid:
            _accum_seen.add(_acc_gid)

        row = data_row_start + _visible_row
        _visible_row += 1
        ws.row_dimensions[row].height = 20

        if _acc_gid:
            _group = [d for d in detail_results if d.get('accumulated_group_id') == _acc_gid]
            _group.sort(key=lambda d: d.get('accumulated_sub_idx', 0))
            _total_base = sum(d.get('array_info', {}).get('table_qty', 1) for d in _group)
            _first = _group[0]
            arr = _first.get('array_info', {})
            rows_v = arr.get('rows', 0)
            cols_v = arr.get('cols', 0)
            detail_config = _first.get('config') or {}
            info_missing = arr.get('missing_per_table', 0) or 0
            missing_boards = info_missing if info_missing else 0
            inverter_count = 0
            inv_note = ''
            for _gd in _group:
                _gd_inv = _gd.get('inverter_count', 0)
                if _gd_inv and int(_gd_inv) > 0:
                    inverter_count = _gd_inv
                _gd_note = _gd.get('inv_note', '') or (_gd.get('array_info', {}).get('note') or '')
                if _gd_note.strip():
                    inv_note = _gd_note
            angle_val = _first.get('angle', '')
            if isinstance(angle_val, float) and angle_val != angle_val:
                angle_val = ''
            if str(angle_val).lower() == 'nan' or str(angle_val).lower().strip('\u00b0') == 'nan':
                angle_val = ''
            angle_clean = str(angle_val).rstrip('\u00b0').strip() if angle_val else ''
            angle_text = f'{angle_clean}\u00b0' if angle_clean else ''
            note = arr.get('note') or ''

            _s(row, 1, _visible_row)
            _s(row, 2, rows_v)
            ws.cell(row=row, column=2).number_format = '0"段"'
            _s(row, 3, cols_v)
            ws.cell(row=row, column=3).number_format = '0"列"'
            _s(row, 4, missing_boards if missing_boards else '/')
            ws.cell(row=row, column=4).alignment = center
            _s(row, 5, inv_note or (f'{int(inverter_count)}台ＰＣＳ' if inverter_count and int(inverter_count) > 0 else ''), align=left_a)
            angle_display = f'横置き {angle_text}' if angle_clean else ''
            _s(row, 6, angle_display)
            _s(row, 7, '')
            ws.cell(row=row, column=7).fill = GRAY_FILL
            _s(row, 8, _total_base)
            ws.cell(row=row, column=8).number_format = '0"基"'
            mb_formula = missing_boards if missing_boards else 0
            ws.cell(row=row, column=9, value=f'=(B{row}*C{row}+{mb_formula})*{_total_base}*{module_wattage}/1000').font = SM_FONT
            ws.cell(row=row, column=9).alignment = center
            ws.cell(row=row, column=9).number_format = '#,##0.00'
            ws.cell(row=row, column=9).border = THIN_BORDER
            _s(row, 10, '')
            ws.cell(row=row, column=10).fill = GRAY_FILL
            _k_parts = []
            for _gd in _group:
                _sn = _gd.get('sheet_name', '')
                _sr = _gd.get('sub_total_row', 24)
                _gd_base = _gd.get('array_info', {}).get('table_qty', 1)
                _k_parts.append(f"'{_sn}'!I{_sr}*{_gd_base}")
            _k_formula = f'=({"+".join(_k_parts)})/{_total_base}'
            ws.cell(row=row, column=11, value=_k_formula).font = SM_FONT
            ws.cell(row=row, column=11).number_format = NUM_FMT
            ws.cell(row=row, column=11).alignment = center
            ws.cell(row=row, column=11).border = THIN_BORDER
            ws.cell(row=row, column=12, value=f'=K{row}*H{row}').font = SM_FONT
            ws.cell(row=row, column=12).alignment = center
            ws.cell(row=row, column=12).border = THIN_BORDER
            ws.cell(row=row, column=12).number_format = NUM_FMT
            _s(row, 13, '')
            ws.merge_cells(f'F{row}:G{row}')
            ws.merge_cells(f'I{row}:J{row}')
            ws.merge_cells(f'L{row}:M{row}')
        else:
            arr = detail.get('array_info', {})
            base_count = arr.get('table_qty', 1)
            rows_v = arr.get('rows', 0)
            cols_v = arr.get('cols', 0)
            detail_config = detail.get('config') or {}
            info_missing = arr.get('missing_per_table', 0) or 0
            bom_missing = detail_config.get('missing_boards', 0) or 0
            missing_boards = info_missing if info_missing else (bom_missing // base_count if base_count and bom_missing else bom_missing)
            note = arr.get('note') or ''
            inverter_count = detail.get('inverter_count', 0)

            angle_val = detail.get('angle', '')
            if isinstance(angle_val, float) and angle_val != angle_val:
                angle_val = ''
            if str(angle_val).lower() == 'nan' or str(angle_val).lower().strip('\u00b0') == 'nan':
                angle_val = ''
            angle_clean = str(angle_val).rstrip('\u00b0').strip() if angle_val else ''
            angle_text = f'{angle_clean}\u00b0' if angle_clean else ''

            _s(row, 1, _visible_row)
            _s(row, 2, rows_v)
            ws.cell(row=row, column=2).number_format = '0"段"'
            _s(row, 3, cols_v)
            ws.cell(row=row, column=3).number_format = '0"列"'
            _s(row, 4, missing_boards if missing_boards else '/')
            ws.cell(row=row, column=4).alignment = center
            inv_note = detail.get('inv_note', '') or note
            if inv_note:
                _s(row, 5, inv_note, align=left_a)
            else:
                if inverter_count and int(inverter_count) > 0:
                    _s(row, 5, f'{int(inverter_count)}台ＰＣＳ', align=left_a)
                else:
                    _s(row, 5, '', align=left_a)
            angle_display = f'横置き {angle_text}' if angle_clean else ''
            _s(row, 6, angle_display)
            _s(row, 7, '')
            ws.cell(row=row, column=7).fill = GRAY_FILL
            _s(row, 8, base_count)
            ws.cell(row=row, column=8).number_format = '0"基"'
            mb_formula = missing_boards if missing_boards else 0
            ws.cell(row=row, column=9, value=f'=(B{row}*C{row}+{mb_formula})*{base_count}*{module_wattage}/1000').font = SM_FONT
            ws.cell(row=row, column=9).alignment = center
            ws.cell(row=row, column=9).number_format = '#,##0.00'
            ws.cell(row=row, column=9).border = THIN_BORDER
            _s(row, 10, '')
            ws.cell(row=row, column=10).fill = GRAY_FILL
            detail_sheet_name = detail.get('sheet_name', '')
            sub_total_row = detail.get('sub_total_row', 24)
            ws.cell(row=row, column=11, value=f"='{detail_sheet_name}'!I{sub_total_row}").font = SM_FONT
            ws.cell(row=row, column=11).number_format = NUM_FMT
            ws.cell(row=row, column=11).alignment = center
            ws.cell(row=row, column=11).border = THIN_BORDER
            ws.cell(row=row, column=12, value=f'=K{row}*H{row}').font = SM_FONT
            ws.cell(row=row, column=12).alignment = center
            ws.cell(row=row, column=12).border = THIN_BORDER
            ws.cell(row=row, column=12).number_format = NUM_FMT
            _s(row, 13, '')
            ws.merge_cells(f'F{row}:G{row}')
            ws.merge_cells(f'I{row}:J{row}')
            ws.merge_cells(f'L{row}:M{row}')

    data_row_end = data_row_start + _visible_row - 1

    _inv_insert_end = data_row_end
    if inverter_detail:
        inv_details = inverter_detail if isinstance(inverter_detail, list) else [inverter_detail]
        for inv_idx, inv_item in enumerate(inv_details):
            row = data_row_end + 1 + inv_idx
            ws.row_dimensions[row].height = 20
            inv_count = inv_item.get('inverter_count', 0)
            inv_sheet = inv_item.get('sheet_name', '')
            inv_sub_row = inv_item.get('sub_total_row', 8)
            inv_arr = inv_item.get('array_info', {})
            inv_qty = inv_arr.get('table_qty', 1)

            _s(row, 1, _visible_row + 1 + inv_idx)
            _s(row, 2, 'パワコン独立架台')
            _s(row, 3, '/')
            _s(row, 4, '/')
            _s(row, 5, '/')
            _s(row, 6, '/')
            _s(row, 7, '')
            _s(row, 8, inv_qty)
            ws.cell(row=row, column=8).number_format = '0"基"'
            _s(row, 9, '/')
            _s(row, 10, '')
            ws.cell(row=row, column=11, value=f"='{inv_sheet}'!I{inv_sub_row}").font = SM_FONT
            ws.cell(row=row, column=11).number_format = NUM_FMT
            ws.cell(row=row, column=11).alignment = center
            ws.cell(row=row, column=11).border = THIN_BORDER
            ws.cell(row=row, column=12, value=f'=K{row}*H{row}').font = SM_FONT
            ws.cell(row=row, column=12).alignment = center
            ws.cell(row=row, column=12).border = THIN_BORDER
            ws.cell(row=row, column=12).number_format = NUM_FMT
            _s(row, 13, '')
            ws.merge_cells(f'B{row}:D{row}')
            ws.merge_cells(f'F{row}:G{row}')
            ws.merge_cells(f'I{row}:J{row}')
            ws.merge_cells(f'L{row}:M{row}')
            _inv_insert_end = row
        data_row_end = _inv_insert_end

    post_spare_count = int(nv_params.get('post_spare_count', 0) or 0)
    post_spare_price = float(nv_params.get('post_spare_price', 0) or 0)
    pile_spare_count = int(nv_params.get('pile_spare_count', 0) or 0)
    pile_spare_price = float(nv_params.get('pile_spare_price', 0) or 0)

    _spare_fill = PatternFill(start_color='FFDDB3', end_color='FFDDB3', fill_type='solid')

    r_post_spare = None
    if post_spare_count > 0 and post_spare_price > 0:
        r_post_spare = data_row_end + 1
        for c in range(1, 14):
            ws.cell(row=r_post_spare, column=c).border = THIN_BORDER
        ws.cell(row=r_post_spare, column=1, value='柱材予備品').font = SM_FONT
        ws.cell(row=r_post_spare, column=1).alignment = center
        ws.cell(row=r_post_spare, column=9, value=post_spare_count).font = SM_FONT
        ws.cell(row=r_post_spare, column=9).alignment = center
        ws.cell(row=r_post_spare, column=9).number_format = '#,##0"本"'
        ws.cell(row=r_post_spare, column=10).fill = GRAY_FILL
        _s(r_post_spare, 11, post_spare_price)
        ws.cell(row=r_post_spare, column=11).number_format = NUM_FMT
        ws.cell(row=r_post_spare, column=11).alignment = center
        ws.cell(row=r_post_spare, column=12, value=f'=K{r_post_spare}*I{r_post_spare}').font = SM_FONT
        ws.cell(row=r_post_spare, column=12).alignment = center
        ws.cell(row=r_post_spare, column=12).number_format = NUM_FMT
        ws.cell(row=r_post_spare, column=13).fill = GRAY_FILL
        ws.merge_cells(f'A{r_post_spare}:H{r_post_spare}')
        ws.merge_cells(f'I{r_post_spare}:J{r_post_spare}')
        ws.merge_cells(f'L{r_post_spare}:M{r_post_spare}')
        ws.row_dimensions[r_post_spare].height = 20
        _bracket_sum_end = r_post_spare
    else:
        _bracket_sum_end = data_row_end

    r_spare_bracket = None
    if spare_parts_count and spare_parts_count > 0:
        r_spare_bracket = _bracket_sum_end + 1
        for c in range(1, 14):
            ws.cell(row=r_spare_bracket, column=c).border = THIN_BORDER
        ws.cell(row=r_spare_bracket, column=1, value='無料予備品').font = SM_FONT
        ws.cell(row=r_spare_bracket, column=1).alignment = center
        for c in range(1, 12):
            ws.cell(row=r_spare_bracket, column=c).fill = _spare_fill
        ws.cell(row=r_spare_bracket, column=12, value=0).font = SM_FONT
        ws.cell(row=r_spare_bracket, column=12).alignment = center
        ws.cell(row=r_spare_bracket, column=12).number_format = NUM_FMT
        ws.merge_cells(f'A{r_spare_bracket}:K{r_spare_bracket}')
        ws.merge_cells(f'L{r_spare_bracket}:M{r_spare_bracket}')
        ws.row_dimensions[r_spare_bracket].height = 20

    r_btotal = (r_spare_bracket if r_spare_bracket else _bracket_sum_end) + 1
    for c in range(1, 14):
        ws.cell(row=r_btotal, column=c).border = THIN_BORDER
    ws.cell(row=r_btotal, column=1, value='合計').font = SM_FONT_BOLD
    ws.cell(row=r_btotal, column=1).alignment = center
    ws.cell(row=r_btotal, column=8, value=f'=SUM(H{data_row_start}:H{_bracket_sum_end})').font = SM_FONT_BOLD
    ws.cell(row=r_btotal, column=8).alignment = center
    ws.cell(row=r_btotal, column=8).number_format = '0"基"'
    ws.cell(row=r_btotal, column=9, value=f'=SUM(I{data_row_start}:I{_bracket_sum_end})').font = SM_FONT_BOLD
    ws.cell(row=r_btotal, column=9).alignment = center
    ws.cell(row=r_btotal, column=9).number_format = '#,##0.00'
    ws.cell(row=r_btotal, column=11, value='-').font = SM_FONT_BOLD
    ws.cell(row=r_btotal, column=11).alignment = center
    ws.cell(row=r_btotal, column=12, value=f'=SUM(L{data_row_start}:L{_bracket_sum_end})').font = SM_FONT_BOLD
    ws.cell(row=r_btotal, column=12).alignment = center
    ws.cell(row=r_btotal, column=12).number_format = NUM_FMT
    ws.merge_cells(f'A{r_btotal}:G{r_btotal}')
    ws.merge_cells(f'I{r_btotal}:J{r_btotal}')
    ws.merge_cells(f'L{r_btotal}:M{r_btotal}')
    ws.row_dimensions[r_btotal].height = 20

    r_disc = r_btotal + 1
    ws.cell(row=r_disc, column=1, value='特別値引き後金額').font = SM_FONT_BOLD
    ws.cell(row=r_disc, column=1).alignment = right_a
    ws.cell(row=r_disc, column=1).border = THIN_BORDER
    ws.merge_cells(f'A{r_disc}:K{r_disc}')

    _disc_parts = []
    _all_rates = set()
    for _di, _detail in enumerate(detail_results):
        _srow = data_row_start + _di
        _sn = _detail.get('sheet_name', '')
        _cr = _detail.get('category_rows', {})
        _cat_groups = {}
        for _r, _cat in _cr.items():
            _cat_groups.setdefault(_cat, []).append(_r)
        _cat_rates = {'standard': discount_rate_pct, 'steel': steel_discount_rate, 'purchased': purchased_discount_rate}
        for _cat, _rows in _cat_groups.items():
            if not _rows:
                continue
            _rate = _cat_rates.get(_cat, discount_rate_pct)
            _all_rates.add(_rate)
            _refs = '+'.join(f"'{_sn}'!I{_r}" for _r in _rows)
            _disc_parts.append(f"({_refs})*H{_srow}*{_rate}/100")

    if _disc_parts:
        if len(_all_rates) == 1:
            _disc_formula = f'=L{r_btotal}*{_all_rates.pop()}/100'
        else:
            _disc_formula = '=' + '+'.join(_disc_parts)
    else:
        _disc_formula = f'=L{r_btotal}*{discount_rate_pct}/100'
    ws.cell(row=r_disc, column=12, value=_disc_formula).font = SM_FONT_BOLD
    ws.cell(row=r_disc, column=12).alignment = center
    ws.cell(row=r_disc, column=12).border = THIN_BORDER
    ws.cell(row=r_disc, column=12).number_format = NUM_FMT
    ws.merge_cells(f'L{r_disc}:M{r_disc}')
    ws.row_dimensions[r_disc].height = 20

    r_watt = r_disc + 1
    ws.cell(row=r_watt, column=1, value='ワット当たり単価').font = SM_FONT_BOLD
    ws.cell(row=r_watt, column=1).alignment = right_a
    ws.cell(row=r_watt, column=1).border = THIN_BORDER
    ws.merge_cells(f'A{r_watt}:K{r_watt}')
    ws.cell(row=r_watt, column=12, value=f'=L{r_disc}/(I{r_btotal}*1000)').font = SM_FONT_BOLD
    ws.cell(row=r_watt, column=12).alignment = center
    ws.cell(row=r_watt, column=12).border = THIN_BORDER
    ws.cell(row=r_watt, column=12).number_format = '0.000'
    ws.cell(row=r_watt, column=13, value='').font = SM_FONT
    ws.cell(row=r_watt, column=13).border = THIN_BORDER
    ws.cell(row=r_watt, column=13).fill = GRAY_FILL
    ws.merge_cells(f'L{r_watt}:M{r_watt}')
    ws.row_dimensions[r_watt].height = 20

    has_pile_data = any(
        (detail.get('pile_total_qty', 0) or detail.get('pile_total_per_base', 0))
        for detail in detail_results
    )

    if has_pile_data:
        # ═══ 二、杭金額 ═══

        r = r_watt + 1
        ws.merge_cells(f'A{r}:M{r}')
        ws.cell(row=r, column=1, value='二、杭金額').font = SM_FONT_BOLD_10
        ws.cell(row=r, column=1).alignment = left_a
        ws.row_dimensions[r].height = 22

        r += 1
        pile_hdr = r
        hdr_pile = ['NO.', 'パネル枚数', '', '', 'その他', '設置角度', '', 'アレイ数', '杭本数（PCS）', '', '1アレイ単価(USD)', '総金額(USD)', '']
        for i, h in enumerate(hdr_pile, 1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.font = SM_FONT_HEADER
            cell.alignment = center
            cell.border = THIN_BORDER
            if h == '':
                cell.fill = GRAY_FILL
        ws.merge_cells(f'B{r}:D{r}')
        ws.merge_cells(f'F{r}:G{r}')
        ws.merge_cells(f'I{r}:J{r}')
        ws.merge_cells(f'L{r}:M{r}')
        ws.row_dimensions[r].height = 22

        pile_data_start = pile_hdr + 1
        pile_qty_sum = 0
        pile_cost_sum = Decimal('0')
        pile_row_idx = 0
        pile_data_end = pile_data_start - 1
        for detail in detail_results:
            arr = detail.get('array_info', {})
            base_count = arr.get('table_qty', 1)
            rows_v = arr.get('rows', 0)
            cols_v = arr.get('cols', 0)
            pile_qty = detail.get('pile_total_qty', 0)
            pile_cost = detail.get('pile_total_per_base', 0)

            if not pile_qty and not pile_cost:
                continue

            row = pile_data_start + pile_row_idx
            pile_row_idx += 1
            ws.row_dimensions[row].height = 20

            detail_config = detail.get('config') or {}
            info_missing = arr.get('missing_per_table', 0) or 0
            bom_missing = detail_config.get('missing_boards', 0) or 0
            missing_boards = info_missing if info_missing else (bom_missing // base_count if base_count and bom_missing else bom_missing)

            angle_val = detail.get('angle', '')
            if isinstance(angle_val, float) and angle_val != angle_val:
                angle_val = ''
            if str(angle_val).lower() == 'nan' or str(angle_val).lower().strip('\u00b0') == 'nan':
                angle_val = ''
            angle_clean = str(angle_val).rstrip('\u00b0').strip() if angle_val else ''
            angle_text = f'{angle_clean}\u00b0' if angle_clean else ''

            _s(row, 1, pile_row_idx)
            _s(row, 2, rows_v)
            ws.cell(row=row, column=2).number_format = '0"段"'
            _s(row, 3, cols_v)
            ws.cell(row=row, column=3).number_format = '0"列"'
            _s(row, 4, missing_boards if missing_boards else '/')
            ws.cell(row=row, column=4).alignment = center
            pile_inv_note = detail.get('inv_note', '') or (arr.get('note') or '')
            _s(row, 5, pile_inv_note, align=left_a)
            angle_display = f'横置き {angle_text}' if angle_clean else ''
            _s(row, 6, angle_display)
            _s(row, 7, '')
            ws.cell(row=row, column=7).fill = GRAY_FILL
            _s(row, 8, base_count)
            ws.cell(row=row, column=8).number_format = '0"基"'
            pile_total_qty_for_row = int(pile_qty) * base_count if pile_qty else 0
            _s(row, 9, pile_total_qty_for_row)
            ws.cell(row=row, column=9).number_format = '#,##0"本"'
            _s(row, 10, '')
            ws.cell(row=row, column=10).fill = GRAY_FILL
            detail_sheet_name = detail.get('sheet_name', '')
            pile_start = detail.get('pile_data_start_row', 0)
            pile_end = detail.get('pile_data_end_row', 0)
            if detail_sheet_name and pile_start and pile_end and pile_start <= pile_end:
                ws.cell(row=row, column=11, value=f"=SUM('{detail_sheet_name}'!I{pile_start}:I{pile_end})").font = SM_FONT
            else:
                _s(row, 11, float(pile_cost) if pile_cost else 0)
            ws.cell(row=row, column=11).number_format = NUM_FMT
            ws.cell(row=row, column=11).alignment = center
            ws.cell(row=row, column=12, value=f'=K{row}*H{row}').font = SM_FONT
            ws.cell(row=row, column=12).alignment = center
            ws.cell(row=row, column=12).border = THIN_BORDER
            ws.cell(row=row, column=12).number_format = NUM_FMT
            _s(row, 13, '')
            ws.cell(row=row, column=13).fill = GRAY_FILL
            ws.merge_cells(f'F{row}:G{row}')
            ws.merge_cells(f'I{row}:J{row}')
            ws.merge_cells(f'L{row}:M{row}')

            pile_qty_sum += pile_total_qty_for_row
            pile_cost_sum += Decimal(str(pile_cost)) * Decimal(str(base_count))
            pile_data_end = row

        if pile_summary and pile_summary.get('total_qty'):
            ps_qty = int(pile_summary['total_qty']) if pile_summary['total_qty'] == int(pile_summary['total_qty']) else pile_summary['total_qty']
            reserve_qty = max(0, ps_qty - pile_qty_sum)
            if reserve_qty > 0:
                row = pile_data_end + 1
                pile_data_end = row
                for c in range(1, 14):
                    ws.cell(row=row, column=c).border = THIN_BORDER
                ws.cell(row=row, column=1, value='杭予備品').font = SM_FONT
                ws.cell(row=row, column=1).alignment = center
                _s(row, 9, reserve_qty)
                ws.cell(row=row, column=9).number_format = '#,##0"本"'
                reserve_price = max(0, float(pile_summary.get('total_price', 0)) - float(pile_cost_sum))
                reserve_unit = round(reserve_price / reserve_qty, 2) if reserve_qty > 0 else 0
                _s(row, 11, reserve_unit)
                ws.cell(row=row, column=11).number_format = NUM_FMT
                ws.cell(row=row, column=11).alignment = center
                ws.cell(row=row, column=12, value=round(reserve_price, 2)).font = SM_FONT
                ws.cell(row=row, column=12).alignment = center
                ws.cell(row=row, column=12).number_format = NUM_FMT
                ws.cell(row=row, column=13).fill = GRAY_FILL
                ws.merge_cells(f'A{row}:H{row}')
                ws.merge_cells(f'I{row}:J{row}')
                ws.merge_cells(f'L{row}:M{row}')

        _pile_sum_end = pile_data_end

        if pile_spare_count > 0 and pile_spare_price > 0:
            _ps_row = _pile_sum_end + 1
            _pile_sum_end = _ps_row
            for c in range(1, 14):
                ws.cell(row=_ps_row, column=c).border = THIN_BORDER
            ws.cell(row=_ps_row, column=1, value='杭予備品').font = SM_FONT
            ws.cell(row=_ps_row, column=1).alignment = center
            ws.cell(row=_ps_row, column=9, value=pile_spare_count).font = SM_FONT
            ws.cell(row=_ps_row, column=9).alignment = center
            ws.cell(row=_ps_row, column=9).number_format = '#,##0"本"'
            ws.cell(row=_ps_row, column=10).fill = GRAY_FILL
            _s(_ps_row, 11, pile_spare_price)
            ws.cell(row=_ps_row, column=11).number_format = NUM_FMT
            ws.cell(row=_ps_row, column=11).alignment = center
            ws.cell(row=_ps_row, column=12, value=f'=K{_ps_row}*I{_ps_row}').font = SM_FONT
            ws.cell(row=_ps_row, column=12).alignment = center
            ws.cell(row=_ps_row, column=12).number_format = NUM_FMT
            ws.cell(row=_ps_row, column=13).fill = GRAY_FILL
            ws.merge_cells(f'A{_ps_row}:H{_ps_row}')
            ws.merge_cells(f'I{_ps_row}:J{_ps_row}')
            ws.merge_cells(f'L{_ps_row}:M{_ps_row}')
            ws.row_dimensions[_ps_row].height = 20
            pile_data_end = _pile_sum_end

        r_spare_pile = None
        if dz_spare_count and dz_spare_count > 0:
            r_spare_pile = pile_data_end + 1
            for c in range(1, 14):
                ws.cell(row=r_spare_pile, column=c).border = THIN_BORDER
            ws.cell(row=r_spare_pile, column=1, value='無料予備品').font = SM_FONT
            ws.cell(row=r_spare_pile, column=1).alignment = center
            for c in range(1, 12):
                ws.cell(row=r_spare_pile, column=c).fill = _spare_fill
            ws.cell(row=r_spare_pile, column=12, value=0).font = SM_FONT
            ws.cell(row=r_spare_pile, column=12).alignment = center
            ws.cell(row=r_spare_pile, column=12).number_format = NUM_FMT
            ws.merge_cells(f'A{r_spare_pile}:K{r_spare_pile}')
            ws.merge_cells(f'L{r_spare_pile}:M{r_spare_pile}')
            ws.row_dimensions[r_spare_pile].height = 20

        r_ptotal = (r_spare_pile if r_spare_pile else pile_data_end) + 1
        for c in range(1, 14):
            ws.cell(row=r_ptotal, column=c).border = THIN_BORDER
        ws.cell(row=r_ptotal, column=1, value='合計').font = SM_FONT_BOLD
        ws.cell(row=r_ptotal, column=1).alignment = center
        ws.cell(row=r_ptotal, column=9, value=f'=SUM(I{pile_data_start}:I{pile_data_end})').font = SM_FONT_BOLD
        ws.cell(row=r_ptotal, column=9).alignment = center
        ws.cell(row=r_ptotal, column=9).number_format = '#,##0"本"'
        ws.cell(row=r_ptotal, column=11, value='-').font = SM_FONT_BOLD
        ws.cell(row=r_ptotal, column=11).alignment = center
        ws.cell(row=r_ptotal, column=12, value=f'=SUM(L{pile_data_start}:L{pile_data_end})').font = SM_FONT_BOLD
        ws.cell(row=r_ptotal, column=12).alignment = center
        ws.cell(row=r_ptotal, column=12).number_format = NUM_FMT
        ws.merge_cells(f'A{r_ptotal}:H{r_ptotal}')
        ws.merge_cells(f'I{r_ptotal}:J{r_ptotal}')
        ws.merge_cells(f'L{r_ptotal}:M{r_ptotal}')
        ws.row_dimensions[r_ptotal].height = 20

        r_bracket_pile_total = r_ptotal
    else:
        r_bracket_pile_total = r_disc

    YELLOW_FILL_TOTAL = PatternFill(start_color='FFDDB3', end_color='FFDDB3', fill_type='solid')
    r_bp_total = r_bracket_pile_total + 1
    for c in range(1, 14):
        ws.cell(row=r_bp_total, column=c).border = THIN_BORDER
    ws.cell(row=r_bp_total, column=1, value='架台と杭のトータル費用(USD)').font = SM_FONT_BOLD
    ws.cell(row=r_bp_total, column=1).alignment = right_a
    ws.merge_cells(f'A{r_bp_total}:K{r_bp_total}')
    if has_pile_data:
        ws.cell(row=r_bp_total, column=12, value=f'=L{r_disc}+L{r_bracket_pile_total}').font = SM_FONT_BOLD
    else:
        ws.cell(row=r_bp_total, column=12, value=f'=L{r_disc}').font = SM_FONT_BOLD
    ws.cell(row=r_bp_total, column=12).alignment = center
    ws.cell(row=r_bp_total, column=12).number_format = NUM_FMT
    ws.merge_cells(f'L{r_bp_total}:M{r_bp_total}')
    for c in range(1, 12):
        ws.cell(row=r_bp_total, column=c).fill = YELLOW_FILL_TOTAL
    ws.row_dimensions[r_bp_total].height = 20

    r_bp_note = r_bp_total + 1
    ws.merge_cells(f'A{r_bp_note}:M{r_bp_note}')
    ws.cell(row=r_bp_note, column=1, value='各アレイの部材リストは次のページに記載しております。').font = SM_FONT
    ws.cell(row=r_bp_note, column=1).alignment = left_a

    # ═══ フェンス金額 ═══

    r_fence_title = r_bp_note + 1
    nv_fgg = nv_fence_gate_data or {}
    detail_fence_rows = []
    detail_gate_rows = []
    fence_sections = []
    gate_sections = []
    fence_height_mm = ''
    fence_len_m = 0
    fence_color_jp = '白色'
    gate_type_label = ''
    gate_qty_val = 0

    multi_fences = nv_fgg.get('fences') if isinstance(nv_fgg, dict) else None
    multi_gates = nv_fgg.get('gates') if isinstance(nv_fgg, dict) else None

    has_fence_data = bool(nv_fgg and (nv_fgg.get('fence') or nv_fgg.get('gate') or multi_fences or multi_gates)) or bool(fence_data)

    _auto_sec = 2
    if has_pile_data:
        _auto_sec += 1
    _fence_sec_num = _auto_sec
    if has_fence_data:
        _auto_sec += 1
    if is_cif:
        sec_ocean = _get_cjk_num(_auto_sec)
        _auto_sec += 1
        sec_cif = _get_cjk_num(_auto_sec)
        _auto_sec += 1
    else:
        sec_ocean = None
        sec_cif = None
    if _show_ddp:
        sec_truck = _get_cjk_num(_auto_sec)
        _auto_sec += 1
        sec_tax = None
    else:
        sec_truck = None
        sec_tax = None

    if has_fence_data:
        if nv_fgg and (multi_fences or multi_gates):
            fence_surface = nv_fgg.get('surface', '白色')
            fence_color_jp = _surface_to_jp_color(fence_surface)
            if multi_gates:
                for g_sec in multi_gates:
                    g_color = g_sec.get('gateColor', '')
                    if g_color:
                        fence_color_jp = g_color
                        break
            if multi_fences:
                for fi, f_sec in enumerate(multi_fences):
                    section_rows = f_sec.get('rows', [])
                    detail_fence_rows.extend(section_rows)
                    section_len = float(f_sec.get('totalLength', 0) or 0)
                    fence_len_m += int(section_len)
                    f_style = f_sec.get('style', '')
                    section_height_mm = _extract_fence_height_mm_from_style(f_style)
                    fence_sections.append({
                        'rows': section_rows,
                        'length_m': section_len,
                        'height_mm': section_height_mm,
                        'surface': f_sec.get('surface', fence_surface),
                        'style': f_style,
                    })
                    if fi == 0 and section_height_mm:
                        fence_height_mm = section_height_mm
            if multi_gates:
                for g_sec in multi_gates:
                    section_rows = g_sec.get('rows', [])
                    section_qty = int(g_sec.get('gateQty', 0) or 0)
                    detail_gate_rows.extend(section_rows)
                    gate_qty_val += section_qty
                    g_style = g_sec.get('gateStyle', '')
                    gate_sections.append({
                        'rows': section_rows,
                        'qty': section_qty,
                        'gate_style': g_style,
                    })
                    if g_style and len(g_style) >= 6:
                        g_prefix = g_style[:3].lower()
                        if g_prefix.startswith('tl'):
                            gate_type_label = '引き戸'
                        elif g_prefix.startswith('tf'):
                            gate_type_label = '折りたたみ扉'
                        elif g_style[3:6] != '120':
                            gate_type_label = '両開き門'
                        else:
                            gate_type_label = '片開き門'
        elif nv_fgg:
            fence_section = nv_fgg.get('fence') or {}
            gate_section = nv_fgg.get('gate') or {}
            detail_fence_rows = fence_section.get('rows', [])
            detail_gate_rows = gate_section.get('rows', [])
            fence_surface = fence_section.get('surface', '白色')
            fence_color_jp = _surface_to_jp_color(fence_surface)
            gate_color = gate_section.get('gateColor', '')
            if gate_color:
                fence_color_jp = gate_color
            fence_len_value = float(fence_section.get('totalLength', 0) or 0)
            fence_len_m = int(fence_len_value)
            f_style = fence_section.get('style', '')
            fence_height_mm = _extract_fence_height_mm_from_style(f_style)
            fence_sections.append({
                'rows': detail_fence_rows,
                'length_m': fence_len_value,
                'height_mm': fence_height_mm,
                'surface': fence_surface,
                'style': f_style,
            })
            gate_qty_val = int(gate_section.get('gateQty', 0) or 0)
            g_style = gate_section.get('gateStyle', '')
            gate_sections.append({
                'rows': detail_gate_rows,
                'qty': gate_qty_val,
                'gate_style': g_style,
            })
            if g_style and len(g_style) >= 6:
                g_prefix = g_style[:3].lower()
                if g_prefix.startswith('tl'):
                    gate_type_label = '引き戸'
                elif g_prefix.startswith('tf'):
                    gate_type_label = '折りたたみ扉'
                elif g_style[3:6] != '120':
                    gate_type_label = '両開き門'
                else:
                    gate_type_label = '片開き門'

    all_fence_items = detail_fence_rows + detail_gate_rows
    fence_only_amount = Decimal(str(sum(r.get('amount', 0) for r in detail_fence_rows)))
    gate_only_amount = Decimal(str(sum(r.get('amount', 0) for r in detail_gate_rows)))

    if all_fence_items:
        coating = (nv_fgg or {}).get('浸塑') if (nv_fgg or {}).get('coating', '浸塑') == '热镀锌' else 'ディップコディング'
        if (nv_fgg or {}).get('coating', '浸塑') == '热镀锌':
            coating_jp = '溶融亜鉛めっき'
        else:
            coating_jp = 'ディップコディング'

        fence_desc_parts = []
        if fence_height_mm:
            fence_desc_parts.append(f'H{fence_height_mm}mm')
        fence_desc_parts.append(fence_color_jp)
        _fence_base_type = '一本打ち込み式'
        _first_fence_style = ''
        if fence_sections:
            _first_fence_style = str(fence_sections[0].get('style', '') or '').upper()
        elif nv_fgg:
            _f_single = (nv_fgg.get('fence') or {}).get('style', '')
            if _f_single:
                _first_fence_style = str(_f_single).upper()
        if 'CG' in _first_fence_style:
            _fence_base_type = '杭支柱分離式'
        elif 'CC' in _first_fence_style:
            _fence_base_type = 'コンクリート基礎'
        elif 'CP' in _first_fence_style:
            _fence_base_type = '一本打ち込み式'
        fence_title_text = f'{_get_cjk_num(_fence_sec_num)}、フェンス金額 （{_fence_base_type} ' + ' '.join(fence_desc_parts) + '）'

        ws.merge_cells(f'A{r_fence_title}:M{r_fence_title}')
        ws.cell(row=r_fence_title, column=1, value=fence_title_text).font = SM_FONT_BOLD_10
        ws.cell(row=r_fence_title, column=1).alignment = left_a
        ws.row_dimensions[r_fence_title].height = 22

        r_fence_hdr = r_fence_title + 1
        fence_hdr = ['番号', '部品名称', '', '', '材質', '', '規格', '', '', '単価(US$)', '数量（PCS)', '総金額(US$)', '']
        for i, h in enumerate(fence_hdr, 1):
            cell = ws.cell(row=r_fence_hdr, column=i, value=h)
            cell.font = SM_FONT_HEADER
            cell.alignment = center
            cell.border = THIN_BORDER
        ws.merge_cells(f'B{r_fence_hdr}:D{r_fence_hdr}')
        ws.merge_cells(f'E{r_fence_hdr}:F{r_fence_hdr}')
        ws.merge_cells(f'G{r_fence_hdr}:I{r_fence_hdr}')
        ws.merge_cells(f'L{r_fence_hdr}:M{r_fence_hdr}')
        ws.row_dimensions[r_fence_hdr].height = 22

        fence_data_start = r_fence_hdr + 1
        fence_agg_rows = []
        seq = 1

        if fence_sections:
            for fence_section_info in fence_sections:
                section_rows = fence_section_info.get('rows') or []
                if not section_rows:
                    continue
                row = fence_data_start + seq - 1
                ws.row_dimensions[row].height = 30
                section_length_m = float(fence_section_info.get('length_m', 0) or 0)
                fence_qty = round(section_length_m / 2, 2) if section_length_m > 0 else round(float(len(section_rows)), 2)
                section_amount = float(sum(r.get('amount', 0) or 0 for r in section_rows))
                fence_unit = section_amount / fence_qty if fence_qty > 0 else 0
                fence_h_val = fence_section_info.get('height_mm') or fence_height_mm or '1500'
                length_display = f'{section_length_m:.2f}' if section_length_m % 1 else str(int(section_length_m))
                fence_name = f'フェンス{length_display}m\n（網・柱材セット）'
                fence_material = f'スチールQ235B\n表面：{coating_jp}'
                fence_spec = f'H{fence_h_val}*W2000'

                _s(row, 1, seq)
                _s(row, 2, fence_name, align=Alignment(horizontal='center', vertical='center', wrap_text=True))
                _s(row, 3, '')
                _s(row, 4, '')
                _s(row, 5, fence_material, align=Alignment(horizontal='center', vertical='center', wrap_text=True))
                _s(row, 6, '')
                _s(row, 7, fence_spec, align=Alignment(horizontal='center', vertical='center', wrap_text=True))
                _s(row, 8, '')
                _s(row, 9, '')
                _s(row, 10, round(fence_unit, 2))
                ws.cell(row=row, column=10).number_format = NUM_FMT
                _s(row, 11, fence_qty)
                ws.cell(row=row, column=11).number_format = '0.00'
                ws.cell(row=row, column=12, value=f'=K{row}*J{row}').font = SM_FONT
                ws.cell(row=row, column=12).alignment = center
                ws.cell(row=row, column=12).border = THIN_BORDER
                ws.cell(row=row, column=12).number_format = NUM_FMT
                _s(row, 13, '')
                ws.merge_cells(f'B{row}:D{row}')
                ws.merge_cells(f'E{row}:F{row}')
                ws.merge_cells(f'G{row}:I{row}')
                ws.merge_cells(f'L{row}:M{row}')
                fence_agg_rows.append(row)
                seq += 1

        if gate_sections:
            for g_sec in gate_sections:
                g_rows = g_sec.get('rows') or []
                g_qty = int(g_sec.get('qty', 0) or 0)
                g_style = str(g_sec.get('gate_style', '') or '')

                if not g_rows or g_qty <= 0:
                    continue

                row = fence_data_start + seq - 1
                ws.row_dimensions[row].height = 30
                g_amount = float(sum(r.get('amount', 0) or 0 for r in g_rows))
                g_unit = g_amount / g_qty if g_qty > 0 else 0

                gate_name = '片開き門'
                w_label = 'W1200'
                g_prefix_lower = g_style[:3].lower()
                if g_prefix_lower.startswith('tl'):
                    gate_name = '引き戸'
                    w_code = g_style[3:6]
                    w_label = f'W{int(w_code) * 10}' if w_code.isdigit() else 'W2400'
                elif g_prefix_lower.startswith('tf'):
                    gate_name = '折りたたみ扉'
                    w_code = g_style[3:6]
                    w_label = f'W{int(w_code) * 10}' if w_code.isdigit() else 'W2400'
                elif g_style.startswith('td'):
                    gate_name = '両開き門'
                    if '420' in g_style:
                        w_label = 'W4200'
                    else:
                        w_label = 'W2400'
                elif g_style.startswith('ts'):
                    gate_name = '片開き門'
                    w_label = 'W1200'

                gate_material = f'スチールQ235B\n表面：{coating_jp}'
                db_h, db_w = _lookup_gate_spec_from_db(g_style)
                if db_h is not None and db_w is not None:
                    gate_spec = f'H{db_h}*W{db_w}'
                else:
                    gate_h_val = fence_height_mm or '1500'
                    gate_spec = f'H{gate_h_val}*{w_label}'

                _s(row, 1, seq)
                _s(row, 2, gate_name, align=Alignment(horizontal='center', vertical='center', wrap_text=True))
                _s(row, 3, '')
                _s(row, 4, '')
                _s(row, 5, gate_material, align=Alignment(horizontal='center', vertical='center', wrap_text=True))
                _s(row, 6, '')
                _s(row, 7, gate_spec, align=Alignment(horizontal='center', vertical='center', wrap_text=True))
                _s(row, 8, '')
                _s(row, 9, '')
                _s(row, 10, round(g_unit, 2))
                ws.cell(row=row, column=10).number_format = NUM_FMT
                _s(row, 11, g_qty)
                ws.cell(row=row, column=12, value=f'=K{row}*J{row}').font = SM_FONT
                ws.cell(row=row, column=12).alignment = center
                ws.cell(row=row, column=12).border = THIN_BORDER
                ws.cell(row=row, column=12).number_format = NUM_FMT
                _s(row, 13, '')
                ws.merge_cells(f'B{row}:D{row}')
                ws.merge_cells(f'E{row}:F{row}')
                ws.merge_cells(f'G{row}:I{row}')
                ws.merge_cells(f'L{row}:M{row}')
                fence_agg_rows.append(row)
                seq += 1
        elif detail_gate_rows:
            row = fence_data_start + seq - 1
            ws.row_dimensions[row].height = 30
            gate_qty = gate_qty_val if gate_qty_val > 0 else 1
            gate_unit = float(gate_only_amount) / gate_qty if gate_qty > 0 else 0
            gate_name = gate_type_label or '片開き扉'
            gate_material = f'スチールQ235B\n表面：{coating_jp}'
            w_label = 'W1200' if '片開き' in gate_name else 'W4200'
            fallback_g_style = ''
            if gate_sections:
                fallback_g_style = str(gate_sections[0].get('gate_style', '') or '')
            db_h, db_w = _lookup_gate_spec_from_db(fallback_g_style)
            if db_h is not None and db_w is not None:
                gate_spec = f'H{db_h}*W{db_w}'
            else:
                gate_h_val = fence_height_mm or '1500'
                gate_spec = f'H{gate_h_val}*{w_label}'

            _s(row, 1, seq)
            _s(row, 2, gate_name, align=Alignment(horizontal='center', vertical='center', wrap_text=True))
            _s(row, 3, '')
            _s(row, 4, '')
            _s(row, 5, gate_material, align=Alignment(horizontal='center', vertical='center', wrap_text=True))
            _s(row, 6, '')
            _s(row, 7, gate_spec, align=Alignment(horizontal='center', vertical='center', wrap_text=True))
            _s(row, 8, '')
            _s(row, 9, '')
            _s(row, 10, round(gate_unit, 2))
            ws.cell(row=row, column=10).number_format = NUM_FMT
            _s(row, 11, gate_qty)
            ws.cell(row=row, column=12, value=f'=K{row}*J{row}').font = SM_FONT
            ws.cell(row=row, column=12).alignment = center
            ws.cell(row=row, column=12).border = THIN_BORDER
            ws.cell(row=row, column=12).number_format = NUM_FMT
            _s(row, 13, '')
            ws.merge_cells(f'B{row}:D{row}')
            ws.merge_cells(f'E{row}:F{row}')
            ws.merge_cells(f'G{row}:I{row}')
            ws.merge_cells(f'L{row}:M{row}')
            fence_agg_rows.append(row)
            seq += 1

        fence_data_end = fence_data_start + len(fence_agg_rows) - 1

        r_fence_sub = fence_data_end + 1
        ws.cell(row=r_fence_sub, column=1, value='フェンス金額').font = SM_FONT_BOLD
        ws.cell(row=r_fence_sub, column=1).alignment = right_a
        ws.cell(row=r_fence_sub, column=1).border = THIN_BORDER
        ws.merge_cells(f'A{r_fence_sub}:K{r_fence_sub}')
        ws.cell(row=r_fence_sub, column=12, value=f'=SUM(L{fence_data_start}:L{fence_data_end})').font = SM_FONT_BOLD
        ws.cell(row=r_fence_sub, column=12).alignment = center
        ws.cell(row=r_fence_sub, column=12).border = THIN_BORDER
        ws.cell(row=r_fence_sub, column=12).number_format = NUM_FMT
        ws.merge_cells(f'L{r_fence_sub}:M{r_fence_sub}')
        ws.row_dimensions[r_fence_sub].height = 20

        r_fence_disc = r_fence_sub + 1
        ws.cell(row=r_fence_disc, column=1, value='特別値引き後金額').font = SM_FONT_BOLD
        ws.cell(row=r_fence_disc, column=1).alignment = right_a
        ws.cell(row=r_fence_disc, column=1).border = THIN_BORDER
        ws.merge_cells(f'A{r_fence_disc}:K{r_fence_disc}')
        ws.cell(row=r_fence_disc, column=12, value=f'=L{r_fence_sub}*{fence_discount_rate}/100').font = SM_FONT_BOLD
        ws.cell(row=r_fence_disc, column=12).alignment = center
        ws.cell(row=r_fence_disc, column=12).border = THIN_BORDER
        ws.cell(row=r_fence_disc, column=12).number_format = NUM_FMT
        ws.merge_cells(f'L{r_fence_disc}:M{r_fence_disc}')
        ws.row_dimensions[r_fence_disc].height = 20

        if detail_fence_rows or detail_gate_rows:
            try:
                _create_fence_detail_sheet(workbook, detail_fence_rows, detail_gate_rows, nv_fgg,
                                           matrix_data=matrix_data, nv_params=nv_params,
                                           image_temp_dir=image_temp_dir, image_cache=image_cache)
            except Exception as e:
                print(f"   ⚠ フェンス物料明細シート生成失敗: {e}")

        r_next = r_fence_disc + 1
    else:
        r_next = r_fence_title

    r_ocean_total = None

    if is_cif:
        # ═══ 海上運賃 ═══

        r_ocean_title = r_next
        ws.merge_cells(f'A{r_ocean_title}:M{r_ocean_title}')
        ws.cell(row=r_ocean_title, column=1, value=f'{sec_ocean}、海上運賃').font = SM_FONT_BOLD_10
        ws.cell(row=r_ocean_title, column=1).alignment = left_a
        ws.row_dimensions[r_ocean_title].height = 22

        r_ocean_hdr = r_ocean_title + 1
        ship_hdr = ['NO.', '仕上げ港', '', '', 'コンテナ情報', '', '', '', '台数', '単価(USD)', '', '総金額(USD)', '']
        for i, h in enumerate(ship_hdr, 1):
            cell = ws.cell(row=r_ocean_hdr, column=i, value=h)
            cell.font = SM_FONT_HEADER
            cell.alignment = center
            cell.border = THIN_BORDER
            if h == '':
                cell.fill = GRAY_FILL
        ws.merge_cells(f'B{r_ocean_hdr}:D{r_ocean_hdr}')
        ws.merge_cells(f'E{r_ocean_hdr}:H{r_ocean_hdr}')
        ws.merge_cells(f'J{r_ocean_hdr}:K{r_ocean_hdr}')
        ws.merge_cells(f'L{r_ocean_hdr}:M{r_ocean_hdr}')
        ws.row_dimensions[r_ocean_hdr].height = 22

        r_ocean_data_start = r_ocean_hdr + 1
        ocean_data_rows = []
        for ci, ct_info in enumerate(containers):
            r_d = r_ocean_data_start + ci
            ws.row_dimensions[r_d].height = 20
            ct_type = ct_info.get('type', '混載便')
            ct_qty = int(ct_info.get('qty', 1))
            ct_freight = float(ct_info.get('freight', 0))
            if ct_freight <= 0 and ct_qty <= 0:
                continue
            _s(r_d, 1, ci + 1)
            _s(r_d, 2, dest_port)
            _s(r_d, 3, '')
            ws.cell(row=r_d, column=3).fill = GRAY_FILL
            _s(r_d, 4, '')
            ws.cell(row=r_d, column=4).fill = GRAY_FILL
            _s(r_d, 5, ct_type)
            _s(r_d, 6, '')
            ws.cell(row=r_d, column=6).fill = GRAY_FILL
            _s(r_d, 7, '')
            ws.cell(row=r_d, column=7).fill = GRAY_FILL
            _s(r_d, 8, '')
            ws.cell(row=r_d, column=8).fill = GRAY_FILL
            _s(r_d, 9, ct_qty)
            _s(r_d, 10, ct_freight)
            ws.cell(row=r_d, column=10).number_format = NUM_FMT
            _s(r_d, 11, '')
            ws.cell(row=r_d, column=11).fill = GRAY_FILL
            ws.cell(row=r_d, column=12, value=f'=J{r_d}*I{r_d}').font = SM_FONT
            ws.cell(row=r_d, column=12).alignment = center
            ws.cell(row=r_d, column=12).border = THIN_BORDER
            ws.cell(row=r_d, column=12).number_format = NUM_FMT
            _s(r_d, 13, '')
            ws.merge_cells(f'B{r_d}:D{r_d}')
            ws.merge_cells(f'E{r_d}:H{r_d}')
            ws.merge_cells(f'J{r_d}:K{r_d}')
            ws.merge_cells(f'L{r_d}:M{r_d}')
            ocean_data_rows.append(r_d)

        r_ocean_total = (ocean_data_rows[-1] if ocean_data_rows else r_ocean_hdr) + 1
        ws.cell(row=r_ocean_total, column=1, value='海上運賃合計金額').font = SM_FONT_BOLD
        ws.cell(row=r_ocean_total, column=1).alignment = right_a
        ws.cell(row=r_ocean_total, column=1).border = THIN_BORDER
        ws.merge_cells(f'A{r_ocean_total}:K{r_ocean_total}')
        if ocean_data_rows:
            ws.cell(row=r_ocean_total, column=12, value=f'=SUM(L{ocean_data_rows[0]}:L{ocean_data_rows[-1]})').font = SM_FONT_BOLD
        else:
            ws.cell(row=r_ocean_total, column=12, value=0).font = SM_FONT_BOLD
        ws.cell(row=r_ocean_total, column=12).alignment = center
        ws.cell(row=r_ocean_total, column=12).border = THIN_BORDER
        ws.cell(row=r_ocean_total, column=12).number_format = NUM_FMT
        ws.merge_cells(f'L{r_ocean_total}:M{r_ocean_total}')

        r_cif_title = r_ocean_total + 1
    else:
        r_cif_title = r_next

    cif_formula = ''
    if has_pile_data:
        cif_formula = f'L{r_disc}+L{r_bracket_pile_total}'
    else:
        cif_formula = f'L{r_bracket_pile_total}'
    if all_fence_items:
        cif_formula += f'+L{r_fence_disc}'
    if is_cif and r_ocean_total:
        cif_formula += f'+L{r_ocean_total}'

    r_cif_val = None
    YELLOW_FILL_2 = PatternFill(start_color='FFDDB3', end_color='FFDDB3', fill_type='solid')

    need_jpy_quote = bool(nv_params.get('need_jpy_quote'))
    r_jpy = None

    if is_cif:
        # ═══ CIF最終価格 ═══

        ws.merge_cells(f'A{r_cif_title}:M{r_cif_title}')
        ws.cell(row=r_cif_title, column=1, value=f'{sec_cif}、CIF最終価格').font = SM_FONT_BOLD_10
        ws.cell(row=r_cif_title, column=1).alignment = left_a
        ws.row_dimensions[r_cif_title].height = 22

        r_cif_val = r_cif_title + 1
        ws.cell(row=r_cif_val, column=1, value='CIFトータル費用(USD)').font = SM_FONT_BOLD
        ws.cell(row=r_cif_val, column=1).alignment = right_a
        ws.cell(row=r_cif_val, column=1).border = THIN_BORDER
        ws.merge_cells(f'A{r_cif_val}:K{r_cif_val}')
        ws.cell(row=r_cif_val, column=12, value=f'=FLOOR({cif_formula},10)').font = SM_FONT_BOLD
        ws.cell(row=r_cif_val, column=12).alignment = center
        ws.cell(row=r_cif_val, column=12).border = THIN_BORDER
        ws.cell(row=r_cif_val, column=12).number_format = '#,##0.00'
        ws.merge_cells(f'L{r_cif_val}:M{r_cif_val}')
        for c in range(1, 12):
            ws.cell(row=r_cif_val, column=c).fill = YELLOW_FILL_2

        r_last_value = r_cif_val
    else:
        r_cif_val = None
        r_last_value = r_next - 1

    if is_cif and need_jpy_quote and not _show_ddp:
        r_jpy = r_last_value + 1
        ws.cell(row=r_jpy, column=1, value='最終価格調整日本円(JPY)').font = SM_FONT_BOLD
        ws.cell(row=r_jpy, column=1).alignment = right_a
        ws.cell(row=r_jpy, column=1).border = THIN_BORDER
        ws.merge_cells(f'A{r_jpy}:K{r_jpy}')
        ws.cell(row=r_jpy, column=12, value=f'=L{r_cif_val}*{exchange_rate}').font = SM_FONT_BOLD
        ws.cell(row=r_jpy, column=12).alignment = center
        ws.cell(row=r_jpy, column=12).border = THIN_BORDER
        ws.cell(row=r_jpy, column=12).number_format = '#,##0'
        ws.merge_cells(f'L{r_jpy}:M{r_jpy}')
        r_last_value = r_jpy

    if _show_ddp:
        if r_cif_val is None:
            r_ddp_base = r_next
        else:
            r_ddp_base = r_cif_val

        # ═══ 陸送 ═══

        r_parent = r_last_value + 1
        ws.merge_cells(f'A{r_parent}:M{r_parent}')
        ws.cell(row=r_parent, column=1, value=f'{sec_truck}、輸入諸経費（輸入関税・消費税）＆現場直送迄の諸費用').font = SM_FONT_BOLD_10
        ws.cell(row=r_parent, column=1).alignment = left_a
        ws.row_dimensions[r_parent].height = 22

        r_truck = r_parent + 1
        ws.cell(row=r_truck, column=1, value=1).font = SM_FONT
        ws.cell(row=r_truck, column=1).border = THIN_BORDER
        ws.cell(row=r_truck, column=1).alignment = center
        ws.cell(row=r_truck, column=2, value=f'現場まで陸送の諸費用（{truck_desc}）').font = SM_FONT
        ws.cell(row=r_truck, column=2).border = THIN_BORDER
        ws.cell(row=r_truck, column=2).alignment = center
        ws.merge_cells(f'B{r_truck}:K{r_truck}')
        ws.cell(row=r_truck, column=12, value=truck_fee).font = SM_FONT_BOLD
        ws.cell(row=r_truck, column=12).alignment = center
        ws.cell(row=r_truck, column=12).border = THIN_BORDER
        ws.cell(row=r_truck, column=12).number_format = NUM_FMT
        ws.merge_cells(f'L{r_truck}:M{r_truck}')
        ws.row_dimensions[r_truck].height = 22

        r_tax = r_truck + 1
        ws.cell(row=r_tax, column=1, value=2).font = SM_FONT
        ws.cell(row=r_tax, column=1).border = THIN_BORDER
        ws.cell(row=r_tax, column=1).alignment = center
        ws.cell(row=r_tax, column=2, value=f'税金（消費税{consumption_tax_pct}%・関税{tariff_rate_pct}%）').font = SM_FONT
        ws.cell(row=r_tax, column=2).border = THIN_BORDER
        ws.cell(row=r_tax, column=2).alignment = center
        ws.merge_cells(f'B{r_tax}:K{r_tax}')
        _bp_base = f'L{r_disc}+L{r_bp_total}'
        _fence_part = f'+L{r_fence_disc}' if all_fence_items else ''
        _ct_rate = round(consumption_tax_pct / 100, 4)
        _tr_rate = round(tariff_rate_pct / 100, 4)
        tax_formula = f'=({_bp_base})*{_ct_rate}+({_bp_base}{_fence_part})*{_tr_rate}'
        ws.cell(row=r_tax, column=12, value=tax_formula).font = SM_FONT_BOLD
        ws.cell(row=r_tax, column=12).alignment = center
        ws.cell(row=r_tax, column=12).border = THIN_BORDER
        ws.cell(row=r_tax, column=12).number_format = NUM_FMT
        ws.merge_cells(f'L{r_tax}:M{r_tax}')
        ws.row_dimensions[r_tax].height = 22

        # ═══ DDP ═══

        r_ddp = r_tax + 1
        ws.cell(row=r_ddp, column=1, value='DDPトータル費用(USD)').font = SM_FONT_BOLD
        ws.cell(row=r_ddp, column=1).alignment = right_a
        ws.cell(row=r_ddp, column=1).border = THIN_BORDER
        ws.merge_cells(f'A{r_ddp}:K{r_ddp}')
        if r_cif_val is not None:
            ws.cell(row=r_ddp, column=12, value=f'=L{r_cif_val}+L{r_truck}+L{r_tax}').font = SM_FONT_BOLD
        else:
            ws.cell(row=r_ddp, column=12, value=f'={cif_formula}+L{r_truck}+L{r_tax}').font = SM_FONT_BOLD
        ws.cell(row=r_ddp, column=12).alignment = center
        ws.cell(row=r_ddp, column=12).border = THIN_BORDER
        ws.cell(row=r_ddp, column=12).number_format = NUM_FMT
        ws.merge_cells(f'L{r_ddp}:M{r_ddp}')
        ws.row_dimensions[r_ddp].height = 20

        r_adjust = r_ddp + 1
        ws.cell(row=r_adjust, column=1, value='DDP最終価格調整（USD）').font = SM_FONT_BOLD
        ws.cell(row=r_adjust, column=1).alignment = right_a
        ws.cell(row=r_adjust, column=1).border = THIN_BORDER
        ws.merge_cells(f'A{r_adjust}:K{r_adjust}')
        if final_price_usd:
            _adj_val = int(float(final_price_usd) // 10) * 10
            ws.cell(row=r_adjust, column=12, value=_adj_val).font = SM_FONT_BOLD
        else:
            ws.cell(row=r_adjust, column=12, value=f'=FLOOR(L{r_ddp},10)').font = SM_FONT_BOLD
        ws.cell(row=r_adjust, column=12).alignment = center
        ws.cell(row=r_adjust, column=12).border = THIN_BORDER
        ws.cell(row=r_adjust, column=12).number_format = '#,##0.00'
        ws.merge_cells(f'L{r_adjust}:M{r_adjust}')
        ws.row_dimensions[r_adjust].height = 20
        for c in range(1, 12):
            ws.cell(row=r_adjust, column=c).fill = YELLOW_FILL_2

        r_last_value = r_adjust

        if need_jpy_quote:
            r_jpy = r_last_value + 1
            ws.cell(row=r_jpy, column=1, value='最終価格調整日本円(JPY)').font = SM_FONT_BOLD
            ws.cell(row=r_jpy, column=1).alignment = right_a
            ws.cell(row=r_jpy, column=1).border = THIN_BORDER
            ws.merge_cells(f'A{r_jpy}:K{r_jpy}')
            ws.cell(row=r_jpy, column=12, value=f'=L{r_adjust}*{exchange_rate}').font = SM_FONT_BOLD
            ws.cell(row=r_jpy, column=12).alignment = center
            ws.cell(row=r_jpy, column=12).border = THIN_BORDER
            ws.cell(row=r_jpy, column=12).number_format = '#,##0'
            ws.merge_cells(f'L{r_jpy}:M{r_jpy}')
            r_last_value = r_jpy

    # ═══ 備考 ═══

    NOTE_FONT = Font(name='Yu Gothic UI', size=10)

    if mitsumori_condition == 'NV':
        r_note1 = r_last_value + 1
        ws.merge_cells(f'A{r_note1}:M{r_note1}')
        ws.cell(row=r_note1, column=1, value='＊備考：日本港で抜取り検査が発生した場合、実際の費用は別途請求させて頂きます。').font = NOTE_FONT
        ws.cell(row=r_note1, column=1).alignment = left_a
    elif is_cif:
        r_note1 = r_last_value + 1
        ws.merge_cells(f'A{r_note1}:M{r_note1}')
        ws.cell(row=r_note1, column=1, value='*備考：EBS BAF YAS LSS 日本現地支払いとなります。').font = NOTE_FONT
        ws.cell(row=r_note1, column=1).alignment = left_a

        r_note2 = r_note1 + 1
        ws.merge_cells(f'A{r_note2}:M{r_note2}')
        ws.cell(row=r_note2, column=1, value='*備考：混載便　東京港FT3日/その他5日（デバン日含む）').font = NOTE_FONT
        ws.cell(row=r_note2, column=1).alignment = left_a

        r_note3 = r_note2 + 1
        ws.merge_cells(f'A{r_note3}:M{r_note3}')
        ws.cell(row=r_note3, column=1, value='*備考：コンテナ　FT期間 14+7日').font = NOTE_FONT
        ws.cell(row=r_note3, column=1).alignment = left_a

        r_note4 = r_note3 + 1
        ws.merge_cells(f'A{r_note4}:M{r_note4}')
        ws.cell(row=r_note4, column=1, value='14：ETA日から計算して14日以内に港にコンテナを取る').font = NOTE_FONT
        ws.cell(row=r_note4, column=1).alignment = left_a

        r_note5 = r_note4 + 1
        ws.merge_cells(f'A{r_note5}:M{r_note5}')
        ws.cell(row=r_note5, column=1, value='7：コンテナを取る日から計算して7日以内にコンテナを返送する').font = NOTE_FONT
        ws.cell(row=r_note5, column=1).alignment = left_a
    else:
        r_note1 = r_last_value + 1
        ws.merge_cells(f'A{r_note1}:M{r_note1}')
        ws.cell(row=r_note1, column=1, value=f'＊備考：為替レートは本日の1USD={exchange_rate}JPY，変動が±2%以上の場合、その場のレートで改めて換算させて頂きます。').font = NOTE_FONT
        ws.cell(row=r_note1, column=1).alignment = left_a

        r_note2 = r_note1 + 1
        ws.merge_cells(f'A{r_note2}:M{r_note2}')
        ws.cell(row=r_note2, column=1, value='＊備考：日本港で抜取り検査が発生した場合、実際の費用は別途請求させて頂きます。').font = NOTE_FONT
        ws.cell(row=r_note2, column=1).alignment = left_a

    # ═══ Back-fill header formulas ═══

    if need_jpy_quote and r_jpy is not None:
        ws.cell(row=total_amount_cell_r, column=3, value=f'=L{r_jpy}').font = _FONT_12_BOLD
        ws.cell(row=total_amount_cell_r, column=3).alignment = _bottom_center
        ws.cell(row=total_amount_cell_r, column=3).number_format = '"JPY" #,##0'
    elif _show_ddp:
        ws.cell(row=total_amount_cell_r, column=3, value=f'=L{r_adjust}').font = _FONT_12_BOLD
        ws.cell(row=total_amount_cell_r, column=3).alignment = _bottom_center
        ws.cell(row=total_amount_cell_r, column=3).number_format = '"US$" #,##0'
    else:
        ws.cell(row=total_amount_cell_r, column=3, value=f'=L{r_cif_val}').font = _FONT_12_BOLD
        ws.cell(row=total_amount_cell_r, column=3).alignment = _bottom_center
        ws.cell(row=total_amount_cell_r, column=3).number_format = '"US$" #,##0'
    ws.cell(row=kw_price_cell_r, column=3, value=f'=C{total_amount_cell_r}/I{r_btotal}').font = _FONT_12_BOLD
    ws.cell(row=kw_price_cell_r, column=3).alignment = _bottom_center
    if need_jpy_quote and r_jpy is not None:
        ws.cell(row=kw_price_cell_r, column=3).number_format = '"JPY" #,##0'
    else:
        ws.cell(row=kw_price_cell_r, column=3).number_format = '"US$" #,##0.000'

    # ═══ 统一设置所有边框（内部细边框 + 外部粗边框） ═══
    
    # 定义边框范围
    box_top = 19
    box_bottom = r_last_value
    box_left = 1
    box_right = 13
    
    thin_side = Side(style='thin', color='000000')
    thick_side = Side(style='medium', color='000000')
    
    print(f"设置边框: 行范围 {box_top} 到 {box_bottom}, 列范围 {box_left} 到 {box_right}")
    box_left = 1
    box_right = 13
    
    thin_side = Side(style='thin', color='000000')
    thick_side = Side(style='medium', color='000000')
    
    print(f"设置边框: 行范围 {box_top} 到 {box_bottom}, 列范围 {box_left} 到 {box_right}")
    
    # ============================================================
    # 第1步：为所有单元格设置细边框
    # ============================================================
    for row in range(box_top, box_bottom + 1):
        for col in range(box_left, box_right + 1):
            cell = ws.cell(row=row, column=col)
            
            # 检查是否是合并单元格中的非起始单元格
            is_merged_non_start = False
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    if cell.coordinate != merged_range.start_cell.coordinate:
                        is_merged_non_start = True
                    break
            
            if not is_merged_non_start:
                cell.border = Border(
                    left=thin_side, right=thin_side,
                    top=thin_side, bottom=thin_side
                )
    
    print("第1步完成: 内部细边框已设置")
    
    # ============================================================
    # 第2步：设置顶部粗边框
    # ============================================================
    for col in range(box_left, box_right + 1):
        cell = ws.cell(row=box_top, column=col)
        
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                cell = merged_range.start_cell
                break
        
        old_border = cell.border
        if old_border is None:
            old_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        cell.border = Border(
            left=old_border.left,
            right=old_border.right,
            top=thick_side,
            bottom=old_border.bottom
        )
    
    print("第2步完成: 顶部粗边框已设置")
    
    # ============================================================
    # 第3步：设置底部粗边框（到 r_jpy 结束）
    # ============================================================
    for col in range(box_left, box_right + 1):
        cell = ws.cell(row=box_bottom, column=col)
        
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                cell = merged_range.start_cell
                break
        
        old_border = cell.border
        if old_border is None:
            old_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        cell.border = Border(
            left=old_border.left,
            right=old_border.right,
            top=old_border.top,
            bottom=thick_side
        )
    
    print("第3步完成: 底部粗边框已设置")
    
    # ============================================================
    # 第4步：设置左侧粗边框
    # ============================================================
    for row in range(box_top, box_bottom + 1):
        cell = ws.cell(row=row, column=box_left)
        
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                cell = merged_range.start_cell
                break
        
        old_border = cell.border
        if old_border is None:
            old_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        cell.border = Border(
            left=thick_side,
            right=old_border.right,
            top=old_border.top,
            bottom=old_border.bottom
        )
    
    print("第4步完成: 左侧粗边框已设置")
    
    # ============================================================
    # 第5步：设置右侧粗边框
    # ============================================================
    for row in range(box_top, box_bottom + 1):
        cell = ws.cell(row=row, column=box_right)
        
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                cell = merged_range.start_cell
                break
        
        old_border = cell.border
        if old_border is None:
            old_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        cell.border = Border(
            left=old_border.left,
            right=thick_side,
            top=old_border.top,
            bottom=old_border.bottom
        )
    
    print("第5步完成: 右侧粗边框已设置")
    print(f"边框设置完成，底部行: {box_bottom}")

    # ═══ Page setup ═══

    sheet_names = workbook.sheetnames
    actual_title = sheet_title or '合計'
    if actual_title in sheet_names:
        idx = sheet_names.index(actual_title)
        workbook.move_sheet(actual_title, offset=-idx)

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(top=0.75, bottom=0.75, left=0.25, right=0.25, header=0.3, footer=0.3)

    return ws.title


def create_spare_parts_sheet(workbook, spare_products, price_mapping,
                              coating_thickness=10, need_weight_code=False,
                              need_weight=False, need_code=False,
                              code_to_images=None, image_temp_dir=None, image_cache=None):
    from backend.core.shared.image_utils import (
        prepare_image_for_excel, add_image_centered_in_cell,
    )

    code_to_images = code_to_images or {}
    image_cache = image_cache or {}

    _sp_add_code = need_code or need_weight_code
    _sp_add_weight = need_weight or need_weight_code
    _sp_code_col = 10 if _sp_add_code else None
    _sp_weight_col = (10 + (1 if _sp_add_code else 0)) if _sp_add_weight else None
    _sp_col_end = 9 + (1 if _sp_add_code else 0) + (1 if _sp_add_weight else 0)
    _sp_merge_end = chr(ord('A') + _sp_col_end - 1)

    ws = workbook.create_sheet(title='部材リスト-無料予備品')

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 16
    if _sp_add_code:
        ws.column_dimensions[chr(ord('A') + _sp_code_col - 1)].width = 14
    if _sp_add_weight:
        ws.column_dimensions[chr(ord('A') + _sp_weight_col - 1)].width = 10

    ws.merge_cells(f'A1:{_sp_merge_end}1')
    title_cell = ws.cell(row=1, column=1, value='部材リスト-無料予備品')
    title_cell.font = SM_FONT_TITLE
    title_cell.alignment = CENTER

    headers = ['部品名称', '材質', '写真', '規格', '単価\n(EXW US$)\nUnit Price', '数量（PCS)\nQTY', '総金額\n(EXW US$)\nAmount Price']
    if _sp_add_code:
        headers += ['品番']
    if _sp_add_weight:
        headers += ['重量(KG)']

    header_row = 3
    ws.merge_cells(f'A{header_row}:B{header_row}')
    cell_num = ws.cell(row=header_row, column=1, value='番号\nItem No.')
    cell_num.font = Font(name='Yu Gothic UI', bold=True, size=8)
    cell_num.alignment = CENTER
    cell_num.border = THIN_BORDER
    cell_num.fill = LIGHT_BLUE_FILL
    ws.cell(row=header_row, column=2).border = THIN_BORDER
    ws.cell(row=header_row, column=2).fill = LIGHT_BLUE_FILL
    for i, h in enumerate(headers, 3):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = Font(name='Yu Gothic UI', bold=True, size=8)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        cell.fill = LIGHT_BLUE_FILL
    ws.row_dimensions[header_row].height = 40

    img_w, img_h, img_col, img_padding = 65, 50, 5, 2

    def col_width_to_px(cw):
        return int(cw * 7.5 + 5)

    def row_height_to_px(rh):
        return int(rh * 1.33)

    def fit_image(ws, r, c, mw, mh, pad=img_padding):
        cw = ws.column_dimensions[get_column_letter(c)].width or 10
        rh = ws.row_dimensions[r].height or 40
        cell_w = col_width_to_px(cw) - 2 * pad
        cell_h = row_height_to_px(rh) - 2 * pad
        fw = min(mw, cell_w)
        fh = min(mh, cell_h)
        return fw, fh

    data_start = header_row + 1
    name_field = 'name_ja'

    for idx, product in enumerate(spare_products):
        row = data_start + idx
        ws.row_dimensions[row].height = 40

        product_code = product.get('code', '')
        price_info = resolve_price_info(price_mapping, product_code, spec=product.get('spec', ''))

        display_name = (
            price_info.get(name_field)
            or price_info.get('name_ko')
            or price_info.get('name')
            or product.get('name', '')
        ) if price_info else product.get('name', '')

        raw_material = (price_info.get('db_material') if price_info and price_info.get('db_material') else None) or product.get('material', '')
        display_material = ''
        if raw_material:
            mat = str(raw_material).replace('&', '+')
            if coating_thickness in (15, 18):
                mat = mat.replace('304', '316')
                mat = mat.replace('/316', '')
            else:
                mat = mat.replace('/316', '')
            mat = translate_material(mat, 'ja')
            display_material = mat
            if coating_thickness in (15, 18):
                display_material += f'  {coating_thickness}um'

        spec = product.get('spec', '')
        quantity = int(float(product.get('quantity', 0) or 0))

        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row=row, column=1, value=idx + 1).font = SM_FONT
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=3, value=display_name).font = SM_FONT
        ws.cell(row=row, column=3).alignment = CENTER
        ws.cell(row=row, column=4, value=display_material).font = SM_FONT
        ws.cell(row=row, column=4).alignment = CENTER
        _set(ws, row, 5, '')

        _set(ws, row, 6, spec)

        display_unit_price = 0.0
        if price_info and has_valid_price_info(price_info):
            display_unit_price = float(price_info['price'])
            pricing_unit = price_info.get('unit', '')
            if pricing_unit == '米':
                length_mm = float(product.get('length', 0) or 0)
                if not length_mm:
                    length_mm = float(extract_length_from_spec(product.get('spec', '')) or 0)
                if length_mm > 0:
                    display_unit_price = display_unit_price * length_mm / 1000

        ws.cell(row=row, column=7, value=display_unit_price).font = SM_FONT
        ws.cell(row=row, column=7).alignment = CENTER
        ws.cell(row=row, column=7).number_format = NUM_FMT

        ws.cell(row=row, column=8, value=quantity).font = SM_FONT
        ws.cell(row=row, column=8).alignment = CENTER

        ws.cell(row=row, column=9, value='無料な予備品').font = SM_FONT
        ws.cell(row=row, column=9).alignment = CENTER

        if _sp_add_code:
            ws.cell(row=row, column=_sp_code_col, value=product_code).font = SM_FONT
            ws.cell(row=row, column=_sp_code_col).alignment = CENTER
            ws.cell(row=row, column=_sp_code_col).border = THIN_BORDER
        if _sp_add_weight:
            _pile_weight = '/'
            if price_info:
                _pw = price_info.get('db_weight')
                _pu = price_info.get('unit', '')
                if _pw:
                    if _pu in ['米', 'm', 'M', 'meter', 'Meter']:
                        _pl = extract_length_from_spec(product.get('spec', ''))
                        _pile_weight = round(float(_pw) / 1000 * _pl, 4) if _pl and _pl > 0 else float(_pw)
                    else:
                        _pile_weight = float(_pw)
            ws.cell(row=row, column=_sp_weight_col, value=_pile_weight).font = SM_FONT
            ws.cell(row=row, column=_sp_weight_col).alignment = CENTER
            ws.cell(row=row, column=_sp_weight_col).border = THIN_BORDER

        _border_end = _sp_col_end
        for c in range(1, _border_end + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER

        img_path = None
        normalized_code = normalize_lookup_code(product_code)
        if product_code in code_to_images and code_to_images[product_code]:
            img_path = code_to_images[product_code][0]
        elif normalized_code in code_to_images and code_to_images[normalized_code]:
            img_path = code_to_images[normalized_code][0]

        if not img_path and price_mapping:
            pi = resolve_price_info(price_mapping, product_code)
            if pi and pi.get('image_bytes'):
                img_bytes = pi['image_bytes']
                img_ext = pi.get('image_ext', '.png')
                if image_temp_dir:
                    db_img_name = f"db_{normalized_code or product_code}{img_ext}"
                    db_img_path = os.path.join(image_temp_dir, db_img_name)
                    if db_img_path not in image_cache:
                        with open(db_img_path, 'wb') as _f:
                            _f.write(img_bytes)
                        image_cache.pop(db_img_path, None)
                    img_path = db_img_path

        if img_path:
            fit_w, fit_h = fit_image(ws, row, img_col, img_w, img_h)
            temp_img_path = prepare_image_for_excel(
                img_path,
                target_width=fit_w,
                target_height=fit_h,
                temp_dir=image_temp_dir,
                cache=image_cache,
            )
            final_img_path = temp_img_path if temp_img_path else img_path
            success = add_image_centered_in_cell(
                ws, final_img_path, row, img_col,
                img_width=fit_w, img_height=fit_h,
            )
            if not success:
                ws.cell(row=row, column=5, value='/').alignment = CENTER
        else:
            ws.cell(row=row, column=5, value='/').alignment = CENTER

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(top=0.75, bottom=0.75, left=0.25, right=0.25, header=0.3, footer=0.3)

    return ws.title
