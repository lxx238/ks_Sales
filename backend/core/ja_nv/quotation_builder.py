import os
import re
import shutil
import uuid
from decimal import Decimal

from openpyxl import Workbook

from backend.excel.reader import excel_file_compat
from backend.core.shared.bom_utils import (
    discover_sheet_bom_starts, extract_bom_dataframe, get_bom_processing_rules,
    normalize_selected_bom_keys, quick_scan_bom_sheets, parse_array_to_rows_cols,
    read_bom_from_dataframe,
)
from backend.core.shared.image_utils import scan_images, find_latest_image_log, load_image_mapping_from_log
from backend.core.shared.product_utils import _match_exclude_group
from backend.core.shared.price_utils import resolve_price_info, has_valid_price_info, get_temp_adjusted_base_price
from backend.core.inquiry_builder import create_inquiry_sheet
from backend.core.shared.bom_zip_parser import _build_products_by_key
from backend.core.shared.product_utils import _split_pile_products, _is_valid_product_code, normalize_preinstall
from backend.core.shared.weight_utils import extract_length_from_spec
from backend.core.array_matcher import (
    find_matching_matrix_array,
    build_bom_matrix_data,
    merge_bom_products,
    find_accumulated_match,
    find_info_accumulated_match,
)
from backend.core.shared.sheet_utils import reorder_sheets_by_matrix_array, set_page_break_preview
from backend.core.ja_nv.quotation_engine import (
    create_nv_detail_sheet,
    create_nv_summary_sheet,
    create_spare_parts_sheet,
)
from backend.services.translate_service import translate_notes_in_details


def split_and_create_quotations(
        input_file,
        price_file_path,
        output_dir=None,
        image_path=None,
        image_folder=None,
        price_mapping_override=None,
        contact_info=None,
        matrix_data=None,
        return_details=False,
        selected_bom_keys=None,
        group=None,
        pre_parsed_products=None,
        pre_parsed_bom_info=None,
        pre_parsed_bom_configs=None,
        exclude_options=None,
        exclude_delete_options=None,
        pre_parsed_inverter_products=None,
        fence_data=None,
        need_weight_code=False,
        need_weight=False,
        need_code=False,
        coating_thickness=10,
        nv_params=None,
        nv_fence_gate_data=None,
        case_type=None,
        exchange_rate=None,
        tariff_rate=None,
        consumption_tax=None,
        fence_tax=None,
        discount_rate=None,
        truck_desc=None,
        truck_fee=None,
        need_total_qty=False,
):
    nv_params = nv_params or {}

    if exchange_rate is not None:
        nv_params['exchange_rate'] = exchange_rate

    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(input_file), 'quotation_output')
    os.makedirs(output_dir, exist_ok=True)

    code_to_images = {}
    image_cache = {}
    image_temp_dir = None

    log_search_dirs = [output_dir, os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")]
    latest_log = find_latest_image_log(log_search_dirs)
    if latest_log:
        log_mapping = load_image_mapping_from_log(latest_log)
        if log_mapping:
            code_to_images.update(log_mapping)

    if image_folder and os.path.exists(image_folder):
        scanned = scan_images(image_folder)
        for code, paths in scanned.items():
            if code not in code_to_images:
                code_to_images[code] = paths
        if code_to_images:
            image_temp_dir = os.path.join(output_dir, f"temp_images_{uuid.uuid4().hex}")
            os.makedirs(image_temp_dir, exist_ok=True)
    elif image_folder:
        image_folder = None

    if price_mapping_override is not None:
        price_mapping = price_mapping_override
    else:
        from backend.core.shared.price_utils import load_price_mapping
        price_mapping = load_price_mapping(price_file_path)

    selected_key_set = normalize_selected_bom_keys(selected_bom_keys)
    column_mapping, skip_keywords, non_bom_sheet_keywords = get_bom_processing_rules()

    arrays = (matrix_data or {}).get('arrays') or []
    all_detail_results = []
    all_unmatched_products = []
    pile_products_all = []
    matrix_array_entries = arrays or []
    has_explicit_matrix_arrays = bool(matrix_array_entries)
    used_matrix_array_indices = set()
    pending_boms = []
    accumulated_results = []
    _pile_include_dz = bool(nv_params and nv_params.get('pile_include_dz'))
    info_acc_results = []

    all_excluded_products = []
    all_spare_products = []

    def _filter_excluded(products):
        _delete_merged = {}
        _spare_merged = {}
        if exclude_delete_options:
            for k, v in exclude_delete_options.items():
                if v:
                    _delete_merged[k] = v
        if exclude_options:
            for k, v in exclude_options.items():
                if v:
                    _spare_merged[k] = v
        _all_merged = {}
        _all_merged.update(_delete_merged)
        _all_merged.update(_spare_merged)
        print(f'[DEBUG_CAP] exclude_delete_options={exclude_delete_options} exclude_options={exclude_options}')
        _dg_in_input = [p for p in products if str(p.get('code', '') or '').strip().upper().startswith('DG-')]
        print(f'[DEBUG_CAP] DG- products in this BOM: {len(_dg_in_input)} -> {[(p.get("code"), p.get("quantity")) for p in _dg_in_input]}')
        if not _all_merged or not any(_all_merged.values()):
            print(f'[DEBUG_CAP] no exclude options active -> all products pass (cap will NOT be deleted here)')
            return products, []
        filtered = []
        excluded = []
        for p in products:
            if _match_exclude_group(p, price_mapping, _delete_merged, prefix_only=True):
                print(f'[DEBUG_CAP] DELETED by delete_option: {p.get("code")}')
                continue
            elif _match_exclude_group(p, price_mapping, _spare_merged, prefix_only=True):
                excluded.append(p)
                filtered.append(p)
            else:
                filtered.append(p)
        return filtered, excluded

    if pre_parsed_products and pre_parsed_bom_info:
        products_by_key = pre_parsed_products
        bom_info_by_key = pre_parsed_bom_info
        inverter_products_all = list(pre_parsed_inverter_products) if pre_parsed_inverter_products else []
    else:
        products_by_key, bom_info_by_key, _, inverter_products_all, _ = _build_products_by_key(
            input_file, selected_key_set, column_mapping, skip_keywords, non_bom_sheet_keywords
        )

    master_wb = Workbook()
    default_sheet = master_wb.active
    master_wb.remove(default_sheet)

    # 布板方式（BOM 的 layout）为工程权威来源，先于明细生成覆盖信息表朝向（横置き/縦置き）
    if isinstance(matrix_data, dict):
        for _bi in bom_info_by_key.values():
            _bl = str((_bi.get('config') or {}).get('layout') or '').strip()
            if _bl in ('横放', '横'):
                matrix_data['layout'] = '横置き'; break
            if _bl in ('竖放', '纵放', '竖', '纵'):
                matrix_data['layout'] = '縦置き'; break

    if not products_by_key and arrays:
        if return_details:
            return {
                'output_file': None,
                'inquiry_file': None,
                'unmatched_count': 0,
                'unmatched_products': [],
                'quotation_product_codes': set(),
                'missing_ja_list': [],
                'missing_image_codes': [],
            }
        return None

    for key, products in products_by_key.items():
        bom_info = bom_info_by_key.get(key)
        if not bom_info:
            continue

        config = bom_info.get('config', {})
        bom_array_str = config.get('array', '')
        bom_rows, bom_cols = parse_array_to_rows_cols(bom_array_str)

        bom_base = config.get('base_count', 0) or 0
        bom_miss = config.get('missing_boards', 0) or 0
        bom_has_inv = any(
            (p.get('quantity') or 0) > 0 and p.get('bom_key') == key
            for p in inverter_products_all
        )
        matrix_has_inv_info = any(me.get('has_inverter') for me in matrix_array_entries)
        inv_check = (bom_has_inv if has_explicit_matrix_arrays and matrix_has_inv_info else None)
        matched_idx, matched_array = find_matching_matrix_array(
            matrix_array_entries,
            bom_rows,
            bom_cols,
            used_indices=used_matrix_array_indices if has_explicit_matrix_arrays else None,
            bom_missing=bom_miss if bom_miss != 0 else None,
            bom_base_count=bom_base,
            strict_only=has_explicit_matrix_arrays,
            bom_has_inverter=inv_check,
        )
        if matched_array is None:
            if has_explicit_matrix_arrays:
                inv_for_pending = [ip for ip in inverter_products_all if ip.get('bom_key') == key]
                pending_boms.append({
                    'sheet_name': bom_info.get('variant_name', ''),
                    'bom_info': bom_info,
                    'products': products,
                    'rows': bom_rows,
                    'cols': bom_cols,
                    'base_count': bom_base,
                    'missing': bom_miss,
                    'config': config,
                    'bom_key': key,
                    'inverter_products': inv_for_pending,
                })
                for ip in inv_for_pending:
                    if ip in inverter_products_all:
                        inverter_products_all.remove(ip)
            continue
        if has_explicit_matrix_arrays and matched_idx is not None:
            used_matrix_array_indices.add(matched_idx)

        matched_rows = matched_array.get('rows', '')
        matched_cols = matched_array.get('cols', '')
        matched_qty = matched_array.get('table_qty', 1)
        info_missing = matched_array.get('missing_per_table', 0) or 0
        effective_missing = info_missing if info_missing else bom_miss
        sheet_prefix = f'({matched_idx + 1}){matched_rows}x{matched_cols}_{matched_qty}'

        filtered, excluded = _filter_excluded(products)
        if excluded:
            all_excluded_products.extend(excluded)
        if not filtered:
            continue

        inv_prods_for_key = [ip for ip in inverter_products_all if ip.get('bom_key') == key]
        if inv_prods_for_key:
            filtered = filtered + inv_prods_for_key
            for ip in inv_prods_for_key:
                if ip in inverter_products_all:
                    inverter_products_all.remove(ip)

        bracket_prods, pile_prods = _split_pile_products(filtered, price_mapping, include_dz=_pile_include_dz)
        if pile_prods:
            for pp in pile_prods:
                scaled = dict(pp)
                scaled['quantity'] = float(pp.get('quantity', 0)) * matched_qty
                pile_products_all.append(scaled)

        try:
            detail = create_nv_detail_sheet(
                master_wb,
                matched_array,
                bracket_prods if bracket_prods else filtered,
                price_mapping,
                sheet_prefix=sheet_prefix,
                matrix_data=matrix_data,
                unmatched_products_out=all_unmatched_products,
                coating_thickness=coating_thickness,
                nv_params=nv_params,
                pile_products=pile_prods,
                code_to_images=code_to_images,
                image_temp_dir=image_temp_dir,
                image_cache=image_cache,
                angle_override=config.get('angle', '') or None,
                need_weight_code=need_weight_code,
                need_weight=need_weight,
                need_code=need_code,
                missing_boards=effective_missing,
                span_ew_override=config.get('cross_span', '') or None,
            )
            detail['angle'] = config.get('angle', '') or detail.get('angle', '') or (matrix_data or {}).get('angle', '')
            detail['matched_array'] = matched_array
            detail['matrix_data'] = build_bom_matrix_data(matrix_data, matched_array, bom_config=config)
            detail['set_count'] = (detail.get('matrix_data') or {}).get('set_count') or matched_qty or 1
            detail['_matrix_idx'] = matched_idx
            if inv_prods_for_key:
                inv_counts = set(
                    int(p.get('inverter_total_count', 0))
                    for p in inv_prods_for_key
                    if int(p.get('inverter_total_count', 0)) > 0
                )
                if inv_counts:
                    detail['inverter_count'] = inv_counts.pop()
            all_detail_results.append(detail)
        except Exception as e:
            import traceback
            print(f"   ❌ NV明細シート生成失敗: {e}")
            traceback.print_exc()

    if pending_boms and has_explicit_matrix_arrays and matrix_array_entries:
        used_sheet_names = set()
        for result in all_detail_results:
            sheet_name = result.get('sheet_name', '')
            if sheet_name:
                used_sheet_names.add(sheet_name)

        accumulated_results = find_accumulated_match(
            matrix_array_entries, pending_boms,
            used_matrix_array_indices, used_sheet_names,
        )

        for acc in accumulated_results:
            matched_idx = acc['matrix_idx']
            matrix_entry = acc['matrix_idx']
            matrix_entry_data = acc['matrix_entry']
            selected_boms = acc.get('selected_boms') or []
            used_matrix_array_indices.add(matched_idx)

            _r, _c = matrix_entry_data.get('rows'), matrix_entry_data.get('cols')
            _total_base = matrix_entry_data.get('table_qty', 1)
            _accum_group_id = f"{_r}x{_c}_{_total_base}"

            selected_boms.sort(key=lambda sb: sb.get('base_count', 0) or 0)

            for _sub_idx, _sb in enumerate(selected_boms):
                _sb_base = _sb.get('base_count', 0) or 1
                _sheet_prefix = f'({matched_idx + 1}_{_sub_idx + 1}){_r}x{_c}_{_sb_base}'
                _sub_array_info = dict(matrix_entry_data)
                _sub_array_info['table_qty'] = _sb_base

                _inv_prods_sub = _sb.get('inverter_products') or []
                _bom_products_sub = _sb.get('products') or []
                _filtered_sub, _excluded_sub = _filter_excluded(_bom_products_sub)
                if _excluded_sub:
                    all_excluded_products.extend(_excluded_sub)
                if not _filtered_sub and not _inv_prods_sub:
                    continue

                _bom_config_sub = _sb.get('config') or {}
                _effective_matrix_data = build_bom_matrix_data(matrix_data, _sub_array_info, bom_config=_bom_config_sub)

                _bracket_prods_sub, _pile_prods_sub = _split_pile_products(_filtered_sub, price_mapping, include_dz=_pile_include_dz)
                if _inv_prods_sub:
                    _bracket_prods_sub = (_bracket_prods_sub or []) + _inv_prods_sub
                if _pile_prods_sub:
                    for _pp in _pile_prods_sub:
                        _scaled = dict(_pp)
                        _scaled['quantity'] = float(_pp.get('quantity', 0)) * _sb_base
                        pile_products_all.append(_scaled)

                try:
                    detail = create_nv_detail_sheet(
                        master_wb,
                        _sub_array_info,
                        _bracket_prods_sub if _bracket_prods_sub else _filtered_sub,
                        price_mapping,
                        sheet_prefix=_sheet_prefix,
                        matrix_data=_effective_matrix_data,
                        unmatched_products_out=all_unmatched_products,
                        coating_thickness=coating_thickness,
                        nv_params=nv_params,
                        pile_products=_pile_prods_sub,
                        code_to_images=code_to_images,
                        image_temp_dir=image_temp_dir,
                        image_cache=image_cache,
                        angle_override=_bom_config_sub.get('angle', '') or None,
                        need_weight_code=need_weight_code,
                        need_weight=need_weight,
                        need_code=need_code,
                        missing_boards=_sub_array_info.get('missing_per_table', 0) or _bom_config_sub.get('missing_boards', 0) or 0,
                        span_ew_override=_bom_config_sub.get('cross_span', '') or None,
                    )
                    detail['config'] = _bom_config_sub
                    detail['bom_key'] = _sb.get('bom_key', '') or 'accumulated'
                    detail['angle'] = _bom_config_sub.get('angle', '') or (matrix_data or {}).get('angle', '')
                    detail['matched_array'] = _sub_array_info
                    detail['matrix_data'] = _effective_matrix_data
                    detail['set_count'] = _sb_base
                    detail['_matrix_idx'] = matched_idx
                    detail['accumulated_group_id'] = _accum_group_id
                    detail['accumulated_sub_idx'] = _sub_idx
                    if _inv_prods_sub:
                        _inv_counts = set(
                            int(p.get('inverter_total_count', 0))
                            for p in _inv_prods_sub
                            if int(p.get('inverter_total_count', 0)) > 0
                        )
                        if _inv_counts:
                            detail['inverter_count'] = _inv_counts.pop()
                    all_detail_results.append(detail)
                    print(f"   ✅ NV detail (accumulated): {detail['sheet_name']} "
                          f"({_r}x{_c}_{_sb_base}, group={_accum_group_id})")
                except Exception as e:
                    import traceback
                    print(f"   ❌ NV累加明細生成失敗: {e}")
                    traceback.print_exc()

    if pending_boms and has_explicit_matrix_arrays and matrix_array_entries:
        used_bom_indices_in_acc = set()
        for acc in accumulated_results:
            for selected_bom in acc.get('selected_boms', []):
                for index, pending in enumerate(pending_boms):
                    if pending is selected_bom:
                        used_bom_indices_in_acc.add(index)
                        break

        info_acc_results = find_info_accumulated_match(
            matrix_array_entries, pending_boms,
            used_matrix_array_indices,
            used_bom_indices=used_bom_indices_in_acc,
        )

        for acc in info_acc_results:
            matrix_indices = acc['matrix_indices']
            matched_bom = acc['matched_bom']
            total_info_qty = acc['total_info_qty']

            for matched_idx in matrix_indices:
                used_matrix_array_indices.add(matched_idx)

            bom_products = matched_bom.get('products') or []
            inv_prods_info = matched_bom.get('inverter_products') or []
            filtered_info, excluded_info = _filter_excluded(bom_products)
            if excluded_info:
                all_excluded_products.extend(excluded_info)
            if not filtered_info:
                continue
            bom_config = matched_bom.get('config') or matched_bom
            bom_base = matched_bom.get('base_count', 0) or 0
            first_entry = acc['matrix_entries'][0]
            matched_rows = first_entry.get('rows', '')
            matched_cols = first_entry.get('cols', '')
            sheet_prefix = f'({matrix_indices[0] + 1}){matched_rows}x{matched_cols}_{bom_base or total_info_qty}'

            accumulated_entry = dict(first_entry)
            accumulated_entry['table_qty'] = bom_base or total_info_qty
            effective_matrix_data = build_bom_matrix_data(matrix_data, accumulated_entry, bom_config=bom_config)

            bracket_prods_info, pile_prods_info = _split_pile_products(filtered_info, price_mapping, include_dz=_pile_include_dz)
            if inv_prods_info:
                bracket_prods_info = (bracket_prods_info or []) + inv_prods_info
            info_qty = bom_base or total_info_qty
            if pile_prods_info:
                for pp in pile_prods_info:
                    scaled = dict(pp)
                    scaled['quantity'] = float(pp.get('quantity', 0)) * info_qty
                    pile_products_all.append(scaled)

            try:
                detail = create_nv_detail_sheet(
                    master_wb,
                    accumulated_entry,
                    bracket_prods_info if bracket_prods_info else filtered_info,
                    price_mapping,
                    sheet_prefix=sheet_prefix,
                    matrix_data=effective_matrix_data,
                    unmatched_products_out=all_unmatched_products,
                    coating_thickness=coating_thickness,
                    nv_params=nv_params,
                    pile_products=pile_prods_info,
                    code_to_images=code_to_images,
                    image_temp_dir=image_temp_dir,
                    image_cache=image_cache,
                    angle_override=bom_config.get('angle', '') or None,
                    need_weight_code=need_weight_code,
                    need_weight=need_weight,
                    need_code=need_code,
                    missing_boards=first_entry.get('missing_per_table', 0) or bom_config.get('missing_boards', 0) or 0,
                    span_ew_override=bom_config.get('cross_span', '') or None,
                )
                detail['config'] = bom_config
                detail['bom_key'] = 'info_accumulated'
                detail['angle'] = bom_config.get('angle', '') or (matrix_data or {}).get('angle', '')
                detail['matched_array'] = first_entry
                detail['matrix_data'] = effective_matrix_data
                detail['set_count'] = effective_matrix_data.get('set_count') or 1
                detail['_matrix_idx'] = matrix_indices[0]
                all_detail_results.append(detail)
            except Exception as e:
                import traceback
                print(f"   ❌ NV信息表累加明細生成失敗: {e}")
                traceback.print_exc()

    if pending_boms and has_explicit_matrix_arrays and matrix_array_entries:
        used_pending_indices = set()
        for acc in accumulated_results:
            for selected_bom in acc.get('selected_boms', []):
                for index, pending in enumerate(pending_boms):
                    if pending is selected_bom:
                        used_pending_indices.add(index)
        for acc in info_acc_results:
            matched_bom = acc.get('matched_bom')
            if matched_bom:
                for index, pending in enumerate(pending_boms):
                    if pending is matched_bom:
                        used_pending_indices.add(index)

        remaining_info_indices = [
            index for index in range(len(matrix_array_entries))
            if index not in used_matrix_array_indices
        ]

        for matrix_index in remaining_info_indices:
            info_entry = matrix_array_entries[matrix_index]
            info_rows = info_entry.get('rows')
            info_cols = info_entry.get('cols')
            info_qty = info_entry.get('table_qty', 1)
            if info_rows is None or info_cols is None:
                continue

            matched_pending = None
            matched_pending_index = None
            for index, pending in enumerate(pending_boms):
                if index in used_pending_indices:
                    continue
                if pending.get('rows') == info_rows and pending.get('cols') == info_cols:
                    matched_pending = pending
                    matched_pending_index = index
                    break

            if matched_pending is None:
                continue

            used_pending_indices.add(matched_pending_index)
            used_matrix_array_indices.add(matrix_index)

            bom_products = matched_pending.get('products') or []
            inv_prods_fb = matched_pending.get('inverter_products') or []
            filtered_fb, excluded_fb = _filter_excluded(bom_products)
            if excluded_fb:
                all_excluded_products.extend(excluded_fb)
            if not filtered_fb:
                continue
            bom_config = matched_pending.get('config') or {}
            sheet_prefix = f'({matrix_index + 1}){info_rows}x{info_cols}_{info_qty}'
            effective_matrix_data = build_bom_matrix_data(matrix_data, info_entry, bom_config=bom_config)

            bracket_prods_fb, pile_prods_fb = _split_pile_products(filtered_fb, price_mapping, include_dz=_pile_include_dz)
            if inv_prods_fb:
                bracket_prods_fb = (bracket_prods_fb or []) + inv_prods_fb
            if pile_prods_fb:
                for pp in pile_prods_fb:
                    scaled = dict(pp)
                    scaled['quantity'] = float(pp.get('quantity', 0)) * info_qty
                    pile_products_all.append(scaled)

            try:
                detail = create_nv_detail_sheet(
                    master_wb,
                    info_entry,
                    bracket_prods_fb if bracket_prods_fb else filtered_fb,
                    price_mapping,
                    sheet_prefix=sheet_prefix,
                    matrix_data=effective_matrix_data,
                    unmatched_products_out=all_unmatched_products,
                    coating_thickness=coating_thickness,
                    nv_params=nv_params,
                    pile_products=pile_prods_fb,
                    code_to_images=code_to_images,
                    image_temp_dir=image_temp_dir,
                    image_cache=image_cache,
                    angle_override=bom_config.get('angle', '') or None,
                    need_weight_code=need_weight_code,
                    need_weight=need_weight,
                    need_code=need_code,
                    missing_boards=info_entry.get('missing_per_table', 0) or bom_config.get('missing_boards', 0) or 0,
                    span_ew_override=bom_config.get('cross_span', '') or None,
                )
                detail['config'] = bom_config
                detail['bom_key'] = 'fallback_rows_cols'
                detail['angle'] = bom_config.get('angle', '') or (matrix_data or {}).get('angle', '')
                detail['matched_array'] = info_entry
                detail['matrix_data'] = effective_matrix_data
                detail['set_count'] = effective_matrix_data.get('set_count') or 1
                detail['_matrix_idx'] = matrix_index
                all_detail_results.append(detail)
            except Exception as e:
                import traceback
                print(f"   ❌ NV同行列兜底明細生成失敗: {e}")
                traceback.print_exc()

    if not all_detail_results and has_explicit_matrix_arrays and matrix_array_entries and pending_boms:
        fallback_pending_indices = set()
        for matrix_index, info_entry in enumerate(matrix_array_entries):
            if matrix_index in used_matrix_array_indices:
                continue
            if matrix_index >= len(pending_boms):
                break
            pending = pending_boms[matrix_index]
            if pending_boms.index(pending) in fallback_pending_indices:
                continue

            fallback_pending_indices.add(pending_boms.index(pending))
            bom_products = pending.get('products') or []
            inv_prods_ff = pending.get('inverter_products') or []
            bom_config = pending.get('config') or {}
            if not bom_products:
                continue

            filtered_ff, excluded_ff = _filter_excluded(bom_products)
            if excluded_ff:
                all_excluded_products.extend(excluded_ff)
            if not filtered_ff:
                continue

            info_rows = info_entry.get('rows')
            info_cols = info_entry.get('cols')
            info_qty = info_entry.get('table_qty', 1)
            sheet_prefix = f'({matrix_index + 1}){info_rows}x{info_cols}_{info_qty}'
            effective_matrix_data = build_bom_matrix_data(matrix_data, info_entry, bom_config=bom_config)

            bracket_prods_ff, pile_prods_ff = _split_pile_products(filtered_ff, price_mapping, include_dz=_pile_include_dz)
            if inv_prods_ff:
                bracket_prods_ff = (bracket_prods_ff or []) + inv_prods_ff
            if pile_prods_ff:
                for pp in pile_prods_ff:
                    scaled = dict(pp)
                    scaled['quantity'] = float(pp.get('quantity', 0)) * info_qty
                    pile_products_all.append(scaled)

            try:
                detail = create_nv_detail_sheet(
                    master_wb,
                    info_entry,
                    bracket_prods_ff if bracket_prods_ff else filtered_ff,
                    price_mapping,
                    sheet_prefix=sheet_prefix,
                    matrix_data=effective_matrix_data,
                    unmatched_products_out=all_unmatched_products,
                    coating_thickness=coating_thickness,
                    nv_params=nv_params,
                    pile_products=pile_prods_ff,
                    code_to_images=code_to_images,
                    image_temp_dir=image_temp_dir,
                    image_cache=image_cache,
                    angle_override=bom_config.get('angle', '') or None,
                    need_weight_code=need_weight_code,
                    need_weight=need_weight,
                    need_code=need_code,
                    missing_boards=info_entry.get('missing_per_table', 0) or bom_config.get('missing_boards', 0) or 0,
                    span_ew_override=bom_config.get('cross_span', '') or None,
                )
                detail['config'] = bom_config
                detail['bom_key'] = f'forced_fallback_{matrix_index}'
                detail['angle'] = bom_config.get('angle', '') or (matrix_data or {}).get('angle', '')
                detail['matched_array'] = info_entry
                detail['matrix_data'] = effective_matrix_data
                detail['set_count'] = effective_matrix_data.get('set_count') or 1
                detail['_matrix_idx'] = matrix_index
                all_detail_results.append(detail)
            except Exception as e:
                import traceback
                print(f"   ❌ NV强制兜底明細生成失敗: {e}")
                traceback.print_exc()

    inverter_detail_results = []
    standalone_inv_entries = [
        arr for arr in matrix_array_entries
        if arr.get('is_standalone_inverter')
    ]
    if standalone_inv_entries and inverter_products_all:
        from collections import OrderedDict
        inv_bom_groups = OrderedDict()
        for p in inverter_products_all:
            bk = p.get('bom_key', '') or '__no_key__'
            inv_bom_groups.setdefault(bk, []).append(p)

        for info_entry in standalone_inv_entries:
            info_inv_count = info_entry.get('inverter_count', 1)
            info_base = info_entry.get('table_qty', 1)
            matched_inv_prods = []
            matched_bk = None
            for bk, prods in inv_bom_groups.items():
                for p in prods:
                    p_count = p.get('inverter_total_count', 0)
                    p_base = p.get('inverter_base_count', 0)
                    if int(p_count) == int(info_inv_count):
                        if p_base > 0 and int(p_base) == int(info_base):
                            matched_inv_prods = prods
                            matched_bk = bk
                            break
                        elif p_base == 0 and not matched_inv_prods:
                            matched_inv_prods = prods
                            matched_bk = bk
                if matched_inv_prods and matched_bk == bk:
                    break

            if not matched_inv_prods:
                for bk, prods in inv_bom_groups.items():
                    if any((p.get('quantity') or 0) > 0 for p in prods):
                        p_count = int(prods[0].get('inverter_total_count', 0))
                        if p_count == int(info_inv_count):
                            matched_inv_prods = prods
                            matched_bk = bk
                            break

            if not matched_inv_prods:
                continue

            inv_bom_groups.pop(matched_bk, None)
            inv_filtered, inv_excluded = _filter_excluded(matched_inv_prods)
            if inv_excluded:
                all_excluded_products.extend(inv_excluded)
            if not inv_filtered:
                continue

            inv_array_info = {
                'no': info_entry.get('no', '0'),
                'rows': '',
                'cols': '',
                'table_qty': info_base,
                'note': info_entry.get('note', ''),
            }
            inv_counts = set(
                int(p.get('inverter_total_count', 0))
                for p in inv_filtered
                if int(p.get('inverter_total_count', 0)) > 0
            )
            total_inv_qty = inv_counts.pop() if inv_counts else info_inv_count
            sheet_prefix = f"パワコン独立架台_{info_base}"

            try:
                inv_detail = create_nv_detail_sheet(
                    master_wb,
                    inv_array_info,
                    inv_filtered,
                    price_mapping,
                    sheet_prefix=sheet_prefix,
                    matrix_data=matrix_data,
                    unmatched_products_out=all_unmatched_products,
                    coating_thickness=coating_thickness,
                    nv_params=nv_params,
                    code_to_images=code_to_images,
                    image_temp_dir=image_temp_dir,
                    image_cache=image_cache,
                    is_inverter=True,
                )
                if inv_detail:
                    inv_detail['bom_key'] = matched_bk
                    inv_detail['inverter_count'] = total_inv_qty
                    inv_detail['inv_note'] = 'パワコン独立架台'
                    inv_detail['is_standalone_inverter'] = True
                    inv_detail['array_info'] = inv_array_info
                    inverter_detail_results.append(inv_detail)
                    print(f"   ⚡ パワコン明細シート生成: {inv_detail.get('sheet_name')} ({matched_bk})")
            except Exception as e:
                import traceback
                print(f"   ❌ パワコン明細シート生成失敗 ({matched_bk}): {e}")
                traceback.print_exc()

    for detail in all_detail_results:
        dk = detail.get('bom_key', '')
        if not dk:
            continue
        inv_count = detail.get('inverter_count', 0)
        if inv_count and int(inv_count) > 0 and not detail.get('is_standalone_inverter'):
            detail['inv_note'] = f'{int(inv_count)}台ＰＣＳ'
        matched_arr = detail.get('matched_array') or detail.get('array_info') or {}
        arr_note = str(matched_arr.get('note') or '').strip()
        if arr_note:
            if not detail.get('inv_note'):
                nums = re.findall(r'(\d+)', arr_note)
                if nums:
                    total_pcs = sum(int(n) for n in nums)
                    detail['inv_note'] = f'{total_pcs}ＰＣＳ'

    if all_detail_results and matrix_array_entries:
        def _sort_key(detail):
            explicit_pos = detail.get('_matrix_idx')
            if explicit_pos is not None:
                return explicit_pos
            _array_order = []
            for arr in matrix_array_entries:
                _r = str(arr.get('rows', '') or '')
                _c = str(arr.get('cols', '') or '')
                _q = int(arr.get('table_qty', 1) or 1)
                _array_order.append((_r, _c, _q))
            ai = detail.get('array_info', {})
            r = str(ai.get('rows', '') or '')
            c = str(ai.get('cols', '') or '')
            q = int(ai.get('table_qty', 1) or 1)
            try:
                return _array_order.index((r, c, q))
            except ValueError:
                return len(_array_order)
        all_detail_results.sort(key=_sort_key)

    if all_detail_results and matrix_array_entries:
        reorder_sheets_by_matrix_array(master_wb, all_detail_results, matrix_array_entries, log_prefix='[NV]')

    translate_notes_in_details(all_detail_results)
    translate_notes_in_details(inverter_detail_results)

    if all_detail_results:
        pile_summary = None
        if pile_products_all:
            import re as _re
            pile_filtered, pile_excluded = _filter_excluded(pile_products_all)
            if pile_excluded:
                all_excluded_products.extend(pile_excluded)
            if pile_filtered:
                pile_filtered = [pp for pp in pile_filtered if float(pp.get('quantity', 0) or 0) > 0]
            if pile_filtered:
                _pq = 0
                _pa = 0.0
                for _pp in pile_filtered:
                    _pp_qty = float(_pp.get('quantity', 0))
                    _pp_code = str(_pp.get('code', '') or '').strip()
                    _pp_pi = resolve_price_info(price_mapping, _pp_code, spec=_pp.get('spec', '')) if price_mapping else None
                    if not _pp_pi or not _pp_pi.get('price'):
                        _norm = _re.sub(r'[-_\s]', '', _pp_code.upper())
                        for _key, _val in (price_mapping or {}).items():
                            if _re.sub(r'[-_\s]', '', str(_key).strip().upper()) == _norm:
                                if _val.get('price'):
                                    _pp_pi = _val
                                    print(f"   🔧 builder杭フォールバック: {_pp_code} → {_key}")
                                    break
                    _pp_price = get_temp_adjusted_base_price(_pp_pi, _pp, group or '日语组', 'export') if _pp_pi and _pp_pi.get('price') else 0
                    _pp_unit = (_pp_pi.get('unit', '') if _pp_pi else '') or ''
                    if _pp_price == 0 and _pp_qty > 0 and _is_valid_product_code(_pp_code):
                        all_unmatched_products.append({
                            'code': _pp_code,
                            'name': _pp.get('name', ''),
                            'spec': _pp.get('spec', ''),
                            'material': _pp.get('material', ''),
                            'quantity': _pp_qty,
                            'weight': _pp.get('weight', 0),
                            'preinstall': normalize_preinstall(_pp.get('preinstall')),
                        })
                        print(f"   ⚠ 杭価格未検出（询价追加）: code={_pp_code}, spec={_pp.get('spec', '')}")
                    if _pp_unit in ('米', 'm', 'M', 'meter'):
                        _len = float(_pp.get('length', 0) or 0)
                        if _len <= 0:
                            _len = float(extract_length_from_spec(_pp.get('spec', '')) or 0)
                        if _len <= 0:
                            _m = _re.search(r'(\d+\.?\d*)', _pp.get('spec', ''))
                            _len = float(_m.group(1)) if _m else 0
                        if _len > 0:
                            _pp_price = _pp_price * _len / 1000
                    print(f"   🔍 builder杭: code={_pp_code}, price={_pp_price}, unit={_pp_unit}, qty={_pp_qty}")
                    _pq += _pp_qty
                    _pa += _pp_price * _pp_qty
                pile_summary = {'total_qty': _pq, 'total_price': round(_pa, 2)}

        _spare_count = len(all_excluded_products)
        _pile_spare_count = sum(1 for sp in all_excluded_products if str(sp.get('code', '') or '').strip().upper().startswith('DZ-'))
        try:
            mc = nv_params.get('mitsumori_condition', 'CIF')
            if mc == 'CIF_DDP':
                cif_params = dict(nv_params)
                cif_params['mitsumori_condition'] = 'CIF'
                create_nv_summary_sheet(
                    master_wb, all_detail_results,
                    matrix_data=matrix_data, fence_data=fence_data,
                    nv_params=cif_params, nv_fence_gate_data=nv_fence_gate_data,
                    pile_summary=pile_summary,
                    image_temp_dir=image_temp_dir, image_cache=image_cache,
                    sheet_title='CIF合計',
                    spare_parts_count=_spare_count,
                    dz_spare_count=_pile_spare_count,
                    inverter_detail=inverter_detail_results if inverter_detail_results else None,
                )
                ddp_params = dict(nv_params)
                ddp_params['mitsumori_condition'] = 'DDP'
                create_nv_summary_sheet(
                    master_wb, all_detail_results,
                    matrix_data=matrix_data, fence_data=fence_data,
                    nv_params=ddp_params, nv_fence_gate_data=nv_fence_gate_data,
                    pile_summary=pile_summary,
                    image_temp_dir=image_temp_dir, image_cache=image_cache,
                    sheet_title='DDP合計',
                    spare_parts_count=_spare_count,
                    dz_spare_count=_pile_spare_count,
                    inverter_detail=inverter_detail_results if inverter_detail_results else None,
                )
            else:
                create_nv_summary_sheet(
                    master_wb,
                    all_detail_results,
                    matrix_data=matrix_data,
                    fence_data=fence_data,
                    nv_params=nv_params,
                    nv_fence_gate_data=nv_fence_gate_data,
                    pile_summary=pile_summary,
                    image_temp_dir=image_temp_dir,
                    image_cache=image_cache,
                    spare_parts_count=_spare_count,
                    dz_spare_count=_pile_spare_count,
                    inverter_detail=inverter_detail_results if inverter_detail_results else None,
                )
        except Exception as e:
            import traceback
            print(f"   ❌ NV合計シート生成失敗: {e}")
            traceback.print_exc()
        mc = nv_params.get('mitsumori_condition', 'CIF')
        if mc == 'CIF_DDP':
            for sn in ('DDP合計', 'CIF合計'):
                if sn in master_wb.sheetnames:
                    master_wb.move_sheet(sn, offset=-master_wb.sheetnames.index(sn))
        elif '合計' in master_wb.sheetnames:
            master_wb.move_sheet('合計', offset=-master_wb.sheetnames.index('合計'))

        if all_excluded_products:
            try:
                import re as _re2
                from collections import defaultdict as _dd
                def _extract_spec_num_for_spare(spec):
                    m = _re2.search(r'(\d+(?:\.\d+)?)', str(spec))
                    return float(m.group(1)) if m else 0
                code_groups = _dd(list)
                for sp in all_excluded_products:
                    code = str(sp.get('code', '') or '').strip().upper()
                    code_groups[code].append(sp)
                spare_products_filtered = []
                for code, prods in code_groups.items():
                    total_qty = sum(float(p.get('quantity', 0) or 0) for p in prods)
                    best = max(prods, key=lambda p: _extract_spec_num_for_spare(p.get('spec', '')))
                    selected = dict(best)
                    new_qty = max(1, round(total_qty * 0.01))
                    selected['quantity'] = new_qty
                    spare_products_filtered.append(selected)
                create_spare_parts_sheet(
                    master_wb,
                    spare_products_filtered,
                    price_mapping,
                    coating_thickness=coating_thickness,
                    need_weight_code=need_weight_code,
                    need_weight=need_weight,
                    need_code=need_code,
                    code_to_images=code_to_images,
                    image_temp_dir=image_temp_dir,
                    image_cache=image_cache,
                )
                print(f"   📦 無料予備品シート生成: {len(all_excluded_products)} 件")
            except Exception as e:
                import traceback
                print(f"   ❌ 無料予備品シート生成失敗: {e}")
                traceback.print_exc()

        input_basename = os.path.splitext(os.path.basename(input_file))[0]
        output_file = os.path.join(output_dir, f"{input_basename}_NV見積.xlsx")
        set_page_break_preview(master_wb)
        master_wb.save(output_file)

        inquiry_file = None
        if all_unmatched_products:
            try:
                inquiry_wb = Workbook()
                inquiry_wb.remove(inquiry_wb.active)
                requester = ''
                if contact_info and contact_info.get('inquiry_requester'):
                    requester = contact_info['inquiry_requester']
                create_inquiry_sheet(
                    inquiry_wb,
                    all_unmatched_products,
                    'NV',
                    inquiry_requester=requester,
                )
                inquiry_file = os.path.join(output_dir, f"{input_basename}_询价表.xlsx")
                inquiry_wb.save(inquiry_file)
                print(f"   📋 询价表: {inquiry_file}")
            except Exception as e:
                print(f"   ❌ 询价表生成失敗: {e}")
                inquiry_file = None

        if image_temp_dir and os.path.exists(image_temp_dir):
            try:
                pile_tmp = os.path.join(image_temp_dir, 'pile_imgs')
                if os.path.exists(pile_tmp):
                    shutil.rmtree(pile_tmp)
                shutil.rmtree(image_temp_dir)
            except Exception:
                pass

        print(f"\n✅ NV処理完了：{len(all_detail_results)} 件の明細シートを生成")
        print(f"📁 出力ファイル: {output_file}")

        if return_details:
            detail_sheet_products = []
            for prods in products_by_key.values():
                detail_sheet_products.extend(prods)
            missing_ja_list = []
            missing_image_codes = []
            seen_ja = set()
            seen_img = set()
            for p in detail_sheet_products:
                code = str(p.get('code', '') or '').strip()
                if not code:
                    continue
                pi = resolve_price_info(price_mapping, code, spec=p.get('spec', ''))
                norm = code.upper().replace(' ', '')
                if norm not in seen_ja:
                    seen_ja.add(norm)
                    if pi and not pi.get('name_ja', '').strip():
                        missing_ja_list.append({
                            'code': pi.get('db_code', code),
                            'name': pi.get('name', '') or p.get('name', ''),
                        })
                if code not in seen_img:
                    seen_img.add(code)
                    if not pi or pi.get('image_status') in ('missing', 'invalid'):
                        missing_image_codes.append(code)
            quotation_product_codes = set()
            for p in detail_sheet_products:
                c = str(p.get('code', '') or '').strip()
                if c:
                    quotation_product_codes.add(c)
            return {
                'output_file': output_file,
                'inquiry_file': inquiry_file,
                'unmatched_count': len(all_unmatched_products),
                'unmatched_products': all_unmatched_products,
                'quotation_product_codes': quotation_product_codes,
                'missing_ja_list': missing_ja_list,
                'missing_image_codes': missing_image_codes,
            }
        return output_file
    else:
        print("\n❌ 明細シートが生成されませんでした")
        if image_temp_dir and os.path.exists(image_temp_dir):
            try:
                pile_tmp = os.path.join(image_temp_dir, 'pile_imgs')
                if os.path.exists(pile_tmp):
                    shutil.rmtree(pile_tmp)
                shutil.rmtree(image_temp_dir)
            except Exception:
                pass
        return None
