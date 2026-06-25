from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins
from decimal import Decimal
import os
import re

_CJK_RE = re.compile(
    r'[\u2e80-\u9fff\uac00-\ud7af\uff00-\uffef]'
    r'|[\u3400-\u4dbf\U00020000-\U0002a6df]'
)


def _strip_cjk_spec(spec):
    if not spec:
        return spec
    s = str(spec)
    s = re.sub(r'\([^)\u2e80-\u9fff\uac00-\ud7af\uff00-\uffef]*'
               r'[\u2e80-\u9fff\uac00-\ud7af\uff00-\uffef][^)]*\)', '', s)
    s = _CJK_RE.sub('', s)
    return s.strip()

from backend.core.quotation_engine import (
    resolve_price_info,
    has_valid_price_info,
    normalize_lookup_code,
    round_to_2_decimal,
    extract_length_from_spec,
    _is_valid_product_code,
)
from backend.core.material_translate import translate_material

SM_FONT = Font(name='Arial', size=8)
SM_FONT_BOLD = Font(name='Arial', size=8, bold=True)
SM_FONT_BOLD_10 = Font(name='Arial', size=10, bold=True)
SM_FONT_TITLE = Font(name='Arial', size=16, bold=True)
SM_FONT_RED = Font(name='Arial', size=8, bold=True, color='FF0000')
SM_FONT_RED_10 = Font(name='Arial', size=10, bold=True, color='FF0000')
SM_FONT_HEADER = Font(name='Arial', size=8, bold=True)

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT_A = Alignment(horizontal='right', vertical='center', wrap_text=True)
LEFT_A = Alignment(horizontal='left', vertical='center', wrap_text=True)

CURRENCY_FMT = '"US$" #,##0.00'

DEFAULT_NV_DISCOUNT_RATE = Decimal('0.72')
DEFAULT_NV_FENCE_DISCOUNT_RATE = Decimal('0.94')

DETAIL_COL_WIDTHS = {
    'A': 5, 'B': 12, 'C': 18, 'D': 15, 'E': 10, 'F': 10, 'G': 13, 'H': 15,
}

SUMMARY_COL_WIDTHS = {
    'A': 5, 'B': 8, 'C': 8, 'D': 8, 'E': 8, 'F': 8,
    'G': 8, 'H': 8, 'I': 8, 'J': 13, 'K': 13, 'L': 8, 'M': 8,
}

GRAY_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
LIGHT_BLUE_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

_LOGO_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))), 'input', '集团标1.png')
_FENCE_LOGO_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))), 'input', '集团标2.png')


def _insert_logo(ws, cell_ref, width=120, height=40):
    if not os.path.isfile(_LOGO_PATH):
        return
    try:
        img = XLImage(_LOGO_PATH)
        img.width = width
        img.height = height
        ws.add_image(img, cell_ref)
    except Exception:
        pass


def _set(ws, r, c, val, font=SM_FONT, align=CENTER, border=True):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font
    cell.alignment = align
    if border:
        cell.border = THIN_BORDER
    return cell


def create_nv_detail_sheet(workbook, array_info, bom_products, price_mapping,
                           sheet_prefix=None, matrix_data=None,
                           unmatched_products_out=None,
                           coating_thickness=10, nv_params=None,
                           pile_products=None,
                           code_to_images=None, image_temp_dir=None,
                           image_cache=None):
    from backend.core.quotation_engine import (
        prepare_image_for_excel, add_image_centered_in_cell,
    )

    nv_params = nv_params or {}
    matrix_data = matrix_data or {}

    img_w, img_h, img_col, img_padding = 65, 50, 4, 2

    def col_width_to_px(cw):
        return int(cw * 7.5 + 5)

    def row_height_to_px(rh):
        return int(rh * 1.33)

    def fit_image(ws, r, c, mw, mh, pad=img_padding):
        cw_val = ws.column_dimensions[get_column_letter(c)].width or 8.43
        rh_val = ws.row_dimensions[r].height or 15
        aw = col_width_to_px(cw_val) - pad * 2
        ah = row_height_to_px(rh_val) - pad * 2
        aw = max(aw, 20)
        ah = max(ah, 20)
        scale = min(aw / mw, ah / mh, 1.0)
        return int(mw * scale), int(mh * scale)

    if sheet_prefix:
        sheet_name = sheet_prefix
    else:
        sheet_name = str(array_info.get('table_qty', 1))
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    original = sheet_name
    counter = 1
    while sheet_name in workbook.sheetnames:
        sheet_name = f"{original}_{counter}"
        counter += 1

    ws = workbook.create_sheet(title=sheet_name)

    for col, width in DETAIL_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 30
    for row in range(2, 7):
        ws.row_dimensions[row].height = 19
    ws.row_dimensions[7].height = 5
    ws.row_dimensions[8].height = 25

    ws.merge_cells('A2:B4')
    for r in range(2, 5):
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = THIN_BORDER
    if os.path.isfile(_LOGO_PATH):
        try:
            from backend.core.quotation_engine import add_image_centered_in_range

            _logo_img = XLImage(_LOGO_PATH)
            _logo_img.width = 100
            _logo_img.height = 50
            add_image_centered_in_range(ws, _logo_img, start_col=1, start_row=3, end_col=2, end_row=4)
        except Exception:
            _insert_logo(ws, 'A3', width=100, height=50)

    ws.merge_cells('A1:H1')
    ws['A1'] = '各アレイ部材リスト'
    ws['A1'].font = Font(name='Arial', bold=True, size=16)
    ws['A1'].alignment = CENTER
    for c in range(1, 9):
        ws.cell(row=1, column=c).border = THIN_BORDER

    rows_val = array_info.get('rows', '')
    cols_val = array_info.get('cols', '')
    qty_val = array_info.get('table_qty', 1)
    panel_count = int(rows_val) * int(cols_val) if rows_val and cols_val else 0
    module_wattage = matrix_data.get('module_wattage') or 0
    panel_size = matrix_data.get('panel_spec') or matrix_data.get('module_size') or ''
    wind_speed = matrix_data.get('max_wind_speed') or ''
    snow_load = matrix_data.get('max_snow_load') or ''
    angle = matrix_data.get('angle') or ''
    ground_height = nv_params.get('ground_height') or matrix_data.get('ground_height') or ''
    span_ew = nv_params.get('span_ew') or ''
    sales_name = nv_params.get('sales_name') or 'Nanami'
    sales_phone = nv_params.get('sales_phone') or '+86-137-7466-5835'
    sales_fax = nv_params.get('sales_fax') or '0086-592-5738212'
    sales_tel = nv_params.get('sales_tel') or '0086-592-5767152'

    ws['A5'] = '〒361-009 中国厦門湖里区枋湖北二路891号 匯鑫財富大厦11-12階'
    ws['A5'].font = SM_FONT
    ws['A5'].alignment = CENTER
    ws.merge_cells('A5:B6')
    for r in range(5, 7):
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = THIN_BORDER

    _set(ws, 2, 3, f'Sales man:{sales_name}')
    _set(ws, 3, 3, f'Mob:{sales_phone}')
    _set(ws, 4, 3, f'Tel : {sales_tel}')
    _set(ws, 5, 3, f'Fax: {sales_fax}')

    _set(ws, 2, 4, 'パネル設置角度')
    angle_display = angle.rstrip('°').strip() if angle else ''
    _set(ws, 2, 5, f'横置き{angle_display}°' if angle_display else '')
    _set(ws, 2, 6, 'パネルサイズ')
    _set(ws, 2, 7, panel_size)

    ws.merge_cells('H2:H6')
    quality_text = (
        '品質保証:\n10年間品質保証、それに、20年以上使える\n'
        '参考標準:\nJIS C 8955太陽電池アレイ用支持物設計標準'
    )
    _set(ws, 2, 8, quality_text)
    for r in range(2, 7):
        ws.cell(row=r, column=8).border = THIN_BORDER

    _set(ws, 3, 4, '最大風速')
    _set(ws, 3, 5, wind_speed)
    _set(ws, 3, 6, '発電量/PC')
    _set(ws, 3, 7, module_wattage)

    _set(ws, 4, 4, '垂直積雪量')
    _set(ws, 4, 5, snow_load)
    _set(ws, 4, 6, '発電量(W)/基')
    if module_wattage and rows_val and cols_val:
        panel_per_set = int(rows_val) * int(cols_val) if rows_val and cols_val else 0
        _set(ws, 4, 7, f'=G3*({panel_per_set}-G6)')
    else:
        _set(ws, 4, 7, '')

    _set(ws, 5, 4, '地上高さ')
    _set(ws, 5, 5, ground_height)
    _set(ws, 5, 6, 'スパン(E/W)')
    _set(ws, 5, 7, span_ew)

    _set(ws, 6, 3, 'パネル数')
    _set(ws, 6, 4, f'{rows_val}段' if rows_val else '')
    _set(ws, 6, 5, f'{cols_val}列' if cols_val else '')
    _set(ws, 6, 6, 'セット数')
    _set(ws, 6, 7, qty_val)

    ws.merge_cells('A7:H7')
    for c in range(1, 9):
        ws.cell(row=7, column=c).border = THIN_BORDER

    YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    headers = ['番号', '部品名称', '材質', '写真', '規格', '単価\n(EXW US$)', '数量（PCS)', '総金額\n(EXW US$)']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=8, column=i, value=h)
        cell.font = Font(name='Arial', bold=True, size=8)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.fill = LIGHT_BLUE_FILL

    YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    data_font = SM_FONT
    data_start = 9
    total_price_sum = Decimal('0')
    matched_count = 0
    name_field = 'name_ja'

    for row_idx in range(data_start, data_start + len(bom_products)):
        ws.row_dimensions[row_idx].height = 60

    for idx, product in enumerate(bom_products):
        row = data_start + idx
        product_code = product.get('code', '')
        price_info = resolve_price_info(price_mapping, product_code, spec=product.get('spec', ''))

        display_name = (
            price_info.get(name_field)
            or price_info.get('name_ko')
            or price_info.get('name')
            or product.get('name', '')
        ) if price_info else product.get('name', '')

        _set(ws, row, 1, idx + 1)
        _set(ws, row, 2, display_name)

        raw_material = product.get('material', '')
        display_material = ''
        if raw_material:
            mat = str(raw_material).replace('&', '+')
            if coating_thickness in (15, 18):
                mat = mat.replace('304', '316')
                mat = mat.replace('/316', '')
            else:
                mat = mat.replace('/316', '')
            mat = translate_material(mat, 'ja')
            display_material = mat
            if coating_thickness in (15, 18):
                display_material += f'  {coating_thickness}um'

        _set(ws, row, 3, display_material)
        _set(ws, row, 4, '')
        _set(ws, row, 5, product.get('spec', ''))

        quantity = product.get('quantity', 0)
        unit_price = 0
        display_unit_price = 0
        price_unit = ''
        is_matched = False

        if price_info and has_valid_price_info(price_info):
            unit_price = float(price_info['price'])
            price_unit = price_info.get('unit', '')
            matched_count += 1
            is_matched = True
        else:
            if unmatched_products_out is not None and quantity > 0 and _is_valid_product_code(product_code):
                unmatched_products_out.append({
                    'code': product_code,
                    'name': product.get('name', ''),
                    'spec': product.get('spec', ''),
                    'material': product.get('material', ''),
                    'quantity': quantity,
                })

        is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
        length_mm = Decimal('0')
        if is_meter:
            length_mm = Decimal(str(extract_length_from_spec(product.get('spec')) or 0))

        if is_meter and length_mm > 0:
            display_unit_price = float(Decimal(str(unit_price)) * length_mm / Decimal('1000'))
        else:
            display_unit_price = unit_price

        if display_unit_price > 0:
            cell = ws.cell(row=row, column=6, value=display_unit_price)
            cell.border = THIN_BORDER
            cell.number_format = CURRENCY_FMT
            cell.font = data_font
            cell.alignment = CENTER
        else:
            _set(ws, row, 6, '')

        if quantity > 0:
            _set(ws, row, 7, int(quantity) if quantity % 1 == 0 else quantity)
        else:
            _set(ws, row, 7, '')

        total_price = Decimal('0')
        if display_unit_price > 0 and quantity > 0:
            total_price = Decimal(str(display_unit_price)) * Decimal(str(quantity))

        total_price_rounded = round_to_2_decimal(total_price)
        if total_price_rounded > 0:
            cell = ws.cell(row=row, column=8, value=float(total_price))
            cell.border = THIN_BORDER
            cell.number_format = CURRENCY_FMT
            cell.font = data_font
            cell.alignment = CENTER
            total_price_sum += total_price_rounded
        else:
            _set(ws, row, 8, '')

        if not is_matched:
            for c in range(1, 9):
                ws.cell(row=row, column=c).fill = YELLOW_FILL

    rows_to_delete = []
    for r in range(data_start, data_start + len(bom_products)):
        name_val = ws.cell(row=r, column=2).value
        if name_val is None or name_val == '':
            continue
        qty_val_cell = ws.cell(row=r, column=7).value
        if qty_val_cell is None or qty_val_cell == '' or (isinstance(qty_val_cell, (int, float)) and qty_val_cell == 0):
            rows_to_delete.append(r)

    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r)

    deleted_count = len(rows_to_delete)
    if deleted_count:
        print(f"   🗑️ 数量=0の行を {deleted_count} 行削除")

    new_seq = 1
    for r in range(data_start, ws.max_row + 1):
        name_val = ws.cell(row=r, column=2).value
        if name_val and name_val != '':
            ws.cell(row=r, column=1, value=new_seq)
            new_seq += 1

    data_end = data_start + len(bom_products) - 1 - deleted_count

    row_product_map = {}
    for idx, product in enumerate(bom_products):
        orig_row = data_start + idx
        pc = product.get('code', '')
        row_product_map[orig_row] = str(pc).strip() if pc else ''

    survived = [r for r in sorted(row_product_map.keys()) if r not in set(rows_to_delete)]
    row_remap = {}
    for orig_row in survived:
        offset = sum(1 for d in rows_to_delete if d < orig_row)
        row_remap[orig_row - offset] = row_product_map[orig_row]

    image_found_count = 0
    image_not_found_count = 0
    for row, product_code in row_remap.items():
        if row < data_start or row > data_end:
            continue
        if not product_code:
            continue
        normalized_code = normalize_lookup_code(product_code)

        img_path = None
        if code_to_images and product_code in code_to_images and code_to_images[product_code]:
            img_path = code_to_images[product_code][0]
        elif code_to_images and normalized_code in code_to_images and code_to_images[normalized_code]:
            img_path = code_to_images[normalized_code][0]

        if not img_path and price_mapping:
            pi = resolve_price_info(price_mapping, product_code)
            if pi and pi.get('image_bytes'):
                img_bytes = pi['image_bytes']
                img_ext = pi.get('image_ext', '.png')
                if image_temp_dir:
                    db_img_name = f"db_{normalized_code or product_code}{img_ext}"
                    db_img_path = os.path.join(image_temp_dir, db_img_name)
                    if db_img_path not in (image_cache or {}):
                        with open(db_img_path, 'wb') as _f:
                            _f.write(img_bytes)
                        if image_cache is not None:
                            image_cache[db_img_path] = True
                    img_path = db_img_path

        if img_path:
            fit_w, fit_h = fit_image(ws, row, img_col, img_w, img_h)
            temp_img_path = prepare_image_for_excel(
                img_path,
                target_width=fit_w,
                target_height=fit_h,
                temp_dir=image_temp_dir,
                cache=image_cache,
            )
            final_img_path = temp_img_path if temp_img_path else img_path
            success = add_image_centered_in_cell(
                ws,
                final_img_path,
                row,
                img_col,
                img_width=fit_w,
                img_height=fit_h,
            )
            if success:
                image_found_count += 1
                continue
        image_not_found_count += 1
        ws.cell(row=row, column=4, value='/').alignment = CENTER

    if image_found_count or image_not_found_count:
        print(f"   🖼️ {sheet_name}: 挿入 {image_found_count} 枚, 未検出 {image_not_found_count} 枚")

    sub_row = data_end + 1
    ws.merge_cells(f'A{sub_row}:G{sub_row}')
    _set(ws, sub_row, 1, 'SUB-TOTAL-（EXW）1基架台合計', font=SM_FONT_BOLD, align=Alignment(horizontal='right', vertical='center'))
    for c in range(1, 8):
        ws.cell(row=sub_row, column=c).border = THIN_BORDER
    cell_h = ws.cell(row=sub_row, column=8, value=f'=SUM(H{data_start}:H{data_end})')
    cell_h.font = SM_FONT_BOLD
    cell_h.border = THIN_BORDER
    cell_h.alignment = CENTER
    cell_h.number_format = CURRENCY_FMT
    ws.row_dimensions[sub_row].height = 25

    total_row = sub_row + 1
    ws.merge_cells(f'A{total_row}:G{total_row}')
    _set(ws, total_row, 1, f'TOTAL-（EXW）{qty_val}基架台合計', font=SM_FONT_BOLD, align=Alignment(horizontal='right', vertical='center'))
    for c in range(1, 8):
        ws.cell(row=total_row, column=c).border = THIN_BORDER
    cell_total = ws.cell(row=total_row, column=8, value=f'=G6*H{sub_row}')
    cell_total.font = SM_FONT_BOLD
    cell_total.border = THIN_BORDER
    cell_total.alignment = CENTER
    cell_total.number_format = CURRENCY_FMT
    ws.row_dimensions[total_row].height = 25

    watt_row = total_row + 1
    ws.merge_cells(f'A{watt_row}:G{watt_row}')
    _set(ws, watt_row, 1, '1Ｗあたり金額', font=SM_FONT_BOLD, align=Alignment(horizontal='right', vertical='center'))
    for c in range(1, 8):
        ws.cell(row=watt_row, column=c).border = THIN_BORDER
    cell_watt = ws.cell(row=watt_row, column=8, value=f'=H{sub_row}/G4')
    cell_watt.font = SM_FONT_BOLD
    cell_watt.border = THIN_BORDER
    cell_watt.alignment = CENTER
    cell_watt.number_format = '0.00'
    ws.row_dimensions[watt_row].height = 25

    pile_data_start_row = 0
    pile_data_end_row = 0
    pile_total_per_base = 0.0
    pile_total_qty = 0
    pile_products = pile_products or []
    _pile_img_cache = {}
    pile_display_name = ''
    if pile_products:
        import tempfile as _tf
        _pile_tmp = _tf.mkdtemp(prefix='nv_pile_img_')
        pile_start_row = watt_row + 1

        pile_seq = 1
        pile_data_start_row = pile_start_row
        rendered_pile_count = 0
        for p in pile_products:
            qty_check = int(p.get('quantity', 0) or 0)
            if qty_check <= 0:
                continue
            row = pile_data_start_row + pile_seq - 1
            ws.row_dimensions[row].height = 60
            code = str(p.get('code', '') or '')
            price_info = resolve_price_info(price_mapping, code, spec=p.get('spec', '')) if price_mapping else None
            display_name = 'スクリュー杭'
            if not pile_display_name:
                pile_display_name = 'スクリュー杭'
            material = p.get('material', '')
            spec = _strip_cjk_spec(p.get('spec', ''))
            if len(spec) > 10:
                spec = spec[:10]
            qty = qty_check

            unit_price = 0.0
            display_unit_price = 0.0
            pile_is_matched = False
            if price_info and has_valid_price_info(price_info):
                unit_price = float(price_info['price'])
                pile_is_matched = True
            else:
                if unmatched_products_out is not None and qty > 0 and _is_valid_product_code(code):
                    unmatched_products_out.append({
                        'code': code,
                        'name': p.get('full_name') or p.get('name', ''),
                        'spec': spec,
                        'material': material,
                        'quantity': qty,
                    })

            total_price = 0.0
            pricing_unit = (price_info.get('unit', '') if price_info else '') or ''
            length_mm = float(p.get('length', 0) or 0)
            if not pile_is_matched:
                display_unit_price = ''
            elif pricing_unit == '米' and length_mm > 0:
                display_unit_price = (length_mm / 1000) * unit_price
            else:
                display_unit_price = unit_price
            if isinstance(display_unit_price, (int, float)) and display_unit_price > 0 and qty > 0:
                total_price = display_unit_price * qty

            pile_total_per_base += total_price
            pile_total_qty += qty

            ws.cell(row=row, column=1, value=pile_seq).font = SM_FONT
            ws.cell(row=row, column=1).alignment = CENTER
            ws.cell(row=row, column=2, value=display_name).font = SM_FONT
            ws.cell(row=row, column=2).alignment = CENTER
            ws.cell(row=row, column=3, value=translate_material(material, 'ja')).font = SM_FONT
            ws.cell(row=row, column=3).alignment = CENTER
            _set(ws, row, 4, '')
            ws.cell(row=row, column=5, value=spec).font = SM_FONT
            ws.cell(row=row, column=5).alignment = CENTER

            if display_unit_price > 0:
                ws.cell(row=row, column=6, value=display_unit_price).font = SM_FONT
                ws.cell(row=row, column=6).alignment = CENTER
                ws.cell(row=row, column=6).number_format = CURRENCY_FMT
            else:
                _set(ws, row, 6, '')

            ws.cell(row=row, column=7, value=qty).font = SM_FONT
            ws.cell(row=row, column=7).alignment = CENTER

            if total_price > 0:
                ws.cell(row=row, column=8, value=total_price).font = SM_FONT
                ws.cell(row=row, column=8).alignment = CENTER
                ws.cell(row=row, column=8).number_format = CURRENCY_FMT
            else:
                _set(ws, row, 8, '')

            for c in range(1, 9):
                ws.cell(row=row, column=c).border = THIN_BORDER

            if not pile_is_matched:
                for c in range(1, 9):
                    ws.cell(row=row, column=c).fill = YELLOW_FILL

            img_inserted = False
            if price_info and price_info.get('image_bytes'):
                try:
                    import os as _os
                    _img_bytes = price_info['image_bytes']
                    _img_ext = price_info.get('image_ext', '.png') or '.png'
                    _db_img_name = f"pile_{code}{_img_ext}"
                    _db_img_path = _os.path.join(_pile_tmp, _db_img_name)
                    if _db_img_path not in _pile_img_cache:
                        with open(_db_img_path, 'wb') as _f:
                            _f.write(_img_bytes)
                        _pile_img_cache[_db_img_path] = True
                    _fit_w, _fit_h = fit_image(ws, row, img_col, img_w, img_h)
                    _prep = prepare_image_for_excel(
                        _db_img_path,
                        target_width=_fit_w,
                        target_height=_fit_h,
                        temp_dir=_pile_tmp,
                        cache=_pile_img_cache,
                    )
                    _final = _prep if _prep else _db_img_path
                    img_inserted = add_image_centered_in_cell(
                        ws, _final, row, img_col,
                        img_width=_fit_w, img_height=_fit_h,
                    )
                except Exception:
                    pass
            if not img_inserted:
                ws.cell(row=row, column=4, value='/').alignment = CENTER

            pile_seq += 1
            rendered_pile_count += 1

        try:
            import shutil as _sh
            _sh.rmtree(_pile_tmp, ignore_errors=True)
        except Exception:
            pass

        pile_data_end_row = pile_data_start_row + rendered_pile_count - 1 if rendered_pile_count > 0 else 0

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(top=0.75, bottom=0.75, left=0.25, right=0.25, header=0.3, footer=0.3)

    discount_total_row = None

    return {
        'sheet_name': sheet_name,
        'discount_total_row': discount_total_row,
        'sub_total_row': sub_row,
        'pile_data_start_row': pile_data_start_row,
        'pile_data_end_row': pile_data_end_row,
        'array_info': array_info,
        'angle': angle,
    }


def _create_fence_detail_sheet(workbook, fence_rows, gate_rows, nv_fgg,
                               matrix_data=None, nv_params=None):
    matrix_data = matrix_data or {}
    nv_params = nv_params or {}

    coating = (nv_fgg or {}).get('coating', '浸塑')
    fence_surface = ((nv_fgg or {}).get('fence') or {}).get('surface', '白色')
    fence_material_map = {}
    all_codes = list(set(
        str(r.get('code', '')).strip()
        for r in (fence_rows + gate_rows)
        if str(r.get('code', '')).strip()
    ))
    if all_codes:
        try:
            from backend.repositories.fence_gate_material_repository import get_material
            surface_prefix_map = {
                '白色浸塑': 'FN01', '咖啡色浸塑': 'FN02', '绿色浸塑': 'FN03',
                '灰褐色浸塑': 'FN04', '深茶色浸塑': 'FN05', '银灰色浸塑': 'FN06',
                '黑色浸塑': 'FN07', '深咖色浸塑': 'FN08', '热镀锌': 'FN11',
                '咖啡浸塑': 'FN02', '绿浸塑': 'FN03',
                '灰褐浸塑': 'FN04', '深茶浸塑': 'FN05', '银灰浸塑': 'FN06',
                '黑浸塑': 'FN07', '深咖浸塑': 'FN08',
            }
            color_prefix = surface_prefix_map.get(fence_surface, 'FN01')
            for code in all_codes:
                try:
                    mat = get_material(code)
                    if mat:
                        fence_material_map[code] = mat
                        continue
                except Exception:
                    pass
                if code.startswith('M') and '-' in code:
                    for prefix in [color_prefix, 'FN01', 'FN11']:
                        full_code = f'{prefix}-{code}'
                        try:
                            mat = get_material(full_code)
                            if mat:
                                fence_material_map[code] = mat
                                break
                        except Exception:
                            pass
        except Exception:
            pass

    fence_section = (nv_fgg or {}).get('fence') or {}
    gate_section = (nv_fgg or {}).get('gate') or {}
    fence_style = fence_section.get('style', '')
    height = '1500'
    if fence_style:
        parts = fence_style.split('-')
        if len(parts) >= 2:
            height = parts[-1]

    sheet_name = f'フェンス H{height}'
    if sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
    else:
        ws = workbook.create_sheet(title=sheet_name)

    for col, width in DETAIL_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30

    fence_surface = fence_section.get('surface', '白色')
    surface_jp = '茶色' if '茶' in fence_surface else '白色'
    fence_len = float(fence_section.get('totalLength', 0) or 0)
    corner_qty = int(fence_section.get('cornerQty', 0) or 0)

    gate_style_code = gate_section.get('gateStyle', '')
    gate_qty_val = int(gate_section.get('gateQty', 0) or 0)

    foundation_type = 'コンクリート基礎'
    gate_type_label = ''
    single_gate_qty = 0
    double_gate_qty = 0
    double_gate_width = 4200

    if gate_style_code and len(gate_style_code) >= 6:
        prefix = gate_style_code[:3]
        width_code = gate_style_code[3:6]

        if prefix.endswith('p'):
            foundation_type = '一体式基礎'
        elif prefix.endswith('g'):
            foundation_type = '杭基礎'

        if width_code == '120':
            single_gate_qty = gate_qty_val
            if gate_qty_val > 0:
                gate_type_label = '片開き門扉'
        else:
            double_gate_qty = gate_qty_val
            try:
                double_gate_width = int(width_code) * 10
            except (ValueError, TypeError):
                double_gate_width = 4200
            if gate_qty_val > 0:
                gate_type_label = '両開き門扉'

    ws.merge_cells('F1:G1')
    ws['F1'] = '見積日：'
    ws['F1'].font = SM_FONT
    ws['F1'].alignment = RIGHT_A
    ws['H1'] = '=TODAY()'
    ws['H1'].number_format = 'yyyy"年"m"月"dd"日"'
    ws['H1'].font = SM_FONT

    ws.merge_cells('A2:H2')
    ws['A2'] = 'フェンス御見積書'
    ws['A2'].font = Font(name='Arial', bold=True, size=16)
    ws['A2'].alignment = CENTER

    if os.path.isfile(_FENCE_LOGO_PATH):
        try:
            _fimg = XLImage(_FENCE_LOGO_PATH)
            _fimg.width = 240
            _fimg.height = 80
            ws.add_image(_fimg, 'A1')
        except Exception:
            pass

    ws.delete_rows(3)

    ws.merge_cells('G3:H9')
    desc_text = (
        f'1.ディップコーディング　{surface_jp}\n'
        f'2.{foundation_type}\n'
        f'3.H={height}ｍｍ　{int(fence_len)}ｍ\n'
        f'4.片開き門扉幅W1200　{single_gate_qty}ヶ所\n'
        f'5.両開き門扉幅W{double_gate_width}　{double_gate_qty}ヶ所\n'
        f'6.柱材コーナー部　{corner_qty}箇所\n'
        f'（0枚予備）'
    )
    _set(ws, 3, 7, desc_text, align=LEFT_A, border=False)

    bottom_border = Border(bottom=Side(style='thin'))

    project_name = str(matrix_data.get('project_name') or '').strip()

    def _info_row(row_num, label, value):
        ws.merge_cells(f'A{row_num}:B{row_num}')
        ws[f'A{row_num}'] = label
        ws[f'A{row_num}'].font = SM_FONT
        ws.merge_cells(f'C{row_num}:F{row_num}')
        ws[f'C{row_num}'] = value
        ws[f'C{row_num}'].font = SM_FONT

    _info_row(4, '案件名：', f'{project_name} フェンス H{height}  {int(fence_len)}M （{surface_jp}）')
    _info_row(5, '見積条件：', f'DDP現場（船便）--{project_name}')
    _info_row(6, '納入期限：', '発注後2－3週後工場から出荷')
    _info_row(7, '取引条件：', 'TTで30%予約金が前払い、B/L発行後引き渡し前に70%お支払')
    _info_row(8, '有効期限：', '御見積後2週間')

    r = 10

    subtitle_parts = [p for p in [foundation_type, gate_type_label] if p]
    subtitle_suffix = '　'.join(subtitle_parts)
    ws.merge_cells(f'A{r}:H{r}')
    ws[f'A{r}'] = f'一、フェンス金額（Ex Works）----{subtitle_suffix}' if subtitle_suffix else '一、フェンス金額（Ex Works）'
    ws[f'A{r}'].font = Font(name='Arial', bold=True, size=10)
    ws[f'A{r}'].alignment = LEFT_A

    r += 1
    header_row = r
    headers = ['番号', '部品名称', '材質', '写真', '規格', '単価\n(EXW US$)', '数量（PCS)', '総金額\n(EXW US$)']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = Font(name='Arial', bold=True, size=8)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.fill = LIGHT_BLUE_FILL
    ws.row_dimensions[header_row].height = 25

    all_items = []
    seq = 1
    for row_data in fence_rows:
        all_items.append({**row_data, '_seq': seq})
        seq += 1
    for row_data in gate_rows:
        all_items.append({**row_data, '_seq': seq})
        seq += 1

    data_start = header_row + 1
    for idx, item in enumerate(all_items):
        row = data_start + idx
        ws.row_dimensions[row].height = 60
        for c in range(1, 9):
            ws.cell(row=row, column=c).border = THIN_BORDER

        _set(ws, row, 1, item.get('_seq', idx + 1))

        item_code = str(item.get('code', '')).strip()
        mat_info = fence_material_map.get(item_code)
        if mat_info:
            display_name = mat_info.get('日语名称', '') or item.get('name', '')
        else:
            display_name = item.get('name', '')
        _set(ws, row, 2, display_name)

        if mat_info:
            if coating == '热镀锌':
                display_material = mat_info.get('材質表面処理_热镀锌', '')
            else:
                display_material = mat_info.get('材質表面処理_浸塑', '')
        else:
            display_material = ''
        _set(ws, row, 3, display_material)
        _set(ws, row, 4, '/')
        _set(ws, row, 5, item.get('spec', ''))

        unit_price = item.get('unit_price', 0)
        qty = item.get('qty', 0)

        ws.cell(row=row, column=6, value=unit_price).font = SM_FONT
        ws.cell(row=row, column=6).alignment = CENTER
        ws.cell(row=row, column=6).number_format = CURRENCY_FMT

        _set(ws, row, 7, qty)

        ws.cell(row=row, column=8, value=f'=F{row}*G{row}').border = THIN_BORDER
        ws.cell(row=row, column=8).alignment = CENTER
        ws.cell(row=row, column=8).number_format = CURRENCY_FMT
        ws.cell(row=row, column=8).font = SM_FONT

    data_end = data_start + len(all_items) - 1
    discount_total_row = None
    if all_items:
        sum_row = data_end + 1
        ws.merge_cells(f'A{sum_row}:G{sum_row}')
        _set(ws, sum_row, 1, 'SUB-TOTAL-（EXW）フェンス合計', font=SM_FONT_BOLD, align=Alignment(horizontal='right', vertical='center'))
        for c in range(1, 8):
            ws.cell(row=sum_row, column=c).border = THIN_BORDER
        ws.cell(row=sum_row, column=8, value=f'=SUM(H{data_start}:H{data_end})').border = THIN_BORDER
        ws.cell(row=sum_row, column=8).font = SM_FONT_BOLD
        ws.cell(row=sum_row, column=8).alignment = CENTER
        ws.cell(row=sum_row, column=8).number_format = CURRENCY_FMT
        ws.row_dimensions[sum_row].height = 25

        fence_discount_rate = nv_params.get('fence_discount_rate') or nv_params.get('nv_fence_discount_rate') or 94

        pile_total = Decimal('0')
        for item in all_items:
            code = str(item.get('code', '')).strip()
            if code.startswith('FN-D') or code.startswith('FN-YG'):
                pile_total += Decimal(str(item.get('amount', 0) or 0))

        total_row = sum_row + 1
        ws.merge_cells(f'A{total_row}:G{total_row}')
        _set(ws, total_row, 1, '合計金額(USD)', font=SM_FONT_BOLD, align=Alignment(horizontal='right', vertical='center'))
        for c in range(1, 8):
            ws.cell(row=total_row, column=c).border = THIN_BORDER
        ws.cell(row=total_row, column=8, value=f'=H{sum_row}').border = THIN_BORDER
        ws.cell(row=total_row, column=8).font = SM_FONT_BOLD
        ws.cell(row=total_row, column=8).alignment = CENTER
        ws.cell(row=total_row, column=8).number_format = CURRENCY_FMT
        ws.row_dimensions[total_row].height = 25

        all_total = Decimal(str(sum(r.get('amount', 0) for r in all_items)))
        non_pile_total = all_total - pile_total
        discounted_total = non_pile_total * Decimal(str(fence_discount_rate)) / Decimal('100') + pile_total

        discount_total_row = total_row + 1
        ws.merge_cells(f'A{discount_total_row}:G{discount_total_row}')
        _set(ws, discount_total_row, 1, '特別値引き後合計金額(USD)', font=SM_FONT_RED, align=Alignment(horizontal='right', vertical='center'))
        for c in range(1, 8):
            ws.cell(row=discount_total_row, column=c).border = THIN_BORDER
        ws.cell(row=discount_total_row, column=8, value=float(round(discounted_total, 2))).border = THIN_BORDER
        ws.cell(row=discount_total_row, column=8).font = SM_FONT_RED
        ws.cell(row=discount_total_row, column=8).alignment = CENTER
        ws.cell(row=discount_total_row, column=8).number_format = CURRENCY_FMT
        ws.row_dimensions[discount_total_row].height = 25

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(top=0.75, bottom=0.75, left=0.25, right=0.25, header=0.3, footer=0.3)

    return {'sheet_name': sheet_name, 'discount_total_row': discount_total_row}