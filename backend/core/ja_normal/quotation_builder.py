import os
import re
import shutil
import uuid
from decimal import Decimal

from backend.core.shared.text_utils import _CJK_RE, _strip_cjk_spec

from openpyxl import Workbook

from backend.core.material_translate import translate_material
from backend.excel.reader import excel_file_compat
from backend.core.shared.bom_utils import (
    discover_sheet_bom_starts, extract_bom_dataframe, get_bom_processing_rules,
    normalize_selected_bom_keys, quick_scan_bom_sheets, parse_array_to_rows_cols,
    read_bom_from_dataframe,
)
from backend.core.shared.image_utils import scan_images, find_latest_image_log, load_image_mapping_from_log
from backend.core.shared.price_utils import resolve_price_info, has_valid_price_info, round_to_2_decimal
from backend.core.shared.product_utils import _match_exclude_group
from backend.core.inquiry_builder import create_inquiry_sheet
from backend.core.shared.bom_zip_parser import _build_products_by_key
from backend.core.shared.product_utils import _split_pile_products
from backend.core.ja_nv.quotation_engine import create_nv_detail_sheet
from backend.core.ja_normal.quotation_engine import create_normal_summary_sheet
from backend.core.shared.sheet_utils import set_page_break_preview


def split_and_create_quotations(
        input_file,
        price_file_path,
        output_dir=None,
        image_path=None,
        image_folder=None,
        price_mapping_override=None,
        contact_info=None,
        matrix_data=None,
        return_details=False,
        selected_bom_keys=None,
        group=None,
        pre_parsed_products=None,
        pre_parsed_bom_info=None,
        pre_parsed_bom_configs=None,
        exclude_options=None,
        pre_parsed_inverter_products=None,
        fence_data=None,
        need_weight_code=False,
        coating_thickness=10,
        case_type=None,
        normal_params=None,
        nv_fence_gate_data=None,
        tariff_rate=None,
        consumption_tax=None,
        fence_tax=None,
        discount_rate=None,
        truck_desc=None,
        truck_fee=None,
        need_total_qty=False,
        exclude_delete_options=None,
):
    normal_params = normal_params or {}

    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(input_file), 'quotation_output')
    os.makedirs(output_dir, exist_ok=True)

    code_to_images = {}
    image_cache = {}
    image_temp_dir = None

    log_search_dirs = [output_dir, os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")]
    latest_log = find_latest_image_log(log_search_dirs)
    if latest_log:
        log_mapping = load_image_mapping_from_log(latest_log)
        if log_mapping:
            code_to_images.update(log_mapping)

    if image_folder and os.path.exists(image_folder):
        scanned = scan_images(image_folder)
        for code, paths in scanned.items():
            if code not in code_to_images:
                code_to_images[code] = paths
        if code_to_images:
            image_temp_dir = os.path.join(output_dir, f"temp_images_{uuid.uuid4().hex}")
            os.makedirs(image_temp_dir, exist_ok=True)
    elif image_folder:
        image_folder = None

    if price_mapping_override is not None:
        price_mapping = price_mapping_override
    else:
        from backend.core.shared.price_utils import load_price_mapping
        price_mapping = load_price_mapping(price_file_path)

    selected_key_set = normalize_selected_bom_keys(selected_bom_keys)
    column_mapping, skip_keywords, non_bom_sheet_keywords = get_bom_processing_rules()

    arrays = (matrix_data or {}).get('arrays') or []
    all_detail_results = []
    all_unmatched_products = []

    used_array_indices = set()

    def find_matching_array(bom_rows, bom_cols, bom_base_count=0):
        if bom_base_count:
            for i, arr in enumerate(arrays):
                if i in used_array_indices:
                    continue
                if bom_rows == arr.get('rows') and bom_cols == arr.get('cols'):
                    if bom_base_count == (arr.get('table_qty') or 1):
                        used_array_indices.add(i)
                        return arr
            for i, arr in enumerate(arrays):
                if i in used_array_indices:
                    continue
                if bom_rows == arr.get('rows') and bom_cols == arr.get('cols'):
                    used_array_indices.add(i)
                    return arr
            return None
        for i, arr in enumerate(arrays):
            if i in used_array_indices:
                continue
            if bom_rows == arr.get('rows') and bom_cols == arr.get('cols'):
                used_array_indices.add(i)
                return arr
        return None

    all_excluded_products = []

    def _filter_excluded(products):
        if not exclude_options or not any(exclude_options.values()):
            return products, []
        filtered = []
        excluded = []
        for p in products:
            if _match_exclude_group(p, price_mapping, exclude_options):
                excluded.append(p)
            else:
                filtered.append(p)
        return filtered, excluded

    if pre_parsed_products and pre_parsed_bom_info:
        products_by_key = pre_parsed_products
        bom_info_by_key = pre_parsed_bom_info
    else:
        products_by_key, bom_info_by_key, _, _, _ = _build_products_by_key(
            input_file, selected_key_set, column_mapping, skip_keywords, non_bom_sheet_keywords
        )

    pile_products_all = []
    detail_sheet_products = []
    matched_inv_keys = set()

    master_wb = Workbook()
    default_sheet = master_wb.active
    master_wb.remove(default_sheet)

    for key, products in products_by_key.items():
        bom_info = bom_info_by_key.get(key)
        if not bom_info:
            continue

        config = bom_info.get('config', {})
        bom_array_str = config.get('array', '')
        bom_rows, bom_cols = parse_array_to_rows_cols(bom_array_str)

        matched_array = find_matching_array(bom_rows, bom_cols, bom_base_count=config.get('base_count', 0))
        if matched_array is None:
            continue

        matched_rows = matched_array.get('rows', '')
        matched_cols = matched_array.get('cols', '')
        matched_qty = matched_array.get('table_qty', 1)
        sheet_prefix = f'{matched_rows}x{matched_cols}_{matched_qty}'

        filtered, excluded = _filter_excluded(products)
        if excluded:
            all_excluded_products.extend(excluded)
        if not filtered:
            continue

        bracket_prods, pile_prods = _split_pile_products(filtered, price_mapping)
        if pile_prods:
            for pp in pile_prods:
                scaled = dict(pp)
                scaled['quantity'] = float(pp.get('quantity', 0)) * matched_qty
                pile_products_all.append(scaled)

        inv_prods_for_key = []
        for ip in (pre_parsed_inverter_products or []):
            if ip.get('bom_key') == key:
                if int(ip.get('quantity', 0) or 0) > 0:
                    inv_prods_for_key.append(ip)
                    matched_inv_keys.add(id(ip))

        all_prods = list(bracket_prods) + inv_prods_for_key
        if not all_prods and not pile_prods:
            continue

        detail_sheet_products.extend(all_prods)
        if pile_prods:
            for pp in pile_prods:
                _pp_copy = dict(pp)
                _pp_copy['_is_pile'] = True
                detail_sheet_products.append(_pp_copy)

        try:
            detail = create_nv_detail_sheet(
                master_wb,
                matched_array,
                all_prods,
                price_mapping,
                sheet_prefix=sheet_prefix,
                matrix_data=matrix_data,
                unmatched_products_out=all_unmatched_products,
                coating_thickness=coating_thickness,
                nv_params=normal_params,
                pile_products=pile_prods,
                code_to_images=code_to_images,
                image_temp_dir=image_temp_dir,
                image_cache=image_cache,
                angle_override=config.get('angle', '') or None,
            )
            detail['config'] = config
            detail['variant_name'] = bom_info.get('variant_name', '')
            detail['bom_key'] = key
            detail['angle'] = config.get('angle', '') or (matrix_data or {}).get('angle', '')
            inv_count = 0
            for ip in (pre_parsed_inverter_products or []):
                if ip.get('bom_key') == key:
                    ic = int(ip.get('inverter_total_count', 0) or 0)
                    if ic > 0:
                        inv_count = ic
                        break
                    inv_count += int(ip.get('quantity', 0) or 0)
            detail['inverter_count'] = inv_count
            all_detail_results.append(detail)
        except Exception as e:
            import traceback
            print(f"   ❌ 普通案件明細シート生成失敗: {e}")
            traceback.print_exc()

    unmatched_inv = [
        ip for ip in (pre_parsed_inverter_products or [])
        if id(ip) not in matched_inv_keys and int(ip.get('quantity', 0) or 0) > 0
    ]
    if unmatched_inv and all_detail_results:
        first_detail = all_detail_results[0]
        first_sn = first_detail.get('sheet_name', '')
        ws_first = master_wb[first_sn] if first_sn in master_wb.sheetnames else None
        if ws_first:
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            inv_start = ws_first.max_row + 1
            sm_font = Font(name='Yu Gothic UI', size=8)
            center_a = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_b = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'),
            )
            YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for idx, ip in enumerate(unmatched_inv):
                row = inv_start + idx
                code = str(ip.get('code', '') or '')
                pi = resolve_price_info(price_mapping, code, spec=ip.get('spec', '')) if price_mapping else None
                display_name = (
                    (pi.get('name_ja') or pi.get('name_ko') or pi.get('name'))
                    if pi else (ip.get('full_name') or ip.get('name', ''))
                ) or ''
                qty = int(ip.get('quantity', 0) or 0)
                inv_is_matched = False
                unit_price = 0
                if pi and has_valid_price_info(pi):
                    unit_price = float(pi['price'])
                    inv_is_matched = True
                else:
                    if qty > 0:
                        all_unmatched_products.append({
                            'code': code,
                            'name': ip.get('full_name') or ip.get('name', ''),
                            'spec': ip.get('spec', ''),
                            'material': (pi.get('db_material') if pi and pi.get('db_material') else None) or ip.get('material', ''),
                            'quantity': qty,
                        })
                total_price = unit_price * qty if unit_price > 0 else 0
                ws_first.cell(row=row, column=1, value=row - inv_start + 1).font = sm_font
                ws_first.cell(row=row, column=1).alignment = center_a
                ws_first.cell(row=row, column=2, value=display_name).font = sm_font
                ws_first.cell(row=row, column=2).alignment = center_a
                _ja_normal_mat = (pi.get('db_material') if pi and pi.get('db_material') else None) or ip.get('material', '')
                ws_first.cell(row=row, column=3, value=translate_material(_ja_normal_mat, 'ja')).font = sm_font
                ws_first.cell(row=row, column=3).alignment = center_a
                ws_first.cell(row=row, column=4).font = sm_font
                ws_first.cell(row=row, column=5, value=_strip_cjk_spec(ip.get('spec', ''))).font = sm_font
                ws_first.cell(row=row, column=5).alignment = center_a
                if unit_price > 0:
                    ws_first.cell(row=row, column=6, value=unit_price).font = sm_font
                    ws_first.cell(row=row, column=6).alignment = center_a
                    ws_first.cell(row=row, column=6).number_format = '#,##0.00'
                else:
                    ws_first.cell(row=row, column=6, value='').font = sm_font
                    ws_first.cell(row=row, column=6).alignment = center_a
                ws_first.cell(row=row, column=7, value=qty).font = sm_font
                ws_first.cell(row=row, column=7).alignment = center_a
                if total_price > 0:
                    ws_first.cell(row=row, column=8, value=total_price).font = sm_font
                    ws_first.cell(row=row, column=8).alignment = center_a
                    ws_first.cell(row=row, column=8).number_format = '#,##0.00'
                else:
                    ws_first.cell(row=row, column=8, value='').font = sm_font
                    ws_first.cell(row=row, column=8).alignment = center_a
                ws_first.row_dimensions[row].height = 60
                for c in range(1, 9):
                    ws_first.cell(row=row, column=c).border = thin_b
                if not inv_is_matched:
                    for c in range(1, 9):
                        ws_first.cell(row=row, column=c).fill = YELLOW_FILL
            first_detail['inv_note'] = f"パワコン取付バー  {sum(int(ip.get('quantity', 0) or 0) for ip in unmatched_inv)}台"

    if all_detail_results:
        pile_summary = None
        if pile_products_all:
            import re as _re
            pile_filtered, pile_excluded = _filter_excluded(pile_products_all)
            if pile_excluded:
                all_excluded_products.extend(pile_excluded)
            if pile_filtered:
                pile_filtered = [pp for pp in pile_filtered if float(pp.get('quantity', 0) or 0) > 0]
            if pile_filtered:
                _pq = 0
                _pa = 0.0
                for _pp in pile_filtered:
                    _pp_qty = float(_pp.get('quantity', 0))
                    _pp_code = str(_pp.get('code', '') or '').strip()
                    _pp_pi = resolve_price_info(price_mapping, _pp_code, spec=_pp.get('spec', '')) if price_mapping else None
                    _pp_price = float(_pp_pi.get('price', 0)) if _pp_pi and _pp_pi.get('price') else 0
                    _pp_unit = (_pp_pi.get('unit', '') if _pp_pi else '') or ''
                    if _pp_unit in ('米', 'm', 'M', 'meter'):
                        _len = float(_pp.get('length', 0) or 0)
                        if _len <= 0:
                            _m = _re.search(r'(\d+\.?\d*)', _pp.get('spec', ''))
                            _len = float(_m.group(1)) if _m else 0
                        if _len > 0:
                            _pp_price = _pp_price * _len / 1000
                    _pq += _pp_qty
                    _pa += _pp_price * _pp_qty
                pile_summary = {'total_qty': _pq, 'total_price': round(_pa, 2)}
        try:
            create_normal_summary_sheet(
                master_wb,
                all_detail_results,
                matrix_data=matrix_data,
                image_path=image_path,
                fence_data=fence_data,
                normal_params=normal_params,
                nv_fence_gate_data=nv_fence_gate_data,
                pile_summary=pile_summary,
                image_temp_dir=image_temp_dir,
                image_cache=image_cache,
            )
        except Exception as e:
            import traceback
            print(f"   ❌ 普通案件合計シート生成失敗: {e}")
            traceback.print_exc()

        input_basename = os.path.splitext(os.path.basename(input_file))[0]
        output_file = os.path.join(output_dir, f"{input_basename}_普通案件見積.xlsx")
        set_page_break_preview(master_wb)
        master_wb.save(output_file)

        inquiry_file = None
        if all_unmatched_products:
            try:
                inquiry_wb = Workbook()
                inquiry_wb.remove(inquiry_wb.active)
                requester = ''
                if contact_info and contact_info.get('inquiry_requester'):
                    requester = contact_info['inquiry_requester']
                create_inquiry_sheet(
                    inquiry_wb,
                    all_unmatched_products,
                    '普通案件',
                    inquiry_requester=requester,
                )
                inquiry_file = os.path.join(output_dir, f"{input_basename}_询价表.xlsx")
                inquiry_wb.save(inquiry_file)
                print(f"   📋 询价表: {inquiry_file}")
            except Exception as e:
                print(f"   ❌ 询价表生成失敗: {e}")
                inquiry_file = None

        if image_temp_dir and os.path.exists(image_temp_dir):
            try:
                shutil.rmtree(image_temp_dir)
            except Exception:
                pass

        print(f"\n✅ 普通案件処理完了：{len(all_detail_results)} 件の明細シートを生成")
        print(f"📁 出力ファイル: {output_file}")

        if return_details:
            missing_ja_list = []
            missing_image_codes = []
            seen_ja = set()
            seen_img = set()
            for p in detail_sheet_products:
                code = str(p.get('code', '') or '').strip()
                if not code:
                    continue
                pi = resolve_price_info(price_mapping, code, spec=p.get('spec', ''))
                norm = code.upper().replace(' ', '')
                if norm not in seen_ja:
                    seen_ja.add(norm)
                    if pi and not pi.get('name_ja', '').strip():
                        missing_ja_list.append({
                            'code': pi.get('db_code', code),
                            'name': pi.get('name', '') or p.get('name', ''),
                        })
                if code not in seen_img:
                    seen_img.add(code)
                    if not pi or pi.get('image_status') in ('missing', 'invalid'):
                        missing_image_codes.append(code)
            quotation_product_codes = set()
            for p in detail_sheet_products:
                c = str(p.get('code', '') or '').strip()
                if c:
                    quotation_product_codes.add(c)
            for p in all_excluded_products:
                c = str(p.get('code', '') or '').strip()
                if c:
                    quotation_product_codes.add(c)
            return {
                'output_file': output_file,
                'inquiry_file': inquiry_file,
                'unmatched_count': len(all_unmatched_products),
                'unmatched_products': all_unmatched_products,
                'quotation_product_codes': quotation_product_codes,
                'missing_ja_list': missing_ja_list,
                'missing_image_codes': missing_image_codes,
            }
        return output_file
    else:
        print("\n❌ 明細シートが生成されませんでした")
        if image_temp_dir and os.path.exists(image_temp_dir):
            try:
                shutil.rmtree(image_temp_dir)
            except Exception:
                pass
        return None
