"""Normal案件用 フェンス明细表（围栏明细）生成。

NV の _create_fence_detail_sheet を複製して normal 専用に独立化したもの。
ヘッダや塗りつぶしの調整はこの関数内で行うことで、
NV 本体（_create_fence_detail_sheet）には一切影響しない。
共有ヘルパー・定数は NV モジュールから import して再利用する。
"""
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins
from backend.core.print_settings import apply_print_setup
from decimal import Decimal
import os
import re

from backend.core.ja_nv.quotation_engine import (
    CENTER, LEFT_A, RIGHT_A,
    DETAIL_COL_WIDTHS, NUM_FMT,
    THIN_BORDER,
    _extract_fence_height_mm_from_style,
    _format_mm_to_meter_text,
    _get_nv_rate_value,
    _lookup_gate_spec_from_db,
    _surface_to_jp_color,
    _LOGO_PATH,
)

SM_FONT = Font(name='微软雅黑', size=12)
_FK_TITLE_FONT = Font(name='微软雅黑', size=22, bold=True)
_FK_BLUE_FONT = Font(name='微软雅黑', size=12, color='0000FF')
_FK_MAT_FILL = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')


def _set(ws, r, c, val, font=SM_FONT, align=CENTER, border=True):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font
    cell.alignment = align
    if border:
        cell.border = THIN_BORDER
    return cell


def create_normal_fence_detail_sheet(workbook, fence_rows, gate_rows, nv_fgg,
                               matrix_data=None, nv_params=None,
                               image_temp_dir=None, image_cache=None):
    import tempfile
    from backend.image.processor import decode_image_base64
    from backend.core.shared.image_utils import (
        prepare_image_for_excel, add_image_centered_in_cell,
    )

    matrix_data = matrix_data or {}
    nv_params = nv_params or {}
    if image_temp_dir is None:
        image_temp_dir = tempfile.gettempdir()
    if image_cache is None:
        image_cache = {}

    img_w, img_h, img_col, img_padding = 65, 50, 5, 2

    def _col_width_to_px(cw):
        return int(cw * 7.5 + 5)

    def _row_height_to_px(rh):
        return int(rh * 1.33)

    def _fit_image(ws, r, c, mw, mh, pad=img_padding):
        cw_val = ws.column_dimensions[get_column_letter(c)].width or 8.43
        rh_val = ws.row_dimensions[r].height or 15
        aw = _col_width_to_px(cw_val) - pad * 2
        ah = _row_height_to_px(rh_val) - pad * 2
        aw = max(aw, 20)
        ah = max(ah, 20)
        scale = min(aw / mw, ah / mh, 1.0)
        return int(mw * scale), int(mh * scale)

    coating = (nv_fgg or {}).get('coating', '浸塑')
    fence_surface = ((nv_fgg or {}).get('fence') or {}).get('surface', '白色')
    fence_material_map = {}

    def _resolve_fence_material(code, color_prefix):
        normalized_code = str(code or '').strip()
        if not normalized_code:
            return None

        candidate_codes = []

        def _add_candidate(value):
            candidate = str(value or '').strip()
            if candidate and candidate not in candidate_codes:
                candidate_codes.append(candidate)

        _add_candidate(normalized_code)

        if '-' in normalized_code:
            first_part = normalized_code.split('-', 1)[0]
            if first_part.startswith('FN') and len(first_part) == 4:
                _add_candidate(normalized_code.split('-', 1)[1])

        if '-' in normalized_code:
            base_with_dash = normalized_code if normalized_code.startswith('M') else normalized_code.split('-', 1)[1] if normalized_code.startswith('FN') else ''
            if base_with_dash.startswith('M'):
                for prefix in [color_prefix, 'FN01', 'FN11']:
                    _add_candidate(f'{prefix}-{base_with_dash}')

        try:
            from backend.repositories.fence_gate_material_repository import get_material
        except Exception:
            return None

        for candidate_code in candidate_codes:
            try:
                material = get_material(candidate_code)
                if material:
                    return material
            except Exception:
                continue

        return None
    all_codes = list(set(
        str(r.get('code', '')).strip()
        for r in (fence_rows + gate_rows)
        if str(r.get('code', '')).strip()
    ))
    _codes_not_found = []
    _codes_no_image = []
    if all_codes:
        try:
            surface_prefix_map = {
                '白色浸塑': 'FN01', '咖啡色浸塑': 'FN02', '绿色浸塑': 'FN03',
                '灰褐色浸塑': 'FN04', '深茶色浸塑': 'FN05', '银灰色浸塑': 'FN06',
                '黑色浸塑': 'FN07', '深咖色浸塑': 'FN08', '热镀锌': 'FN11',
                '咖啡浸塑': 'FN02', '绿浸塑': 'FN03',
                '灰褐浸塑': 'FN04', '深茶浸塑': 'FN05', '银灰浸塑': 'FN06',
                '黑浸塑': 'FN07', '深咖浸塑': 'FN08',
            }
            color_prefix = surface_prefix_map.get(fence_surface, 'FN01')
            for code in all_codes:
                mat = _resolve_fence_material(code, color_prefix)
                if mat:
                    fence_material_map[code] = mat
                else:
                    _codes_not_found.append(code)
            for code, mat in fence_material_map.items():
                if not mat.get('image_base64'):
                    _codes_no_image.append(code)
        except Exception:
            pass

    if _codes_not_found:
        print(f"   ⚠ フェンス物料DB未検出コード: {_codes_not_found}")
    if _codes_no_image:
        print(f"   ⚠ フェンス物料画像なしコード: {_codes_no_image}")

    fence_section = (nv_fgg or {}).get('fence') or {}
    gate_section = (nv_fgg or {}).get('gate') or {}
    multi_fences = (nv_fgg or {}).get('fences') or []
    multi_gates = (nv_fgg or {}).get('gates') or []
    fence_style = fence_section.get('style', '')
    height_code = '150'
    height_mm = '1500'

    if multi_fences:
        first_fence = multi_fences[0] if multi_fences else {}
        fence_style = first_fence.get('style', fence_style)
        fence_section = first_fence if not fence_section.get('style') else fence_section

    if fence_style:
        parts = fence_style.split('-')
        if len(parts) >= 2:
            height_code = parts[-1]
        height_mm = _extract_fence_height_mm_from_style(fence_style) or height_mm

    if multi_fences and len(multi_fences) > 1:
        sheet_name = 'フェンス明細'
    else:
        sheet_name = f'フェンス H{height_code}'
    if sheet_name in workbook.sheetnames:
        return {'sheet_name': sheet_name, 'discount_total_row': None}
    ws = workbook.create_sheet(title=sheet_name)

    for col, width in DETAIL_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['F'].width = 20

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 40

    fence_surface = fence_section.get('surface', (nv_fgg or {}).get('surface', '白色'))
    surface_jp = _surface_to_jp_color(fence_surface)
    if multi_gates:
        for g_sec in multi_gates:
            g_color = g_sec.get('gateColor', '')
            if g_color:
                surface_jp = g_color
                break
    elif gate_section:
        g_color = gate_section.get('gateColor', '')
        if g_color:
            surface_jp = g_color
    fence_len = float(fence_section.get('totalLength', 0) or 0)
    if multi_fences:
        fence_len = sum(float(f.get('totalLength', 0) or 0) for f in multi_fences)
    corner_qty = int(fence_section.get('cornerQty', 0) or 0)
    if multi_fences:
        corner_qty = sum(int(f.get('cornerQty', 0) or 0) for f in multi_fences)

    if multi_gates:
        first_gate = multi_gates[0] if multi_gates else {}
        gate_style_code = first_gate.get('gateStyle', '')
        gate_qty_val = sum(int(g.get('gateQty', 0) or 0) for g in multi_gates)
    else:
        gate_style_code = gate_section.get('gateStyle', '')
        gate_qty_val = int(gate_section.get('gateQty', 0) or 0)

    foundation_type = 'コンクリート基礎'
    gate_type_label = ''
    single_gate_qty = 0
    double_gate_qty = 0
    double_gate_width = 4200

    if gate_style_code and len(gate_style_code) >= 6:
        prefix = gate_style_code[:3]
        width_code = gate_style_code[3:6]

        if prefix.endswith('p'):
            foundation_type = '一体式基礎'
        elif prefix.endswith('g'):
            foundation_type = '杭基礎'

        _NV_GATE_TYPE_LABELS = {
            '片開き門扉', '両開き門扉', '引き戸扉', '折りたたみ扉', '伸縮門扉',
        }

        if prefix.lower().startswith('tl'):
            if gate_qty_val > 0:
                gate_type_label = '引き戸扉'
            double_gate_qty = gate_qty_val
            double_gate_width = int(width_code) * 10 if width_code.isdigit() else 2400
        elif prefix.lower().startswith('tf'):
            if gate_qty_val > 0:
                gate_type_label = '折りたたみ扉'
            double_gate_qty = gate_qty_val
            double_gate_width = int(width_code) * 10 if width_code.isdigit() else 2400
        elif width_code == '120':
            single_gate_qty = gate_qty_val
            if gate_qty_val > 0:
                gate_type_label = '片開き門扉'
        else:
            double_gate_qty = gate_qty_val
            db_h_fd, db_w_fd = _lookup_gate_spec_from_db(gate_style_code)
            if db_w_fd is not None:
                double_gate_width = db_w_fd
            else:
                try:
                    double_gate_width = int(width_code) * 10
                except (ValueError, TypeError):
                    double_gate_width = 4200
            if gate_qty_val > 0:
                gate_type_label = '両開き門扉'

    ws.merge_cells('G1:H1')
    ws['G1'] = '見積日：'
    ws['G1'].font = SM_FONT
    ws['G1'].alignment = RIGHT_A
    ws['I1'] = '=TODAY()'
    ws['I1'].number_format = 'yyyy"年"m"月"dd"日"'
    ws['I1'].font = SM_FONT
    ws['I1'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('A2:I2')
    ws['A2'] = 'フェンス御見積書'
    ws['A2'].font = _FK_TITLE_FONT
    ws['A2'].alignment = CENTER

    if os.path.isfile(_LOGO_PATH):
        try:
            _fimg = XLImage(_LOGO_PATH)
            _fimg.width = 160
            _fimg.height = 53
            ws.add_image(_fimg, 'A1')
        except Exception:
            pass

    ws.delete_rows(3)

    ws.merge_cells('H3:I8')
    _height_m = _format_mm_to_meter_text(height_mm) or '1.5'
    _desc_text = (
        f'1.ディップコーディング　{surface_jp}\n'
        f'2.{foundation_type}\n'
        f'3.H={_height_m}m　L={int(fence_len)}ｍ\n'
        f'4.片開き門扉幅W1200　{single_gate_qty}ヶ所\n'
        f'5.両開き門扉幅W4200　{double_gate_qty}ヶ所\n'
        f'6.柱材コーナー部　{corner_qty}箇所\n'
        f'（0枚予備）'
    )
    _cell_desc = ws.cell(row=3, column=8, value=_desc_text)
    _cell_desc.font = SM_FONT
    _cell_desc.alignment = LEFT_A

    bottom_border = Border(bottom=Side(style='thin'))

    project_name = str(matrix_data.get('project_name') or '').strip()

    fence_mitsumori_condition = nv_params.get('mitsumori_condition', 'CIF')
    _fence_mc_label = fence_mitsumori_condition
    if fence_mitsumori_condition == 'CIF_DDP':
        _fence_mc_label = 'CIF+DDP'
    elif fence_mitsumori_condition == 'NV':
        _fence_mc_label = 'DDP（船便）'
    fence_torihiki_condition = nv_params.get('torihiki_condition', 'TTで30%予約金が前払い、B/L発行後引き渡し前に70%お支払')
    if fence_mitsumori_condition == 'NV':
        fence_torihiki_condition = '納品月末締め翌月末払い'
    fence_nounyu_period = '発注後2週後工場から出荷'
    fence_yuko_period = '御見積後1週間'

    def _info_row(row_num, label, value):
        ws.merge_cells(f'A{row_num}:B{row_num}')
        ws[f'A{row_num}'] = label
        ws[f'A{row_num}'].font = SM_FONT
        ws.merge_cells(f'C{row_num}:E{row_num}')
        ws[f'C{row_num}'] = value
        ws[f'C{row_num}'].font = SM_FONT

    _info_row(4, '案件名：', f'{project_name} フェンス H{height_mm}  {int(fence_len)}M （{surface_jp}）')
    ws['C4'].font = _FK_BLUE_FONT
    _info_row(5, '見積条件：', _fence_mc_label)
    _info_row(6, '納入期限：', fence_nounyu_period)
    _info_row(7, '取引条件：', fence_torihiki_condition)
    _info_row(8, '有効期限：', fence_yuko_period)

    for _ir in range(4, 9):
        ws.row_dimensions[_ir].height = 30

    r = 10

    subtitle_parts = [p for p in [foundation_type, gate_type_label] if p]
    subtitle_suffix = '　'.join(subtitle_parts)
    ws.merge_cells(f'A{r}:I{r}')
    ws[f'A{r}'] = f'一、フェンス金額----{subtitle_suffix}' if subtitle_suffix else '一、フェンス金額'
    ws[f'A{r}'].font = SM_FONT
    ws[f'A{r}'].alignment = LEFT_A
    ws.row_dimensions[r].height = 30
    for _sc in range(1, 10):
        ws.cell(row=r, column=_sc).border = THIN_BORDER

    r += 1
    header_row = r
    headers = [
        '部品名称\nProduct Name',
        '材質&表面処理\nMaterial',
        '写真\nPhoto',
        '規格\nSpec.',
        '単価(US＄)\nUnit Price',
        '数量（PCS)\nQTY',
        '総金額(US＄)\nAmount Price',
    ]
    ws.merge_cells(f'A{header_row}:B{header_row}')
    cell_num = ws.cell(row=header_row, column=1, value='番号\nItem No.')
    cell_num.font = SM_FONT
    cell_num.alignment = CENTER
    cell_num.border = THIN_BORDER
    ws.cell(row=header_row, column=2).border = THIN_BORDER
    for i, h in enumerate(headers, 3):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = SM_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 36

    all_items = []
    seq = 1
    for row_data in fence_rows:
        all_items.append({**row_data, '_seq': seq})
        seq += 1
    for row_data in gate_rows:
        all_items.append({**row_data, '_seq': seq})
        seq += 1

    data_start = header_row + 1
    for idx, item in enumerate(all_items):
        row = data_start + idx
        ws.row_dimensions[row].height = 42
        for c in range(1, 10):
            ws.cell(row=row, column=c).border = THIN_BORDER

        ws.merge_cells(f'A{row}:B{row}')
        _set(ws, row, 1, item.get('_seq', idx + 1))

        item_code = str(item.get('code', '')).strip()
        mat_info = fence_material_map.get(item_code)
        if mat_info:
            display_name = mat_info.get('日语名称', '') or item.get('name', '')
        else:
            display_name = item.get('name', '')
        _set(ws, row, 3, display_name)

        if mat_info:
            if coating == '热镀锌':
                display_material = mat_info.get('材質表面処理_热镀锌', '')
            else:
                display_material = mat_info.get('材質表面処理_浸塑', '')
        else:
            display_material = ''
        _set(ws, row, 4, display_material)
        ws.cell(row=row, column=4).fill = _FK_MAT_FILL

        img_inserted = False
        if mat_info and mat_info.get('image_base64'):
            img_bytes, img_ext, img_status = decode_image_base64(mat_info['image_base64'])
            if img_status == 'ready' and img_bytes:
                db_img_name = f"fence_{item_code}.{img_ext}"
                db_img_path = os.path.join(image_temp_dir, db_img_name)
                if db_img_path not in image_cache:
                    with open(db_img_path, 'wb') as _f:
                        _f.write(img_bytes)
                    image_cache.pop(db_img_path, None)
                fit_w, fit_h = _fit_image(ws, row, img_col, img_w, img_h)
                temp_img_path = prepare_image_for_excel(
                    db_img_path,
                    target_width=fit_w,
                    target_height=fit_h,
                    temp_dir=image_temp_dir,
                    cache=image_cache,
                )
                final_img_path = temp_img_path if temp_img_path else db_img_path
                success = add_image_centered_in_cell(
                    ws,
                    final_img_path,
                    row,
                    img_col,
                    img_width=fit_w,
                    img_height=fit_h,
                )
                if success:
                    img_inserted = True
        if not img_inserted:
            _set(ws, row, 5, '/')

        _set(ws, row, 6, item.get('spec', ''))

        unit_price = item.get('unit_price', 0)
        qty = item.get('qty', 0)

        ws.cell(row=row, column=7, value=unit_price).font = SM_FONT
        ws.cell(row=row, column=7).alignment = CENTER
        ws.cell(row=row, column=7).number_format = NUM_FMT

        _set(ws, row, 8, qty)

        ws.cell(row=row, column=9, value=f'=G{row}*H{row}').border = THIN_BORDER
        ws.cell(row=row, column=9).alignment = CENTER
        ws.cell(row=row, column=9).number_format = NUM_FMT
        ws.cell(row=row, column=9).font = SM_FONT

    data_end = data_start + len(all_items) - 1
    _img_ok = sum(1 for item in all_items
                  if fence_material_map.get(str(item.get('code', '')).strip())
                  and fence_material_map[str(item.get('code', '')).strip()].get('image_base64'))
    _img_ng = len(all_items) - _img_ok
    print(f"   🖼️ フェンス明細: 画像あり {_img_ok}件, 画像なし {_img_ng}件")
    discount_total_row = None
    if all_items:
        try:
            fence_discount_rate = Decimal(str(_get_nv_rate_value(nv_params, 'fence_discount_rate', 'nv_fence_discount_rate', 94)))
        except Exception:
            fence_discount_rate = Decimal('94')

        total_row = data_end + 1
        ws.merge_cells(f'A{total_row}:H{total_row}')
        _set(ws, total_row, 1, '合計金額(USD)', font=SM_FONT, align=Alignment(horizontal='right', vertical='center'))
        for c in range(1, 9):
            ws.cell(row=total_row, column=c).border = THIN_BORDER
        ws.cell(row=total_row, column=9, value=f'=SUM(I{data_start}:I{data_end})').border = THIN_BORDER
        ws.cell(row=total_row, column=9).font = SM_FONT
        ws.cell(row=total_row, column=9).alignment = CENTER
        ws.cell(row=total_row, column=9).number_format = NUM_FMT
        ws.row_dimensions[total_row].height = 25

        discount_total_row = total_row + 1
        ws.merge_cells(f'A{discount_total_row}:H{discount_total_row}')
        _set(ws, discount_total_row, 1, '特別値引き後合計金額(USD)', font=SM_FONT, align=Alignment(horizontal='right', vertical='center'))
        for c in range(1, 9):
            ws.cell(row=discount_total_row, column=c).border = THIN_BORDER
        ws.cell(row=discount_total_row, column=9, value=f'=I{total_row}*{fence_discount_rate}/100').border = THIN_BORDER
        ws.cell(row=discount_total_row, column=9).font = SM_FONT
        ws.cell(row=discount_total_row, column=9).alignment = CENTER
        ws.cell(row=discount_total_row, column=9).number_format = NUM_FMT
        ws.row_dimensions[discount_total_row].height = 25

    apply_print_setup(ws, 'ja_normal')

    return {'sheet_name': sheet_name, 'discount_total_row': discount_total_row}
