from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XlImage
from openpyxl.worksheet.page import PageMargins
from backend.core.print_settings import apply_print_setup
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
import os
from decimal import Decimal

from backend.core.shared.price_utils import round_to_2_decimal

SM_FONT = Font(name='微软雅黑', size=12)
SM_FONT_BOLD = Font(name='微软雅黑', size=12, bold=True)
SM_FONT_BOLD_BLUE = Font(name='微软雅黑', size=12, bold=True, color='FF2630F0')
SM_FONT_BOLD_10 = Font(name='微软雅黑', size=12)
SM_FONT_HEADER = Font(name='微软雅黑', size=12)
SM_FONT_RED = Font(name='微软雅黑', size=12, bold=True, color='FF0000')
SM_FONT_RED_10 = Font(name='微软雅黑', size=12, bold=True, color='FF0000')


_FK_LABEL_BLUE = InlineFont(rFont='微软雅黑', sz=12, b=True, color='FF0000FF')
_FK_LABEL_NORMAL = InlineFont(rFont='微软雅黑', sz=12)


def _fk_label(suffix):
    return CellRichText([
        TextBlock(_FK_LABEL_BLUE, 'フェンス+架台'),
        TextBlock(_FK_LABEL_NORMAL, suffix),
    ])

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
THICK_SIDE = Side(style='medium', color='000000')

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT_A = Alignment(horizontal='right', vertical='center', wrap_text=True)
LEFT_A = Alignment(horizontal='left', vertical='center', wrap_text=True)

NUM_FMT = '"US$" #,##0.00'
NUM_FMT = '#,##0.00'

BLUE_FILL = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid')
HEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
GREEN_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

SUMMARY_COL_WIDTHS = {
    'A': 7, 'B': 7, 'C': 11, 'D': 11, 'E': 11,
    'F': 11, 'G': 16, 'H': 16, 'I': 16, 'J': 20, 'K': 16, 'L': 16,
}
def create_normal_summary_sheet(workbook, detail_results, matrix_data=None,
                                 image_path=None, fence_data=None,
                                 normal_params=None, nv_fence_gate_data=None,
                                 pile_summary=None,
                                 image_temp_dir=None, image_cache=None,
                                 sheet_title=None):
    from datetime import datetime

    normal_params = normal_params or {}
    discount_rate_pct = normal_params.get('discount_rate', 71)
    fence_discount_rate = normal_params.get('fence_discount_rate', 94)
    consumption_tax_pct = normal_params.get('consumption_tax', 10)
    tariff_rate_pct = normal_params.get('tariff_rate', 3)
    shipping_fee = normal_params.get('shipping_fee', 0)

    ws = workbook.create_sheet(title=sheet_title or '合計')

    for col, width in SUMMARY_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    center = CENTER
    right_a = RIGHT_A
    left_a = LEFT_A
    thin_border = THIN_BORDER

    matrix_data = matrix_data or {}
    project_name = str(matrix_data.get('project_name') or '').strip()
    output_kw = matrix_data.get('output_kw') or 0
    module_wattage = matrix_data.get('module_wattage') or 0

    fence_data = fence_data or {}

    _base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
    _logo = os.path.join(_base_dir, 'input', '集团标2.png')
    logo_to_use = _logo if os.path.isfile(_logo) else image_path
    if logo_to_use and os.path.isfile(str(logo_to_use)):
        try:
            img = XlImage(str(logo_to_use))
            img.width = 302
            img.height = 134
            ws.add_image(img, 'A1')
        except Exception:
            pass

    ws['J1'] = '見積日：'
    ws['J1'].font = SM_FONT
    ws['J1'].alignment = center
    ws.merge_cells('K1:L1')
    ws['K1'] = datetime.now().date()
    ws['K1'].font = SM_FONT
    ws['K1'].number_format = 'yyyy"年"m"月"d"日"'
    ws['K1'].alignment = center
    ws.row_dimensions[1].height = 30

    ws['D2'] = '架台御見積書'
    ws.merge_cells('D2:J2')
    ws['D2'].font = Font(name='微软雅黑', size=18, bold=True)
    ws['D2'].alignment = center
    ws.row_dimensions[2].height = 91
    for _r in range(3, 8):
        ws.row_dimensions[_r].height = 30

    ws['A3'] = '案件名：'
    ws['A3'].font = SM_FONT
    ws['A3'].alignment = center
    ws.merge_cells('A3:B3')
    ws.merge_cells('C3:H3')
    ws['C3'] = project_name
    ws['C3'].font = SM_FONT
    ws['C3'].alignment = left_a

    ws['A4'] = '見積条件：'
    ws['A4'].font = SM_FONT
    ws['A4'].alignment = center
    ws.merge_cells('A4:B4')
    ws.merge_cells('C4:H4')
    mitsumori_condition = normal_params.get('mitsumori_condition', '')
    _MC_LABEL = {
        'CIF': 'CIF（船便）',
        'DDP': 'DDP現場（船便）',
        'CIF_DDP': 'DDP現場（船便）\nCIF（船便）',
    }
    mitsumori_label_display = _MC_LABEL.get(mitsumori_condition, mitsumori_condition)
    if mitsumori_label_display:
        ws['C4'] = mitsumori_label_display
        ws['C4'].font = SM_FONT
        ws['C4'].alignment = left_a
        if mitsumori_condition == 'CIF_DDP':
            ws.row_dimensions[4].height = 36

    ws['A5'] = '納入期限：'
    ws['A5'].font = SM_FONT
    ws['A5'].alignment = center
    ws.merge_cells('A5:B5')
    ws['C5'] = '発注後10-14日間後工場から出荷'
    ws['C5'].font = SM_FONT
    ws['C5'].alignment = left_a
    ws.merge_cells('C5:H5')

    ws['A6'] = '取引条件：'
    ws['A6'].font = SM_FONT
    ws['A6'].alignment = center
    ws.merge_cells('A6:B6')
    _trade_condition = normal_params.get('trade_condition') or '取引基本契約書に基づく'
    ws['C6'] = _trade_condition
    ws['C6'].font = SM_FONT
    ws['C6'].alignment = left_a
    ws.merge_cells('C6:H6')

    ws['A7'] = '有効期限：'
    ws['A7'].font = SM_FONT
    ws['A7'].alignment = center
    ws.merge_cells('A7:B7')
    ws['C7'] = '御見積後1日間'
    ws['C7'].font = SM_FONT
    ws['C7'].alignment = left_a
    ws.merge_cells('C7:H7')

    r8 = 8
    ws[f'A{r8}'] = '一、架台本体金額（Ex Works）'
    ws[f'A{r8}'].font = SM_FONT_BOLD_10
    ws[f'A{r8}'].alignment = left_a
    ws.merge_cells(f'A{r8}:L{r8}')
    ws.row_dimensions[r8].height = 36
    thin_s = Side(style='thin')
    ws.cell(row=r8, column=1).border = Border(left=THICK_SIDE, top=THICK_SIDE, right=thin_s, bottom=thin_s)
    for c in range(2, 13):
        ws.cell(row=r8, column=c).border = Border(left=thin_s, top=THICK_SIDE, right=THICK_SIDE if c == 12 else thin_s, bottom=thin_s)

    r9 = r8 + 1
    ws[f'A{r9}'] = '序号'
    ws[f'A{r9}'].font = SM_FONT_HEADER
    ws[f'A{r9}'].alignment = center

    ws.merge_cells(f'B{r9}:C{r9}')
    ws[f'B{r9}'] = 'パネル数'
    ws[f'B{r9}'].font = SM_FONT_HEADER
    ws[f'B{r9}'].alignment = center

    ws[f'D{r9}'] = 'セット数'
    ws[f'D{r9}'].font = SM_FONT_HEADER
    ws[f'D{r9}'].alignment = center

    ws[f'E{r9}'] = 'パワコン'
    ws[f'E{r9}'].font = SM_FONT_HEADER
    ws[f'E{r9}'].alignment = center
    ws[f'E{r9}'].fill = BLUE_FILL

    ws[f'F{r9}'] = '集電箱'
    ws[f'F{r9}'].font = SM_FONT_HEADER
    ws[f'F{r9}'].alignment = center

    ws[f'G{r9}'] = '角度'
    ws[f'G{r9}'].font = SM_FONT_HEADER
    ws[f'G{r9}'].alignment = center
    ws[f'G{r9}'].fill = BLUE_FILL

    ws[f'H{r9}'] = '発電量（KW）'
    ws[f'H{r9}'].font = SM_FONT_HEADER
    ws[f'H{r9}'].alignment = center

    ws[f'I{r9}'] = '単価(USD)/基'
    ws[f'I{r9}'].font = SM_FONT_HEADER
    ws[f'I{r9}'].alignment = center

    ws[f'J{r9}'] = '特別値引き後金額'
    ws[f'J{r9}'].font = SM_FONT_HEADER
    ws[f'J{r9}'].alignment = center
    ws[f'J{r9}'].fill = BLUE_FILL

    ws[f'K{r9}'] = '総金額(USD)'
    ws[f'K{r9}'].font = SM_FONT_HEADER
    ws[f'K{r9}'].alignment = center

    ws[f'L{r9}'] = 'ワットあたりの価格(USD)'
    ws[f'L{r9}'].font = SM_FONT_HEADER
    ws[f'L{r9}'].alignment = right_a
    ws.row_dimensions[r9].height = 36

    data_row_start = r9 + 1
    grand_total = Decimal('0')

    for i, detail in enumerate(detail_results):
        row = data_row_start + i
        ws.row_dimensions[row].height = 36
        arr = detail.get('array_info', {})
        base_count = arr.get('table_qty', 1)
        panel_count = arr.get('rows', 0) * arr.get('cols', 0)
        detail_config = detail.get('config') or {}
        info_missing = arr.get('missing_per_table', 0) or 0
        bom_missing = detail_config.get('missing_boards', 0) or 0
        missing_boards = info_missing if info_missing else (bom_missing // base_count if base_count and bom_missing else bom_missing)
        note = arr.get('note') or ''

        ws.cell(row=row, column=1, value=i + 1).font = SM_FONT
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=2, value=arr.get('rows', 0)).font = SM_FONT
        ws.cell(row=row, column=2).alignment = center
        ws.cell(row=row, column=2).number_format = '0"段"'
        ws.cell(row=row, column=3, value=arr.get('cols', 0)).font = SM_FONT
        ws.cell(row=row, column=3).alignment = center
        ws.cell(row=row, column=3).number_format = '0"列"'
        ws.cell(row=row, column=4, value=base_count).font = SM_FONT
        ws.cell(row=row, column=4).alignment = center
        missing_val = missing_boards if missing_boards else '/'
        ws.cell(row=row, column=5, value=missing_val).font = SM_FONT
        ws.cell(row=row, column=5).alignment = center
        ws.cell(row=row, column=5).fill = BLUE_FILL
        angle = detail.get('angle', '')
        angle_clean = str(angle).rstrip('°').strip() if angle else ''
        angle_text = f'{angle_clean}°' if angle_clean else ''
        ws.cell(row=row, column=6).font = SM_FONT
        ws.cell(row=row, column=6).alignment = center
        ws.cell(row=row, column=7, value=angle_text).font = SM_FONT
        ws.cell(row=row, column=7).alignment = center
        ws.cell(row=row, column=7).fill = BLUE_FILL
        _miss_op = '-' if missing_boards >= 0 else '+'
        ws.cell(row=row, column=8, value=f'=(B{row}*C{row}{_miss_op}{missing_boards})*D{row}*{module_wattage}/1000').font = SM_FONT
        ws.cell(row=row, column=8).alignment = center
        ws.cell(row=row, column=8).number_format = '#,##0.00'

        detail_sheet_name = detail.get('sheet_name', '')
        sub_total_row = detail.get('sub_total_row', 24)
        ws.cell(row=row, column=9, value=f"='{detail_sheet_name}'!J{sub_total_row}").font = SM_FONT
        ws.cell(row=row, column=9).number_format = '#,##0.00'
        ws.cell(row=row, column=9).alignment = center
        ws.cell(row=row, column=10, value=f'=I{row}*{discount_rate_pct}/100').font = SM_FONT
        ws.cell(row=row, column=10).number_format = '#,##0.00'
        ws.cell(row=row, column=10).alignment = center
        ws.cell(row=row, column=10).fill = BLUE_FILL
        ws.cell(row=row, column=11, value=f'=J{row}*D{row}').font = SM_FONT
        ws.cell(row=row, column=11).number_format = '#,##0.00'
        ws.cell(row=row, column=11).alignment = center
        ws.cell(row=row, column=12, value=f'=K{row}/H{row}/1000').font = SM_FONT
        ws.cell(row=row, column=12).number_format = '#,##0.0000'
        ws.cell(row=row, column=12).alignment = right_a

        for c in range(1, 13):
            ws.cell(row=row, column=c).border = thin_border

    data_end = data_row_start + len(detail_results) - 1

    r_sub = data_end + 1
    ws[f'A{r_sub}'] = '架台総金額(USD)'
    ws[f'A{r_sub}'].font = SM_FONT
    ws[f'A{r_sub}'].alignment = right_a
    ws.merge_cells(f'A{r_sub}:K{r_sub}')
    ws[f'L{r_sub}'] = f'=SUM(K{data_row_start}:K{data_end})'
    ws[f'L{r_sub}'].font = SM_FONT
    ws[f'L{r_sub}'].alignment = right_a
    ws[f'L{r_sub}'].number_format = NUM_FMT
    for c in range(1, 13):
        ws.cell(row=r_sub, column=c).fill = BLUE_FILL
    ws.row_dimensions[r_sub].height = 36

    r_pile = r_sub + 1

    pile_label = 'スクリュー杭'

    ws[f'A{r_pile}'] = pile_label
    ws[f'A{r_pile}'].font = SM_FONT
    ws[f'A{r_pile}'].alignment = center
    ws.merge_cells(f'A{r_pile}:D{r_pile}')
    ws[f'E{r_pile}'] = '単価'
    ws[f'E{r_pile}'].font = SM_FONT
    ws[f'E{r_pile}'].alignment = center
    ws.merge_cells(f'E{r_pile}:F{r_pile}')
    ws.merge_cells(f'G{r_pile}:H{r_pile}')
    ws[f'I{r_pile}'] = '数量'
    ws[f'I{r_pile}'].font = SM_FONT
    ws[f'I{r_pile}'].alignment = center
    ws[f'K{r_pile}'] = '金額'
    ws[f'K{r_pile}'].font = SM_FONT
    ws[f'K{r_pile}'].alignment = center
    ws.row_dimensions[r_pile].height = 36

    pile_qty_parts = []
    _first_pile_ref = ''
    for d in detail_results:
        ps = d.get('pile_data_start_row', 0)
        pe = d.get('pile_data_end_row', 0)
        sn = d.get('sheet_name', '')
        tqty = d.get('array_info', {}).get('table_qty', 1)
        if ps and pe and sn:
            pile_qty_parts.append(f"SUM('{sn}'!I{ps}:I{pe})*{tqty}")
            if not _first_pile_ref:
                _first_pile_ref = f"'{sn}'!H{ps}"

    ws[f'G{r_pile}'].font = SM_FONT
    ws[f'G{r_pile}'].alignment = center
    ws[f'G{r_pile}'].number_format = NUM_FMT
    ws[f'J{r_pile}'].font = SM_FONT
    ws[f'J{r_pile}'].alignment = center
    ws[f'L{r_pile}'].font = SM_FONT
    ws[f'L{r_pile}'].alignment = right_a
    ws[f'L{r_pile}'].number_format = NUM_FMT

    if _first_pile_ref:
        ws[f'G{r_pile}'] = f'={_first_pile_ref}'
    else:
        ws[f'G{r_pile}'] = 0

    if pile_qty_parts:
        ws[f'J{r_pile}'] = f'={"+".join(pile_qty_parts)}'
    else:
        ws[f'J{r_pile}'] = 0

    ws[f'L{r_pile}'] = f'=G{r_pile}*J{r_pile}'

    r_1 = r_pile + 1
    ws[f'A{r_1}'] = '総金額(USD)'
    ws[f'A{r_1}'].font = SM_FONT
    ws[f'A{r_1}'].alignment = right_a
    ws.merge_cells(f'A{r_1}:J{r_1}')
    ws.merge_cells(f'K{r_1}:L{r_1}')
    ws[f'K{r_1}'] = f'=L{r_sub}+L{r_pile}'
    ws[f'K{r_1}'].font = SM_FONT
    ws[f'K{r_1}'].alignment = right_a
    ws[f'K{r_1}'].number_format = '#,##0.00'
    for c in range(1, 13):
        ws.cell(row=r_1, column=c).fill = BLUE_FILL
    ws.row_dimensions[r_1].height = 36

    r_kw = r_1 + 1
    ws[f'A{r_kw}'] = '発電量(KW)'
    ws[f'A{r_kw}'].font = SM_FONT
    ws[f'A{r_kw}'].alignment = right_a
    ws.merge_cells(f'A{r_kw}:J{r_kw}')
    ws.merge_cells(f'K{r_kw}:L{r_kw}')
    ws[f'K{r_kw}'] = f'=SUM(H{data_row_start}:H{data_end})'
    ws[f'K{r_kw}'].font = SM_FONT
    ws[f'K{r_kw}'].alignment = right_a
    ws[f'K{r_kw}'].number_format = '#,##0.00'
    ws.row_dimensions[r_kw].height = 36

    r_wp = r_kw + 1
    ws[f'A{r_wp}'] = 'ワットあたりの価格'
    ws[f'A{r_wp}'].font = SM_FONT
    ws[f'A{r_wp}'].alignment = right_a
    ws.merge_cells(f'A{r_wp}:J{r_wp}')
    ws.merge_cells(f'K{r_wp}:L{r_wp}')
    ws[f'K{r_wp}'] = f'=K{r_1}/K{r_kw}/1000'
    ws[f'K{r_wp}'].font = SM_FONT
    ws[f'K{r_wp}'].alignment = right_a
    ws[f'K{r_wp}'].number_format = '#,##0.0000'
    ws.row_dimensions[r_wp].height = 36

    r_fence = r_wp + 1
    nv_fgg = nv_fence_gate_data or {}
    fence_data = fence_data or {}
    fence_title_text = '二、フェンスの金額（Ex Works）'

    detail_fence_rows = []
    detail_gate_rows = []
    fence_height_mm = ''
    fence_len_m = 0
    fence_color_jp = '白色'
    gate_type_label = '片開き門'
    GATE_TYPE_LABEL_MAP = {
        'single': '片開き門', 'double': '両開き門',
        'sliding': '引き戸', 'folding': '折りたたみ扉',
        'telescopic': '伸縮門', 'custom': 'カスタム門',
    }

    def _resolve_gate_label(style_code):
        if not style_code or len(style_code) < 6:
            return '片開き門'
        prefix = style_code[:3].lower()
        w_code = style_code[3:6]
        if prefix.startswith('tl'):
            gt = 'sliding'
        elif prefix.startswith('tf'):
            gt = 'folding'
        elif prefix.startswith('tx'):
            gt = 'custom'
        elif w_code == '120':
            gt = 'single'
        else:
            gt = 'double'
        return GATE_TYPE_LABEL_MAP.get(gt, '片開き門')

    gate_qty_val = 0
    fence_subtotal_amount = Decimal('0')

    multi_fences = nv_fgg.get('fences') if isinstance(nv_fgg, dict) else None
    multi_gates = nv_fgg.get('gates') if isinstance(nv_fgg, dict) else None

    if nv_fgg and (multi_fences or multi_gates):
        fence_color_jp = '白色'
        fence_surface = nv_fgg.get('surface', '白色')
        if fence_surface and '茶' in fence_surface:
            fence_color_jp = '茶色'

        if multi_fences:
            for fi, f_sec in enumerate(multi_fences):
                f_rows = f_sec.get('rows', [])
                detail_fence_rows.extend(f_rows)
                f_len = int(float(f_sec.get('totalLength', 0) or 0))
                if fi == 0:
                    fence_len_m = f_len
                else:
                    fence_len_m += f_len
                f_style = f_sec.get('style', '')
                if fi == 0 and f_style and '-' in f_style:
                    parts = f_style.split('-')
                    height_code = parts[-1] if parts else ''
                    if height_code:
                        try:
                            fence_height_mm = str(int(height_code) * 10)
                        except (ValueError, TypeError):
                            fence_height_mm = ''

        if multi_gates:
            for gi, g_sec in enumerate(multi_gates):
                g_rows = g_sec.get('rows', [])
                detail_gate_rows.extend(g_rows)
                g_style = g_sec.get('gateStyle', '')
                g_qty = int(g_sec.get('gateQty', 0) or 0)
                gate_qty_val += g_qty
                g_color = g_sec.get('gateColor', '')
                if g_color:
                    fence_color_jp = g_color
                if g_style and len(g_style) >= 6:
                    gate_type_label = _resolve_gate_label(g_style)

        fence_amount_raw = Decimal('0')
        if detail_fence_rows:
            fence_amount_raw += Decimal(str(sum(r.get('amount', 0) for r in detail_fence_rows)))
        if detail_gate_rows:
            fence_amount_raw += Decimal(str(sum(r.get('amount', 0) for r in detail_gate_rows)))
        fence_subtotal_amount = fence_amount_raw

    elif nv_fgg:
        fence_section = nv_fgg.get('fence') or {}
        gate_section = nv_fgg.get('gate') or {}
        f_rows = fence_section.get('rows', [])
        g_rows = gate_section.get('rows', [])
        detail_fence_rows = f_rows
        detail_gate_rows = g_rows
        fence_style = fence_section.get('style', '')
        fence_surface = fence_section.get('surface', '白色')
        fence_color_jp = '茶色' if '茶' in fence_surface else '白色'
        fence_len_m = int(float(fence_section.get('totalLength', 0) or 0))

        if fence_style and '-' in fence_style:
            parts = fence_style.split('-')
            height_code = parts[-1] if parts else ''
            if height_code:
                try:
                    fence_height_mm = str(int(height_code) * 10)
                except (ValueError, TypeError):
                    fence_height_mm = ''

        gate_style_code = gate_section.get('gateStyle', '')
        gate_qty_val = int(gate_section.get('gateQty', 0) or 0)
        gate_color = gate_section.get('gateColor', '')
        if gate_color:
            fence_color_jp = gate_color
        if gate_style_code and len(gate_style_code) >= 6:
            gate_type_label = _resolve_gate_label(gate_style_code)

        fence_amount_raw = Decimal('0')
        if f_rows:
            fence_amount_raw += Decimal(str(sum(r.get('amount', 0) for r in f_rows)))
        if g_rows:
            fence_amount_raw += Decimal(str(sum(r.get('amount', 0) for r in g_rows)))
        fence_subtotal_amount = fence_amount_raw

    elif fence_data:
        fence_length = str(fence_data.get('length', 140)).replace('M', '').replace('m', '').strip()
        try:
            fence_len_m = int(float(fence_length or '0'))
        except (ValueError, TypeError):
            fence_len_m = 0

    _show_ddp = mitsumori_condition in ('DDP', 'CIF_DDP')
    _show_cif = mitsumori_condition in ('CIF', 'CIF_DDP')
    if not _show_ddp and not _show_cif:
        _show_cif = True

    has_fence_data = bool(detail_fence_rows or detail_gate_rows) or fence_subtotal_amount > 0

    if has_fence_data:
        ws[f'A{r_fence}'] = fence_title_text
        ws[f'A{r_fence}'].font = SM_FONT_BOLD_10
        ws[f'A{r_fence}'].alignment = left_a
        ws.merge_cells(f'A{r_fence}:L{r_fence}')
        ws.row_dimensions[r_fence].height = 36
        for c in range(1, 13):
            ws.cell(row=r_fence, column=c).fill = ORANGE_FILL

        r_fs = r_fence + 1
        height_display = f'H{fence_height_mm}' if fence_height_mm else 'H'
        has_fence = fence_subtotal_amount > 0

        fence_detail_info = None
        if detail_fence_rows or detail_gate_rows:
            try:
                from backend.core.ja_normal.fence_detail_sheet import create_normal_fence_detail_sheet
                fence_detail_info = create_normal_fence_detail_sheet(workbook, detail_fence_rows, detail_gate_rows, nv_fgg,
                                           matrix_data=matrix_data, nv_params=normal_params,
                                           image_temp_dir=image_temp_dir, image_cache=image_cache)
            except Exception as e:
                print(f"   ⚠ フェンス物料明細シート生成失敗: {e}")

        fence_discount_ref = ''
        if has_fence and fence_detail_info and fence_detail_info.get('discount_total_row'):
            ds_name = fence_detail_info.get('sheet_name', '').replace("'", "''")
            ds_row = fence_detail_info['discount_total_row']
            fence_discount_ref = f"'{ds_name}'!I{ds_row}"

        fence_desc = f'【フェンス {height_display}  {fence_len_m}M    {gate_type_label}{gate_qty_val}セット  {fence_color_jp}】  フェンス総金額(USD)'
        ws[f'A{r_fs}'] = fence_desc
        ws[f'A{r_fs}'].font = SM_FONT_BOLD_BLUE
        ws[f'A{r_fs}'].alignment = right_a
        ws.merge_cells(f'A{r_fs}:J{r_fs}')
        ws.merge_cells(f'K{r_fs}:L{r_fs}')
        if fence_discount_ref:
            ws[f'K{r_fs}'] = f'={fence_discount_ref}'
        else:
            ws[f'K{r_fs}'] = float(fence_subtotal_amount) if fence_subtotal_amount > 0 else 0
        ws[f'K{r_fs}'].font = SM_FONT
        ws[f'K{r_fs}'].alignment = right_a
        ws[f'K{r_fs}'].number_format = NUM_FMT
        for c in range(1, 13):
            ws.cell(row=r_fs, column=c).border = thin_border
            ws.cell(row=r_fs, column=c).fill = ORANGE_FILL
        ws.row_dimensions[r_fs].height = 36

        fence_total_ref = f'K{r_fs}'
        cur = r_fs + 1
        _sec_no = '三'
    else:
        fence_total_ref = '0'
        cur = r_fence
        _sec_no = '二'

    delivery_location = str(normal_params.get('delivery_location', '') or '').strip()
    _loc_segment = f'  {delivery_location}  ' if delivery_location else '  '

    if _show_ddp:
        r_ddp_title = cur
        ddp_title_text = f'{_sec_no}、DDP現場（船便）'
        if mitsumori_condition == 'CIF_DDP':
            ddp_title_text = f'{_sec_no}、DDP現場（船便）--(税金＆消費税含み場合）'
        ws[f'A{r_ddp_title}'] = ddp_title_text
        ws[f'A{r_ddp_title}'].font = SM_FONT
        ws[f'A{r_ddp_title}'].alignment = left_a
        ws.merge_cells(f'A{r_ddp_title}:L{r_ddp_title}')
        ws.row_dimensions[r_ddp_title].height = 36
        if mitsumori_condition == 'CIF_DDP':
            for c in range(1, 13):
                ws.cell(row=r_ddp_title, column=c).fill = BLUE_FILL
        cur += 1

        r_ddp_tax = cur
        ws[f'A{r_ddp_tax}'] = _fk_label(f' 製品金額＊{consumption_tax_pct + tariff_rate_pct}％（消費税＋関税）')
        ws[f'A{r_ddp_tax}'].font = SM_FONT
        ws[f'A{r_ddp_tax}'].alignment = right_a
        ws.merge_cells(f'A{r_ddp_tax}:J{r_ddp_tax}')
        ws.merge_cells(f'K{r_ddp_tax}:L{r_ddp_tax}')
        ws[f'K{r_ddp_tax}'] = f'=K{r_1}*{(consumption_tax_pct + tariff_rate_pct) / 100}+{fence_total_ref}*{consumption_tax_pct / 100}'
        ws[f'K{r_ddp_tax}'].font = SM_FONT
        ws[f'K{r_ddp_tax}'].alignment = right_a
        ws[f'K{r_ddp_tax}'].number_format = NUM_FMT
        ws.row_dimensions[r_ddp_tax].height = 36
        cur += 1

        r_ddp_ship = cur
        ws[f'A{r_ddp_ship}'] = _fk_label(f'{_loc_segment}現場までの全て費用\n4Tユニック 配送')
        ws[f'A{r_ddp_ship}'].font = SM_FONT
        ws[f'A{r_ddp_ship}'].alignment = right_a
        ws.merge_cells(f'A{r_ddp_ship}:J{r_ddp_ship}')
        ws.merge_cells(f'K{r_ddp_ship}:L{r_ddp_ship}')
        ws[f'K{r_ddp_ship}'] = shipping_fee
        ws[f'K{r_ddp_ship}'].font = SM_FONT
        ws[f'K{r_ddp_ship}'].alignment = right_a
        ws[f'K{r_ddp_ship}'].number_format = NUM_FMT
        ws.row_dimensions[r_ddp_ship].height = 36
        cur += 1

        r_ddp_total = cur
        ws[f'A{r_ddp_total}'] = _fk_label(' 総金額(USD)')
        ws[f'A{r_ddp_total}'].font = SM_FONT
        ws[f'A{r_ddp_total}'].alignment = right_a
        ws.merge_cells(f'A{r_ddp_total}:J{r_ddp_total}')
        ws.merge_cells(f'K{r_ddp_total}:L{r_ddp_total}')
        ws[f'K{r_ddp_total}'] = f'=K{r_1}+{fence_total_ref}+K{r_ddp_tax}+K{r_ddp_ship}'
        ws[f'K{r_ddp_total}'].font = SM_FONT
        ws[f'K{r_ddp_total}'].alignment = right_a
        ws[f'K{r_ddp_total}'].number_format = NUM_FMT
        ws.row_dimensions[r_ddp_total].height = 36
        cur += 1

        for rr in range(r_ddp_tax, r_ddp_total + 1):
            for c in range(1, 13):
                ws.cell(row=rr, column=c).fill = BLUE_FILL
        _sec_no = {'二': '三', '三': '四', '四': '五'}.get(_sec_no, '四')

    if _show_cif:
        r_cif_title = cur
        ws[f'A{r_cif_title}'] = f'{_sec_no}、CIF（船便）'
        ws[f'A{r_cif_title}'].font = SM_FONT
        ws[f'A{r_cif_title}'].alignment = left_a
        ws.merge_cells(f'A{r_cif_title}:L{r_cif_title}')
        ws.row_dimensions[r_cif_title].height = 36
        if mitsumori_condition == 'CIF_DDP':
            for c in range(1, 13):
                ws.cell(row=r_cif_title, column=c).fill = ORANGE_FILL
        cur += 1

        r_cif_ship = cur
        ws[f'A{r_cif_ship}'] = '海上運賃（USD)'
        ws[f'A{r_cif_ship}'].font = SM_FONT
        ws[f'A{r_cif_ship}'].alignment = right_a
        ws.merge_cells(f'A{r_cif_ship}:J{r_cif_ship}')
        ws.merge_cells(f'K{r_cif_ship}:L{r_cif_ship}')
        ws[f'K{r_cif_ship}'] = shipping_fee
        ws[f'K{r_cif_ship}'].font = SM_FONT
        ws[f'K{r_cif_ship}'].alignment = right_a
        ws[f'K{r_cif_ship}'].number_format = NUM_FMT
        ws.row_dimensions[r_cif_ship].height = 36
        cur += 1

        r_cif_total = cur
        ws[f'A{r_cif_total}'] = _fk_label(' 総金額(USD)')
        ws[f'A{r_cif_total}'].font = SM_FONT
        ws[f'A{r_cif_total}'].alignment = right_a
        ws.merge_cells(f'A{r_cif_total}:J{r_cif_total}')
        ws.merge_cells(f'K{r_cif_total}:L{r_cif_total}')
        ws[f'K{r_cif_total}'] = f'=K{r_1}+{fence_total_ref}+K{r_cif_ship}'
        ws[f'K{r_cif_total}'].font = SM_FONT
        ws[f'K{r_cif_total}'].alignment = right_a
        ws[f'K{r_cif_total}'].number_format = NUM_FMT
        ws.row_dimensions[r_cif_total].height = 36
        cur += 1

        _cif_fill = ORANGE_FILL if mitsumori_condition == 'CIF_DDP' else BLUE_FILL
        for rr in range(r_cif_ship, r_cif_total + 1):
            for c in range(1, 13):
                ws.cell(row=rr, column=c).fill = _cif_fill

    # ========== 新的边框设置（参考成功代码） ==========
    table_start = r8
    table_end = cur - 1
    left_col = 1
    right_col = 12
    
    thin_side = Side(style='thin', color='000000')
    thick_side = Side(style='medium', color='000000')
    
    # 1. 先设置所有内部细边框（跳过合并单元格的非起始单元格）
    for row in range(table_start, table_end + 1):
        for col in range(left_col, right_col + 1):
            cell = ws.cell(row=row, column=col)
            # 检查是否是合并单元格中的非起始单元格
            is_skip = False
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range and cell.coordinate != merged_range.start_cell.coordinate:
                    is_skip = True
                    break
            if not is_skip:
                cell.border = Border(
                    left=thin_side, right=thin_side,
                    top=thin_side, bottom=thin_side
                )
    
    # 2. 设置左侧粗边框
    for row in range(table_start, table_end + 1):
        target_cell = ws.cell(row=row, column=left_col)
        # 如果是合并单元格，取起始单元格
        for merged_range in ws.merged_cells.ranges:
            if target_cell.coordinate in merged_range:
                target_cell = merged_range.start_cell
                break
        old_border = target_cell.border
        if old_border is None:
            old_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        target_cell.border = Border(
            left=thick_side,
            right=old_border.right,
            top=old_border.top,
            bottom=old_border.bottom
        )
    
    # 3. 设置右侧粗边框
    for row in range(table_start, table_end + 1):
        target_cell = ws.cell(row=row, column=right_col)
        for merged_range in ws.merged_cells.ranges:
            if target_cell.coordinate in merged_range:
                target_cell = merged_range.start_cell
                break
        old_border = target_cell.border
        if old_border is None:
            old_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        target_cell.border = Border(
            left=old_border.left,
            right=thick_side,
            top=old_border.top,
            bottom=old_border.bottom
        )
    
    # 4. 设置顶部粗边框
    for col in range(left_col, right_col + 1):
        target_cell = ws.cell(row=table_start, column=col)
        for merged_range in ws.merged_cells.ranges:
            if target_cell.coordinate in merged_range:
                target_cell = merged_range.start_cell
                break
        old_border = target_cell.border
        if old_border is None:
            old_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        target_cell.border = Border(
            left=old_border.left,
            right=old_border.right,
            top=thick_side,
            bottom=old_border.bottom
        )
    
    # 5. 设置底部粗边框
    for col in range(left_col, right_col + 1):
        target_cell = ws.cell(row=table_end, column=col)
        for merged_range in ws.merged_cells.ranges:
            if target_cell.coordinate in merged_range:
                target_cell = merged_range.start_cell
                break
        old_border = target_cell.border
        if old_border is None:
            old_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        target_cell.border = Border(
            left=old_border.left,
            right=old_border.right,
            top=old_border.top,
            bottom=thick_side
        )
            

    sheet_names = workbook.sheetnames
    if '合計' in sheet_names:
        idx = sheet_names.index('合計')
        workbook.move_sheet('合計', offset=-idx)

    apply_print_setup(ws, 'ja_normal')

    return ws.title
