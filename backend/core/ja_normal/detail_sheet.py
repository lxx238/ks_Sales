"""Normal案件用 物料明细表（明细シート）生成。

NV の create_nv_detail_sheet を複製して normal 専用に独立化したもの。
ヘッダ（抬头）や塗りつぶし（颜色填充）の調整はこの関数内で行うことで、
NV 本体（create_nv_detail_sheet）には一切影響しない。
共有ヘルパー・定数は NV モジュールから import して再利用する。
"""
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins
from backend.core.print_settings import apply_print_setup
from decimal import Decimal
import os
import re

from backend.core.shared.price_utils import (
    resolve_price_info, has_valid_price_info, round_to_2_decimal,
    get_temp_adjusted_base_price,
    get_temp_base_price, apply_temp_preinstall_adjustment,
)
from backend.core.shared.text_utils import normalize_lookup_code
from backend.core.shared.product_utils import _is_valid_product_code, normalize_preinstall
from backend.core.shared.weight_utils import extract_length_from_spec
from backend.core.material_translate import translate_material

# 共有ヘルパー・定数は NV モジュールから再利用（これらは変更しない）
from backend.core.ja_nv.quotation_engine import (
    _strip_decimal_zero, _insert_logo,
    _format_pile_spec_display, _resolve_pile_price,
    THIN_BORDER, CENTER, LEFT_A, RIGHT_A,
    NUM_FMT, DETAIL_COL_WIDTHS,
    DETAIL_LOGO_WIDTH, DETAIL_LOGO_HEIGHT,
    _LOGO_PATH, LIGHT_BLUE_FILL,
)

# normal 案件のフォントは全て「微软雅黑」に統一（タイトル=16、それ以外=12）
SM_FONT = Font(name='微软雅黑', size=12)
SM_FONT_BOLD = Font(name='微软雅黑', size=12, bold=True)
FONT_10_BOLD = Font(name='微软雅黑', size=12, bold=True)
TITLE_FONT = Font(name='微软雅黑', size=16, bold=True)


def _set(ws, r, c, val, font=SM_FONT, align=CENTER, border=True):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font
    cell.alignment = align
    if border:
        cell.border = THIN_BORDER
    return cell


def create_normal_detail_sheet(workbook, array_info, bom_products, price_mapping,
                           sheet_prefix=None, matrix_data=None,
                           unmatched_products_out=None,
                           coating_thickness=10, nv_params=None,
                           pile_products=None,
                           code_to_images=None, image_temp_dir=None,
                           image_cache=None, angle_override=None,
                           need_weight_code=False, need_weight=False, need_code=False, missing_boards=0,
                           is_inverter=False, span_ew_override=None,
                           steel_discount_rate=None, purchased_discount_rate=None):
    from backend.core.shared.image_utils import (
        prepare_image_for_excel, add_image_centered_in_cell,
    )

    nv_params = nv_params or {}
    matrix_data = matrix_data or {}

    _add_code = need_code or need_weight_code
    _add_weight = need_weight or need_weight_code  # 参照互換のため残す（重量は常時表示）
    # 重量は固定でG列に常時表示。単価/数量/総金額は右へ1列ずつシフト。
    WEIGHT_COL = 7       # G: 重量(KG)
    PRICE_COL = 8        # H: 単価
    QTY_COL = 9          # I: 数量
    TOTAL_COL = 10       # J: 総金額
    _code_col = 11 if _add_code else None
    _weight_col = WEIGHT_COL
    _col_end = TOTAL_COL + (1 if _add_code else 0)
    _merge_end = get_column_letter(_col_end)

    img_w, img_h, img_col, img_padding = 65, 50, 5, 2

    def col_width_to_px(cw):
        return int(cw * 7.5 + 5)

    def row_height_to_px(rh):
        return int(rh * 1.33)

    def fit_image(ws, r, c, mw, mh, pad=img_padding):
        cw_val = ws.column_dimensions[get_column_letter(c)].width or 8.43
        rh_val = ws.row_dimensions[r].height or 15
        aw = col_width_to_px(cw_val) - pad * 2
        ah = row_height_to_px(rh_val) - pad * 2
        aw = max(aw, 20)
        ah = max(ah, 20)
        scale = min(aw / mw, ah / mh, 1.0)
        return int(mw * scale), int(mh * scale)

    if sheet_prefix:
        sheet_name = sheet_prefix
    else:
        sheet_name = str(array_info.get('table_qty', 1))
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    original = sheet_name
    counter = 1
    while sheet_name in workbook.sheetnames:
        sheet_name = f"{original}_{counter}"
        counter += 1

    ws = workbook.create_sheet(title=sheet_name)

    for col, width in DETAIL_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # normal 案件の列幅を個別指定（NV には影響しない）
    ws.column_dimensions['D'].width = 31
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 22
    ws.column_dimensions['J'].width = 25

    ws.row_dimensions[1].height = 38
    for row in range(2, 6):
        ws.row_dimensions[row].height = 23
    ws.row_dimensions[6].height = 40
    ws.row_dimensions[7].height = 12
    ws.row_dimensions[8].height = 25

    ws.merge_cells('A2:C5')
    for r in range(2, 6):
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN_BORDER
    if os.path.isfile(_LOGO_PATH):
        try:
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils import column_width_to_pixels, row_height_to_pixels

            _logo_img = XLImage(_LOGO_PATH)
            _logo_w, _logo_h = DETAIL_LOGO_WIDTH, DETAIL_LOGO_HEIGHT
            _logo_img.width = _logo_w
            _logo_img.height = _logo_h

            _range_w = sum(
                column_width_to_pixels(ws.column_dimensions[get_column_letter(c)].width or 8.43)
                for c in range(1, 4)
            )
            _range_h = sum(
                row_height_to_pixels(ws.row_dimensions[r].height or 15)
                for r in range(3, 6)
            )

            EMU = 9525
            _logo_img.anchor = OneCellAnchor(
                _from=AnchorMarker(
                    col=1,
                    colOff=int((_range_w - _logo_w) / 2 * EMU),
                    row=2,
                    rowOff=int((_range_h - _logo_h) / 2 * EMU),
                ),
                ext=XDRPositiveSize2D(
                    cx=int(_logo_w * EMU),
                    cy=int(_logo_h * EMU),
                ),
            )
            ws.add_image(_logo_img)
        except Exception:
            _insert_logo(ws, 'B3', width=DETAIL_LOGO_WIDTH, height=DETAIL_LOGO_HEIGHT)

    _title_merge_end = _merge_end
    _title_col_end = _col_end
    ws.merge_cells(f'A1:{_title_merge_end}1')
    if is_inverter:
        _detail_title = str((matrix_data or {}).get('project_name') or '').strip() or 'パワコン部材リスト'
    else:
        _detail_title = str((matrix_data or {}).get('project_name') or '').strip() or '各アレイ部材リスト'
    ws['A1'] = _detail_title
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER
    for c in range(1, _title_col_end + 1):
        ws.cell(row=1, column=c).border = THIN_BORDER

    rows_val = array_info.get('rows', '')
    cols_val = array_info.get('cols', '')
    qty_val = array_info.get('table_qty', 1)
    panel_count = int(rows_val) * int(cols_val) if rows_val and cols_val else 0
    module_wattage = matrix_data.get('module_wattage') or 0
    panel_size = matrix_data.get('panel_spec') or matrix_data.get('module_size') or ''
    wind_speed = matrix_data.get('max_wind_speed') or ''
    snow_load = matrix_data.get('max_snow_load') or ''
    angle = angle_override if angle_override is not None else (matrix_data.get('angle') or '')
    if isinstance(angle, float) and angle != angle:
        angle = ''
    if str(angle).lower() == 'nan' or str(angle).lower().strip('°') == 'nan':
        angle = ''
    ground_height = nv_params.get('ground_height') or matrix_data.get('ground_height') or ''
    span_ew = span_ew_override if span_ew_override is not None else (nv_params.get('span_ew') or '')
    sales_name = nv_params.get('sales_name') or 'Nanami'
    sales_phone = nv_params.get('sales_phone') or '+86-137-7466-5835'
    sales_fax = nv_params.get('sales_fax') or '0086-592-5738212'
    sales_tel = nv_params.get('sales_tel') or '0086-592-5767152'

    ws['A6'] = '〒361-000　\n中国厦門湖里区枋湖北二路891号匯鑫財富大厦11-12階'
    ws['A6'].font = SM_FONT
    ws['A6'].alignment = LEFT_A
    ws.merge_cells('A6:D6')
    for c in range(1, 5):
        ws.cell(row=6, column=c).border = THIN_BORDER

    _set(ws, 2, 4, f'Sales man:{sales_name}', align=LEFT_A)
    _set(ws, 3, 4, f'Mob:{sales_phone}', align=LEFT_A)
    _set(ws, 4, 4, f'Tel : {sales_tel}', align=LEFT_A)
    _set(ws, 5, 4, f'Fax: {sales_fax}', align=LEFT_A)

    # ===== pair-1 ラベル(E列) + 値(F:G結合) / pair-2 ラベル(H列) + 値(I列) =====
    # Row 2
    _set(ws, 2, 5, (matrix_data or {}).get('layout') or '縦置き', align=CENTER)
    if is_inverter:
        _set(ws, 2, 6, '/')
    else:
        angle_display = angle.rstrip('°').strip() if angle else ''
        angle_display = _strip_decimal_zero(angle_display)
        _set(ws, 2, 6, f'{angle_display}°' if angle_display else '', align=CENTER)
    ws.merge_cells('F2:G2')
    ws.cell(row=2, column=7).border = THIN_BORDER
    _set(ws, 2, 8, 'サイズ', align=RIGHT_A)
    if is_inverter:
        _set(ws, 2, 9, '/')
    else:
        _set(ws, 2, 9, f'{panel_size}mm' if panel_size else '', align=CENTER)

    # 品質保証テキストは J列のみ使用（J2:J6 結合）
    ws.merge_cells('J2:J6')
    quality_text = (
        '品質保証:\n10年間品質保証、それに、20年以上使える\n'
        '参考標準:\nJIS C 8955太陽電池アレイ用支持物設計標準'
    )
    _set(ws, 2, 10, quality_text, font=SM_FONT, align=LEFT_A)
    for r in range(2, 7):
        ws.cell(row=r, column=10).border = THIN_BORDER
        if _code_col:
            ws.cell(row=r, column=_code_col).border = THIN_BORDER

    # Row 3
    _set(ws, 3, 5, '風速', align=CENTER)
    _set(ws, 3, 6, _strip_decimal_zero(wind_speed), align=CENTER)
    ws.merge_cells('F3:G3')
    ws.cell(row=3, column=7).border = THIN_BORDER
    _set(ws, 3, 8, '発電量/PC', align=RIGHT_A)
    if is_inverter:
        _set(ws, 3, 9, '/')
    else:
        _cell_mw = _set(ws, 3, 9, module_wattage, align=CENTER)
        if module_wattage:
            _cell_mw.number_format = '#,##0 "Wp"'

    # Row 4
    _set(ws, 4, 5, '積雪', align=CENTER)
    _set(ws, 4, 6, _strip_decimal_zero(snow_load), align=CENTER)
    ws.merge_cells('F4:G4')
    ws.cell(row=4, column=7).border = THIN_BORDER
    _set(ws, 4, 8, '発電量(W)/基', align=RIGHT_A)
    if is_inverter:
        _set(ws, 4, 9, '/')
    elif module_wattage and rows_val and cols_val:
        panel_per_set = int(rows_val) * int(cols_val) if rows_val and cols_val else 0
        mb_val = int(missing_boards) if missing_boards else 0
        _cell_gen = _set(ws, 4, 9, f'=I3*({panel_per_set}+{mb_val})', align=CENTER)
        _cell_gen.number_format = '#,##0 "Wp"'
    else:
        _set(ws, 4, 9, '')

    # Row 5
    _set(ws, 5, 5, 'パネル高さ', align=CENTER)
    _set(ws, 5, 6, ground_height, align=CENTER)
    ws.merge_cells('F5:G5')
    ws.cell(row=5, column=7).border = THIN_BORDER
    _set(ws, 5, 8, '')
    _set(ws, 5, 9, '')

    # Row 6
    if is_inverter:
        _set(ws, 6, 5, '/')
        _set(ws, 6, 6, '')
    else:
        _set(ws, 6, 5, f'{rows_val}段' if rows_val else '', align=RIGHT_A)
        _set(ws, 6, 6, f'{cols_val}列' if cols_val else '', align=LEFT_A)
    ws.merge_cells('F6:G6')
    ws.cell(row=6, column=7).border = THIN_BORDER
    _set(ws, 6, 8, 'セット数')
    _set(ws, 6, 9, qty_val)

    _hdr_merge_end = _merge_end
    _hdr_col_end = _col_end
    ws.merge_cells(f'A7:{_hdr_merge_end}7')
    for c in range(1, _hdr_col_end + 1):
        ws.cell(row=7, column=c).border = THIN_BORDER

    YELLOW_FILL = PatternFill(start_color='FFDDB3', end_color='FFDDB3', fill_type='solid')
    headers = ['品名', '材質', '写真', '規格', '重量(KG)', '単価（USD）', '数量（PCS)', '総金額（USD）']
    if _add_code:
        headers += ['品番']
    ws.merge_cells('A8:B8')
    cell_num = ws.cell(row=8, column=1, value='序号')
    cell_num.font = SM_FONT_BOLD
    cell_num.alignment = CENTER
    cell_num.border = THIN_BORDER
    ws.cell(row=8, column=2).border = THIN_BORDER
    for i, h in enumerate(headers, 3):
        cell = ws.cell(row=8, column=i, value=h)
        cell.font = SM_FONT_BOLD
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    YELLOW_FILL = PatternFill(start_color='FFDDB3', end_color='FFDDB3', fill_type='solid')
    data_font = SM_FONT
    data_start = 9
    total_price_sum = Decimal('0')
    matched_count = 0
    name_field = 'name_ja'
    _detail_table_qty = int(array_info.get('table_qty', 1)) if array_info else 1
    from backend.core.shared.price_utils import _get_discount_category
    _cat_row_map = {}
    _category_row_lists = {
        'standard': [],
        'steel': [],
        'purchased': [],
    }
    _category_sums = {
        'standard': Decimal('0'),
        'steel': Decimal('0'),
        'purchased': Decimal('0'),
    }

    for row_idx in range(data_start, data_start + len(bom_products)):
        ws.row_dimensions[row_idx].height = 40

    for idx, product in enumerate(bom_products):
        row = data_start + idx
        product_code = product.get('code', '')
        price_info = resolve_price_info(price_mapping, product_code, spec=product.get('spec', ''))

        display_name = (
            price_info.get(name_field)
            or price_info.get('name_ko')
            or price_info.get('name')
            or product.get('name', '')
        ) if price_info else product.get('name', '')

        ws.merge_cells(f'A{row}:B{row}')
        _set(ws, row, 1, idx + 1)
        _set(ws, row, 3, display_name)

        raw_material = (price_info.get('db_material') if price_info and price_info.get('db_material') else None) or product.get('material', '')
        display_material = ''
        if raw_material:
            mat = str(raw_material).replace('&', '+')
            if coating_thickness in (15, 18):
                mat = mat.replace('304', '316')
                mat = mat.replace('/316', '')
            else:
                mat = mat.replace('/316', '')
            mat = translate_material(mat, 'ja')
            display_material = mat
            if coating_thickness in (15, 18):
                display_material += f'  {coating_thickness}um'

        _set(ws, row, 4, display_material)
        _set(ws, row, 5, '')
        _spec_clean = re.sub(r'[\u4e00-\u9fff]+', '', str(product.get('spec', '') or '')).strip()
        _spec_cell = ws.cell(row=row, column=6)
        try:
            _spec_num = float(_spec_clean)
            _spec_cell.value = _spec_num
            _spec_cell.number_format = '"L"!=#"mm"'
        except (ValueError, TypeError):
            _spec_cell.value = _spec_clean
        _spec_cell.font = SM_FONT
        _spec_cell.alignment = CENTER
        _spec_cell.border = THIN_BORDER

        quantity = product.get('quantity', 0)
        unit_price = 0
        display_unit_price = 0
        price_unit = ''
        is_matched = False

        if price_info and has_valid_price_info(price_info):
            unit_price = get_temp_base_price(price_info, product, '日语组', 'export')
            price_unit = price_info.get('unit', '')
            matched_count += 1
            is_matched = True
            _cat_row_map[row] = _get_discount_category(price_info, product)
        else:
            _cat_row_map[row] = _get_discount_category(price_info, product)
            if unmatched_products_out is not None and quantity > 0 and _is_valid_product_code(product_code):
                unmatched_products_out.append({
                    'code': product_code,
                    'name': product.get('name', ''),
                    'spec': product.get('spec', ''),
                    'material': raw_material,
                    'quantity': quantity * _detail_table_qty,
                    'weight': product.get('weight', 0),
                    'preinstall': normalize_preinstall(product.get('preinstall')),
                })

        is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
        length_mm = Decimal('0')
        if is_meter:
            length_mm = Decimal(str(extract_length_from_spec(product.get('spec')) or 0))

        if is_meter and length_mm > 0:
            display_unit_price = float(Decimal(str(unit_price)) * length_mm / Decimal('1000'))
        else:
            display_unit_price = unit_price
        display_unit_price = apply_temp_preinstall_adjustment(price_info, display_unit_price, product, '日语组', 'export')

        if display_unit_price > 0:
            cell = ws.cell(row=row, column=PRICE_COL, value=display_unit_price)
            cell.border = THIN_BORDER
            cell.number_format = NUM_FMT
            cell.font = data_font
            cell.alignment = CENTER
        else:
            _set(ws, row, PRICE_COL, '')

        if quantity > 0:
            _set(ws, row, QTY_COL, int(quantity) if quantity % 1 == 0 else quantity)
        else:
            _set(ws, row, QTY_COL, '')

        total_price = Decimal('0')
        cat = _get_discount_category(price_info, product)
        _category_row_lists[cat].append(row)
        if display_unit_price > 0 and quantity > 0:
            total_price = Decimal(str(display_unit_price)) * Decimal(str(quantity))

        total_price_rounded = round_to_2_decimal(total_price)
        if total_price_rounded > 0:
            cell = ws.cell(row=row, column=TOTAL_COL, value=float(total_price))
            cell.border = THIN_BORDER
            cell.number_format = NUM_FMT
            cell.font = data_font
            cell.alignment = CENTER
            total_price_sum += total_price_rounded
            _category_sums[cat] += total_price_rounded
        else:
            _set(ws, row, TOTAL_COL, '')

        weight_val = '/'
        if price_info:
            w = price_info.get('db_weight')
            price_unit = price_info.get('unit', '')
            if w:
                is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
                if is_meter:
                    length_mm = extract_length_from_spec(product.get('spec'))
                    if length_mm and length_mm > 0:
                        weight_val = round(float(w) / 1000 * length_mm, 4)
                    else:
                        weight_val = float(w)
                else:
                    weight_val = float(w)
        ws.cell(row=row, column=WEIGHT_COL, value=weight_val).border = THIN_BORDER
        ws.cell(row=row, column=WEIGHT_COL).alignment = CENTER
        ws.cell(row=row, column=WEIGHT_COL).font = data_font

        if _add_code:
            ws.cell(row=row, column=_code_col, value=product_code).border = THIN_BORDER
            ws.cell(row=row, column=_code_col).alignment = CENTER
            ws.cell(row=row, column=_code_col).font = data_font

        if not is_matched:
            _fill_end = _col_end
            for c in range(1, _fill_end + 1):
                ws.cell(row=row, column=c).fill = YELLOW_FILL

    rows_to_delete = []
    for r in range(data_start, data_start + len(bom_products)):
        name_val = ws.cell(row=r, column=3).value
        if name_val is None or name_val == '':
            continue
        qty_val_cell = ws.cell(row=r, column=QTY_COL).value
        if qty_val_cell is None or qty_val_cell == '' or (isinstance(qty_val_cell, (int, float)) and qty_val_cell == 0):
            rows_to_delete.append(r)

    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r)

    deleted_count = len(rows_to_delete)
    if deleted_count:
        print(f"   🗑️ 数量=0の行を {deleted_count} 行削除")

    new_seq = 1
    for r in range(data_start, ws.max_row + 1):
        name_val = ws.cell(row=r, column=3).value
        if name_val and name_val != '':
            ws.merge_cells(f'A{r}:B{r}')
            ws.cell(row=r, column=1, value=new_seq)
            ws.cell(row=r, column=1).font = SM_FONT
            ws.cell(row=r, column=1).alignment = CENTER
            new_seq += 1

    for r in range(data_start, ws.max_row + 1):
        name_val = ws.cell(row=r, column=3).value
        if name_val and name_val != '':
            ws.cell(row=r, column=TOTAL_COL, value=f'=H{r}*I{r}')
            ws.cell(row=r, column=TOTAL_COL).font = SM_FONT
            ws.cell(row=r, column=TOTAL_COL).alignment = CENTER
            ws.cell(row=r, column=TOTAL_COL).number_format = NUM_FMT

    data_end = data_start + len(bom_products) - 1 - deleted_count

    _cat_row_remapped = {}
    for orig_row, cat in _cat_row_map.items():
        if orig_row in set(rows_to_delete):
            continue
        offset = sum(1 for d in rows_to_delete if d < orig_row)
        new_row = orig_row - offset
        _cat_row_remapped[new_row] = cat

    row_product_map = {}
    for idx, product in enumerate(bom_products):
        orig_row = data_start + idx
        pc = product.get('code', '')
        row_product_map[orig_row] = str(pc).strip() if pc else ''

    survived = [r for r in sorted(row_product_map.keys()) if r not in set(rows_to_delete)]
    row_remap = {}
    for orig_row in survived:
        offset = sum(1 for d in rows_to_delete if d < orig_row)
        row_remap[orig_row - offset] = row_product_map[orig_row]

    image_found_count = 0
    image_not_found_count = 0
    for row, product_code in row_remap.items():
        if row < data_start or row > data_end:
            continue
        if not product_code:
            continue
        normalized_code = normalize_lookup_code(product_code)

        img_path = None
        if code_to_images and product_code in code_to_images and code_to_images[product_code]:
            img_path = code_to_images[product_code][0]
        elif code_to_images and normalized_code in code_to_images and code_to_images[normalized_code]:
            img_path = code_to_images[normalized_code][0]

        if not img_path and price_mapping:
            pi = resolve_price_info(price_mapping, product_code)
            if pi and pi.get('image_bytes'):
                img_bytes = pi['image_bytes']
                img_ext = pi.get('image_ext', '.png')
                if image_temp_dir:
                    db_img_name = f"db_{normalized_code or product_code}{img_ext}"
                    db_img_path = os.path.join(image_temp_dir, db_img_name)
                    if db_img_path not in (image_cache or {}):
                        with open(db_img_path, 'wb') as _f:
                            _f.write(img_bytes)
                        if image_cache is not None:
                            image_cache.pop(db_img_path, None)
                    img_path = db_img_path

        if img_path:
            fit_w, fit_h = fit_image(ws, row, img_col, img_w, img_h)
            temp_img_path = prepare_image_for_excel(
                img_path,
                target_width=fit_w,
                target_height=fit_h,
                temp_dir=image_temp_dir,
                cache=image_cache,
            )
            final_img_path = temp_img_path if temp_img_path else img_path
            success = add_image_centered_in_cell(
                ws,
                final_img_path,
                row,
                img_col,
                img_width=fit_w,
                img_height=fit_h,
            )
            if success:
                image_found_count += 1
                continue
        image_not_found_count += 1
        ws.cell(row=row, column=5, value='/').alignment = CENTER

    if image_found_count or image_not_found_count:
        print(f"   🖼️ {sheet_name}: 挿入 {image_found_count} 枚, 未検出 {image_not_found_count} 枚")

    # 修复 create_nv_detail_sheet 函数中的这三行（大约在第 558-597 行附近）

    sub_row = data_end + 1
    _sub_border_end = _col_end
    _set(ws, sub_row, 1, '')
    _set(ws, sub_row, 2, '')
    ws.merge_cells(f'C{sub_row}:I{sub_row}')
    _set(ws, sub_row, 3, 'SUB-TOTAL-（EXW）1基パワコン合計' if is_inverter else 'SUB-TOTAL-（EXW）1基架台合計', font=SM_FONT, align=Alignment(horizontal='right', vertical='center'))
    ws.cell(row=sub_row, column=1).border = Border(left=THIN_BORDER.left, right=Side(style=None), top=THIN_BORDER.top, bottom=THIN_BORDER.bottom)
    ws.cell(row=sub_row, column=2).border = Border(left=Side(style=None), right=Side(style=None), top=THIN_BORDER.top, bottom=THIN_BORDER.bottom)
    ws.cell(row=sub_row, column=3).border = Border(left=Side(style=None), right=Side(style=None), top=THIN_BORDER.top, bottom=THIN_BORDER.bottom)
    for c in range(4, _sub_border_end + 1):
        ws.cell(row=sub_row, column=c).border = THIN_BORDER
    ws.cell(row=sub_row, column=9).border = Border(left=Side(style=None), right=THIN_BORDER.right, top=THIN_BORDER.top, bottom=THIN_BORDER.bottom)
    cell_h = ws.cell(row=sub_row, column=TOTAL_COL, value=f'=SUM({get_column_letter(TOTAL_COL)}{data_start}:{get_column_letter(TOTAL_COL)}{data_end})')
    cell_h.font = SM_FONT
    cell_h.border = THIN_BORDER
    cell_h.alignment = CENTER
    cell_h.number_format = NUM_FMT
    if _add_code:
        ws.cell(row=sub_row, column=_code_col).border = THIN_BORDER
    ws.row_dimensions[sub_row].height = 28

    # Hidden category subtotal rows (for summary sheet discount formula references)
    _std_disc_rate = float(steel_discount_rate if steel_discount_rate is not None else 84)
    _purch_disc_rate = float(purchased_discount_rate if purchased_discount_rate is not None else 94)
    _disc_rate_pct = float(nv_params.get('discount_rate', 71) if nv_params else 71)
    cat_label_map = [
        ('standard', '標準部品小計'),
        ('steel', '碳鋼部品小計'),
        ('purchased', '外購部品小計'),
    ]
    category_subtotal_rows = {}
    cat_start_row = sub_row + 1
    _label_merge_end = 9 if not _add_code else 10
    for cat_idx, (cat, label) in enumerate(cat_label_map):
        r = cat_start_row + cat_idx
        category_subtotal_rows[cat] = r
        ws.merge_cells(f'A{r}:{get_column_letter(_label_merge_end)}{r}')
        cell_label = ws.cell(row=r, column=1, value=label)
        cell_label.font = data_font
        cell_label.alignment = Alignment(horizontal='right', vertical='center')
        cell_label.border = THIN_BORDER
        for c in range(2, _col_end + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
        cat_rows = _category_row_lists.get(cat, [])
        if cat_rows:
            refs = ','.join(f'{get_column_letter(TOTAL_COL)}{row}' for row in cat_rows)
            cat_formula = f'=SUM({refs})'
            if len(cat_formula) > 8000:
                cat_formula = float(_category_sums[cat])
        else:
            cat_formula = 0
        cell_cat = ws.cell(row=r, column=TOTAL_COL, value=cat_formula)
        cell_cat.border = THIN_BORDER
        cell_cat.number_format = NUM_FMT
        cell_cat.font = data_font
        cell_cat.alignment = CENTER
        ws.row_dimensions[r].hidden = True
        ws.row_dimensions[r].height = 24

    discounted_row = cat_start_row + len(cat_label_map)
    ws.merge_cells(f'A{discounted_row}:{get_column_letter(_label_merge_end)}{discounted_row}')
    cell_disc_label = ws.cell(row=discounted_row, column=1, value='特別値引き後合計')
    cell_disc_label.font = data_font
    cell_disc_label.alignment = Alignment(horizontal='right', vertical='center')
    cell_disc_label.border = THIN_BORDER
    for c in range(2, _col_end + 1):
        ws.cell(row=discounted_row, column=c).border = THIN_BORDER
    _std_r = category_subtotal_rows['standard']
    _steel_r = category_subtotal_rows['steel']
    _purch_r = category_subtotal_rows['purchased']
    ws.cell(
        row=discounted_row, column=TOTAL_COL,
        value=(
            f'={get_column_letter(TOTAL_COL)}{_std_r}*{_disc_rate_pct}/100'
            f'+{get_column_letter(TOTAL_COL)}{_steel_r}*{_std_disc_rate}/100'
            f'+{get_column_letter(TOTAL_COL)}{_purch_r}*{_purch_disc_rate}/100'
        )
    )
    ws.cell(row=discounted_row, column=TOTAL_COL).border = THIN_BORDER
    ws.cell(row=discounted_row, column=TOTAL_COL).number_format = NUM_FMT
    ws.cell(row=discounted_row, column=TOTAL_COL).font = data_font
    ws.cell(row=discounted_row, column=TOTAL_COL).alignment = CENTER
    ws.row_dimensions[discounted_row].height = 28

    pile_data_start_row = 0
    pile_data_end_row = 0
    pile_total_per_base = 0.0
    pile_total_qty = 0
    pile_products = pile_products or []
    _pile_img_cache = {}
    pile_display_name = ''
    if pile_products:
        import tempfile as _tf
        _pile_tmp = _tf.mkdtemp(prefix='nv_pile_img_')
        pile_start_row = discounted_row + 1

        pile_seq = 1
        pile_data_start_row = pile_start_row
        rendered_pile_count = 0
        first_pile_unit_price = None
        for p in pile_products:
            qty_check = int(p.get('quantity', 0) or 0)
            if qty_check <= 0:
                continue
            row = pile_data_start_row + pile_seq - 1
            ws.row_dimensions[row].height = 40
            code = str(p.get('code', '') or '')
            price_info = _resolve_pile_price(price_mapping, code, spec=p.get('spec', ''))
            display_name = (
                price_info.get('name_ja')
                or price_info.get('name')
                or 'スクリュー杭'
            ) if price_info else 'スクリュー杭'
            if not pile_display_name:
                pile_display_name = display_name
            raw_material = (price_info.get('db_material') if price_info and price_info.get('db_material') else None) or p.get('material', '')
            display_material = ''
            if raw_material:
                mat = str(raw_material).replace('&', '+')
                if coating_thickness in (15, 18):
                    mat = mat.replace('304', '316')
                    mat = mat.replace('/316', '')
                else:
                    mat = mat.replace('/316', '')
                mat = translate_material(mat, 'ja')
                display_material = mat
                if coating_thickness in (15, 18):
                    display_material += f'  {coating_thickness}um'
            spec = _format_pile_spec_display(p.get('spec', ''))
            qty = qty_check

            unit_price = 0.0
            display_unit_price = 0.0
            pile_is_matched = False
            YELLOW_FILL = PatternFill(start_color='FFDDB3', end_color='FFDDB3', fill_type='solid')
            if price_info and has_valid_price_info(price_info):
                unit_price = get_temp_base_price(price_info, p, '日语组', 'export')
                pile_is_matched = True
            else:
                print(f"   ⚠ 杭価格未検出: code={code}, spec={p.get('spec', '')}, price_info={'found' if price_info else 'None'}")
            pricing_unit = (price_info.get('unit', '') if price_info else '') or ''
            length_mm = float(p.get('length', 0) or 0)
            if not length_mm:
                length_mm = float(extract_length_from_spec(p.get('spec', '')) or 0)
            if not length_mm:
                _spec_raw = str(p.get('spec', '') or '').strip()
                _m = re.search(r'(\d+(?:\.\d+)?)', _spec_raw)
                if _m:
                    length_mm = float(_m.group(1))
            if not pile_is_matched:
                display_unit_price = ''
            elif pricing_unit == '米' and length_mm > 0:
                display_unit_price = (length_mm / 1000) * unit_price
            elif pricing_unit == '米' and length_mm <= 0:
                display_unit_price = unit_price
                print(f"   ⚠ 杭長さ抽出失敗、メートル単価をそのまま使用: code={code}, spec={p.get('spec', '')}, unit_price={unit_price}")
            else:
                display_unit_price = unit_price
            if isinstance(display_unit_price, (int, float)):
                display_unit_price = apply_temp_preinstall_adjustment(price_info, display_unit_price, p, '日语组', 'export')
            print(f"   🔍 杭計算: code={code}, unit_price={unit_price}, pricing_unit={pricing_unit}, length_mm={length_mm}, qty={qty}, display_unit_price={display_unit_price}")
            _real_total = display_unit_price * qty if isinstance(display_unit_price, (int, float)) and display_unit_price > 0 and qty > 0 else 0.0
            pile_total_per_base += _real_total
            pile_total_qty += qty

            if first_pile_unit_price is None:
                first_pile_unit_price = display_unit_price if isinstance(display_unit_price, (int, float)) else 0.0
            _pile_row_price = first_pile_unit_price

            ws.merge_cells(f'A{row}:B{row}')
            ws.cell(row=row, column=1, value=pile_seq).font = SM_FONT
            ws.cell(row=row, column=1).alignment = CENTER
            ws.cell(row=row, column=3, value=display_name).font = SM_FONT
            ws.cell(row=row, column=3).alignment = CENTER
            ws.cell(row=row, column=4, value=display_material).font = SM_FONT
            ws.cell(row=row, column=4).alignment = CENTER
            _set(ws, row, 5, '')
            try:
                _spec_num = float(spec)
                _spec_display = f'L={int(_spec_num)}mm' if _spec_num == int(_spec_num) else f'L={_spec_num}mm'
            except (ValueError, TypeError):
                _spec_display = str(spec)
            ws.cell(row=row, column=6, value=_spec_display).font = SM_FONT
            ws.cell(row=row, column=6).alignment = CENTER

            _pile_weight = '/'
            if price_info:
                _pw = price_info.get('db_weight')
                _pu = price_info.get('unit', '')
                if _pw:
                    if _pu in ['米', 'm', 'M', 'meter', 'Meter']:
                        _pl = extract_length_from_spec(p.get('spec'))
                        _pile_weight = round(float(_pw) / 1000 * _pl, 4) if _pl and _pl > 0 else float(_pw)
                    else:
                        _pile_weight = float(_pw)
            ws.cell(row=row, column=WEIGHT_COL, value=_pile_weight).font = SM_FONT
            ws.cell(row=row, column=WEIGHT_COL).alignment = CENTER
            ws.cell(row=row, column=WEIGHT_COL).border = THIN_BORDER

            ws.cell(row=row, column=PRICE_COL, value=_pile_row_price).font = SM_FONT
            ws.cell(row=row, column=PRICE_COL).alignment = CENTER
            ws.cell(row=row, column=PRICE_COL).number_format = NUM_FMT

            ws.cell(row=row, column=QTY_COL, value=qty).font = SM_FONT
            ws.cell(row=row, column=QTY_COL).alignment = CENTER

            ws.cell(row=row, column=TOTAL_COL, value=0).font = SM_FONT
            ws.cell(row=row, column=TOTAL_COL).alignment = CENTER
            ws.cell(row=row, column=TOTAL_COL).number_format = NUM_FMT

            if _add_code:
                ws.cell(row=row, column=_code_col, value=code).font = SM_FONT
                ws.cell(row=row, column=_code_col).alignment = CENTER
                ws.cell(row=row, column=_code_col).border = THIN_BORDER

            _pile_border_end = _col_end
            for c in range(1, _pile_border_end + 1):
                ws.cell(row=row, column=c).border = THIN_BORDER

            img_inserted = False
            if price_info and price_info.get('image_bytes'):
                try:
                    import os as _os
                    _img_bytes = price_info['image_bytes']
                    _img_ext = price_info.get('image_ext', '.png') or '.png'
                    _db_img_name = f"pile_{code}{_img_ext}"
                    _db_img_path = _os.path.join(_pile_tmp, _db_img_name)
                    if _db_img_path not in _pile_img_cache:
                        with open(_db_img_path, 'wb') as _f:
                            _f.write(_img_bytes)
                        _pile_img_cache.pop(_db_img_path, None)
                    _fit_w, _fit_h = fit_image(ws, row, img_col, img_w, img_h)
                    _prep = prepare_image_for_excel(
                        _db_img_path,
                        target_width=_fit_w,
                        target_height=_fit_h,
                        temp_dir=_pile_tmp,
                        cache=_pile_img_cache,
                    )
                    _final = _prep if _prep else _db_img_path
                    img_inserted = add_image_centered_in_cell(
                        ws, _final, row, img_col,
                        img_width=_fit_w, img_height=_fit_h,
                    )
                except Exception:
                    pass
            if not img_inserted:
                ws.cell(row=row, column=5, value='/').alignment = CENTER

            pile_seq += 1
            rendered_pile_count += 1

        pile_data_end_row = pile_data_start_row + rendered_pile_count - 1 if rendered_pile_count > 0 else 0

    if bool(nv_params.get('remove_weight')):
        ws.column_dimensions[get_column_letter(WEIGHT_COL)].hidden = True
        _last_content_row = max(sub_row, pile_data_end_row or 0)
        for _r in range(8, _last_content_row + 1):
            _cell = ws.cell(row=_r, column=WEIGHT_COL)
            _is_merged_member = False
            for _mr in ws.merged_cells.ranges:
                if _cell.coordinate in _mr and _cell.coordinate != _mr.start_cell.coordinate:
                    _is_merged_member = True
                    break
            if not _is_merged_member:
                _cell.value = None

    apply_print_setup(ws, 'ja_normal')

    discount_total_row = None

    return {
        'sheet_name': sheet_name,
        'discount_total_row': discount_total_row,
        'sub_total_row': sub_row,
        'category_subtotal_rows': category_subtotal_rows,
        'discounted_total_row': discounted_row,
        'pile_data_start_row': pile_data_start_row,
        'pile_data_end_row': pile_data_end_row,
        'pile_total_per_base': pile_total_per_base,
        'pile_total_qty': pile_total_qty,
        'array_info': array_info,
        'angle': angle,
        'category_rows': _cat_row_remapped,
    }
