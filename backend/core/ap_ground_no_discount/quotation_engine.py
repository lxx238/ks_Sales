"""亚太组（Asia-Pacific）地面支架报价单引擎 —— 无折扣版（隔离模块）。

与 ap_ground（有折扣版）完全隔离，方法名独立：
    - collect_ap_ground_no_disc_array
    - create_ap_ground_no_disc_detail_sheet
    - create_ap_ground_no_disc_summary_sheet

币种：USD（美元）。

输出结构（无折扣，简化版）：
    1) 每个阵列一个明细页，BOM 表 9 列（A~I）：
            Number | Item Number | Product Name | Material | Picture | Spec. |
            Unit Price (USD) EXW | QTY (PCS) | Total price(USD) EXW
       其中 Total price = Unit Price × QTY（每基），不做折扣。
    2) 一个汇总页（Total）：每阵列一行
       NO. | Row | Column | Installation Angle | Table | Power（kw） |
       Price(USD/Table) | Amount(USD) + 抬头 + EXW TOTAL AMOUNT + Unite Price/W。

计价说明：
    - 单价直接取数据库价格（USD，长度物料按规格折算），不做汇率换算、不做折扣；
    - Sales / 联系人信息由前端 contact_info 提供，未提供则留空。

本模块与 ap_ground / ap_common / en_simple 完全隔离，自行实现分类、取价、渲染逻辑。
"""

import os
import sys
from decimal import Decimal

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from backend.core.print_settings import apply_print_setup

from backend.core.shared.bom_utils import resolve_products_and_array
from backend.core.shared.price_utils import (
    resolve_price_info,
    has_valid_price_info,
    get_temp_base_price,
    apply_temp_preinstall_adjustment,
)
from backend.core.shared.weight_utils import extract_length_from_spec
from backend.core.shared.text_utils import (
    extract_main_name,
    normalize_lookup_code,
    _strip_cjk_spec,
    normalize_angle,
)
from backend.core.shared.product_utils import (
    _is_valid_product_code,
    _match_exclude_group,
    normalize_preinstall,
)
from backend.core.shared.constants import (
    EXCLUDE_ITEM_GROUPS,
    IMAGE_PADDING,
)
from backend.core.shared.image_utils import (
    prepare_image_for_excel,
    add_image_centered_in_cell,
)
from backend.core.material_translate import translate_material, adjust_material_by_coating

if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

CURRENCY_LABEL = 'USD'
CURRENCY_FMT = '"$"#,##0.00'

# 图片所在列：E（第 5 列）
IMAGE_COL_INDEX = 6
AP_GROUND_IMAGE_WIDTH = 372
AP_GROUND_IMAGE_HEIGHT = 92

# BOM 明细表头（10 列，A~J，Number 合并 A-B）—— 无折扣
DETAIL_HEADERS = [
    'Number',
    'Item Number',
    'Product Name',
    'Material',
    'Picture',
    'Spec.',
    'Unit Price (USD)\nEXW',
    'QTY (PCS)',
    'Total price(USD)\nEXW',
]
# 列定义（1-based）：
# Number=A~B(合并), Item Number=C(3), Product Name=D(4), Material=E(5),
# Picture=F(6), Spec.=G(7), Unit Price=H(8), QTY=I(9), Total price=J(10)
MAX_COL = 10  # A~J（Number 合并 A-B）
MAX_COL_LETTER = get_column_letter(MAX_COL)
COL_ITEM_NUMBER = get_column_letter(3)    # C 物料编码
COL_PICTURE = get_column_letter(6)        # F
COL_UNIT_PRICE = get_column_letter(8)     # H 原价（即展示单价）
COL_QTY = get_column_letter(9)            # I 每基数量
COL_TOTAL_PRICE = get_column_letter(10)   # J 每基合计 = H*I


def _set_cell(ws, row, col, val, font=None, align=None, border=None,
              number_format=None, fill=None):
    cell = ws.cell(row=row, column=col, value=val)
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format
    if fill is not None:
        cell.fill = fill
    return cell


def _apply_outer_frame(ws, top_row, bottom_row, left_col, right_col,
                       weight='medium', color='000000'):
    """在指定矩形外缘绘制粗边框（合并单元格重定向到锚点单元格，保留内部边框）。"""
    side = Side(style=weight, color=color)

    def _anchor(coord):
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                return rng.start_cell
        return ws[coord]

    def _merge_border(cell, **edges):
        ex = cell.border
        cell.border = Border(
            left=edges.get('left', ex.left),
            right=edges.get('right', ex.right),
            top=edges.get('top', ex.top),
            bottom=edges.get('bottom', ex.bottom),
        )

    for r in range(top_row, bottom_row + 1):
        lc = _anchor(ws.cell(row=r, column=left_col).coordinate)
        rc = _anchor(ws.cell(row=r, column=right_col).coordinate)
        _merge_border(lc, left=side)
        _merge_border(rc, right=side)
    for c in range(left_col, right_col + 1):
        tc = _anchor(ws.cell(row=top_row, column=c).coordinate)
        bc = _anchor(ws.cell(row=bottom_row, column=c).coordinate)
        _merge_border(tc, top=side)
        _merge_border(bc, bottom=side)


def _classify_products(all_products, price_mapping, delete_options=None,
                       always_exclude=False, ap_exclude_options=None):
    """分类：主物料 / 排除(可选)项 / 地桩（与 ap_common 一致的隔离副本）。"""
    delete_options = delete_options or {}
    ap_exclude_options = ap_exclude_options or {}
    all_exclude_opts = {k: True for k in EXCLUDE_ITEM_GROUPS} if always_exclude else {}
    for k in list(ap_exclude_options.keys()) + list(delete_options.keys()):
        all_exclude_opts[k] = True

    main = []
    excluded = []
    pile_products = []
    price_info_cache = {}
    for p in all_products:
        code = str(p.get('code', '')).strip()
        spec = str(p.get('spec', '')).strip()
        cache_key = (code, spec) if spec else code
        if cache_key not in price_info_cache:
            price_info_cache[cache_key] = (
                resolve_price_info(price_mapping, code, spec=spec) if price_mapping else None
            )
        pi = price_info_cache[cache_key]
        p['_price_info'] = pi
        attr = (pi.get('attribute', '') if pi else '') or ''
        if attr.strip() in ('地桩', '地盤杭'):
            p['_is_pile'] = True
            pile_products.append(p)
            continue
        if not _is_valid_product_code(code):
            continue
        qty = p.get('quantity', 0)
        if not qty or qty <= 0:
            continue
        matched_group = _match_exclude_group(p, price_mapping, all_exclude_opts)
        if matched_group:
            if delete_options.get(matched_group):
                continue
            if ap_exclude_options.get(matched_group):
                excluded.append(p)
            else:
                main.append(p)
        else:
            main.append(p)
    return main, excluded, pile_products, price_info_cache


def _calc_base_display_price(product, price_mapping, group, sale_type):
    """基础单价(长度折算后、未折扣) → 原价展示单价（直接取数据库价格）。"""
    code = product.get('code', '')
    price_info = resolve_price_info(price_mapping, code, spec=product.get('spec', ''))
    if not price_info or not has_valid_price_info(price_info):
        return 0.0, '', price_info
    base_price = float(get_temp_base_price(price_info, product, group, sale_type))
    price_unit = price_info.get('unit', '')
    is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
    if is_meter:
        length_mm = extract_length_from_spec(product.get('spec', '')) or 0
        if length_mm > 0:
            base_price = base_price * length_mm / 1000.0
    base_price = apply_temp_preinstall_adjustment(price_info, base_price, product, group, sale_type)
    return base_price, price_unit, price_info


def collect_ap_ground_no_disc_array(df, price_mapping, config=None, matrix_data=None,
                                    pre_parsed_products=None, sale_type='export',
                                    coating_thickness=10, delete_options=None,
                                    always_exclude_extra_items=False, ap_exclude_options=None,
                                    unmatched_products_list=None, module_wattage=None, **kwargs):
    """收集单个阵列的产品与站点信息（不渲染）。无折扣版：仅保留 unit_price / qty_per_table。"""
    group = '亚太组'
    all_products, array_info, span_info, rows, cols = resolve_products_and_array(
        pre_parsed_products, df, matrix_data,
    )
    main_products, excluded_products, pile_products, _pic = _classify_products(
        all_products, price_mapping, delete_options or {},
        always_exclude_extra_items, ap_exclude_options or {},
    )

    matrix_data = matrix_data or {}
    set_count = matrix_data.get('set_count') or 1
    if not isinstance(set_count, int) or set_count <= 0:
        set_count = 1
    output_kw = matrix_data.get('output_kw') or 0
    if not output_kw and module_wattage and rows and cols:
        try:
            output_kw = float(rows) * float(cols) * float(set_count) * float(module_wattage) / 1000.0
        except (TypeError, ValueError):
            output_kw = 0
    project_name = str(matrix_data.get('project_name') or '').strip()

    render_products = main_products + excluded_products
    items = []
    matched_count = 0
    unmatched_count = 0
    local_unmatched = []
    for product in render_products:
        code = str(product.get('code', '')).strip()
        quantity = product.get('quantity', 0)

        display_price, price_unit, price_info = _calc_base_display_price(
            product, price_mapping, group, sale_type
        )
        is_matched = bool(price_info and has_valid_price_info(price_info))
        if is_matched:
            matched_count += 1
        else:
            unmatched_count += 1
            if unmatched_products_list is not None and _is_valid_product_code(code):
                local_unmatched.append({
                    'code': code,
                    'name': product.get('name', ''),
                    'spec': product.get('spec', ''),
                    'quantity': quantity,
                    'unit': price_info.get('unit', '') if price_info else '',
                    'preinstall': normalize_preinstall(product.get('preinstall')),
                    'issue_reason': '数据库无法匹配',
                })

        en_name = ''
        _name_src = price_info if price_info else (resolve_price_info(price_mapping, code) if price_mapping else None)
        if _name_src:
            en_name = (_name_src.get('name_en') or _name_src.get('name_ko') or '')
        raw_mat = ((price_info.get('db_material') if price_info and price_info.get('db_material') else None)
                   or (_name_src.get('db_material') if _name_src and _name_src.get('db_material') else None)
                   or product.get('material', ''))
        material = adjust_material_by_coating(translate_material(raw_mat, 'en'), coating_thickness)
        spec_val = product.get('spec', '')

        qty_per_table = float(quantity or 0)
        items.append({
            'code': code,
            'name': en_name,
            'material': material,
            'spec': spec_val,
            'spec_norm': _strip_cjk_spec(spec_val),
            'unit_price': display_price,
            'qty_per_table': qty_per_table,
            'is_matched': is_matched,
        })

    pile_scaled = []
    for p in pile_products:
        qty = float(p.get('quantity', 0) or 0)
        if qty <= 0:
            continue
        code = str(p.get('code', '')).strip()
        spec = p.get('spec', '')
        pi = resolve_price_info(price_mapping, code, spec=spec) if price_mapping else None
        en_name = ''
        if pi:
            en_name = (pi.get('name_en') or pi.get('name_ko') or '')
        raw_mat = (pi.get('db_material') if pi and pi.get('db_material') else None) or p.get('material', '')
        material = adjust_material_by_coating(translate_material(raw_mat, 'en'), coating_thickness)
        pile_scaled.append({
            'code': code,
            'name': en_name,
            'material': material,
            'spec': spec,
            'spec_norm': _strip_cjk_spec(spec),
            'unit_price': float(pi.get('base_price', 0) or 0) if pi else 0,
            'qty_per_table': qty,
            'is_matched': bool(pi and has_valid_price_info(pi)),
        })

    site = {
        'rows': rows,
        'cols': cols,
        'tables': set_count,
        'kw': output_kw,
        'span_info': span_info,
        'pile_products': pile_scaled,
        'project_name': project_name,
    }
    return {
        'site': site,
        'items': items,
        'matched': matched_count,
        'unmatched': unmatched_count,
    }


# ------------------------------------------------------------------ 明细页

def create_ap_ground_no_disc_detail_sheet(
        workbook, site, items, price_mapping=None,
        matrix_data=None, config=None, contact_info=None,
        sale_type='export', coating_thickness=10,
        image_path=None, image_folder=None, code_to_images=None,
        image_temp_dir=None, image_cache=None,
        sheet_prefix=None, module_wattage=None,
        **kwargs
):
    """创建单个阵列的明细工作表（9 列 A~I，抬头 + BOM，无折扣）。"""
    matrix_data = matrix_data or {}
    config = config or {}
    contact_info = contact_info or {}

    rows = site.get('rows')
    cols = site.get('cols')
    tables = site.get('tables') or 1
    try:
        tables = int(tables)
    except (TypeError, ValueError):
        tables = 1

    sheet_name = sheet_prefix or ''
    if not sheet_name:
        sheet_name = f'{rows}x{cols}x{tables}' if rows and cols else 'Detail'
    sheet_name = extract_main_name(sheet_name) or sheet_name
    # 清洗 Excel sheet 名非法字符（[ ] : \ / ? *）
    for _bad in '[]:\\/?*':
        sheet_name = sheet_name.replace(_bad, '')
    sheet_name = sheet_name.strip() or 'Detail'
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    original = sheet_name
    counter = 1
    while sheet_name in workbook.sheetnames:
        sheet_name = f"{original}_{counter}"
        counter += 1
    ws = workbook.create_sheet(title=sheet_name)
    print(f"[AP-GROUND-ND] Creating detail sheet: {sheet_name}")

    thin = Side(style='thin', color='000000')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name='Calibri', size=26, bold=True, color='000000')
    info_font = Font(name='Calibri', size=14, color='000000')
    info_bold = Font(name='Calibri', size=14, bold=True, color='000000')
    header_font = Font(name='Calibri', size=14, bold=True, color='000000')
    normal_font = Font(name='Calibri', size=14, color='000000')
    bold_font = Font(name='Calibri', size=14, bold=True, color='000000')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_a = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_a = Alignment(horizontal='right', vertical='center', wrap_text=True)

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # 列宽（参考无折扣样例）
    _widths = {'A': 12, 'B': 22, 'C': 30, 'D': 35, 'E': 25, 'F': 24,
               'G': 24, 'H': 21, 'I': 28, 'J': 22}
    for col_letter, _w in _widths.items():
        ws.column_dimensions[col_letter].width = _w

    # ---- 抬头：标题 ----
    ws.merge_cells(f'A1:{MAX_COL_LETTER}1')
    _set_cell(ws, 1, 1, 'Solar PV Mounting System', font=title_font, align=center)
    ws.row_dimensions[1].height = 60

    # ---- 抬头信息块（行 2~6）----
    angle_val = config.get('angle', '') or matrix_data.get('angle', '')
    angle_display = normalize_angle(angle_val) if angle_val else ''
    panel_size = matrix_data.get('module_size') or config.get('panel_spec', '') or ''
    power_pc = matrix_data.get('module_wattage') or module_wattage or ''
    wind = matrix_data.get('max_wind_speed') or ''
    snow = matrix_data.get('max_snow_load') or ''
    watt_per_table = ''
    if power_pc and rows and cols:
        try:
            watt_per_table = int(float(power_pc) * float(rows) * float(cols))
        except (TypeError, ValueError):
            watt_per_table = ''
    span_info = site.get('span_info') or config.get('cross_span', '') or ''
    span_sn = span_ew = ''
    if span_info:
        span_ew = span_info

    contact_name = contact_info.get('contact_name') or ''
    phone = contact_info.get('phone') or ''
    tel = contact_info.get('tel') or ''

    # A2:C6 合并（Logo 区，A 为窄边距），图片从 B3 开始
    ws.column_dimensions['A'].width = 1
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 20
    ws.merge_cells('A2:C6')
    if image_path and os.path.exists(image_path):
        try:
            img = XLImage(img=image_path)
            _tot_w_px = sum(float(ws.column_dimensions[c].width or 8.43) for c in ('A', 'B', 'C')) * 7.0
            _tot_h_px = sum(float(ws.row_dimensions[r].height or 20) for r in range(2, 7)) * 1.33
            _ow, _oh = img.width, img.height
            _ratio = _ow / _oh
            _tr = _tot_w_px / _tot_h_px
            if _ratio > _tr:
                img.width = _tot_w_px; img.height = _tot_w_px / _ratio
            else:
                img.height = _tot_h_px; img.width = _tot_h_px * _ratio
            ws.add_image(img, 'B3')
        except Exception:
            _set_cell(ws, 3, 2, '', font=info_font, align=center, border=thin_border)

    def _h(r, c, val, font=None):
        _set_cell(ws, r, c, val, font=font or info_font, align=center, border=thin_border)

    # D 列：联系人 + Array
    _h(2, 4, f'Sales:{contact_name}' if contact_name else '')
    _h(3, 4, f'Mob: {phone}' if phone else '')
    _h(4, 4, f'Tel: {tel}' if tel else '')
    _h(6, 4, 'Array')

    # E / F 列：标签1 + 值1（行2-5），Array 行号(段)/列号(列)
    _h(2, 5, 'Installation Angle')
    _h(2, 6, angle_display)
    _h(3, 5, 'Max Wind Load ')
    _h(3, 6, str(wind))
    _h(4, 5, 'Max Snow Load ')
    _h(4, 6, str(snow))
    _h(5, 5, 'Span(S/N)')
    _h(5, 6, str(span_sn))
    if rows:
        _h(6, 5, f'{rows} rows')
    if cols:
        _h(6, 6, f'{cols} Columns')

    # G / H 列：标签2 + 值2（行2-5），Array 台数(table)
    _h(2, 7, 'Panel Size')
    _h(2, 8, f'{panel_size} mm' if panel_size else '')
    _h(3, 7, 'Power/PC')
    _h(3, 8, f'{power_pc} WP' if power_pc else '')
    _h(4, 7, 'Watt(W)/Table')
    _h(4, 8, f'{watt_per_table} WP' if watt_per_table != '' else '')
    _h(5, 7, 'Span(E/W)')
    _h(5, 8, f'{span_ew} mm' if span_ew else '')
    if tables:
        ws.merge_cells('G6:H6')
        _h(6, 7, f'{tables} table')

    # I2:J6 合并：质保
    ws.merge_cells('I2:J6')
    _set_cell(ws, 2, 9, '10 years warranty \n20 years service life', font=info_font,
              align=center, border=thin_border)

    # 抬头网格所有单元格补齐边框
    for _r in range(2, 7):
        for _c in range(1, MAX_COL + 1):
            ws.cell(row=_r, column=_c).border = thin_border
    for _r in range(2, 7):
        ws.row_dimensions[_r].height = 20

    # ---- BOM 表头（行 8）----
    header_row = 8
    ws.row_dimensions[header_row].height = 46
    # Number 占 A-B 两列（合并），其余表头从 C 列开始
    ws.merge_cells(f'A{header_row}:B{header_row}')
    for ci, label in enumerate(DETAIL_HEADERS):
        if ci == 0:
            col = 1  # Number -> A（合并 A-B）
        else:
            col = ci + 2  # 其余顺移 +1
        _set_cell(ws, header_row, col, label, font=header_font, align=center,
                  border=thin_border)

    # ---- BOM 数据（行 9+）----
    data_start = 9
    matched_count = 0
    unmatched_count = 0
    row_code_map = {}
    image_found_count = 0
    image_not_found_count = 0

    for idx, item in enumerate(items):
        row = data_start + idx
        ws.row_dimensions[row].height = 60
        code = item['code']
        unit_price = item['unit_price']
        qty_per_table = item['qty_per_table']
        is_matched = item['is_matched']

        if is_matched:
            matched_count += 1
        else:
            unmatched_count += 1

        for c in range(1, MAX_COL + 1):
            _set_cell(ws, row, c, None, font=normal_font, align=center, border=thin_border)

        # Number 占 A-B 两列（合并）
        ws.merge_cells(f'A{row}:B{row}')
        _set_cell(ws, row, 1, idx + 1, font=normal_font, align=center, border=thin_border)
        _set_cell(ws, row, 3, code, font=normal_font, align=center, border=thin_border)
        _set_cell(ws, row, 4, item['name'], font=normal_font, align=center, border=thin_border)
        _set_cell(ws, row, 5, item['material'], font=normal_font, align=center, border=thin_border)
        _set_cell(ws, row, 6, '', font=normal_font, align=center, border=thin_border)  # Picture
        _set_cell(ws, row, 7, _strip_cjk_spec(item['spec']), font=normal_font, align=center, border=thin_border)

        # H 单价
        if is_matched and unit_price > 0:
            _set_cell(ws, row, 8, round(unit_price, 6), font=normal_font, align=center,
                      border=thin_border, number_format=CURRENCY_FMT)
        else:
            _set_cell(ws, row, 8, '', font=normal_font, align=center, border=thin_border,
                      number_format=CURRENCY_FMT)
        # I 每基数量
        qty_disp = int(qty_per_table) if float(qty_per_table).is_integer() else qty_per_table
        _set_cell(ws, row, 9, qty_disp, font=normal_font, align=center, border=thin_border)
        # J 每基合计 = H × I
        _set_cell(ws, row, 10, f'=H{row}*I{row}', font=normal_font, align=center,
                  border=thin_border, number_format=CURRENCY_FMT)

        row_code_map[row] = code
        if not is_matched:
            for c in range(1, MAX_COL + 1):
                ws.cell(row=row, column=c).fill = yellow_fill

    data_end = data_start + len(items) - 1 if items else data_start - 1
    has_data = bool(items) and data_end >= data_start

    # ---- 末尾合计行 ----
    total_table_row = data_end + 1
    total_all_row = data_end + 2
    alt_note_row = data_end + 3

    def _total_row(r, label, multiplier=None, amount_font=None):
        amt_font = amount_font or bold_font
        ws.merge_cells(f'A{r}:I{r}')
        _set_cell(ws, r, 1, label, font=bold_font, align=right_a, border=thin_border)
        for ci in range(2, 10):
            ws.cell(row=r, column=ci).border = thin_border
        if has_data:
            expr = f'SUM(J{data_start}:J{data_end})'
            if multiplier:
                expr = f'{expr}*{multiplier}'
            _set_cell(ws, r, 10, f'={expr}', font=amt_font, align=center,
                      border=thin_border, number_format=CURRENCY_FMT)
        else:
            _set_cell(ws, r, 10, 0, font=amt_font, align=center, border=thin_border,
                      number_format=CURRENCY_FMT)
        ws.row_dimensions[r].height = 24

    _total_row(total_table_row, 'TOTAL AMOUNT/TABLE')
    _total_row(total_all_row, 'TOTAL AMOUNT OF ALL TABLE', multiplier=tables, amount_font=bold_font)

    # 备选产品提示行（外框之外）
    _red_info = Font(name='Calibri', size=14, color='FFFF0000')
    ws.merge_cells(f'A{alt_note_row}:J{alt_note_row}')
    _set_cell(ws, alt_note_row, 1,
              'The following products are alternative, please let me know if you need them.',
              font=_red_info, align=left_a)
    ws.row_dimensions[alt_note_row].height = 24

    # ---- 图片插入（F 列）----
    def _col_width_to_px(char_width):
        return int(char_width * 7.5 + 5)

    def _row_height_to_px(pt_height):
        return int(pt_height * 1.33)

    def _image_src_size(path):
        try:
            from PIL import Image as _PIL
            with _PIL.open(path) as _im:
                return _im.size
        except Exception:
            return None, None

    def _mark_missing_image(row_number):
        cell = ws.cell(row=row_number, column=IMAGE_COL_INDEX)
        cell.value = '/'
        cell.alignment = center
        cell.font = normal_font

    for row in range(data_start, data_end + 1):
        product_code = str(row_code_map.get(row, '') or '').strip()
        if not product_code:
            continue
        normalized_code = normalize_lookup_code(product_code)
        img_path = None
        if code_to_images:
            img_path = (code_to_images.get(product_code) or code_to_images.get(normalized_code) or [None])[0]
        if img_path and image_folder:
            col_w = ws.column_dimensions[get_column_letter(IMAGE_COL_INDEX)].width or 8.43
            row_h = ws.row_dimensions[row].height or 15
            avail_w = max(_col_width_to_px(col_w) - IMAGE_PADDING * 2, 20)
            avail_h = max(_row_height_to_px(row_h) - IMAGE_PADDING * 2, 20)
            src_w, src_h = _image_src_size(img_path)
            if not src_w or not src_h:
                src_w, src_h = AP_GROUND_IMAGE_WIDTH, AP_GROUND_IMAGE_HEIGHT
            scale = min(avail_w / src_w, avail_h / src_h, 1.0)
            disp_w = max(int(src_w * scale), 20)
            disp_h = max(int(src_h * scale), 20)
            temp_img_path = prepare_image_for_excel(
                img_path, target_width=disp_w, target_height=disp_h,
                temp_dir=image_temp_dir, cache=image_cache,
            )
            final_img_path = temp_img_path or img_path
            if add_image_centered_in_cell(ws, final_img_path, row, IMAGE_COL_INDEX,
                                          img_width=disp_w, img_height=disp_h):
                image_found_count += 1
            else:
                image_not_found_count += 1
                _mark_missing_image(row)
        else:
            image_not_found_count += 1
            _mark_missing_image(row)
    print(f"   [AP-GROUND-ND] Images: found={image_found_count}, not_found={image_not_found_count}")

    # ---- 外边缘框 ----
    _apply_outer_frame(ws, 1, total_all_row, 1, MAX_COL)

    apply_print_setup(ws, 'ap_ground')

    quotation_product_codes = {str(it['code']).strip() for it in items if it.get('code')}
    return {
        'sheet_name': sheet_name,
        'site': site,
        'quotation_product_codes': quotation_product_codes,
        'valid_products': len(items),
        'matched_count': matched_count,
        'unmatched_count': unmatched_count,
        'image_found_count': image_found_count,
        'image_not_found_count': image_not_found_count,
        'data_start_row': data_start,
        'data_end_row': data_end,
        'tables': tables,
        'rows': rows,
        'cols': cols,
        'pile_products': site.get('pile_products') or [],
    }


# ------------------------------------------------------------------ 汇总页

def create_ap_ground_no_disc_summary_sheet(
        workbook, all_detail_results, matrix_data=None, contact_info=None,
        trade_method='EXW', dest_port='', production_lead_time='4-6 Weeks after receiving deposit',
        payment_term='T/T 30% deposit,70% balance before shipping',
        validity_days=7, module_wattage=None,
        image_path=None, container_details=None,
):
    """创建汇总页 Total（无折扣版）。表头：
    NO. | Row | Column | Installation Angle | Table | Power（kw） |
    Price(USD/Table) | Amount(USD)
    """
    from datetime import datetime
    matrix_data = matrix_data or {}
    contact_info = contact_info or {}
    project_name = str(matrix_data.get('project_name') or '').strip()

    ws = workbook.create_sheet(title='Total')

    thin = Side(style='thin', color='000000')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bottom_border = Border(bottom=thin)
    title_font = Font(name='Calibri', size=26, bold=True)
    info_font = Font(name='Calibri', size=11)
    info_bold = Font(name='Calibri', size=11, bold=True)
    header_font = Font(name='Calibri', size=12, bold=True, color='000000')
    normal_font = Font(name='Calibri', size=12)
    bold_font = Font(name='Calibri', size=12, bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_a = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_a = Alignment(horizontal='right', vertical='center', wrap_text=True)

    # 汇总列宽 A~K（F-G / H-I / J-K 合并）
    _summary_widths = {'A': 6, 'B': 8, 'C': 10, 'D': 18, 'E': 8, 'F': 10,
                       'G': 10, 'H': 16, 'I': 10, 'J': 16, 'K': 10}
    for col_letter, _w in _summary_widths.items():
        ws.column_dimensions[col_letter].width = _w

    # ---- 标题 ----
    ws.merge_cells('A3:K3')
    _set_cell(ws, 3, 1, 'Solar PV Mounting System Quotation', font=title_font, align=center)
    ws.row_dimensions[3].height = 60

    if image_path and os.path.exists(image_path):
        try:
            img = XLImage(img=image_path)
            img.width = 240
            img.height = 60
            ws.add_image(img, 'H4')
        except Exception:
            pass

    # ---- 抬头信息（左字段带下框线；右：公司信息 J-K，从第7行起）----
    red_amount_font = Font(name='Calibri', size=13, bold=True, color='FFFF0000')

    def _left_field(r, label, value, wide=False, val_font=None, val_fmt=None, label_font=None):
        if wide:
            ws.merge_cells(f'A{r}:C{r}'); lcol, vcol = 1, 4
            ws.merge_cells(f'D{r}:F{r}')
        else:
            ws.merge_cells(f'A{r}:B{r}'); lcol, vcol = 1, 3
            ws.merge_cells(f'C{r}:F{r}')
        _set_cell(ws, r, lcol, label, font=label_font or info_font, align=left_a, border=bottom_border)
        _set_cell(ws, r, vcol, value, font=val_font or info_font, align=left_a,
                  border=bottom_border, number_format=val_fmt)

    _left_field(5, 'Project:', project_name)
    _left_field(6, 'Price Term：', f'{trade_method} {dest_port}'.strip().upper())
    _left_field(7, 'Production Delivery Time：', production_lead_time, wide=True)
    _left_field(8, 'Payment Term: ', payment_term)
    # EXW TOTAL AMOUNT / Unite Price/W（标签与值均为红色，值稍后回填）
    _left_field(9, 'EXW TOTAL AMOUNT:', '', wide=True, label_font=red_amount_font,
                val_font=red_amount_font, val_fmt=CURRENCY_FMT)
    _left_field(10, 'Unite Price/W:', '', wide=True, label_font=red_amount_font,
                val_font=red_amount_font, val_fmt='$#,##0.0000')
    exw_val_col, upw_val_col = 4, 4
    for _r in range(5, 11):
        ws.row_dimensions[_r].height = 20

    # DATE：I2 右对齐，=TODAY() J2 左对齐
    _set_cell(ws, 2, 9, 'DATE：', font=info_font, align=right_a)
    _set_cell(ws, 2, 10, '=TODAY()', font=info_font, align=left_a, number_format='yyyy-mm-dd')

    # 公司信息（右侧 H-K 合并，第7行起，左对齐）
    company_lines = [
        'Xiamen Kseng Metal Tech Co.,Ltd',
        'Add.: RM 601, Huixin Wealth Centre, No. 891,',
        'Fanghu North 2nd Rd, Huli Dist, Xiamen, Fujian, China',
    ]
    email = contact_info.get('tel') or 'abby.yan@xmkseng.com'
    phone = contact_info.get('phone') or '18050039573'
    company_lines.append(f'Email： {email}')
    company_lines.append(f'Phone/WhatsApp: {phone}')
    for i, line in enumerate(company_lines):
        cr = 7 + i
        ws.merge_cells(f'H{cr}:K{cr}')
        _set_cell(ws, cr, 8, line,
                  font=Font(name='Calibri', size=13, bold=True) if i == 0 else info_font,
                  align=left_a)

    # ---- Part 1 标题（A-K 合并）----
    part1_start = 14
    ws.merge_cells(f'A{part1_start}:K{part1_start}')
    _set_cell(ws, part1_start, 1, 'Part 1: Solar Mounting System',
              font=Font(name='Calibri', size=14, bold=True), align=left_a,
              border=thin_border)
    for ci in range(2, 12):
        ws.cell(row=part1_start, column=ci).border = thin_border
    ws.row_dimensions[part1_start].height = 28

    hdr_row = part1_start + 1
    headers = ['NO.', 'Row', 'Column', 'Installation Angle', 'Table', 'Power（kw）',
               '', 'Price(USD/Table)', '', 'Amount(USD)', '']
    for ci, h in enumerate(headers):
        _set_cell(ws, hdr_row, ci + 1, h, font=header_font, align=center, border=thin_border)
    ws.merge_cells(f'F{hdr_row}:G{hdr_row}')   # Power（kw）
    ws.merge_cells(f'H{hdr_row}:I{hdr_row}')   # Price(USD/Table)
    ws.merge_cells(f'J{hdr_row}:K{hdr_row}')   # Amount(USD)
    ws.row_dimensions[hdr_row].height = 30

    # ---- 数据行 ----
    data_row = hdr_row + 1
    amount_cells = []   # J 列（Amount）
    power_cells = []    # F 列（Power）
    for i, dr in enumerate(all_detail_results):
        r = data_row + i
        sn = dr.get('sheet_name', '')
        sn_esc = sn.replace("'", "''")
        ds = dr.get('data_start_row')
        de = dr.get('data_end_row')
        a_rows = dr.get('rows') or ''
        a_cols = dr.get('cols') or ''
        a_tables = dr.get('tables') or 1
        site_kw = (dr.get('site') or {}).get('kw') or 0
        _angle_val = (dr.get('site') or {}).get('angle')
        installation = normalize_angle(_angle_val) if _angle_val else (matrix_data.get('panel_orientation') or '')

        for ci in range(1, 12):
            ws.cell(row=r, column=ci).font = normal_font
            ws.cell(row=r, column=ci).border = thin_border

        _set_cell(ws, r, 1, i + 1, font=normal_font, align=center, border=thin_border)
        _set_cell(ws, r, 2, a_rows, font=normal_font, align=center, border=thin_border)
        _set_cell(ws, r, 3, a_cols, font=normal_font, align=center, border=thin_border)
        _set_cell(ws, r, 4, installation, font=normal_font, align=center, border=thin_border)
        _set_cell(ws, r, 5, a_tables, font=normal_font, align=center, border=thin_border)
        # F-G Power（kw）
        ws.merge_cells(f'F{r}:G{r}')
        if site_kw:
            _set_cell(ws, r, 6, site_kw, font=normal_font, align=center, border=thin_border,
                      number_format='0.00')
            power_cells.append(f'F{r}')
        else:
            _set_cell(ws, r, 6, '', font=normal_font, align=center, border=thin_border)
        # H-I Price(USD/Table) = 明细 I 列（每基合计）求和
        ws.merge_cells(f'H{r}:I{r}')
        if ds and de and de >= ds:
            _set_cell(ws, r, 8, f"=SUM('{sn_esc}'!{COL_TOTAL_PRICE}{ds}:{COL_TOTAL_PRICE}{de})",
                      font=normal_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
        else:
            _set_cell(ws, r, 8, 0, font=normal_font, align=center, border=thin_border,
                      number_format=CURRENCY_FMT)
        # J-K Amount(USD) = Price × Table
        ws.merge_cells(f'J{r}:K{r}')
        _set_cell(ws, r, 10, f'=H{r}*E{r}', font=normal_font, align=center,
                  border=thin_border, number_format=CURRENCY_FMT)
        amount_cells.append(f'J{r}')
        ws.row_dimensions[r].height = 24

    data_end = data_row + len(all_detail_results) - 1 if all_detail_results else data_row - 1

    # ---- Total 行 ----
    total_row = data_end + 1
    ws.merge_cells(f'A{total_row}:E{total_row}')
    _set_cell(ws, total_row, 1, 'Total', font=bold_font, align=center, border=thin_border)
    ws.merge_cells(f'F{total_row}:G{total_row}')
    if power_cells:
        _set_cell(ws, total_row, 6, f'={"+".join(power_cells)}', font=bold_font, align=center,
                  border=thin_border, number_format='0.00')
    else:
        _set_cell(ws, total_row, 6, '', font=bold_font, align=center, border=thin_border)
    ws.merge_cells(f'H{total_row}:I{total_row}')
    _set_cell(ws, total_row, 8, '', font=bold_font, align=center, border=thin_border)
    ws.merge_cells(f'J{total_row}:K{total_row}')
    if amount_cells:
        _set_cell(ws, total_row, 10, f'={"+".join(amount_cells)}', font=red_amount_font, align=center,
                  border=thin_border, number_format=CURRENCY_FMT)
    else:
        _set_cell(ws, total_row, 10, 0, font=red_amount_font, align=center, border=thin_border,
                  number_format=CURRENCY_FMT)
    ws.row_dimensions[total_row].height = 24

    orig_total_ref = f'J{total_row}'
    power_inner = "+".join(power_cells) if power_cells else None

    # ---- per/w 行 ----
    perw_row = total_row + 1
    ws.merge_cells(f'A{perw_row}:I{perw_row}')
    _set_cell(ws, perw_row, 1, 'per/w', font=bold_font, align=right_a, border=thin_border)
    for ci in range(2, 10):
        ws.cell(row=perw_row, column=ci).border = thin_border
    ws.merge_cells(f'J{perw_row}:K{perw_row}')
    if power_inner:
        _set_cell(ws, perw_row, 10, f'=ROUND({orig_total_ref}/(({power_inner})*1000),4)',
                  font=red_amount_font, align=center, border=thin_border, number_format='$#,##0.0000')
    else:
        _set_cell(ws, perw_row, 10, 0, font=red_amount_font, align=center, border=thin_border,
                  number_format='$#,##0.0000')
    ws.row_dimensions[perw_row].height = 24

    # ---- TOTAL AMOUNT(EXW) 行（仅 EXW 时显示；FOB/CIF 不显示此行）----
    is_exw = str(trade_method or '').upper() == 'EXW'
    if is_exw:
        total_exw_row = total_row + 2
        ws.merge_cells(f'A{total_exw_row}:I{total_exw_row}')
        _set_cell(ws, total_exw_row, 1, 'TOTAL AMOUNT(EXW)', font=bold_font, align=right_a, border=thin_border)
        for ci in range(2, 10):
            ws.cell(row=total_exw_row, column=ci).border = thin_border
        ws.merge_cells(f'J{total_exw_row}:K{total_exw_row}')
        _set_cell(ws, total_exw_row, 10, f'={orig_total_ref}', font=red_amount_font, align=center,
                  border=thin_border, number_format=CURRENCY_FMT)
        ws.row_dimensions[total_exw_row].height = 24
        frame_bottom = total_exw_row
    else:
        frame_bottom = perw_row

    # ---- 回填抬头 EXW TOTAL AMOUNT / Unite Price/W（红色，D 列，行与标签一致）----
    _set_cell(ws, 9, exw_val_col, f'=ROUND({orig_total_ref},2)', font=red_amount_font,
              align=left_a, number_format=CURRENCY_FMT)
    if power_inner:
        upw = f'=ROUND({orig_total_ref}/(({power_inner})*1000),4)'
    else:
        upw = 0
    _set_cell(ws, 10, upw_val_col, upw, font=red_amount_font, align=left_a, number_format='$#,##0.0000')

    # ---- Part 2: Freight & Insurance（仅 FOB/CIF 且有柜型时）----
    containers = [c for c in (container_details or []) if c.get('qty')]
    if containers and trade_method in ('FOB', 'CIF'):
        port_name = str(dest_port or 'Xiamen').upper()
        p2_title = frame_bottom + 2
        p2_hdr = p2_title + 1
        p2_val = p2_hdr + 1
        ws.merge_cells(f'A{p2_title}:K{p2_title}')
        _set_cell(ws, p2_title, 1, 'Part 2: Freight & Insurance',
                  font=Font(name='Calibri', size=14, bold=True), align=left_a,
                  border=thin_border)
        for ci in range(2, 12):
            ws.cell(row=p2_title, column=ci).border = thin_border
        ws.row_dimensions[p2_title].height = 26

        label_text = f'{trade_method} {port_name}'
        # 标签 CIF JEDDAH 占 A-C（跨表头/数值两行）
        ws.merge_cells(f'A{p2_hdr}:C{p2_val}')
        _set_cell(ws, p2_hdr, 1, label_text, font=bold_font, align=center, border=thin_border)

        # 列头：D:F Container | G Quantity | H:I Unit price | J:K Amount(USD)
        ws.merge_cells(f'D{p2_hdr}:F{p2_hdr}')
        ws.merge_cells(f'H{p2_hdr}:I{p2_hdr}')
        ws.merge_cells(f'J{p2_hdr}:K{p2_hdr}')
        for cc, txt in {4: 'Container', 7: 'Quantity', 8: 'Unit price', 10: 'Amount(USD)'}.items():
            _set_cell(ws, p2_hdr, cc, txt, font=header_font, align=center, border=thin_border)
        ws.row_dimensions[p2_hdr].height = 24

        def _num(v):
            try:
                f = float(v or 0)
            except (TypeError, ValueError):
                f = 0.0
            return int(f) if f == int(f) else f

        if len(containers) == 1:
            cd = containers[0]
            container_text = str(cd.get('type', '')).upper()
            qty_val = cd.get('qty', 0)
            unit_price_formula = f'={_num(cd.get("freight_per_unit", 0))}'
        else:
            container_text = '+'.join(
                f'{str(c.get("type", "")).upper()}*{c.get("qty", 0)}' for c in containers)
            qty_val = 1
            unit_price_formula = '=' + '+'.join(
                f'{_num(c.get("freight_per_unit", 0))}*{c.get("qty", 0)}' for c in containers)

        ws.merge_cells(f'D{p2_val}:F{p2_val}')
        _set_cell(ws, p2_val, 4, container_text, font=normal_font, align=center, border=thin_border)
        _set_cell(ws, p2_val, 7, qty_val, font=normal_font, align=center, border=thin_border)
        ws.merge_cells(f'H{p2_val}:I{p2_val}')
        _set_cell(ws, p2_val, 8, unit_price_formula, font=normal_font, align=center,
                  border=thin_border, number_format=CURRENCY_FMT)
        ws.merge_cells(f'J{p2_val}:K{p2_val}')
        _set_cell(ws, p2_val, 10, f'=G{p2_val}*H{p2_val}', font=normal_font, align=center,
                  border=thin_border, number_format=CURRENCY_FMT)
        ws.row_dimensions[p2_val].height = 30

        for _r in (p2_hdr, p2_val):
            for _ci in range(1, 12):
                ws.cell(row=_r, column=_ci).border = thin_border

        p2_total = p2_val + 1
        ws.merge_cells(f'A{p2_total}:I{p2_total}')
        _set_cell(ws, p2_total, 1, f'TOTAL AMOUNT({trade_method})', font=bold_font,
                  align=right_a, border=thin_border)
        for ci in range(2, 10):
            ws.cell(row=p2_total, column=ci).border = thin_border
        ws.merge_cells(f'J{p2_total}:K{p2_total}')
        _set_cell(ws, p2_total, 10, f'={orig_total_ref}+J{p2_val}', font=red_amount_font,
                  align=center, border=thin_border, number_format=CURRENCY_FMT)
        ws.row_dimensions[p2_total].height = 24
        frame_bottom = p2_total

    # ---- 汇总表外边缘框 ----
    _apply_outer_frame(ws, part1_start, frame_bottom, 1, 11)

    # ---- 最后一行：提示语（Calibri 18 红色，始终显示）----
    note_row = frame_bottom + 2
    ws.merge_cells(f'A{note_row}:K{note_row}')
    _red_note_font = Font(name='Calibri', size=18, color='FFFF0000')
    _set_cell(ws, note_row, 1, 'Please kindly check the detailed prices of components in the following sheet',
              font=_red_note_font, align=left_a)

    apply_print_setup(ws, 'ap_ground')

    # 汇总页置顶
    sheet_names = workbook.sheetnames
    if 'Total' in sheet_names:
        idx = sheet_names.index('Total')
        workbook.move_sheet('Total', offset=-idx)
    print(f"   [AP-GROUND-ND] Summary sheet(Total) created")
    return 'Total'
