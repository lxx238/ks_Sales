import openpyxl
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from backend.core.print_settings import apply_print_setup
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import tempfile
import shutil
import os
import uuid
from pathlib import Path
from decimal import Decimal, ROUND_HALF_UP
import re
import pandas as pd
from collections import defaultdict
from datetime import datetime

from backend.excel.reader import excel_file_compat, read_excel_compat
from backend.core.material_translate import translate_material, adjust_material_by_coating
from backend.core.shared.sheet_utils import reorder_sheets_by_matrix_array, set_page_break_preview

if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

from backend.core.shared.text_utils import (
    _strip_cjk_spec,
    extract_main_name,
    normalize_lookup_code,
    parse_decimal_number,
    normalize_excel_cell_text,
    normalize_lookup_token,
    format_number_text,
    parse_numeric_cell_text,
)
from backend.core.shared.price_utils import (
    _is_standard_priced,
    _get_discount_category,
    _fallback_category,
    _get_discount_rate,
    load_price_mapping,
    resolve_price_info,
    has_valid_price_info,
    round_to_2_decimal,
    get_temp_adjusted_base_price,
    get_temp_base_price, apply_temp_preinstall_adjustment,
)
from backend.core.shared.weight_utils import (
    extract_length_from_spec,
    calculate_bom_total_weight,
    log_weight_formula,
    calculate_report_total_weight,
    calculate_report_unit_weight,
    lookup_unit_weight_from_material_db,
)
from backend.core.shared.image_utils import (
    extract_code_from_filename,
    scan_images,
    find_latest_image_log,
    load_image_mapping_from_log,
    prepare_image_for_excel,
    add_image_centered_in_cell,
    add_image_centered_in_range,
    center_images_in_column,
)
from backend.core.shared.bom_utils import (
    get_bom_processing_rules,
    normalize_selected_bom_keys,
    build_bom_selection_key,
    extract_sheet_names_from_keys,
    quick_scan_bom_sheets,
    discover_sheet_bom_starts,
    list_bom_tables,
    collect_bom_products,
    read_bom_from_dataframe,
    parse_array_to_rows_cols,
    build_config_region,
    find_anchor_positions,
    collect_text_candidates,
    collect_numeric_candidates,
    validate_numeric_candidates,
    build_field_value_from_rule,
    extract_config_value,
    score_header_candidate,
    find_header_mapping_for_row,
    is_valid_header_mapping,
    extract_config_info,
    extract_bom_dataframe,
    CONFIG_FIELD_RULES,
    HEADER_FIELD_RULES,
)
from backend.core.shared.bom_zip_parser import (
    _build_ss_lookup,
    _resolve_cell_text,
    _resolve_all_values,
    _list_bom_tables_zip,
    _quick_get_sheet_names,
    _parse_bom_sheets_zip,
)
from backend.core.shared.product_utils import (
    _is_valid_product_code,
    _get_product_name_candidates,
    _match_exclude_group,
    _split_pile_products,
    normalize_preinstall,
)
from backend.core.en_common.quotation_engine import create_total_materials_sheet
from backend.core.shared.constants import (
    IMAGE_WIDTH,
    IMAGE_HEIGHT,
    IMAGE_COLUMN_INDEX,
    IMAGE_PADDING,
    CARBON_STEEL_PRICING_ATTRS,
    PURCHASED_PRICING_ATTRS,
    EXCLUDE_ITEM_GROUPS,
)
from backend.core.shared.cache_utils import (
    _parse_md_cache_key,
    _store_parse_md,
    _get_parse_md,
)

_KO_THIN = Side(style='thin', color='000000')
_KO_THICK = Side(style='medium', color='000000')
_KO_THIN_BORDER = Border(left=_KO_THIN, right=_KO_THIN, top=_KO_THIN, bottom=_KO_THIN)
_KO_NO_BORDER = Border()
_KO_BORDER_TOP = Border(left=_KO_THIN, right=_KO_THIN, top=_KO_THICK, bottom=_KO_THIN)
_KO_BORDER_BOTTOM = Border(left=_KO_THIN, right=_KO_THIN, top=_KO_THIN, bottom=_KO_THICK)
_KO_BORDER_LEFT = Border(left=_KO_THICK, right=_KO_THIN, top=_KO_THIN, bottom=_KO_THIN)
_KO_BORDER_RIGHT = Border(left=_KO_THIN, right=_KO_THICK, top=_KO_THIN, bottom=_KO_THIN)
_KO_BORDER_TL = Border(left=_KO_THICK, right=_KO_THIN, top=_KO_THICK, bottom=_KO_THIN)
_KO_BORDER_TR = Border(left=_KO_THIN, right=_KO_THICK, top=_KO_THICK, bottom=_KO_THIN)
_KO_BORDER_BL = Border(left=_KO_THICK, right=_KO_THIN, top=_KO_THIN, bottom=_KO_THICK)
_KO_BORDER_BR = Border(left=_KO_THIN, right=_KO_THICK, top=_KO_THIN, bottom=_KO_THICK)

_KO_TITLE_FONT = Font(name='Malgun Gothic', size=36, bold=True, color='000000')
_KO_NORMAL_FONT = Font(name='Malgun Gothic', size=16, color='000000')
_KO_HEADER_FONT = Font(name='Malgun Gothic', size=16, bold=True, color='000000')
_KO_SMALL_BOLD_FONT = Font(name='Malgun Gothic', size=16, bold=True, color='000000')
_KO_RED_SMALL_FONT = Font(name='Malgun Gothic', size=16, bold=True, color='FF0000')
_KO_A7_FONT = Font(name='Malgun Gothic', size=16, bold=True, color='FFFFFF')
_KO_A7_FILL = PatternFill(start_color='E7F5FF', end_color='E7F5FF', fill_type='solid')
_KO_YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
_KO_CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_KO_LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
_KO_RIGHT_ALIGN = Alignment(horizontal='right', vertical='center', wrap_text=True)

_KO_INQUIRY_HEADER_FILL_1 = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_KO_INQUIRY_HEADER_FILL_2 = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_KO_INQUIRY_HEADER_FILL_3 = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_KO_INQUIRY_HEADER_FONT = Font(bold=True)
_KO_INQUIRY_NORMAL_FONT = Font(bold=False)
_KO_INQUIRY_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
_KO_INQUIRY_LEFT_ALIGN = Alignment(horizontal="left", vertical="center")

_KO_SUMMARY_MALGUN = Font(name='Malgun Gothic')
_KO_SUMMARY_MALGUN_BOLD = Font(name='Malgun Gothic', bold=True)
_KO_SUMMARY_MALGUN_BOLD_16 = Font(name='Malgun Gothic', bold=True, size=16)
_KO_SUMMARY_MALGUN_SMALL = Font(name='Malgun Gothic', size=8)
_KO_SUMMARY_MALGUN_SMALL_RED = Font(name='Malgun Gothic', size=8, color='FF0000')
_KO_SUMMARY_LIGHT_BLUE = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

from backend.core.array_matcher import (
    build_matrix_array_entries,
    find_matching_matrix_array,
    build_bom_matrix_data,
    merge_bom_products,
    find_accumulated_match,
    find_info_accumulated_match,
)

data_start_row = 10


def delete_empty_weight_rows_and_renumber(ws, data_start_row):
    """删除重量（I列）为空的行，并重新编号"""
    print("\n🗑️ 开始清理重量为空的行...")

    rows_to_delete = []
    current_row = data_start_row
    while current_row <= ws.max_row:
        weight_cell = ws.cell(row=current_row, column=9)
        weight_value = weight_cell.value

        is_empty = False
        if weight_value is None or weight_value == '':
            is_empty = True
        elif isinstance(weight_value, (int, float)) and weight_value == 0:
            is_empty = True
        elif isinstance(weight_value, str):
            try:
                match = re.search(r'(\d+(?:\.\d+)?)', weight_value)
                if match and float(match.group(1)) == 0:
                    is_empty = True
            except:
                pass

        if is_empty:
            name_cell = ws.cell(row=current_row, column=2)
            if name_cell.value is None or name_cell.value == '':
                rows_to_delete.append(current_row)
            else:
                rows_to_delete.append(current_row)
        current_row += 1

    deleted_count = 0
    if rows_to_delete:
        sorted_rows = sorted(rows_to_delete)
        i = 0
        while i < len(sorted_rows):
            start = sorted_rows[i]
            count = 1
            while i + count < len(sorted_rows) and sorted_rows[i + count] == start + count:
                count += 1
            ws.delete_rows(start, count)
            deleted_count += count
            i += count

    print(f"✅ 已删除 {deleted_count} 行")

    print("\n🔢 重新编号序号列...")
    new_seq = 1
    current_row = data_start_row
    while current_row <= ws.max_row:
        name_cell = ws.cell(row=current_row, column=2)
        if name_cell.value and name_cell.value != '':
            ws.cell(row=current_row, column=1, value=new_seq)
            new_seq += 1
        current_row += 1

    # ========== 公式修复 ==========
    print("\n📐 修复公式引用...")
    data_rows = []
    current_row = data_start_row
    while current_row <= ws.max_row:
        name_cell = ws.cell(row=current_row, column=2)
        if name_cell.value and name_cell.value != '':
            data_rows.append(current_row)
        current_row += 1

    for r in data_rows:
        ws.cell(row=r, column=8, value=f"=F{r}*G{r}")

    current_row = data_start_row
    while current_row <= ws.max_row:
        h_cell = ws.cell(row=current_row, column=8)
        if h_cell.value and isinstance(h_cell.value, str) and h_cell.value.startswith('=SUM(H'):
            group_start = None
            group_end = None
            scan = current_row - 1
            while scan >= data_start_row:
                name_cell = ws.cell(row=scan, column=2)
                if name_cell.value and name_cell.value != '':
                    if group_end is None:
                        group_end = scan
                    group_start = scan
                else:
                    break
                scan -= 1
            if group_start and group_end:
                h_cell.value = f'=SUM(H{group_start}:H{group_end})'
        current_row += 1
    print(f"✅ 公式修复完成")

    print(f"✅ 清理完成，剩余 {new_seq - 1} 个有效产品")
    return ws.max_row


def create_inquiry_sheet(workbook, unmatched_products, source_sheet_name, inquiry_requester=''):
    """
    创建询价表，将未匹配价格的产品填入
    返回创建的工作表对象
    """
    seen_keys = {}
    deduped_products = []
    for p in unmatched_products:
        code = p.get('code', '')
        spec = p.get('spec', '')
        key = (code, spec)
        if key in seen_keys:
            idx = seen_keys[key]
            deduped_products[idx]['quantity'] = float(deduped_products[idx].get('quantity', 0)) + float(p.get('quantity', 0))
        else:
            seen_keys[key] = len(deduped_products)
            deduped_products.append(dict(p))
            deduped_products[-1]['quantity'] = float(p.get('quantity', 0))

    print(f"\n📋 创建询价表，包含 {len(deduped_products)} 个未匹配价格的产品（去重前 {len(unmatched_products)} 个）")

    sheet_name = f"询价表"
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]

    original_sheet_name = sheet_name
    counter = 1
    while sheet_name in workbook.sheetnames:
        sheet_name = f"{original_sheet_name}_{counter}"
        counter += 1

    ws = workbook.create_sheet(title=sheet_name)

    thin_border = _KO_THIN_BORDER
    header_fill_1 = _KO_INQUIRY_HEADER_FILL_1
    header_fill_2 = _KO_INQUIRY_HEADER_FILL_2
    header_fill_3 = _KO_INQUIRY_HEADER_FILL_3
    header_font = _KO_INQUIRY_HEADER_FONT
    normal_font = _KO_INQUIRY_NORMAL_FONT
    center_alignment = _KO_INQUIRY_CENTER_ALIGN
    left_alignment = _KO_INQUIRY_LEFT_ALIGN

    # 设置列宽
    column_widths = [
        13.0, 18, 22.175, 74.875, 59.8333333333333, 26.4083333333333, 32.1833333333333, 36.0916666666667,
        19.5, 13.0, 5.25, 4.25, 9.375, 13.0, 12.5, 13.0, 20.375,
        15.125, 13.0, 7.125, 10.5, 9.0, 15.125, 7.125, 30.25,
        10.5, 13.0, 13.0, 10.375, 11.125, 7.125, 9.375
    ]

    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 设置第一行表头（三个合并区域）
    ws.merge_cells('A1:H1')
    ws['A1'] = "业务/研发沟通确定"
    ws['A1'].fill = header_fill_1
    ws['A1'].font = header_font
    ws['A1'].alignment = center_alignment
    ws['A1'].border = thin_border

    ws.merge_cells('I1:R1')
    ws['I1'] = "研发提供"
    ws['I1'].fill = header_fill_2
    ws['I1'].font = header_font
    ws['I1'].alignment = center_alignment
    ws['I1'].border = thin_border

    ws.merge_cells('S1:AF1')
    ws['S1'] = "采购提供"
    ws['S1'].fill = header_fill_3
    ws['S1'].font = header_font
    ws['S1'].alignment = center_alignment
    ws['S1'].border = thin_border

    # 设置第二行表头
    header_row2 = [
        "询价人", "询价日期", "产品编码", "产品名称", "规格 (mm)", "数量", "销售单位", "备注",
        "产品类别", "售价", "售价单位", "", "报价日期", "表面处理", "重量属性 (kg)", "物料总重(kg)", "备注",
        "是否带图档附件", "产品归属项目", "采购员", "采购单价", "采购单位", "是否含税含运费",
        "供应商", "备注 (不同起订量不同单价之类)", "挤压模", "费用", "冲孔模", "费用",
        "最低采购量", "上机费", "其它备注"
    ]

    # 需要加粗的列
    bold_columns = [3, 7, 9, 10, 21]

    for col_idx, value in enumerate(header_row2, start=1):
        if col_idx == 12:  # 跳过 L2（因为要与K列合并）
            continue

        cell = ws.cell(row=2, column=col_idx, value=value)

        # 根据列所属区域设置背景色
        if col_idx <= 8:  # A-H 列：业务/研发沟通确定
            cell.fill = header_fill_1
        elif col_idx <= 18:  # I-R 列：研发提供
            cell.fill = header_fill_2
        else:  # S-AF 列：采购提供
            cell.fill = header_fill_3

        if col_idx in bold_columns:
            cell.font = header_font
        else:
            cell.font = normal_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # 合并 K2:L2
    ws.merge_cells('K2:L2')

    # 获取当前日期
    current_date = datetime.now().strftime("%Y.%m.%d")

    # 填入数据行
    current_row = 3
    inquiry_requester = str(inquiry_requester or '').strip()

    for product in deduped_products:
        # A列: 询价人
        ws.cell(row=current_row, column=1, value=inquiry_requester)
        # B列: 询价日期
        ws.cell(row=current_row, column=2, value=current_date)
        # C列: 产品编码
        ws.cell(row=current_row, column=3, value=product['code'])
        # D列: 产品名称
        ws.cell(row=current_row, column=4, value=product['name'])
        # E列: 规格 (mm)
        try:
            _spec_num = float(product['spec'])
            ws.cell(row=current_row, column=5, value=_spec_num)
            ws.cell(row=current_row, column=5).number_format = '"L"!=#"mm"'
        except (ValueError, TypeError):
            ws.cell(row=current_row, column=5, value=_strip_cjk_spec(product['spec']))
        # F列: 数量
        quantity = product['quantity']
        if quantity > 0:
            ws.cell(row=current_row, column=6, value=int(quantity) if quantity % 1 == 0 else quantity)
        else:
            ws.cell(row=current_row, column=6, value="")
        # G列: 销售单位（留空）
        # H列: 备注

        # 为这一行的所有单元格设置边框和对齐
        ws.cell(row=current_row, column=7, value=product.get('unit', ''))
        ws.cell(row=current_row, column=8, value='铝价数据库无法匹配')

        # O列: 单重（重量属性）— 查询物料库单重，按规格折算，回退 BOM 单重
        _uw = lookup_unit_weight_from_material_db(
            product.get('code', ''), product.get('spec', ''), product.get('weight', 0)
        )
        if _uw:
            ws.cell(row=current_row, column=15, value=float(round(_uw, 4)))
        # P列: 物料总重 = 单重 × 数量
        if _uw and quantity:
            try:
                _qty_f = float(quantity)
            except (TypeError, ValueError):
                _qty_f = 0.0
            if _qty_f > 0:
                ws.cell(row=current_row, column=16, value=round(_uw * _qty_f, 2))

        for col in range(1, 33):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = left_alignment

        current_row += 1

    # 设置行高
    ws.row_dimensions[1].height = 20.1
    ws.row_dimensions[2].height = 20.1
    for row in range(3, current_row):
        ws.row_dimensions[row].height = 25.0

    # 冻结窗格
    ws.freeze_panes = 'B3'

    print(f"   ✅ 询价表创建成功: {sheet_name} (共 {len(deduped_products)} 行数据)")

    return ws


def save_inquiry_sheet_to_file(unmatched_products, output_dir, input_filename=None, inquiry_requester=''):
    """
    将询价表保存为独立的Excel文件
    按产品编码去重，相同编码只保留一条
    """
    if not unmatched_products:
        return None

    # 创建新的工作簿
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 创建询价表
    ws = create_inquiry_sheet(wb, unmatched_products, "汇总", inquiry_requester=inquiry_requester)

    # 生成文件名 - 直接命名为"询价表"
    filename = f"询价表.xlsx"

    # 如果文件已存在，添加时间戳
    output_file = os.path.join(output_dir, filename)
    counter = 1
    while os.path.exists(output_file):
        filename = f"询价表_{counter}.xlsx"
        output_file = os.path.join(output_dir, filename)
        counter += 1

    # 保存文件
    wb.save(output_file)
    print(f"\n📋 询价表已保存为独立文件: {output_file}")
    print(f"   📊 包含 {len(unmatched_products)} 个未匹配价格的产品")

    return output_file


def _create_pile_detail_sheet(workbook, pile_products, price_mapping,
                              image_path=None, image_folder=None,
                              code_to_images=None, image_temp_dir=None,
                              image_cache=None, group=None,
                              unmatched_products_out=None):
    sheet_name = '지주(地桩)'
    if sheet_name in workbook.sheetnames:
        idx = 2
        while f'{sheet_name}_{idx}' in workbook.sheetnames:
            idx += 1
        sheet_name = f'{sheet_name}_{idx}'

    ws = workbook.create_sheet(title=sheet_name)

    thin_border = _KO_THIN_BORDER
    normal_font = _KO_NORMAL_FONT
    header_font = _KO_HEADER_FONT
    small_bold_font = _KO_SMALL_BOLD_FONT
    center_align = _KO_CENTER_ALIGN
    left_align = _KO_LEFT_ALIGN
    yellow_fill = _KO_YELLOW_FILL

    column_widths = {
        'A': 15.77, 'B': 28.67, 'C': 45, 'D': 30.03, 'E': 34.58,
        'F': 24.58, 'G': 25.49, 'H': 24.13, 'I': 22.31, 'J': 25.49,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    def _col_width_to_px(char_width):
        return int(char_width * 7.5 + 5)

    def _row_height_to_px(pt_height):
        return int(pt_height * 1.33)

    def _fit_image_to_cell(cell_row, cell_col, max_w, max_h, padding=IMAGE_PADDING):
        col_w = ws.column_dimensions[get_column_letter(cell_col)].width or 8.43
        row_h = ws.row_dimensions[cell_row].height or 15
        avail_w = _col_width_to_px(col_w) - padding * 2
        avail_h = _row_height_to_px(row_h) - padding * 2
        avail_w = max(avail_w, 20)
        avail_h = max(avail_h, 20)
        scale = min(avail_w / max_w, avail_h / max_h, 1.0)
        return int(max_w * scale), int(max_h * scale)

    headers = [
        ('번호\nItem No.', 'A'), ('상품명\nProduct Name', 'B'), ('제조 재료\nMaterial', 'C'),
        ('사진\nPicture', 'D'), ('규격\nSpec.', 'E'), ('단가\nUnit Price (US$)\nEX Works', 'F'),
        ('수량\nQTY (PCS)', 'G'), ('총가격\nTotal Price (US$)\nEX Works', 'H'),
        ('무게\nWeight(KG)', 'I'), ('제품코드\nRemark', 'J')
    ]
    for text, col in headers:
        cell = ws[f'{col}{data_start_row - 1}']
        cell.value = text
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    name_field = {
        '英语组': 'name_en',
        '日语组': 'name_ja',
    }.get(group or '', 'name_ko')

    pile_products = [p for p in pile_products if p.get('quantity') and p.get('quantity', 0) > 0]

    total_price_sum = Decimal('0')
    total_weight_sum = Decimal('0')
    total_weight_actual_sum = Decimal('0')
    matched_count = 0
    unmatched_count = 0
    image_found_count = 0
    image_not_found_count = 0

    for idx, product in enumerate(pile_products):
        row = data_start_row + idx
        ws.row_dimensions[row].height = 61
        product_code = product.get('code', '')
        price_info = product.get('_price_info') or resolve_price_info(price_mapping, product_code, spec=product.get('spec', ''))
        if not product.get('_price_info') and price_info:
            product['_price_info'] = price_info
        _name_info = price_info or product.get('_price_info_no_spec') or resolve_price_info(price_mapping, product_code)
        if not product.get('_price_info_no_spec') and _name_info:
            product['_price_info_no_spec'] = _name_info
        display_name = (
            _name_info.get(name_field)
            or _name_info.get('name_ko')
            or _name_info.get('name')
            or product.get('name', '')
        ) if _name_info else product.get('name', '')

        ws.cell(row=row, column=1, value='=ROW()-9')
        ws.cell(row=row, column=2, value=display_name)
        _raw_mat = (price_info.get('db_material') if price_info and price_info.get('db_material') else None) or product.get('material', '')
        ws.cell(row=row, column=3, value=adjust_material_by_coating(translate_material(_raw_mat, 'ko'), coating_thickness))
        ws.cell(row=row, column=4, value="")
        try:
            _spec_num = float(product.get('spec', ''))
            ws.cell(row=row, column=5, value=_spec_num)
            ws.cell(row=row, column=5).number_format = '"L"!=#"mm"'
        except (ValueError, TypeError):
            ws.cell(row=row, column=5, value=_strip_cjk_spec(product.get('spec', '')))


        unit_price = 0
        is_matched = False
        if price_info and has_valid_price_info(price_info):
            unit_price = float(price_info['price'])
            matched_count += 1
            is_matched = True
        else:
            unmatched_count += 1
            if unmatched_products_out is not None and product.get('quantity', 0) > 0 and _is_valid_product_code(product_code):
                unmatched_products_out.append({
                    'code': product_code,
                    'name': (_name_info.get('name') if _name_info else None) or product.get('name', ''),
                    'spec': product.get('spec', ''),
                    'quantity': product.get('quantity', 0),
                    'weight': product.get('weight', 0),
                    'preinstall': normalize_preinstall(product.get('preinstall')),
                    'issue_reason': '地桩产品无匹配价格',
                })

        if unit_price > 0:
            ws.cell(row=row, column=6, value=float(unit_price))
        else:
            ws.cell(row=row, column=6, value="")
        ws.cell(row=row, column=6).number_format = currency_number_format

        quantity = product.get('quantity', 0)
        if quantity > 0:
            ws.cell(row=row, column=7, value=int(quantity) if quantity % 1 == 0 else quantity)
        else:
            ws.cell(row=row, column=7, value="")

        total_price = Decimal('0')
        if unit_price > 0 and quantity > 0:
            total_price = Decimal(str(unit_price)) * Decimal(str(quantity))
            ws.cell(row=row, column=8, value=f"=F{row}*G{row}")
            total_price_sum += total_price
        else:
            ws.cell(row=row, column=8, value=f"=F{row}*G{row}")
        ws.cell(row=row, column=8).number_format = currency_number_format

        weight_cell = ws.cell(row=row, column=9)
        unit_weight = calculate_report_unit_weight(product, price_info)
        if unit_weight is not None and unit_weight > 0:
            weight_cell.value = float(round_to_2_decimal(unit_weight))
            weight_cell.number_format = '#,##0.00'
            weight_cell.alignment = center_align
            total_weight_sum += unit_weight
            total_weight_actual_sum += unit_weight * Decimal(str(quantity))
        else:
            weight_cell.value = ""
            weight_cell.alignment = center_align

        ws.cell(row=row, column=10, value=product_code)

        for col in [1, 2, 3, 4, 5, 6, 7, 8, 10]:
            c = ws.cell(row=row, column=col)
            c.alignment = center_align
            c.font = normal_font
        weight_cell.font = normal_font

        if not is_matched:
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = yellow_fill

    data_end_row = data_start_row + len(pile_products) - 1

    for row in range(data_start_row, data_end_row + 1):
        product_code = str(ws.cell(row=row, column=10).value or '').strip()
        if not product_code:
            continue
        normalized_code = normalize_lookup_code(product_code)
        img_path = None
        if image_folder and code_to_images:
            if product_code in code_to_images and code_to_images[product_code]:
                img_path = code_to_images[product_code][0]
            elif normalized_code in code_to_images and code_to_images[normalized_code]:
                img_path = code_to_images[normalized_code][0]
        if img_path:
            fit_w, fit_h = _fit_image_to_cell(row, IMAGE_COLUMN_INDEX, IMAGE_WIDTH, IMAGE_HEIGHT)
            temp_img_path = prepare_image_for_excel(
                img_path, target_width=fit_w, target_height=fit_h,
                temp_dir=image_temp_dir, cache=image_cache,
            )
            final_img_path = temp_img_path if temp_img_path else img_path
            success = add_image_centered_in_cell(
                ws, final_img_path, row, IMAGE_COLUMN_INDEX,
                img_width=fit_w, img_height=fit_h,
            )
            if success:
                image_found_count += 1
                continue
            image_not_found_count += 1
        else:
            image_not_found_count += 1
        ws.cell(row=row, column=IMAGE_COLUMN_INDEX).value = "/"
        ws.cell(row=row, column=IMAGE_COLUMN_INDEX).alignment = center_align
        ws.cell(row=row, column=IMAGE_COLUMN_INDEX).font = normal_font

    sub_row = data_end_row + 1
    ws.merge_cells(f'A{sub_row}:G{sub_row}')
    ws[f'A{sub_row}'] = 'SUB-TOTAL AMOUNT (지주/말뚝)'
    ws[f'A{sub_row}'].font = small_bold_font
    ws[f'A{sub_row}'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    ws.cell(row=sub_row, column=8, value=f'=SUM(H{data_start_row}:H{data_end_row})')
    ws.cell(row=sub_row, column=8).number_format = currency_number_format
    ws.cell(row=sub_row, column=8).font = small_bold_font
    ws.cell(row=sub_row, column=8).alignment = center_align

    for r in range(2, sub_row + 1):
        for c in range(1, 11):
            ws.cell(row=r, column=c).border = thin_border

    print(f"   🖼️ 地桩시트: 삽입 {image_found_count} 장, 미검출 {image_not_found_count} 장")

    apply_print_setup(ws, 'ko_normal')

    return {
        'sheet_name': sheet_name,
        'sub_total_row': sub_row,
        'total_price_per_base': float(round_to_2_decimal(total_price_sum)),
        'total_weight_per_base': float(round_to_2_decimal(total_weight_actual_sum)),
        'matched_count': matched_count,
        'unmatched_count': unmatched_count,
        'image_found_count': image_found_count,
        'image_not_found_count': image_not_found_count,
    }


def _split_pile_products(products, price_mapping):
    bracket_products = []
    pile_products = []
    for p in products:
        code = p.get('code', '')
        price_info = resolve_price_info(price_mapping, code, spec=p.get('spec', '')) if price_mapping else None
        attr = (price_info.get('attribute', '') if price_info else '') or ''
        if attr.strip() in ('地桩', '地盤杭'):
            pile_products.append(p)
        else:
            bracket_products.append(p)
    return bracket_products, pile_products


def create_quotation_from_dataframe(
        df,
        workbook,
        sheet_name_prefix,
        price_mapping,
        image_path=None,
        image_folder=None,
        code_to_images=None,
        image_temp_dir=None,
        image_cache=None,
        unmatched_products_list=None,
        contact_info=None,
        config=None,
        matrix_data=None,
        group=None,
        exclude_options=None,
        sale_type='export',
        ko_discount_rate=100,
        ko_steel_discount_rate=84,
        ko_purchased_discount_rate=94,
        coating_thickness=10,
        delete_options=None,
        always_exclude_extra_items=False,
        ko_exclude_options=None,
        pre_parsed_products=None,
        ko_discount_in_detail=True,
):
    """
    从DataFrame创建报价表（直接添加到工作簿中）
    计价单位从价格表获取
    计价逻辑：
    - 价格表单位=米：总价 = (长度/1000) × 单价 × 数量
    - 价格表单位=个/套/支：总价 = 单价 × 数量
    """
    from backend.core.shared.bom_utils import resolve_products_and_array as _resolve
    all_products, array_info, span_info, rows, cols = _resolve(
        pre_parsed_products, df, matrix_data,
    )

    from backend.core.ko_ksd.quotation_engine import _classify_products_single_pass

    delete_options = delete_options or {}
    ko_exclude_options = ko_exclude_options or {}
    if exclude_options and any(exclude_options.values()):
        for k, v in exclude_options.items():
            if v:
                ko_exclude_options[k] = True

    aluminum, carbon_steel, excluded_products, pile_products, price_info_cache = _classify_products_single_pass(
        all_products, price_mapping, delete_options,
        always_exclude_extra_items, ko_exclude_options,
    )
    all_products = aluminum + carbon_steel

    _codes_with_spec = set()
    for _dp in all_products:
        _dp_spec = str(_dp.get('spec', '')).strip()
        if _dp_spec and _dp_spec not in ('0', '请输入规格'):
            _codes_with_spec.add(_dp.get('code', ''))
    if _codes_with_spec:
        _before_dedup = len(all_products)
        all_products = [
            p for p in all_products
            if p.get('code', '') not in _codes_with_spec or str(p.get('spec', '')).strip()
        ]
        _dedup_removed = _before_dedup - len(all_products)
        if _dedup_removed:
            print(f'   🧹 去重移除空规格重复产品: {_dedup_removed} 项')

    if excluded_products:
        print(f'[EXCLUDE] {len(excluded_products)} items moved to extra section')
    if pile_products:
        print(f"   🔩 地桩产品分离: {len(pile_products)} 项地桩, {len(all_products)} 项架台")

    _pre_name_field = {
        '英语组': 'name_en',
        '日语组': 'name_ja',
    }.get(group or '', 'name_ko')

    _empty_weight_products = []
    for _fwp in all_products:
        if _fwp.get('_is_pile'):
            continue
        _fwp_code = _fwp.get('code', '')
        _fwp_spec = _fwp.get('spec', '')
        _fwp_pi = _fwp.get('_price_info') or resolve_price_info(price_mapping, _fwp_code, spec=_fwp_spec)
        _fwp_uw = calculate_report_unit_weight(_fwp, _fwp_pi)
        if _fwp_uw is None or _fwp_uw <= 0:
            _empty_weight_products.append((_fwp, _fwp_pi))

    total_products = len(all_products)

    matrix_data = matrix_data or {}
    matrix_project_name = str(matrix_data.get('project_name') or '').strip()
    matrix_output_wp = matrix_data.get('output_wp')
    matrix_set_count = matrix_data.get('set_count')
    matrix_max_wind = matrix_data.get('max_wind_speed')
    matrix_max_snow = matrix_data.get('max_snow_load')
    matrix_module_watt = matrix_data.get('module_wattage')
    matrix_module_size = matrix_data.get('module_size')
    if not isinstance(matrix_set_count, int) or matrix_set_count <= 0:
        matrix_set_count = None
    total_table_count = matrix_set_count or 1

    sheet_name = extract_main_name(sheet_name_prefix)
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]

    original_sheet_name = sheet_name
    counter = 1
    while sheet_name in workbook.sheetnames:
        sheet_name = f"{original_sheet_name}_{counter}"
        counter += 1

    ws = workbook.create_sheet(title=sheet_name)
    print(f"📝 创建工作表: {sheet_name}")

    # ========== 样式定义 ==========
    thin = _KO_THIN
    thick = _KO_THICK
    thin_border = _KO_THIN_BORDER
    no_border = _KO_NO_BORDER

    title_font = _KO_TITLE_FONT
    normal_font = _KO_NORMAL_FONT
    header_font = _KO_HEADER_FONT
    small_bold_font = _KO_SMALL_BOLD_FONT
    red_small_font = _KO_RED_SMALL_FONT
    if sale_type == 'domestic':
        currency_number_format = '"¥" #,##0.00'
    else:
        currency_number_format = '#,##0.00'

    a7_font = _KO_A7_FONT
    a7_fill = _KO_A7_FILL

    yellow_fill = _KO_YELLOW_FILL

    center_align = _KO_CENTER_ALIGN
    left_align = _KO_LEFT_ALIGN
    right_align = _KO_RIGHT_ALIGN

    # ========== 设置列宽 ==========
    column_widths = {
        'A': 15.77, 'B': 28.67, 'C': 43, 'D': 30.03, 'E': 34.58,
        'F': 35, 'G': 25.49, 'H': 35, 'I': 22.31, 'J': 25.49,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    def _col_width_to_px(char_width):
        return int(char_width * 7.5 + 5)

    def _row_height_to_px(pt_height):
        return int(pt_height * 1.33)

    def _fit_image_to_cell(cell_row, cell_col, max_w, max_h, padding=IMAGE_PADDING):
        col_w = ws.column_dimensions[get_column_letter(cell_col)].width or 8.43
        row_h = ws.row_dimensions[cell_row].height or 15
        avail_w = _col_width_to_px(col_w) - padding * 2
        avail_h = _row_height_to_px(row_h) - padding * 2
        avail_w = max(avail_w, 20)
        avail_h = max(avail_h, 20)
        scale = min(avail_w / max_w, avail_h / max_h, 1.0)
        return int(max_w * scale), int(max_h * scale)

    # ========== 先临时创建所有数据行 ==========
    data_start_row = 10

    extra_rows = len(excluded_products) + 6
    for row in range(1, data_start_row + total_products + extra_rows):
        if row not in [1, 2, 3, 4, 5, 6, 7, 8, 9]:
            ws.row_dimensions[row].height = 61

    row_heights_temp = {
        1: 18, 2: 66, 3: 30, 4: 30, 5: 30, 6: 30, 7: 30, 8: 20, 9: 75,
    }
    for row, height in row_heights_temp.items():
        ws.row_dimensions[row].height = height

    # ========== 第1行: 空行 ==========
    for col in range(1, 11):
        ws.cell(row=1, column=col)

    # ========== 第2行: 标题 ==========
    ws.merge_cells('A2:J2')
    ws['A2'] = matrix_project_name or '태양광 시스템 견적서'
    ws['A2'].font = title_font
    ws['A2'].alignment = center_align

    # ========== 合并 A3:B6 区域并嵌入图片 ==========
    ws.merge_cells('A3:B6')
    if image_path and os.path.exists(image_path):
        try:
            img = XLImage(img=image_path)
            width_a = ws.column_dimensions['A'].width
            width_b = ws.column_dimensions['B'].width
            total_width_pixels = (width_a + width_b) * 7

            height_4 = ws.row_dimensions[4].height
            height_5 = ws.row_dimensions[5].height
            height_6 = ws.row_dimensions[6].height
            total_height_pts = height_4 + height_5 + height_6
            total_height_pixels = total_height_pts * 1.33

            target_width = total_width_pixels * 1.05
            target_height = total_height_pixels * 1.8

            original_width = img.width
            original_height = img.height
            original_ratio = original_width / original_height
            target_ratio = target_width / target_height

            if original_ratio > target_ratio:
                img.width = target_width
                img.height = target_width / original_ratio
            else:
                img.height = target_height
                img.width = target_height * original_ratio

            min_size = 50
            if img.width < min_size or img.height < min_size:
                img.width = max(img.width, min_size)
                img.height = max(img.height, min_size)

            ws.add_image(img, 'A4')
        except Exception as e:
            ws['A4'] = '[图片未找到或加载失败]'
            ws['A4'].font = normal_font
            ws['A4'].alignment = center_align
    else:
        ws['A3'] = ''
        ws['A3'].alignment = center_align

    # ========== 第3-7行: 项目信息区域 ==========
    contact_defaults = {
        'contact_name': '진설정',
        'phone': '0086-18050053693',
        'tel': 'seol@xmkseng.com',
        'fax': ''
    }
    contact_info = contact_info or {}
    contact_name = contact_info.get('contact_name') or contact_defaults['contact_name']
    phone = contact_info.get('phone') or contact_defaults['phone']
    tel = contact_info.get('tel') or contact_defaults['tel']
    fax = contact_info.get('fax') or contact_defaults['fax']

    ws['C3'] = f"담당자:{contact_name}"
    ws['C3'].font = normal_font
    ws['C3'].alignment = center_align

    ws['D3'] = '설치각도'
    ws['D3'].font = normal_font
    ws['D3'].alignment = center_align

    angle_val = config.get('angle', '') if config else ''
    angle_clean = str(angle_val).rstrip('°').strip() if angle_val else ''
    ws['E3'] = f"{angle_clean}°" if angle_clean else ''
    ws['E3'].font = normal_font
    ws['E3'].alignment = center_align

    ws['F3'] = '모듈사이즈'
    ws['F3'].font = normal_font
    ws['F3'].alignment = center_align

    ws['C4'] = f"Tel: {phone}"
    ws['C4'].font = normal_font
    ws['C4'].alignment = center_align

    ws['D4'] = '최대풍속'
    ws['D4'].font = normal_font
    ws['D4'].alignment = center_align

    ws['E4'] = matrix_max_wind or ''
    ws['E4'].font = normal_font
    ws['E4'].alignment = center_align

    ws['F4'] = '모듈 발전량'
    ws['F4'].font = normal_font
    ws['F4'].alignment = center_align

    ws['G4'] = f"{matrix_module_watt}w" if matrix_module_watt else ''
    ws['G4'].font = normal_font
    ws['G4'].alignment = center_align

    ws['C5'] = f"Email: {tel}"
    ws['C5'].font = normal_font
    ws['C5'].alignment = center_align

    ws['D5'] = '최대 적설량'
    ws['D5'].font = normal_font
    ws['D5'].alignment = center_align

    ws['E5'] = matrix_max_snow or ''
    ws['E5'].font = normal_font
    ws['E5'].alignment = center_align

    ws['F5'] = '전체 발전량'
    ws['F5'].font = normal_font
    ws['F5'].alignment = center_align

    if matrix_output_wp:
        ws['G5'] = f"{matrix_output_wp} Wp"
    else:
        ws['G5'] = ''
    ws['G5'].font = normal_font
    ws['G5'].alignment = center_align

    missing_boards_val = matrix_data.get('missing_per_table', 0) or 0
    if missing_boards_val == 0:
        _raw_mb = int((config or {}).get('missing_boards', 0) or 0)
        missing_boards_val = -_raw_mb if _raw_mb > 0 else _raw_mb
    ws['C6'] = '빠진 모듈개수'
    ws['C6'].font = normal_font
    ws['C6'].alignment = center_align

    ws['D6'] = str(missing_boards_val) if missing_boards_val != 0 else '/'
    ws['D6'].font = normal_font
    ws['D6'].alignment = center_align

    if span_info:
        ws['G6'] = f"{span_info} mm"
    else:
        ws['G6'] = ''
    ws['G6'].font = normal_font
    ws['G6'].alignment = center_align

    ws['F6'] = '경간(E/W)'
    ws['F6'].font = normal_font
    ws['F6'].alignment = center_align

    # ========== 新增：填写板规到 G3 ==========
    if matrix_module_size:
        panel_spec = re.sub(r'[-×xX]', '*', str(matrix_module_size))
        if not panel_spec.lower().endswith('mm'):
            panel_spec = panel_spec + 'mm'
    else:
        panel_spec = config.get('panel_spec', '') if config else ''
    ws['G3'] = panel_spec
    ws['G3'].font = normal_font
    ws['G3'].alignment = center_align

    ws['C7'] = '배열'
    ws['C7'].font = normal_font
    ws['C7'].alignment = center_align

    ws.merge_cells('A7:B7')
    ws['A7'] = ''
    ws['A7'].font = a7_font
    ws['A7'].alignment = center_align
    ws['A7'].fill = a7_fill

    if rows is not None:
        ws['D7'] = f"{rows} 행"
    else:
        ws['D7'] = ''
    if cols is not None:
        ws['E7'] = f"{cols} 열"
    else:
        ws['E7'] = ''
    ws['D7'].alignment = center_align
    ws['E7'].alignment = center_align
    ws['D7'].font = normal_font
    ws['E7'].font = normal_font

    ws['F7'] = '어레이'
    ws['F7'].font = normal_font
    ws['F7'].alignment = center_align

    if matrix_set_count:
        ws['G7'] = matrix_set_count
    else:
        ws['G7'] = ''
    ws['G7'].font = normal_font
    ws['G7'].alignment = center_align

    ws.merge_cells('H3:J7')
    ws['H3'] = '품질보증: 10년\n사용수명: 25년'
    ws['H3'].font = normal_font
    ws['H3'].alignment = center_align

    # ========== 合并 A8:J8 单元格 ==========
    ws.merge_cells('A8:J8')
    currency_label = 'RMB' if sale_type == 'domestic' else 'US$'

    for col in range(1, 11):
        cell = ws.cell(row=8, column=col)
        cell.alignment = center_align
        cell.font = normal_font

    headers = [
        ('번호\nItem No.', 'A'), ('상품명\nProduct Name', 'B'), ('제조 재료\nMaterial', 'C'),
        ('사진\nPicture', 'D'), ('규격\nSpec.', 'E'), (f'단가\nUnit Price ({currency_label})\nEX Works', 'F'),
        (f'수량\nQTY (PCS)', 'G'), (f'총가격\nTotal Price ({currency_label})\nEX Works', 'H'),
        ('무게\nWeight(KG)', 'I'), ('제품코드\nRemark', 'J')
    ]
    for text, col in headers:
        cell = ws[f'{col}9']
        cell.value = text
        cell.font = header_font
        cell.alignment = center_align

    # ========== 填入所有产品数据，并匹配价格（单位从价格表获取）==========
    total_weight_sum = Decimal('0')
    total_weight_actual_sum = Decimal('0')
    total_price_sum = Decimal('0')
    matched_count = 0
    unmatched_count = 0
    unmatched_codes = []
    meter_unit_count = 0  # 统计按米计价的产品数量
    piece_unit_count = 0  # 统计按个计价的产品数量
    length_extract_fail_count = 0  # 统计长度提取失败的数量
    image_found_count = 0  # 统计找到图片的数量
    image_not_found_count = 0  # 统计未找到图片的数量
    # ========== 预扫描：判断简单模式 vs 复杂模式 ==========
    is_complex = False
    for p in all_products:
        if p.get('_is_pile'):
            continue
        pi = p.get('_price_info') or resolve_price_info(price_mapping, p.get('code', ''), spec=p.get('spec', ''))
        if not pi or not has_valid_price_info(pi):
            is_complex = True
            break
        cat = _get_discount_category(pi, p)
        if cat != 'standard':
            is_complex = True
            break

    standard_rows = []
    steel_rows = []
    purchased_rows = []
    inquiry_rows = []
    unmatched_rows = []
    standard_price_sum = Decimal('0')
    steel_price_sum = Decimal('0')
    purchased_price_sum = Decimal('0')
    inquiry_price_sum = Decimal('0')

    # 收集未匹配的产品信息（用于生成询价表）
    local_unmatched_products = []
    row_product_map = {}
    row_price_info_map = {}
    row_issue_map = {}
    weight_log_buffer = []

    def register_issue(row_number, product_data, price_data, reason):
        if not reason:
            return

        issue_item = row_issue_map.get(row_number)
        if issue_item is None:
            _ri_name_info = price_data or product_data.get('_price_info_no_spec') or resolve_price_info(price_mapping, product_data.get('code', ''))
            display_name = (
                (_ri_name_info or {}).get(name_field)
                or (_ri_name_info or {}).get('name_ko')
                or (_ri_name_info or {}).get('name')
                or product_data.get('name', '')
            )

            issue_item = {
                'code': product_data.get('code', ''),
                'name': (_ri_name_info.get('name') if _ri_name_info else None) or product_data.get('name', ''),
                'spec': product_data.get('spec', ''),
                'quantity': product_data.get('quantity', 0),
                'unit': price_data.get('unit', '') if price_data else '',
                'preinstall': normalize_preinstall(product_data.get('preinstall')),
                'reasons': [],
                'needs_inquiry': False,
            }
            row_issue_map[row_number] = issue_item

        if reason not in issue_item['reasons']:
            issue_item['reasons'].append(reason)

        if price_data is None or not has_valid_price_info(price_data):
            issue_item['needs_inquiry'] = True

    name_field = {
        '英语组': 'name_en',
        '日语组': 'name_ja',
    }.get(group or '', 'name_ko')

    for idx, product in enumerate(all_products):
        row = data_start_row + idx
        product_code = product['code']
        price_info = product.get('_price_info') or resolve_price_info(price_mapping, product_code, spec=product.get('spec', ''))
        if not product.get('_price_info') and price_info:
            product['_price_info'] = price_info
        _name_info = price_info or product.get('_price_info_no_spec') or resolve_price_info(price_mapping, product_code)
        if not product.get('_price_info_no_spec') and _name_info:
            product['_price_info_no_spec'] = _name_info
        display_name = (
            _name_info.get(name_field)
            or _name_info.get('name_ko')
            or _name_info.get('name')
            or product['name']
        ) if _name_info else product['name']

        row_product_map[row] = product
        row_price_info_map[row] = price_info

        ws.cell(row=row, column=1, value='=ROW()-9')
        ws.cell(row=row, column=2, value=display_name)
        _raw_mat = (price_info.get('db_material') if price_info and price_info.get('db_material') else None) or product['material']
        ws.cell(row=row, column=3, value=adjust_material_by_coating(translate_material(_raw_mat, 'ko'), coating_thickness))
        ws.cell(row=row, column=4, value="")
        try:
            _spec_num = float(product['spec'])
            ws.cell(row=row, column=5, value=_spec_num)
            ws.cell(row=row, column=5).number_format = '"L"!=#"mm"'
        except (ValueError, TypeError):
            ws.cell(row=row, column=5, value=_strip_cjk_spec(product['spec']))

        unit_price = 0
        display_unit_price = 0
        price_unit = ''
        is_matched = False

        if product.get('_is_pile'):
            unit_price = 0
            display_unit_price = 0
            is_matched = True
            matched_count += 1
            ws.cell(row=row, column=6, value=0)
            ws.cell(row=row, column=6).number_format = currency_number_format

            quantity = product['quantity']
            if quantity > 0:
                ws.cell(row=row, column=7, value=int(quantity) if quantity % 1 == 0 else quantity)
            else:
                ws.cell(row=row, column=7, value="")

            ws.cell(row=row, column=8, value=0)
            ws.cell(row=row, column=8).number_format = currency_number_format
            ws.cell(row=row, column=10, value=product['code'])

            weight_cell = ws.cell(row=row, column=9)
            unit_weight = calculate_report_unit_weight(product, price_info)
            if unit_weight is not None and unit_weight > 0:
                weight_cell.value = float(round_to_2_decimal(unit_weight))
                weight_cell.number_format = '#,##0.00'
            else:
                weight_cell.value = ""
            for col in [1, 2, 3, 4, 5, 6, 7, 8, 10]:
                cell = ws.cell(row=row, column=col)
                cell.alignment = center_align
                cell.font = normal_font
            weight_cell.font = normal_font
            continue

        if price_info and has_valid_price_info(price_info):
            unit_price = get_temp_base_price(price_info, product, group or '韩语组', sale_type)
            price_unit = price_info.get('unit', '')
            matched_count += 1
            is_matched = True
        else:
            unmatched_count += 1
            if product_code:
                unmatched_codes.append(product_code)
            unit_price = 0
            price_unit = ''
            is_matched = False
            # 收集未匹配的产品信息
            if price_info:
                register_issue(row, product, price_info, price_info.get('issue_reason') or '数据库缺少价格')
                price_unit = price_info.get('unit', '')
            else:
                register_issue(row, product, None, '数据库无匹配')

        # 设置单价（第6列）
        if price_info and price_info.get('image_status') == 'missing':
            register_issue(row, product, price_info, '数据库缺少图片')
        elif price_info and price_info.get('image_status') == 'invalid':
            register_issue(row, product, price_info, '图片数据无效')

        is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
        length_mm = Decimal('0')
        if is_meter:
            length_mm = Decimal(str(extract_length_from_spec(product['spec']) or 0))

        if is_meter and length_mm > 0:
            display_unit_price = float(Decimal(str(unit_price)) * length_mm / Decimal('1000'))
        else:
            display_unit_price = unit_price
        display_unit_price = apply_temp_preinstall_adjustment(price_info, display_unit_price, product, group or '韩语组', sale_type)

        if display_unit_price > 0:
            if ko_discount_in_detail and is_matched and is_complex:
                cat = _get_discount_category(price_info, product)
                rate = _get_discount_rate(cat, ko_discount_rate, ko_steel_discount_rate, ko_purchased_discount_rate)
                factor = rate / 100.0
                ws.cell(row=row, column=6, value=f"={display_unit_price}*{factor}")
            else:
                ws.cell(row=row, column=6, value=float(display_unit_price))
        else:
            ws.cell(row=row, column=6, value="")
        ws.cell(row=row, column=6).number_format = currency_number_format

        # 数量（第7列）- 单基数量
        quantity = product['quantity']
        if quantity > 0:
            ws.cell(row=row, column=7,
                    value=int(quantity) if quantity % 1 == 0 else quantity)
        else:
            ws.cell(row=row, column=7, value="")

        # ========== 根据价格表的单位计算总价 ==========
        total_price = Decimal('0')

        if display_unit_price > 0 and quantity > 0:
            if is_meter and length_mm > 0:
                meter_unit_count += 1
            else:
                piece_unit_count += 1

            total_price = Decimal(str(display_unit_price)) * Decimal(str(quantity))
            if ko_discount_in_detail and is_matched and is_complex:
                cat = _get_discount_category(price_info, product)
                rate = _get_discount_rate(cat, ko_discount_rate, ko_steel_discount_rate, ko_purchased_discount_rate)
                total_price = total_price * Decimal(str(rate)) / Decimal('100')
            ws.cell(row=row, column=8, value=f"=F{row}*G{row}")
            ws.cell(row=row, column=8).number_format = currency_number_format
            total_price_sum += total_price
            if _is_standard_priced(price_info):
                cat = _get_discount_category(price_info)
                if cat == 'standard':
                    standard_rows.append(row)
                    standard_price_sum += total_price
                elif cat == 'steel':
                    steel_rows.append(row)
                    steel_price_sum += total_price
                elif cat == 'purchased':
                    purchased_rows.append(row)
                    purchased_price_sum += total_price
                else:
                    inquiry_rows.append(row)
                    inquiry_price_sum += total_price
        else:
            ws.cell(row=row, column=8, value=f"=F{row}*G{row}")
            ws.cell(row=row, column=8).number_format = currency_number_format
            if not is_matched:
                fallback_cat = _get_discount_category(None, product)
                if fallback_cat == 'steel':
                    steel_rows.append(row)
                elif fallback_cat == 'purchased':
                    purchased_rows.append(row)
                else:
                    unmatched_rows.append(row)

        # 重量（第9列）
        weight_cell = ws.cell(row=row, column=9)
        unit_weight = calculate_report_unit_weight(product, price_info)

        if unit_weight is not None and unit_weight > 0:
            weight_cell.value = float(round_to_2_decimal(unit_weight))
            weight_cell.number_format = '#,##0.00'
            weight_cell.alignment = center_align

            total_weight_sum += unit_weight
            total_weight_actual_sum += unit_weight * Decimal(str(quantity))
        else:
            weight_cell.value = ""
            weight_cell.alignment = center_align

        # 产品编码（第10列）
        ws.cell(row=row, column=10, value=product['code'])

        # 设置对齐和字体
        for col in [1, 2, 3, 4, 5, 6, 7, 8, 10]:
            cell = ws.cell(row=row, column=col)
            cell.alignment = center_align
            cell.font = normal_font
        weight_cell.font = normal_font

        # ========== 为未匹配价格的行添加黄色背景 ==========
        if not is_matched:
            for col in range(1, 11):  # A-J列
                cell = ws.cell(row=row, column=col)
                cell.fill = yellow_fill

    # 图片检查完成后再统一汇总问题项

    if weight_log_buffer:
        print(f"[WEIGHT] sheet={sheet_name} rows={len(weight_log_buffer)}")
        print('\n'.join(weight_log_buffer))

    # 打印价格匹配统计
    print(f"   💰 价格匹配统计: 成功 {matched_count} 项, 失败 {unmatched_count} 项")
    print(f"   📏 计价单位统计: 米计价 {meter_unit_count} 项, 个/套计价 {piece_unit_count} 项")
    if length_extract_fail_count > 0:
        print(f"   ⚠️ 长度提取失败 {length_extract_fail_count} 项 (已按个计价)")
    if unmatched_codes:
        print(f"   ⚠️ 未匹配的编码: {unmatched_codes[:10]}{'...' if len(unmatched_codes) > 10 else ''}")

    last_data_row = data_start_row + total_products - 1 if total_products > 0 else data_start_row - 1

    # ========== 删除行后再插入图片，避免错位 ==========
    def mark_missing_image(row_number):
        image_cell = ws.cell(row=row_number, column=IMAGE_COLUMN_INDEX)
        image_cell.value = "/"
        image_cell.alignment = center_align
        image_cell.font = normal_font

    for row in range(data_start_row, last_data_row + 1):
        product_code = ws.cell(row=row, column=10).value
        if product_code is None or str(product_code).strip() == '':
            continue
        product_code = str(product_code).strip()
        normalized_code = normalize_lookup_code(product_code)
        price_info_for_row = row_price_info_map.get(row)

        img_path = None
        if image_folder and code_to_images and product_code in code_to_images and code_to_images[product_code]:
            img_path = code_to_images[product_code][0]
        elif image_folder and code_to_images and normalized_code in code_to_images and code_to_images[normalized_code]:
            img_path = code_to_images[normalized_code][0]

        if img_path:
            fit_w, fit_h = _fit_image_to_cell(row, IMAGE_COLUMN_INDEX, IMAGE_WIDTH, IMAGE_HEIGHT)
            temp_img_path = prepare_image_for_excel(
                img_path,
                target_width=fit_w,
                target_height=fit_h,
                temp_dir=image_temp_dir,
                cache=image_cache
            )
            final_img_path = temp_img_path if temp_img_path else img_path
            success = add_image_centered_in_cell(
                ws,
                final_img_path,
                row,
                IMAGE_COLUMN_INDEX,
                img_width=fit_w,
                img_height=fit_h
            )
            if success:
                image_found_count += 1
                continue

            image_not_found_count += 1
            mark_missing_image(row)
            if price_info_for_row and price_info_for_row.get('image_status') == 'ready':
                register_issue(row, row_product_map.get(row, {}), price_info_for_row, '数据库图片未落盘')
            continue

        image_not_found_count += 1
        mark_missing_image(row)
        if price_info_for_row and price_info_for_row.get('image_status') == 'ready':
            register_issue(row, row_product_map.get(row, {}), price_info_for_row, '数据库图片未落盘')
        if image_not_found_count <= 5:
            print(f"   ⚠️ 未找到图片: {product_code}")

    print(f"   🖼️ 图片统计: 成功插入 {image_found_count} 张, 未找到 {image_not_found_count} 张")

    local_unmatched_products = []
    for issue_item in row_issue_map.values():
        if not _is_valid_product_code(issue_item.get('code', '')):
            continue
        reasons = issue_item.pop('reasons', [])
        needs_inquiry = bool(issue_item.pop('needs_inquiry', False))
        issue_item['issue_reason'] = '铝价数据库无法匹配'
        if needs_inquiry:
            local_unmatched_products.append(issue_item)

    for _fwp, _fwp_pi in _empty_weight_products:
        _fwp_code = _fwp.get('code', '')
        if not _is_valid_product_code(_fwp_code):
            continue
        if _fwp_pi is not None and has_valid_price_info(_fwp_pi):
            continue
        _fwp_ni = _fwp_pi or _fwp.get('_price_info_no_spec') or resolve_price_info(price_mapping, _fwp_code)
        _fwp_name = (
            (_fwp_ni.get(_pre_name_field) or _fwp_ni.get('name_ko')
             or _fwp_ni.get('name') or _fwp.get('name', ''))
        ) if _fwp_ni else _fwp.get('name', '')
        local_unmatched_products.append({
            'code': _fwp_code,
            'name': (_fwp_ni.get('name') if _fwp_ni else None) or _fwp.get('name', ''),
            'spec': _fwp.get('spec', ''),
            'quantity': _fwp.get('quantity', 0),
            'unit': _fwp_pi.get('unit', '') if _fwp_pi else '',
            'weight': _fwp.get('weight', 0),
            'preinstall': normalize_preinstall(_fwp.get('preinstall')),
            'issue_reason': '铝价数据库无法匹配',
        })

    if unmatched_products_list is not None:
        for _up in local_unmatched_products:
            _up['quantity'] = float(_up.get('quantity', 0)) * total_table_count
        unmatched_products_list.extend(local_unmatched_products)

    valid_product_count = total_products

    total_weight_sum_rounded = round_to_2_decimal(total_weight_sum)
    total_weight_actual_rounded = round_to_2_decimal(total_weight_actual_sum)
    total_price_sum_rounded = round_to_2_decimal(total_price_sum)
    total_row_price_sum_rounded = round_to_2_decimal(total_price_sum_rounded * Decimal(str(total_table_count)))
    total_row_weight_sum_rounded = round_to_2_decimal(total_weight_actual_sum * Decimal(str(total_table_count)))
    total_amount_label = f'TOTAL AMOUNT OF {total_table_count} TABLES'

    multiplier = Decimal('1')

    # ========== 计算小计和总计行的位置 ==========
    new_data_end_row = last_data_row
    subtotal_row_1 = new_data_end_row + 1
    total_row_1 = subtotal_row_1 + 1
    note_row = total_row_1 + 1

    ws.row_dimensions[subtotal_row_1].height = 40
    ws.row_dimensions[total_row_1].height = 40
    ws.row_dimensions[note_row].height = 28

    # ========== 第一个小计和总计 ==========
    ws.merge_cells(f'A{subtotal_row_1}:G{subtotal_row_1}')
    subtotal_a = ws[f'A{subtotal_row_1}']
    subtotal_a.value = 'SUB-TOTAL AMOUNT/TABLE'
    subtotal_a.font = small_bold_font
    subtotal_a.alignment = right_align

    subtotal_h = ws.cell(row=subtotal_row_1, column=8)
    if total_price_sum > 0:
        ws.cell(row=subtotal_row_1, column=8, value=f'=SUM(H{data_start_row}:H{new_data_end_row})')
        subtotal_h.number_format = currency_number_format
    else:
        subtotal_h.value = ""
    subtotal_h.font = small_bold_font
    subtotal_h.alignment = center_align

    subtotal_i = ws.cell(row=subtotal_row_1, column=9)
    if total_weight_sum_rounded > 0:
        ws.cell(row=subtotal_row_1, column=9, value=f'=SUMPRODUCT(I{data_start_row}:I{new_data_end_row},G{data_start_row}:G{new_data_end_row})')
        subtotal_i.number_format = numbers.FORMAT_NUMBER_00
    else:
        subtotal_i.value = ""
    subtotal_i.font = small_bold_font
    subtotal_i.alignment = center_align

    subtotal_j = ws.cell(row=subtotal_row_1, column=10)
    subtotal_j.value = 'KG'
    subtotal_j.font = small_bold_font
    subtotal_j.alignment = center_align

    ws.merge_cells(f'A{total_row_1}:G{total_row_1}')
    total_a = ws[f'A{total_row_1}']
    total_a.value = total_amount_label
    total_a.font = small_bold_font
    total_a.alignment = right_align

    total_h = ws.cell(row=total_row_1, column=8)
    if total_row_price_sum_rounded > 0:
        if total_table_count > 1:
            ws.cell(row=total_row_1, column=8, value=f'=H{subtotal_row_1}*{total_table_count}')
        else:
            ws.cell(row=total_row_1, column=8, value=f'=H{subtotal_row_1}')
        total_h.number_format = currency_number_format
    else:
        total_h.value = ""
    total_h.font = small_bold_font
    total_h.alignment = center_align

    total_i = ws.cell(row=total_row_1, column=9)
    if total_row_weight_sum_rounded > 0:
        if total_table_count > 1:
            ws.cell(row=total_row_1, column=9, value=f'=I{subtotal_row_1}*{total_table_count}')
        else:
            ws.cell(row=total_row_1, column=9, value=f'=I{subtotal_row_1}')
        total_i.number_format = numbers.FORMAT_NUMBER_00
    else:
        total_i.value = ""
    total_i.font = small_bold_font
    total_i.alignment = center_align

    total_j = ws.cell(row=total_row_1, column=10)
    total_j.value = 'KG'
    total_j.font = small_bold_font
    total_j.alignment = center_align

    # ========== 附加说明行 ==========
    ws.merge_cells(f'A{note_row}:J{note_row}')
    ws[f'A{note_row}'] = '아래 제품은 견적서에 미포함입니다. 필요시 추가 주문 가능합니다.'
    ws[f'A{note_row}'].font = red_small_font
    ws[f'A{note_row}'].alignment = left_align

    # ========== 附加产品数据行（排除的导电片/接地铜线夹） ==========
    excluded_price_sum = Decimal('0')
    excluded_weight_sum = Decimal('0')
    excluded_weight_actual_sum = Decimal('0')

    if excluded_products:
        excluded_products = [p for p in excluded_products if p.get('quantity') and p.get('quantity', 0) > 0]
    if excluded_products:
        excluded_start = note_row + 1
        for ei, eproduct in enumerate(excluded_products):
            e_row = excluded_start + ei
            ws.row_dimensions[e_row].height = 61

            e_code = eproduct.get('code', '')
            e_price_info = eproduct.get('_price_info') or resolve_price_info(price_mapping, e_code, spec=eproduct.get('spec', ''))
            if not eproduct.get('_price_info') and e_price_info:
                eproduct['_price_info'] = e_price_info
            _e_name_info = e_price_info or eproduct.get('_price_info_no_spec') or resolve_price_info(price_mapping, e_code)
            if not eproduct.get('_price_info_no_spec') and _e_name_info:
                eproduct['_price_info_no_spec'] = _e_name_info
            e_display_name = (
                (_e_name_info.get(name_field) or _e_name_info.get('name_ko')
                 or _e_name_info.get('name') or eproduct.get('name', ''))
            ) if _e_name_info else eproduct.get('name', '')

            e_unit_price = 0
            e_display_unit_price = 0
            e_price_unit = ''
            e_is_matched = False

            if e_price_info and has_valid_price_info(e_price_info):
                e_unit_price = get_temp_base_price(e_price_info, eproduct, group or '韩语组', sale_type)
                e_price_unit = e_price_info.get('unit', '')
                e_is_matched = True

            e_seq = ei + 1
            e_qty = eproduct.get('quantity', 0)

            ws.cell(row=e_row, column=1, value=e_seq)
            ws.cell(row=e_row, column=2, value=e_display_name)
            _e_raw_mat = (e_price_info.get('db_material') if e_price_info and e_price_info.get('db_material') else None) or eproduct.get('material', '')
            ws.cell(row=e_row, column=3, value=adjust_material_by_coating(translate_material(_e_raw_mat, 'ko'), coating_thickness))
            ws.cell(row=e_row, column=4, value="")
            try:
                _spec_num = float(eproduct.get('spec', ''))
                ws.cell(row=e_row, column=5, value=_spec_num)
                ws.cell(row=e_row, column=5).number_format = '"L"!=#"mm"'
            except (ValueError, TypeError):
                ws.cell(row=e_row, column=5, value=_strip_cjk_spec(eproduct.get('spec', '')))

            e_is_meter = e_price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
            e_length_mm = Decimal('0')
            if e_is_meter:
                e_length_mm = Decimal(str(extract_length_from_spec(eproduct.get('spec', '')) or 0))

            if e_is_meter and e_length_mm > 0:
                e_display_unit_price = float(Decimal(str(e_unit_price)) * e_length_mm / Decimal('1000'))
            else:
                e_display_unit_price = e_unit_price
            e_display_unit_price = apply_temp_preinstall_adjustment(e_price_info, e_display_unit_price, eproduct, group or '韩语组', sale_type)

            if e_display_unit_price > 0:
                ws.cell(row=e_row, column=6, value=float(e_display_unit_price))
            else:
                ws.cell(row=e_row, column=6, value="")
            ws.cell(row=e_row, column=6).number_format = currency_number_format

            if e_qty > 0:
                ws.cell(row=e_row, column=7, value=int(e_qty) if e_qty % 1 == 0 else e_qty)
            else:
                ws.cell(row=e_row, column=7, value="")

            e_total_price = Decimal('0')
            if e_display_unit_price > 0 and e_qty > 0:
                e_total_price = Decimal(str(e_display_unit_price)) * Decimal(str(e_qty))
                ws.cell(row=e_row, column=8, value=f"=F{e_row}*G{e_row}")
                excluded_price_sum += e_total_price
            else:
                ws.cell(row=e_row, column=8, value=f"=F{e_row}*G{e_row}")
            ws.cell(row=e_row, column=8).number_format = currency_number_format

            e_unit_weight = calculate_report_unit_weight(eproduct, e_price_info)
            e_weight_cell = ws.cell(row=e_row, column=9)
            if e_unit_weight is not None and e_unit_weight > 0:
                e_weight_cell.value = float(round_to_2_decimal(e_unit_weight))
                e_weight_cell.number_format = '#,##0.00'
                e_weight_cell.alignment = center_align
                excluded_weight_sum += e_unit_weight
                excluded_weight_actual_sum += e_unit_weight * Decimal(str(e_qty))
            else:
                e_weight_cell.value = ""
                e_weight_cell.alignment = center_align

            ws.cell(row=e_row, column=10, value=e_code)

            for ecol in [1, 2, 3, 4, 5, 6, 7, 8, 10]:
                c = ws.cell(row=e_row, column=ecol)
                c.alignment = center_align
                c.font = normal_font
            e_weight_cell.font = normal_font

            if not e_is_matched:
                for ecol in range(1, 11):
                    ws.cell(row=e_row, column=ecol).fill = yellow_fill

            if (e_price_info is None or not has_valid_price_info(e_price_info)) and _is_valid_product_code(e_code):
                local_unmatched_products.append({
                    'code': e_code,
                    'name': (_e_name_info.get('name') if _e_name_info else None) or eproduct.get('name', ''),
                    'spec': eproduct.get('spec', ''),
                    'quantity': e_qty,
                    'unit': e_price_info.get('unit', '') if e_price_info else '',
                    'weight': eproduct.get('weight', 0),
                    'preinstall': normalize_preinstall(eproduct.get('preinstall')),
                    'issue_reason': '铝价数据库无法匹配',
                })

        excluded_end = excluded_start + len(excluded_products) - 1

        for e_row in range(excluded_start, excluded_end + 1):
            e_code = str(ws.cell(row=e_row, column=10).value or '').strip()
            if not e_code:
                continue
            e_norm_code = normalize_lookup_code(e_code)
            e_img_path = None
            if image_folder and code_to_images and e_code in code_to_images and code_to_images[e_code]:
                e_img_path = code_to_images[e_code][0]
            elif image_folder and code_to_images and e_norm_code in code_to_images and code_to_images[e_norm_code]:
                e_img_path = code_to_images[e_norm_code][0]

            if e_img_path:
                e_fit_w, e_fit_h = _fit_image_to_cell(e_row, IMAGE_COLUMN_INDEX, IMAGE_WIDTH, IMAGE_HEIGHT)
                e_temp = prepare_image_for_excel(
                    e_img_path, target_width=e_fit_w, target_height=e_fit_h,
                    temp_dir=image_temp_dir, cache=None,
                )
                e_final = e_temp if e_temp else e_img_path
                e_ok = add_image_centered_in_cell(
                    ws, e_final, e_row, IMAGE_COLUMN_INDEX,
                    img_width=e_fit_w, img_height=e_fit_h,
                )
                if e_ok:
                    image_found_count += 1
                    continue
                image_not_found_count += 1
                ws.cell(row=e_row, column=IMAGE_COLUMN_INDEX).value = "/"
                ws.cell(row=e_row, column=IMAGE_COLUMN_INDEX).alignment = center_align
                ws.cell(row=e_row, column=IMAGE_COLUMN_INDEX).font = normal_font
            else:
                image_not_found_count += 1
                ws.cell(row=e_row, column=IMAGE_COLUMN_INDEX).value = "/"
                ws.cell(row=e_row, column=IMAGE_COLUMN_INDEX).alignment = center_align
                ws.cell(row=e_row, column=IMAGE_COLUMN_INDEX).font = normal_font

        extra_subtotal_row = excluded_end + 1
        extra_total_row = excluded_end + 2

        ws.row_dimensions[extra_subtotal_row].height = 40
        ws.row_dimensions[extra_total_row].height = 40

        ws.merge_cells(f'A{extra_subtotal_row}:G{extra_subtotal_row}')
        ex_sub_a = ws[f'A{extra_subtotal_row}']
        ex_sub_a.value = 'SUB-TOTAL AMOUNT/TABLE'
        ex_sub_a.font = small_bold_font
        ex_sub_a.alignment = right_align

        ex_sub_h = ws.cell(row=extra_subtotal_row, column=8)
        ex_sub_i = ws.cell(row=extra_subtotal_row, column=9)
        ws.cell(row=extra_subtotal_row, column=8, value=f'=SUM(H{excluded_start}:H{excluded_end})')
        ex_sub_h.number_format = currency_number_format
        ex_weight_rounded = round_to_2_decimal(excluded_weight_sum)
        ws.cell(row=extra_subtotal_row, column=9, value=f'=SUMPRODUCT(I{excluded_start}:I{excluded_end},G{excluded_start}:G{excluded_end})')
        if ex_weight_rounded > 0:
            ex_sub_i.number_format = numbers.FORMAT_NUMBER_00
        ex_sub_h.font = small_bold_font
        ex_sub_h.alignment = center_align
        ex_sub_i.font = small_bold_font
        ex_sub_i.alignment = center_align

        ex_sub_j = ws.cell(row=extra_subtotal_row, column=10)
        ex_sub_j.value = 'KG'
        ex_sub_j.font = small_bold_font
        ex_sub_j.alignment = center_align

        ws.merge_cells(f'A{extra_total_row}:G{extra_total_row}')
        ex_tot_a = ws[f'A{extra_total_row}']
        ex_tot_a.value = total_amount_label
        ex_tot_a.font = small_bold_font
        ex_tot_a.alignment = right_align

        ex_tot_h = ws.cell(row=extra_total_row, column=8)
        ex_tot_i = ws.cell(row=extra_total_row, column=9)
        if total_table_count > 1:
            ws.cell(row=extra_total_row, column=8, value=f'=H{extra_subtotal_row}*{total_table_count}')
        else:
            ws.cell(row=extra_total_row, column=8, value=f'=H{extra_subtotal_row}')
        ex_tot_h.number_format = currency_number_format
        ex_total_weight = round_to_2_decimal(excluded_weight_sum)
        if ex_total_weight > 0:
            if total_table_count > 1:
                ws.cell(row=extra_total_row, column=9, value=f'=I{extra_subtotal_row}*{total_table_count}')
            else:
                ws.cell(row=extra_total_row, column=9, value=f'=I{extra_subtotal_row}')
            ex_tot_i.number_format = numbers.FORMAT_NUMBER_00
        ex_tot_h.font = small_bold_font
        ex_tot_h.alignment = center_align
        ex_tot_i.font = small_bold_font
        ex_tot_i.alignment = center_align

        ex_tot_j = ws.cell(row=extra_total_row, column=10)
        ex_tot_j.value = 'KG'
        ex_tot_j.font = small_bold_font
        ex_tot_j.alignment = center_align
    else:
        extra_total_row = note_row

    # ========== 统一设置边框（单遍） ==========
    top_border = _KO_BORDER_TOP
    bottom_border = _KO_BORDER_BOTTOM
    left_border = _KO_BORDER_LEFT
    right_border = _KO_BORDER_RIGHT
    top_left_border = _KO_BORDER_TL
    top_right_border = _KO_BORDER_TR
    bottom_left_border = _KO_BORDER_BL
    bottom_right_border = _KO_BORDER_BR

    for row in range(2, extra_total_row + 1):
        for col in range(1, 11):
            is_top = (row == 2)
            is_bottom = (row == extra_total_row)
            is_left = (col == 1)
            is_right = (col == 10)
            if is_top and is_left:
                b = top_left_border
            elif is_top and is_right:
                b = top_right_border
            elif is_bottom and is_left:
                b = bottom_left_border
            elif is_bottom and is_right:
                b = bottom_right_border
            elif is_top:
                b = top_border
            elif is_bottom:
                b = bottom_border
            elif is_left:
                b = left_border
            elif is_right:
                b = right_border
            else:
                b = thin_border
            ws.cell(row=row, column=col).border = b

    ws['A7'].border = thin_border
    ws['B7'].border = thin_border
    ws['D7'].border = thin_border
    ws['E7'].border = thin_border
    ws['G6'].border = thin_border

    for col in range(1, 11):
        ws.cell(row=1, column=col).border = no_border

    # ========== 页面设置 ==========
    apply_print_setup(ws, 'ko_normal')
    ws.sheet_view.zoomScale = 50

    quotation_product_codes = set()
    for p in all_products:
        c = str(p.get('code', '')).strip()
        if c:
            quotation_product_codes.add(c)
    for p in excluded_products:
        c = str(p.get('code', '')).strip()
        if c:
            quotation_product_codes.add(c)

    return {
        'sheet_name': sheet_name,
        'quotation_product_codes': quotation_product_codes,
        'valid_products': valid_product_count,
        'total_weight': float(total_row_weight_sum_rounded),
        'total_price': float(total_row_price_sum_rounded),
        'matched_count': matched_count,
        'unmatched_count': unmatched_count,
        'meter_unit_count': meter_unit_count,
        'piece_unit_count': piece_unit_count,
        'length_extract_fail_count': length_extract_fail_count,
        'unmatched_products_count': len(local_unmatched_products),
        'image_found_count': image_found_count,
        'image_not_found_count': image_not_found_count,
        'pile_products': pile_products,
        'all_render_products': aluminum + carbon_steel + excluded_products,
        'sub_total_row': subtotal_row_1,
        'total_row': total_row_1,
        'detail_data_end_row': new_data_end_row,
        'excluded_sub_total_row': extra_total_row,
        'standard_rows': standard_rows,
        'steel_rows': steel_rows,
        'purchased_rows': purchased_rows,
        'inquiry_rows': inquiry_rows,
        'unmatched_rows': unmatched_rows,
        'is_complex': is_complex,
        'standard_total': float(standard_price_sum),
        'steel_total': float(steel_price_sum),
        'purchased_total': float(purchased_price_sum),
        'inquiry_total': float(inquiry_price_sum),
    }


def create_summary_quotation_sheet(workbook, all_quotation_results, matrix_data=None,
                                   image_path=None, contact_info=None, pile_summary=None,
                                   trade_method='CIF', dest_port='부산',
                                   container_type='40HQ', container_qty=1,
                                   ko_case_type='NORMAL', ko_discount_rate=100,
                                   ko_steel_discount_rate=84,
                                   ko_purchased_discount_rate=94,
                                    ko_tariff_rate=1.6, ko_consumption_tax=10,
                                   sale_type='export', ko_freight=0,
                                   ko_cif_freight=0, ko_ddp_address='',
                                   ko_discount_in_detail=True):
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from datetime import datetime

    ws = workbook.create_sheet(title='견적서')

    malgun = _KO_SUMMARY_MALGUN
    malgun_bold = _KO_SUMMARY_MALGUN_BOLD
    malgun_bold_16 = _KO_SUMMARY_MALGUN_BOLD_16
    malgun_small = _KO_SUMMARY_MALGUN_SMALL
    malgun_small_red = _KO_SUMMARY_MALGUN_SMALL_RED
    summary_currency = '#,##0.00'
    summary_currency_label = 'RMB' if sale_type == 'domestic' else 'USD'

    red_inline = InlineFont(rFont='Malgun Gothic', color='FF0000')
    normal_inline = InlineFont(rFont='Malgun Gothic')

    light_blue = _KO_SUMMARY_LIGHT_BLUE
    center = _KO_CENTER_ALIGN
    left = _KO_LEFT_ALIGN
    right_a = _KO_RIGHT_ALIGN

    thin_side = _KO_THIN
    thick_side = _KO_THICK
    thin_border = _KO_THIN_BORDER

    col_widths = {'A': 5, 'B': 5, 'C': 8, 'D': 7, 'E': 5, 'F': 6,
                  'G': 15.5, 'H': 13, 'I': 7, 'J': 7, 'K': 9.1, 'L': 9.1}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    matrix_data = matrix_data or {}
    project_name = str(matrix_data.get('project_name') or '').strip()
    output_kw = matrix_data.get('output_kw') or 0

    # ========== Row 1 ==========
    ws.merge_cells('A1:G1')
    ws.row_dimensions[1].height = 70
    if image_path and os.path.exists(image_path):
        try:
            img = XLImage(img=image_path)
            original_width = img.width
            original_height = img.height

            merged_width = sum(col_widths.get(get_column_letter(c), 0) for c in range(1, 8))
            cell_width_px = merged_width * 7.5
            row1_height = ws.row_dimensions[1].height or 70
            cell_height_px = row1_height * 1.33

            target_scale = 0.6
            target_w = int(original_width * target_scale)
            target_h = int(original_height * target_scale)

            margin = 10
            max_w = cell_width_px - margin * 2
            max_h = cell_height_px - margin * 2
            if target_w > max_w or target_h > max_h:
                scale_w = max_w / original_width if original_width > 0 else 1
                scale_h = max_h / original_height if original_height > 0 else 1
                fit_scale = min(scale_w, scale_h)
                target_w = int(original_width * fit_scale)
                target_h = int(original_height * fit_scale)

            offset_x = max(0, (cell_width_px - target_w) / 2)
            offset_y = max(0, (cell_height_px - target_h) / 2)

            EMU_PER_PIXEL = 9525
            marker = AnchorMarker(
                col=0, colOff=int(offset_x * EMU_PER_PIXEL),
                row=0, rowOff=int(offset_y * EMU_PER_PIXEL),
            )
            img.anchor = OneCellAnchor(
                _from=marker,
                ext=XDRPositiveSize2D(
                    cx=int(target_w * EMU_PER_PIXEL),
                    cy=int(target_h * EMU_PER_PIXEL),
                ),
            )
            ws.add_image(img)
        except Exception:
            ws['A1'] = 'KSENG'
            ws['A1'].font = malgun_bold_16
            ws['A1'].alignment = center
    else:
        ws['A1'] = 'KSENG'
        ws['A1'].font = malgun_bold_16
        ws['A1'].alignment = center

    ws.merge_cells('H1:L1')
    ws['H1'] = f'견적일：\t\t{datetime.now().strftime("%Y/%m/%d")}'
    ws['H1'].alignment = center
    ws['H1'].font = malgun

    # ========== Row 2: 标题 ==========
    ws.merge_cells('A2:L2')
    ws['A2'] = '태양광 지붕 시스템 견적서'
    ws['A2'].font = malgun_bold_16
    ws['A2'].alignment = center
    ws['A2'].fill = light_blue
    ws.row_dimensions[2].height = 30

    # ========== Row 3-7: 项目信息 ==========
    info_rows = [
        (3, '프로젝트：\nProject Name:', project_name),
        (4, '무역거래조건：\nPrice Term：', ''),
        (5, '납기일：\nDelivery time：', '계약금 수금후 10~14일뒤 출고 가능합니다 (현장 상황에 따라 다소 변경될 수도 있습니다)'),
        (6, '지불조건：\nPayment Term:', '계약서 체결후 30% 예약금 선불(T/T),나머지 70%요금 선적전 지불입니다.'),
        (7, '유효기간：\nValidity Date:', '본 견적 요금은 당일 기준으로만 유효하며, 실제 발주 시점에서 가격을 재확인해 주시기 바랍니다.')
    ]
    for row_num, label, value in info_rows:
        ws.merge_cells(f'A{row_num}:C{row_num}')
        ws[f'A{row_num}'] = label
        ws[f'A{row_num}'].alignment = left
        ws[f'A{row_num}'].font = malgun
        ws.merge_cells(f'D{row_num}:L{row_num}')
        ws[f'D{row_num}'] = value
        ws[f'D{row_num}'].alignment = left
        ws[f'D{row_num}'].font = malgun
        ws.row_dimensions[row_num].height = 35

    trade_display = trade_method
    if trade_method == 'CIF':
        trade_display = f'CIF {dest_port}'
    elif trade_method == 'FOB':
        trade_display = 'FOB'
    elif trade_method == 'EXW':
        trade_display = 'EXW'

    if trade_method == 'DDP':
        if ko_ddp_address:
            ws[f'D4'] = CellRichText(
                TextBlock(red_inline, 'DDP '),
                TextBlock(normal_inline, ko_ddp_address),
            )
        else:
            ws[f'D4'] = 'DDP'
            ws[f'D4'].font = Font(name='Malgun Gothic', color='FF0000')
    else:
        ws[f'D4'] = trade_display
        ws[f'D4'].font = Font(name='Malgun Gothic', color='FF0000')
    ws[f'D4'].alignment = left

    # ========== Row 8: Part 1 ==========
    ws.merge_cells('A8:L8')
    ws['A8'] = 'Part 1:구조물 요금(Solar Mounting System)'
    ws['A8'].font = malgun_bold
    ws['A8'].alignment = left
    ws.row_dimensions[8].height = 30

    # ========== Row 9: 表头 ==========
    header_items = [
        (1, '번호\nNo.'), (2, '행\nRow'), (3, '열\nColumn'),
        (4, '세트\nSet'),
    ]
    for col_idx, header in header_items:
        cell = ws.cell(row=9, column=col_idx, value=header)
        cell.alignment = center
        cell.border = thin_border
        cell.font = malgun

    ws.merge_cells('E9:F9')
    cell = ws.cell(row=9, column=5, value='각도\nAngle')
    cell.alignment = center
    cell.border = thin_border
    cell.font = malgun

    cell = ws.cell(row=9, column=7, value='출력량(KW)\nOutput volume')
    cell.alignment = center
    cell.border = thin_border
    cell.font = malgun

    cell = ws.cell(row=9, column=8, value=f'1세트요금\n({summary_currency_label}/Table)')
    cell.alignment = center
    cell.border = thin_border
    cell.font = malgun

    ws.merge_cells('I9:L9')
    ws['I9'] = f'합계({summary_currency_label})\nprice'
    ws['I9'].alignment = center
    ws['I9'].border = thin_border
    ws['I9'].font = malgun
    ws.row_dimensions[9].height = 50

    # ========== Row 10+: 每个BOM一行数据（累加匹配的多个sheet合并为一行） ==========
    data_start_row = 10
    total_price_all = Decimal('0')
    total_output_kw = Decimal('0')
    total_weight_all = Decimal('0')

    _arrays_for_sort = matrix_data.get('arrays') or []
    if _arrays_for_sort:
        _arr_pos = {}
        for _ai, _ae in enumerate(_arrays_for_sort):
            _k = (_ae.get('rows'), _ae.get('cols'), _ae.get('table_qty', 1), _ae.get('missing_per_table', 0) or 0)
            if _k not in _arr_pos:
                _arr_pos[_k] = _ai
        def _sort_by_array(item):
            _idx, _d = item
            _explicit = _d.get('_matrix_idx')
            if _explicit is not None:
                return (_explicit, _d.get('accumulated_sub_idx', 0), _idx)
            _ma = _d.get('matched_array') or _d.get('matrix_data') or {}
            _k = (_ma.get('rows'), _ma.get('cols'), _ma.get('table_qty', 1) or 1, _ma.get('missing_per_table', 0) or 0)
            return (_arr_pos.get(_k, len(_arrays_for_sort)), 0, _idx)
        _indexed = list(enumerate(all_quotation_results))
        _indexed.sort(key=_sort_by_array)
        all_quotation_results[:] = [_r for _, _r in _indexed]

    _summary_rows = []
    _seen_acc_groups = set()
    for qr in all_quotation_results:
        gid = qr.get('_acc_group_id')
        if gid is not None and gid not in _seen_acc_groups:
            _seen_acc_groups.add(gid)
            _group = [r for r in all_quotation_results if r.get('_acc_group_id') == gid]
            _summary_rows.append(('accumulated', _group))
        elif gid is None:
            _summary_rows.append(('single', [qr]))

    for i, (entry_type, items) in enumerate(_summary_rows):
        row_num = data_start_row + i

        if entry_type == 'single':
            qr = items[0]
            config = qr.get('config') or {}
            row_matrix_data = qr.get('matrix_data') or matrix_data
            row_matched_array = qr.get('matched_array') or {}
            angle_val = config.get('angle', '')
            row_array_rows = row_matched_array.get('rows', row_matrix_data.get('array_rows'))
            row_array_cols = row_matched_array.get('cols', row_matrix_data.get('array_cols'))
            row_set_count = row_matrix_data.get('set_count') or 1
            if not row_set_count:
                row_set_count = 1
            row_output_kw = row_matrix_data.get('output_kw')
            if row_output_kw in (None, ''):
                row_output_kw = 0

            per_set_price = Decimal(str(qr.get('total_price', 0))) / Decimal(str(row_set_count))
            total_price = Decimal(str(qr.get('total_price', 0)))
            total_price_all += total_price
            row_weight = Decimal(str(qr.get('total_weight', 0)))
            total_weight_all += row_weight
            try:
                total_output_kw += Decimal(str(row_output_kw))
            except (ArithmeticError, ValueError):
                pass

            _m_idx = qr.get('_matrix_idx')
            _no_label = f"({_m_idx + 1})" if _m_idx is not None else f'{i + 1}'
            ws.cell(row=row_num, column=1, value=_no_label).alignment = center
            ws.cell(row=row_num, column=2, value=row_array_rows or '')
            ws.cell(row=row_num, column=3, value=row_array_cols or '')
            ws.cell(row=row_num, column=4, value=row_set_count)
            ws.merge_cells(f'E{row_num}:F{row_num}')
            ws.cell(row=row_num, column=5, value=angle_val or '')
            ws.cell(row=row_num, column=7, value=row_output_kw)

            detail_sheet_name = qr.get('sheet_name', '')
            detail_sub_row = qr.get('sub_total_row')
            detail_total_row = qr.get('total_row')
            if detail_sheet_name and detail_sub_row:
                ws.cell(row=row_num, column=8, value=f"='{detail_sheet_name}'!H{detail_sub_row}")
            else:
                ws.cell(row=row_num, column=8, value=float(per_set_price))
            ws.cell(row=row_num, column=8).number_format = summary_currency

            ws.merge_cells(f'I{row_num}:L{row_num}')
            if detail_sheet_name and detail_total_row:
                ws.cell(row=row_num, column=9, value=f"='{detail_sheet_name}'!H{detail_total_row}")
            elif row_set_count and row_set_count != 1:
                ws.cell(row=row_num, column=9, value=f'=H{row_num}*{row_set_count}')
            elif detail_sheet_name and detail_sub_row:
                ws.cell(row=row_num, column=9, value=f"='{detail_sheet_name}'!H{detail_sub_row}")
            else:
                ws.cell(row=row_num, column=9, value=float(total_price))
            ws.cell(row=row_num, column=9).number_format = summary_currency
            ws.cell(row=row_num, column=9).fill = light_blue

        elif entry_type == 'accumulated':
            _first = items[0]
            _matched_array = _first.get('matched_array') or {}
            _group_total_qty = _first.get('_acc_group_total_qty', 1)
            _group_rows = _matched_array.get('rows', '')
            _group_cols = _matched_array.get('cols', '')
            _group_angle = (_first.get('config') or {}).get('angle', '')

            _group_output_kw = Decimal('0')
            _group_total_price = Decimal('0')
            _group_total_weight = Decimal('0')
            for _gr in items:
                _gmd = _gr.get('matrix_data') or {}
                _gkw = _gmd.get('output_kw')
                if _gkw not in (None, ''):
                    try:
                        _group_output_kw += Decimal(str(_gkw))
                    except (ArithmeticError, ValueError):
                        pass
                _group_total_price += Decimal(str(_gr.get('total_price', 0)))
                _group_total_weight += Decimal(str(_gr.get('total_weight', 0)))
            total_price_all += _group_total_price
            total_weight_all += _group_total_weight
            total_output_kw += _group_output_kw

            _m_idx_first = items[0].get('_matrix_idx')
            _no_label = f"({_m_idx_first + 1})" if _m_idx_first is not None else f'{i + 1}'
            ws.cell(row=row_num, column=1, value=_no_label).alignment = center
            ws.cell(row=row_num, column=2, value=_group_rows)
            ws.cell(row=row_num, column=3, value=_group_cols)
            ws.cell(row=row_num, column=4, value=_group_total_qty)
            ws.merge_cells(f'E{row_num}:F{row_num}')
            ws.cell(row=row_num, column=5, value=_group_angle)
            ws.cell(row=row_num, column=7, value=float(_group_output_kw))

            _sum_refs = []
            for _gr in items:
                _sn = _gr.get('sheet_name', '')
                _tr = _gr.get('total_row')
                if _sn and _tr:
                    _sn_esc = _sn.replace("'", "''")
                    _sum_refs.append(f"'{_sn_esc}'!H{_tr}")
            ws.merge_cells(f'I{row_num}:L{row_num}')
            if _sum_refs:
                ws.cell(row=row_num, column=9, value=f'={"+".join(_sum_refs)}')
            else:
                ws.cell(row=row_num, column=9, value=float(_group_total_price))
            ws.cell(row=row_num, column=9).number_format = summary_currency
            ws.cell(row=row_num, column=9).fill = light_blue

            if _group_total_qty and _group_total_qty > 1:
                ws.cell(row=row_num, column=8, value=f'=I{row_num}/D{row_num}')
            elif _sum_refs:
                ws.cell(row=row_num, column=8, value=f'=I{row_num}')
            else:
                ws.cell(row=row_num, column=8, value=float(_group_total_price))
            ws.cell(row=row_num, column=8).number_format = summary_currency

        for col in range(1, 13):
            c = ws.cell(row=row_num, column=col)
            c.border = thin_border
            c.font = malgun
            if not c.alignment or c.alignment.horizontal is None:
                c.alignment = center
        ws.row_dimensions[row_num].height = 30

    data_end_row = data_start_row + len(_summary_rows) - 1

    # ========== 汇总行 ==========
    summary_row_1 = data_end_row + 1  # 합계

    ws.merge_cells(f'A{summary_row_1}:H{summary_row_1}')
    ws[f'A{summary_row_1}'] = f'합계({summary_currency_label})'
    ws[f'A{summary_row_1}'].alignment = right_a
    ws[f'A{summary_row_1}'].border = thin_border
    ws[f'A{summary_row_1}'].font = malgun
    ws.merge_cells(f'I{summary_row_1}:L{summary_row_1}')
    ws.cell(row=summary_row_1, column=9, value=f'=SUM(I{data_start_row}:I{data_end_row})')
    ws.cell(row=summary_row_1, column=9).number_format = summary_currency
    ws.cell(row=summary_row_1, column=9).border = thin_border
    ws.cell(row=summary_row_1, column=9).font = malgun
    ws.cell(row=summary_row_1, column=9).alignment = center

    grand_total = total_price_all
    global_is_complex = any(qr.get('is_complex', False) for qr in all_quotation_results)

    if ko_discount_in_detail and global_is_complex:
        # 折扣已在明细表单价中，汇总表不需要할인된 금액行
        discount_row = summary_row_1
        summary_row_2 = data_end_row + 2  # 총 출력량
        summary_row_3 = summary_row_2 + 1  # 1W당 비용
    else:
        # ========== 할인된 금액(USD) Discounted amount ==========
        discount_row = data_end_row + 2
        if global_is_complex:
            formula_parts = []
            for qr in all_quotation_results:
                sn = qr.get('sheet_name', '')
                if not sn:
                    continue
                sn_esc = sn.replace("'", "''")
                sc = qr.get('set_count', 1) or 1
                cat_configs = [
                    ('standard_rows', ko_discount_rate),
                    ('steel_rows', ko_steel_discount_rate),
                    ('purchased_rows', ko_purchased_discount_rate),
                    ('inquiry_rows', 100),
                    ('unmatched_rows', 100),
                ]
                for rows_key, rate in cat_configs:
                    rows = qr.get(rows_key, [])
                    if not rows:
                        continue
                    refs = ','.join(f"'{sn_esc}'!H{r}" for r in rows)
                    part = f'SUM({refs})'
                    if rate != 100:
                        part += f'*{rate}/100'
                    formula_parts.append(part)

            if formula_parts:
                discount_value = '=ROUND(' + '+'.join(formula_parts) + ',0)'
            else:
                discount_value = 0
        else:
            if ko_discount_rate != 100:
                discount_value = f'=ROUND(I{summary_row_1}*{ko_discount_rate}/100,0)'
            else:
                discount_value = f'=I{summary_row_1}'

        ws.merge_cells(f'A{discount_row}:H{discount_row}')
        ws[f'A{discount_row}'] = f'할인된 금액({summary_currency_label})\nDiscounted amount'
        ws[f'A{discount_row}'].alignment = right_a
        ws[f'A{discount_row}'].border = thin_border
        ws[f'A{discount_row}'].font = malgun
        ws[f'A{discount_row}'].fill = light_blue
        ws.merge_cells(f'I{discount_row}:L{discount_row}')
        ws.cell(row=discount_row, column=9, value=discount_value)
        ws.cell(row=discount_row, column=9).number_format = '#,##0'
        ws.cell(row=discount_row, column=9).border = thin_border
        ws.cell(row=discount_row, column=9).font = malgun
        ws.cell(row=discount_row, column=9).alignment = center
        ws.cell(row=discount_row, column=9).fill = light_blue
        ws.row_dimensions[discount_row].height = 40

        summary_row_2 = data_end_row + 3  # 총 출력량
        summary_row_3 = summary_row_2 + 1  # 1W당 비용

    ws.merge_cells(f'A{summary_row_2}:H{summary_row_2}')
    ws[f'A{summary_row_2}'] = '총 출력량(KW)'
    ws[f'A{summary_row_2}'].alignment = right_a
    ws[f'A{summary_row_2}'].border = thin_border
    ws[f'A{summary_row_2}'].font = malgun
    ws.merge_cells(f'I{summary_row_2}:L{summary_row_2}')
    summary_output_kw = float(total_output_kw) if total_output_kw > 0 else output_kw
    ws.cell(row=summary_row_2, column=9, value=f'=SUM(G{data_start_row}:G{data_end_row})')
    ws.cell(row=summary_row_2, column=9).border = thin_border
    ws.cell(row=summary_row_2, column=9).font = malgun
    ws.cell(row=summary_row_2, column=9).alignment = center

    ws.merge_cells(f'A{summary_row_3}:H{summary_row_3}')
    ws[f'A{summary_row_3}'] = f'1W당비용({summary_currency_label})'
    ws[f'A{summary_row_3}'].alignment = right_a
    ws[f'A{summary_row_3}'].border = thin_border
    ws[f'A{summary_row_3}'].font = malgun
    ws.merge_cells(f'I{summary_row_3}:L{summary_row_3}')
    ws.cell(row=summary_row_3, column=9, value=f'=ROUND(IF(I{summary_row_2}=0,0,I{discount_row}/(I{summary_row_2}*1000)),3)')
    ws.cell(row=summary_row_3, column=9).number_format = '#,##0.000'
    ws.cell(row=summary_row_3, column=9).border = thin_border
    ws.cell(row=summary_row_3, column=9).font = malgun
    ws.cell(row=summary_row_3, column=9).alignment = center

    if ko_discount_in_detail:
        all_summary_rows = [summary_row_1, summary_row_2, summary_row_3]
        for sr in all_summary_rows:
            ws.row_dimensions[sr].height = 30
    else:
        all_summary_rows = [summary_row_1, discount_row, summary_row_2, summary_row_3]
        for sr in all_summary_rows:
            ws.row_dimensions[sr].height = 30
        ws.row_dimensions[discount_row].height = 40

    # ========== Part 2: 运费 (根据贸易方式决定是否显示) ==========
    if trade_method == 'EXW':
        part2_row = summary_row_3 + 1
        weight_row = part2_row
        ws.merge_cells(f'A{weight_row}:L{weight_row}')
        ws[f'A{weight_row}'] = f'총무게 ~ {round(float(total_weight_all), 2)} kg'
        ws[f'A{weight_row}'].alignment = center
        ws[f'A{weight_row}'].font = malgun
        ws.row_dimensions[weight_row].height = 30

        hint_row = weight_row + 1
        ws.merge_cells(f'A{hint_row}:L{hint_row}')
        ws[f'A{hint_row}'] = '자세한 가격은 다음 시트에 확인해 주세요.'
        ws[f'A{hint_row}'].alignment = left
        ws[f'A{hint_row}'].font = malgun_small_red
        ws.row_dimensions[hint_row].height = 30
        last_row = hint_row
    elif trade_method in ('CIF', 'FOB'):
        part2_row = summary_row_3 + 1
        ws.merge_cells(f'A{part2_row}:L{part2_row}')
        if ko_case_type == 'KSD':
            ws[f'A{part2_row}'] = 'Part 2: KSD 운임비 (Freight Cost)'
        else:
            ws[f'A{part2_row}'] = 'Part 2: 운임비 (Freight Cost)'
        ws[f'A{part2_row}'].font = malgun_bold
        ws[f'A{part2_row}'].alignment = left
        ws.row_dimensions[part2_row].height = 30

        cif_row = part2_row + 1
        ws.merge_cells(f'A{cif_row}:H{cif_row}')
        if trade_method == 'CIF':
            ws[f'A{cif_row}'] = f'CIF {dest_port}({summary_currency_label})'
        elif ko_case_type == 'KSD':
            ws[f'A{cif_row}'] = f'FOB({container_type})({summary_currency_label})'
        else:
            ws[f'A{cif_row}'] = f'FOB({summary_currency_label})'
        ws[f'A{cif_row}'].alignment = right_a
        ws[f'A{cif_row}'].border = thin_border
        ws[f'A{cif_row}'].font = malgun
        ws.merge_cells(f'I{cif_row}:L{cif_row}')
        ws.cell(row=cif_row, column=9, value=ko_freight)
        ws.cell(row=cif_row, column=9).border = thin_border
        ws.cell(row=cif_row, column=9).font = malgun
        ws.cell(row=cif_row, column=9).alignment = center
        ws.row_dimensions[cif_row].height = 30

        container_detail_row = cif_row + 1
        ws.merge_cells(f'A{container_detail_row}:G{container_detail_row}')
        ws[f'A{container_detail_row}'] = '컨테이너 : 20피트/40피트/LCL'
        ws[f'A{container_detail_row}'].alignment = center
        ws[f'A{container_detail_row}'].font = malgun
        ws[f'H{container_detail_row}'] = '사용되는 수량'
        ws[f'H{container_detail_row}'].alignment = center
        ws[f'H{container_detail_row}'].font = malgun
        ws.merge_cells(f'I{container_detail_row}:L{container_detail_row}')
        ws[f'I{container_detail_row}'] = f'{container_type}'
        ws[f'I{container_detail_row}'].alignment = center
        ws[f'I{container_detail_row}'].font = malgun
        ws.row_dimensions[container_detail_row].height = 30

        total_all_row = container_detail_row + 1
        ws.merge_cells(f'A{total_all_row}:H{total_all_row}')
        ws[f'A{total_all_row}'] = ('토탈(RMB)' if sale_type == 'domestic' else '토탈(USD)')
        ws[f'A{total_all_row}'].alignment = right_a
        ws[f'A{total_all_row}'].fill = light_blue
        ws[f'A{total_all_row}'].font = malgun_bold
        ws[f'A{total_all_row}'].border = thin_border
        ws.merge_cells(f'I{total_all_row}:L{total_all_row}')
        ws.cell(row=total_all_row, column=9, value=f'=ROUND(I{discount_row}+I{cif_row},0)')
        ws.cell(row=total_all_row, column=9).number_format = '#,##0'
        ws.cell(row=total_all_row, column=9).fill = light_blue
        ws.cell(row=total_all_row, column=9).font = malgun_bold
        ws.cell(row=total_all_row, column=9).border = thin_border
        ws.cell(row=total_all_row, column=9).alignment = center
        ws.row_dimensions[total_all_row].height = 30

        weight_row = total_all_row + 1
        ws.merge_cells(f'A{weight_row}:L{weight_row}')
        ws[f'A{weight_row}'] = f'총무게 ~ {round(float(total_weight_all), 2)} kg'
        ws[f'A{weight_row}'].alignment = center
        ws[f'A{weight_row}'].font = malgun
        ws.row_dimensions[weight_row].height = 30

        hint_row = weight_row + 1
        ws.merge_cells(f'A{hint_row}:L{hint_row}')
        ws[f'A{hint_row}'] = '자세한 가격은 다음 시트에 확인해 주세요.'
        ws[f'A{hint_row}'].alignment = left
        ws[f'A{hint_row}'].font = malgun_small_red
        ws.row_dimensions[hint_row].height = 30
        last_row = hint_row
    else:
        part2_row = summary_row_3 + 1
        ws.merge_cells(f'A{part2_row}:L{part2_row}')
        if ko_case_type == 'KSD':
            ws[f'A{part2_row}'] = 'Part 2: KSD 운임비 (Freight Cost)+세금(Tax)'
        else:
            ws[f'A{part2_row}'] = 'Part 2: 운임비 (Freight Cost)+세금(Tax)'
        ws[f'A{part2_row}'].font = malgun_bold
        ws[f'A{part2_row}'].alignment = left
        ws.row_dimensions[part2_row].height = 30

        cif_freight_row = part2_row + 1
        ws.merge_cells(f'A{cif_freight_row}:H{cif_freight_row}')
        ws[f'A{cif_freight_row}'] = f'DDP 현장({container_type}):'
        ws[f'A{cif_freight_row}'].alignment = right_a
        ws[f'A{cif_freight_row}'].border = thin_border
        ws[f'A{cif_freight_row}'].font = malgun
        ws.merge_cells(f'I{cif_freight_row}:L{cif_freight_row}')
        ws.cell(row=cif_freight_row, column=9, value=ko_cif_freight + ko_freight)
        ws.cell(row=cif_freight_row, column=9).border = thin_border
        ws.cell(row=cif_freight_row, column=9).font = malgun
        ws.cell(row=cif_freight_row, column=9).alignment = center
        ws.row_dimensions[cif_freight_row].height = 30

        tariff_row = cif_freight_row + 1
        ws.merge_cells(f'A{tariff_row}:H{tariff_row}')
        ws[f'A{tariff_row}'] = '관세+부가세:'
        ws[f'A{tariff_row}'].alignment = right_a
        ws[f'A{tariff_row}'].border = thin_border
        ws[f'A{tariff_row}'].font = malgun
        ws.merge_cells(f'I{tariff_row}:L{tariff_row}')
        ws.cell(row=tariff_row, column=9, value=f'={ko_tariff_rate}/100*({ko_cif_freight}+I{discount_row})+{ko_consumption_tax}/100*(I{discount_row}+{ko_cif_freight}+{ko_tariff_rate}/100*({ko_cif_freight}+I{discount_row}))')
        ws.cell(row=tariff_row, column=9).number_format = summary_currency
        ws.cell(row=tariff_row, column=9).border = thin_border
        ws.cell(row=tariff_row, column=9).font = malgun
        ws.cell(row=tariff_row, column=9).alignment = center
        ws.row_dimensions[tariff_row].height = 30
        total_all_row = tariff_row + 1

        ws.merge_cells(f'A{total_all_row}:H{total_all_row}')
        ws[f'A{total_all_row}'] = ('토탈(RMB)' if sale_type == 'domestic' else '토탈(USD)')
        ws[f'A{total_all_row}'].alignment = right_a
        ws[f'A{total_all_row}'].fill = light_blue
        ws[f'A{total_all_row}'].font = malgun_bold
        ws[f'A{total_all_row}'].border = thin_border
        ws.merge_cells(f'I{total_all_row}:L{total_all_row}')
        ws.cell(row=total_all_row, column=9, value=f'=ROUND(I{discount_row}+I{cif_freight_row}+I{tariff_row},0)')
        ws.cell(row=total_all_row, column=9).number_format = '#,##0'
        ws.cell(row=total_all_row, column=9).fill = light_blue
        ws.cell(row=total_all_row, column=9).font = malgun_bold
        ws.cell(row=total_all_row, column=9).border = thin_border
        ws.cell(row=total_all_row, column=9).alignment = center
        ws.row_dimensions[total_all_row].height = 30

        weight_row = total_all_row + 1
        ws.merge_cells(f'A{weight_row}:L{weight_row}')
        ws[f'A{weight_row}'] = f'총무게 ~ {round(float(total_weight_all), 2)} kg'
        ws[f'A{weight_row}'].alignment = center
        ws[f'A{weight_row}'].font = malgun
        ws.row_dimensions[weight_row].height = 30

        hint_row = weight_row + 1
        ws.merge_cells(f'A{hint_row}:L{hint_row}')
        ws[f'A{hint_row}'] = '자세한 가격은 다음 시트에 확인해 주세요.'
        ws[f'A{hint_row}'].alignment = left
        ws[f'A{hint_row}'].font = malgun_small_red
        ws.row_dimensions[hint_row].height = 30
        last_row = hint_row

    # ========== 外围粗边框 ==========
    _s_t_t = Border(left=thick_side, right=thick_side, top=thick_side, bottom=thin_side)
    _s_t_b = Border(left=thick_side, right=thin_side, top=thick_side, bottom=thin_side)
    _s_b_t = Border(left=thin_side, right=thick_side, top=thick_side, bottom=thin_side)
    _s_t_l = Border(left=thick_side, right=thin_side, top=thin_side, bottom=thick_side)
    _s_b_l = Border(left=thick_side, right=thick_side, top=thin_side, bottom=thick_side)
    _s_b_r = Border(left=thin_side, right=thick_side, top=thin_side, bottom=thick_side)
    _s_c = Border(left=thick_side, right=thick_side, top=thin_side, bottom=thin_side)
    _s_l = Border(left=thick_side, right=thin_side, top=thin_side, bottom=thin_side)
    _s_r = Border(left=thin_side, right=thick_side, top=thin_side, bottom=thin_side)
    _s_inner = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    _s_br = Border(left=thick_side, right=thick_side, top=thick_side, bottom=thick_side)

    for r in range(1, last_row + 1):
        for c in range(1, 13):
            is_top = (r == 1)
            is_bottom = (r == last_row)
            is_left = (c == 1)
            is_right = (c == 12)
            if is_top and is_left:
                b = _s_t_t
            elif is_top and is_right:
                b = _s_b_t
            elif is_bottom and is_left:
                b = _s_b_l
            elif is_bottom and is_right:
                b = _s_b_r
            elif is_top:
                b = _s_c
            elif is_left:
                b = _s_l
            elif is_right:
                b = _s_r
            elif is_bottom:
                b = _s_t_l
            else:
                b = _s_inner
            ws.cell(row=r, column=c).border = b

    # ========== 页面设置 ==========
    apply_print_setup(ws, 'ko_normal')
    ws.sheet_view.zoomScale = 100

    # 移到第一个位置
    sheet_names = workbook.sheetnames
    if '견적서' in sheet_names:
        idx = sheet_names.index('견적서')
        workbook.move_sheet('견적서', offset=-idx)

    print(f"   ✅ 报价单(견적서)已生成")
    return ws.title


def split_and_create_quotations(
        input_file,
        price_file_path,
        output_dir=None,
        image_path=None,
        image_folder=None,
        price_mapping_override=None,
        contact_info=None,
        matrix_data=None,
        return_details=False,
        selected_bom_keys=None,
        group=None,
        need_weight_code=False,
        exclude_options=None,
        trade_method='CIF',
        dest_port='부산',
        container_type='40HQ',
        container_qty=1,
        ko_case_type='NORMAL',
        sale_type='export',
        ko_discount_rate=100,
        ko_steel_discount_rate=84,
        ko_purchased_discount_rate=94,
        ko_tariff_rate=1.6,
        ko_consumption_tax=10,
        ko_freight=0,
        ko_cif_freight=0,
        ko_ddp_address='',
        coating_thickness=10,
        delete_options=None,
        always_exclude_extra_items=False,
        ko_exclude_options=None,
        pre_parsed_products_by_key=None,
        pre_parsed_bom_struct_meta=None,
        ko_discount_in_detail=True,
        need_total_qty=False,
        exclude_delete_options=None,
        need_total_materials=False,
):
    """
    主函数：拆分多工作表BOM并生成韩语报价表（所有报价表在同一个Excel文件的不同Sheet中）
    计价单位从价格表获取
    同时生成独立的询价表文件，包含所有未匹配价格的产品（不去重）
    """
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(input_file), 'quotation_output')
    os.makedirs(output_dir, exist_ok=True)

    # 预扫描图片并准备临时目录
    code_to_images = {}
    image_cache = {}
    image_temp_dir = None
    # 先尝试从日志加载匹配结果
    log_search_dirs = [output_dir, os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")]
    latest_log = find_latest_image_log(log_search_dirs)
    if latest_log:
        log_mapping = load_image_mapping_from_log(latest_log)
        if log_mapping:
            code_to_images.update(log_mapping)
            print(f"已从日志加载图片匹配: {latest_log}")

    if image_folder and os.path.exists(image_folder):
        print(f"图片文件夹: {image_folder}")
        scanned = scan_images(image_folder)
        # 用扫描结果补全日志中缺失的编码
        for code, paths in scanned.items():
            if code not in code_to_images:
                code_to_images[code] = paths
        if code_to_images:
            image_temp_dir = os.path.join(output_dir, f"temp_images_{uuid.uuid4().hex}")
            os.makedirs(image_temp_dir, exist_ok=True)
    elif image_folder:
        print(f"⚠️ 图片文件夹不存在: {image_folder}")
        image_folder = None

    print(f"\n{'=' * 80}")
    print(f"开始处理文件: {input_file}")
    print(f"价格文件: {price_file_path}")
    print(f"输出目录: {output_dir}")
    print(f"{'=' * 80}\n")

    # 加载价格映射
    if price_mapping_override is not None:
        price_mapping = price_mapping_override
    else:
        price_mapping = load_price_mapping(price_file_path)
    selected_key_set = normalize_selected_bom_keys(selected_bom_keys)
    matrix_array_entries = build_matrix_array_entries(matrix_data)
    has_explicit_matrix_arrays = bool((matrix_data or {}).get('arrays'))
    used_matrix_array_indices = set()

    column_mapping = {
        '编码': '编码', 'Part No.': '编码',
        '名称': '名称', 'Part Name': '名称',
        '规格': '规格', 'Spec': '规格',
        '材质': '材质', 'Material': '材质',
        '数量': '数量', 'Qty.': '数量',
        '备注': '备注', 'Remark': '备注',
        '单重': '单重', '(Kg)': '单重'
    }

    skip_keywords = [
        'BOM表-', '导轨切法', '架台角度', '布板方式', '架台类型', '是否内缩',
        '切法辅助', '导轨压块配套方案', '导轨定长方案', '中压块类型',
        '中侧压块配套方案', '导轨总长', '东西板间距', '板规', '阵列',
        '是否东西可调', '导轨伸出面板长度', '跨距', '基础类型',
        '备注：', '基数：', '单基总重', '多基总重'
    ]

    print(f"正在读取文件: {input_file}")

    _cached_md = _get_parse_md(input_file, selected_bom_keys)
    _skip_kw_for_scan = ['对照表', '物料总表', '定价表', '套件明细', 'Sheet']

    _use_zip_fast_path = False
    _zip_bom_info_by_key = {}
    _zip_bom_configs = []

    if pre_parsed_bom_struct_meta and pre_parsed_bom_struct_meta.get('source') == 'zip':
        _zip_bom_info_by_key = pre_parsed_bom_struct_meta.get('bom_info_by_key', {})
        _zip_bom_configs = pre_parsed_bom_struct_meta.get('bom_configs', [])
        _all_covered = all(
            k in (pre_parsed_products_by_key or {})
            for k in _zip_bom_info_by_key
        )
        if _all_covered and _zip_bom_info_by_key:
            _use_zip_fast_path = True
            print(f"   ⚡ 预解析快速路径: {len(_zip_bom_info_by_key)} 个BOM直接复用(零文件读取)")
    elif pre_parsed_products_by_key:
        try:
            _non_bom_kw = ['对照表', '物料总表', '定价表', '套件明细', 'Sheet']
            _zip_products_by_key, _zip_info_by_key, _, _zip_inv, _zip_configs = _parse_bom_sheets_zip(
                input_file,
                normalize_selected_bom_keys(selected_bom_keys),
                _non_bom_kw,
            )
            _all_covered = all(
                k in pre_parsed_products_by_key
                for k in _zip_products_by_key
            )
            if _all_covered and _zip_products_by_key:
                _use_zip_fast_path = True
                _zip_bom_info_by_key = _zip_info_by_key
                _zip_bom_configs = _zip_configs
                print(f"   ⚡ ZIP快速路径: {_zip_bom_info_by_key.__len__()} 个BOM已通过ZIP解析")
        except Exception as _zip_err:
            print(f"   ⚠️ ZIP快速路径失败: {_zip_err}")

    if _use_zip_fast_path:
        _cached_parsed = {}
        _cached_bom_starts = {}
        bom_sheet_names = []
        _sheet_bom_map = {}
        for _bk, _bi in _zip_bom_info_by_key.items():
            _sn = _bi.get('sheet_name', '') or _bk.split('::')[0]
            if _sn not in _sheet_bom_map:
                _sheet_bom_map[_sn] = []
            _sheet_bom_map[_sn].append(_bi)
        bom_sheet_names = list(_sheet_bom_map.keys())
        for _sn, _bis in _sheet_bom_map.items():
            _cached_bom_starts[_sn] = _bis
        xls = None
    elif _cached_md:
        xls = _cached_md['xls']
        _cached_parsed = _cached_md.get('parsed_sheets', {})
        _cached_bom_starts = _cached_md.get('bom_starts_map', {})
        target_sheets_cache = extract_sheet_names_from_keys(selected_key_set)
        bom_sheet_names = []
        for sn in _cached_bom_starts:
            if any(kw in sn for kw in _skip_kw_for_scan):
                continue
            if target_sheets_cache is not None and sn not in target_sheets_cache:
                continue
            if _cached_bom_starts[sn]:
                bom_sheet_names.append(sn)
        print(f"   📦 使用缓存的BOM解析数据 ({len(_cached_parsed)} 个工作表已缓存)")
    else:
        xls = excel_file_compat(input_file)
        _cached_parsed = {}
        _cached_bom_starts = {}
        bom_sheet_names = quick_scan_bom_sheets(xls, _skip_kw_for_scan, selected_key_set)

    master_wb = Workbook()
    default_sheet = master_wb.active
    master_wb.remove(default_sheet)

    all_quotation_results = []
    all_unmatched_products = []
    pile_products_all = []
    inquiry_file = None
    pending_boms = []
    accumulated_results = []

    _all_sheet_names = bom_sheet_names
    if xls is not None:
        _all_sheet_names = xls.sheet_names
        print(f"\n发现 {len(xls.sheet_names)} 个工作表: {xls.sheet_names}")
    else:
        print(f"\nZIP快速路径: 处理 {len(bom_sheet_names)} 个BOM工作表")
    print("=" * 80)

    if has_explicit_matrix_arrays and matrix_array_entries and not _use_zip_fast_path:
        _sheet_match_counts = {}
        for _sn in bom_sheet_names:
            if _sn in _cached_parsed:
                _df = _cached_parsed[_sn]
            else:
                _df = xls.parse(sheet_name=_sn, header=None)
            if _df.empty or len(_df) < 5:
                continue
            if _sn in _cached_bom_starts:
                _bstarts = _cached_bom_starts[_sn]
            else:
                _bstarts = discover_sheet_bom_starts(_df, len(_df), sheet_name=_sn)
                _cached_bom_starts[_sn] = _bstarts
            _sel = [b for b in _bstarts if not selected_key_set or b.get('key') in selected_key_set]
            _cnt = 0
            for _bi in _sel:
                _as = str((_bi.get('config') or {}).get('array', ''))
                _r, _c = parse_array_to_rows_cols(_as)
                _m = (_bi.get('config') or {}).get('missing_boards', 0) or 0
                _b = (_bi.get('config') or {}).get('base_count', 0) or 0
                _, _ma = find_matching_matrix_array(
                    matrix_array_entries, _r, _c, bom_missing=_m, bom_base_count=_b,
                )
                if _ma:
                    _cnt += 1
            _sheet_match_counts[_sn] = _cnt
        if _sheet_match_counts:
            bom_sheet_names = sorted(bom_sheet_names, key=lambda sn: _sheet_match_counts.get(sn, 0), reverse=True)
            print(f"Sheet order by match count: {[(sn, _sheet_match_counts.get(sn, 0)) for sn in bom_sheet_names if sn in _sheet_match_counts]}")

    for sheet_name in bom_sheet_names:
        if _use_zip_fast_path:
            bom_starts = _cached_bom_starts.get(sheet_name, [])
        else:
            if sheet_name in _cached_parsed:
                df = _cached_parsed[sheet_name]
            else:
                df = xls.parse(sheet_name=sheet_name, header=None)
            total_rows = len(df)

            if df.empty or total_rows < 5:
                print(f"\n⏭️  跳过空表: {sheet_name}")
                continue

            print(f"\n📋 处理工作表: {sheet_name}")
            print("-" * 80)
            print(f"   工作表形状: {df.shape}")

            if sheet_name in _cached_bom_starts:
                bom_starts = _cached_bom_starts[sheet_name]
            else:
                bom_starts = discover_sheet_bom_starts(df, total_rows, sheet_name=sheet_name)
        selected_bom_starts = [
            bom_info for bom_info in bom_starts
            if not selected_key_set or bom_info.get('key') in selected_key_set
        ]

        if not _use_zip_fast_path:
            print(f"   发现 {len(bom_starts)} 个BOM表，选中 {len(selected_bom_starts)} 个")

        if matrix_array_entries and not _use_zip_fast_path:
            matched = []
            for bom_info in selected_bom_starts:
                bom_array_str = str((bom_info.get('config') or {}).get('array', ''))
                bom_r, bom_c = parse_array_to_rows_cols(bom_array_str)
                bom_miss = (bom_info.get('config') or {}).get('missing_boards', 0) or 0
                bom_base = (bom_info.get('config') or {}).get('base_count', 0) or 0
                _, matched_array = find_matching_matrix_array(
                    matrix_array_entries, bom_r, bom_c, bom_missing=bom_miss, bom_base_count=bom_base,
                )
                if matched_array:
                    matched.append(bom_info)
                else:
                    print(f"   ⏭️  跳过阵列不匹配的BOM: {bom_info.get('variant_name', '')} "
                          f"(阵列 {bom_array_str} 缺板={bom_miss} 未出现在信息表阵列中)")
            skipped_count = len(selected_bom_starts) - len(matched)
            selected_bom_starts = matched
            if skipped_count:
                matrix_desc = ', '.join(
                    f"{item.get('rows')}x{item.get('cols')}缺板={item.get('missing_per_table',0)}"
                    for item in matrix_array_entries
                )
                print(f"   按信息表阵列({matrix_desc})过滤，跳过 {skipped_count} 个不匹配的BOM")

        if not _use_zip_fast_path:
            print(f"   发现 {len(bom_starts)} 个BOM表，选中 {len(selected_bom_starts)} 个")

        selected_index = 0
        for original_index, bom_info in enumerate(bom_starts, 1):
            if selected_key_set and bom_info.get('key') not in selected_key_set:
                continue
            if bom_info not in selected_bom_starts:
                continue

            selected_index += 1
            print(f"\n   处理第 {selected_index}/{len(selected_bom_starts)} 个BOM表...")

            bom_key = bom_info.get('key', '')
            _pre_parsed_result = None
            if pre_parsed_products_by_key and bom_key in pre_parsed_products_by_key:
                _pre_parsed_result = pre_parsed_products_by_key[bom_key]

            if _pre_parsed_result is not None:
                bom_df = None
            elif _use_zip_fast_path:
                bom_df = None
            else:
                bom_df = extract_bom_dataframe(
                    df, bom_info, original_index, bom_starts, total_rows,
                    column_mapping, skip_keywords
                )

            if bom_df is not None or _pre_parsed_result is not None:
                sheet_prefix = f"{sheet_name}_{bom_info['variant_name']}"

                try:
                    bom_array_str = str((bom_info.get('config') or {}).get('array', ''))
                    bom_rows, bom_cols = parse_array_to_rows_cols(bom_array_str)
                    bom_miss = (bom_info.get('config') or {}).get('missing_boards', 0) or 0
                    bom_base = (bom_info.get('config') or {}).get('base_count', 0) or 0
                    matched_idx, matched_array = find_matching_matrix_array(
                        matrix_array_entries,
                        bom_rows,
                        bom_cols,
                        used_indices=used_matrix_array_indices if has_explicit_matrix_arrays else None,
                        bom_missing=bom_miss if bom_miss != 0 else None,
                        bom_base_count=bom_base,
                        strict_only=has_explicit_matrix_arrays,
                    )
                    if matched_array is None and has_explicit_matrix_arrays:
                        if _pre_parsed_result is not None:
                            _pb_products = _pre_parsed_result[0]
                        else:
                            _pb_products, _, _ = read_bom_from_dataframe(bom_df) if bom_df is not None else ([], None, None)
                        pending_boms.append({
                            'sheet_name': sheet_name,
                            'bom_info': bom_info,
                            'products': _pb_products or [],
                            'rows': bom_rows,
                            'cols': bom_cols,
                            'base_count': bom_base,
                            'missing': bom_miss,
                            'config': bom_info.get('config', {}),
                        })
                        continue
                    if has_explicit_matrix_arrays and matched_idx is not None:
                        used_matrix_array_indices.add(matched_idx)

                    ma_qty = (matched_array or {}).get('table_qty', 1) if matched_array else 1
                    if matched_array:
                        ma_rows = matched_array.get('rows', '')
                        ma_cols = matched_array.get('cols', '')
                        ma_miss = matched_array.get('missing_per_table', 0) or 0
                        ma_no = matched_array.get('no', '')
                        base_prefix = f"{ma_rows}×{ma_cols}_{ma_qty}"
                        same_dim_entries = [
                            e for e in (matrix_array_entries or [])
                            if e.get('rows') == ma_rows and e.get('cols') == ma_cols
                            and (e.get('table_qty') or 1) == ma_qty
                        ]
                        if len(same_dim_entries) > 1:
                            if ma_miss != 0:
                                sheet_prefix = f"{base_prefix}_缺{ma_miss}"
                            elif ma_no:
                                sheet_prefix = f"{base_prefix}_{ma_no}"
                            else:
                                sheet_prefix = base_prefix
                        else:
                            sheet_prefix = base_prefix
                        if matched_idx is not None:
                            sheet_prefix = f"({matched_idx + 1}){sheet_prefix}"

                    effective_matrix_data = build_bom_matrix_data(
                        matrix_data, matched_array, bom_config=bom_info.get('config'),
                    )
                    _cpp = _pre_parsed_result if _pre_parsed_result is not None else None
                    result = create_quotation_from_dataframe(
                        bom_df,
                        master_wb,
                        sheet_prefix,
                        price_mapping,
                        image_path,
                        image_folder,
                        code_to_images,
                        image_temp_dir,
                        image_cache,
                        all_unmatched_products,
                        contact_info,
                        config=bom_info['config'],
                        matrix_data=effective_matrix_data,
                        group=group,
                        exclude_options=exclude_options,
                        sale_type=sale_type,
                    ko_discount_rate=ko_discount_rate,
                    ko_steel_discount_rate=ko_steel_discount_rate,
                    ko_purchased_discount_rate=ko_purchased_discount_rate,
                        coating_thickness=coating_thickness,
                        delete_options=delete_options,
                        always_exclude_extra_items=always_exclude_extra_items,
                        ko_exclude_options=ko_exclude_options,
                        pre_parsed_products=_cpp,
                        ko_discount_in_detail=ko_discount_in_detail,
                    )
                    result['config'] = bom_info['config']
                    result['matched_array'] = matched_array
                    result['matrix_data'] = effective_matrix_data
                    result['set_count'] = effective_matrix_data.get('set_count') or 1
                    if matched_idx is not None:
                        result['_matrix_idx'] = matched_idx
                    bom_cfg = bom_info.get('config') or {}
                    result['matched_info'] = {
                        'bom_key': bom_info.get('key', ''),
                        'bom_variant': bom_info.get('variant_name', ''),
                        'bom_array': bom_cfg.get('array', ''),
                        'bom_missing': bom_cfg.get('missing_boards', 0),
                        'bom_base_count': bom_cfg.get('base_count', 0),
                        'matched_array_no': (matched_array or {}).get('no', ''),
                    }
                    all_quotation_results.append(result)
                    pile_prods = result.get('pile_products', [])
                    if pile_prods:
                        for pp in pile_prods:
                            scaled = dict(pp)
                            scaled['quantity'] = float(pp.get('quantity', 0))
                            pile_products_all.append(scaled)
                    print(f"   ✅ 报价表生成成功: {result['sheet_name']} "
                          f"(匹配{result['matched_count']}项, 未匹配{result['unmatched_count']}项, "
                          f"米计价{result['meter_unit_count']}项, 个/套计价{result['piece_unit_count']}项, "
                          f"图片{result['image_found_count']}/{result['image_found_count'] + result['image_not_found_count']})")
                except Exception as e:
                    print(f"   ❌ 报价表生成失败: {e}")
            else:
                print(f"   ❌ BOM数据提取失败")

    # ========== 累加匹配 ==========
    if pending_boms and has_explicit_matrix_arrays and matrix_array_entries:
        used_sheet_names = set()
        for r in all_quotation_results:
            _sn = r.get('sheet_name', '')
            if _sn:
                used_sheet_names.add(_sn)

        accumulated_results = find_accumulated_match(
            matrix_array_entries, pending_boms,
            used_matrix_array_indices, used_sheet_names,
        )

        for acc in accumulated_results:
            m_idx = acc['matrix_idx']
            matrix_entry = acc['matrix_entry']
            accumulated_base = acc['accumulated_base']
            merged_products = acc.get('merged_products', [])
            selected_boms = acc.get('selected_boms', [])

            used_matrix_array_indices.add(m_idx)

            ma_rows = matrix_entry.get('rows', '')
            ma_cols = matrix_entry.get('cols', '')
            ma_qty = matrix_entry.get('table_qty', 1)
            _accum_group_id = f"{ma_rows}×{ma_cols}_{ma_qty}"

            sheet_prefix = f"{ma_rows}×{ma_cols}_{ma_qty}"
            sheet_prefix = f"({m_idx + 1}){sheet_prefix}"

            _pb_config = (selected_boms[0].get('config') if selected_boms else None) or {}
            _pb_array_info = f"{ma_rows}×{ma_cols}" if ma_rows and ma_cols else ''

            _pb_effective_md = build_bom_matrix_data(
                matrix_data, matrix_entry, bom_config=_pb_config,
            )

            try:
                _pb_result = create_quotation_from_dataframe(
                    None, master_wb, sheet_prefix, price_mapping,
                    image_path=image_path,
                    image_folder=image_folder,
                    code_to_images=code_to_images,
                    image_temp_dir=image_temp_dir,
                    image_cache=image_cache,
                    unmatched_products_list=all_unmatched_products,
                    contact_info=contact_info,
                    config=_pb_config,
                    matrix_data=_pb_effective_md,
                    group=group,
                    exclude_options=exclude_options,
                    sale_type=sale_type,
                    ko_discount_rate=ko_discount_rate,
                    ko_steel_discount_rate=ko_steel_discount_rate,
                    ko_purchased_discount_rate=ko_purchased_discount_rate,
                    coating_thickness=coating_thickness,
                    delete_options=delete_options,
                    always_exclude_extra_items=always_exclude_extra_items,
                    ko_exclude_options=ko_exclude_options,
                    pre_parsed_products=(merged_products, _pb_array_info, _pb_config.get('cross_span', '')),
                    ko_discount_in_detail=ko_discount_in_detail,
                )
                _pb_result['config'] = _pb_config
                _pb_result['matched_array'] = matrix_entry
                _pb_result['matrix_data'] = _pb_effective_md
                _pb_result['set_count'] = ma_qty
                _pb_result['_matrix_idx'] = m_idx
                _pb_result['_acc_group_id'] = _accum_group_id
                _pb_result['accumulated_sub_idx'] = 0
                _pb_result['_acc_group_total_qty'] = ma_qty
                _pb_result['matched_info'] = {
                    'bom_key': '',
                    'bom_variant': f"acc_merged_{len(selected_boms)}boms",
                    'bom_array': f"{ma_rows}×{ma_cols}",
                    'bom_missing': _pb_config.get('missing_boards', 0) or 0,
                    'bom_base_count': ma_qty,
                    'matched_array_no': '',
                }
                all_quotation_results.append(_pb_result)

                _pb_pile = _pb_result.get('pile_products', [])
                if _pb_pile:
                    for pp in _pb_pile:
                        scaled = dict(pp)
                        scaled['quantity'] = float(pp.get('quantity', 0))
                        pile_products_all.append(scaled)

                print(f"   ✅ 报价表生成成功(累加合并): {_pb_result['sheet_name']} "
                      f"(合并{len(selected_boms)}个BOM, base={ma_qty}, "
                      f"匹配{_pb_result['matched_count']}项)")
            except Exception as e:
                import traceback
                print(f"   ❌ 累加报价表生成失败: {e}")
                traceback.print_exc()

    # ========== 信息表累加匹配 (多个小信息表 → 一个大BOM) ==========
    if pending_boms and has_explicit_matrix_arrays and matrix_array_entries:
        used_bom_indices_in_acc = set()
        for acc in accumulated_results:
            for sel_bom in acc.get('selected_boms', []):
                for k, pb in enumerate(pending_boms):
                    if pb is sel_bom:
                        used_bom_indices_in_acc.add(k)
                        break

        info_acc_results = find_info_accumulated_match(
            matrix_array_entries, pending_boms,
            used_matrix_array_indices,
            used_bom_indices=used_bom_indices_in_acc,
        )

        for acc in info_acc_results:
            m_indices = acc['matrix_indices']
            matched_bom = acc['matched_bom']
            total_info_qty = acc['total_info_qty']

            for m_idx in m_indices:
                used_matrix_array_indices.add(m_idx)

            bom_products = matched_bom.get('products') or []
            bom_config = matched_bom.get('config') or matched_bom
            bom_base = matched_bom.get('base_count', 0) or 0

            first_entry = acc['matrix_entries'][0]
            ma_rows = first_entry.get('rows', '')
            ma_cols = first_entry.get('cols', '')
            sheet_prefix = f"{ma_rows}×{ma_cols}_{bom_base}"
            if m_indices:
                sheet_prefix = f"({m_indices[0] + 1}){sheet_prefix}"

            _acc_entry = dict(first_entry)
            _acc_entry['table_qty'] = bom_base

            effective_matrix_data = build_bom_matrix_data(
                matrix_data, _acc_entry, bom_config=bom_config,
            )

            try:
                result = create_quotation_from_dataframe(
                    None, master_wb, sheet_prefix, price_mapping,
                    image_path=image_path,
                    image_folder=image_folder,
                    code_to_images=code_to_images,
                    image_temp_dir=image_temp_dir,
                    image_cache=image_cache,
                    unmatched_products_list=all_unmatched_products,
                    contact_info=contact_info,
                    config=bom_config,
                    matrix_data=effective_matrix_data,
                    group=group,
                    exclude_options=exclude_options,
                    sale_type=sale_type,
                    ko_discount_rate=ko_discount_rate,
                    ko_steel_discount_rate=ko_steel_discount_rate,
                    ko_purchased_discount_rate=ko_purchased_discount_rate,
                    coating_thickness=coating_thickness,
                    delete_options=delete_options,
                    always_exclude_extra_items=always_exclude_extra_items,
                    ko_exclude_options=ko_exclude_options,
                    pre_parsed_products=(bom_products, f"{ma_rows}×{ma_cols}" if ma_rows and ma_cols else '', bom_config.get('cross_span', '')),
                    ko_discount_in_detail=ko_discount_in_detail,
                )
                result['config'] = bom_config
                result['matched_array'] = first_entry
                result['matrix_data'] = effective_matrix_data
                result['set_count'] = bom_base
                result['_matrix_idx'] = m_indices[0] if m_indices else None
                result['matched_info'] = {
                    'bom_key': '',
                    'bom_variant': f"info_accumulated_{len(m_indices)}entries",
                    'bom_array': f"{ma_rows}×{ma_cols}",
                    'bom_missing': bom_config.get('missing_boards', 0) or 0,
                    'bom_base_count': bom_base,
                    'matched_array_no': '',
                }
                all_quotation_results.append(result)
                print(f"   ✅ 报价表生成成功(信息表累加): {result['sheet_name']} "
                      f"({len(m_indices)}个信息表 → BOM base={bom_base}, "
                      f"匹配{result['matched_count']}项)")
            except Exception as e:
                import traceback
                print(f"   ❌ 信息表累加报价表生成失败: {e}")
                traceback.print_exc()

    # ========== Pass 3: 间接匹配兜底 (rows×cols相同即可) ==========
    if pending_boms and has_explicit_matrix_arrays and matrix_array_entries:
        _used_bom_set = set()
        for acc in accumulated_results:
            for sel_bom in acc.get('selected_boms', []):
                for k, pb in enumerate(pending_boms):
                    if pb is sel_bom:
                        _used_bom_set.add(k)
        for acc in info_acc_results if 'info_acc_results' in dir() else []:
            matched = acc.get('matched_bom')
            if matched:
                for k, pb in enumerate(pending_boms):
                    if pb is matched:
                        _used_bom_set.add(k)

        remaining_info_indices = [
            i for i in range(len(matrix_array_entries))
            if i not in used_matrix_array_indices
        ]

        for m_idx in remaining_info_indices:
            info_entry = matrix_array_entries[m_idx]
            info_rows = info_entry.get('rows')
            info_cols = info_entry.get('cols')
            info_qty = info_entry.get('table_qty', 1)
            if info_rows is None or info_cols is None:
                continue

            matched_pb = None
            matched_pb_k = None
            for k, pb in enumerate(pending_boms):
                if k in _used_bom_set:
                    continue
                if pb.get('rows') == info_rows and pb.get('cols') == info_cols:
                    matched_pb = pb
                    matched_pb_k = k
                    break

            if matched_pb is None:
                continue

            _used_bom_set.add(matched_pb_k)
            used_matrix_array_indices.add(m_idx)

            bom_products = matched_pb.get('products') or []
            bom_config = matched_pb.get('config') or {}
            bom_base = matched_pb.get('base_count', 0) or 0

            sheet_prefix = f"{info_rows}×{info_cols}_{info_qty}"
            sheet_prefix = f"({m_idx + 1}){sheet_prefix}"
            same_dim_entries = [
                e for e in matrix_array_entries
                if e.get('rows') == info_rows and e.get('cols') == info_cols
                and (e.get('table_qty') or 1) == info_qty
            ]
            if len(same_dim_entries) > 1:
                info_miss = info_entry.get('missing_per_table', 0) or 0
                info_no = info_entry.get('no', '')
                if info_miss != 0:
                    sheet_prefix = f"{sheet_prefix}_缺{info_miss}"
                elif info_no:
                    sheet_prefix = f"{sheet_prefix}_{info_no}"

            effective_matrix_data = build_bom_matrix_data(
                matrix_data, info_entry, bom_config=bom_config,
            )

            try:
                result = create_quotation_from_dataframe(
                    None, master_wb, sheet_prefix, price_mapping,
                    image_path=image_path,
                    image_folder=image_folder,
                    code_to_images=code_to_images,
                    image_temp_dir=image_temp_dir,
                    image_cache=image_cache,
                    unmatched_products_list=all_unmatched_products,
                    contact_info=contact_info,
                    config=bom_config,
                    matrix_data=effective_matrix_data,
                    group=group,
                    exclude_options=exclude_options,
                    sale_type=sale_type,
                    ko_discount_rate=ko_discount_rate,
                    ko_steel_discount_rate=ko_steel_discount_rate,
                    ko_purchased_discount_rate=ko_purchased_discount_rate,
                    coating_thickness=coating_thickness,
                    delete_options=delete_options,
                    always_exclude_extra_items=always_exclude_extra_items,
                    ko_exclude_options=ko_exclude_options,
                    pre_parsed_products=(bom_products, f"{info_rows}×{info_cols}" if info_rows and info_cols else '', bom_config.get('cross_span', '')),
                    ko_discount_in_detail=ko_discount_in_detail,
                )
                result['config'] = bom_config
                result['matched_array'] = info_entry
                result['matrix_data'] = effective_matrix_data
                result['set_count'] = effective_matrix_data.get('set_count') or 1
                result['_matrix_idx'] = m_idx
                result['matched_info'] = {
                    'bom_key': '',
                    'bom_variant': 'indirect',
                    'bom_array': f"{info_rows}×{info_cols}",
                    'bom_missing': bom_config.get('missing_boards', 0) or 0,
                    'bom_base_count': bom_base,
                    'matched_array_no': '',
                }
                all_quotation_results.append(result)
                print(f"   ✅ 报价表生成成功(间接): {result['sheet_name']} "
                      f"(信息{info_rows}×{info_cols}_{info_qty} → BOM base={bom_base}, "
                      f"匹配{result['matched_count']}项)")
            except Exception as e:
                import traceback
                print(f"   ❌ 间接匹配报价表生成失败: {e}")
                traceback.print_exc()

    pile_summary = None
    if pile_products_all:
        from decimal import Decimal as _Dec
        _pile_total_qty = 0
        _pile_total_price = 0.0
        for _pp in pile_products_all:
            _pp_code = _pp.get('code', '')
            _pp_qty = float(_pp.get('quantity', 0))
            _pp_pi = resolve_price_info(price_mapping, _pp_code, spec=_pp.get('spec', '')) if price_mapping else None
            _pp_price = get_temp_adjusted_base_price(_pp_pi, _pp, '韩语组', sale_type) if _pp_pi and has_valid_price_info(_pp_pi) else 0
            _pp_unit = _pp_pi.get('unit', '') if _pp_pi else ''
            _is_m = _pp_unit in ('米', 'm', 'M', 'meter', 'Meter', 'meters')
            if _is_m:
                _len = float(extract_length_from_spec(_pp.get('spec', '')) or 0)
                if _len > 0:
                    _pp_price = _pp_price * _len / 1000
            _pile_total_qty += _pp_qty
            _pile_total_price += _pp_price * _pp_qty
        pile_summary = {
            'total_qty': _pile_total_qty,
            'total_price': round(_pile_total_price, 2),
        }
        if pile_summary:
            print(f"   🔩 地桩汇总: 数量={_pile_total_qty}, 金额={_pile_total_price:.2f}")

    # 按 matrix_array_entries（信息表）顺序重排 sheet
    if all_quotation_results and has_explicit_matrix_arrays and matrix_array_entries:
        reorder_sheets_by_matrix_array(master_wb, all_quotation_results, matrix_array_entries, log_prefix='[NORMAL]')

    # 保存报价汇总文件
    if all_quotation_results:
        if matrix_data:
            try:
                create_summary_quotation_sheet(
                    master_wb, all_quotation_results,
                    matrix_data=matrix_data,
                    image_path=image_path,
                    contact_info=contact_info,
                    pile_summary=pile_summary,
                    trade_method=trade_method,
                    dest_port=dest_port,
                    container_type=container_type,
                    container_qty=container_qty,
                    ko_case_type=ko_case_type,
                    ko_discount_rate=ko_discount_rate,
                    ko_tariff_rate=ko_tariff_rate,
                    ko_consumption_tax=ko_consumption_tax,
                    sale_type=sale_type,
                    ko_freight=ko_freight,
                    ko_cif_freight=ko_cif_freight,
                    ko_ddp_address=ko_ddp_address,
                    ko_steel_discount_rate=ko_steel_discount_rate,
                    ko_purchased_discount_rate=ko_purchased_discount_rate,
                    ko_discount_in_detail=ko_discount_in_detail,
                )
            except Exception as e:
                print(f"   ❌ 报价单(견적서)生成失败: {e}")

        if need_total_materials:
            try:
                create_total_materials_sheet(
                    master_wb,
                    all_quotation_results,
                    price_mapping=price_mapping,
                    sale_type=sale_type,
                    coating_thickness=coating_thickness,
                    lang='ko',
                    need_weight_code=need_weight_code,
                    need_total_qty=need_total_qty,
                    discount_method='project',
                    ko_discount_rate=ko_discount_rate,
                    ko_steel_discount_rate=ko_steel_discount_rate,
                    ko_purchased_discount_rate=ko_purchased_discount_rate,
                    pile_products=pile_products_all,
                    show_code=True,
                )
            except Exception as e:
                import traceback
                print(f"   ❌ 物料汇总页生成失败: {e}")
                traceback.print_exc()

        input_basename = os.path.splitext(os.path.basename(input_file))[0]
        output_file = os.path.join(output_dir, f"{input_basename}_报价汇总.xlsx")

        set_page_break_preview(master_wb)
        master_wb.save(output_file)

        print("\n" + "=" * 80)
        print(f"✅ 处理完成！共生成 {len(all_quotation_results)} 个报价表")
        print(f"📁 报价汇总文件: {output_file}")

        # 如果有未匹配的产品，生成独立的询价表文件（不去重）
        if all_unmatched_products:
            print(f"\n📋 发现 {len(all_unmatched_products)} 个未匹配价格的产品条目...")

            # 统计未匹配的产品（不去重，显示所有条目）
            print(f"   📊 未匹配产品明细（前20条）:")
            for i, p in enumerate(all_unmatched_products[:20], 1):
                print(f"      {i}. 编码:{p['code']} | 名称:{p['name']} | 数量:{p['quantity']}")
            if len(all_unmatched_products) > 20:
                print(f"      ... 还有 {len(all_unmatched_products) - 20} 条")

            # 保存为独立的询价表文件（不去重）
            inquiry_file = save_inquiry_sheet_to_file(
                all_unmatched_products,  # 直接传入，不去重
                output_dir,
                input_file,
                inquiry_requester=(contact_info or {}).get('inquiry_requester', ''),
            )

            if inquiry_file:
                print(f"📋 询价表已单独保存: {inquiry_file}")

        print("\n📊 生成工作表汇总:")
        print("-" * 80)

        total_products = 0
        total_weight = 0
        total_price = 0
        total_matched = 0
        total_unmatched = 0
        total_meter_unit = 0
        total_piece_unit = 0
        total_length_fail = 0
        total_images_found = 0
        total_images_not_found = 0

        for i, item in enumerate(all_quotation_results, 1):
            print(f"  {i:2d}. {item['sheet_name']}")
            print(
                f"      物料项数: {item['valid_products']} 项 | 总重量: {item['total_weight']:.2f} KG | 总价: {item['total_price']:.2f} USD")
            print(f"      价格匹配: 成功 {item['matched_count']} 项, 失败 {item['unmatched_count']} 项")
            print(f"      计价单位: 米计价 {item['meter_unit_count']} 项, 个/套计价 {item['piece_unit_count']} 项")
            print(f"      图片匹配: 成功 {item['image_found_count']} 张, 未找到 {item['image_not_found_count']} 张")
            if item.get('length_extract_fail_count', 0) > 0:
                print(f"      长度提取失败: {item['length_extract_fail_count']} 项 (已按个计价)")
            total_products += item['valid_products']
            total_weight += item['total_weight']
            total_price += item['total_price']
            total_matched += item['matched_count']
            total_unmatched += item['unmatched_count']
            total_meter_unit += item['meter_unit_count']
            total_piece_unit += item['piece_unit_count']
            total_length_fail += item.get('length_extract_fail_count', 0)
            total_images_found += item['image_found_count']
            total_images_not_found += item['image_not_found_count']

        print("\n📊 总体统计:")
        print(f"  总工作表数: {len(all_quotation_results)}")
        print(f"  总物料项数: {total_products}")
        print(f"  总重量: {total_weight:.2f} KG")
        print(f"  总金额: {total_price:.2f} USD")
        print(f"  价格匹配: 成功 {total_matched} 项, 失败 {total_unmatched} 项")
        print(f"  计价单位: 米计价 {total_meter_unit} 项, 个/套计价 {total_piece_unit} 项")
        print(f"  图片匹配: 成功 {total_images_found} 张, 未找到 {total_images_not_found} 张")
        if total_length_fail > 0:
            print(f"  长度提取失败: {total_length_fail} 项 (已按个计价)")

        print("\n" + "=" * 80)
        print("✅ 所有报价表生成完成！")

        # 打印询价表统计
        if all_unmatched_products:
            print(f"\n📋 询价表统计:")
            print(f"  未匹配产品总数: {len(all_unmatched_products)} 条")
            # 按编码统计重复情况
            from collections import Counter
            code_counts = Counter([p['code'] for p in all_unmatched_products if p['code']])
            duplicate_codes = {code: count for code, count in code_counts.items() if count > 1}
            if duplicate_codes:
                print(f"  重复出现的编码: {len(duplicate_codes)} 个")
                for code, count in list(duplicate_codes.items())[:10]:
                    print(f"      {code}: 出现 {count} 次")

        # 清理临时图片目录
        if image_temp_dir and os.path.exists(image_temp_dir):
            try:
                shutil.rmtree(image_temp_dir)
            except:
                pass

        if return_details:
            quotation_product_codes = set()
            for item in all_quotation_results:
                for c in (item.get('quotation_product_codes') or []):
                    quotation_product_codes.add(c)
            return {
                'output_file': output_file,
                'inquiry_file': inquiry_file,
                'unmatched_count': len(all_unmatched_products),
                'unmatched_products': all_unmatched_products,
                'quotation_product_codes': quotation_product_codes,
            }
        return output_file
    else:
        print("\n❌ 没有生成任何报价表")
        if image_temp_dir and os.path.exists(image_temp_dir):
            try:
                shutil.rmtree(image_temp_dir)
            except:
                pass
        return None


def analyze_unmatched_codes(input_file, price_file_path, price_mapping_override=None):
    """
    Use the same BOM parsing logic as quotation generation to find unmatched codes.
    Returns: dict with unmatched_codes, total_products, matched_count, unmatched_count.
    """
    if price_mapping_override is not None:
        price_mapping = price_mapping_override
    else:
        price_mapping = load_price_mapping(price_file_path)

    column_mapping = {
        '编码': '编码', 'Part No.': '编码',
        '名称': '名称', 'Part Name': '名称',
        '规格': '规格', 'Spec': '规格',
        '材质': '材质', 'Material': '材质',
        '数量': '数量', 'Qty.': '数量',
        '备注': '备注', 'Remark': '备注',
        '单重': '单重', '(Kg)': '单重'
    }

    skip_keywords = [
        'BOM表', '导轨切法', '柜台角度', '布局方式', '柜台类型', '是否内缩',
        '切法辅助', '导轨压块配套方案', '导轨定长方案', '中压块类型',
        '中侧压块配套方案', '导轨总长', '东西板间距', '板规', '阵列',
        '是否东西可调', '导轨伸出面板长度', '跨距', '基础类型',
        '备注：', '基数：', '单基总重', '多基总重'
    ]

    xls = excel_file_compat(input_file)
    unmatched_codes = []
    total_products = 0
    matched_count = 0
    unmatched_items_count = 0

    bom_sheet_names = quick_scan_bom_sheets(xls, ['配套', '配置', '价格表', '物料总表'])

    for sheet_name in bom_sheet_names:
        df = xls.parse(sheet_name=sheet_name, header=None)
        total_rows = len(df)
        if df.empty or total_rows < 5:
            continue

        bom_starts = []
        for idx, row in df.iterrows():
            row_str = ' '.join([str(x) if pd.notna(x) else '' for x in row])
            if "BOM表" in row_str:
                config = extract_config_info(df, idx, total_rows)
                bom_starts.append({
                    'row': idx,
                    'config': config,
                    'variant_name': f"{config.get('array', '未知')}_{config.get('variant', '标准')}"
                })

        for i, bom_info in enumerate(bom_starts, 1):
            bom_df = extract_bom_dataframe(
                df, bom_info, i, bom_starts, total_rows,
                column_mapping, skip_keywords
            )
            if bom_df is None or bom_df.empty:
                continue

            products, _, _ = read_bom_from_dataframe(bom_df)
            for product in products:
                product_code = str(product.get('code', '')).strip()
                if not product_code:
                    continue
                total_products += 1
                if product_code in price_mapping:
                    matched_count += 1
                else:
                    unmatched_items_count += 1
                    unmatched_codes.append(product_code)

    unique_unmatched = []
    seen = set()
    for code in unmatched_codes:
        if code not in seen:
            unique_unmatched.append(code)
            seen.add(code)

    return {
        'unmatched_codes': unique_unmatched,
        'unmatched_code_count': len(unique_unmatched),
        'unmatched_items_count': unmatched_items_count,
        'total_products': total_products,
        'matched_count': matched_count,
        'unmatched_count': len(unique_unmatched)
    }

