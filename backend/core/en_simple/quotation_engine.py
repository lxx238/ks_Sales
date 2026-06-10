import math
import os
import re
from decimal import Decimal

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins


def _safe_int(val, default=0):
    try:
        if isinstance(val, float) and math.isnan(val):
            return default
        return int(val)
    except (ValueError, TypeError, OverflowError):
        return default


def _safe_float(val, default=0.0):
    try:
        f = float(val)
        return default if math.isnan(f) else f
    except (ValueError, TypeError):
        return default

from backend.core.shared.price_utils import (
    resolve_price_info,
    has_valid_price_info,
    round_to_2_decimal,
    _get_discount_category,
    _is_standard_priced,
)
from backend.core.shared.weight_utils import (
    extract_length_from_spec,
    WEIGHT_BY_LENGTH_ATTRIBUTES,
)
from backend.core.shared.text_utils import (
    normalize_lookup_code,
    extract_main_name,
    parse_decimal_number,
    _strip_cjk_spec,
    normalize_angle,
)
from backend.core.shared.product_utils import (
    _is_valid_product_code,
    _match_exclude_group,
)
from backend.core.shared.constants import (
    EXCLUDE_ITEM_GROUPS,
    CARBON_STEEL_PRICING_ATTRS,
    IMAGE_WIDTH,
    IMAGE_HEIGHT,
    IMAGE_COLUMN_INDEX,
    IMAGE_PADDING,
)
from backend.core.shared.image_utils import (
    prepare_image_for_excel,
    add_image_centered_in_cell,
)
from backend.core.material_translate import translate_material, adjust_material_by_coating
from backend.core.shared.translations import t as _t

CURRENCY_FMT = '"US$" #,##0.00'

_EN_THIN = Side(style='thin', color='000000')
_EN_THICK = Side(style='medium', color='000000')
_EN_THIN_BORDER = Border(left=_EN_THIN, right=_EN_THIN, top=_EN_THIN, bottom=_EN_THIN)

_EN_TITLE_FONT = Font(name='Malgun Gothic', size=36, bold=True, color='000000')
_EN_NORMAL_FONT = Font(name='Malgun Gothic', size=16, color='000000')
_EN_HEADER_FONT = Font(name='Malgun Gothic', size=16, bold=True, color='000000')
_EN_SMALL_BOLD_FONT = Font(name='Malgun Gothic', size=16, bold=True, color='000000')
_EN_RED_SMALL_FONT = Font(name='Malgun Gothic', size=16, bold=True, color='FF0000')
_EN_A7_FILL = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
_EN_YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
_EN_CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_EN_LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
_EN_RIGHT_ALIGN = Alignment(horizontal='right', vertical='center', wrap_text=True)

_EN_SUMMARY_NORMAL_FONT = Font(name='Arial', size=11)
_EN_SUMMARY_BOLD_FONT = Font(name='Arial', size=11, bold=True)
_EN_SUMMARY_TITLE_FONT = Font(name='Arial', size=26, bold=True)
_EN_SUMMARY_COMPANY_FONT = Font(name='Arial', size=15, bold=True)
_EN_SUMMARY_PART_TITLE_FONT = Font(name='Arial', size=14, bold=True)
_EN_SUMMARY_PART_INNER_FONT = Font(name='Arial', size=12)
_EN_SUMMARY_PART_INNER_BOLD_FONT = Font(name='Arial', size=12, bold=True)
_EN_SUMMARY_AMOUNT_FONT = Font(name='Arial', size=13, bold=True, color='000000')
_EN_SUMMARY_INFO_FONT = Font(name='Arial', size=11)
_EN_SUMMARY_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_EN_SUMMARY_LEFT_BOTTOM = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
_EN_SUMMARY_RIGHT_A = Alignment(horizontal='right', vertical='center', wrap_text=True)
_EN_SUMMARY_BOTTOM_BORDER = Border(bottom=_EN_THIN)
_EN_SUMMARY_YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

_LANG_NAME_KEY_MAP = {
    'en': ('name_en',),
    'fr': ('name_fr', 'name_en'),
    'es': ('name_es', 'name_en'),
}


def _resolve_product_name(price_info, product, lang='en'):
    fallback = product.get('name', '')
    if not price_info:
        return fallback
    for key in _LANG_NAME_KEY_MAP.get(lang, ('name_en',)):
        val = price_info.get(key)
        if val:
            return val
    for key in ('name_ko', 'name'):
        val = price_info.get(key)
        if val:
            return val
    return fallback


def _set_cell(ws, row, col, val, font=None, align=None, border=None, number_format=None, fill=None):
    cell = ws.cell(row=row, column=col, value=val)
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format
    if fill is not None:
        cell.fill = fill
    return cell


def _is_carbon_steel_cached(product, price_info):
    if price_info:
        attr = str(price_info.get('pricing_attribute', '')).strip().upper()
        if attr in CARBON_STEEL_PRICING_ATTRS:
            return True
    return False


def _classify_products_single_pass(all_products, price_mapping, delete_options, always_exclude, ko_exclude_options=None):
    delete_options = delete_options or {}
    ko_exclude_options = ko_exclude_options or {}
    all_exclude_opts = {k: True for k in EXCLUDE_ITEM_GROUPS} if always_exclude else {}
    for k in list(ko_exclude_options.keys()) + list(delete_options.keys()):
        all_exclude_opts[k] = True
    aluminum = []
    carbon_steel = []
    excluded = []
    pile_products = []
    price_info_cache = {}
    for p in all_products:
        code = str(p.get('code', '')).strip()
        spec = str(p.get('spec', '')).strip()
        cache_key = (code, spec) if spec else code
        if cache_key not in price_info_cache:
            price_info_cache[cache_key] = resolve_price_info(price_mapping, code, spec=spec) if price_mapping else None
        pi = price_info_cache[cache_key]
        p['_price_info'] = pi
        attr = (pi.get('attribute', '') if pi else '') or ''
        if attr.strip() in ('地桩', '地盤杭'):
            p['_is_pile'] = True
            pile_products.append(p)
            continue
        if not _is_valid_product_code(code):
            continue
        qty = p.get('quantity', 0)
        if not qty or qty <= 0:
            continue
        matched_group = _match_exclude_group(p, price_mapping, all_exclude_opts)
        is_cs = _is_carbon_steel_cached(p, pi)
        if matched_group:
            if delete_options.get(matched_group):
                continue
            if ko_exclude_options.get(matched_group):
                excluded.append(p)
            else:
                if is_cs:
                    carbon_steel.append(p)
                else:
                    aluminum.append(p)
        elif is_cs:
            carbon_steel.append(p)
        else:
            aluminum.append(p)
    return aluminum, carbon_steel, excluded, pile_products, price_info_cache


def create_detail_sheet(workbook, array_info, bom_products, price_mapping,
                        sheet_prefix=None,
                        image_path=None, image_folder=None, code_to_images=None,
                        image_temp_dir=None, image_cache=None, matrix_data=None,
                        group=None, unmatched_products_out=None,
                        contact_info=None, config=None,
                        sale_type='export',
                        ko_discount_rate=100, ko_steel_discount_rate=84,
                        ko_purchased_discount_rate=94, coating_thickness=10,
                        delete_options=None, always_exclude_extra_items=False,
                        ko_exclude_options=None, need_weight_code=False,
                        need_total_qty=False,
                        lang='en',
                        discount_method='project',
                        **kwargs):
    all_products = list(bom_products) if bom_products else []

    aluminum, carbon_steel, excluded, pile_products, price_info_cache = _classify_products_single_pass(
        all_products, price_mapping, delete_options or {},
        always_exclude_extra_items, ko_exclude_options or {}
    )

    matrix_data = matrix_data or {}
    matrix_project_name = str(matrix_data.get('project_name') or '').strip()
    matrix_output_wp = matrix_data.get('output_wp') or 0
    matrix_max_wind = matrix_data.get('max_wind_speed') or ''
    matrix_max_snow = matrix_data.get('max_snow_load') or ''
    matrix_module_watt = matrix_data.get('module_wattage') or ''
    matrix_module_size = matrix_data.get('module_size') or ''
    matrix_set_count = matrix_data.get('set_count') or 1
    if not isinstance(matrix_set_count, int) or matrix_set_count <= 0:
        matrix_set_count = 1
    total_table_count = matrix_set_count
    config = config or {}
    angle_val = config.get('angle', '') or matrix_data.get('angle', '')
    rows = array_info.get('rows') if isinstance(array_info, dict) else None
    cols = array_info.get('cols') if isinstance(array_info, dict) else None
    span_info = config.get('cross_span', '') or kwargs.get('span_info', '')

    sheet_name = sheet_prefix or ''
    if not sheet_name:
        sheet_name = f'{rows}x{cols}' if rows and cols else 'Detail'
    sheet_name = extract_main_name(sheet_name)
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    original = sheet_name
    counter = 1
    while sheet_name in workbook.sheetnames:
        sheet_name = f"{original}_{counter}"
        counter += 1

    ws = workbook.create_sheet(title=sheet_name)
    print(f"[EN-SIMPLE] Creating detail sheet: {sheet_name}")

    thin = _EN_THIN
    thick = _EN_THICK
    thin_border = _EN_THIN_BORDER

    title_font = _EN_TITLE_FONT
    normal_font = _EN_NORMAL_FONT
    header_font = _EN_HEADER_FONT
    small_bold_font = _EN_SMALL_BOLD_FONT
    red_small_font = _EN_RED_SMALL_FONT
    if sale_type == 'domestic':
        currency_number_format = '#,##0.00'
        currency_label = 'RMB'
    elif sale_type == 'euro':
        currency_number_format = '#,##0.00'
        currency_label = 'EUR'
    else:
        currency_number_format = '#,##0.00'
        currency_label = 'US$'
    a7_fill = _EN_A7_FILL
    yellow_fill = _EN_YELLOW_FILL

    center_align = _EN_CENTER_ALIGN
    left_align = _EN_LEFT_ALIGN
    right_align = _EN_RIGHT_ALIGN

    col_qty = 7
    if need_total_qty:
        col_total_qty = 8
        col_total_price = 9
    else:
        col_total_qty = None
        col_total_price = 8
    if need_weight_code:
        col_weight = col_total_price + 1
    else:
        col_weight = None
    col_remark = col_total_price + (1 if need_weight_code else 0) + 1
    max_col = col_remark
    max_col_letter = get_column_letter(max_col)
    tpc_letter = get_column_letter(col_total_price)
    qty_formula_col = col_qty
    qty_formula_letter = get_column_letter(qty_formula_col)
    merge_end_col = col_total_price - 1
    merge_end_letter = get_column_letter(merge_end_col)

    _widths = [15.77, 28.67, 45, 30.03, 34.58, 24.58, 25.49]
    if need_total_qty:
        _widths.append(25.49)
    _widths.append(24.13)
    if need_weight_code:
        _widths.append(22.31)
    _widths.append(25.49)
    for _i, _w in enumerate(_widths):
        ws.column_dimensions[get_column_letter(_i + 1)].width = _w

    combined_products = aluminum + carbon_steel
    all_render_products = combined_products + excluded
    data_start_row = 9
    row_heights_temp = {1: 18, 2: 66, 3: 30, 4: 30, 5: 30, 6: 30, 7: 30, 8: 20, 9: 70.7}
    for row, height in row_heights_temp.items():
        ws.row_dimensions[row].height = height

    for col in range(1, max_col + 1):
        ws.cell(row=1, column=col)

    for _hdr_row in range(2, data_start_row):
        for _hdr_col in range(1, max_col + 1):
            ws.cell(row=_hdr_row, column=_hdr_col).border = thin_border

    ws.merge_cells(f'A2:{max_col_letter}2')
    ws['A2'] = _t('detail_title', lang)
    ws['A2'].font = title_font
    ws['A2'].alignment = center_align

    ws.merge_cells('A3:B7')
    if image_path and os.path.exists(image_path):
        try:
            img = XLImage(img=image_path)
            width_a = ws.column_dimensions['A'].width
            width_b = ws.column_dimensions['B'].width
            total_width_pixels = (width_a + width_b) * 7
            height_4 = ws.row_dimensions[4].height
            height_5 = ws.row_dimensions[5].height
            height_6 = ws.row_dimensions[6].height
            total_height_pts = height_4 + height_5 + height_6
            total_height_pixels = total_height_pts * 1.33
            target_width = total_width_pixels * 1.05
            target_height = total_height_pixels * 1.8
            original_width = img.width
            original_height = img.height
            original_ratio = original_width / original_height
            target_ratio = target_width / target_height
            if original_ratio > target_ratio:
                img.width = target_width
                img.height = target_width / original_ratio
            else:
                img.height = target_height
                img.width = target_height * original_ratio
            min_size = 50
            if img.width < min_size or img.height < min_size:
                img.width = max(img.width, min_size)
                img.height = max(img.height, min_size)
            ws.add_image(img, 'A4')
        except Exception:
            ws['A4'] = '[Image not found]'
            ws['A4'].font = normal_font
            ws['A4'].alignment = center_align
    else:
        ws['A3'] = ''
        ws['A3'].alignment = center_align

    contact_defaults = {
        'contact_name': '진설정',
        'phone': '0086-18050053693',
        'tel': 'seol@xmkseng.com',
        'fax': ''
    }
    contact_info = contact_info or {}
    contact_name = contact_info.get('contact_name') or contact_defaults['contact_name']
    phone = contact_info.get('phone') or contact_defaults['phone']
    tel = contact_info.get('tel') or contact_defaults['tel']

    ws['C3'] = _t('detail_sales', lang).format(contact_name)
    ws['C3'].font = normal_font
    ws['C3'].alignment = center_align
    ws['D3'] = _t('detail_installation_angle', lang)
    ws['D3'].font = normal_font
    ws['D3'].alignment = center_align
    ws['E3'] = normalize_angle(angle_val)
    ws['E3'].font = normal_font
    ws['E3'].alignment = center_align
    ws['F3'] = _t('detail_panel_size', lang)
    ws['F3'].font = normal_font
    ws['F3'].alignment = center_align

    ws['C4'] = _t('detail_phone', lang).format(phone)
    ws['C4'].font = normal_font
    ws['C4'].alignment = center_align
    ws['D4'] = _t('detail_max_wind_load', lang)
    ws['D4'].font = normal_font
    ws['D4'].alignment = center_align
    ws['E4'] = matrix_max_wind or ''
    ws['E4'].font = normal_font
    ws['E4'].alignment = center_align
    ws['F4'] = _t('detail_power_pc', lang)
    ws['F4'].font = normal_font
    ws['F4'].alignment = center_align
    ws['G4'] = f"{matrix_module_watt}w" if matrix_module_watt else ''
    ws['G4'].font = normal_font
    ws['G4'].alignment = center_align

    ws['C5'] = _t('detail_email', lang).format(tel)
    ws['C5'].font = normal_font
    ws['C5'].alignment = center_align
    ws['D5'] = _t('detail_max_snow_load', lang)
    ws['D5'].font = normal_font
    ws['D5'].alignment = center_align
    ws['E5'] = matrix_max_snow or ''
    ws['E5'].font = normal_font
    ws['E5'].alignment = center_align
    ws['F5'] = _t('detail_total_output', lang)
    ws['F5'].font = normal_font
    ws['F5'].alignment = center_align
    ws['G5'] = f"{matrix_output_wp} Wp" if matrix_output_wp else ''
    ws['G5'].font = normal_font
    ws['G5'].alignment = center_align

    missing_boards_val = _safe_int(config.get('missing_boards', 0) or 0)
    if missing_boards_val == 0:
        missing_boards_val = matrix_data.get('missing_per_table', 0) or 0
    ws['C6'] = _t('detail_missing_modules', lang)
    ws['C6'].font = normal_font
    ws['C6'].alignment = center_align
    ws['D6'] = str(missing_boards_val) if missing_boards_val != 0 else '/'
    ws['D6'].font = normal_font
    ws['D6'].alignment = center_align
    ws['F6'] = _t('detail_span_ew', lang)
    ws['F6'].font = normal_font
    ws['F6'].alignment = center_align
    ws['G6'] = f"{span_info} mm" if span_info else ''
    ws['G6'].font = normal_font
    ws['G6'].alignment = center_align

    if matrix_module_size:
        panel_spec = re.sub(r'[-×xX]', '*', str(matrix_module_size))
        if not panel_spec.lower().endswith('mm'):
            panel_spec = panel_spec + 'mm'
    else:
        panel_spec = config.get('panel_spec', '')
    ws['G3'] = panel_spec
    ws['G3'].font = normal_font
    ws['G3'].alignment = center_align

    ws['C7'] = _t('detail_array', lang)
    ws['C7'].font = normal_font
    ws['C7'].alignment = center_align
    if rows is not None:
        ws['D7'] = _t('detail_row', lang).format(rows)
    else:
        ws['D7'] = ''
    if cols is not None:
        ws['E7'] = _t('detail_col', lang).format(cols)
    else:
        ws['E7'] = ''
    ws['D7'].alignment = center_align
    ws['E7'].alignment = center_align
    ws['D7'].font = normal_font
    ws['E7'].font = normal_font
    ws['F7'] = _t('detail_tables', lang)
    ws['F7'].font = normal_font
    ws['F7'].alignment = center_align
    if matrix_set_count:
        ws['G7'] = matrix_set_count
    else:
        ws['G7'] = ''
    ws['G7'].font = normal_font
    ws['G7'].alignment = center_align

    ws.merge_cells(f'H3:{max_col_letter}7')
    ws['H3'] = _t('detail_warranty', lang)
    ws['H3'].font = normal_font
    ws['H3'].alignment = center_align

    ws.merge_cells(f'A8:{max_col_letter}8')

    row_product_map = {}
    row_price_info_map = {}
    matched_count = 0
    unmatched_count = 0
    image_found_count = 0
    image_not_found_count = 0
    local_unmatched = []

    def _col_width_to_px(char_width):
        return int(char_width * 7.5 + 5)

    def _row_height_to_px(pt_height):
        return int(pt_height * 1.33)

    def _fit_image_to_cell(cell_row, cell_col, max_w, max_h, padding=IMAGE_PADDING):
        col_w = ws.column_dimensions[get_column_letter(cell_col)].width or 8.43
        row_h = ws.row_dimensions[cell_row].height or 15
        avail_w = _col_width_to_px(col_w) - padding * 2
        avail_h = _row_height_to_px(row_h) - padding * 2
        avail_w = max(avail_w, 20)
        avail_h = max(avail_h, 20)
        scale = min(avail_w / max_w, avail_h / max_h, 1.0)
        return int(max_w * scale), int(max_h * scale)

    def _write_product_rows(start_row, products):
        nonlocal matched_count, unmatched_count, local_unmatched
        sub_total_price = Decimal('0')
        sub_total_weight = Decimal('0') if need_weight_code else None
        standard_rows = []
        steel_rows = []
        purchased_rows = []
        inquiry_rows = []
        unmatched_rows = []
        standard_price_sum = Decimal('0')
        steel_price_sum = Decimal('0')
        purchased_price_sum = Decimal('0')
        inquiry_price_sum = Decimal('0')
        is_complex = False
        for p in products:
            pi = p.get('_price_info')
            if pi is None:
                pi = resolve_price_info(price_mapping, p.get('code', ''), spec=p.get('spec', ''))
                p['_price_info'] = pi
            if not pi or not has_valid_price_info(pi):
                is_complex = True
                break
            cat = _get_discount_category(pi, p)
            if cat != 'standard':
                is_complex = True
                break
        for idx, product in enumerate(products):
            row = start_row + idx
            ws.row_dimensions[row].height = 61
            product_code = product['code']
            price_info = product.get('_price_info')
            if price_info is None:
                price_info = resolve_price_info(price_mapping, product_code, spec=product.get('spec', ''))
                product['_price_info'] = price_info
            en_name = _resolve_product_name(price_info, product, lang)
            cn_name = (
                price_info.get('name') or product.get('name', '')
            ) if price_info else product.get('name', '')
            if not price_info:
                _name_info = product.get('_price_info_no_spec') or resolve_price_info(price_mapping, product_code)
                if _name_info and not product.get('_price_info_no_spec'):
                    product['_price_info_no_spec'] = _name_info
                if _name_info:
                    en_name = _resolve_product_name(_name_info, product, lang)
                    cn_name = _name_info.get('name') or cn_name

            row_product_map[row] = product
            row_price_info_map[row] = price_info

            _set_cell(ws, row, 1, f'=ROW()-{start_row - 1}', font=normal_font, align=center_align, border=thin_border)
            _set_cell(ws, row, 2, en_name, font=normal_font, align=center_align, border=thin_border)
            _raw_mat = (price_info.get('db_material') if price_info and price_info.get('db_material') else None) or product.get('material', '')
            _set_cell(ws, row, 3, adjust_material_by_coating(translate_material(_raw_mat, 'en'), coating_thickness),
                      font=normal_font, align=center_align, border=thin_border)
            _set_cell(ws, row, 4, "", font=normal_font, align=center_align, border=thin_border)
            try:
                _spec_num = float(product['spec'])
                _set_cell(ws, row, 5, _spec_num,
                          font=normal_font, align=center_align, border=thin_border)
                ws.cell(row=row, column=5).number_format = '"L"!=#"mm"'
            except (ValueError, TypeError):
                _set_cell(ws, row, 5, _strip_cjk_spec(product['spec']),
                          font=normal_font, align=center_align, border=thin_border)

            unit_price = 0
            display_unit_price = 0
            price_unit = ''
            is_matched = False

            if price_info and has_valid_price_info(price_info):
                unit_price = float(price_info['price'])
                price_unit = price_info.get('unit', '')
                matched_count += 1
                is_matched = True
            else:
                unmatched_count += 1
                if unmatched_products_out is not None and _is_valid_product_code(product_code):
                    local_unmatched.append({
                        'code': product_code,
                        'name': cn_name,
                        'spec': product.get('spec', ''),
                        'quantity': product.get('quantity', 0),
                        'unit': price_info.get('unit', '') if price_info else '',
                        'issue_reason': 'No price match',
                    })

            is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
            length_mm = Decimal('0')
            if is_meter:
                length_mm = Decimal(str(extract_length_from_spec(product['spec']) or 0))
            if is_meter and length_mm > 0:
                display_unit_price = float(Decimal(str(unit_price)) * length_mm / Decimal('1000'))
            else:
                display_unit_price = unit_price

            _original_unit_price = display_unit_price
            if discount_method == 'unit_price' and display_unit_price > 0 and is_matched:
                _cat = _get_discount_category(price_info)
                if _cat == 'steel':
                    _rate = ko_steel_discount_rate
                elif _cat == 'purchased':
                    _rate = ko_purchased_discount_rate
                else:
                    _rate = ko_discount_rate
                display_unit_price = display_unit_price * _rate / 100

            if display_unit_price > 0:
                if discount_method == 'unit_price' and _original_unit_price > 0 and is_matched:
                    _set_cell(ws, row, 6, f'={round(_original_unit_price, 6)}*{_rate}/100',
                              font=normal_font, align=center_align, border=thin_border,
                              number_format=currency_number_format)
                else:
                    _set_cell(ws, row, 6, float(display_unit_price),
                              font=normal_font, align=center_align, border=thin_border,
                              number_format=currency_number_format)
            else:
                _set_cell(ws, row, 6, "",
                          font=normal_font, align=center_align, border=thin_border,
                          number_format=currency_number_format)

            per_table_qty = _safe_float(product['quantity'])
            total_qty = per_table_qty * total_table_count
            _qty_val = per_table_qty
            if _qty_val > 0:
                _set_cell(ws, row, col_qty, _safe_int(_qty_val) if _qty_val % 1 == 0 else _qty_val,
                          font=normal_font, align=center_align, border=thin_border)
            else:
                _set_cell(ws, row, col_qty, "",
                          font=normal_font, align=center_align, border=thin_border)

            if need_total_qty:
                if total_qty > 0:
                    _set_cell(ws, row, col_total_qty, _safe_int(total_qty) if total_qty % 1 == 0 else total_qty,
                              font=normal_font, align=center_align, border=thin_border)
                else:
                    _set_cell(ws, row, col_total_qty, "",
                              font=normal_font, align=center_align, border=thin_border)

            total_price = Decimal('0')
            _price_qty = per_table_qty
            if display_unit_price > 0 and _price_qty > 0:
                total_price = Decimal(str(display_unit_price)) * Decimal(str(_price_qty))
                _set_cell(ws, row, col_total_price, f"=F{row}*{qty_formula_letter}{row}",
                          font=normal_font, align=center_align, border=thin_border,
                          number_format=currency_number_format)
                sub_total_price += total_price
                cat = _get_discount_category(price_info)
                if cat == 'standard':
                    standard_rows.append(row)
                    standard_price_sum += total_price
                elif cat == 'steel':
                    steel_rows.append(row)
                    steel_price_sum += total_price
                elif cat == 'purchased':
                    purchased_rows.append(row)
                    purchased_price_sum += total_price
                else:
                    inquiry_rows.append(row)
                    inquiry_price_sum += total_price
            else:
                _set_cell(ws, row, col_total_price, f"=F{row}*{qty_formula_letter}{row}",
                          font=normal_font, align=center_align, border=thin_border,
                          number_format=currency_number_format)
                if not is_matched:
                    unmatched_rows.append(row)

            if need_weight_code:
                weight_cell = ws.cell(row=row, column=col_weight)
                unit_weight = None
                if price_info:
                    db_w = parse_decimal_number(price_info.get('db_weight'))
                    if db_w is not None and db_w > 0:
                        code_attribute = str(price_info.get('code_attribute') or '').strip().upper()
                        unit_weight = db_w
                        if code_attribute in WEIGHT_BY_LENGTH_ATTRIBUTES:
                            length_mm_w = extract_length_from_spec(product.get('spec'))
                            if length_mm_w and length_mm_w > 0:
                                unit_weight = db_w * Decimal(str(length_mm_w)) / Decimal('1000')
                            else:
                                unit_weight = None
                if unit_weight is None:
                    bom_w = product.get('weight', 0)
                    if bom_w and bom_w > 0:
                        unit_weight = Decimal(str(bom_w))
                if unit_weight is not None and unit_weight > 0:
                    weight_cell.value = float(round_to_2_decimal(unit_weight))
                    weight_cell.alignment = center_align
                    weight_cell.font = normal_font
                    weight_cell.border = thin_border
                    weight_cell.number_format = '#,##0.00'
                    sub_total_weight += round_to_2_decimal(unit_weight)
                else:
                    weight_cell.value = ""
                    weight_cell.alignment = center_align
                    weight_cell.font = normal_font
                    weight_cell.border = thin_border

            _set_cell(ws, row, col_remark, product_code,
                      font=normal_font, align=center_align, border=thin_border)

            if not is_matched:
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).fill = yellow_fill

        data_end = start_row + len(products) - 1 if products else start_row - 1
        _sw = float(sub_total_weight) if need_weight_code else 0.0
        return (data_end, float(sub_total_price), _sw,
                standard_rows, steel_rows, purchased_rows, inquiry_rows,
                unmatched_rows,
                float(standard_price_sum), float(steel_price_sum),
                float(purchased_price_sum), float(inquiry_price_sum),
                is_complex)

    def _write_group_header(row):
        group_headers = [
            (_t('hdr_item_no', lang), 'A'), (_t('hdr_product_name', lang), 'B'), (_t('hdr_material', lang), 'C'),
            (_t('hdr_picture', lang), 'D'), (_t('hdr_spec', lang), 'E'),
            (_t('hdr_unit_price_exw', lang).format(currency=currency_label), 'F'),
            (_t('hdr_qty_pcs', lang), 'G'),
        ]
        if need_total_qty:
            group_headers.append((_t('hdr_total_qty_pcs', lang), get_column_letter(col_total_qty)))
        group_headers.append((_t('hdr_total_price_exw', lang).format(currency=currency_label), tpc_letter))
        if need_weight_code:
            group_headers.append((_t('hdr_weight', lang), get_column_letter(col_weight)))
        group_headers.append((_t('hdr_remark', lang), get_column_letter(col_remark)))
        green_fill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
        for text, col in group_headers:
            cell = ws[f'{col}{row}']
            cell.value = text
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            if need_weight_code and col == get_column_letter(col_weight):
                cell.fill = green_fill
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = thin_border

    def _write_group_totals(row, sub_total, total_all, kw, sub_weight=0, total_weight=0, data_first_row=None, data_last_row=None):
        ws.row_dimensions[row].height = 40
        ws.merge_cells(f'A{row}:{merge_end_letter}{row}')
        ws[f'A{row}'] = _t('total_per_table', lang).format(total_table_count)
        ws[f'A{row}'].font = small_bold_font
        ws[f'A{row}'].alignment = right_align
        if data_first_row and data_last_row and data_last_row >= data_first_row:
            ws.cell(row=row, column=col_total_price, value=f'=SUM({tpc_letter}{data_first_row}:{tpc_letter}{data_last_row})')
        else:
            ws.cell(row=row, column=col_total_price, value=sub_total)
        ws.cell(row=row, column=col_total_price).number_format = currency_number_format
        ws.cell(row=row, column=col_total_price).font = small_bold_font
        ws.cell(row=row, column=col_total_price).alignment = center_align
        if need_weight_code:
            wc_letter = get_column_letter(col_weight)
            if data_first_row and data_last_row and data_last_row >= data_first_row:
                ws.cell(row=row, column=col_weight, value=f'=SUMPRODUCT({wc_letter}{data_first_row}:{wc_letter}{data_last_row},{qty_formula_letter}{data_first_row}:{qty_formula_letter}{data_last_row})')
            else:
                ws.cell(row=row, column=col_weight, value=round(sub_weight, 2))
            ws.cell(row=row, column=col_weight).number_format = '#,##0.00'
            ws.cell(row=row, column=col_weight).font = small_bold_font
            ws.cell(row=row, column=col_weight).alignment = center_align
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = thin_border

        sub_total_row = row
        row += 1
        ws.row_dimensions[row].height = 40
        ws.merge_cells(f'A{row}:{merge_end_letter}{row}')
        total_label = _t('total_per_table', lang).format(total_table_count)
        ws[f'A{row}'] = total_label
        ws[f'A{row}'].font = small_bold_font
        ws[f'A{row}'].alignment = right_align
        ws.cell(row=row, column=col_total_price, value=f'={tpc_letter}{sub_total_row}*G7')
        ws.cell(row=row, column=col_total_price).number_format = currency_number_format
        ws.cell(row=row, column=col_total_price).font = small_bold_font
        ws.cell(row=row, column=col_total_price).alignment = center_align
        if need_weight_code:
            wc_letter = get_column_letter(col_weight)
            ws.cell(row=row, column=col_weight, value=f'={wc_letter}{sub_total_row}*G7')
            ws.cell(row=row, column=col_weight).number_format = '#,##0.00'
            ws.cell(row=row, column=col_weight).font = small_bold_font
            ws.cell(row=row, column=col_weight).alignment = center_align
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = thin_border

        all_tables_row = row
        row += 1
        ws.row_dimensions[row].height = 40
        ws.merge_cells(f'A{row}:{merge_end_letter}{row}')
        ws[f'A{row}'] = _t('total_per_watt', lang).format(currency=currency_label)
        ws[f'A{row}'].font = small_bold_font
        ws[f'A{row}'].alignment = right_align
        if kw and kw > 0:
            ws.cell(row=row, column=col_total_price, value=f'={tpc_letter}{all_tables_row}/({kw}*1000)')
        else:
            ws.cell(row=row, column=col_total_price, value=0)
        ws.cell(row=row, column=col_total_price).number_format = '#,##0.000'
        ws.cell(row=row, column=col_total_price).font = small_bold_font
        ws.cell(row=row, column=col_total_price).alignment = center_align
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = thin_border

        return row

    multiplier = Decimal(str(total_table_count)) if total_table_count > 1 else Decimal('1')
    output_kw = matrix_data.get('output_kw') or 0
    current_row = data_start_row

    part1_sub_total = 0.0
    part1_total_all = 0.0
    part1_sub_weight = 0.0
    part1_total_weight = 0.0
    part1_total_table_row = 0
    part1_std_rows = []
    part1_steel_rows = []
    part1_purchased_rows = []
    part1_inq_rows = []
    part1_unmatched_rows = []
    part1_std_price = 0.0
    part1_steel_price = 0.0
    part1_purchased_price = 0.0
    part1_inq_price = 0.0
    is_complex = False
    if combined_products:
        _write_group_header(current_row)
        current_row += 1
        part1_data_start = current_row
        (data_end, part1_sub_total, part1_sub_weight,
         part1_std_rows, part1_steel_rows, part1_purchased_rows, part1_inq_rows,
         part1_unmatched_rows,
         part1_std_price, part1_steel_price, part1_purchased_price, part1_inq_price,
         is_complex) = _write_product_rows(current_row, combined_products)
        part1_data_end = data_end
        current_row = data_end + 1
        part1_total_all = part1_sub_total
        part1_total_weight = part1_sub_weight
        part1_total_table_row = current_row
        current_row = _write_group_totals(current_row, part1_sub_total, part1_total_all, output_kw, part1_sub_weight, part1_total_weight, data_first_row=part1_data_start, data_last_row=part1_data_end) + 1

    part3_sub_total = 0.0
    part3_total_all = 0.0
    part3_sub_weight = 0.0
    part3_total_weight = 0.0
    if excluded:
        ws.merge_cells(f'A{current_row}:{max_col_letter}{current_row}')
        ws[f'A{current_row}'] = _t('excluded_notice', lang)
        ws[f'A{current_row}'].font = red_small_font
        ws[f'A{current_row}'].alignment = left_align
        for c in range(1, max_col + 1):
            ws.cell(row=current_row, column=c).border = thin_border
        current_row += 1
        ws.merge_cells(f'A{current_row}:{max_col_letter}{current_row}')
        ws[f'A{current_row}'] = _t('excluded_title', lang)
        ws[f'A{current_row}'].font = small_bold_font
        ws[f'A{current_row}'].alignment = left_align
        ws.row_dimensions[current_row].height = 40
        for c in range(1, max_col + 1):
            ws.cell(row=current_row, column=c).border = thin_border
        current_row += 1
        _write_group_header(current_row)
        current_row += 1
        part3_data_start = current_row
        data_end, part3_sub_total, part3_sub_weight, _, _, _, _, _, _, _, _, _, _ = _write_product_rows(current_row, excluded)
        part3_data_end = data_end
        current_row = data_end + 1
        part3_total_all = part3_sub_total
        part3_total_weight = part3_sub_weight
        current_row = _write_group_totals(current_row, part3_sub_total, part3_total_all, output_kw, part3_sub_weight, part3_total_weight, data_first_row=part3_data_start, data_last_row=part3_data_end) + 1

    last_data_row = current_row - 1

    def mark_missing_image(row_number):
        image_cell = ws.cell(row=row_number, column=IMAGE_COLUMN_INDEX)
        image_cell.value = "/"
        image_cell.alignment = center_align
        image_cell.font = normal_font

    for row in range(data_start_row, last_data_row + 1):
        if row not in row_product_map:
            continue
        product_code = ws.cell(row=row, column=max_col).value
        if product_code is None or str(product_code).strip() == '':
            continue
        product_code = str(product_code).strip()
        normalized_code = normalize_lookup_code(product_code)

        img_path = None
        if code_to_images and product_code in code_to_images and code_to_images[product_code]:
            img_path = code_to_images[product_code][0]
        elif code_to_images and normalized_code in code_to_images and code_to_images[normalized_code]:
            img_path = code_to_images[normalized_code][0]

        if img_path and image_folder:
            fit_w, fit_h = _fit_image_to_cell(row, IMAGE_COLUMN_INDEX, IMAGE_WIDTH, IMAGE_HEIGHT)
            temp_img_path = prepare_image_for_excel(
                img_path, target_width=fit_w, target_height=fit_h,
                temp_dir=image_temp_dir, cache=image_cache
            )
            final_img_path = temp_img_path if temp_img_path else img_path
            success = add_image_centered_in_cell(
                ws, final_img_path, row, IMAGE_COLUMN_INDEX,
                img_width=fit_w, img_height=fit_h
            )
            if success:
                image_found_count += 1
                continue
            image_not_found_count += 1
            mark_missing_image(row)
        else:
            image_not_found_count += 1
            mark_missing_image(row)

    print(f"   [EN-SIMPLE] Images: found={image_found_count}, not_found={image_not_found_count}")

    grand_total = part1_total_all + part3_total_all

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(top=0.75, bottom=0.75, left=0.7, right=0.7, header=0.3, footer=0.3)

    if unmatched_products_out is not None:
        for _up in local_unmatched:
            _up['quantity'] = float(_up.get('quantity', 0)) * total_table_count
        unmatched_products_out.extend(local_unmatched)

    quotation_product_codes = set()
    for p in combined_products + excluded:
        c = str(p.get('code', '')).strip()
        if c:
            quotation_product_codes.add(c)

    return {
        'sheet_name': sheet_name,
        'quotation_product_codes': quotation_product_codes,
        'valid_products': len(combined_products) + len(excluded),
        'total_weight': 0,
        'total_price': grand_total,
        'part1_price_per_table': part1_sub_total,
        'part1_price_all_tables': part1_total_all,
        'part1_total_table_row': part1_total_table_row,
        'part1_standard_rows': part1_std_rows,
        'part1_steel_rows': part1_steel_rows,
        'part1_purchased_rows': part1_purchased_rows,
        'part1_inquiry_rows': part1_inq_rows,
        'part1_unmatched_rows': part1_unmatched_rows,
        'part1_standard_total': part1_std_price,
        'part1_steel_total': part1_steel_price,
        'part1_purchased_total': part1_purchased_price,
        'part1_inquiry_total': part1_inq_price,
        'part2_price_per_table': 0,
        'part2_price_all_tables': 0,
        'part2_total_table_row': 0,
        'part2_standard_rows': [],
        'part2_steel_rows': [],
        'part2_purchased_rows': [],
        'part2_inquiry_rows': [],
        'part2_unmatched_rows': [],
        'is_complex': is_complex,
        'part2_standard_total': 0,
        'part2_steel_total': 0,
        'part2_purchased_total': 0,
        'part2_inquiry_total': 0,
        'matched_count': matched_count,
        'unmatched_count': unmatched_count,
        'unmatched_products': local_unmatched,
        'image_found_count': image_found_count,
        'image_not_found_count': image_not_found_count,
        'pile_products': pile_products,
        'all_render_products': all_render_products,
        'sub_total_row': last_data_row,
        'total_row': last_data_row,
        'detail_data_end_row': last_data_row,
        'config': config,
        'matrix_data': matrix_data,
        'detail_total_price_col_letter': tpc_letter,
        'discount_method': discount_method,
    }


def create_summary_sheet(
        workbook, all_quotation_results, matrix_data=None, image_path=None,
        ko_discount_rate=100, ko_steel_discount_rate=84,
        ko_purchased_discount_rate=94, sale_type='export',
        contact_info=None, trade_method='EXW',
        ko_ddp_address='',
        pile_summary=None,
        dest_port='BUSAN', container_type='40HQ',
        container_qty=1, ko_cif_freight=0,
        skip_freight=True, quote_validity='7d', lang='en', discount_method='project',
        payment_term='3070shipment', **kwargs):
    from datetime import datetime

    _validity_map = {
        '1d': _t('validity_1d', lang),
        '7d': _t('validity_7d', lang),
        'today': _t('validity_today', lang),
    }
    validity_text = _validity_map.get(quote_validity, None)
    if validity_text is None:
        m = re.match(r'^(\d+)d$', str(quote_validity))
        if m:
            n = int(m.group(1))
            validity_text = f'{n} days'
        else:
            validity_text = _t('validity_7d', lang)
    _payment_map = {
        '100advance': _t('payment_100advance', lang),
        '3070shipment': _t('payment_3070shipment', lang),
        '3070bl': _t('payment_3070bl', lang),
    }
    payment_text = _payment_map.get(payment_term, _t('payment_3070shipment', lang))

    if sale_type == 'domestic':
        ksd_currency_fmt = '¥#,##0.00'
        ksd_currency_label = 'RMB'
    elif sale_type == 'euro':
        ksd_currency_fmt = '€#,##0.00'
        ksd_currency_label = 'EUR'
    else:
        ksd_currency_fmt = '$#,##0.00'
        ksd_currency_label = 'USD'
    matrix_data = matrix_data or {}
    project_name = str(matrix_data.get('project_name') or '').strip()
    output_kw = matrix_data.get('output_kw') or 0

    ws = workbook.create_sheet(title='Total')

    normal_font = _EN_SUMMARY_NORMAL_FONT
    bold_font = _EN_SUMMARY_BOLD_FONT
    title_font = _EN_SUMMARY_TITLE_FONT
    company_font = _EN_SUMMARY_COMPANY_FONT
    part_title_font = _EN_SUMMARY_PART_TITLE_FONT
    part_inner_font = _EN_SUMMARY_PART_INNER_FONT
    part_inner_bold_font = _EN_SUMMARY_PART_INNER_BOLD_FONT
    summary_amount_font = _EN_SUMMARY_AMOUNT_FONT
    center = _EN_SUMMARY_CENTER
    left_bottom = _EN_SUMMARY_LEFT_BOTTOM
    right_a = _EN_SUMMARY_RIGHT_A
    thin_border = _EN_THIN_BORDER
    bottom_border = _EN_SUMMARY_BOTTOM_BORDER
    yellow_fill = _EN_SUMMARY_YELLOW_FILL

    for col_letter in 'ABCDEFGHIJKLM':
        if col_letter in ('A', 'B', 'C'):
            ws.column_dimensions[col_letter].width = 11
        elif col_letter in ('D', 'E', 'F'):
            ws.column_dimensions[col_letter].width = 15
        else:
            ws.column_dimensions[col_letter].width = 8

    _qno = matrix_data.get('quotation_no', '') if matrix_data else ''
    _info_font = Font(name='Arial', size=11)
    ws.merge_cells('H1:M1')
    ws['H1'] = f'Quotation NO.:{_qno}' if _qno else 'Quotation NO.:'
    ws['H1'].font = _info_font
    ws['H1'].alignment = right_a

    ws.merge_cells('H2:M2')
    ws['H2'] = f'DATE:{datetime.now().strftime("%Y/%m/%d")}'
    ws['H2'].font = _info_font
    ws['H2'].alignment = right_a
    ws['J2'].font = _info_font
    ws['J2'].alignment = center

    ws.merge_cells('A3:M3')
    ws['A3'] = _t('summary_title', lang)
    ws['A3'].font = title_font
    ws['A3'].alignment = center
    ws.row_dimensions[3].height = 30

    if image_path and os.path.exists(image_path):
        try:
            img = XLImage(img=image_path)
            img.width = 280
            img.height = 70
            ws.add_image(img, 'H4')
        except Exception:
            pass

    ws.merge_cells('A6:B6')
    ws['A6'] = _t('summary_project', lang)
    ws['A6'].font = normal_font
    ws['A6'].alignment = left_bottom
    project_name_font = Font(name='Arial', size=11)
    ws.merge_cells('C6:F6')
    ws['C6'] = project_name
    ws['C6'].font = project_name_font
    ws['C6'].alignment = left_bottom

    ws.merge_cells('A7:B7')
    ws['A7'] = _t('summary_price_term', lang)
    ws['A7'].font = normal_font
    ws['A7'].alignment = left_bottom
    ws.merge_cells('C7:F7')
    ws['C7'] = trade_method
    ws['C7'].font = Font(name='Arial', size=11)
    ws['C7'].alignment = left_bottom

    ws.merge_cells('A8:B8')
    ws['A8'] = _t('summary_delivery_time', lang)
    ws['A8'].font = normal_font
    ws['A8'].alignment = left_bottom
    ws.merge_cells('C8:F8')
    ws['C8'] = _t('summary_delivery_value', lang)
    ws['C8'].font = normal_font
    ws['C8'].alignment = left_bottom
    ws.merge_cells('H8:M8')
    ws['H8'] = 'Xiamen Kseng Metal Tech Co.,Ltd'
    ws['H8'].font = Font(name='Arial', size=15, bold=True)
    ws['H8'].alignment = left_bottom

    ws.merge_cells('A9:B9')
    ws['A9'] = _t('summary_payment_term', lang)
    ws['A9'].font = normal_font
    ws['A9'].alignment = left_bottom
    ws.merge_cells('C9:F9')
    ws['C9'] = payment_text
    ws['C9'].font = normal_font
    ws['C9'].alignment = left_bottom
    ws.merge_cells('H9:M9')
    ws['H9'] = 'Add.: RM 302, Huixin Wealth Centre, No. 891, '
    ws['H9'].font = normal_font
    ws['H9'].alignment = left_bottom

    ws.merge_cells('A10:B10')
    ws['A10'] = _t('summary_validity_date', lang)
    ws['A10'].font = normal_font
    ws['A10'].alignment = left_bottom
    ws.merge_cells('C10:F10')
    ws['C10'] = validity_text
    ws['C10'].font = normal_font
    ws['C10'].alignment = left_bottom
    ws.merge_cells('H10:M10')
    ws['H10'] = 'Fanghu North 2nd Rd, Huli Dist, Xiamen, Fujian, China'
    ws['H10'].font = normal_font
    ws['H10'].alignment = left_bottom

    part1_total_all = Decimal('0')
    part1_per_table = Decimal('0')
    part1_std_total = Decimal('0')
    part1_steel_total = Decimal('0')
    part1_purchased_total = Decimal('0')
    part1_inq_total = Decimal('0')
    total_output_kw = Decimal('0')
    for qr in all_quotation_results:
        part1_total_all += Decimal(str(qr.get('part1_price_all_tables', 0)))
        part1_per_table += Decimal(str(qr.get('part1_price_per_table', 0)))
        part1_std_total += Decimal(str(qr.get('part1_standard_total', 0)))
        part1_steel_total += Decimal(str(qr.get('part1_steel_total', 0)))
        part1_purchased_total += Decimal(str(qr.get('part1_purchased_total', 0)))
        part1_inq_total += Decimal(str(qr.get('part1_inquiry_total', 0)))
        md = qr.get('matrix_data') or matrix_data
        kw = md.get('output_kw') or 0
        try:
            total_output_kw += Decimal(str(kw))
        except (ArithmeticError, ValueError):
            pass

    if discount_method == 'unit_price':
        grand_total = float(part1_std_total + part1_steel_total + part1_purchased_total + part1_inq_total)
    else:
        grand_total = (float(part1_std_total) * ko_discount_rate / 100.0
                       + float(part1_steel_total) * ko_steel_discount_rate / 100.0
                       + float(part1_purchased_total) * ko_purchased_discount_rate / 100.0
                       + float(part1_inq_total))
    unit_per_w = round(grand_total / (float(total_output_kw) * 1000), 7) if total_output_kw > 0 else 0

    for r in range(6, 13):
        ws.row_dimensions[r].height = 20
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = bottom_border

    part1_start = 14
    ws.merge_cells(f'A{part1_start}:M{part1_start}')
    title_cell = ws.cell(row=part1_start, column=1, value=_t('summary_part1_title', lang))
    title_cell.font = part_title_font
    title_cell.alignment = left_bottom
    title_cell.border = thin_border
    for ci in range(2, 14):
        ws.cell(row=part1_start, column=ci).border = thin_border
    ws.row_dimensions[part1_start].height = 30

    hdr_row = part1_start + 1
    part1_headers = [_t('summary_hdr_no', lang), _t('summary_hdr_row', lang), _t('summary_hdr_column', lang), _t('summary_hdr_angle', lang), '', _t('summary_hdr_table', lang), _t('summary_hdr_power', lang), '', _t('summary_hdr_price_per_table', lang).format(currency=ksd_currency_label), '', '', _t('summary_hdr_amount', lang).format(currency=ksd_currency_label), '']
    for ci, h in enumerate(part1_headers):
        cell = ws.cell(row=hdr_row, column=ci + 1, value=h)
        cell.font = part_inner_bold_font
        cell.alignment = center
        cell.border = thin_border
    ws.merge_cells(f'D{hdr_row}:E{hdr_row}')
    ws.merge_cells(f'G{hdr_row}:H{hdr_row}')
    ws.merge_cells(f'I{hdr_row}:K{hdr_row}')
    ws.merge_cells(f'L{hdr_row}:M{hdr_row}')

    data_row = hdr_row + 1
    for i, qr in enumerate(all_quotation_results):
        r = data_row + i
        cfg = qr.get('config') or {}
        md = qr.get('matrix_data') or matrix_data
        ma = qr.get('matched_array') or {}
        r_rows = ma.get('rows', md.get('array_rows'))
        r_cols = ma.get('cols', md.get('array_cols'))
        r_set = _safe_int(md.get('set_count', 1), 1)
        angle_val = cfg.get('angle', '') or md.get('angle', '')
        angle_display = normalize_angle(angle_val)
        r_kw = md.get('output_kw') or 0
        detail_sheet = qr.get('sheet_name', '')
        p1_total_table_row = qr.get('part1_total_table_row', 0)
        _tpc = qr.get('detail_total_price_col_letter', 'H')

        ws.cell(row=r, column=1, value=i + 1).alignment = center
        ws.cell(row=r, column=2, value=r_rows or '').alignment = center
        ws.cell(row=r, column=3, value=r_cols or '').alignment = center
        ws.cell(row=r, column=4, value=angle_display).alignment = center
        ws.merge_cells(f'D{r}:E{r}')
        ws.cell(row=r, column=6, value=r_set or '').alignment = center
        ws.cell(row=r, column=7, value=r_kw).alignment = center
        ws.merge_cells(f'G{r}:H{r}')
        if detail_sheet and p1_total_table_row:
            ws.cell(row=r, column=9, value=f"='{detail_sheet}'!{_tpc}{p1_total_table_row}").number_format = ksd_currency_fmt
        else:
            ws.cell(row=r, column=9, value=qr.get('part1_price_per_table', 0)).number_format = ksd_currency_fmt
        ws.cell(row=r, column=9).alignment = center
        ws.merge_cells(f'I{r}:K{r}')
        ws.cell(row=r, column=12, value=f'=I{r}').number_format = ksd_currency_fmt
        ws.cell(row=r, column=12).alignment = center
        ws.merge_cells(f'L{r}:M{r}')
        for ci in range(1, 13):
            ws.cell(row=r, column=ci).font = part_inner_font
            ws.cell(row=r, column=ci).border = thin_border
        if not qr.get('part1_price_per_table', 0):
            for ci in range(1, 13):
                ws.cell(row=r, column=ci).fill = yellow_fill
        ws.row_dimensions[r].height = 25

    data_end = data_row + len(all_quotation_results) - 1

    total_row = data_end + 1
    ws.merge_cells(f'A{total_row}:F{total_row}')
    ws.cell(row=total_row, column=1, value=_t('summary_total_label', lang)).font = part_inner_bold_font
    ws.cell(row=total_row, column=1).alignment = center
    ws.merge_cells(f'G{total_row}:H{total_row}')
    ws.cell(row=total_row, column=7, value=f'=SUM(G{data_row}:G{data_end})').alignment = center
    ws.cell(row=total_row, column=7).font = part_inner_font
    ws.merge_cells(f'I{total_row}:K{total_row}')
    ws.merge_cells(f'L{total_row}:M{total_row}')
    ws.cell(row=total_row, column=12, value=f'=SUM(L{data_row}:L{data_end})').number_format = ksd_currency_fmt
    ws.cell(row=total_row, column=12).alignment = center
    ws.cell(row=total_row, column=12).font = part_inner_font
    for ci in range(1, 13):
        ws.cell(row=total_row, column=ci).border = thin_border
    ws.row_dimensions[total_row].height = 25

    has_steel = any(qr.get('part1_steel_rows') for qr in all_quotation_results)
    has_purchased = any(qr.get('part1_purchased_rows') for qr in all_quotation_results)
    if discount_method == 'unit_price':
        discount_formula = f'=L{total_row}'
    elif ko_discount_rate == ko_steel_discount_rate == ko_purchased_discount_rate or (not has_steel and not has_purchased):
        discount_formula = f'=L{total_row}*{ko_discount_rate}/100'
    else:
        discount_parts = []
        for qr in all_quotation_results:
            sn = qr.get('sheet_name', '')
            sn_esc = sn.replace("'", "''")
            md = qr.get('matrix_data') or matrix_data
            sc = md.get('set_count') or qr.get('set_count') or 1
            sc = _safe_int(sc, 1)
            for rows_key, rate in [
                ('part1_standard_rows', ko_discount_rate),
                ('part1_steel_rows', ko_steel_discount_rate),
                ('part1_purchased_rows', ko_purchased_discount_rate),
                ('part1_inquiry_rows', 100),
            ]:
                rows = qr.get(rows_key, [])
                if not rows:
                    continue
                refs = ','.join(f"'{sn_esc}'!{qr.get('detail_total_price_col_letter', 'H')}{r}" for r in rows)
                if rate == 100:
                    discount_parts.append(f"SUM({refs})")
                else:
                    discount_parts.append(f"SUM({refs})*{rate}/100")
        if discount_parts:
            discount_formula = '=' + '+'.join(discount_parts)
            if len(discount_formula) > 8000:
                discount_formula = grand_total
        else:
            discount_formula = grand_total

    disc_p1_row = total_row + 1
    ws.merge_cells(f'A{disc_p1_row}:K{disc_p1_row}')
    ws.cell(row=disc_p1_row, column=1, value=f'Discount Price ({ksd_currency_label}):').font = summary_amount_font
    ws.cell(row=disc_p1_row, column=1).alignment = right_a
    ws.merge_cells(f'L{disc_p1_row}:M{disc_p1_row}')
    ws.cell(row=disc_p1_row, column=12, value=discount_formula).number_format = ksd_currency_fmt
    ws.cell(row=disc_p1_row, column=12).font = summary_amount_font
    ws.cell(row=disc_p1_row, column=12).alignment = center
    for ci in range(1, 14):
        ws.cell(row=disc_p1_row, column=ci).border = thin_border
    ws.row_dimensions[disc_p1_row].height = 25

    perw_row = disc_p1_row + 1
    ws.merge_cells(f'A{perw_row}:K{perw_row}')
    ws.cell(row=perw_row, column=1, value='Unite Price/W:').font = summary_amount_font
    ws.cell(row=perw_row, column=1).alignment = right_a
    ws.merge_cells(f'L{perw_row}:M{perw_row}')
    ws.cell(row=perw_row, column=12, value=f'=ROUND(IF(G{total_row}>0,L{disc_p1_row}/(G{total_row}*1000),0),3)').number_format = '$#,##0.000'
    ws.cell(row=perw_row, column=12).font = summary_amount_font
    ws.cell(row=perw_row, column=12).alignment = center
    for ci in range(1, 14):
        ws.cell(row=perw_row, column=ci).border = thin_border
    ws.row_dimensions[perw_row].height = 25

    if skip_freight:
        freight_cif_row = disc_p1_row
        freight_end_row = perw_row
    else:
        freight_title_row = perw_row + 1
        ws.merge_cells(f'A{freight_title_row}:M{freight_title_row}')
        ft_cell = ws.cell(row=freight_title_row, column=1, value=_t('summary_part2_title', lang).format(trade_method))
        ft_cell.font = part_title_font
        ft_cell.alignment = left_bottom
        ft_cell.border = thin_border
        for ci in range(2, 14):
            ws.cell(row=freight_title_row, column=ci).border = thin_border
        ws.row_dimensions[freight_title_row].height = 30

        freight_hdr_row = freight_title_row + 1
        freight_headers = [_t('summary_hdr_no', lang), _t('summary_freight_hdr_port', lang), '', _t('summary_freight_hdr_container', lang), '', '', _t('summary_freight_hdr_quantity', lang), '', _t('summary_freight_hdr_unit_price', lang), '', '', _t('summary_freight_hdr_amount_usd', lang), '']
        for ci, h in enumerate(freight_headers):
            cell = ws.cell(row=freight_hdr_row, column=ci + 1, value=h)
            cell.font = part_inner_bold_font
            cell.alignment = center
            cell.border = thin_border
        ws.merge_cells(f'B{freight_hdr_row}:C{freight_hdr_row}')
        ws.merge_cells(f'D{freight_hdr_row}:F{freight_hdr_row}')
        ws.merge_cells(f'G{freight_hdr_row}:H{freight_hdr_row}')
        ws.merge_cells(f'I{freight_hdr_row}:K{freight_hdr_row}')
        ws.merge_cells(f'L{freight_hdr_row}:M{freight_hdr_row}')
        ws.row_dimensions[freight_hdr_row].height = 25

        _container_details = kwargs.get('container_details') or []
        freight_amount_cells = []

        if _container_details:
            dr = freight_hdr_row + 1
            seq = 1
            for cd in _container_details:
                ct = str(cd.get('type', '')).upper()
                cq = _safe_float(cd.get('qty', 0))
                cpu = cd.get('freight_per_unit', 0)
                camt = cd.get('amount', 0) if cd.get('amount') else cpu * cq

                ws.cell(row=dr, column=1, value=seq).alignment = center
                ws.cell(row=dr, column=2, value=dest_port).alignment = center
                ws.merge_cells(f'B{dr}:C{dr}')
                ws.cell(row=dr, column=4, value=ct).alignment = center
                ws.merge_cells(f'D{dr}:F{dr}')
                ws.cell(row=dr, column=7, value=_safe_int(cq) if isinstance(cq, float) and cq % 1 == 0 else cq).alignment = center
                ws.merge_cells(f'G{dr}:H{dr}')
                ws.cell(row=dr, column=9, value=float(cpu) if cpu else 0).alignment = center
                ws.cell(row=dr, column=9).number_format = ksd_currency_fmt
                ws.merge_cells(f'I{dr}:K{dr}')
                ws.cell(row=dr, column=12, value=float(camt) if camt else 0).alignment = center
                ws.cell(row=dr, column=12).number_format = ksd_currency_fmt
                ws.merge_cells(f'L{dr}:M{dr}')
                for ci in range(1, 14):
                    ws.cell(row=dr, column=ci).font = part_inner_font
                    ws.cell(row=dr, column=ci).border = thin_border
                freight_amount_cells.append(f'L{dr}')
                ws.row_dimensions[dr].height = 25
                seq += 1
                dr += 1
            freight_data_end = dr - 1
        else:
            freight_data_row = freight_hdr_row + 1
            ws.cell(row=freight_data_row, column=1, value=1).alignment = center
            ws.cell(row=freight_data_row, column=2, value=dest_port).alignment = center
            ws.merge_cells(f'B{freight_data_row}:C{freight_data_row}')
            ws.cell(row=freight_data_row, column=4, value=container_type).alignment = center
            ws.merge_cells(f'D{freight_data_row}:F{freight_data_row}')
            ws.cell(row=freight_data_row, column=7, value=1).alignment = center
            ws.merge_cells(f'G{freight_data_row}:H{freight_data_row}')
            ws.cell(row=freight_data_row, column=9, value=ko_cif_freight).alignment = center
            ws.cell(row=freight_data_row, column=9).number_format = ksd_currency_fmt
            ws.merge_cells(f'I{freight_data_row}:K{freight_data_row}')
            ws.cell(row=freight_data_row, column=12, value=f'=I{freight_data_row}*G{freight_data_row}').alignment = center
            ws.cell(row=freight_data_row, column=12).number_format = ksd_currency_fmt
            ws.merge_cells(f'L{freight_data_row}:M{freight_data_row}')
            for ci in range(1, 14):
                ws.cell(row=freight_data_row, column=ci).font = part_inner_font
                ws.cell(row=freight_data_row, column=ci).border = thin_border
            freight_amount_cells.append(f'L{freight_data_row}')
            ws.row_dimensions[freight_data_row].height = 25
            freight_data_end = freight_data_row

        freight_subtotal_row = freight_data_end + 1
        ws.merge_cells(f'A{freight_subtotal_row}:K{freight_subtotal_row}')
        ws.cell(row=freight_subtotal_row, column=1, value=f'{trade_method} {_t("common_freight_subtotal", lang)}').font = summary_amount_font
        ws.cell(row=freight_subtotal_row, column=1).alignment = right_a
        ws.merge_cells(f'L{freight_subtotal_row}:M{freight_subtotal_row}')
        if freight_amount_cells:
            ws.cell(row=freight_subtotal_row, column=12, value=f'={"+".join(freight_amount_cells)}').number_format = ksd_currency_fmt
        else:
            ws.cell(row=freight_subtotal_row, column=12, value=0).number_format = ksd_currency_fmt
        ws.cell(row=freight_subtotal_row, column=12).font = summary_amount_font
        ws.cell(row=freight_subtotal_row, column=12).alignment = center
        for ci in range(1, 14):
            ws.cell(row=freight_subtotal_row, column=ci).border = thin_border
        ws.row_dimensions[freight_subtotal_row].height = 25

        freight_cif_row = freight_subtotal_row + 1
        ws.merge_cells(f'A{freight_cif_row}:K{freight_cif_row}')
        ws.cell(row=freight_cif_row, column=1, value=f'TOTAL AMOUNT:').font = summary_amount_font
        ws.cell(row=freight_cif_row, column=1).alignment = right_a
        ws.merge_cells(f'L{freight_cif_row}:M{freight_cif_row}')
        ws.cell(row=freight_cif_row, column=12, value=f'=L{disc_p1_row}+L{freight_subtotal_row}').number_format = ksd_currency_fmt
        ws.cell(row=freight_cif_row, column=12).font = summary_amount_font
        ws.cell(row=freight_cif_row, column=12).alignment = center
        for ci in range(1, 14):
            ws.cell(row=freight_cif_row, column=ci).border = thin_border
        ws.row_dimensions[freight_cif_row].height = 25

        freight_end_row = freight_cif_row

    ws.merge_cells('A11:B11')
    ws['A11'] = 'TOTAL AMOUNT:'
    ws['A11'].font = summary_amount_font
    ws['A11'].alignment = left_bottom
    ws.merge_cells('C11:F11')
    ws.cell(row=11, column=3, value=f'=L{freight_cif_row}').number_format = ksd_currency_fmt
    ws.cell(row=11, column=3).font = summary_amount_font
    ws.cell(row=11, column=3).alignment = left_bottom
    ws.merge_cells('H11:M11')
    phone = (contact_info or {}).get('phone', '0086-18050053693')
    email = (contact_info or {}).get('tel', 'using@xmkseng.com')
    ws['H11'] = _t('detail_phone', lang).format(phone)
    ws['H11'].font = normal_font
    ws['H11'].alignment = left_bottom
    ws.merge_cells('A12:B12')
    ws['A12'] = 'Unite Price/W:'
    ws['A12'].font = summary_amount_font
    ws['A12'].alignment = left_bottom
    ws.merge_cells('C12:F12')
    ws.cell(row=12, column=3, value=f'=ROUND(IF(G{total_row}>0,L{freight_cif_row}/(G{total_row}*1000),0),3)').number_format = '$#,##0.000'
    ws.cell(row=12, column=3).font = summary_amount_font
    ws.cell(row=12, column=3).alignment = left_bottom
    ws.merge_cells('H12:M12')
    ws['H12'] = _t('detail_email', lang).format(email)
    ws['H12'].font = normal_font
    ws['H12'].alignment = left_bottom

    outer_start_row = part1_start
    if skip_freight:
        outer_end_row = perw_row
    else:
        outer_end_row = freight_end_row
    outer_left_col = 1
    outer_right_col = 13

    thin_side = Side(style='thin', color='000000')
    thick_side = Side(style='medium', color='000000')

    for r in range(outer_start_row, outer_end_row + 1):
        for c in range(outer_left_col, outer_right_col + 1):
            cell = ws.cell(row=r, column=c)
            is_skip = False
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range and cell.coordinate != merged_range.start_cell.coordinate:
                    is_skip = True
                    break
            if not is_skip:
                cell.border = Border(
                    left=thin_side, right=thin_side,
                    top=thin_side, bottom=thin_side
                )

    for r in range(outer_start_row, outer_end_row + 1):
        target_cell = ws.cell(row=r, column=outer_left_col)
        for merged_range in ws.merged_cells.ranges:
            if target_cell.coordinate in merged_range:
                target_cell = merged_range.start_cell
                break
        old_border = target_cell.border
        if old_border is None:
            old_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        target_cell.border = Border(
            left=thick_side,
            right=old_border.right,
            top=old_border.top,
            bottom=old_border.bottom
        )

    for r in range(outer_start_row, outer_end_row + 1):
        target_cell = ws.cell(row=r, column=outer_right_col)
        for merged_range in ws.merged_cells.ranges:
            if target_cell.coordinate in merged_range:
                target_cell = merged_range.start_cell
                break
        old_border = target_cell.border
        if old_border is None:
            old_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        target_cell.border = Border(
            left=old_border.left,
            right=thick_side,
            top=old_border.top,
            bottom=old_border.bottom
        )

    for c in range(outer_left_col, outer_right_col + 1):
        target_cell = ws.cell(row=outer_start_row, column=c)
        for merged_range in ws.merged_cells.ranges:
            if target_cell.coordinate in merged_range:
                target_cell = merged_range.start_cell
                break
        old_border = target_cell.border
        if old_border is None:
            old_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        target_cell.border = Border(
            left=old_border.left,
            right=old_border.right,
            top=thick_side,
            bottom=old_border.bottom
        )

    for c in range(outer_left_col, outer_right_col + 1):
        target_cell = ws.cell(row=outer_end_row, column=c)
        for merged_range in ws.merged_cells.ranges:
            if target_cell.coordinate in merged_range:
                target_cell = merged_range.start_cell
                break
        old_border = target_cell.border
        if old_border is None:
            old_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        target_cell.border = Border(
            left=old_border.left,
            right=old_border.right,
            top=old_border.top,
            bottom=thick_side
        )

    hint_row = outer_end_row + 2
    ws.merge_cells(f'A{outer_end_row + 1}:L{outer_end_row + 1}')
    ws.merge_cells(f'A{hint_row}:L{hint_row}')
    ws.cell(row=hint_row, column=1,
            value=_t('summary_hint', lang)).font = Font(
        name='Arial', size=12, color='FF0000')
    ws.cell(row=hint_row, column=1).alignment = left_bottom

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(top=0.75, bottom=0.75, left=0.7, right=0.7, header=0.3, footer=0.3)

    sheet_names = workbook.sheetnames
    if 'Total' in sheet_names:
        idx = sheet_names.index('Total')
        workbook.move_sheet('Total', offset=-idx)

    print(f"   [EN-SIMPLE] Summary sheet(Total) created")
    return ws.title
