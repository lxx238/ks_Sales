from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from backend.core.print_settings import apply_print_setup
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter
import os
import re
from decimal import Decimal

from backend.core.shared.text_utils import _CJK_RE, _strip_cjk_spec, normalize_lookup_code
from backend.core.shared.price_utils import resolve_price_info, has_valid_price_info, round_to_2_decimal
from backend.core.shared.product_utils import _is_valid_product_code, normalize_preinstall
from backend.core.shared.weight_utils import extract_length_from_spec
from backend.core.material_translate import translate_material

MEIRYO = Font(name='Meiryo UI', size=11)
MEIRYO_BOLD = Font(name='Meiryo UI', size=11, bold=True)
MEIRYO_TITLE = Font(name='Meiryo UI', size=16, bold=True)
MEIRYO_HEADER = Font(name='Meiryo UI', size=11, bold=True)

SM_FONT = Font(name='Meiryo UI', size=8)
SM_FONT_BOLD = Font(name='Meiryo UI', size=8, bold=True)
SM_FONT_BOLD_10 = Font(name='Meiryo UI', size=10, bold=True)
SM_FONT_TITLE = Font(name='Meiryo UI', size=16, bold=True)
SM_FONT_HEADER = Font(name='Meiryo UI', size=8, bold=True)
SM_FONT_RED = Font(name='Meiryo UI', size=8, bold=True, color='FF0000')
SM_FONT_RED_10 = Font(name='Meiryo UI', size=10, bold=True, color='FF0000')

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

THICK_SIDE = Side(style='medium', color='000000')

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT_A = Alignment(horizontal='right', vertical='center', wrap_text=True)
LEFT_A = Alignment(horizontal='left', vertical='center', wrap_text=True)

CURRENCY_FMT = '"US$" #,##0.00'
JPY_FMT = '[$￥-411]#,##0;-[$￥-411]#,##0'

BLUE_FILL = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
GREEN_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

SUMMARY_COL_WIDTHS = {
    'A': 6, 'B': 6.5, 'C': 6.5, 'D': 9, 'E': 8,
    'F': 8, 'G': 12, 'H': 12, 'I': 15, 'J': 12, 'K': 12,
}

DETAIL_COL_WIDTHS = {
    'A': 5, 'B': 10, 'C': 10, 'D': 15, 'E': 10, 'F': 10, 'G': 8, 'H': 15,
}

DEFAULT_DISCOUNT_RATE = Decimal('0.71')


def _apply_outer_border(ws, min_row, max_row, min_col, max_col):
    saved_merges = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= min_row and mr.max_row <= max_row and mr.min_col >= min_col and mr.max_col <= max_col:
            saved_merges.append(str(mr))
            ws.unmerge_cells(str(mr))

    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            left_s = THICK_SIDE if c == min_col else Side(style='thin')
            right_s = THICK_SIDE if c == max_col else Side(style='thin')
            top_s = THICK_SIDE if r == min_row else Side(style='thin')
            bottom_s = THICK_SIDE if r == max_row else Side(style='thin')
            cell.border = Border(left=left_s, right=right_s, top=top_s, bottom=bottom_s)

    for mr_str in saved_merges:
        ws.merge_cells(mr_str)


def create_detail_sheet(workbook, array_info, bom_products, price_mapping,
                        sheet_prefix=None,
                        image_path=None, image_folder=None, code_to_images=None,
                        image_temp_dir=None, image_cache=None, matrix_data=None,
                        discount_rate=None, group=None,
                        unmatched_products_out=None,
                        excluded_products=None,
                        need_weight_code=True,
                        inv_remark='',
                        coating_thickness=10):
    from backend.core.quotation_engine import (
        prepare_image_for_excel, add_image_centered_in_cell,
    )
    JA_IMG_W = 65
    JA_IMG_H = 50
    JA_IMG_COL = 4
    JA_IMG_PADDING = 2

    def _col_width_to_px(char_width):
        return int(char_width * 7.5 + 5)

    def _row_height_to_px(pt_height):
        return int(pt_height * 1.33)

    def _fit_image_to_cell(ws, cell_row, cell_col, max_w, max_h, padding=JA_IMG_PADDING):
        col_w = ws.column_dimensions[get_column_letter(cell_col)].width or 8.43
        row_h = ws.row_dimensions[cell_row].height or 15
        avail_w = _col_width_to_px(col_w) - padding * 2
        avail_h = _row_height_to_px(row_h) - padding * 2
        avail_w = max(avail_w, 20)
        avail_h = max(avail_h, 20)
        scale = min(avail_w / max_w, avail_h / max_h, 1.0)
        return int(max_w * scale), int(max_h * scale)

    if sheet_prefix:
        sheet_name = sheet_prefix
    else:
        sheet_name = array_info.get('note', '') or f"セット{array_info.get('no', '1')}"
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    original = sheet_name
    counter = 1
    while sheet_name in workbook.sheetnames:
        sheet_name = f"{original}_{counter}"
        counter += 1

    ws = workbook.create_sheet(title=sheet_name)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center = Alignment(horizontal='center', vertical='center')
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    row2_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    max_col = 8 if not need_weight_code else 10
    max_col_letter = get_column_letter(max_col)

    sm_font = Font(name='Meiryo UI', size=6)
    sm_font_bold = Font(name='Meiryo UI', size=8, bold=True)
    col_widths = {'A': 6, 'B': 10, 'C': 15, 'D': 12, 'E': 10, 'F': 11.5, 'G': 8, 'H': 15, 'I': 10, 'J': 10}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    data_font = sm_font
    header_font = sm_font_bold
    data_row_height = 60

    ws.row_dimensions[1].height = 25

    is_inverter_sheet = (sheet_prefix or '') == 'パワコン'
    is_pile_sheet = (sheet_prefix or '') == '地盤杭'

    ws.merge_cells(f'A1:{max_col_letter}1')
    if is_inverter_sheet:
        ws['A1'] = 'パワコン取付バー/集電箱'
        ws['A1'].font = Font(name='Meiryo UI', size=16, bold=True)
    elif is_pile_sheet:
        ws['A1'] = '地盤杭'
        ws['A1'].font = Font(name='Meiryo UI', size=16, bold=True)
    else:
        project_name = (matrix_data or {}).get('project_name', '')
        ws['A1'] = project_name
        ws['A1'].font = Font(name='Meiryo UI', size=16, bold=True)
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 25 if is_inverter_sheet else 40

    if is_inverter_sheet or is_pile_sheet:
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill = row2_fill
            cell.border = thin_border
        ws.row_dimensions[2].height = 3
    else:
        rows_val = array_info.get('rows', '')
        cols_val = array_info.get('cols', '')

        ws.merge_cells('A2:B2')
        ws['A2'] = f'{rows_val}段' if rows_val else ''
        ws['A2'].font = data_font
        ws['A2'].alignment = center
        ws.merge_cells('C2:D2')
        ws['C2'] = f'{cols_val}列' if cols_val else ''
        ws['C2'].font = data_font
        ws['C2'].alignment = center

        ws.merge_cells('E2:F2')
        ws['E2'] = 'セット数'
        ws['E2'].font = data_font
        ws['E2'].alignment = center

        ws.merge_cells(f'G2:{max_col_letter}2')
        ws['G2'] = array_info.get('table_qty', 1)
        ws['G2'].font = data_font
        ws['G2'].alignment = center

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill = row2_fill
            cell.border = thin_border
        ws.row_dimensions[2].height = 15

    ws.merge_cells(f'A3:{max_col_letter}3')
    ws.row_dimensions[3].height = 3

    headers = ['序号', '品名', '材質', '写真', '規格', '単価(USD)', '数量(PCS)', '総金額(USD)']
    if need_weight_code:
        headers += ['品番', '重量(KG)']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.border = thin_border
        cell.alignment = center
        cell.font = header_font
        cell.fill = header_fill
    ws.row_dimensions[4].height = 25

    name_field = 'name_ja'
    total_price_sum = Decimal('0')
    matched_count = 0
    unmatched_count = 0
    data_start = 5
    _detail_table_qty = int(array_info.get('table_qty', 1)) if array_info else 1

    YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    filtered_products = bom_products

    for row in range(data_start, data_start + len(filtered_products)):
        ws.row_dimensions[row].height = data_row_height

    for idx, product in enumerate(filtered_products):
        row = data_start + idx
        product_code = product.get('code', '')
        price_info = resolve_price_info(price_mapping, product_code)

        display_name = (
            price_info.get(name_field)
            or price_info.get('name_ko')
            or price_info.get('name')
            or product.get('name', '')
        ) if price_info else product.get('name', '')

        ws.cell(row=row, column=1, value=idx + 1).border = thin_border
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=1).font = data_font
        ws.cell(row=row, column=2, value=display_name).border = thin_border
        ws.cell(row=row, column=2).font = data_font
        raw_material = (price_info.get('db_material') if price_info and price_info.get('db_material') else None) or product.get('material', '')
        display_material = ''
        if raw_material:
            mat = str(raw_material).replace('&', '+')
            if coating_thickness in (15, 18):
                mat = mat.replace('304', '316')
                mat = mat.replace('/316', '')
            else:
                mat = mat.replace('/316', '')
            mat = translate_material(mat, 'ja')
            display_material = f'材質：{mat}'
            if coating_thickness in (15, 18):
                display_material += f'  {coating_thickness}um'
        ws.cell(row=row, column=3, value=display_material).border = thin_border
        ws.cell(row=row, column=3).font = data_font
        ws.cell(row=row, column=4, value='').border = thin_border
        ws.cell(row=row, column=4).font = data_font
        ws.cell(row=row, column=5, value=_strip_cjk_spec(product.get('spec', ''))).border = thin_border
        ws.cell(row=row, column=5).font = data_font

        quantity = product.get('quantity', 0)
        unit_price = 0
        display_unit_price = 0
        price_unit = ''
        is_matched = False

        if price_info and has_valid_price_info(price_info):
            unit_price = float(price_info['price'])
            price_unit = price_info.get('unit', '')
            matched_count += 1
            is_matched = True
        else:
            unmatched_count += 1
            if unmatched_products_out is not None and quantity > 0 and _is_valid_product_code(product_code):
                unmatched_products_out.append({
                    'code': product_code,
                    'name': product.get('name', ''),
                    'spec': product.get('spec', ''),
                    'material': raw_material,
                    'quantity': quantity * _detail_table_qty,
                    'weight': product.get('weight', 0),
                    'preinstall': normalize_preinstall(product.get('preinstall')),
                })

        is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
        length_mm = Decimal('0')
        if is_meter:
            length_mm = Decimal(str(extract_length_from_spec(product.get('spec')) or 0))

        if is_meter and length_mm > 0:
            display_unit_price = float(Decimal(str(unit_price)) * length_mm / Decimal('1000'))
        else:
            display_unit_price = unit_price

        if display_unit_price > 0:
            ws.cell(row=row, column=6, value=display_unit_price).border = thin_border
            ws.cell(row=row, column=6).number_format = CURRENCY_FMT
            ws.cell(row=row, column=6).font = data_font
        else:
            ws.cell(row=row, column=6, value='').border = thin_border
            ws.cell(row=row, column=6).font = data_font

        if quantity > 0:
            ws.cell(row=row, column=7,
                    value=int(quantity) if quantity % 1 == 0 else quantity).border = thin_border
            ws.cell(row=row, column=7).font = data_font
        else:
            ws.cell(row=row, column=7, value='').border = thin_border
            ws.cell(row=row, column=7).font = data_font

        total_price = Decimal('0')
        if display_unit_price > 0 and quantity > 0:
            total_price = Decimal(str(display_unit_price)) * Decimal(str(quantity))

        total_price_rounded = round_to_2_decimal(total_price)
        if total_price_rounded > 0:
            ws.cell(row=row, column=8, value=float(total_price)).border = thin_border
            ws.cell(row=row, column=8).number_format = CURRENCY_FMT
            ws.cell(row=row, column=8).font = data_font
            total_price_sum += total_price_rounded
        else:
            ws.cell(row=row, column=8, value='').border = thin_border
            ws.cell(row=row, column=8).font = data_font

        ws.cell(row=row, column=2).alignment = center
        ws.cell(row=row, column=3).alignment = center
        ws.cell(row=row, column=5).alignment = center
        ws.cell(row=row, column=6).alignment = center
        ws.cell(row=row, column=7).alignment = center
        ws.cell(row=row, column=8).alignment = center

        if need_weight_code:
            ws.cell(row=row, column=9, value=product_code).border = thin_border
            ws.cell(row=row, column=9).alignment = center
            ws.cell(row=row, column=9).font = data_font
            weight_val = '/'
            if price_info:
                w = price_info.get('db_weight')
                price_unit = price_info.get('unit', '')
                if w:
                    is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
                    if is_meter:
                        length_mm = extract_length_from_spec(product.get('spec'))
                        if length_mm and length_mm > 0:
                            weight_val = round(float(w) / 1000 * length_mm, 4)
                        else:
                            weight_val = float(w)
                    else:
                        weight_val = float(w)
            ws.cell(row=row, column=10, value=weight_val).border = thin_border
            ws.cell(row=row, column=10).alignment = center
            ws.cell(row=row, column=10).font = data_font

        if not is_matched:
            for c in range(1, max_col + 1):
                ws.cell(row=row, column=c).fill = YELLOW_FILL

    # ========== 删除行 ==========
    rows_to_delete = []
    for r in range(data_start, data_start + len(filtered_products)):
        name_val = ws.cell(row=r, column=2).value
        if name_val is None or name_val == '':
            continue
        qty_val = ws.cell(row=r, column=7).value
        if qty_val is None or qty_val == '' or (isinstance(qty_val, (int, float)) and qty_val == 0):
            rows_to_delete.append(r)

    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r)

    deleted_count = len(rows_to_delete)
    if deleted_count:
        print(f"   🗑️ 数量/重量=0/空の行を {deleted_count} 行削除")

    new_seq = 1
    for r in range(data_start, ws.max_row + 1):
        name_val = ws.cell(row=r, column=2).value
        if name_val and name_val != '':
            ws.cell(row=r, column=1, value=new_seq)
            new_seq += 1

    data_end = data_start + len(filtered_products) - 1 - deleted_count

    # ========== 图片插入（用row_product_map/remap处理删除后的行偏移） ==========
    row_product_map = {}
    for idx, product in enumerate(filtered_products):
        orig_row = data_start + idx
        pc = product.get('code', '')
        row_product_map[orig_row] = str(pc).strip() if pc else ''

    survived = [r for r in sorted(row_product_map.keys()) if r not in set(rows_to_delete)]
    row_remap = {}
    for orig_row in survived:
        offset = sum(1 for d in rows_to_delete if d < orig_row)
        row_remap[orig_row - offset] = row_product_map[orig_row]

    image_found_count = 0
    image_not_found_count = 0
    for row, product_code in row_remap.items():
        if row < data_start or row > data_end:
            continue
        if not product_code:
            continue
        normalized_code = normalize_lookup_code(product_code)

        img_path = None
        if code_to_images and product_code in code_to_images and code_to_images[product_code]:
            img_path = code_to_images[product_code][0]
        elif code_to_images and normalized_code in code_to_images and code_to_images[normalized_code]:
            img_path = code_to_images[normalized_code][0]

        if not img_path and price_mapping:
            pi = resolve_price_info(price_mapping, product_code)
            if pi and pi.get('image_bytes'):
                img_bytes = pi['image_bytes']
                img_ext = pi.get('image_ext', '.png')
                if image_temp_dir:
                    import os as _os
                    db_img_name = f"db_{normalized_code or product_code}{img_ext}"
                    db_img_path = _os.path.join(image_temp_dir, db_img_name)
                    if db_img_path not in image_cache:
                        with open(db_img_path, 'wb') as _f:
                            _f.write(img_bytes)
                        image_cache[db_img_path] = True
                    img_path = db_img_path

        if img_path:
            fit_w, fit_h = _fit_image_to_cell(ws, row, JA_IMG_COL, JA_IMG_W, JA_IMG_H)
            temp_img_path = prepare_image_for_excel(
                img_path,
                target_width=fit_w,
                target_height=fit_h,
                temp_dir=image_temp_dir,
                cache=image_cache,
            )
            final_img_path = temp_img_path if temp_img_path else img_path
            success = add_image_centered_in_cell(
                ws,
                final_img_path,
                row,
                JA_IMG_COL,
                img_width=fit_w,
                img_height=fit_h,
            )
            if success:
                image_found_count += 1
                continue
        image_not_found_count += 1
        ws.cell(row=row, column=4, value='/').alignment = center

    if image_found_count or image_not_found_count:
        print(f"   🖼️ {sheet_name}: 挿入 {image_found_count} 枚, 未検出 {image_not_found_count} 枚")

    center_both = Alignment(horizontal='center', vertical='center')
    sub_row = data_end + 1
    ws.merge_cells(f'A{sub_row}:G{sub_row}')
    ws[f'A{sub_row}'] = 'SUB-TOTAL-（FOB）1基スクリュー杭基礎架台合計'
    ws[f'A{sub_row}'].alignment = center_both
    ws[f'A{sub_row}'].font = data_font

    sub_total_merge_end = 8 if not need_weight_code else 10
    if sub_total_merge_end > 8:
        ws.merge_cells(f'H{sub_row}:{get_column_letter(sub_total_merge_end)}{sub_row}')
    ws.cell(row=sub_row, column=8, value=f'=SUM(H5:H{data_end})')
    ws.cell(row=sub_row, column=8).number_format = CURRENCY_FMT
    ws.cell(row=sub_row, column=8).font = data_font
    ws.cell(row=sub_row, column=8).alignment = center_both
    ws.row_dimensions[sub_row].height = 25
    for c in range(1, max_col + 1):
        ws.cell(row=sub_row, column=c).border = thin_border
        ws.cell(row=sub_row, column=c).font = data_font

    _apply_outer_border(ws, 1, sub_row, 1, max_col)

    apply_print_setup(ws, 'ja')

    return {
        'sheet_name': sheet_name,
        'array_info': array_info,
        'total_price_per_base': float(total_price_sum),
        'total_price': float(total_price_sum),
        'matched_count': matched_count,
        'unmatched_count': unmatched_count,
        'sub_total_row': sub_row,
    }


def create_ja_inquiry_sheet(workbook, unmatched_products, price_mapping=None):
    thin = Side(style='thin', color='000000')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(name='Meiryo UI', size=9, bold=True)
    data_font = Font(name='Meiryo UI', size=9)
    center_a = Alignment(horizontal='center', vertical='center')
    left_a = Alignment(horizontal='left', vertical='center')

    ws = workbook.create_sheet(title='询价表')

    col_widths = {'A': 15, 'B': 30, 'C': 30}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    headers = ['工程编码', '工程品名', '工程品名--日语']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = center_a
        cell.fill = header_fill
        cell.border = thin_border

    seen_keys = set()
    current_row = 2
    for product in unmatched_products:
        code = product.get('code', '')
        spec = product.get('spec', '')
        key = (code, spec)
        if key in seen_keys:
            continue
        seen_keys.add(key)

        name = product.get('name', '')
        if price_mapping:
            price_info = resolve_price_info(price_mapping, code)
            if price_info:
                name = price_info.get('name', name)

        ws.cell(row=current_row, column=1, value=code).border = thin_border
        ws.cell(row=current_row, column=1).font = data_font
        ws.cell(row=current_row, column=1).alignment = center_a

        ws.cell(row=current_row, column=2, value=name).border = thin_border
        ws.cell(row=current_row, column=2).font = data_font
        ws.cell(row=current_row, column=2).alignment = left_a

        ws.cell(row=current_row, column=3, value='').border = thin_border
        ws.cell(row=current_row, column=3).font = data_font
        ws.cell(row=current_row, column=3).alignment = left_a

        current_row += 1

    return ws.title


def create_summary_sheet(workbook, detail_results, matrix_data=None,
                         image_path=None, fence_data=None, shipping_data=None,
                         inverter_detail=None, pile_detail=None, pile_summary=None):
    from datetime import datetime

    ws = workbook.create_sheet(title='合計')

    for col, width in SUMMARY_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_a = Alignment(horizontal='right', vertical='center', wrap_text=True)
    left_a = Alignment(horizontal='left', vertical='center', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    matrix_data = matrix_data or {}
    project_name = str(matrix_data.get('project_name') or '').strip()
    output_kw = matrix_data.get('output_kw') or 0
    module_wattage = matrix_data.get('module_wattage') or 0
    arrays = matrix_data.get('arrays') or []
    wind_speed = matrix_data.get('max_wind_speed') or ''
    snow_load = matrix_data.get('max_snow_load') or ''
    angle = matrix_data.get('angle') or ''
    panel_size = matrix_data.get('panel_spec') or matrix_data.get('module_size') or ''

    fence_data = fence_data or {}
    shipping_data = shipping_data or {}

    # ========== Row 1: 見積日 + Logo ==========
    _base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
    _est_logo = os.path.join(_base_dir, 'input', 'ESTlogo.png')
    logo_path = _est_logo if os.path.isfile(_est_logo) else image_path
    if logo_path and os.path.isfile(str(logo_path)):
        try:
            img = XlImage(str(logo_path))
            img.width = 240
            img.height = 60
            ws.add_image(img, 'A2')
        except Exception:
            pass

    ws['I1'] = '見積日：'
    ws['I1'].font = SM_FONT
    ws['I1'].alignment = right_a
    ws.merge_cells('J1:K1')
    ws['J1'] = datetime.now().date()
    ws['J1'].font = SM_FONT
    ws['J1'].number_format = 'yyyy"年"m"月"d"日"'
    ws['J1'].alignment = center
    ws.row_dimensions[1].height = 25

    # ========== Row 2: 架台御見積書 ==========
    ws['D2'] = '架台御見積書'
    ws.merge_cells('D2:J2')
    ws['D2'].font = SM_FONT_TITLE
    ws['D2'].alignment = center
    ws.row_dimensions[2].height = 40

    # ========== Row 3: 案件名 ==========
    ws['A3'] = '案件名：'
    ws['A3'].font = SM_FONT
    ws['A3'].alignment = left_a
    ws.merge_cells('A3:B3')
    ws.merge_cells('C3:H3')
    ws['C3'] = project_name
    ws['C3'].font = SM_FONT
    ws['C3'].alignment = left_a
    if output_kw:
        ws['J3'] = output_kw
        ws['J3'].font = SM_FONT
        ws['J3'].alignment = right_a
    ws['K3'] = 'KW'
    ws['K3'].font = SM_FONT
    ws['K3'].alignment = center

    # ========== Row 4: 見積条件 ==========
    ws['A4'] = '見積条件：'
    ws['A4'].font = SM_FONT
    ws['A4'].alignment = left_a
    ws.merge_cells('A4:B4')
    ws.merge_cells('C4:H4')

    # ========== Row 5: 納入期限 ==========
    ws['A5'] = '納入期限：'
    ws['A5'].font = SM_FONT
    ws['A5'].alignment = left_a
    ws.merge_cells('A5:B5')
    ws['C5'] = '発注後10-14日間後工場から出荷'
    ws['C5'].font = SM_FONT
    ws['C5'].alignment = left_a
    ws.merge_cells('C5:H5')

    # ========== Row 6: 取引条件 ==========
    ws['A6'] = '取引条件：'
    ws['A6'].font = SM_FONT
    ws['A6'].alignment = left_a
    ws.merge_cells('A6:B6')
    ws['C6'] = '取引基本契約書に基づく'
    ws['C6'].font = SM_FONT
    ws['C6'].alignment = left_a
    ws.merge_cells('C6:H6')

    # ========== Row 7: 有効期限 ==========
    ws['A7'] = '有効期限：'
    ws['A7'].font = SM_FONT
    ws['A7'].alignment = left_a
    ws.merge_cells('A7:B7')
    ws['C7'] = '御見積後2日間'
    ws['C7'].font = SM_FONT
    ws['C7'].alignment = left_a
    ws.merge_cells('C7:H7')

    # ========== Row 8 ==========
    ws['A8'] = '※'
    ws['A8'].font = SM_FONT
    ws['A8'].alignment = center
    ws.merge_cells('B8:C8')
    ws['B8'] = '風速 図面通り'
    ws['B8'].font = SM_FONT
    ws['B8'].alignment = center
    ws.merge_cells('D8:E8')
    ws['D8'] = 'パネル高さ 図面通り'
    ws['D8'].font = SM_FONT
    ws['D8'].alignment = center
    ws['F8'] = '※'
    ws['F8'].font = SM_FONT
    ws['F8'].alignment = center
    ws['G8'] = 'パネルサイズ'
    ws['G8'].font = SM_FONT
    ws['G8'].alignment = center
    ws['H8'] = '図面通り'
    ws['H8'].font = SM_FONT
    ws['H8'].alignment = center

    # ========== Row 9 ==========
    ws['A9'] = '※'
    ws['A9'].font = SM_FONT
    ws['A9'].alignment = center
    ws.merge_cells('B9:C9')
    ws['B9'] = '積雪 図面通り'
    ws['B9'].font = SM_FONT
    ws['B9'].alignment = center
    ws.merge_cells('D9:E9')
    ws['D9'] = '傾斜角度 図面通り'
    ws['D9'].font = SM_FONT
    ws['D9'].alignment = center
    ws['F9'] = '※'
    ws['F9'].font = SM_FONT
    ws['F9'].alignment = center
    ws['G9'] = '発電量/PC'
    ws['G9'].font = SM_FONT
    ws['G9'].alignment = center
    ws['H9'] = module_wattage if module_wattage else '図面通り'
    ws['H9'].font = SM_FONT
    ws['H9'].alignment = center

    # ========== Row 10: 一、架台本体金額 ==========
    ws['A10'] = '一、架台本体金額（Ex Works）'
    ws['A10'].font = SM_FONT_BOLD_10
    ws['A10'].alignment = left_a
    ws.merge_cells('A10:K10')
    ws.row_dimensions[10].height = 20

    # ========== Row 11: 表头 ==========
    ws['A11'] = '序号'
    ws['A11'].font = SM_FONT_HEADER
    ws['A11'].alignment = center

    ws.merge_cells('B11:C11')
    ws['B11'] = 'パネル数'
    ws['B11'].font = SM_FONT_HEADER
    ws['B11'].alignment = center

    ws['D11'] = 'セット数'
    ws['D11'].font = SM_FONT_HEADER
    ws['D11'].alignment = center

    ws.merge_cells('E11:F11')
    ws['E11'] = '備考'
    ws['E11'].font = SM_FONT_HEADER
    ws['E11'].alignment = center

    ws['G11'] = '発電量（KW）'
    ws['G11'].font = SM_FONT_HEADER
    ws['G11'].alignment = center

    ws['H11'] = '単価(USD)/基'
    ws['H11'].font = SM_FONT_HEADER
    ws['H11'].alignment = center

    ws['I11'] = '特別値引き後金額'
    ws['I11'].font = SM_FONT_HEADER
    ws['I11'].alignment = center
    ws['I11'].fill = BLUE_FILL

    ws['J11'] = '総金額(USD)'
    ws['J11'].font = SM_FONT_HEADER
    ws['J11'].alignment = center

    ws['K11'] = 'W単価(USD)'
    ws['K11'].font = SM_FONT_HEADER
    ws['K11'].alignment = center

    # ========== Row 12+: 数据行 ==========
    discount_rate_pct = shipping_data.get('discount_rate', 71) if shipping_data else 71
    discount_rate = Decimal(str(discount_rate_pct)) / 100
    grand_total = Decimal('0')
    grand_output_kw = Decimal('0')
    data_row_start = 12

    for i, detail in enumerate(detail_results):
        row = data_row_start + i
        arr = detail.get('array_info', {})
        base_count = arr.get('table_qty', 1)
        panel_count = arr.get('rows', 0) * arr.get('cols', 0)
        note = detail.get('inv_note', '') or (arr.get('note') or '')

        gen_kw = Decimal(str(panel_count)) * Decimal(str(base_count)) * Decimal(str(module_wattage)) / Decimal('1000')
        per_base = Decimal(str(detail.get('total_price_per_base', 0)))
        special_price = round_to_2_decimal(per_base * discount_rate)
        total_amount = round_to_2_decimal(special_price * Decimal(str(base_count)))
        w_unit_price = round_to_2_decimal(special_price / (Decimal(str(panel_count)) * Decimal(str(module_wattage)))) if panel_count > 0 and module_wattage > 0 else Decimal('0')

        ws.cell(row=row, column=1, value=i + 1).font = SM_FONT
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=2, value=arr.get('rows', 0)).font = SM_FONT
        ws.cell(row=row, column=2).alignment = center
        ws.cell(row=row, column=3, value=arr.get('cols', 0)).font = SM_FONT
        ws.cell(row=row, column=3).alignment = center
        ws.cell(row=row, column=4, value=base_count).font = SM_FONT
        ws.cell(row=row, column=4).alignment = center
        ws.merge_cells(f'E{row}:F{row}')
        ws.cell(row=row, column=5, value=note).font = SM_FONT
        ws.cell(row=row, column=5).alignment = center
        ws.cell(row=row, column=7, value=f'=(B{row}*C{row}*D{row})*{module_wattage}/1000').font = SM_FONT
        ws.cell(row=row, column=7).alignment = center
        ws.cell(row=row, column=7).number_format = '#,##0.00'

        detail_sheet_name = detail.get('sheet_name', '')
        sub_total_row = detail.get('sub_total_row', 24)
        ws.cell(row=row, column=8, value=f"='{detail_sheet_name}'!H{sub_total_row}").font = SM_FONT
        ws.cell(row=row, column=8).number_format = CURRENCY_FMT
        ws.cell(row=row, column=8).alignment = center
        ws.cell(row=row, column=9, value=f'=H{row}*{discount_rate_pct}/100').font = SM_FONT
        ws.cell(row=row, column=9).number_format = CURRENCY_FMT
        ws.cell(row=row, column=9).alignment = center
        ws.cell(row=row, column=9).fill = BLUE_FILL
        ws.cell(row=row, column=10, value=f'=I{row}*D{row}').font = SM_FONT
        ws.cell(row=row, column=10).number_format = CURRENCY_FMT
        ws.cell(row=row, column=10).alignment = center
        ws.cell(row=row, column=11, value=f'=J{row}/G{row}/1000').font = SM_FONT
        ws.cell(row=row, column=11).number_format = CURRENCY_FMT
        ws.cell(row=row, column=11).alignment = center

        for c in range(1, 12):
            ws.cell(row=row, column=c).border = thin_border

        grand_total += total_amount
        grand_output_kw += gen_kw

    inverter_row_idx = len(detail_results)
    inverter_grand_amount = Decimal('0')
    if inverter_detail:
        inv_details = inverter_detail if isinstance(inverter_detail, list) else [inverter_detail]
        for inv_idx, inv_detail_item in enumerate(inv_details):
            row = data_row_start + inverter_row_idx
            inv_matched_table_qty = inv_detail_item.get('array_info', {}).get('table_qty', 1)
            inv_per_base = Decimal(str(inv_detail_item.get('total_price_per_base', 0)))
            inv_special = round_to_2_decimal(inv_per_base * discount_rate)
            inverter_grand_amount += inv_special * Decimal(str(inv_matched_table_qty))

            inv_remark_text = inv_detail_item.get('inv_remark', '')
            inv_qty_per_base = inv_detail_item.get('inv_qty_per_base', 0)
            if inv_qty_per_base > 0:
                inv_total_qty = inv_qty_per_base * inv_matched_table_qty
                inv_display_name = f"パワコン取付バー  {inv_total_qty}台"
            elif inv_remark_text:
                inv_display_name = inv_remark_text
            else:
                inv_display_name = 'パワコン取付バー'

            ws.cell(row=row, column=1, value=inverter_row_idx + 1).font = SM_FONT
            ws.cell(row=row, column=1).alignment = center
            ws.merge_cells(f'B{row}:D{row}')
            ws.cell(row=row, column=2, value=inv_display_name).font = SM_FONT
            ws.cell(row=row, column=2).alignment = center
            ws.merge_cells(f'E{row}:F{row}')
            ws.cell(row=row, column=5, value='').font = SM_FONT
            ws.cell(row=row, column=5).alignment = center
            ws.cell(row=row, column=7, value='').font = SM_FONT
            ws.cell(row=row, column=7).alignment = center

            inv_sheet_name = inv_detail_item.get('sheet_name', '')
            inv_sub_row = inv_detail_item.get('sub_total_row', 8)
            ws.cell(row=row, column=8, value=f"='{inv_sheet_name}'!H{inv_sub_row}").font = SM_FONT
            ws.cell(row=row, column=8).number_format = CURRENCY_FMT
            ws.cell(row=row, column=8).alignment = center
            ws.cell(row=row, column=9, value=f'=H{row}*{discount_rate_pct}/100').font = SM_FONT
            ws.cell(row=row, column=9).number_format = CURRENCY_FMT
            ws.cell(row=row, column=9).alignment = center
            ws.cell(row=row, column=9).fill = BLUE_FILL
            ws.cell(row=row, column=10, value=f'=I{row}*{inv_matched_table_qty}').font = SM_FONT
            ws.cell(row=row, column=10).number_format = CURRENCY_FMT
            ws.cell(row=row, column=10).alignment = center
            ws.cell(row=row, column=11, value='').font = SM_FONT
            ws.cell(row=row, column=11).alignment = center

            for c in range(1, 12):
                ws.cell(row=row, column=c).border = thin_border

            grand_total += inv_special * Decimal(str(inv_matched_table_qty))
            inverter_row_idx += 1

    data_end = data_row_start + inverter_row_idx - 1

    # ========== Row 17+: 固定行区域 ==========
    r17 = data_end + 1

    # Row 17: 架台総金額
    ws[f'A{r17}'] = '架台総金額(USD)'
    ws[f'A{r17}'].font = SM_FONT
    ws[f'A{r17}'].alignment = right_a
    ws.merge_cells(f'A{r17}:J{r17}')
    ws[f'K{r17}'] = f'=SUM(J{data_row_start}:J{data_end})'
    ws[f'K{r17}'].font = SM_FONT
    ws[f'K{r17}'].alignment = center
    ws[f'K{r17}'].number_format = CURRENCY_FMT
    for c in range(1, 12):
        ws.cell(row=r17, column=c).fill = BLUE_FILL

    r18 = r17 + 1
    pile_label = 'スクリュー杭'
    _has_pile = pile_summary is not None
    if _has_pile:
        _ps_price = pile_summary.get('total_price', 0)
        _ps_qty = pile_summary.get('total_qty', 0)
        grand_total += Decimal(str(_ps_price))
        ws[f'A{r18}'] = pile_label
        ws[f'A{r18}'].font = SM_FONT
        ws[f'A{r18}'].alignment = center
        ws.merge_cells(f'A{r18}:C{r18}')
        ws[f'D{r18}'] = '単価(USD)'
        ws[f'D{r18}'].font = SM_FONT
        ws[f'D{r18}'].alignment = center
        ws.merge_cells(f'E{r18}:G{r18}')
        ws[f'H{r18}'] = '数量'
        ws[f'H{r18}'].font = SM_FONT
        ws[f'H{r18}'].alignment = center
        ws[f'I{r18}'] = int(_ps_qty) if _ps_qty == int(_ps_qty) else _ps_qty
        ws[f'I{r18}'].font = SM_FONT
        ws[f'I{r18}'].alignment = center
        ws[f'J{r18}'] = '総金額(USD)'
        ws[f'J{r18}'].font = SM_FONT
        ws[f'J{r18}'].alignment = center
        ws[f'K{r18}'] = _ps_price
        ws[f'K{r18}'].font = SM_FONT
        ws[f'K{r18}'].alignment = center
        ws[f'K{r18}'].number_format = CURRENCY_FMT
        ws[f'E{r18}'] = f'=IF(I{r18}=0,"",K{r18}/I{r18})'
        ws[f'E{r18}'].font = SM_FONT
        ws[f'E{r18}'].alignment = center
        ws[f'E{r18}'].number_format = CURRENCY_FMT
    else:
        ws[f'A{r18}'] = pile_label
        ws[f'A{r18}'].font = SM_FONT
        ws[f'A{r18}'].alignment = center
        ws.merge_cells(f'A{r18}:C{r18}')
        ws[f'D{r18}'] = '単価(USD)'
        ws[f'D{r18}'].font = SM_FONT
        ws[f'D{r18}'].alignment = center
        ws.merge_cells(f'E{r18}:G{r18}')
        ws[f'H{r18}'] = '数量'
        ws[f'H{r18}'].font = SM_FONT
        ws[f'H{r18}'].alignment = center
        ws[f'I{r18}'] = 0
        ws[f'I{r18}'].font = SM_FONT
        ws[f'I{r18}'].alignment = center
        ws[f'J{r18}'] = '総金額(USD)'
        ws[f'J{r18}'].font = SM_FONT
        ws[f'J{r18}'].alignment = center
        ws[f'K{r18}'] = 0
        ws[f'K{r18}'].font = SM_FONT
        ws[f'K{r18}'].alignment = center

    r19 = r18 + 1
    ws[f'A{r19}'] = '①架台＋杭本体　総金額(USD)'
    ws[f'A{r19}'].font = SM_FONT
    ws[f'A{r19}'].alignment = right_a
    ws.merge_cells(f'A{r19}:I{r19}')
    ws.merge_cells(f'J{r19}:K{r19}')
    if _has_pile:
        ws[f'J{r19}'] = f'=K{r17}+K{r18}'
    else:
        ws[f'J{r19}'] = f'=K{r17}'
    ws[f'J{r19}'].font = SM_FONT
    ws[f'J{r19}'].alignment = center
    ws[f'J{r19}'].number_format = CURRENCY_FMT

    r20 = r19 + 1
    ws[f'A{r20}'] = '発電量(KW)'
    ws[f'A{r20}'].font = SM_FONT
    ws[f'A{r20}'].alignment = right_a
    ws.merge_cells(f'A{r20}:I{r20}')
    ws.merge_cells(f'J{r20}:K{r20}')
    ws[f'J{r20}'] = f'=SUM(G{data_row_start}:G{data_end})'
    ws[f'J{r20}'].font = SM_FONT
    ws[f'J{r20}'].alignment = center
    ws[f'J{r20}'].number_format = '#,##0.00'

    r21 = r20 + 1
    ws[f'A{r21}'] = 'ワットあたりの価格'
    ws[f'A{r21}'].font = SM_FONT
    ws[f'A{r21}'].alignment = right_a
    ws.merge_cells(f'A{r21}:I{r21}')
    ws.merge_cells(f'J{r21}:K{r21}')
    ws[f'J{r21}'] = f'=J{r19}/J{r20}/1000'
    ws[f'J{r21}'].font = SM_FONT
    ws[f'J{r21}'].alignment = center
    ws[f'J{r21}'].number_format = '#,##0.00'

    ws['J3'] = f'=J{r20}'
    ws['J3'].font = SM_FONT
    ws['J3'].alignment = right_a

    fence_data = fence_data or {}
    fence_items_raw = fence_data.get('items') or []
    has_fence_data = bool(fence_items_raw)

    has_fence = False
    r_f2 = r21

    if has_fence_data:
        r22 = r21 + 1
        ws[f'A{r22}'] = '二、フェンス金額'
        ws[f'A{r22}'].font = SM_FONT_BOLD_10
        ws[f'A{r22}'].alignment = left_a
        ws.merge_cells(f'A{r22}:D{r22}')
        ws.row_dimensions[r22].height = 20
        fence_length = str(fence_data.get('length', 140)).replace('M', '').replace('m', '').strip()
        fence_length = fence_length or '140'
        fence_length_val = 0
        try:
            fence_length_val = float(fence_length)
        except (ValueError, TypeError):
            fence_length_val = 0
        fence_color = fence_data.get('color', '茶')
        fence_corner = fence_data.get('corner', 'コーナー10箇所')
        fence_items = fence_items_raw

        _fence_name_ja_map = {
            '网片': '網', '网扣': '網連結金具', '门扣': '扉連結金具',
            '弯钩螺栓': 'J型ボルト', '端盖': 'キャップ', '橡胶环': 'EPDMリング',
            '地桩': '杭', '外六角螺栓': 'フェンス用固定ボルト',
            '横插销': '水平ピン', '竖插销': '垂直ピン',
            '门网片': '扉用の網', '门网片(含门柱门框)': '片開き扉',
            '平头圆管': '柱材', '削尖圆管': '柱材（先端尖）', '锁地桩圆管': '柱材',
            '折半圆网片': '網（半円折曲）', '折三角网片': '網（三角折曲）',
        }

        def _translate_fence_name(cn_name):
            if not cn_name:
                return cn_name
            if cn_name in _fence_name_ja_map:
                return _fence_name_ja_map[cn_name]
            for cn_key, ja_val in _fence_name_ja_map.items():
                if cn_key in cn_name:
                    return ja_val
            return cn_name

        if fence_length_val == 0:
            fence_items = []
            for item in fence_items_raw:
                zero_item = dict(item)
                zero_item['qty'] = 0
                zero_item['unit_price'] = 0
                fence_items.append(zero_item)
        ws.merge_cells(f'E{r22}:F{r22}')
        ws[f'E{r22}'] = f'{fence_length}M'
        ws[f'E{r22}'].font = SM_FONT
        ws[f'E{r22}'].alignment = center
        ws[f'G{r22}'] = fence_color
        ws[f'G{r22}'].font = SM_FONT
        ws[f'G{r22}'].alignment = center
        ws[f'H{r22}'] = fence_corner
        ws[f'H{r22}'].font = SM_FONT
        ws[f'H{r22}'].alignment = center

        r23 = r22 + 1
        ws[f'A{r23}'] = 'NO.'
        ws[f'A{r23}'].font = SM_FONT_HEADER
        ws[f'A{r23}'].alignment = center
        ws[f'B{r23}'] = '規格・部品名称'
        ws[f'B{r23}'].font = SM_FONT_HEADER
        ws[f'B{r23}'].alignment = center
        ws.merge_cells(f'B{r23}:F{r23}')
        ws[f'G{r23}'] = '単位'
        ws[f'G{r23}'].font = SM_FONT_HEADER
        ws[f'G{r23}'].alignment = center
        ws[f'H{r23}'] = '1Mの単価\nUnit Price'
        ws[f'H{r23}'].font = SM_FONT_HEADER
        ws[f'H{r23}'].alignment = center
        ws[f'I{r23}'] = '数量\nQTY'
        ws[f'I{r23}'].font = SM_FONT_HEADER
        ws[f'I{r23}'].alignment = center
        ws[f'J{r23}'] = '総金額(US＄)\nAmount Price'
        ws[f'J{r23}'].font = SM_FONT_HEADER
        ws[f'J{r23}'].alignment = center
        ws.merge_cells(f'J{r23}:K{r23}')

        fence_total = Decimal('0')
        for fi in range(5):
            r = r23 + 1 + fi
            ws.cell(row=r, column=1, value='').alignment = center
            ws.merge_cells(f'B{r}:F{r}')
            for col in range(7, 10):
                ws.cell(row=r, column=col, value='').alignment = center
            ws.merge_cells(f'J{r}:K{r}')
            if fi < len(fence_items):
                item = fence_items[fi]
                ws.cell(row=r, column=1, value=fi + 1).font = SM_FONT
                ws.cell(row=r, column=1).alignment = center
                ws.cell(row=r, column=2, value=_translate_fence_name(item.get('name', ''))).font = SM_FONT
                ws.cell(row=r, column=7, value=item.get('unit', '')).font = SM_FONT
                ws.cell(row=r, column=7).alignment = center
                ws.cell(row=r, column=8, value=item.get('unit_price', 0)).font = SM_FONT
                ws.cell(row=r, column=8).alignment = center
                ws.cell(row=r, column=8).number_format = CURRENCY_FMT
                ws.cell(row=r, column=9, value=item.get('qty', 0)).font = SM_FONT
                ws.cell(row=r, column=9).alignment = center
                amount = Decimal(str(item.get('unit_price', 0))) * Decimal(str(item.get('qty', 0)))
                fence_total += amount
                ws.cell(row=r, column=10, value=float(round_to_2_decimal(amount))).font = SM_FONT
                ws.cell(row=r, column=10).alignment = center
                ws.cell(row=r, column=10).number_format = CURRENCY_FMT
            for c in range(1, 12):
                ws.cell(row=r, column=c).border = thin_border

        fence_end = r23 + 5
        has_fence = fence_total > 0

        _case_type = shipping_data.get('case_type', 'EST') if shipping_data else 'EST'

        if _case_type == 'EST':
            r_f2 = fence_end + 1
            ws[f'A{r_f2}'] = '②総金額'
            ws[f'A{r_f2}'].font = SM_FONT_RED_10
            ws[f'A{r_f2}'].alignment = right_a
            ws.merge_cells(f'A{r_f2}:I{r_f2}')
            ws[f'J{r_f2}'] = f'=SUM(J{r23}:J{fence_end})'
            ws[f'J{r_f2}'].font = SM_FONT
            ws[f'J{r_f2}'].alignment = center
            ws[f'J{r_f2}'].number_format = CURRENCY_FMT
            ws.merge_cells(f'J{r_f2}:K{r_f2}')
        else:
            fence_discount_rate = shipping_data.get('fence_discount_rate', 94) if shipping_data else 94

            r_f1 = fence_end + 1
            ws[f'A{r_f1}'] = 'フェンス値引き'
            ws[f'A{r_f1}'].font = SM_FONT
            ws[f'A{r_f1}'].alignment = right_a
            ws.merge_cells(f'A{r_f1}:I{r_f1}')
            ws[f'J{r_f1}'] = f'=SUM(J{r23}:J{fence_end})*{fence_discount_rate}/100'
            ws[f'J{r_f1}'].font = SM_FONT
            ws[f'J{r_f1}'].alignment = center
            ws[f'J{r_f1}'].number_format = CURRENCY_FMT
            ws.merge_cells(f'J{r_f1}:K{r_f1}')

            r_f2 = r_f1 + 1
            ws[f'A{r_f2}'] = '②総金額'
            ws[f'A{r_f2}'].font = SM_FONT_RED_10
            ws[f'A{r_f2}'].alignment = right_a
            ws.merge_cells(f'A{r_f2}:I{r_f2}')
            ws[f'J{r_f2}'] = f'=J{r_f1}'
            ws[f'J{r_f2}'].font = SM_FONT
            ws[f'J{r_f2}'].alignment = center
            ws[f'J{r_f2}'].number_format = CURRENCY_FMT
            ws.merge_cells(f'J{r_f2}:K{r_f2}')

    # ========== 二、運賃 ==========
    r_31 = r_f2 + 1
    ws[f'A{r_31}'] = '二、運賃'
    ws[f'A{r_31}'].font = SM_FONT_BOLD_10
    ws[f'A{r_31}'].alignment = left_a
    ws.merge_cells(f'A{r_31}:K{r_31}')
    ws.row_dimensions[r_31].height = 20

    r_32 = r_31 + 1
    ws[f'A{r_32}'] = '1ドル＝'
    ws[f'A{r_32}'].font = SM_FONT
    ws[f'A{r_32}'].alignment = right_a
    ws.merge_cells(f'A{r_32}:D{r_32}')
    exchange_rate = shipping_data.get('exchange_rate', 160)
    tariff_rate_pct = shipping_data.get('tariff_rate', 1.6)
    consumption_tax_pct = shipping_data.get('consumption_tax', 10)
    fence_tax_pct = shipping_data.get('fence_tax', 10)
    ws[f'E{r_32}'] = exchange_rate
    ws[f'E{r_32}'].font = SM_FONT
    ws[f'E{r_32}'].alignment = center
    ws[f'F{r_32}'] = '円にて換算、レートは２％以上増やすと、当日のレートに変更'
    ws[f'F{r_32}'].font = SM_FONT
    ws[f'F{r_32}'].alignment = left_a
    ws.merge_cells(f'F{r_32}:I{r_32}')
    ws[f'J{r_32}'] = '（USD）'
    ws[f'J{r_32}'].font = SM_FONT_RED
    ws[f'J{r_32}'].alignment = center
    ws[f'K{r_32}'] = '(JPY)'
    ws[f'K{r_32}'].font = SM_FONT_RED
    ws[f'K{r_32}'].alignment = center

    r_33 = r_32 + 1
    ws[f'A{r_33}'] = f'架台+杭　関税「H.S.code {float(grand_total):.2f} 構造物及びその部分品（その他のもの）」{tariff_rate_pct}％'
    ws[f'A{r_33}'].font = SM_FONT
    ws[f'A{r_33}'].alignment = right_a
    ws.merge_cells(f'A{r_33}:I{r_33}')
    ws[f'J{r_33}'] = f'=J{r19}*{tariff_rate_pct}/100'
    ws[f'J{r_33}'].font = SM_FONT
    ws[f'J{r_33}'].alignment = center
    ws[f'J{r_33}'].number_format = CURRENCY_FMT
    ws[f'K{r_33}'] = f'=J{r_33}*E{r_32}'
    ws[f'K{r_33}'].font = SM_FONT
    ws[f'K{r_33}'].alignment = center
    ws[f'K{r_33}'].number_format = JPY_FMT

    r_34 = r_33 + 1
    ws[f'A{r_34}'] = f'架台+杭　消費税{consumption_tax_pct}％'
    ws[f'A{r_34}'].font = SM_FONT
    ws[f'A{r_34}'].alignment = right_a
    ws.merge_cells(f'A{r_34}:I{r_34}')
    ws[f'J{r_34}'] = f'=(J{r19}+J{r_33})*{consumption_tax_pct}/100'
    ws[f'J{r_34}'].font = SM_FONT
    ws[f'J{r_34}'].alignment = center
    ws[f'J{r_34}'].number_format = CURRENCY_FMT
    ws[f'K{r_34}'] = f'=J{r_34}*E{r_32}'
    ws[f'K{r_34}'].font = SM_FONT
    ws[f'K{r_34}'].alignment = center
    ws[f'K{r_34}'].number_format = JPY_FMT

    r_35 = r_34 + 1
    ws[f'A{r_35}'] = '③架台+杭　関税'
    ws[f'A{r_35}'].font = SM_FONT_RED
    ws[f'A{r_35}'].alignment = right_a
    ws.merge_cells(f'A{r_35}:I{r_35}')
    ws[f'J{r_35}'] = f'=J{r_33}+J{r_34}'
    ws[f'J{r_35}'].font = SM_FONT
    ws[f'J{r_35}'].alignment = center
    ws[f'J{r_35}'].number_format = CURRENCY_FMT
    ws[f'K{r_35}'] = f'=J{r_35}*E{r_32}'
    ws[f'K{r_35}'].font = SM_FONT
    ws[f'K{r_35}'].alignment = center
    ws[f'K{r_35}'].number_format = JPY_FMT

    r_36 = r_35 + 1
    ws[f'A{r_36}'] = f'④フェンス税金{fence_tax_pct}％'
    ws[f'A{r_36}'].font = SM_FONT_RED
    ws[f'A{r_36}'].alignment = right_a
    ws.merge_cells(f'A{r_36}:I{r_36}')
    ws[f'J{r_36}'] = f'=J{r_f2}*{fence_tax_pct}/100'
    ws[f'J{r_36}'].font = SM_FONT
    ws[f'J{r_36}'].alignment = center
    ws[f'J{r_36}'].number_format = CURRENCY_FMT
    ws[f'K{r_36}'] = f'=J{r_36}*E{r_32}'
    ws[f'K{r_36}'].font = SM_FONT
    ws[f'K{r_36}'].alignment = center
    ws[f'K{r_36}'].number_format = JPY_FMT

    r_37 = r_36 + 1
    ws[f'A{r_37}'] = '⑤'
    ws[f'A{r_37}'].font = SM_FONT_RED
    ws[f'A{r_37}'].alignment = center
    ws.merge_cells(f'B{r_37}:E{r_37}')
    ws[f'B{r_37}'] = f'=C4'
    ws[f'B{r_37}'].font = SM_FONT
    ws[f'B{r_37}'].alignment = center
    ws[f'F{r_37}'] = '混載便'
    ws[f'F{r_37}'].font = SM_FONT
    ws[f'F{r_37}'].alignment = center
    ws[f'G{r_37}'] = shipping_data.get('truck_desc', '') or '4Tユニック+4T平車 配送'
    ws[f'G{r_37}'].font = SM_FONT
    ws[f'G{r_37}'].alignment = center
    ws.merge_cells(f'G{r_37}:I{r_37}')
    ws[f'J{r_37}'] = shipping_data.get('truck_fee', 0) or 0
    ws[f'J{r_37}'].font = SM_FONT
    ws[f'J{r_37}'].alignment = center
    ws[f'J{r_37}'].number_format = CURRENCY_FMT
    ws[f'K{r_37}'] = f'=J{r_37}*E{r_32}'
    ws[f'K{r_37}'].font = SM_FONT
    ws[f'K{r_37}'].alignment = center
    ws[f'K{r_37}'].number_format = JPY_FMT

    r_38 = r_37 + 1
    if has_fence:
        ws[f'A{r_38}'] = '①+②+③+④+⑤総金額(USD)'
    else:
        ws[f'A{r_38}'] = '①+③+⑤総金額(USD)'
    ws[f'A{r_38}'].font = SM_FONT_RED
    ws[f'A{r_38}'].alignment = right_a
    ws.merge_cells(f'A{r_38}:I{r_38}')
    if has_fence:
        ws[f'J{r_38}'] = f'=J{r19}+J{r_f2}+J{r_35}+J{r_36}+J{r_37}'
    else:
        ws[f'J{r_38}'] = f'=J{r19}+J{r_35}+J{r_37}'
    ws[f'J{r_38}'].font = SM_FONT
    ws[f'J{r_38}'].alignment = center
    ws[f'J{r_38}'].number_format = CURRENCY_FMT
    ws[f'K{r_38}'] = f'=J{r_38}*E{r_32}'
    ws[f'K{r_38}'].font = SM_FONT
    ws[f'K{r_38}'].alignment = center
    ws[f'K{r_38}'].number_format = JPY_FMT

    r_39 = r_38 + 1
    ws[f'A{r_39}'] = '請求金額'
    ws[f'A{r_39}'].font = SM_FONT_RED_10
    ws[f'A{r_39}'].alignment = right_a
    ws.merge_cells(f'A{r_39}:I{r_39}')
    if has_fence:
        ws[f'J{r_39}'] = f'=J{r_38}-J{r_f2}'
    else:
        ws[f'J{r_39}'] = f'=J{r_38}'
    ws[f'J{r_39}'].font = SM_FONT
    ws[f'J{r_39}'].alignment = center
    ws[f'J{r_39}'].number_format = CURRENCY_FMT
    ws[f'K{r_39}'] = f'=J{r_39}*E{r_32}'
    ws[f'K{r_39}'].font = SM_FONT
    ws[f'K{r_39}'].alignment = center
    ws[f'K{r_39}'].number_format = JPY_FMT

    for rr in range(r_31, r_39 + 1):
        for c in range(1, 12):
            ws.cell(row=rr, column=c).fill = GREEN_FILL

    # ========== Row 40+: 二、運賃（DDP） ==========
    r_40 = r_39 + 1
    ws[f'A{r_40}'] = '二、運賃'
    ws[f'A{r_40}'].font = SM_FONT_BOLD_10
    ws[f'A{r_40}'].alignment = left_a
    ws.merge_cells(f'A{r_40}:K{r_40}')
    ws.row_dimensions[r_40].height = 20

    r_41 = r_40 + 1
    ws[f'A{r_41}'] = '項目'
    ws[f'A{r_41}'].font = SM_FONT
    ws[f'A{r_41}'].alignment = center
    ws.merge_cells(f'A{r_41}:I{r_41}')
    ws[f'J{r_41}'] = '金額（USD）'
    ws[f'J{r_41}'].font = SM_FONT
    ws[f'J{r_41}'].alignment = center
    ws[f'K{r_41}'] = '金額(JPY)'
    ws[f'K{r_41}'].font = SM_FONT
    ws[f'K{r_41}'].alignment = center

    r_42 = r_41 + 1
    ws.merge_cells(f'A{r_42}:H{r_42}')
    ws[f'I{r_42}'] = '①架台本体一式価格'
    ws[f'I{r_42}'].font = SM_FONT_RED
    ws[f'I{r_42}'].alignment = center
    if has_fence:
        ws[f'J{r_42}'] = f'=J{r19}+J{r_f2}'
    else:
        ws[f'J{r_42}'] = f'=J{r19}'
    ws[f'J{r_42}'].font = SM_FONT
    ws[f'J{r_42}'].alignment = center
    ws[f'J{r_42}'].number_format = CURRENCY_FMT
    ws[f'K{r_42}'] = f'=J{r_42}*E{r_32}'
    ws[f'K{r_42}'].font = SM_FONT
    ws[f'K{r_42}'].alignment = center
    ws[f'K{r_42}'].number_format = JPY_FMT

    r_43 = r_42 + 1
    ws[f'A{r_43}'] = f'=B{r_37}'
    ws[f'A{r_43}'].font = SM_FONT
    ws[f'A{r_43}'].alignment = center
    ws.merge_cells(f'A{r_43}:D{r_43}')
    ws[f'E{r_43}'] = f'=F{r_37}'
    ws[f'E{r_43}'].font = SM_FONT
    ws[f'E{r_43}'].alignment = center
    ws.merge_cells(f'E{r_43}:F{r_43}')
    ws[f'G{r_43}'] = f'=G{r_37}'
    ws[f'G{r_43}'].font = SM_FONT
    ws[f'G{r_43}'].alignment = center
    ws.merge_cells(f'G{r_43}:H{r_43}')
    ws[f'I{r_43}'] = '②DDP現地配送費'
    ws[f'I{r_43}'].font = SM_FONT_RED
    ws[f'I{r_43}'].alignment = center
    ws[f'J{r_43}'] = f'=J{r_43 + 3}-J{r_43 + 2}-J{r_42}'
    ws[f'J{r_43}'].font = SM_FONT
    ws[f'J{r_43}'].alignment = center
    ws[f'J{r_43}'].number_format = CURRENCY_FMT
    ws[f'K{r_43}'] = f'=J{r_43}*E{r_32}'
    ws[f'K{r_43}'].font = SM_FONT
    ws[f'K{r_43}'].alignment = center
    ws[f'K{r_43}'].number_format = JPY_FMT

    r_44 = r_43 + 1
    ws[f'A{r_44}'] = '①+②='
    ws[f'A{r_44}'].font = SM_FONT_RED
    ws[f'A{r_44}'].alignment = right_a
    ws.merge_cells(f'A{r_44}:H{r_44}')
    ws[f'I{r_44}'] = '③小計'
    ws[f'I{r_44}'].font = SM_FONT_RED
    ws[f'I{r_44}'].alignment = center
    ws[f'J{r_44}'] = f'=J{r_42}+J{r_43}'
    ws[f'J{r_44}'].font = SM_FONT
    ws[f'J{r_44}'].alignment = center
    ws[f'J{r_44}'].number_format = CURRENCY_FMT
    ws[f'K{r_44}'] = f'=J{r_44}*E{r_32}'
    ws[f'K{r_44}'].font = SM_FONT
    ws[f'K{r_44}'].alignment = center
    ws[f'K{r_44}'].number_format = JPY_FMT

    r_45 = r_44 + 1
    ws.merge_cells(f'A{r_45}:H{r_45}')
    ws[f'I{r_45}'] = f'④消費税{consumption_tax_pct}％'
    ws[f'I{r_45}'].font = SM_FONT_RED
    ws[f'I{r_45}'].alignment = center
    ws[f'J{r_45}'] = f'=J{r_45 + 1}/11'
    ws[f'J{r_45}'].font = SM_FONT
    ws[f'J{r_45}'].alignment = center
    ws[f'J{r_45}'].number_format = CURRENCY_FMT
    ws[f'K{r_45}'] = f'=J{r_45}*E{r_32}'
    ws[f'K{r_45}'].font = SM_FONT
    ws[f'K{r_45}'].alignment = center
    ws[f'K{r_45}'].number_format = JPY_FMT

    r_46 = r_45 + 1
    ws[f'A{r_46}'] = '①+②+③＋④='
    ws[f'A{r_46}'].font = SM_FONT_RED
    ws[f'A{r_46}'].alignment = right_a
    ws.merge_cells(f'A{r_46}:H{r_46}')
    ws[f'I{r_46}'] = '⑤合計(税込み)'
    ws[f'I{r_46}'].font = SM_FONT_RED
    ws[f'I{r_46}'].alignment = center
    ws[f'J{r_46}'] = f'=J{r_38}'
    ws[f'J{r_46}'].font = SM_FONT
    ws[f'J{r_46}'].alignment = center
    ws[f'J{r_46}'].number_format = CURRENCY_FMT
    ws[f'K{r_46}'] = f'=J{r_46}*E{r_32}'
    ws[f'K{r_46}'].font = SM_FONT
    ws[f'K{r_46}'].alignment = center
    ws[f'K{r_46}'].number_format = JPY_FMT

    for rr in range(r_41, r_46 + 1):
        for c in range(1, 12):
            ws.cell(row=rr, column=c).fill = YELLOW_FILL

    r_47 = r_46 + 1
    ws[f'A{r_47}'] = 'レートウェッブ'
    ws[f'A{r_47}'].font = SM_FONT
    ws[f'A{r_47}'].alignment = center
    ws.merge_cells(f'A{r_47}:B{r_47}')
    ws[f'C{r_47}'] = 'http://www.safe.gov.cn/safe/rmbhlzjj/index.html'
    ws[f'C{r_47}'].font = SM_FONT
    ws[f'C{r_47}'].alignment = left_a
    ws.merge_cells(f'C{r_47}:K{r_47}')

    r_48 = r_47 + 1
    ws[f'A{r_48}'] = '原材料ウェッブ'
    ws[f'A{r_48}'].font = SM_FONT
    ws[f'A{r_48}'].alignment = center
    ws.merge_cells(f'A{r_48}:B{r_48}')
    ws[f'C{r_48}'] = 'https://market.cnal.com/nanhai/'
    ws[f'C{r_48}'].font = SM_FONT
    ws[f'C{r_48}'].alignment = left_a
    ws.merge_cells(f'C{r_48}:K{r_48}')

    # ========== 边框处理 ==========
    for row in range(10, r21 + 1):
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = thin_border

    r_border_start = r22 if has_fence_data else r_31
    for row in range(r_border_start, r_48 + 1):
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = thin_border

    data_regions = [(10, r21, 1, 11), (r_border_start, r_48, 1, 11)]
    for min_row, max_row, min_col, max_col in data_regions:
        region_merges = [
            str(mr) for mr in list(ws.merged_cells.ranges)
            if mr.min_row >= min_row and mr.max_row <= max_row
            and mr.min_col >= min_col and mr.max_col <= max_col
        ]
        for mr_str in region_merges:
            ws.unmerge_cells(mr_str)

        thick_side = Side(style='medium', color='000000')
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                new_left = thick_side if col == min_col else cell.border.left
                new_right = thick_side if col == max_col else cell.border.right
                new_top = thick_side if row == min_row else cell.border.top
                new_bottom = thick_side if row == max_row else cell.border.bottom
                cell.border = Border(left=new_left, right=new_right,
                                     top=new_top, bottom=new_bottom)

        for mr_str in region_merges:
            ws.merge_cells(mr_str)

    sheet_names = workbook.sheetnames
    if '合計' in sheet_names:
        idx = sheet_names.index('合計')
        workbook.move_sheet('合計', offset=-idx)

    apply_print_setup(ws, 'ja')

    return ws.title
