import os
import re
import tempfile
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from PIL import Image as PILImage
from backend.core.shared.constants import IMAGE_WIDTH, IMAGE_HEIGHT, IMAGE_COLUMN_INDEX, IMAGE_PADDING


def extract_code_from_filename(filename):
    name_without_ext = os.path.splitext(filename)[0]

    patterns = [
        r'^([A-Z]+-\d+)(?:-\d+)?[_\-\s\u4e00-\u9fff]',
        r'^([A-Z]+-\d+)(?:-\d+)?$',
        r'^([A-Z]+\d+[xX]\d+)[_\-\s\u4e00-\u9fff]',
        r'^([A-Z]+\d+[xX]\d+)$',
        r'^([A-Z]+-[A-Z0-9]+-\d+)[_\-\s\u4e00-\u9fff]',
        r'^([A-Z]+-[A-Z0-9]+-\d+)$',
        r'^([A-Z]+-\d+)[_\-\s\u4e00-\u9fff]',
        r'^([A-Z]+-\d+)$',
        r'^([A-Z]+\d+)[_\-\s\u4e00-\u9fff]',
        r'^([A-Z]+\d+)$',
    ]

    for pattern in patterns:
        match = re.match(pattern, name_without_ext, re.IGNORECASE)
        if match:
            code = match.group(1).upper()
            if code and code[0].isalpha() and any(c.isdigit() for c in code):
                return code

    code_patterns = [
        r'([A-Z]+-\d+)(?:-\d+)?',
        r'([A-Z]+\d+[xX]\d+)',
        r'([A-Z]+-[A-Z0-9]+-\d+)',
        r'([A-Z]+-\d+)',
        r'([A-Z]+\d+)',
    ]

    for pattern in code_patterns:
        match = re.search(pattern, name_without_ext, re.IGNORECASE)
        if match:
            code = match.group(1).upper()
            if code and code[0].isalpha() and any(c.isdigit() for c in code):
                return code

    return None


def scan_images(image_folder):
    code_to_images = {}
    extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.JPG', '.PNG', '.JPEG'}

    if not image_folder or not os.path.exists(image_folder):
        return code_to_images

    for root, _, files in os.walk(image_folder):
        for filename in files:
            ext = os.path.splitext(filename)[1]
            if ext in extensions:
                code = extract_code_from_filename(filename)
                if code:
                    code_to_images.setdefault(code, []).append(os.path.join(root, filename))

    return code_to_images


def find_latest_image_log(search_dirs):
    latest_path = None
    latest_mtime = 0
    for d in search_dirs:
        if not d or not os.path.exists(d):
            continue
        for name in os.listdir(d):
            if name.startswith("图片匹配记录_") and name.endswith(".txt"):
                path = os.path.join(d, name)
                try:
                    mtime = os.path.getmtime(path)
                except:
                    continue
                if mtime > latest_mtime:
                    latest_mtime = mtime
                    latest_path = path
    return latest_path


def load_image_mapping_from_log(log_path):
    mapping = {}
    if not log_path or not os.path.exists(log_path):
        return mapping

    try:
        with open(log_path, "r", encoding="utf-8", errors="ignore") as f:
            current_code = None
            for line in f:
                line = line.strip()
                if not line:
                    continue

                code_match = re.search(r'^(?:编码|缂栫爜|Code)\\s*[:：]\\s*([A-Za-z0-9\\-]+)', line)
                if code_match:
                    current_code = code_match.group(1).strip()
                    continue

                path_match = re.search(r'([A-Za-z]:\\\\.+?\\.(?:jpg|jpeg|png|gif|bmp|webp))', line, re.IGNORECASE)
                if path_match and current_code:
                    img_path = path_match.group(1).strip()
                    if os.path.exists(img_path):
                        mapping[current_code] = [img_path]
    except Exception:
        return mapping

    return mapping


def prepare_image_for_excel(image_path, target_width=IMAGE_WIDTH, target_height=IMAGE_HEIGHT, temp_dir=None,
                            cache=None):
    if not image_path or not os.path.exists(image_path):
        return None

    if cache is not None and image_path in cache:
        return cache[image_path]

    if temp_dir is None:
        temp_dir = tempfile.gettempdir()
    os.makedirs(temp_dir, exist_ok=True)

    temp_filename = f"temp_{abs(hash(image_path))}_{os.path.basename(image_path)}"
    temp_filename = os.path.splitext(temp_filename)[0] + ".jpg"
    temp_path = os.path.join(temp_dir, temp_filename)

    if os.path.exists(temp_path):
        if cache is not None:
            cache[image_path] = temp_path
        return temp_path

    try:
        img = PILImage.open(image_path)

        if img.mode in ('RGBA', 'LA', 'P'):
            rgb_img = PILImage.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P':
                img = img.convert('RGBA')
            rgb_img.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
            img = rgb_img

        img.thumbnail((target_width, target_height), PILImage.Resampling.LANCZOS)
        img.save(temp_path, "JPEG", quality=85)

        if cache is not None:
            cache[image_path] = temp_path
        return temp_path
    except Exception:
        return None


def _col_width_to_px(ws, cell_row, cell_col):
    col_letter = get_column_letter(cell_col)
    width = ws.column_dimensions[col_letter].width or 8.43
    try:
        from openpyxl.utils import column_width_to_pixels
        return column_width_to_pixels(width)
    except ImportError:
        return int(width * 7.5 + 5)


def _row_height_to_px(ws, cell_row, cell_col):
    height = ws.row_dimensions[cell_row].height or 15
    try:
        from openpyxl.utils import row_height_to_pixels
        return row_height_to_pixels(height)
    except ImportError:
        return int(height * 1.33)


def add_image_centered_in_cell(
        ws,
        img_path,
        cell_row,
        cell_col,
        img_width=IMAGE_WIDTH,
        img_height=IMAGE_HEIGHT,
):
    try:
        img = XLImage(img_path)
        img.width = img_width
        img.height = img_height

        col_width_pixels = _col_width_to_px(ws, cell_row, cell_col)
        row_height_pixels = _row_height_to_px(ws, cell_row, cell_col)

        offset_x = max(0, (col_width_pixels - img_width) / 2)
        offset_y = max(0, (row_height_pixels - img_height) / 2)

        EMU_PER_PIXEL = 9525
        marker = AnchorMarker(
            col=cell_col - 1,
            colOff=int(offset_x * EMU_PER_PIXEL),
            row=cell_row - 1,
            rowOff=int(offset_y * EMU_PER_PIXEL),
        )
        img.anchor = OneCellAnchor(
            _from=marker,
            ext=XDRPositiveSize2D(
                cx=int(img_width * EMU_PER_PIXEL),
                cy=int(img_height * EMU_PER_PIXEL),
            ),
        )

        ws.add_image(img)
        return True
    except Exception:
        try:
            img = XLImage(img_path)
            img.width = img_width
            img.height = img_height
            cell_ref = get_column_letter(cell_col) + str(cell_row)
            ws.add_image(img, cell_ref)
            return True
        except Exception:
            return False


def add_image_centered_in_range(ws, img, start_col, start_row, end_col, end_row):
    try:
        from openpyxl.utils import column_width_to_pixels, row_height_to_pixels
        use_utils = True
    except Exception:
        use_utils = False

    def _col_px(col_idx):
        col_letter = get_column_letter(col_idx)
        width = ws.column_dimensions[col_letter].width or 8.43
        if use_utils:
            try:
                return column_width_to_pixels(width)
            except Exception:
                pass
        return width * 7.5

    def _row_px(row_idx):
        height = ws.row_dimensions[row_idx].height or 15
        if use_utils:
            try:
                return row_height_to_pixels(height)
            except Exception:
                pass
        return height * 1.33

    total_width = sum(_col_px(col) for col in range(start_col, end_col + 1))
    total_height = sum(_row_px(row) for row in range(start_row, end_row + 1))

    img_width = img.width
    img_height = img.height

    offset_x = max(0, (total_width - img_width) / 2)
    offset_y = max(0, (total_height - img_height) / 2)

    EMU_PER_PIXEL = 9525
    offset_x_emu = int(offset_x * EMU_PER_PIXEL)
    offset_y_emu = int(offset_y * EMU_PER_PIXEL)

    anchor = OneCellAnchor(
        _from=AnchorMarker(
            col=start_col - 1,
            colOff=offset_x_emu,
            row=start_row - 1,
            rowOff=offset_y_emu
        ),
        ext=XDRPositiveSize2D(
            cx=int(img.width * EMU_PER_PIXEL),
            cy=int(img.height * EMU_PER_PIXEL),
        ),
    )
    img.anchor = anchor
    ws.add_image(img)
    return True


def center_images_in_column(ws, target_col=IMAGE_COLUMN_INDEX, scale_to_fit=False, max_scale=1.0):
    EMU_PER_PIXEL = 9525
    default_col_width = 8.43
    default_row_height = 15

    try:
        from openpyxl.utils import column_width_to_pixels, row_height_to_pixels
        use_utils = True
    except Exception:
        use_utils = False

    def _col_px(col_idx):
        col_letter = get_column_letter(col_idx)
        width = ws.column_dimensions[col_letter].width
        if width is None:
            width = default_col_width
        if use_utils:
            try:
                return column_width_to_pixels(width)
            except Exception:
                pass
        return width * 7.5

    def _row_px(row_idx):
        height = ws.row_dimensions[row_idx].height
        if height is None:
            height = default_row_height
        if use_utils:
            try:
                return row_height_to_pixels(height)
            except Exception:
                pass
        return height * 1.33

    processed = 0
    for img in list(ws._images):
        anchor = img.anchor

        if hasattr(anchor, '_from'):
            from_col = anchor._from.col + 1
            from_row = anchor._from.row + 1
        elif hasattr(anchor, 'from_col'):
            from_col = anchor.from_col + 1
            from_row = anchor.from_row + 1
        else:
            continue

        if from_col != target_col:
            continue

        cell_width_px = _col_px(from_col)
        cell_height_px = _row_px(from_row)

        img_width = img.width
        img_height = img.height

        if scale_to_fit and img_width and img_height:
            scale_w = cell_width_px / img_width
            scale_h = cell_height_px / img_height
            scale = min(scale_w, scale_h, max_scale)
            if scale < 1.0:
                img_width = int(img_width * scale)
                img_height = int(img_height * scale)
                img.width = img_width
                img.height = img_height

        offset_x = (cell_width_px - img_width) / 2
        offset_y = (cell_height_px - img_height) / 2
        offset_x = max(0, offset_x)
        offset_y = max(0, offset_y)

        offset_x_emu = int(offset_x * EMU_PER_PIXEL)
        offset_y_emu = int(offset_y * EMU_PER_PIXEL)

        if hasattr(anchor, '_from'):
            anchor._from.colOff = offset_x_emu
            anchor._from.rowOff = offset_y_emu
        elif hasattr(anchor, 'from_col_off'):
            anchor.from_col_off = offset_x_emu
            anchor.from_row_off = offset_y_emu

        processed += 1

    return processed
