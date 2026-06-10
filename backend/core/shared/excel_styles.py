from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

THICK_SIDE = Side(style='medium', color='000000')

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT_A = Alignment(horizontal='right', vertical='center', wrap_text=True)
LEFT_A = Alignment(horizontal='left', vertical='center', wrap_text=True)

CURRENCY_FMT = '"US$" #,##0.00'
JPY_FMT = '[$￥-411]#,##0;-[$￥-411]#,##0'

BLUE_FILL = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
GREEN_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
HEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
GRAY_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='F3D4A9', end_color='F3D4A9', fill_type='solid')


def _apply_outer_border(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cur = cell.border
            cell.border = Border(
                left=THICK_SIDE if c == min_col else cur.left,
                right=THICK_SIDE if c == max_col else cur.right,
                top=THICK_SIDE if r == min_row else cur.top,
                bottom=THICK_SIDE if r == max_row else cur.bottom,
            )


def _set_cell(ws, row, col, val, font=None, align=None, border=None, number_format=None, fill=None):
    cell = ws.cell(row=row, column=col, value=val)
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format
    if fill is not None:
        cell.fill = fill
    return cell
