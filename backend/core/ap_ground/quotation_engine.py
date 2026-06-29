"""亚太组（Asia-Pacific）地面支架报价单引擎（隔离模块）。

对应参考样例：D:\\科盛\\input\\亚太\\地面案件模板1\\【KSENG Quotation】...006(1).xlsx
币种：USD（美元）。

输出结构（与 006 样例一致）：
    1) 每个阵列一个明细页（{rows}x{cols}x{tables}），抬头含 Sales / Installation Angle /
       Panel Size / Max Wind Load / Power-PC / Watt(W)/Table / Span / Array / 质保，
       BOM 表 14 列（A~N）：
           Item No. | Product Name | Material | Picture | Spec. |
           Unit Price(USD) EXW | Discount Unit Price(USD) EXW |
           QTY(PCS) | Total QTY(PCS) | Spare Parts |
           Total price(USD) EXW | Discount Total price(USD) EXW |
           Discount Total price Of Spare Parts(USD) EXW | Remark
       其中价格列按「每基（per table）」计算（与 006 一致），汇总页再乘台数。
    2) 一个汇总页（Total）：每阵列一行（NO/Row/Column/Installation/Angle/Table/Power kw/
       Price(USD)/Table / Amount(USD) / Total Amount after Special Discount）+ 抬头 +
       EXW TOTAL AMOUNT + Unite Price/W。

计价说明：
    - 单价(原价) 直接取数据库价格（USD，长度物料按规格折算），不做汇率换算；
    - 折扣单价 = 原价 × 折扣率/100（按 standard/steel/purchased 分类）；
    - 备品 Spare Parts = Total QTY × 1%；
    - Sales / 联系人信息由前端 contact_info 提供，未提供则留空。

本模块与 ap_common / en_simple 完全隔离，自行实现分类、取价、渲染逻辑，
不修改上述既有模块。
"""

import os
import sys
from decimal import Decimal

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from backend.core.print_settings import apply_print_setup

from backend.core.shared.bom_utils import resolve_products_and_array
from backend.core.shared.price_utils import (
    resolve_price_info,
    has_valid_price_info,
    get_temp_base_price,
    apply_temp_preinstall_adjustment,
    _get_discount_category,
)
from backend.core.shared.weight_utils import extract_length_from_spec
from backend.core.shared.text_utils import (
    extract_main_name,
    normalize_lookup_code,
    _strip_cjk_spec,
    normalize_angle,
)
from backend.core.shared.product_utils import (
    _is_valid_product_code,
    _match_exclude_group,
    normalize_preinstall,
)
from backend.core.shared.constants import (
    EXCLUDE_ITEM_GROUPS,
    CARBON_STEEL_PRICING_ATTRS,
    IMAGE_PADDING,
)
from backend.core.shared.image_utils import (
    prepare_image_for_excel,
    add_image_centered_in_cell,
)
from backend.core.material_translate import translate_material, adjust_material_by_coating

if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

CURRENCY_LABEL = 'USD'
CURRENCY_FMT = '"$"#,##0.00'

# 图片所在列：E（第 5 列）——Item No. 占 A-B 两列后顺移
IMAGE_COL_INDEX = 5
AP_GROUND_IMAGE_WIDTH = 372
AP_GROUND_IMAGE_HEIGHT = 92

# 备品比例：Total QTY × 1%
SPARE_RATE = 0.01

# Item No. 占两列（A-B 合并），其余列整体右移 1 列，故明细表共 15 列（A~O）
COL_ITEM_NO_START = get_column_letter(1)    # A
COL_ITEM_NO_END = get_column_letter(2)      # B

# BOM 明细表头（15 列，A~O）：Item No. 占 A-B
DETAIL_HEADERS = [
    'Item No.',
    'Product Name',
    'Material',
    'Picture',
    'Spec.',
    'Unit Price (USD)\nEXW',
    'Discount \nUnit Price\n(USD) EXW',
    'QTY (PCS)',
    'Total QTY (PCS)',
    'Spare Parts',
    'Total price(USD)\nEXW',
    'Discount Total price(USD)\nEXW',
    'Discount Total price Of Spare Parts\n(USD) EXW',
    'Remark',
]
# 各数据列在表头中的逻辑序号（1-based，对应 DETAIL_HEADERS 列表下标+1）：
# Item No.=1-2(合并), Product Name=3, Material=4, Picture=5, Spec.=6,
# Unit Price=7, Discount Unit=8, QTY=9, Total QTY=10, Spare=11,
# Total price=12, Discount Total=13, Discount Spare=14, Remark=15
MAX_COL = 15  # A~O
MAX_COL_LETTER = get_column_letter(MAX_COL)
# 价格列字母（便于汇总页跨表引用）
COL_PRODUCT_NAME = get_column_letter(3)     # C
COL_MATERIAL = get_column_letter(4)         # D
COL_PICTURE = get_column_letter(5)          # E
COL_SPEC = get_column_letter(6)             # F
COL_UNIT_PRICE = get_column_letter(7)       # G 原价
COL_DISC_UNIT = get_column_letter(8)        # H 折扣单价
COL_QTY = get_column_letter(9)              # I 每基数量
COL_TOTAL_QTY = get_column_letter(10)       # J 总数量
COL_SPARE = get_column_letter(11)           # K 备品
COL_TOTAL_PRICE = get_column_letter(12)     # L 每基原价合计
COL_DISC_TOTAL = get_column_letter(13)      # M 每基折扣合计
COL_DISC_SPARE = get_column_letter(14)      # N 备品折扣合计
COL_REMARK = get_column_letter(15)          # O 备注(编码)
# 抬头信息块各合并区右移 1 列后的起止列字母
LOGO_MERGE = 'A2:C6'              # Logo 区（A 列窄边距 + B + C）
CONTACT_MERGE_COL_START = 'D'     # 联系人/Array（D-E 合并）
CONTACT_MERGE_COL_END = 'E'
LABEL1_COL = 'F'                  # Installation Angle / Max Wind / ... 标签
VALUE1_COL = 'G'                  # 对应值
LABEL2_COL = 'H'                  # Panel Size / Power-PC / ... 标签
VALUE2_COL_START = 'I'            # 值2（I-K 合并）
VALUE2_COL_END = 'K'
WARRANTY_COL_START = 'L'          # 质保（L-O 合并）
WARRANTY_COL_END = MAX_COL_LETTER


def _set_cell(ws, row, col, val, font=None, align=None, border=None,
              number_format=None, fill=None):
    cell = ws.cell(row=row, column=col, value=val)
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format
    if fill is not None:
        cell.fill = fill
    return cell


def _apply_outer_frame(ws, top_row, bottom_row, left_col, right_col,
                       weight='medium', color='000000'):
    """在指定矩形外缘绘制粗边框（合并单元格重定向到锚点单元格，保留内部边框）。"""
    side = Side(style=weight, color=color)

    def _anchor(coord):
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                return rng.start_cell
        return ws[coord]

    def _merge_border(cell, **edges):
        ex = cell.border
        cell.border = Border(
            left=edges.get('left', ex.left),
            right=edges.get('right', ex.right),
            top=edges.get('top', ex.top),
            bottom=edges.get('bottom', ex.bottom),
        )

    for r in range(top_row, bottom_row + 1):
        lc = _anchor(ws.cell(row=r, column=left_col).coordinate)
        rc = _anchor(ws.cell(row=r, column=right_col).coordinate)
        _merge_border(lc, left=side)
        _merge_border(rc, right=side)
    for c in range(left_col, right_col + 1):
        tc = _anchor(ws.cell(row=top_row, column=c).coordinate)
        bc = _anchor(ws.cell(row=bottom_row, column=c).coordinate)
        _merge_border(tc, top=side)
        _merge_border(bc, bottom=side)


def _classify_products(all_products, price_mapping, delete_options=None,
                       always_exclude=False, ap_exclude_options=None):
    """分类：主物料 / 排除(可选)项 / 地桩（迁移自 ap_common，隔离副本）。"""
    delete_options = delete_options or {}
    ap_exclude_options = ap_exclude_options or {}
    all_exclude_opts = {k: True for k in EXCLUDE_ITEM_GROUPS} if always_exclude else {}
    for k in list(ap_exclude_options.keys()) + list(delete_options.keys()):
        all_exclude_opts[k] = True

    main = []
    excluded = []
    pile_products = []
    price_info_cache = {}
    for p in all_products:
        code = str(p.get('code', '')).strip()
        spec = str(p.get('spec', '')).strip()
        cache_key = (code, spec) if spec else code
        if cache_key not in price_info_cache:
            price_info_cache[cache_key] = (
                resolve_price_info(price_mapping, code, spec=spec) if price_mapping else None
            )
        pi = price_info_cache[cache_key]
        p['_price_info'] = pi
        attr = (pi.get('attribute', '') if pi else '') or ''
        if attr.strip() in ('地桩', '地盤杭'):
            p['_is_pile'] = True
            pile_products.append(p)
            continue
        if not _is_valid_product_code(code):
            continue
        qty = p.get('quantity', 0)
        if not qty or qty <= 0:
            continue
        matched_group = _match_exclude_group(p, price_mapping, all_exclude_opts)
        if matched_group:
            if delete_options.get(matched_group):
                continue
            if ap_exclude_options.get(matched_group):
                excluded.append(p)
            else:
                main.append(p)
        else:
            main.append(p)
    return main, excluded, pile_products, price_info_cache


def _calc_base_display_price(product, price_mapping, group, sale_type):
    """基础单价(长度折算后、未折扣) → 原价展示单价（直接取数据库价格）。"""
    code = product.get('code', '')
    price_info = resolve_price_info(price_mapping, code, spec=product.get('spec', ''))
    if not price_info or not has_valid_price_info(price_info):
        return 0.0, '', price_info
    base_price = float(get_temp_base_price(price_info, product, group, sale_type))
    price_unit = price_info.get('unit', '')
    is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
    if is_meter:
        length_mm = extract_length_from_spec(product.get('spec', '')) or 0
        if length_mm > 0:
            base_price = base_price * length_mm / 1000.0
    base_price = apply_temp_preinstall_adjustment(price_info, base_price, product, group, sale_type)
    return base_price, price_unit, price_info


def collect_ap_ground_array(df, price_mapping, config=None, matrix_data=None,
                            pre_parsed_products=None, sale_type='export',
                            coating_thickness=10, delete_options=None,
                            always_exclude_extra_items=False, ap_exclude_options=None,
                            unmatched_products_list=None, module_wattage=None, **kwargs):
    """收集单个阵列的产品与站点信息（不渲染）。

    返回 site + items。items 同时保留 qty_per_table（每基数量）与 qty_total（总数量），
    以及未折扣原价 unit_price，供明细页 F/G/K 列使用。
    """
    group = '亚太组'
    all_products, array_info, span_info, rows, cols = resolve_products_and_array(
        pre_parsed_products, df, matrix_data,
    )
    main_products, excluded_products, pile_products, _pic = _classify_products(
        all_products, price_mapping, delete_options or {},
        always_exclude_extra_items, ap_exclude_options or {},
    )

    matrix_data = matrix_data or {}
    set_count = matrix_data.get('set_count') or 1
    if not isinstance(set_count, int) or set_count <= 0:
        set_count = 1
    output_kw = matrix_data.get('output_kw') or 0
    if not output_kw and module_wattage and rows and cols:
        try:
            output_kw = float(rows) * float(cols) * float(set_count) * float(module_wattage) / 1000.0
        except (TypeError, ValueError):
            output_kw = 0
    project_name = str(matrix_data.get('project_name') or '').strip()

    render_products = main_products + excluded_products
    items = []
    matched_count = 0
    unmatched_count = 0
    local_unmatched = []
    for product in render_products:
        code = str(product.get('code', '')).strip()
        quantity = product.get('quantity', 0)

        display_price, price_unit, price_info = _calc_base_display_price(
            product, price_mapping, group, sale_type
        )
        is_matched = bool(price_info and has_valid_price_info(price_info))
        if is_matched:
            matched_count += 1
        else:
            unmatched_count += 1
            if unmatched_products_list is not None and _is_valid_product_code(code):
                local_unmatched.append({
                    'code': code,
                    'name': product.get('name', ''),
                    'spec': product.get('spec', ''),
                    'quantity': quantity,
                    'unit': price_info.get('unit', '') if price_info else '',
                    'preinstall': normalize_preinstall(product.get('preinstall')),
                    'issue_reason': '数据库无法匹配',
                })

        en_name = ''
        _name_src = price_info if price_info else (resolve_price_info(price_mapping, code) if price_mapping else None)
        if _name_src:
            en_name = (_name_src.get('name_en') or _name_src.get('name_ko') or '')
        raw_mat = ((price_info.get('db_material') if price_info and price_info.get('db_material') else None)
                   or (_name_src.get('db_material') if _name_src and _name_src.get('db_material') else None)
                   or product.get('material', ''))
        material = adjust_material_by_coating(translate_material(raw_mat, 'en'), coating_thickness)
        spec_val = product.get('spec', '')

        qty_per_table = float(quantity or 0)
        qty_total = qty_per_table * set_count
        discount_cat = _get_discount_category(price_info, product)
        unit_weight = product.get('weight') or 0
        items.append({
            'code': code,
            'name': en_name,
            'material': material,
            'spec': spec_val,
            'spec_norm': _strip_cjk_spec(spec_val),
            'unit_price': display_price,
            'qty_per_table': qty_per_table,
            'qty_total': qty_total,
            'is_matched': is_matched,
            'discount_cat': discount_cat,
            'unit_weight': unit_weight,
        })

    pile_scaled = []
    for pp in pile_products:
        s = dict(pp)
        s['quantity'] = float(pp.get('quantity', 0) or 0) * set_count
        pile_scaled.append(s)

    if unmatched_products_list is not None:
        for _up in local_unmatched:
            _up['quantity'] = float(_up.get('quantity', 0) or 0) * set_count
        unmatched_products_list.extend(local_unmatched)

    return {
        'site': {
            'project': project_name,
            'rows': rows,
            'cols': cols,
            'tables': set_count,
            'kw': output_kw,
        },
        'items': items,
        'pile_products': pile_scaled,
        'span_info': span_info,
        'matched': matched_count,
        'unmatched': unmatched_count,
    }


# ------------------------------------------------------------------ 明细页

def create_ap_ground_detail_sheet(
        workbook, site, items, price_mapping=None,
        matrix_data=None, config=None, contact_info=None,
        sale_type='export', coating_thickness=10,
        image_path=None, image_folder=None, code_to_images=None,
        image_temp_dir=None, image_cache=None,
        ap_discount_rate=100, ap_steel_discount_rate=100, ap_purchased_discount_rate=100,
        sheet_prefix=None, module_wattage=None,
):
    """创建单个阵列的明细工作表（14 列 A~N，抬头 + BOM）。"""
    matrix_data = matrix_data or {}
    config = config or {}
    contact_info = contact_info or {}

    rows = site.get('rows')
    cols = site.get('cols')
    tables = site.get('tables') or 1
    try:
        tables = int(tables)
    except (TypeError, ValueError):
        tables = 1

    sheet_name = sheet_prefix or ''
    if not sheet_name:
        sheet_name = f'{rows}x{cols}x{tables}' if rows and cols else 'Detail'
    sheet_name = extract_main_name(sheet_name) or sheet_name
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    original = sheet_name
    counter = 1
    while sheet_name in workbook.sheetnames:
        sheet_name = f"{original}_{counter}"
        counter += 1
    ws = workbook.create_sheet(title=sheet_name)
    print(f"[AP-GROUND] Creating detail sheet: {sheet_name}")

    thin = Side(style='thin', color='000000')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # 字体/颜色统一 Arial（表头无填充黑字）
    title_font = Font(name='Arial', size=26, bold=True, color='000000')
    info_font = Font(name='Arial', size=14, color='000000')
    info_bold = Font(name='Arial', size=14, bold=True, color='000000')
    header_font = Font(name='Arial', size=14, bold=True, color='000000')
    normal_font = Font(name='Arial', size=14, color='000000')
    bold_font = Font(name='Arial', size=14, bold=True, color='000000')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_a = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_a = Alignment(horizontal='right', vertical='center', wrap_text=True)

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # 列宽：A=1（窄边距），B=11（Item No. 右半），C~O=25
    ws.column_dimensions['A'].width = 1
    ws.column_dimensions['B'].width = 11
    for col in 'CDEFGHIJKLMNO':
        ws.column_dimensions[col].width = 25

    # ---- 抬头：标题 ----
    ws.merge_cells(f'A1:{MAX_COL_LETTER}1')
    _set_cell(ws, 1, 1, 'Solar PV Mounting System', font=title_font, align=center)
    ws.row_dimensions[1].height = 30

    # ---- 抬头信息块（行 2~6）----
    angle_val = config.get('angle', '') or matrix_data.get('angle', '')
    angle_display = normalize_angle(angle_val) if angle_val else ''
    panel_size = matrix_data.get('module_size') or config.get('panel_spec', '') or ''
    power_pc = matrix_data.get('module_wattage') or module_wattage or ''
    wind = matrix_data.get('max_wind_speed') or ''
    snow = matrix_data.get('max_snow_load') or ''
    watt_per_table = ''
    if power_pc and rows and cols:
        try:
            watt_per_table = int(float(power_pc) * float(rows) * float(cols))
        except (TypeError, ValueError):
            watt_per_table = ''
    span_info = site.get('span_info') or config.get('cross_span', '') or ''
    span_sn = span_ew = ''
    if span_info:
        span_ew = span_info

    contact_name = contact_info.get('contact_name') or ''
    phone = contact_info.get('phone') or ''
    tel = contact_info.get('tel') or ''

    # A2:C6 合并（Logo 区，A 为窄边距）
    ws.merge_cells(LOGO_MERGE)
    if image_path and os.path.exists(image_path):
        try:
            img = XLImage(img=image_path)
            _tot_w_px = sum(float(ws.column_dimensions[c].width or 8.43) for c in ('A', 'B', 'C')) * 7.0
            _tot_h_px = sum(float(ws.row_dimensions[r].height or 20) for r in range(2, 7)) * 1.33
            _ow, _oh = img.width, img.height
            _ratio = _ow / _oh
            _tr = _tot_w_px / _tot_h_px
            if _ratio > _tr:
                img.width = _tot_w_px; img.height = _tot_w_px / _ratio
            else:
                img.height = _tot_h_px; img.width = _tot_h_px * _ratio
            ws.add_image(img, 'A2')
        except Exception:
            _set_cell(ws, 2, 1, '', font=info_font, align=center, border=thin_border)

    # 抬头网格统一样式：框线 + 居中 + 垂直居中
    def _h(r, c, val, font=None):
        _set_cell(ws, r, c, val, font=font or info_font, align=center, border=thin_border)

    # 联系人 + Array（D-E 合并）
    _contact_start = 4  # D
    def _contact(r, val, font=None):
        ws.merge_cells(f'{CONTACT_MERGE_COL_START}{r}:{CONTACT_MERGE_COL_END}{r}')
        _h(r, _contact_start, val, font=font)

    _contact(2, f'Sales:{contact_name}' if contact_name else '')
    _contact(3, f'Mob: {phone}' if phone else '')
    _contact(4, f'Tel: {tel}' if tel else '')
    _contact(6, 'Array', font=info_bold)

    # F / G 列：标签1 + 值1（行2-5），Array 行号(段)/列号(列)
    _h(2, 6, 'Installation Angle')
    _h(2, 7, angle_display)
    _h(3, 6, 'Max Wind Load ')
    _h(3, 7, str(wind))
    _h(4, 6, 'Max Snow Load ')
    _h(4, 7, str(snow))
    _h(5, 6, 'Span(S/N)')
    _h(5, 7, str(span_sn))
    if rows:
        _h(6, 6, f'{rows} rows')
    if cols:
        _h(6, 7, f'{cols} Columns')

    # H 列：标签2（行2-5） + Array 台数(table)
    _h(2, 8, 'Panel Size')
    _h(3, 8, 'Power/PC')
    _h(4, 8, 'Watt(W)/Table')
    _h(5, 8, 'Span(E/W)')
    if tables:
        _h(6, 8, f'{tables} table')

    # I-K 合并：值2（行2-5）
    _val2_start = 9  # I
    def _hj(r, val):
        ws.merge_cells(f'{VALUE2_COL_START}{r}:{VALUE2_COL_END}{r}')
        _h(r, _val2_start, val)

    _hj(2, f'{panel_size} mm' if panel_size else '')
    _hj(3, f'{power_pc} WP' if power_pc else '')
    _hj(4, f'{watt_per_table} WP' if watt_per_table != '' else '')
    _hj(5, f'{span_ew} mm' if span_ew else '')

    # 质保（L:O 合并 L2:O6）
    ws.merge_cells(f'{WARRANTY_COL_START}2:{WARRANTY_COL_END}6')
    _set_cell(ws, 2, 12, '10 years warranty \n20 years service life', font=info_font,
              align=center, border=thin_border)

    # 抬头网格所有单元格补齐边框（含合并区与非数据格）
    # 注意：合并单元格的边框必须对每个构成单元格都强制设置才能形成完整外框，
    #       不能用 cell.border == Border() 这类条件判断（对 MergedCell 会失效）。
    for _r in range(2, 7):
        for _c in range(1, MAX_COL + 1):
            ws.cell(row=_r, column=_c).border = thin_border

    for _r in range(2, 7):
        ws.row_dimensions[_r].height = 20



    # ---- BOM 表头（行 8）----
    header_row = 8
    ws.row_dimensions[header_row].height = 67
    _hdr_black = InlineFont(rFont='Arial', sz=14, b=True, color='000000')
    _hdr_red = InlineFont(rFont='Arial', sz=14, b=True, color='FFFF0000')
    import re as _re_hdr
    # Item No. 占 A-B 两列（合并），其余表头从 C 列开始
    ws.merge_cells(f'{COL_ITEM_NO_START}{header_row}:{COL_ITEM_NO_END}{header_row}')
    for ci, label in enumerate(DETAIL_HEADERS):
        if ci == 0:
            col = 1  # Item No. -> A（合并 A-B）
        else:
            col = ci + 2  # 其余顺移 +1
        if 'EXW' in label or 'Discount' in label:
            blocks = []
            for seg in _re_hdr.split(r'(EXW|Discount)', label):
                if seg == '':
                    continue
                blocks.append(TextBlock(_hdr_red if seg in ('EXW', 'Discount') else _hdr_black, seg))
            _set_cell(ws, header_row, col, CellRichText(blocks), font=header_font,
                      align=center, border=thin_border)
        else:
            _set_cell(ws, header_row, col, label, font=header_font, align=center,
                      border=thin_border)

    # ---- BOM 数据（行 9+）----
    data_start = 9
    matched_count = 0
    unmatched_count = 0
    row_code_map = {}
    image_found_count = 0
    image_not_found_count = 0
    per_table_disc_total = Decimal('0')  # 每基折扣合计累加（供汇总引用前预览）

    for idx, item in enumerate(items):
        row = data_start + idx
        ws.row_dimensions[row].height = 60
        code = item['code']
        unit_price = item['unit_price']
        qty_per_table = item['qty_per_table']
        is_matched = item['is_matched']
        discount_cat = item.get('discount_cat', 'standard')
        if discount_cat == 'steel':
            discount_rate = ap_steel_discount_rate
        elif discount_cat == 'purchased':
            discount_rate = ap_purchased_discount_rate
        else:
            discount_rate = ap_discount_rate

        if is_matched:
            matched_count += 1
        else:
            unmatched_count += 1

        for c in range(1, MAX_COL + 1):
            _set_cell(ws, row, c, None, font=normal_font, align=center, border=thin_border)

        # Item No. 占 A-B 两列（合并）
        ws.merge_cells(f'A{row}:B{row}')
        _set_cell(ws, row, 1, idx + 1, font=normal_font, align=center, border=thin_border)
        _set_cell(ws, row, 3, item['name'], font=normal_font, align=center, border=thin_border)
        _set_cell(ws, row, 4, item['material'], font=normal_font, align=center, border=thin_border)
        _set_cell(ws, row, 5, '', font=normal_font, align=center, border=thin_border)  # Picture
        _set_cell(ws, row, 6, _strip_cjk_spec(item['spec']), font=normal_font, align=center, border=thin_border)

        # G 原价（长度折算、未折扣）
        if is_matched and unit_price > 0:
            _set_cell(ws, row, 7, round(unit_price, 6), font=normal_font, align=center,
                      border=thin_border, number_format=CURRENCY_FMT)
            # H 折扣单价 = G × rate/100
            _set_cell(ws, row, 8, f'=G{row}*{discount_rate}/100', font=normal_font, align=center,
                      border=thin_border, number_format=CURRENCY_FMT)
            per_table_disc_total += (Decimal(str(unit_price)) * Decimal(str(discount_rate)) / Decimal('100')
                                     * Decimal(str(qty_per_table)))
        else:
            _set_cell(ws, row, 7, '', font=normal_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
            _set_cell(ws, row, 8, '', font=normal_font, align=center, border=thin_border, number_format=CURRENCY_FMT)

        # I 每基数量
        qty_disp = int(qty_per_table) if float(qty_per_table).is_integer() else qty_per_table
        _set_cell(ws, row, 9, qty_disp, font=normal_font, align=center, border=thin_border)
        # J 总数量 = I × tables
        _set_cell(ws, row, 10, f'=I{row}*{tables}', font=normal_font, align=center,
                  border=thin_border, number_format='#,##0')
        # K 备品 = J × 1%
        _set_cell(ws, row, 11, f'=J{row}*{SPARE_RATE}', font=normal_font, align=center,
                  border=thin_border, number_format='#,##0.000')
        # L 每基原价合计 = G × I
        _set_cell(ws, row, 12, f'=N(G{row})*I{row}', font=normal_font, align=center,
                  border=thin_border, number_format=CURRENCY_FMT)
        # M 每基折扣合计 = H × I
        _set_cell(ws, row, 13, f'=N(H{row})*I{row}', font=normal_font, align=center,
                  border=thin_border, number_format=CURRENCY_FMT)
        # N 备品折扣合计 = H × K
        _set_cell(ws, row, 14, f'=N(H{row})*K{row}', font=normal_font, align=center,
                  border=thin_border, number_format=CURRENCY_FMT)
        # O 备注(编码)
        _set_cell(ws, row, 15, code, font=normal_font, align=center, border=thin_border)

        row_code_map[row] = code
        if not is_matched:
            for c in range(1, MAX_COL + 1):
                ws.cell(row=row, column=c).fill = yellow_fill

    data_end = data_start + len(items) - 1 if items else data_start - 1

    # ---- 隐藏 J / K / N 列（Total QTY / Spare / 备品折扣合计）----
    for _hc in ('J', 'K', 'N'):
        ws.column_dimensions[_hc].hidden = True

    # ---- 末尾合计行 ----
    total_table_row = data_end + 1
    total_all_row = data_end + 2
    alt_note_row = data_end + 3
    has_data = bool(items) and data_end >= data_start

    def _total_row(r, label, multiplier=None, amount_font=None):
        amt_font = amount_font or bold_font
        ws.merge_cells(f'A{r}:K{r}')
        _set_cell(ws, r, 1, label, font=bold_font, align=right_a, border=thin_border)
        for ci in range(2, 12):
            ws.cell(row=r, column=ci).border = thin_border
        # L 列（原价合计）、M 列（折扣合计）
        if has_data:
            k_expr = f'SUM(L{data_start}:L{data_end})'
            l_expr = f'SUM(M{data_start}:M{data_end})'
            if multiplier:
                k_expr = f'{k_expr}*{multiplier}'
                l_expr = f'{l_expr}*{multiplier}'
            _set_cell(ws, r, 12, f'={k_expr}', font=amt_font, align=center,
                      border=thin_border, number_format=CURRENCY_FMT)
            _set_cell(ws, r, 13, f'={l_expr}', font=amt_font, align=center,
                      border=thin_border, number_format=CURRENCY_FMT)
        for ci in (10, 11, 14, 15):
            ws.cell(row=r, column=ci).border = thin_border
        ws.row_dimensions[r].height = 24

    _total_row(total_table_row, 'TOTAL AMOUNT/TABLE')
    _red_bold = Font(name='Arial', size=14, bold=True, color='FFFF0000')
    _total_row(total_all_row, 'TOTAL AMOUNT OF ALL TABLE', multiplier=tables, amount_font=_red_bold)

    # 备选产品提示行（外框之外）
    _red_info = Font(name='Arial', size=14, color='FFFF0000')
    ws.merge_cells(f'A{alt_note_row}:O{alt_note_row}')
    _set_cell(ws, alt_note_row, 1,
              'The following products are alternative, please let me know if you need them.',
              font=_red_info, align=left_a)
    ws.row_dimensions[alt_note_row].height = 24

    # ---- 图片插入（D 列）----
    def _col_width_to_px(char_width):
        return int(char_width * 7.5 + 5)

    def _row_height_to_px(pt_height):
        return int(pt_height * 1.33)

    def _image_src_size(path):
        try:
            from PIL import Image as _PIL
            with _PIL.open(path) as _im:
                return _im.size
        except Exception:
            return None, None

    def _mark_missing_image(row_number):
        cell = ws.cell(row=row_number, column=IMAGE_COL_INDEX)
        cell.value = '/'
        cell.alignment = center
        cell.font = normal_font

    for row in range(data_start, data_end + 1):
        product_code = str(row_code_map.get(row, '') or '').strip()
        if not product_code:
            continue
        normalized_code = normalize_lookup_code(product_code)
        img_path = None
        if code_to_images:
            img_path = (code_to_images.get(product_code) or code_to_images.get(normalized_code) or [None])[0]
        if img_path and image_folder:
            col_w = ws.column_dimensions[get_column_letter(IMAGE_COL_INDEX)].width or 8.43
            row_h = ws.row_dimensions[row].height or 15
            avail_w = max(_col_width_to_px(col_w) - IMAGE_PADDING * 2, 20)
            avail_h = max(_row_height_to_px(row_h) - IMAGE_PADDING * 2, 20)
            src_w, src_h = _image_src_size(img_path)
            if not src_w or not src_h:
                src_w, src_h = AP_GROUND_IMAGE_WIDTH, AP_GROUND_IMAGE_HEIGHT
            scale = min(avail_w / src_w, avail_h / src_h, 1.0)
            disp_w = max(int(src_w * scale), 20)
            disp_h = max(int(src_h * scale), 20)
            temp_img_path = prepare_image_for_excel(
                img_path, target_width=disp_w, target_height=disp_h,
                temp_dir=image_temp_dir, cache=image_cache,
            )
            final_img_path = temp_img_path or img_path
            if add_image_centered_in_cell(ws, final_img_path, row, IMAGE_COL_INDEX,
                                          img_width=disp_w, img_height=disp_h):
                image_found_count += 1
            else:
                image_not_found_count += 1
                _mark_missing_image(row)
        else:
            image_not_found_count += 1
            _mark_missing_image(row)
    print(f"   [AP-GROUND] Images: found={image_found_count}, not_found={image_not_found_count}")

    # ---- 外边缘框（medium 粗框包围整张表：行1~TOTAL AMOUNT OF ALL TABLE，列1~N）----
    _apply_outer_frame(ws, 1, total_all_row, 1, MAX_COL)

    apply_print_setup(ws, 'ap_ground')

    quotation_product_codes = {str(it['code']).strip() for it in items if it.get('code')}
    return {
        'sheet_name': sheet_name,
        'site': site,
        'quotation_product_codes': quotation_product_codes,
        'valid_products': len(items),
        'matched_count': matched_count,
        'unmatched_count': unmatched_count,
        'image_found_count': image_found_count,
        'image_not_found_count': image_not_found_count,
        'data_start_row': data_start,
        'data_end_row': data_end,
        'per_table_disc_total': float(per_table_disc_total),
        'tables': tables,
        'rows': rows,
        'cols': cols,
        'pile_products': site.get('pile_products') or [],
    }


# ------------------------------------------------------------------ 汇总页

def create_ap_ground_summary_sheet(
        workbook, all_detail_results, matrix_data=None, contact_info=None,
        trade_method='EXW', dest_port='', production_lead_time='30days after receiving deposit',
        payment_term='30% T/T deposit, 70% balance before shipment',
        validity_days=7, ap_special_discount_rate=100, module_wattage=None,
        image_path=None, container_details=None,
):
    """创建汇总页 Total（每阵列一行 + 抬头 + 合计）。"""
    from datetime import datetime
    matrix_data = matrix_data or {}
    contact_info = contact_info or {}
    project_name = str(matrix_data.get('project_name') or '').strip()

    ws = workbook.create_sheet(title='Total')

    thin = Side(style='thin', color='000000')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bottom_border = Border(bottom=thin)
    # 字体/颜色对齐 en_simple 汇总表（Arial，表头无填充黑字）
    title_font = Font(name='Arial', size=26, bold=True)
    info_font = Font(name='Arial', size=11)
    info_bold = Font(name='Arial', size=11, bold=True)
    header_font = Font(name='Arial', size=12, bold=True, color='000000')
    normal_font = Font(name='Arial', size=12)
    bold_font = Font(name='Arial', size=12, bold=True)
    amount_font = Font(name='Arial', size=13, bold=True, color='000000')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_a = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_a = Alignment(horizontal='right', vertical='center', wrap_text=True)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    _summary_widths = {'A': 5, 'B': 10, 'C': 10, 'D': 15, 'E': 10, 'F': 10,
                       'G': 18, 'H': 10, 'I': 10, 'J': 10, 'K': 10, 'L': 15, 'M': 15}
    for col_letter, _w in _summary_widths.items():
        ws.column_dimensions[col_letter].width = _w

    # ---- 标题 ----
    ws.merge_cells('A3:M3')
    _set_cell(ws, 3, 1, 'Solar PV Mounting System Quotation', font=title_font, align=center)
    ws.row_dimensions[3].height = 30

    if image_path and os.path.exists(image_path):
        try:
            img = XLImage(img=image_path)
            img.width = 240
            img.height = 60
            ws.add_image(img, 'J4')
        except Exception:
            pass

    # ---- 抬头信息（左字段带下框线；右：公司信息 J-M，从第7行起）----
    port_label = str(dest_port or 'Xiamen').upper()
    red_amount_font = Font(name='Arial', size=13, bold=True, color='FFFF0000')

    def _left_field(r, label, value, wide=False, val_font=None, val_fmt=None):
        if wide:
            ws.merge_cells(f'A{r}:C{r}'); lcol, vcol = 1, 4
            ws.merge_cells(f'D{r}:G{r}')
        else:
            ws.merge_cells(f'A{r}:B{r}'); lcol, vcol = 1, 3
            ws.merge_cells(f'C{r}:G{r}')
        _set_cell(ws, r, lcol, label, font=info_font, align=left_a, border=bottom_border)
        _set_cell(ws, r, vcol, value, font=val_font or info_font, align=left_a,
                  border=bottom_border, number_format=val_fmt)

    _left_field(5, 'Project:', project_name)
    _left_field(6, 'Price Term：', f'{trade_method} {dest_port}'.strip().upper())
    _left_field(7, 'Production Lead Time：', production_lead_time, wide=True)
    _left_field(8, 'Payment Term: ', payment_term)
    _left_field(9, 'DATE：', datetime.now().strftime('%Y-%m-%d'))
    _left_field(10, 'Validity Date:', f'{validity_days} days')
    # EXW TOTAL AMOUNT / Unite Price/W（宽标签 A-C / 值 D-G，红色，值稍后回填）
    _left_field(11, 'EXW TOTAL AMOUNT:', '', wide=True, val_font=red_amount_font, val_fmt=CURRENCY_FMT)
    _left_field(12, 'Unite Price/W:', '', wide=True, val_font=red_amount_font, val_fmt='$#,##0.0000')
    exw_val_col, upw_val_col = 4, 4
    for _r in range(5, 13):
        ws.row_dimensions[_r].height = 20

    # 公司信息（右侧 J-M，第7行起）
    company_lines = [
        'Xiamen Kseng Metal Tech Co.,Ltd',
        'Add.: RM 601, Huixin Wealth Centre, No. 891,',
        'Fanghu North 2nd Rd, Huli Dist, Xiamen, Fujian, China',
    ]
    email = contact_info.get('tel') or 'abby.yan@xmkseng.com'
    phone = contact_info.get('phone') or '18050039573'
    company_lines.append(f'Email： {email}')
    company_lines.append(f'Phone/WhatsApp: {phone}')
    for i, line in enumerate(company_lines):
        cr = 7 + i
        ws.merge_cells(f'J{cr}:M{cr}')
        _set_cell(ws, cr, 10, line,
                  font=Font(name='Arial', size=13, bold=True) if i == 0 else info_font,
                  align=left_a)

    # ---- Part 1 标题（A-M 合并）----
    part1_start = 14
    ws.merge_cells(f'A{part1_start}:M{part1_start}')
    _set_cell(ws, part1_start, 1, 'Part 1: Solar Mounting System', font=info_bold, align=left_a,
              border=thin_border)
    for ci in range(2, 14):
        ws.cell(row=part1_start, column=ci).border = thin_border
    ws.row_dimensions[part1_start].height = 28

    hdr_row = part1_start + 1
    headers = ['NO.', 'Row', 'Column', 'Installation', 'Angle', 'Table', 'Power(kw)',
               'Price(USD)/Table', '', 'Amount(USD)', '', 'Discount Amount(USD)', '']
    for ci, h in enumerate(headers):
        _set_cell(ws, hdr_row, ci + 1, h, font=header_font, align=center, border=thin_border)
    ws.merge_cells(f'H{hdr_row}:I{hdr_row}')
    ws.merge_cells(f'J{hdr_row}:K{hdr_row}')
    ws.merge_cells(f'L{hdr_row}:M{hdr_row}')
    ws.row_dimensions[hdr_row].height = 30

    # ---- 数据行 ----
    data_row = hdr_row + 1
    amount_orig_cells = []   # J 列（原价 EXW）
    amount_disc_cells = []   # L 列（折扣）
    power_cells = []
    for i, dr in enumerate(all_detail_results):
        r = data_row + i
        sn = dr.get('sheet_name', '')
        sn_esc = sn.replace("'", "''")
        ds = dr.get('data_start_row')
        de = dr.get('data_end_row')
        a_rows = dr.get('rows') or ''
        a_cols = dr.get('cols') or ''
        a_tables = dr.get('tables') or 1
        site_kw = (dr.get('site') or {}).get('kw') or 0
        installation = matrix_data.get('panel_orientation') or 'Portrait'

        for ci in range(1, 14):
            ws.cell(row=r, column=ci).font = normal_font
            ws.cell(row=r, column=ci).border = thin_border

        _set_cell(ws, r, 1, i + 1, font=normal_font, align=center, border=thin_border)
        _set_cell(ws, r, 2, a_rows, font=normal_font, align=center, border=thin_border)
        _set_cell(ws, r, 3, a_cols, font=normal_font, align=center, border=thin_border)
        _set_cell(ws, r, 4, installation, font=normal_font, align=center, border=thin_border)
        angle_val = (dr.get('site') or {}).get('angle')
        _set_cell(ws, r, 5, normalize_angle(angle_val) if angle_val else '', font=normal_font, align=center, border=thin_border)
        _set_cell(ws, r, 6, a_tables, font=normal_font, align=center, border=thin_border)
        if site_kw:
            _set_cell(ws, r, 7, site_kw, font=normal_font, align=center, border=thin_border, number_format='0.00')
            power_cells.append(f'G{r}')
        else:
            _set_cell(ws, r, 7, '', font=normal_font, align=center, border=thin_border)
        # H-I Price(USD)/Table = 明细 L 列（每基折扣合计）求和
        ws.merge_cells(f'H{r}:I{r}')
        if ds and de and de >= ds:
            _set_cell(ws, r, 8, f"=SUM('{sn_esc}'!{COL_DISC_TOTAL}{ds}:{COL_DISC_TOTAL}{de})",
                      font=normal_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
        else:
            _set_cell(ws, r, 8, dr.get('per_table_disc_total', 0), font=normal_font, align=center,
                      border=thin_border, number_format=CURRENCY_FMT)
        # J-K Amount(USD) = 明细 K 列（每基原价合计）求和 × 台数
        ws.merge_cells(f'J{r}:K{r}')
        if ds and de and de >= ds:
            _set_cell(ws, r, 10, f"=SUM('{sn_esc}'!{COL_TOTAL_PRICE}{ds}:{COL_TOTAL_PRICE}{de})*F{r}",
                      font=normal_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
        else:
            _set_cell(ws, r, 10, 0, font=normal_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
        amount_orig_cells.append(f'J{r}')
        # L-M Discount Amount(USD) = Price × Table（折扣）
        ws.merge_cells(f'L{r}:M{r}')
        _set_cell(ws, r, 12, f'=H{r}*F{r}', font=normal_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
        amount_disc_cells.append(f'L{r}')
        ws.row_dimensions[r].height = 24

    data_end = data_row + len(all_detail_results) - 1 if all_detail_results else data_row - 1

    # ---- Total 行 ----
    total_row = data_end + 1
    ws.merge_cells(f'A{total_row}:F{total_row}')
    _set_cell(ws, total_row, 1, 'Total', font=bold_font, align=center, border=thin_border)
    if power_cells:
        _set_cell(ws, total_row, 7, f'={"+".join(power_cells)}', font=bold_font, align=center,
                  border=thin_border, number_format='0.00')
    else:
        _set_cell(ws, total_row, 7, '', font=bold_font, align=center, border=thin_border)
    ws.merge_cells(f'H{total_row}:I{total_row}')
    _set_cell(ws, total_row, 8, '', font=bold_font, align=center, border=thin_border)
    ws.merge_cells(f'J{total_row}:K{total_row}')
    if amount_orig_cells:
        _set_cell(ws, total_row, 10, f'={"+".join(amount_orig_cells)}', font=bold_font, align=center,
                  border=thin_border, number_format=CURRENCY_FMT)
    else:
        _set_cell(ws, total_row, 10, 0, font=bold_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
    ws.merge_cells(f'L{total_row}:M{total_row}')
    if amount_disc_cells:
        _set_cell(ws, total_row, 12, f'={"+".join(amount_disc_cells)}', font=bold_font, align=center,
                  border=thin_border, number_format=CURRENCY_FMT)
    else:
        _set_cell(ws, total_row, 12, 0, font=bold_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
    ws.row_dimensions[total_row].height = 24

    orig_total_ref = f'J{total_row}'
    disc_total_ref = f'L{total_row}'
    power_inner = "+".join(power_cells) if power_cells else None  # 例: G16+G17

    # ---- Total 下方 3 行：per/w、TOTAL AMOUNT(EXW)、Extra Profit Margin ----
    perw_row = total_row + 1
    ws.merge_cells(f'A{perw_row}:I{perw_row}')
    _set_cell(ws, perw_row, 1, 'per/w', font=bold_font, align=right_a, border=thin_border)
    for ci in range(2, 10):
        ws.cell(row=perw_row, column=ci).border = thin_border
    ws.merge_cells(f'J{perw_row}:K{perw_row}')
    if power_inner:
        _set_cell(ws, perw_row, 10, f'=ROUND({orig_total_ref}/(({power_inner})*1000),4)',
                  font=bold_font, align=center, border=thin_border, number_format='$#,##0.0000')
    else:
        _set_cell(ws, perw_row, 10, 0, font=bold_font, align=center, border=thin_border, number_format='$#,##0.0000')
    ws.merge_cells(f'L{perw_row}:M{perw_row}')
    if power_inner:
        _set_cell(ws, perw_row, 12, f'=ROUND({disc_total_ref}/(({power_inner})*1000),4)',
                  font=bold_font, align=center, border=thin_border, number_format='$#,##0.0000')
    else:
        _set_cell(ws, perw_row, 12, 0, font=bold_font, align=center, border=thin_border, number_format='$#,##0.0000')
    ws.row_dimensions[perw_row].height = 24

    total_exw_row = total_row + 2
    ws.merge_cells(f'A{total_exw_row}:I{total_exw_row}')
    _set_cell(ws, total_exw_row, 1, 'TOTAL AMOUNT(EXW)', font=bold_font, align=right_a, border=thin_border)
    for ci in range(2, 10):
        ws.cell(row=total_exw_row, column=ci).border = thin_border
    ws.merge_cells(f'J{total_exw_row}:K{total_exw_row}')
    _set_cell(ws, total_exw_row, 10, f'={orig_total_ref}', font=bold_font, align=center,
              border=thin_border, number_format=CURRENCY_FMT)
    ws.merge_cells(f'L{total_exw_row}:M{total_exw_row}')
    _set_cell(ws, total_exw_row, 12, f'={disc_total_ref}', font=bold_font, align=center,
              border=thin_border, number_format=CURRENCY_FMT)
    ws.row_dimensions[total_exw_row].height = 24

    extra_row = total_row + 3
    ws.merge_cells(f'A{extra_row}:K{extra_row}')
    _set_cell(ws, extra_row, 1, 'Extra Profit Margin', font=bold_font, align=right_a, border=thin_border)
    for ci in range(2, 12):
        ws.cell(row=extra_row, column=ci).border = thin_border
    ws.merge_cells(f'L{extra_row}:M{extra_row}')
    _set_cell(ws, extra_row, 12, f'={orig_total_ref}-{disc_total_ref}', font=bold_font, align=center,
              border=thin_border, number_format=CURRENCY_FMT)
    ws.row_dimensions[extra_row].height = 24

    # ---- 回填抬头 EXW TOTAL AMOUNT / Unite Price/W（红色，D 列，取折扣总额）----
    _set_cell(ws, 11, exw_val_col, f'=ROUND({disc_total_ref},2)', font=red_amount_font,
              align=left_a, number_format=CURRENCY_FMT)
    if power_inner:
        upw = f'=ROUND({disc_total_ref}/(({power_inner})*1000),4)'
    else:
        upw = 0
    _set_cell(ws, 12, upw_val_col, upw, font=red_amount_font, align=left_a, number_format='$#,##0.0000')

    # ---- Part 2: Freight & Insurance（仅 FOB/CIF 且有柜型时）----
    frame_bottom = extra_row
    containers = [c for c in (container_details or []) if c.get('qty')]
    if containers and trade_method in ('FOB', 'CIF'):
        port_name = str(dest_port or 'Xiamen').upper()
        p2_title = extra_row + 2
        p2_hdr = p2_title + 1
        p2_data_start = p2_hdr + 1
        ws.merge_cells(f'A{p2_title}:M{p2_title}')
        _set_cell(ws, p2_title, 1, 'Part 2: Freight & Insurance', font=info_bold, align=left_a, border=thin_border)
        for ci in range(2, 14):
            ws.cell(row=p2_title, column=ci).border = thin_border
        ws.row_dimensions[p2_title].height = 26

        # 列头：A:F 空 | G:H Container | I Quantity | J:K Unit price | L:M Amount(USD)
        ws.merge_cells(f'A{p2_hdr}:F{p2_hdr}')
        ws.merge_cells(f'G{p2_hdr}:H{p2_hdr}')
        ws.merge_cells(f'J{p2_hdr}:K{p2_hdr}')
        ws.merge_cells(f'L{p2_hdr}:M{p2_hdr}')
        _hdrs = {1: '', 7: 'Container', 9: 'Quantity', 10: 'Unit price', 12: 'Amount(USD)'}
        for cc, txt in _hdrs.items():
            _set_cell(ws, p2_hdr, cc, txt, font=header_font, align=center, border=thin_border)
        for ci in range(1, 14):
            ws.cell(row=p2_hdr, column=ci).border = thin_border
        ws.row_dimensions[p2_hdr].height = 24

        freight_amount_cells = []
        for ci_idx, cd in enumerate(containers):
            rr = p2_data_start + ci_idx
            ct = str(cd.get('type', '')).upper()
            cq = cd.get('qty', 0)
            cpu = cd.get('freight_per_unit', 0) or 0
            desc = ''
            if ci_idx == 0:
                desc = f'{trade_method} {port_name}\n+shipping cost from Xiamen to {port_name}(few accessories)'
            ws.merge_cells(f'A{rr}:F{rr}')
            _set_cell(ws, rr, 1, desc, font=normal_font, align=left_a, border=thin_border)
            ws.merge_cells(f'G{rr}:H{rr}')
            _set_cell(ws, rr, 7, ct, font=normal_font, align=center, border=thin_border)
            _set_cell(ws, rr, 9, cq, font=normal_font, align=center, border=thin_border)
            ws.merge_cells(f'J{rr}:K{rr}')
            _set_cell(ws, rr, 10, float(cpu), font=normal_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
            ws.merge_cells(f'L{rr}:M{rr}')
            _set_cell(ws, rr, 12, f'=J{rr}*I{rr}', font=normal_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
            for _ci in range(1, 14):
                ws.cell(row=rr, column=_ci).border = thin_border
            freight_amount_cells.append(f'L{rr}')
            ws.row_dimensions[rr].height = 30

        p2_data_end = p2_data_start + len(containers) - 1
        p2_total = p2_data_end + 1
        ws.merge_cells(f'A{p2_total}:K{p2_total}')
        _set_cell(ws, p2_total, 1, f'TOTAL AMOUNT({trade_method})', font=bold_font, align=right_a, border=thin_border)
        for ci in range(2, 12):
            ws.cell(row=p2_total, column=ci).border = thin_border
        ws.merge_cells(f'L{p2_total}:M{p2_total}')
        if freight_amount_cells:
            _set_cell(ws, p2_total, 12, f'={disc_total_ref}+{"+".join(freight_amount_cells)}',
                      font=bold_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
        else:
            _set_cell(ws, p2_total, 12, f'={disc_total_ref}', font=bold_font, align=center,
                      border=thin_border, number_format=CURRENCY_FMT)
        ws.row_dimensions[p2_total].height = 24
        frame_bottom = p2_total

        p2_note = p2_total + 2
        ws.merge_cells(f'A{p2_note}:M{p2_note}')
        _red_note_font = Font(name='Arial', size=11, color='FFFF0000')
        _set_cell(ws, p2_note, 1, 'Please kindly check the detailed prices of components in the following sheet',
                  font=_red_note_font, align=left_a)

    # ---- 汇总表外边缘框（Part1 标题 ~ 末行，A~M）----
    _apply_outer_frame(ws, part1_start, frame_bottom, 1, 13)

    apply_print_setup(ws, 'ap_ground')

    # 汇总页置顶
    sheet_names = workbook.sheetnames
    if 'Total' in sheet_names:
        idx = sheet_names.index('Total')
        workbook.move_sheet('Total', offset=-idx)
    print(f"   [AP-GROUND] Summary sheet(Total) created")
    return 'Total'
