from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
from decimal import Decimal

from backend.core.quotation_engine import (
    resolve_price_info,
    has_valid_price_info,
    normalize_lookup_code,
    round_to_2_decimal,
    extract_length_from_spec,
    _is_valid_product_code,
)
from backend.core.material_translate import translate_material

MEIRYO = Font(name='Meiryo UI', size=11)
MEIRYO_BOLD = Font(name='Meiryo UI', size=11, bold=True)
MEIRYO_TITLE = Font(name='Meiryo UI', size=22, bold=True)
MEIRYO_HEADER = Font(name='Meiryo UI', size=11, bold=True)

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

THICK_SIDE = Side(style='medium', color='000000')

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT_A = Alignment(horizontal='right', vertical='center', wrap_text=True)
LEFT_A = Alignment(horizontal='left', vertical='center', wrap_text=True)

CURRENCY_FMT = '"US$" #,##0.00'

SUMMARY_COL_WIDTHS = {
    'A': 8, 'B': 15, 'C': 12, 'D': 15, 'E': 12,
    'F': 10, 'G': 12, 'H': 15, 'I': 18, 'J': 15, 'K': 15,
}

DETAIL_COL_WIDTHS = {
    'A': 5, 'B': 10, 'C': 10, 'D': 15, 'E': 10, 'F': 10, 'G': 8, 'H': 15,
}

DEFAULT_DISCOUNT_RATE = Decimal('0.71')


def _apply_outer_border(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            left_s = THICK_SIDE if c == min_col else Side(style='thin')
            right_s = THICK_SIDE if c == max_col else Side(style='thin')
            top_s = THICK_SIDE if r == min_row else Side(style='thin')
            bottom_s = THICK_SIDE if r == max_row else Side(style='thin')
            cell.border = Border(left=left_s, right=right_s, top=top_s, bottom=bottom_s)


def create_detail_sheet(workbook, array_info, bom_products, price_mapping,
                        sheet_prefix=None,
                        image_path=None, image_folder=None, code_to_images=None,
                        image_temp_dir=None, image_cache=None, matrix_data=None,
                        discount_rate=None, group=None,
                        unmatched_products_out=None):
    from backend.core.quotation_engine import (
        prepare_image_for_excel, add_image_centered_in_cell,
    )
    JA_IMG_W = 100
    JA_IMG_H = 65
    JA_IMG_COL = 4
    JA_IMG_PADDING = 2

    def _col_width_to_px(char_width):
        return int(char_width * 7.5 + 5)

    def _row_height_to_px(pt_height):
        return int(pt_height * 1.33)

    def _fit_image_to_cell(ws, cell_row, cell_col, max_w, max_h, padding=JA_IMG_PADDING):
        col_w = ws.column_dimensions[get_column_letter(cell_col)].width or 8.43
        row_h = ws.row_dimensions[cell_row].height or 15
        avail_w = _col_width_to_px(col_w) - padding * 2
        avail_h = _row_height_to_px(row_h) - padding * 2
        avail_w = max(avail_w, 20)
        avail_h = max(avail_h, 20)
        scale = min(avail_w / max_w, avail_h / max_h)
        return int(max_w * scale), int(max_h * scale)

    if sheet_prefix:
        sheet_name = sheet_prefix
    else:
        sheet_name = array_info.get('note', '') or f"セット{array_info.get('no', '1')}"
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    original = sheet_name
    counter = 1
    while sheet_name in workbook.sheetnames:
        sheet_name = f"{original}_{counter}"
        counter += 1

    ws = workbook.create_sheet(title=sheet_name)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    for col, width in DETAIL_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    black_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000'),
    )

    ws.merge_cells('A1:H1')
    project_name = (matrix_data or {}).get('project_name', '')
    ws['A1'] = project_name
    ws['A1'].font = Font(name='Meiryo UI', size=20, bold=True)
    ws['A1'].alignment = center
    ws['A1'].border = black_border
    ws.row_dimensions[1].height = 40
    for c in range(1, 9):
        ws.cell(row=1, column=c).border = black_border

    rows_val = array_info.get('rows', '')
    cols_val = array_info.get('cols', '')
    array_label = f'{rows_val}段{cols_val}列' if rows_val and cols_val else ''

    ws.merge_cells('A2:B2')
    ws['A2'] = array_label
    ws['A2'].font = MEIRYO
    ws['A2'].alignment = center
    ws['A2'].border = black_border
    ws.merge_cells('C2:D2')
    for c in range(3, 5):
        ws.cell(row=2, column=c).border = black_border

    ws.merge_cells('E2:F2')
    ws['E2'] = 'セット数'
    ws['E2'].font = MEIRYO
    ws['E2'].alignment = center
    ws['E2'].border = black_border
    ws.cell(row=2, column=6).border = black_border
    ws.merge_cells('G2:H2')
    ws['G2'] = array_info.get('table_qty', 1)
    ws['G2'].font = MEIRYO
    ws['G2'].alignment = center
    ws['G2'].border = black_border
    ws.cell(row=2, column=8).border = black_border
    ws.row_dimensions[2].height = 15

    ws.merge_cells('A3:H3')
    ws.row_dimensions[3].height = 5

    BLACK_BORDER = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000'),
    )
    headers = ['序号', '品名', '材質', '写真', '規格', '単価\n（USD）', '数量\n（PCS)', '総金額\n（USD）']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.border = BLACK_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = MEIRYO_HEADER
        cell.fill = header_fill
    ws.row_dimensions[4].height = 15

    name_field = 'name_ja'
    total_price_sum = Decimal('0')
    matched_count = 0
    unmatched_count = 0
    _detail_table_qty = int(array_info.get('table_qty', 1)) if array_info else 1
    data_start = 5

    YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    filtered_products = [
        p for p in bom_products
        if p.get('quantity') not in (None, '', 0, 0.0)
    ]

    for row in range(data_start, data_start + len(filtered_products)):
        ws.row_dimensions[row].height = 60

    for idx, product in enumerate(filtered_products):
        row = data_start + idx
        product_code = product.get('code', '')
        price_info = resolve_price_info(price_mapping, product_code)

        display_name = (
            price_info.get(name_field)
            or price_info.get('name_ko')
            or price_info.get('name')
            or product.get('name', '')
        ) if price_info else product.get('name', '')

        ws.cell(row=row, column=1, value=idx + 1).border = thin_border
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=2, value=display_name).border = thin_border
        ws.cell(row=row, column=3, value=translate_material(product.get('material', ''), 'ja')).border = thin_border
        ws.cell(row=row, column=4, value='').border = thin_border
        ws.cell(row=row, column=5, value=product.get('spec', '')).border = thin_border

        quantity = product.get('quantity', 0)
        unit_price = 0
        display_unit_price = 0
        price_unit = ''
        is_matched = False

        if price_info and has_valid_price_info(price_info):
            unit_price = float(price_info['price'])
            price_unit = price_info.get('unit', '')
            matched_count += 1
            is_matched = True
        else:
            unmatched_count += 1
            if unmatched_products_out is not None and _is_valid_product_code(product_code):
                unmatched_products_out.append({
                    'code': product_code,
                    'name': product.get('name', ''),
                    'spec': product.get('spec', ''),
                    'material': product.get('material', ''),
                    'quantity': quantity * _detail_table_qty,
                })

        is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
        length_mm = Decimal('0')
        if is_meter:
            length_mm = Decimal(str(extract_length_from_spec(product.get('spec')) or 0))

        if is_meter and length_mm > 0:
            display_unit_price = float(Decimal(str(unit_price)) * length_mm / Decimal('1000'))
        else:
            display_unit_price = unit_price

        if display_unit_price > 0:
            ws.cell(row=row, column=6, value=display_unit_price).border = thin_border
            ws.cell(row=row, column=6).number_format = CURRENCY_FMT
        else:
            ws.cell(row=row, column=6, value='').border = thin_border

        if quantity > 0:
            ws.cell(row=row, column=7,
                    value=int(quantity) if quantity % 1 == 0 else quantity).border = thin_border
        else:
            ws.cell(row=row, column=7, value='').border = thin_border

        total_price = Decimal('0')
        if display_unit_price > 0 and quantity > 0:
            total_price = Decimal(str(display_unit_price)) * Decimal(str(quantity))

        total_price_rounded = round_to_2_decimal(total_price)
        if total_price_rounded > 0:
            ws.cell(row=row, column=8, value=float(total_price_rounded)).border = thin_border
            ws.cell(row=row, column=8).number_format = CURRENCY_FMT
            total_price_sum += total_price_rounded
        else:
            ws.cell(row=row, column=8, value='').border = thin_border

        ws.cell(row=row, column=2).alignment = center
        ws.cell(row=row, column=3).alignment = center
        ws.cell(row=row, column=5).alignment = center
        ws.cell(row=row, column=6).alignment = center
        ws.cell(row=row, column=7).alignment = center
        ws.cell(row=row, column=8).alignment = center

        if not is_matched:
            for c in range(1, 9):
                ws.cell(row=row, column=c).fill = YELLOW_FILL

    data_end = data_start + len(filtered_products) - 1

    image_found_count = 0
    image_not_found_count = 0
    for row in range(data_start, data_end + 1):
        idx = row - data_start
        if idx >= len(filtered_products):
            continue
        product_code = filtered_products[idx].get('code', '')
        if not product_code:
            continue
        product_code = str(product_code).strip()
        normalized_code = normalize_lookup_code(product_code)

        img_path = None
        if code_to_images and product_code in code_to_images and code_to_images[product_code]:
            img_path = code_to_images[product_code][0]
        elif code_to_images and normalized_code in code_to_images and code_to_images[normalized_code]:
            img_path = code_to_images[normalized_code][0]

        if img_path:
            fit_w, fit_h = _fit_image_to_cell(ws, row, JA_IMG_COL, JA_IMG_W, JA_IMG_H)
            temp_img_path = prepare_image_for_excel(
                img_path,
                target_width=fit_w,
                target_height=fit_h,
                temp_dir=image_temp_dir,
                cache=image_cache,
            )
            final_img_path = temp_img_path if temp_img_path else img_path
            success = add_image_centered_in_cell(
                ws,
                final_img_path,
                row,
                JA_IMG_COL,
                img_width=fit_w,
                img_height=fit_h,
            )
            if success:
                image_found_count += 1
                continue
        image_not_found_count += 1
        ws.cell(row=row, column=4, value='/').alignment = center

    if image_found_count or image_not_found_count:
        print(f"   🖼️ {sheet_name}: 挿入 {image_found_count} 枚, 未検出 {image_not_found_count} 枚")

    sub_row = data_end + 1
    ws.merge_cells(f'A{sub_row}:G{sub_row}')
    ws[f'A{sub_row}'] = 'SUB-TOTAL-（FOB）1基スクリュー杭基礎架台合計'
    ws[f'A{sub_row}'].alignment = right_align
    ws[f'A{sub_row}'].font = MEIRYO
    ws.cell(row=sub_row, column=8, value=float(total_price_sum))
    ws.cell(row=sub_row, column=8).number_format = CURRENCY_FMT
    ws.cell(row=sub_row, column=8).font = MEIRYO
    for c in range(1, 9):
        ws.cell(row=sub_row, column=c).border = thin_border

    per_base_price = total_price_sum

    return {
        'sheet_name': sheet_name,
        'array_info': array_info,
        'total_price_per_base': float(per_base_price),
        'total_price': float(total_price_sum),
        'matched_count': matched_count,
        'unmatched_count': unmatched_count,
    }


def create_summary_sheet(workbook, detail_results, matrix_data=None,
                         image_path=None, fence_data=None, shipping_data=None):
    ws = workbook.create_sheet(title='合計')

    for col, width in SUMMARY_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    matrix_data = matrix_data or {}
    project_name = str(matrix_data.get('project_name') or '').strip()
    output_kw = matrix_data.get('output_kw') or 0
    module_wattage = matrix_data.get('module_wattage') or 0
    arrays = matrix_data.get('arrays') or []
    wind_speed = matrix_data.get('max_wind_speed') or ''
    snow_load = matrix_data.get('max_snow_load') or ''
    angle = matrix_data.get('angle') or ''
    panel_size = matrix_data.get('panel_spec') or matrix_data.get('module_size') or ''

    fence_data = fence_data or {}
    shipping_data = shipping_data or {}

    ws['I1'] = '見積日：'
    ws['I1'].font = MEIRYO
    ws.merge_cells('J1:K1')
    ws['J1'] = '=TODAY()'
    ws['J1'].font = MEIRYO
    ws['J1'].number_format = 'YYYY/MM/DD'

    ws['D2'] = '架台御見積書'
    ws.merge_cells('D2:J2')
    ws['D2'].font = MEIRYO_TITLE
    ws['D2'].alignment = CENTER

    ws['A3'] = '案件名：'
    ws['A3'].font = MEIRYO
    ws.merge_cells('B3:H3')
    ws['B3'] = project_name
    ws['B3'].font = MEIRYO
    ws['K3'] = 'KW'
    ws['K3'].font = MEIRYO
    if output_kw:
        ws['I3'] = output_kw
        ws['I3'].font = MEIRYO

    ws['A4'] = '見積条件：'
    ws['A4'].font = MEIRYO
    ws.merge_cells('B4:K4')

    ws['A5'] = '納入期限：'
    ws['A5'].font = MEIRYO
    ws['C5'] = '発注後10-14日間後工場から出荷'
    ws['C5'].font = MEIRYO
    ws.merge_cells('C5:H5')

    ws['A6'] = '取引条件：'
    ws['A6'].font = MEIRYO
    ws['C6'] = '取引基本契約書に基づく'
    ws['C6'].font = MEIRYO
    ws.merge_cells('C6:H6')

    ws['A7'] = '有効期限：'
    ws['A7'].font = MEIRYO
    ws['C7'] = '御見積後2日間'
    ws['C7'].font = MEIRYO
    ws.merge_cells('C7:H7')

    ws['A8'] = '※'
    ws['A8'].font = MEIRYO
    ws['B8'] = f'風速 {wind_speed}' if wind_speed else '風速 図面通り'
    ws['B8'].font = MEIRYO
    ws['D8'] = 'パネル高さ'
    ws['D8'].font = MEIRYO
    ws['H8'] = '図面通り'
    ws['H8'].font = MEIRYO
    ws['F8'] = '※'
    ws['F8'].font = MEIRYO
    ws['G8'] = f'パネルサイズ {panel_size}' if panel_size else 'パネルサイズ'
    ws['G8'].font = MEIRYO

    ws['A9'] = '※'
    ws['A9'].font = MEIRYO
    ws['B9'] = f'積雪 {snow_load}' if snow_load else '積雪 図面通り'
    ws['B9'].font = MEIRYO
    ws['D9'] = '傾斜角度'
    ws['D9'].font = MEIRYO
    ws['H9'] = angle if angle else '図面通り'
    ws['H9'].font = MEIRYO
    ws['F9'] = '※'
    ws['F9'].font = MEIRYO
    ws['G9'] = f'発電量/PC {module_wattage}W' if module_wattage else '発電量/PC'
    ws['G9'].font = MEIRYO

    ws['A10'] = '一、架台本体金額（Ex Works）'
    ws['A10'].font = MEIRYO_BOLD
    ws.merge_cells('A10:K10')

    ws['A11'] = '序号'
    ws['A11'].font = MEIRYO_HEADER
    ws['A11'].alignment = CENTER
    ws['B11'] = 'パネル数'
    ws['B11'].font = MEIRYO_HEADER
    ws.merge_cells('B11:C11')
    ws['B11'].alignment = CENTER
    ws['D11'] = 'セット数'
    ws['D11'].font = MEIRYO_HEADER
    ws['D11'].alignment = CENTER
    ws['E11'] = '備考'
    ws['E11'].font = MEIRYO_HEADER
    ws.merge_cells('E11:F11')
    ws['E11'].alignment = CENTER
    ws['G11'] = '発電量（KW）'
    ws['G11'].font = MEIRYO_HEADER
    ws['G11'].alignment = CENTER
    ws['H11'] = '単価(USD)/基'
    ws['H11'].font = MEIRYO_HEADER
    ws['H11'].alignment = CENTER
    ws['I11'] = '特別値引き後金額'
    ws['I11'].font = MEIRYO_HEADER
    ws['I11'].alignment = CENTER
    ws['J11'] = '総金額(USD)'
    ws['J11'].font = MEIRYO_HEADER
    ws['J11'].alignment = CENTER
    ws['K11'] = 'W単価(USD)'
    ws['K11'].font = MEIRYO_HEADER
    ws['K11'].alignment = CENTER

    discount_rate = DEFAULT_DISCOUNT_RATE
    grand_total = Decimal('0')
    grand_output_kw = Decimal('0')

    for i, detail in enumerate(detail_results):
        row = 12 + i
        arr = detail.get('array_info', {})
        base_count = arr.get('table_qty', 1)
        panel_count = arr.get('rows', 0) * arr.get('cols', 0)
        note = arr.get('note', '')

        gen_kw = Decimal(str(panel_count)) * Decimal(str(base_count)) * Decimal(str(module_wattage)) / Decimal('1000')
        per_base = Decimal(str(detail.get('total_price_per_base', 0)))
        special_price = round_to_2_decimal(per_base * discount_rate)
        total_amount = round_to_2_decimal(special_price * Decimal(str(base_count)))
        w_unit_price = round_to_2_decimal(special_price / (Decimal(str(panel_count)) * Decimal(str(module_wattage)))) if panel_count > 0 and module_wattage > 0 else Decimal('0')

        ws.cell(row=row, column=1, value=i + 1)
        ws.merge_cells(f'B{row}:C{row}')
        ws.cell(row=row, column=2, value=panel_count)
        ws.cell(row=row, column=4, value=base_count)
        ws.merge_cells(f'E{row}:F{row}')
        ws.cell(row=row, column=5, value=note)
        ws.cell(row=row, column=7, value=float(gen_kw))
        ws.cell(row=row, column=8, value=float(per_base))
        ws.cell(row=row, column=8).number_format = CURRENCY_FMT
        ws.cell(row=row, column=9, value=float(special_price))
        ws.cell(row=row, column=9).number_format = CURRENCY_FMT
        ws.cell(row=row, column=10, value=float(total_amount))
        ws.cell(row=row, column=10).number_format = CURRENCY_FMT
        ws.cell(row=row, column=11, value=float(w_unit_price))
        ws.cell(row=row, column=11).number_format = CURRENCY_FMT

        for c in range(1, 12):
            ws.cell(row=row, column=c).font = MEIRYO
            ws.cell(row=row, column=c).alignment = CENTER
            ws.cell(row=row, column=c).border = THIN_BORDER

        grand_total += total_amount
        grand_output_kw += gen_kw

    data_end = 12 + len(detail_results) - 1

    sum_start = max(data_end + 1, 17)
    ws[f'A{sum_start}'] = '架台総金額(USD)'
    ws[f'A{sum_start}'].font = MEIRYO
    ws[f'A{sum_start}'].alignment = RIGHT_A
    ws.merge_cells(f'A{sum_start}:J{sum_start}')
    ws[f'K{sum_start}'] = float(grand_total)
    ws[f'K{sum_start}'].font = MEIRYO

    row_18 = sum_start + 1
    ws[f'A{row_18}'] = '通常2000ｍｍ'
    ws[f'A{row_18}'].font = MEIRYO
    ws.merge_cells(f'A{row_18}:C{row_18}')
    ws[f'D{row_18}'] = '単価'
    ws[f'D{row_18}'].font = MEIRYO
    ws.merge_cells(f'E{row_18}:G{row_18}')
    ws[f'H{row_18}'] = '数量'
    ws[f'H{row_18}'].font = MEIRYO
    ws[f'I{row_18}'] = '金額'
    ws[f'I{row_18}'].font = MEIRYO
    ws[f'J{row_18}'] = 0
    ws[f'J{row_18}'].font = MEIRYO
    ws[f'K{row_18}'] = 0
    ws[f'K{row_18}'].font = MEIRYO

    row_19 = row_18 + 1
    ws[f'A{row_19}'] = '①架台＋杭本体　総金額(USD)'
    ws[f'A{row_19}'].font = MEIRYO
    ws[f'A{row_19}'].alignment = RIGHT_A
    ws.merge_cells(f'A{row_19}:I{row_19}')
    ws.merge_cells(f'J{row_19}:K{row_19}')
    ws[f'J{row_19}'] = float(grand_total)
    ws[f'J{row_19}'].font = MEIRYO

    row_20 = row_19 + 1
    ws[f'A{row_20}'] = '発電量(KW)'
    ws[f'A{row_20}'].font = MEIRYO
    ws[f'A{row_20}'].alignment = RIGHT_A
    ws.merge_cells(f'A{row_20}:I{row_20}')
    ws.merge_cells(f'J{row_20}:K{row_20}')
    ws[f'J{row_20}'] = float(grand_output_kw)
    ws[f'J{row_20}'].font = MEIRYO

    row_21 = row_20 + 1
    ws[f'A{row_21}'] = 'ワットあたりの価格'
    ws[f'A{row_21}'].font = MEIRYO
    ws[f'A{row_21}'].alignment = RIGHT_A
    ws.merge_cells(f'A{row_21}:I{row_21}')
    ws.merge_cells(f'J{row_21}:K{row_21}')
    w_price = round(float(grand_total / (grand_output_kw * 1000)), 4) if grand_output_kw > 0 else 0
    ws[f'J{row_21}'] = w_price
    ws[f'J{row_21}'].font = MEIRYO

    row_22 = row_21 + 1
    ws[f'A{row_22}'] = '二、フェンス金額'
    ws[f'A{row_22}'].font = MEIRYO_BOLD
    fence_length = fence_data.get('length', 140)
    fence_color = fence_data.get('color', '茶')
    fence_corner = fence_data.get('corner', 'コーナー10箇所')
    ws[f'D{row_22}'] = fence_length
    ws[f'D{row_22}'].font = MEIRYO
    ws[f'E{row_22}'] = 'M'
    ws[f'E{row_22}'].font = MEIRYO
    ws[f'G{row_22}'] = fence_color
    ws[f'G{row_22}'].font = MEIRYO
    ws[f'H{row_22}'] = fence_corner
    ws[f'H{row_22}'].font = MEIRYO
    ws[f'K{row_22}'] = '張様仕入れ'
    ws[f'K{row_22}'].font = MEIRYO

    row_23 = row_22 + 1
    ws[f'A{row_23}'] = 'NO.'
    ws[f'A{row_23}'].font = MEIRYO_HEADER
    ws[f'A{row_23}'].alignment = CENTER
    ws[f'B{row_23}'] = '規格・部品名称'
    ws[f'B{row_23}'].font = MEIRYO_HEADER
    ws[f'B{row_23}'].alignment = CENTER
    ws.merge_cells(f'B{row_23}:F{row_23}')
    ws[f'G{row_23}'] = '単位'
    ws[f'G{row_23}'].font = MEIRYO_HEADER
    ws[f'G{row_23}'].alignment = CENTER
    ws[f'H{row_23}'] = '1Mの単価\nUnit Price'
    ws[f'H{row_23}'].font = MEIRYO_HEADER
    ws[f'H{row_23}'].alignment = CENTER
    ws[f'I{row_23}'] = '数量\nQTY'
    ws[f'I{row_23}'].font = MEIRYO_HEADER
    ws[f'I{row_23}'].alignment = CENTER
    ws[f'J{row_23}'] = '総金額(US＄)\nAmount Price'
    ws[f'J{row_23}'].font = MEIRYO_HEADER
    ws[f'J{row_23}'].alignment = CENTER
    ws.merge_cells(f'J{row_23}:K{row_23}')

    fence_items = fence_data.get('items', [])
    fence_total = Decimal('0')
    for fi in range(5):
        r = row_23 + 1 + fi
        ws.merge_cells(f'B{r}:F{r}')
        ws.merge_cells(f'J{r}:K{r}')
        if fi < len(fence_items):
            item = fence_items[fi]
            ws.cell(row=r, column=1, value=fi + 1).font = MEIRYO
            ws.cell(row=r, column=2, value=item.get('name', '')).font = MEIRYO
            ws.cell(row=r, column=7, value=item.get('unit', '')).font = MEIRYO
            ws.cell(row=r, column=8, value=item.get('unit_price', 0)).font = MEIRYO
            ws.cell(row=r, column=9, value=item.get('qty', 0)).font = MEIRYO
            amount = Decimal(str(item.get('unit_price', 0))) * Decimal(str(item.get('qty', 0)))
            fence_total += amount
            ws.cell(row=r, column=10, value=float(amount)).font = MEIRYO
        for c in range(1, 12):
            ws.cell(row=r, column=c).border = THIN_BORDER

    fence_end = row_23 + 5

    r_add = fence_end + 1
    ws[f'A{r_add}'] = '追加諸係り：'
    ws[f'A{r_add}'].font = MEIRYO
    ws[f'A{r_add}'].alignment = RIGHT_A
    ws.merge_cells(f'A{r_add}:I{r_add}')
    ws[f'J{r_add}'] = 0
    ws[f'J{r_add}'].font = MEIRYO
    ws[f'K{r_add}'] = 0
    ws[f'K{r_add}'].font = MEIRYO

    r_f2 = r_add + 1
    ws[f'A{r_f2}'] = '②総金額'
    ws[f'A{r_f2}'].font = MEIRYO
    ws[f'A{r_f2}'].alignment = RIGHT_A
    ws.merge_cells(f'A{r_f2}:I{r_f2}')
    ws[f'J{r_f2}'] = float(fence_total)
    ws[f'J{r_f2}'].font = MEIRYO
    ws[f'K{r_f2}'] = 0
    ws[f'K{r_f2}'].font = MEIRYO

    r_31 = r_f2 + 1
    ws[f'A{r_31}'] = '二、運賃'
    ws[f'A{r_31}'].font = MEIRYO_BOLD
    ws.merge_cells(f'A{r_31}:K{r_31}')

    r_32 = r_31 + 1
    ws[f'A{r_32}'] = '1ドル＝'
    ws[f'A{r_32}'].font = MEIRYO
    ws[f'A{r_32}'].alignment = RIGHT_A
    ws.merge_cells(f'A{r_32}:E{r_32}')
    exchange_rate = shipping_data.get('exchange_rate', 0)
    ws[f'F{r_32}'] = exchange_rate
    ws[f'F{r_32}'].font = MEIRYO
    ws[f'G{r_32}'] = '円にて換算、レートは２％以上増やすと、当日のレートに変更'
    ws[f'G{r_32}'].font = MEIRYO
    ws.merge_cells(f'G{r_32}:I{r_32}')
    ws[f'J{r_32}'] = '（USD）'
    ws[f'J{r_32}'].font = MEIRYO
    ws[f'K{r_32}'] = '(JPY)'
    ws[f'K{r_32}'].font = MEIRYO

    tariiff_1 = grand_total * Decimal('0.016')
    tax_1 = grand_total * Decimal('0.10')
    fence_tax = fence_total * Decimal('0.10')
    shipping_cost = Decimal(str(shipping_data.get('shipping_usd', 0)))

    r_33 = r_32 + 1
    ws[f'A{r_33}'] = '架台+杭　関税「H.S.code 7610.90 000 構造物及びその部分品（その他のもの）」1.6％'
    ws[f'A{r_33}'].font = MEIRYO
    ws[f'A{r_33}'].alignment = RIGHT_A
    ws.merge_cells(f'A{r_33}:I{r_33}')
    ws.merge_cells(f'J{r_33}:K{r_33}')
    ws[f'J{r_33}'] = float(tariiff_1)
    ws[f'J{r_33}'].font = MEIRYO

    r_34 = r_33 + 1
    ws[f'A{r_34}'] = '架台+杭　消費税10％'
    ws[f'A{r_34}'].font = MEIRYO
    ws[f'A{r_34}'].alignment = RIGHT_A
    ws.merge_cells(f'A{r_34}:I{r_34}')
    ws.merge_cells(f'J{r_34}:K{r_34}')
    ws[f'J{r_34}'] = float(tax_1)
    ws[f'J{r_34}'].font = MEIRYO

    r_35 = r_34 + 1
    ws[f'A{r_35}'] = '③架台+杭　関税'
    ws[f'A{r_35}'].font = MEIRYO
    ws[f'A{r_35}'].alignment = RIGHT_A
    ws.merge_cells(f'A{r_35}:I{r_35}')
    ws.merge_cells(f'J{r_35}:K{r_35}')
    ws[f'J{r_35}'] = float(tariiff_1)
    ws[f'J{r_35}'].font = MEIRYO

    r_36 = r_35 + 1
    ws[f'A{r_36}'] = '④フェンス税金10％'
    ws[f'A{r_36}'].font = MEIRYO
    ws[f'A{r_36}'].alignment = RIGHT_A
    ws.merge_cells(f'A{r_36}:I{r_36}')
    ws.merge_cells(f'J{r_36}:K{r_36}')
    ws[f'J{r_36}'] = float(fence_tax)
    ws[f'J{r_36}'].font = MEIRYO

    r_37 = r_36 + 1
    ws[f'A{r_37}'] = '⑤'
    ws[f'A{r_37}'].font = MEIRYO
    ws[f'B{r_37}'] = 0
    ws[f'B{r_37}'].font = MEIRYO
    ws[f'F{r_37}'] = '混載便'
    ws[f'F{r_37}'].font = MEIRYO
    ws[f'G{r_37}'] = '4Tユニック+4T平車 配送'
    ws[f'G{r_37}'].font = MEIRYO
    ws.merge_cells(f'G{r_37}:H{r_37}')
    ws[f'I{r_37}'] = 0
    ws[f'I{r_37}'].font = MEIRYO
    ws[f'J{r_37}'] = 0
    ws[f'J{r_37}'].font = MEIRYO
    ws[f'K{r_37}'] = 0
    ws[f'K{r_37}'].font = MEIRYO

    grand_all = grand_total + fence_total + tariiff_1 + fence_tax + shipping_cost

    r_38 = r_37 + 1
    ws[f'A{r_38}'] = '①+②+③+④+⑤総金額(USD)'
    ws[f'A{r_38}'].font = MEIRYO
    ws[f'A{r_38}'].alignment = RIGHT_A
    ws.merge_cells(f'A{r_38}:I{r_38}')
    ws.merge_cells(f'J{r_38}:K{r_38}')
    ws[f'J{r_38}'] = float(grand_all)
    ws[f'J{r_38}'].font = MEIRYO

    r_39 = r_38 + 1
    ws[f'A{r_39}'] = '請求金額'
    ws[f'A{r_39}'].font = MEIRYO
    ws[f'A{r_39}'].alignment = RIGHT_A
    ws.merge_cells(f'A{r_39}:I{r_39}')
    ws.merge_cells(f'J{r_39}:K{r_39}')
    ws[f'J{r_39}'] = float(grand_all)
    ws[f'J{r_39}'].font = MEIRYO

    r_40 = r_39 + 1
    ws[f'A{r_40}'] = '二、運賃'
    ws[f'A{r_40}'].font = MEIRYO_BOLD
    ws.merge_cells(f'A{r_40}:K{r_40}')

    r_41 = r_40 + 1
    ws[f'A{r_41}'] = '項目'
    ws[f'A{r_41}'].font = MEIRYO
    ws.merge_cells(f'A{r_41}:I{r_41}')
    ws[f'J{r_41}'] = '金額（USD）'
    ws[f'J{r_41}'].font = MEIRYO
    ws[f'J{r_41}'].alignment = CENTER
    ws[f'K{r_41}'] = '金額(JPY)'
    ws[f'K{r_41}'].font = MEIRYO
    ws[f'K{r_41}'].alignment = CENTER

    r_42 = r_41 + 1
    ws.merge_cells(f'A{r_42}:H{r_42}')
    ws[f'I{r_42}'] = '①架台本体一式価格'
    ws[f'I{r_42}'].font = MEIRYO
    ws[f'J{r_42}'] = float(grand_total)
    ws[f'J{r_42}'].font = MEIRYO
    ws[f'K{r_42}'] = float(grand_total * Decimal(str(exchange_rate))) if exchange_rate else 0
    ws[f'K{r_42}'].font = MEIRYO

    r_43 = r_42 + 1
    ws.merge_cells(f'A{r_43}:H{r_43}')
    ws[f'I{r_43}'] = '②DDP現地配送費'
    ws[f'I{r_43}'].font = MEIRYO
    ws[f'J{r_43}'] = 0
    ws[f'J{r_43}'].font = MEIRYO
    ws[f'K{r_43}'] = 0
    ws[f'K{r_43}'].font = MEIRYO

    ddp_subtotal = grand_total
    r_44 = r_43 + 1
    ws[f'A{r_44}'] = '①+②='
    ws[f'A{r_44}'].font = MEIRYO
    ws[f'A{r_44}'].alignment = RIGHT_A
    ws.merge_cells(f'A{r_44}:H{r_44}')
    ws[f'I{r_44}'] = '③小計'
    ws[f'I{r_44}'].font = MEIRYO
    ws[f'J{r_44}'] = float(ddp_subtotal)
    ws[f'J{r_44}'].font = MEIRYO
    ws[f'K{r_44}'] = float(ddp_subtotal * Decimal(str(exchange_rate))) if exchange_rate else 0
    ws[f'K{r_44}'].font = MEIRYO

    ddp_tax = ddp_subtotal * Decimal('0.10')
    r_45 = r_44 + 1
    ws.merge_cells(f'A{r_45}:H{r_45}')
    ws[f'I{r_45}'] = '④消費税10％'
    ws[f'I{r_45}'].font = MEIRYO
    ws[f'J{r_45}'] = float(ddp_tax)
    ws[f'J{r_45}'].font = MEIRYO
    ws[f'K{r_45}'] = float(ddp_tax * Decimal(str(exchange_rate))) if exchange_rate else 0
    ws[f'K{r_45}'].font = MEIRYO

    ddp_total = ddp_subtotal + ddp_tax
    r_46 = r_45 + 1
    ws[f'A{r_46}'] = '①+②+③＋④='
    ws[f'A{r_46}'].font = MEIRYO
    ws[f'A{r_46}'].alignment = RIGHT_A
    ws.merge_cells(f'A{r_46}:H{r_46}')
    ws[f'I{r_46}'] = '⑤合計(税込み)'
    ws[f'I{r_46}'].font = MEIRYO
    ws[f'J{r_46}'] = float(ddp_total)
    ws[f'J{r_46}'].font = MEIRYO
    ws[f'K{r_46}'] = float(ddp_total * Decimal(str(exchange_rate))) if exchange_rate else 0
    ws[f'K{r_46}'].font = MEIRYO

    r_47 = r_46 + 1
    ws[f'A{r_47}'] = 'レートウェッブ'
    ws[f'A{r_47}'].font = MEIRYO
    ws.merge_cells(f'A{r_47}:B{r_47}')
    ws[f'C{r_47}'] = 'http://www.safe.gov.cn/safe/rmbhlzjj/index.html'
    ws[f'C{r_47}'].font = MEIRYO
    ws.merge_cells(f'C{r_47}:K{r_47}')

    r_48 = r_47 + 1
    ws[f'A{r_48}'] = '原材料ウェッブ'
    ws[f'A{r_48}'].font = MEIRYO
    ws.merge_cells(f'A{r_48}:B{r_48}')
    ws[f'C{r_48}'] = 'https://market.cnal.com/nanhai/'
    ws[f'C{r_48}'].font = MEIRYO
    ws.merge_cells(f'C{r_48}:K{r_48}')

    _apply_outer_border(ws, 1, r_48, 1, 11)

    for r in range(10, r_48 + 1):
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            if cell.border == Border():
                cell.border = THIN_BORDER

    sheet_names = workbook.sheetnames
    if '合計' in sheet_names:
        idx = sheet_names.index('合計')
        workbook.move_sheet('合計', offset=-idx)

    return ws.title
