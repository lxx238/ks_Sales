import os
import shutil
import uuid

from openpyxl import Workbook

from backend.excel.reader import excel_file_compat
from backend.core.shared.bom_utils import (
    discover_sheet_bom_starts, extract_bom_dataframe, get_bom_processing_rules,
    normalize_selected_bom_keys, quick_scan_bom_sheets, parse_array_to_rows_cols,
    read_bom_from_dataframe,
)
from backend.core.shared.image_utils import scan_images, find_latest_image_log, load_image_mapping_from_log
from backend.core.inquiry_builder import create_inquiry_sheet
from backend.core.array_matcher import build_bom_matrix_data
from backend.core.en_common.quotation_engine import (
    create_detail_sheet,
    create_summary_sheet,
    create_total_materials_sheet,
)
from backend.core.shared.sheet_utils import set_page_break_preview


def split_and_create_quotations(
        input_file,
        price_file_path,
        output_dir=None,
        image_path=None,
        image_folder=None,
        price_mapping_override=None,
        contact_info=None,
        matrix_data=None,
        return_details=False,
        selected_bom_keys=None,
        group=None,
        pre_parsed_products=None,
        pre_parsed_bom_info=None,
        need_weight_code=False,
        sale_type='export',
        ko_discount_rate=100,
        ko_steel_discount_rate=84,
        ko_purchased_discount_rate=94,
        coating_thickness=10,
        trade_method='EXW',
        dest_port='BUSAN',
        container_type='40HQ',
        container_qty=1,
        ko_cif_freight=0,
        pre_parsed_bom_data=None,
        quote_validity='7d',
        en_lang='en',
        container_details=None,
        discount_method='project',
        need_total_qty=False,
        need_total_materials=False,
        **kwargs
):
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(input_file), 'quotation_output')
    os.makedirs(output_dir, exist_ok=True)

    code_to_images = {}
    image_cache = {}
    image_temp_dir = None

    log_search_dirs = [output_dir, os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")]
    latest_log = find_latest_image_log(log_search_dirs)
    if latest_log:
        log_mapping = load_image_mapping_from_log(latest_log)
        if log_mapping:
            code_to_images.update(log_mapping)

    if image_folder and os.path.exists(image_folder):
        scanned = scan_images(image_folder)
        for code, paths in scanned.items():
            if code not in code_to_images:
                code_to_images[code] = paths
        if code_to_images:
            image_temp_dir = os.path.join(output_dir, f"temp_images_{uuid.uuid4().hex}")
            os.makedirs(image_temp_dir, exist_ok=True)
    elif image_folder:
        image_folder = None

    if price_mapping_override is not None:
        price_mapping = price_mapping_override
    else:
        from backend.core.shared.price_utils import load_price_mapping
        price_mapping = load_price_mapping(price_file_path)

    selected_key_set = normalize_selected_bom_keys(selected_bom_keys)
    column_mapping, skip_keywords, non_bom_sheet_keywords = get_bom_processing_rules()

    arrays = (matrix_data or {}).get('arrays') or []
    all_detail_results = []
    all_unmatched_products = []

    used_array_indices = set()

    def find_matching_array(bom_rows, bom_cols, bom_base_count=0):
        if bom_base_count:
            for i, arr in enumerate(arrays):
                if i in used_array_indices:
                    continue
                if bom_rows == arr.get('rows') and bom_cols == arr.get('cols'):
                    if bom_base_count == (arr.get('table_qty') or 1):
                        used_array_indices.add(i)
                        return arr
        for i, arr in enumerate(arrays):
            if i in used_array_indices:
                continue
            if bom_rows == arr.get('rows') and bom_cols == arr.get('cols'):
                used_array_indices.add(i)
                return arr
        return None

    if pre_parsed_products is not None and pre_parsed_bom_info is not None:
        products_by_key = pre_parsed_products
        bom_info_by_key = pre_parsed_bom_info
        span_info_by_key = kwargs.get('pre_parsed_span_info') or {}
    else:
        products_by_key = {}
        bom_info_by_key = {}
        span_info_by_key = {}

        xls = excel_file_compat(input_file)

        bom_sheet_names = quick_scan_bom_sheets(xls, non_bom_sheet_keywords, selected_key_set)

        for sheet_name in bom_sheet_names:
            df = xls.parse(sheet_name=sheet_name, header=None)
            total_rows = len(df)
            if df.empty or total_rows < 5:
                continue
            bom_starts = discover_sheet_bom_starts(df, total_rows, sheet_name=sheet_name)
            for original_index, bom_info in enumerate(bom_starts, 1):
                key = bom_info.get('key', '')
                if selected_key_set and key not in selected_key_set:
                    continue
                bom_df = extract_bom_dataframe(
                    df, bom_info, original_index, bom_starts, total_rows,
                    column_mapping, skip_keywords,
                )
                if bom_df is None or bom_df.empty:
                    continue
                products, array_info, span_info = read_bom_from_dataframe(bom_df)
                if not products:
                    continue
                products_by_key[key] = products
                bom_info_by_key[key] = bom_info
                if span_info:
                    span_info_by_key[key] = span_info

    master_wb = Workbook()
    default_sheet = master_wb.active
    master_wb.remove(default_sheet)

    for key, products in products_by_key.items():
        bom_info = bom_info_by_key.get(key)
        if not bom_info:
            print(f"   [EN-COMMON] skip {key}: no bom_info (keys={list(bom_info_by_key.keys())[:3]})")
            continue

        config = bom_info.get('config', {})
        bom_array_str = config.get('array', '')
        bom_rows, bom_cols = parse_array_to_rows_cols(bom_array_str)

        matched_array = find_matching_array(bom_rows, bom_cols, bom_base_count=config.get('base_count', 0))
        if matched_array is None:
            print(f"   [EN-COMMON] skip {key}: array='{bom_array_str}' rows={bom_rows} cols={bom_cols} base={config.get('base_count',0)} no match in {len(arrays)} arrays")
            continue

        r = matched_array.get('rows', '')
        c = matched_array.get('cols', '')
        sheet_prefix = f'{r}x{c}' if r and c else key.split('::')[0]

        effective_matrix_data = build_bom_matrix_data(
            matrix_data, matched_array, bom_config=config,
        )

        try:
            detail = create_detail_sheet(
                master_wb,
                matched_array,
                products,
                price_mapping,
                sheet_prefix=sheet_prefix,
                image_path=image_path,
                image_folder=image_folder,
                code_to_images=code_to_images,
                image_temp_dir=image_temp_dir,
                image_cache=image_cache,
                matrix_data=effective_matrix_data,
                group=group,
                unmatched_products_out=all_unmatched_products,
                contact_info=contact_info,
                config=config,
                sale_type=sale_type,
                ko_discount_rate=ko_discount_rate,
                ko_steel_discount_rate=ko_steel_discount_rate,
                ko_purchased_discount_rate=ko_purchased_discount_rate,
                coating_thickness=coating_thickness,
                need_weight_code=need_weight_code,
                lang=en_lang,
                discount_method=discount_method,
            )
            detail['config'] = config
            detail['matched_array'] = matched_array
            detail['matrix_data'] = effective_matrix_data
            detail['set_count'] = effective_matrix_data.get('set_count') or 1
            if key in span_info_by_key and span_info_by_key[key]:
                if not config.get('cross_span'):
                    detail['span_info'] = span_info_by_key[key]
            all_detail_results.append(detail)
        except Exception as e:
            print(f"   detail sheet failed: {e}")

    inquiry_file = None
    if all_unmatched_products:
        try:
            inquiry_wb = Workbook()
            inquiry_wb.remove(inquiry_wb.active)
            requester = ''
            if contact_info and contact_info.get('inquiry_requester'):
                requester = contact_info['inquiry_requester']
            create_inquiry_sheet(
                inquiry_wb,
                all_unmatched_products,
                source_sheet_name='询价表',
                inquiry_requester=requester,
            )
            input_basename = os.path.splitext(os.path.basename(input_file))[0]
            inquiry_file = os.path.join(output_dir, f"{input_basename}_询价表.xlsx")
            inquiry_wb.save(inquiry_file)
        except Exception as e:
            print(f"   inquiry sheet failed: {e}")
            inquiry_file = None

    if all_detail_results:
        try:
            create_summary_sheet(
                master_wb,
                all_detail_results,
                matrix_data=matrix_data,
                image_path=image_path,
                ko_discount_rate=ko_discount_rate,
                ko_steel_discount_rate=ko_steel_discount_rate,
                ko_purchased_discount_rate=ko_purchased_discount_rate,
                sale_type=sale_type,
                contact_info=contact_info,
                trade_method=trade_method,
                dest_port=dest_port,
                container_type=container_type,
                container_qty=container_qty,
                ko_cif_freight=ko_cif_freight,
                skip_freight=False,
                quote_validity=quote_validity,
                lang=en_lang,
                container_details=container_details,
                discount_method=discount_method,
                payment_term=kwargs.get('payment_term', '3070shipment'),
                seller_name=kwargs.get('seller_name', ''),
            )
        except Exception as e:
            print(f"   summary sheet failed: {e}")

        try:
            if need_total_materials:
                _tm_title = create_total_materials_sheet(
                    master_wb,
                    all_detail_results,
                    price_mapping=price_mapping,
                    sale_type=sale_type,
                    coating_thickness=coating_thickness,
                    lang=en_lang,
                    need_weight_code=need_weight_code,
                    need_total_qty=need_total_qty,
                    discount_method=discount_method,
                    ko_discount_rate=ko_discount_rate,
                    ko_steel_discount_rate=ko_steel_discount_rate,
                    ko_purchased_discount_rate=ko_purchased_discount_rate,
                )
                if _tm_title and _tm_title in master_wb.sheetnames:
                    _tm_idx = master_wb.sheetnames.index(_tm_title)
                    if _tm_idx > 1:
                        master_wb.move_sheet(_tm_title, offset=-(_tm_idx - 1))
        except Exception as e:
            print(f"   total materials sheet failed: {e}")

        input_basename = os.path.splitext(os.path.basename(input_file))[0]
        output_file = os.path.join(output_dir, f"{input_basename}_Quotation.xlsx")
        set_page_break_preview(master_wb)
        master_wb.save(output_file)

        if image_temp_dir and os.path.exists(image_temp_dir):
            try:
                shutil.rmtree(image_temp_dir)
            except Exception:
                pass

        print(f"\n   {len(all_detail_results)} detail sheets generated")
        print(f"   output: {output_file}")

        if return_details:
            quotation_product_codes = set()
            for r in all_detail_results:
                for c in (r.get('quotation_product_codes') or []):
                    quotation_product_codes.add(c)
            return {
                'output_file': output_file,
                'inquiry_file': inquiry_file,
                'unmatched_count': len(all_unmatched_products),
                'unmatched_products': all_unmatched_products,
                'quotation_product_codes': quotation_product_codes,
            }
        return output_file
    else:
        print(f"\n   no detail sheets generated (products_by_key={len(products_by_key)}, arrays={len(arrays)})")
        if not arrays:
            print(f"   [EN] HINT: no info sheet arrays - check if matrix file was uploaded correctly")
        if image_temp_dir and os.path.exists(image_temp_dir):
            try:
                shutil.rmtree(image_temp_dir)
            except Exception:
                pass
        return None
