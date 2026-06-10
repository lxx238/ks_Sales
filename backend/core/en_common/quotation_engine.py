import math
import os
from decimal import Decimal
from datetime import datetime


def _safe_int(val, default=0):
    try:
        if isinstance(val, float) and math.isnan(val):
            return default
        return int(val)
    except (ValueError, TypeError, OverflowError):
        return default


def _safe_float(val, default=0.0):
    try:
        f = float(val)
        return default if math.isnan(f) else f
    except (ValueError, TypeError):
        return default

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils.units import points_to_pixels

from backend.core.shared.price_utils import (
    resolve_price_info,
    has_valid_price_info,
    round_to_2_decimal,
    _get_discount_category,
)
from backend.core.shared.weight_utils import (
    extract_length_from_spec,
    WEIGHT_BY_LENGTH_ATTRIBUTES,
)
from backend.core.shared.text_utils import (
    normalize_lookup_code,
    extract_main_name,
    parse_decimal_number,
    _strip_cjk_spec,
)
from backend.core.shared.product_utils import (
    _is_valid_product_code,
    _match_exclude_group,
)
from backend.core.shared.constants import (
    EXCLUDE_ITEM_GROUPS,
    CARBON_STEEL_PRICING_ATTRS,
    IMAGE_WIDTH,
    IMAGE_HEIGHT,
    IMAGE_COLUMN_INDEX,
    IMAGE_PADDING,
)
from backend.core.shared.image_utils import (
    prepare_image_for_excel,
    add_image_centered_in_cell,
)
from backend.core.material_translate import translate_material, adjust_material_by_coating
from backend.core.shared.translations import t as _t

CURRENCY_FMT = '#,##0.00'
BLUE_FILL = PatternFill(patternType='solid', fgColor=Color(theme=3, tint=0.6))

_LANG_NAME_KEY_MAP = {
    'en': ('name_en',),
    'fr': ('name_fr', 'name_en'),
    'es': ('name_es', 'name_en'),
}


def _resolve_product_name(price_info, product, lang='en'):
    fallback = product.get('name', '')
    if not price_info:
        return fallback
    for key in _LANG_NAME_KEY_MAP.get(lang, ('name_en',)):
        val = price_info.get(key)
        if val:
            return val
    for key in ('name_ko', 'name'):
        val = price_info.get(key)
        if val:
            return val
    return fallback


def _sc(ws, row, col, val, font=None, align=None, border=None, number_format=None, fill=None):
    cell = ws.cell(row=row, column=col, value=val)
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format
    if fill is not None:
        cell.fill = fill
    return cell


def _is_carbon_steel_cached(product, price_info):
    if price_info:
        attr = str(price_info.get('pricing_attribute', '')).strip().upper()
        if attr in CARBON_STEEL_PRICING_ATTRS:
            return True
    return False


def _classify_products_single_pass(all_products, price_mapping, delete_options, always_exclude, ko_exclude_options=None):
    delete_options = delete_options or {}
    ko_exclude_options = ko_exclude_options or {}
    all_exclude_opts = {k: True for k in EXCLUDE_ITEM_GROUPS} if always_exclude else {}
    for k in list(ko_exclude_options.keys()) + list(delete_options.keys()):
        all_exclude_opts[k] = True
    aluminum = []
    carbon_steel = []
    excluded = []
    pile_products = []
    price_info_cache = {}
    for p in all_products:
        code = str(p.get('code', '')).strip()
        spec = str(p.get('spec', '')).strip()
        cache_key = (code, spec) if spec else code
        if cache_key not in price_info_cache:
            price_info_cache[cache_key] = resolve_price_info(price_mapping, code, spec=spec) if price_mapping else None
        pi = price_info_cache[cache_key]
        p['_price_info'] = pi
        attr = (pi.get('attribute', '') if pi else '') or ''
        if attr.strip() in ('地桩', '地盤杭'):
            p['_is_pile'] = True
            pile_products.append(p)
            continue
        if not _is_valid_product_code(code):
            continue
        qty = p.get('quantity', 0)
        if not qty or qty <= 0:
            continue
        matched_group = _match_exclude_group(p, price_mapping, all_exclude_opts)
        is_cs = _is_carbon_steel_cached(p, pi)
        if matched_group:
            if delete_options.get(matched_group):
                continue
            if ko_exclude_options.get(matched_group):
                excluded.append(p)
            else:
                if is_cs:
                    carbon_steel.append(p)
                else:
                    aluminum.append(p)
        elif is_cs:
            carbon_steel.append(p)
        else:
            aluminum.append(p)
    return aluminum, carbon_steel, excluded, pile_products, price_info_cache


def create_detail_sheet(workbook, array_info, bom_products, price_mapping,
                        sheet_prefix=None,
                        image_path=None, image_folder=None, code_to_images=None,
                        image_temp_dir=None, image_cache=None, matrix_data=None,
                        group=None, unmatched_products_out=None,
                        contact_info=None, config=None,
                        sale_type='export',
                        ko_discount_rate=100, ko_steel_discount_rate=84,
                        ko_purchased_discount_rate=94, coating_thickness=10,
                        delete_options=None, always_exclude_extra_items=False,
                        ko_exclude_options=None, need_weight_code=False,
                        lang='en',
                        discount_method='project',
                        **kwargs):
    all_products = list(bom_products) if bom_products else []

    if sale_type == 'domestic':
        currency_label = 'RMB'
    elif sale_type == 'euro':
        currency_label = 'EUR'
    else:
        currency_label = 'USD'

    aluminum, carbon_steel, excluded, pile_products, price_info_cache = _classify_products_single_pass(
        all_products, price_mapping, delete_options or {},
        always_exclude_extra_items, ko_exclude_options or {}
    )

    matrix_data = matrix_data or {}
    rows = array_info.get('rows') if isinstance(array_info, dict) else None
    cols = array_info.get('cols') if isinstance(array_info, dict) else None
    table_qty = array_info.get('table_qty', 1) if isinstance(array_info, dict) else 1
    table_qty = _safe_int(table_qty, 1)
    config = config or {}

    sheet_name = sheet_prefix or ''
    if not sheet_name:
        sheet_name = f'{rows}x{cols}' if rows and cols else 'Detail'
    sheet_name = extract_main_name(sheet_name)
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    original = sheet_name
    counter = 1
    while sheet_name in workbook.sheetnames:
        sheet_name = f"{original}_{counter}"
        counter += 1

    ws = workbook.create_sheet(title=sheet_name)
    print(f"[EN-COMMON] Creating detail sheet: {sheet_name}")

    thin = Side(style='thin', color='000000')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(name='Arial', size=22, bold=True)
    normal_font = Font(name='Arial', size=11)
    bold_font = Font(name='Arial', size=11, bold=True)
    total_font = Font(name='Arial', size=12, bold=True)

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)

    col_widths = {
        'A': 8.125, 'B': 21.875, 'C': 18.5, 'D': 20,
        'E': 13.875, 'F': 20, 'G': 20,
        'H': 20, 'I': 20,
    }
    if need_weight_code:
        col_widths['J'] = 20
        col_widths['K'] = 22.31
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    max_col = 9
    if need_weight_code:
        max_col = 11

    max_col_letter = get_column_letter(max_col)

    ws.merge_cells(f'A1:{max_col_letter}1')
    _sc(ws, 1, 1, _t('common_detail_title', lang), font=title_font, align=center_align, border=thin_border, fill=BLUE_FILL)
    for c in range(2, max_col + 1):
        ws.cell(row=1, column=c).border = thin_border
    ws.row_dimensions[1].height = 53

    header_row = 2
    qty_header = _t('common_hdr_qty_in_table', lang).format(rows, cols) if rows and cols else _t('common_hdr_qty_in_table_simple', lang)
    headers = [_t('common_hdr_no', lang), _t('common_hdr_item', lang), _t('common_hdr_picture', lang), _t('common_hdr_material', lang), _t('common_hdr_spec', lang), qty_header, _t('common_hdr_total_qty', lang), _t('common_hdr_unit_price', lang).format(currency=currency_label), _t('common_hdr_total_amount', lang).format(currency=currency_label)]
    if need_weight_code:
        headers.append(_t('hdr_weight', lang))
        headers.append(_t('hdr_remark', lang))
    for ci, h in enumerate(headers):
        cell = ws.cell(row=header_row, column=ci + 1, value=h)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = BLUE_FILL
    ws.row_dimensions[header_row].height = 45

    combined_products = aluminum + carbon_steel
    all_render_products = combined_products + excluded
    data_start_row = 3
    matched_count = 0
    unmatched_count = 0
    local_unmatched = []
    row_product_map = {}
    standard_rows = []
    steel_rows = []
    purchased_rows = []
    inquiry_rows = []
    unmatched_rows = []
    standard_price_sum = Decimal('0')
    steel_price_sum = Decimal('0')
    purchased_price_sum = Decimal('0')
    inquiry_price_sum = Decimal('0')

    current_row = data_start_row

    if combined_products:
        for idx, product in enumerate(combined_products):
            row = current_row
            ws.row_dimensions[row].height = 60
            product_code = product['code']
            price_info = product.get('_price_info')
            if price_info is None:
                price_info = resolve_price_info(price_mapping, product_code, spec=product.get('spec', ''))
                product['_price_info'] = price_info
            en_name = _resolve_product_name(price_info, product, lang)

            row_product_map[row] = product

            _sc(ws, row, 1, idx + 1, font=normal_font, align=center_align, border=thin_border)
            _sc(ws, row, 2, en_name, font=normal_font, align=center_align, border=thin_border)
            _sc(ws, row, 3, '', font=normal_font, align=center_align, border=thin_border)

            _raw_mat = (price_info.get('db_material') if price_info and price_info.get('db_material') else None) or product.get('material', '')
            _sc(ws, row, 4, adjust_material_by_coating(translate_material(_raw_mat, 'en'), coating_thickness),
                font=normal_font, align=center_align, border=thin_border)

            ws.cell(row=row, column=5).number_format = '@'
            _sc(ws, row, 5, _strip_cjk_spec(str(product['spec']).strip()),
                font=normal_font, align=center_align, border=thin_border)

            quantity = _safe_float(product['quantity'])
            if quantity > 0:
                _sc(ws, row, 6, _safe_int(quantity) if quantity % 1 == 0 else quantity,
                    font=normal_font, align=center_align, border=thin_border)
            else:
                _sc(ws, row, 6, "", font=normal_font, align=center_align, border=thin_border)

            total_qty = quantity * table_qty
            if total_qty > 0:
                _sc(ws, row, 7, _safe_int(total_qty) if total_qty % 1 == 0 else total_qty,
                    font=normal_font, align=center_align, border=thin_border)
            else:
                _sc(ws, row, 7, "", font=normal_font, align=center_align, border=thin_border)

            unit_price = 0
            display_unit_price = 0
            price_unit = ''
            is_matched = False

            if price_info and has_valid_price_info(price_info):
                unit_price = float(price_info['price'])
                price_unit = price_info.get('unit', '')
                matched_count += 1
                is_matched = True
            else:
                unmatched_count += 1
                if unmatched_products_out is not None and _is_valid_product_code(product_code):
                    local_unmatched.append({
                        'code': product_code,
                        'name': price_info.get('name', '') if price_info else product.get('name', ''),
                        'spec': product.get('spec', ''),
                        'quantity': product.get('quantity', 0),
                        'unit': price_info.get('unit', '') if price_info else '',
                        'issue_reason': 'No price match',
                    })

            is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
            length_mm = Decimal('0')
            if is_meter:
                length_mm = Decimal(str(extract_length_from_spec(product['spec']) or 0))
            if is_meter and length_mm > 0:
                display_unit_price = float(Decimal(str(unit_price)) * length_mm / Decimal('1000'))
            else:
                display_unit_price = unit_price

            _original_unit_price = display_unit_price
            if discount_method == 'unit_price' and display_unit_price > 0 and is_matched:
                _cat = _get_discount_category(price_info)
                if _cat == 'steel':
                    _rate = ko_steel_discount_rate
                elif _cat == 'purchased':
                    _rate = ko_purchased_discount_rate
                else:
                    _rate = ko_discount_rate
                display_unit_price = display_unit_price * _rate / 100

            if display_unit_price > 0:
                if discount_method == 'unit_price' and _original_unit_price > 0 and is_matched:
                    _sc(ws, row, 8, f'={round(_original_unit_price, 6)}*{_rate}/100',
                        font=normal_font, align=center_align, border=thin_border,
                        number_format=CURRENCY_FMT)
                else:
                    _sc(ws, row, 8, float(display_unit_price),
                        font=normal_font, align=center_align, border=thin_border,
                        number_format=CURRENCY_FMT)
            else:
                _sc(ws, row, 8, "", font=normal_font, align=center_align, border=thin_border)

            if display_unit_price > 0 and total_qty > 0:
                _sc(ws, row, 9, f"=H{row}*G{row}",
                    font=normal_font, align=center_align, border=thin_border,
                    number_format=CURRENCY_FMT)
                total_price = Decimal(str(display_unit_price)) * Decimal(str(total_qty))
                cat = _get_discount_category(price_info)
                if cat == 'standard':
                    standard_rows.append(row)
                    standard_price_sum += total_price
                elif cat == 'steel':
                    steel_rows.append(row)
                    steel_price_sum += total_price
                elif cat == 'purchased':
                    purchased_rows.append(row)
                    purchased_price_sum += total_price
                else:
                    inquiry_rows.append(row)
                    inquiry_price_sum += total_price
            else:
                _sc(ws, row, 9, f"=H{row}*G{row}",
                    font=normal_font, align=center_align, border=thin_border,
                    number_format=CURRENCY_FMT)
                if not is_matched:
                    unmatched_rows.append(row)
                    for c in range(1, max_col + 1):
                        ws.cell(row=row, column=c).fill = yellow_fill

            if need_weight_code:
                weight_cell = ws.cell(row=row, column=10)
                unit_weight = None
                if price_info:
                    db_w = parse_decimal_number(price_info.get('db_weight'))
                    if db_w is not None and db_w > 0:
                        code_attribute = str(price_info.get('code_attribute') or '').strip().upper()
                        unit_weight = db_w
                        if code_attribute in WEIGHT_BY_LENGTH_ATTRIBUTES:
                            length_mm_w = extract_length_from_spec(product.get('spec'))
                            if length_mm_w and length_mm_w > 0:
                                unit_weight = db_w * Decimal(str(length_mm_w)) / Decimal('1000')
                            else:
                                unit_weight = None
                if unit_weight is None:
                    bom_w = product.get('weight', 0)
                    if bom_w and bom_w > 0:
                        unit_weight = Decimal(str(bom_w))
                if unit_weight is not None and unit_weight > 0:
                    weight_cell.value = float(round_to_2_decimal(unit_weight))
                    weight_cell.alignment = center_align
                    weight_cell.font = normal_font
                    weight_cell.border = thin_border
                    weight_cell.number_format = '#,##0.00'
                else:
                    weight_cell.value = ""
                    weight_cell.alignment = center_align
                    weight_cell.font = normal_font
                    weight_cell.border = thin_border
                _sc(ws, row, 11, product_code, font=normal_font, align=center_align, border=thin_border)

            current_row += 1

    part1_data_end = current_row - 1

    if excluded:
        current_row += 1
        ws.merge_cells(f'A{current_row}:{max_col_letter}{current_row}')
        _sc(ws, current_row, 1, _t('common_excluded_title', lang),
            font=Font(name='Arial', size=11, bold=True, color='FF0000'), align=left_align, fill=BLUE_FILL)
        current_row += 1
        for ci, h in enumerate(headers):
            cell = ws.cell(row=current_row, column=ci + 1, value=h)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = BLUE_FILL
        ws.row_dimensions[current_row].height = 36
        current_row += 1

        for product in excluded:
            row = current_row
            ws.row_dimensions[row].height = 40
            product_code = product['code']
            price_info = product.get('_price_info')
            if price_info is None:
                price_info = resolve_price_info(price_mapping, product_code, spec=product.get('spec', ''))
                product['_price_info'] = price_info
            en_name = _resolve_product_name(price_info, product, lang)

            _sc(ws, row, 1, '', font=normal_font, align=center_align, border=thin_border)
            _sc(ws, row, 2, en_name, font=normal_font, align=center_align, border=thin_border)
            _raw_mat = (price_info.get('db_material') if price_info and price_info.get('db_material') else None) or product.get('material', '')
            _sc(ws, row, 4, adjust_material_by_coating(translate_material(_raw_mat, 'en'), coating_thickness),
                font=normal_font, align=center_align, border=thin_border)
            ws.cell(row=row, column=5).number_format = '@'
            _sc(ws, row, 5, _strip_cjk_spec(str(product['spec']).strip()), font=normal_font, align=center_align, border=thin_border)

            quantity = _safe_float(product['quantity'])
            total_qty = quantity * table_qty
            _sc(ws, row, 6, _safe_int(quantity) if quantity % 1 == 0 else quantity,
                font=normal_font, align=center_align, border=thin_border)
            _sc(ws, row, 7, _safe_int(total_qty) if total_qty % 1 == 0 else total_qty,
                font=normal_font, align=center_align, border=thin_border)

            unit_price = 0
            display_unit_price = 0
            price_unit = ''
            ex_is_matched = False
            if price_info and has_valid_price_info(price_info):
                unit_price = float(price_info['price'])
                price_unit = price_info.get('unit', '')
                ex_is_matched = True
            is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
            length_mm = Decimal(str(extract_length_from_spec(product['spec']) or 0))
            if is_meter and length_mm > 0:
                display_unit_price = float(Decimal(str(unit_price)) * length_mm / Decimal('1000'))
            else:
                display_unit_price = unit_price

            _ex_original_price = display_unit_price
            if discount_method == 'unit_price' and display_unit_price > 0 and ex_is_matched:
                _ex_cat = _get_discount_category(price_info)
                if _ex_cat == 'steel':
                    _ex_rate = ko_steel_discount_rate
                elif _ex_cat == 'purchased':
                    _ex_rate = ko_purchased_discount_rate
                else:
                    _ex_rate = ko_discount_rate
                display_unit_price = display_unit_price * _ex_rate / 100

            if display_unit_price > 0:
                if discount_method == 'unit_price' and _ex_original_price > 0 and ex_is_matched:
                    _sc(ws, row, 8, f'={round(_ex_original_price, 6)}*{_ex_rate}/100', font=normal_font, align=center_align, border=thin_border, number_format=CURRENCY_FMT)
                else:
                    _sc(ws, row, 8, float(display_unit_price), font=normal_font, align=center_align, border=thin_border, number_format=CURRENCY_FMT)
            else:
                _sc(ws, row, 8, "", font=normal_font, align=center_align, border=thin_border)
            _sc(ws, row, 9, f"=H{row}*G{row}", font=normal_font, align=center_align, border=thin_border, number_format=CURRENCY_FMT)

            if need_weight_code:
                weight_cell = ws.cell(row=row, column=10)
                unit_weight = None
                if price_info:
                    db_w = parse_decimal_number(price_info.get('db_weight'))
                    if db_w is not None and db_w > 0:
                        code_attribute = str(price_info.get('code_attribute') or '').strip().upper()
                        unit_weight = db_w
                        if code_attribute in WEIGHT_BY_LENGTH_ATTRIBUTES:
                            length_mm_w = extract_length_from_spec(product.get('spec'))
                            if length_mm_w and length_mm_w > 0:
                                unit_weight = db_w * Decimal(str(length_mm_w)) / Decimal('1000')
                            else:
                                unit_weight = None
                if unit_weight is None:
                    bom_w = product.get('weight', 0)
                    if bom_w and bom_w > 0:
                        unit_weight = Decimal(str(bom_w))
                if unit_weight is not None and unit_weight > 0:
                    weight_cell.value = float(round_to_2_decimal(unit_weight))
                    weight_cell.alignment = center_align
                    weight_cell.font = normal_font
                    weight_cell.border = thin_border
                    weight_cell.number_format = '#,##0.00'
                else:
                    weight_cell.value = ""
                    weight_cell.alignment = center_align
                    weight_cell.font = normal_font
                    weight_cell.border = thin_border
                _sc(ws, row, 11, product_code, font=normal_font, align=center_align, border=thin_border)

            if not ex_is_matched:
                for c in range(1, max_col + 1):
                    ws.cell(row=row, column=c).fill = yellow_fill

            current_row += 1

    total_row = current_row
    ws.merge_cells(f'A{total_row}:G{total_row}')
    _sc(ws, total_row, 1, _t('common_total_amount_label', lang).format(currency=currency_label), font=total_font, align=right_align, border=thin_border, fill=BLUE_FILL)
    for c in range(2, 8):
        ws.cell(row=total_row, column=c).border = thin_border
    if part1_data_end >= data_start_row:
        _sc(ws, total_row, 8, f'=SUM(I{data_start_row}:I{current_row - 1})', font=total_font, align=center_align, border=thin_border, number_format=CURRENCY_FMT)
    else:
        _sc(ws, total_row, 8, 0, font=total_font, align=center_align, border=thin_border, number_format=CURRENCY_FMT)
    ws.merge_cells(f'H{total_row}:I{total_row}')
    for c in range(5, max_col + 1):
        ws.cell(row=total_row, column=c).border = thin_border
    ws.row_dimensions[total_row].height = 33

    image_found_count = 0
    image_not_found_count = 0
    last_data_row = current_row - 1

    for row in range(data_start_row, last_data_row + 1):
        if row not in row_product_map:
            continue
        product = row_product_map[row]
        product_code = str(product.get('code', '')).strip()
        if not product_code:
            continue
        normalized_code = normalize_lookup_code(product_code)

        img_path = None
        if code_to_images and product_code in code_to_images and code_to_images[product_code]:
            img_path = code_to_images[product_code][0]
        elif code_to_images and normalized_code in code_to_images and code_to_images[normalized_code]:
            img_path = code_to_images[normalized_code][0]

        col_w = ws.column_dimensions['C'].width or 18.5
        row_h = ws.row_dimensions[row].height or 60
        avail_w = int(col_w * 7.5 + 5) - IMAGE_PADDING * 2
        avail_h = points_to_pixels(row_h) - IMAGE_PADDING * 2
        avail_w = max(avail_w, 20)
        avail_h = max(avail_h, 20)

        if img_path and image_folder:
            scale = min(avail_w / IMAGE_WIDTH, avail_h / IMAGE_HEIGHT, 1.0)
            fit_w = int(IMAGE_WIDTH * scale)
            fit_h = int(IMAGE_HEIGHT * scale)
            temp_img_path = prepare_image_for_excel(
                img_path, target_width=fit_w, target_height=fit_h,
                temp_dir=image_temp_dir, cache=image_cache
            )
            final_img_path = temp_img_path if temp_img_path else img_path
            success = add_image_centered_in_cell(
                ws, final_img_path, row, 3,
                img_width=fit_w, img_height=fit_h
            )
            if success:
                image_found_count += 1
                continue
            image_not_found_count += 1
            ws.cell(row=row, column=3, value='/').alignment = center_align
        else:
            image_not_found_count += 1
            ws.cell(row=row, column=3, value='/').alignment = center_align

    print(f"   [EN-COMMON] Images: found={image_found_count}, not_found={image_not_found_count}")

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(top=0.5, bottom=0.5, left=0.25, right=0.25, header=0.3, footer=0.3)

    if unmatched_products_out is not None:
        for _up in local_unmatched:
            _up['quantity'] = float(_up.get('quantity', 0)) * table_qty
        unmatched_products_out.extend(local_unmatched)

    quotation_product_codes = set()
    for p in combined_products + excluded:
        c = str(p.get('code', '')).strip()
        if c:
            quotation_product_codes.add(c)

    part1_price_per_table = float(standard_price_sum + steel_price_sum + purchased_price_sum + inquiry_price_sum) / table_qty if table_qty > 1 else float(standard_price_sum + steel_price_sum + purchased_price_sum + inquiry_price_sum)
    part1_price_all_tables = float(standard_price_sum + steel_price_sum + purchased_price_sum + inquiry_price_sum)

    return {
        'sheet_name': sheet_name,
        'quotation_product_codes': quotation_product_codes,
        'valid_products': len(combined_products) + len(excluded),
        'total_weight': 0,
        'total_price': part1_price_all_tables,
        'part1_price_per_table': part1_price_per_table,
        'part1_price_all_tables': part1_price_all_tables,
        'part1_total_table_row': total_row,
        'part1_standard_rows': standard_rows,
        'part1_steel_rows': steel_rows,
        'part1_purchased_rows': purchased_rows,
        'part1_inquiry_rows': inquiry_rows,
        'part1_unmatched_rows': unmatched_rows,
        'part1_standard_total': float(standard_price_sum),
        'part1_steel_total': float(steel_price_sum),
        'part1_purchased_total': float(purchased_price_sum),
        'part1_inquiry_total': float(inquiry_price_sum),
        'part2_price_per_table': 0,
        'part2_price_all_tables': 0,
        'part2_total_table_row': 0,
        'part2_standard_rows': [],
        'part2_steel_rows': [],
        'part2_purchased_rows': [],
        'part2_inquiry_rows': [],
        'part2_unmatched_rows': [],
        'is_complex': False,
        'part2_standard_total': 0,
        'part2_steel_total': 0,
        'part2_purchased_total': 0,
        'part2_inquiry_total': 0,
        'matched_count': matched_count,
        'unmatched_count': unmatched_count,
        'unmatched_products': local_unmatched,
        'image_found_count': image_found_count,
        'image_not_found_count': image_not_found_count,
        'pile_products': pile_products,
        'all_render_products': all_render_products,
        'sub_total_row': total_row,
        'total_row': total_row,
        'detail_data_end_row': last_data_row,
        'config': config,
        'matrix_data': matrix_data,
        'discount_method': discount_method,
    }


def create_total_materials_sheet(workbook, all_quotation_results, price_mapping=None,
                                  sale_type='export', coating_thickness=10, lang='en',
                                  need_weight_code=False, need_total_qty=False,
                                  discount_method='project',
                                  ko_discount_rate=100, ko_steel_discount_rate=84,
                                  ko_purchased_discount_rate=94, **kwargs):
    print(f"   [EN-COMMON] Total materials sheet: need_weight_code={need_weight_code}, need_total_qty={need_total_qty}")
    if sale_type == 'domestic':
        currency_label = 'RMB'
    elif sale_type == 'euro':
        currency_label = 'EUR'
    else:
        currency_label = 'USD'

    consolidated = {}
    for qr in all_quotation_results:
        products = qr.get('all_render_products') or []
        config = qr.get('config') or {}
        table_qty = 1
        matched_array = qr.get('matched_array') or {}
        if matched_array:
            table_qty = _safe_int(matched_array.get('table_qty', 1), 1)
        md = qr.get('matrix_data') or {}
        if md.get('set_count'):
            table_qty = _safe_int(md['set_count'], table_qty)

        for p in products:
            code = str(p.get('code', '')).strip()
            if not code or not _is_valid_product_code(code):
                continue
            qty = p.get('quantity', 0)
            if not qty or qty <= 0:
                continue
            per_table_qty = qty
            total_qty = qty * table_qty
            spec = str(p.get('spec', '')).strip()
            key = (code, spec)
            price_info = p.get('_price_info')
            if price_info is None and price_mapping:
                price_info = resolve_price_info(price_mapping, code, spec=spec)
            en_name = _resolve_product_name(price_info, p, lang)
            _raw_mat = (price_info.get('db_material') if price_info and price_info.get('db_material') else None) or p.get('material', '')
            material = adjust_material_by_coating(translate_material(_raw_mat, 'en'), coating_thickness)
            unit_price = 0
            price_unit = ''
            if price_info and has_valid_price_info(price_info):
                unit_price = float(price_info['price'])
                price_unit = price_info.get('unit', '')
            is_meter = price_unit in ['米', 'm', 'M', 'meter', 'Meter', 'METERS', 'meters']
            length_mm = Decimal('0')
            if is_meter:
                length_mm = Decimal(str(extract_length_from_spec(spec) or 0))
            if is_meter and length_mm > 0:
                display_unit_price = float(Decimal(str(unit_price)) * length_mm / Decimal('1000'))
            else:
                display_unit_price = unit_price

            if discount_method == 'unit_price' and display_unit_price > 0 and price_info and has_valid_price_info(price_info):
                _tm_cat = _get_discount_category(price_info)
                if _tm_cat == 'steel':
                    _tm_rate = ko_steel_discount_rate
                elif _tm_cat == 'purchased':
                    _tm_rate = ko_purchased_discount_rate
                else:
                    _tm_rate = ko_discount_rate
                display_unit_price = display_unit_price * _tm_rate / 100

            weight_val = None
            if need_weight_code:
                if price_info:
                    db_w = parse_decimal_number(price_info.get('db_weight'))
                    if db_w is not None and db_w > 0:
                        code_attribute = str(price_info.get('code_attribute') or '').strip().upper()
                        weight_val = db_w
                        if code_attribute in WEIGHT_BY_LENGTH_ATTRIBUTES:
                            length_mm_w = extract_length_from_spec(spec)
                            if length_mm_w and length_mm_w > 0:
                                weight_val = db_w * Decimal(str(length_mm_w)) / Decimal('1000')
                            else:
                                weight_val = None
                if weight_val is None:
                    bom_w = p.get('weight', 0)
                    if bom_w and bom_w > 0:
                        weight_val = Decimal(str(bom_w))

            if key in consolidated:
                consolidated[key]['total_qty'] += total_qty
                consolidated[key]['amount'] += Decimal(str(display_unit_price)) * Decimal(str(total_qty))
                if need_total_qty:
                    consolidated[key]['per_table_qty'] += per_table_qty
                if need_weight_code and weight_val is not None and weight_val > 0:
                    consolidated[key]['weight'] = (consolidated[key].get('weight') or Decimal('0')) + weight_val * Decimal(str(total_qty))
            else:
                consolidated[key] = {
                    'code': code,
                    'name': en_name,
                    'material': material,
                    'spec': spec,
                    'total_qty': total_qty,
                    'unit_price': display_unit_price,
                    'amount': Decimal(str(display_unit_price)) * Decimal(str(total_qty)),
                }
                if need_total_qty:
                    consolidated[key]['per_table_qty'] = per_table_qty
                if need_weight_code and weight_val is not None and weight_val > 0:
                    consolidated[key]['weight'] = weight_val * Decimal(str(total_qty))
                else:
                    consolidated[key]['weight'] = None

    sorted_items = sorted(consolidated.values(), key=lambda x: x['code'])

    ws = workbook.create_sheet(title=_t('total_title', lang))

    title_font = Font(name='Arial', size=16, bold=True)
    normal_font = Font(name='Arial', size=11)
    bold_font = Font(name='Arial', size=11, bold=True)
    total_font = Font(name='Arial', size=12, bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_idx = 1
    col_no = col_idx; col_idx += 1
    col_item = col_idx; col_idx += 1
    col_material = col_idx; col_idx += 1
    col_spec = col_idx; col_idx += 1
    if need_total_qty:
        col_per_qty = col_idx; col_idx += 1
    col_qty = col_idx; col_idx += 1
    col_unit_price = col_idx; col_idx += 1
    col_amount = col_idx; col_idx += 1
    if need_weight_code:
        col_weight = col_idx; col_idx += 1
        col_code = col_idx; col_idx += 1
    max_col = col_idx - 1

    col_widths = {}
    col_widths[get_column_letter(col_no)] = 8
    col_widths[get_column_letter(col_item)] = 28
    col_widths[get_column_letter(col_material)] = 18
    col_widths[get_column_letter(col_spec)] = 24
    ws.column_dimensions['C'].width = 27
    if need_total_qty:
        col_widths[get_column_letter(col_per_qty)] = 14
    col_widths[get_column_letter(col_qty)] = 14
    col_widths[get_column_letter(col_unit_price)] = 16
    col_widths[get_column_letter(col_amount)] = 18
    if need_weight_code:
        col_widths[get_column_letter(col_weight)] = 14
        col_widths[get_column_letter(col_code)] = 22
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    max_col_letter = get_column_letter(max_col)
    ws.merge_cells(f'A1:{max_col_letter}1')
    ws.cell(row=1, column=1, value=_t('total_title', lang)).font = title_font
    ws.cell(row=1, column=1).alignment = center_align
    ws.cell(row=1, column=1).border = thin_border
    for c in range(2, max_col + 1):
        ws.cell(row=1, column=c).border = thin_border
    ws.row_dimensions[1].height = 36

    header_row = 2
    headers = {
        col_no: _t('total_hdr_no', lang),
        col_item: _t('total_hdr_item', lang),
        col_material: _t('total_hdr_material', lang),
        col_spec: _t('total_hdr_spec', lang),
    }
    if need_total_qty:
        from backend.core.shared.translations import t as _t_tm
        headers[col_per_qty] = 'QTY/Table'
    headers[col_qty] = _t('total_hdr_qty', lang)
    headers[col_unit_price] = _t('total_hdr_unit_price', lang).format(currency=currency_label)
    headers[col_amount] = _t('total_hdr_amount', lang).format(currency=currency_label)
    if need_weight_code:
        headers[col_weight] = _t('hdr_weight', lang)
        headers[col_code] = _t('hdr_remark', lang)
    for ci, h in headers.items():
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = BLUE_FILL
    ws.row_dimensions[header_row].height = 30

    data_start = 3
    grand_total = Decimal('0')
    for i, item in enumerate(sorted_items):
        r = data_start + i

        ws.cell(row=r, column=col_no, value=i + 1).font = normal_font
        ws.cell(row=r, column=col_no).alignment = center_align
        ws.cell(row=r, column=col_no).border = thin_border

        ws.cell(row=r, column=col_item, value=item['name']).font = normal_font
        ws.cell(row=r, column=col_item).alignment = center_align
        ws.cell(row=r, column=col_item).border = thin_border

        ws.cell(row=r, column=col_material, value=item['material']).font = normal_font
        ws.cell(row=r, column=col_material).alignment = center_align
        ws.cell(row=r, column=col_material).border = thin_border

        ws.cell(row=r, column=col_spec).number_format = '@'
        ws.cell(row=r, column=col_spec, value=_strip_cjk_spec(str(item['spec']).strip())).font = normal_font
        ws.cell(row=r, column=col_spec).alignment = center_align
        ws.cell(row=r, column=col_spec).border = thin_border

        if need_total_qty:
            pq = _safe_float(item.get('per_table_qty', 0))
            ws.cell(row=r, column=col_per_qty, value=_safe_int(pq) if pq % 1 == 0 else pq).font = normal_font
            ws.cell(row=r, column=col_per_qty).alignment = center_align
            ws.cell(row=r, column=col_per_qty).border = thin_border

        tq = _safe_float(item['total_qty'])
        ws.cell(row=r, column=col_qty, value=_safe_int(tq) if tq % 1 == 0 else tq).font = normal_font
        ws.cell(row=r, column=col_qty).alignment = center_align
        ws.cell(row=r, column=col_qty).border = thin_border

        up = item['unit_price']
        if up > 0:
            ws.cell(row=r, column=col_unit_price, value=float(up)).font = normal_font
            ws.cell(row=r, column=col_unit_price).number_format = CURRENCY_FMT
        else:
            ws.cell(row=r, column=col_unit_price, value='').font = normal_font
        ws.cell(row=r, column=col_unit_price).alignment = center_align
        ws.cell(row=r, column=col_unit_price).border = thin_border

        amt = item['amount']
        grand_total += amt
        if amt > 0:
            ws.cell(row=r, column=col_amount, value=float(amt)).font = normal_font
            ws.cell(row=r, column=col_amount).number_format = CURRENCY_FMT
        else:
            ws.cell(row=r, column=col_amount, value='').font = normal_font
        ws.cell(row=r, column=col_amount).alignment = center_align
        ws.cell(row=r, column=col_amount).border = thin_border

        if need_weight_code:
            wv = item.get('weight')
            if wv is not None and float(wv) > 0:
                ws.cell(row=r, column=col_weight, value=float(round_to_2_decimal(wv))).font = normal_font
                ws.cell(row=r, column=col_weight).number_format = '#,##0.00'
            else:
                ws.cell(row=r, column=col_weight, value='').font = normal_font
            ws.cell(row=r, column=col_weight).alignment = center_align
            ws.cell(row=r, column=col_weight).border = thin_border

            ws.cell(row=r, column=col_code, value=item.get('code', '')).font = normal_font
            ws.cell(row=r, column=col_code).alignment = center_align
            ws.cell(row=r, column=col_code).border = thin_border

        ws.row_dimensions[r].height = 22

    data_end = data_start + len(sorted_items) - 1

    if sorted_items:
        total_row_num = data_end + 1
        ws.merge_cells(f'A{total_row_num}:{get_column_letter(col_amount - 1)}{total_row_num}')
        ws.cell(row=total_row_num, column=1, value=_t('total_grand_total', lang)).font = total_font
        ws.cell(row=total_row_num, column=1).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=total_row_num, column=1).border = thin_border
        for cc in range(2, max_col + 1):
            ws.cell(row=total_row_num, column=cc).border = thin_border
        ws.cell(row=total_row_num, column=col_amount, value=float(grand_total)).font = total_font
        ws.cell(row=total_row_num, column=col_amount).alignment = center_align
        ws.cell(row=total_row_num, column=col_amount).border = thin_border
        ws.cell(row=total_row_num, column=col_amount).number_format = CURRENCY_FMT
        ws.row_dimensions[total_row_num].height = 28

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    from openpyxl.worksheet.page import PageMargins
    ws.page_margins = PageMargins(top=0.5, bottom=0.5, left=0.4, right=0.4, header=0.3, footer=0.3)

    print(f"   [EN-COMMON] Total materials sheet created ({len(sorted_items)} items)")
    return ws.title


def create_summary_sheet(
        workbook, all_quotation_results, matrix_data=None, image_path=None,
        ko_discount_rate=100, ko_steel_discount_rate=84,
        ko_purchased_discount_rate=94, sale_type='export',
        contact_info=None, trade_method='EXW',
        ko_ddp_address='',
        pile_summary=None,
        dest_port='BUSAN', container_type='40HQ',
        container_qty=1, ko_cif_freight=0,
        skip_freight=True, quote_validity='7d', lang='en', discount_method='project',
        payment_term='3070shipment', **kwargs):

    matrix_data = matrix_data or {}
    _validity_map = {'1d': _t('validity_1d', lang), '7d': _t('validity_1w', lang), 'today': _t('validity_today', lang)}
    validity_text = _validity_map.get(quote_validity, _t('validity_1w', lang))
    _payment_map = {
        '100advance': _t('payment_100advance', lang),
        '3070shipment': _t('payment_3070shipment', lang),
        '3070bl': _t('payment_3070bl', lang),
    }
    payment_text = _payment_map.get(payment_term, _t('payment_3070shipment', lang))
    if sale_type == 'domestic':
        currency_label = 'RMB'
    elif sale_type == 'euro':
        currency_label = 'EUR'
    else:
        currency_label = 'USD'
    project_name = str(matrix_data.get('project_name') or '').strip()
    output_wp = matrix_data.get('output_wp') or 0
    module_wattage = matrix_data.get('module_wattage') or 0
    module_size = matrix_data.get('module_size') or ''
    angle_val = matrix_data.get('angle') or 0
    max_wind = matrix_data.get('max_wind_speed') or ''
    max_snow = matrix_data.get('max_snow_load') or ''
    arrays = matrix_data.get('arrays') or []

    ws = workbook.create_sheet(title=_t('common_summary_title', lang))

    title_font = Font(name='Arial', size=22, bold=True)
    section_font = Font(name='Arial', size=12, bold=True)
    label_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=11)
    bold_font = Font(name='Arial', size=11, bold=True)
    red_bold_font = Font(name='Arial', size=13, bold=True, color='FF0000')
    black_bold_font = Font(name='Arial', size=13, bold=True)
    company_font = Font(name='Arial', size=14, bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_a = Alignment(horizontal='right', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_letter in 'ABCDEFGHIJKLMNO':
        ws.column_dimensions[col_letter].width = 13

    ws.column_dimensions['A'].width = 8.625
    ws.column_dimensions['B'].width = 9.0
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['J'].width = 8.625

    row = 1
    ws.merge_cells(f'A{row}:O{row}')
    ws.cell(row=row, column=1, value=_t('common_summary_title', lang)).font = title_font
    ws.cell(row=row, column=1).alignment = center
    ws.row_dimensions[row].height = 36

    row = 2
    ws.merge_cells(f'K{row}:L{row}')
    ws.cell(row=row, column=11, value=_t('common_quote_no', lang)).font = label_font
    ws.cell(row=row, column=11).alignment = right_a
    _qno = matrix_data.get('quotation_no', '') if matrix_data else ''
    if _qno:
        ws.cell(row=row, column=13, value=_qno).font = normal_font

    row = 3
    ws.merge_cells(f'K{row}:L{row}')
    ws.cell(row=row, column=11, value=_t('common_date', lang)).font = label_font
    ws.cell(row=row, column=11).alignment = right_a
    ws.cell(row=row, column=13, value=datetime.now().strftime('%Y-%m-%d')).font = normal_font

    row = 4
    ws.merge_cells(f'K{row}:L{row}')
    ws.cell(row=row, column=11, value=_t('common_version', lang)).font = label_font
    ws.cell(row=row, column=11).alignment = right_a
    ws.cell(row=row, column=13, value='V1.0').font = normal_font

    for r in range(2, 5):
        ws.row_dimensions[r].height = 15

    row = 5
    ws.cell(row=row, column=1, value=_t('common_seller', lang)).font = company_font
    ws.cell(row=row, column=10, value=_t('common_buyer', lang)).font = company_font
    ws.row_dimensions[row].height = 23

    row = 6
    ws.cell(row=row, column=1, value='Xiamen Kseng Metal Tech. Co., Ltd').font = company_font
    ws.row_dimensions[row].height = 23

    contact_defaults = {
        'contact_name': 'Samantha Ruan',
        'phone': '+86-18050060639',
        'tel_num': '',
        'tel': 'samantha@xmkseng.com',
    }
    contact_info = contact_info or {}
    contact_name = contact_info.get('contact_name') or contact_defaults['contact_name']
    phone = contact_info.get('phone') or contact_defaults['phone']
    tel_num = contact_info.get('tel_num') or contact_defaults['tel_num']
    email = contact_info.get('tel') or contact_defaults['tel']

    row = 7
    ws.cell(row=row, column=1, value=_t('common_from', lang)).font = label_font
    ws.cell(row=row, column=2, value=contact_name).font = normal_font
    ws.cell(row=row, column=10, value=_t('common_attn', lang)).font = label_font

    row = 8
    ws.cell(row=row, column=1, value=_t('common_tel', lang)).font = label_font
    ws.cell(row=row, column=2, value=tel_num or phone).font = normal_font
    ws.cell(row=row, column=10, value=_t('common_tel', lang)).font = label_font

    row = 9
    ws.cell(row=row, column=1, value=_t('common_email', lang)).font = label_font
    ws.cell(row=row, column=2, value=email).font = normal_font
    ws.cell(row=row, column=10, value=_t('common_email', lang)).font = label_font

    row = 10
    ws.cell(row=row, column=1, value=_t('common_phone', lang)).font = label_font
    ws.cell(row=row, column=2, value=phone).font = normal_font
    ws.cell(row=row, column=10, value=_t('common_phone', lang)).font = label_font

    row = 11
    ws.cell(row=row, column=1, value=_t('common_web', lang)).font = label_font
    ws.cell(row=row, column=2, value='www.xmkseng.com').font = normal_font
    ws.cell(row=row, column=10, value=_t('common_web', lang)).font = label_font

    row = 12
    ws.cell(row=row, column=1, value=_t('common_add', lang)).font = label_font
    ws.cell(row=row, column=2, value='6F, No.891, Fanghu North 2nd Rd, Huli Dist, Xiamen City, China').font = normal_font
    ws.cell(row=row, column=10, value=_t('common_add', lang)).font = label_font

    row = 13
    ws.row_dimensions[row].height = 8

    row = 14
    ws.merge_cells(f'A{row}:C{row}')
    _sc(ws, row, 1, _t('common_trade_term', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'D{row}:G{row}')
    _sc(ws, row, 4, trade_method, font=normal_font, align=center, border=thin_border)
    ws.merge_cells(f'H{row}:J{row}')
    _sc(ws, row, 8, _t('common_lead_time', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'K{row}:O{row}')
    _sc(ws, row, 11, _t('common_lead_time_value', lang), font=normal_font, align=center, border=thin_border)
    ws.row_dimensions[row].height = 25

    row = 15
    ws.merge_cells(f'A{row}:C{row}')
    _sc(ws, row, 1, _t('common_payment_term', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'D{row}:G{row}')
    _sc(ws, row, 4, payment_text, font=normal_font, align=center, border=thin_border)
    ws.merge_cells(f'H{row}:J{row}')
    _sc(ws, row, 8, _t('common_quote_validity', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'K{row}:O{row}')
    _sc(ws, row, 11, validity_text, font=normal_font, align=center, border=thin_border)
    ws.row_dimensions[row].height = 25

    row = 16
    ws.row_dimensions[row].height = 8

    row = 17
    ws.merge_cells(f'A{row}:C{row}')
    _sc(ws, row, 1, _t('common_project_name', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'D{row}:O{row}')
    _sc(ws, row, 4, project_name, font=normal_font, align=center, border=thin_border)
    ws.row_dimensions[row].height = 25

    row = 18
    ws.merge_cells(f'A{row}:C{row}')
    _sc(ws, row, 1, _t('common_snow_load', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'D{row}:G{row}')
    _sc(ws, row, 4, f'{max_snow} kN/m²' if max_snow else '', font=normal_font, align=center, border=thin_border)
    ws.merge_cells(f'H{row}:J{row}')
    _sc(ws, row, 8, _t('common_module_dimension', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'K{row}:O{row}')
    _sc(ws, row, 11, module_size, font=normal_font, align=center, border=thin_border)
    ws.row_dimensions[row].height = 25

    row = 19
    ws.merge_cells(f'A{row}:C{row}')
    _sc(ws, row, 1, _t('common_wind_load', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'D{row}:G{row}')
    _sc(ws, row, 4, f'{max_wind} m/s' if max_wind else '', font=normal_font, align=center, border=thin_border)
    ws.merge_cells(f'H{row}:J{row}')
    _sc(ws, row, 8, _t('common_module_capacity', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'K{row}:O{row}')
    _sc(ws, row, 11, module_wattage, font=normal_font, align=center, border=thin_border)
    ws.row_dimensions[row].height = 25

    row = 20
    ws.merge_cells(f'A{row}:C{row}')
    _sc(ws, row, 1, _t('common_ground_clearance', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'D{row}:G{row}')
    _sc(ws, row, 4, '', font=normal_font, align=center, border=thin_border)
    ws.merge_cells(f'H{row}:J{row}')
    _sc(ws, row, 8, _t('common_module_quantity', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'K{row}:O{row}')
    total_module_qty = sum(
        (a.get('rows', 0) * a.get('cols', 0) - abs(_safe_int(
            (matrix_data.get('arrays') or [{}])[0].get('missing_per_table', 0) if matrix_data.get('arrays') else 0
        ))) * a.get('table_qty', 1)
        for a in arrays
    ) if arrays else 0
    _sc(ws, row, 11, total_module_qty if total_module_qty else '', font=normal_font, align=center, border=thin_border)
    ws.row_dimensions[row].height = 25

    if not output_wp and module_wattage and total_module_qty:
        output_wp = module_wattage * total_module_qty

    row = 21
    ws.merge_cells(f'A{row}:C{row}')
    _sc(ws, row, 1, _t('common_installation_angle', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'D{row}:G{row}')
    _sc(ws, row, 4, angle_val, font=normal_font, align=center, border=thin_border)
    ws.merge_cells(f'H{row}:J{row}')
    _sc(ws, row, 8, _t('common_total_capacity', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'K{row}:O{row}')
    _sc(ws, row, 11, output_wp, font=normal_font, align=center, border=thin_border)
    ws.row_dimensions[row].height = 25

    row = 22
    ws.merge_cells(f'A{row}:C{row}')
    _sc(ws, row, 1, _t('common_main_material', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'D{row}:G{row}')
    _sc(ws, row, 4, 'AL6005-T5', font=normal_font, align=center, border=thin_border)
    ws.merge_cells(f'H{row}:J{row}')
    _sc(ws, row, 8, _t('common_layout', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'K{row}:O{row}')
    layout_parts = []
    for a in arrays:
        layout_parts.append(f"{a.get('rows', '')} {_t('layout_row', lang)} × {a.get('cols', '')} {_t('layout_column', lang)} {_t('layout_tables', lang)} {a.get('table_qty', 1)} {_t('layout_tables', lang)}")
    layout_cell = _sc(ws, row, 11, '\n'.join(layout_parts) if layout_parts else '', font=normal_font, align=center, border=thin_border)
    layout_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    base_layout_height = 25
    if layout_parts and len(layout_parts) > 1:
        ws.row_dimensions[row].height = base_layout_height * (1 + (len(layout_parts) - 1) * 0.7)
    else:
        ws.row_dimensions[row].height = base_layout_height

    row = 23
    ws.row_dimensions[row].height = 8

    row = 24
    ws.merge_cells(f'A{row}:O{row}')
    _sc(ws, row, 1, _t('common_part1_title', lang), font=section_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.row_dimensions[row].height = 27

    row = 25
    _sc(ws, row, 1, _t('common_part1_hdr_no', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'B{row}:D{row}')
    _sc(ws, row, 2, _t('common_part1_hdr_item', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    _sc(ws, row, 5, _t('common_part1_hdr_missing_panel', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'F{row}:H{row}')
    _sc(ws, row, 6, _t('common_part1_hdr_picture', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'I{row}:J{row}')
    _sc(ws, row, 9, _t('common_part1_hdr_total_capacity', lang), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'K{row}:L{row}')
    _sc(ws, row, 11, _t('common_part1_hdr_exw_unit_price', lang).format(currency=currency_label), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.merge_cells(f'M{row}:O{row}')
    _sc(ws, row, 13, _t('common_part1_hdr_exw_amount', lang).format(currency=currency_label), font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
    ws.row_dimensions[row].height = 35

    data_row = 26
    total_base_count = 0
    for i, qr in enumerate(all_quotation_results):
        r = data_row + i
        cfg = qr.get('config') or {}
        md = qr.get('matrix_data') or matrix_data
        ma = qr.get('matched_array') or {}
        r_rows = ma.get('rows', md.get('array_rows'))
        r_cols = ma.get('cols', md.get('array_cols'))
        r_set = _safe_int(md.get('set_count', 1), 1)
        r_angle = cfg.get('angle', '') or md.get('angle', '')
        detail_sheet = qr.get('sheet_name', '')
        r_kw = md.get('output_kw') or 0
        r_wp = _safe_int(_safe_float(r_kw) * 1000) if r_kw else 0
        total_base_count += r_set

        missing = cfg.get('missing_boards', 0)
        if not missing:
            missing = _safe_int(ma.get('missing_per_table', 0))

        ws.cell(row=r, column=1, value=i + 1).font = normal_font
        ws.cell(row=r, column=1).alignment = center
        ws.cell(row=r, column=1).border = thin_border

        ws.merge_cells(f'B{r}:D{r}')
        desc = f'{r_rows} {_t("layout_row", lang)} × {r_cols} {_t("layout_column", lang)} {_t("layout_tables", lang)} {r_set} {_t("layout_tables", lang)}' if r_rows and r_cols else detail_sheet
        ws.cell(row=r, column=2, value=desc).font = normal_font
        ws.cell(row=r, column=2).alignment = center
        ws.cell(row=r, column=2).border = thin_border
        for ci in (3, 4):
            ws.cell(row=r, column=ci).border = thin_border

        _sc(ws, r, 5, missing if missing else '/', font=normal_font, align=center, border=thin_border)

        ws.merge_cells(f'F{r}:H{r}')
        for ci in (6, 7, 8):
            ws.cell(row=r, column=ci).border = thin_border

        ws.merge_cells(f'I{r}:J{r}')
        ws.cell(row=r, column=9, value=r_wp).font = normal_font
        ws.cell(row=r, column=9).alignment = center
        ws.cell(row=r, column=9).border = thin_border
        ws.cell(row=r, column=10).border = thin_border

        p1_per_table = qr.get('part1_price_per_table', 0)
        ws.merge_cells(f'K{r}:L{r}')
        if r_set > 0 and p1_per_table:
            ws.cell(row=r, column=11, value=f'=M{r}/{r_set}').font = normal_font
            ws.cell(row=r, column=11).number_format = '#,##0.000'
        else:
            ws.cell(row=r, column=11, value='').font = normal_font
        ws.cell(row=r, column=11).alignment = center
        ws.cell(row=r, column=11).border = thin_border
        ws.cell(row=r, column=12).border = thin_border

        ws.merge_cells(f'M{r}:O{r}')
        total_table_row = qr.get('part1_total_table_row', 0)
        if detail_sheet and total_table_row:
            ws.cell(row=r, column=13, value=f"='{detail_sheet}'!H{total_table_row}").number_format = CURRENCY_FMT
        else:
            ws.cell(row=r, column=13, value=p1_per_table).number_format = CURRENCY_FMT
        ws.cell(row=r, column=13).font = normal_font
        ws.cell(row=r, column=13).alignment = center
        ws.cell(row=r, column=13).border = thin_border
        for ci in (14, 15):
            ws.cell(row=r, column=ci).border = thin_border

        ws.row_dimensions[r].height = 82

    data_end = data_row + len(all_quotation_results) - 1

    if discount_method == 'unit_price':
        total_exw_row = data_end + 1

        discount_row = total_exw_row
        ws.merge_cells(f'A{discount_row}:L{discount_row}')
        _sc(ws, discount_row, 1, _t('common_discount_exw_price', lang).format(currency=currency_label), font=black_bold_font, align=right_a, border=thin_border, fill=BLUE_FILL)
        for ci in range(2, 13):
            ws.cell(row=discount_row, column=ci).border = thin_border
        ws.merge_cells(f'M{discount_row}:O{discount_row}')
        if all_quotation_results:
            refs = '+'.join(f'M{data_row + i}' for i in range(len(all_quotation_results)))
            _sc(ws, discount_row, 13, f'={refs}', font=black_bold_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
        else:
            _sc(ws, discount_row, 13, 0, font=black_bold_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
        ws.row_dimensions[discount_row].height = 30

        unit_price_row = discount_row + 1
        ws.merge_cells(f'A{unit_price_row}:L{unit_price_row}')
        _sc(ws, unit_price_row, 1, _t('common_discount_exw_unit_price', lang).format(currency=currency_label), font=black_bold_font, align=right_a, border=thin_border, fill=BLUE_FILL)
        for ci in range(2, 13):
            ws.cell(row=unit_price_row, column=ci).border = thin_border
        ws.merge_cells(f'M{unit_price_row}:O{unit_price_row}')
        if total_base_count > 0:
            _sc(ws, unit_price_row, 13, f'=M{discount_row}/{total_base_count}', font=black_bold_font, align=center, border=thin_border, number_format='#,##0.000')
        else:
            _sc(ws, unit_price_row, 13, 0, font=black_bold_font, align=center, border=thin_border, number_format='#,##0.000')
        ws.row_dimensions[unit_price_row].height = 30
    else:
        total_exw_row = data_end + 1
        ws.merge_cells(f'A{total_exw_row}:L{total_exw_row}')
        _sc(ws, total_exw_row, 1, _t('common_total_exw_amount', lang).format(currency=currency_label), font=black_bold_font, align=right_a, border=thin_border, fill=BLUE_FILL)
        for ci in range(2, 13):
            ws.cell(row=total_exw_row, column=ci).border = thin_border
        ws.merge_cells(f'M{total_exw_row}:O{total_exw_row}')
        if all_quotation_results:
            refs = '+'.join(f'M{data_row + i}' for i in range(len(all_quotation_results)))
            _sc(ws, total_exw_row, 13, f'={refs}', font=black_bold_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
        else:
            _sc(ws, total_exw_row, 13, 0, font=black_bold_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
        ws.row_dimensions[total_exw_row].height = 30

        discount_row = total_exw_row + 1
        ws.merge_cells(f'A{discount_row}:L{discount_row}')
        _sc(ws, discount_row, 1, _t('common_discount_exw_price', lang).format(currency=currency_label), font=black_bold_font, align=right_a, border=thin_border, fill=BLUE_FILL)
        for ci in range(2, 13):
            ws.cell(row=discount_row, column=ci).border = thin_border
        ws.merge_cells(f'M{discount_row}:O{discount_row}')

        discount_parts = []
        for qr in all_quotation_results:
            sn = qr.get('sheet_name', '')
            sn_esc = sn.replace("'", "''")
            for rows_key, rate in [
                ('part1_standard_rows', ko_discount_rate),
                ('part1_steel_rows', ko_steel_discount_rate),
                ('part1_purchased_rows', ko_purchased_discount_rate),
                ('part1_inquiry_rows', 100),
            ]:
                rows_list = qr.get(rows_key, [])
                if not rows_list:
                    continue
                refs = ','.join(f"'{sn_esc}'!I{r}" for r in rows_list)
                if rate == 100:
                    discount_parts.append(f"SUM({refs})")
                else:
                    discount_parts.append(f"SUM({refs})*{rate}/100")

        if discount_parts:
            discount_formula = '=' + '+'.join(discount_parts)
            if len(discount_formula) > 8000:
                discount_formula = sum(
                    float(qr.get('part1_standard_total', 0)) * ko_discount_rate / 100 +
                    float(qr.get('part1_steel_total', 0)) * ko_steel_discount_rate / 100 +
                    float(qr.get('part1_purchased_total', 0)) * ko_purchased_discount_rate / 100 +
                    float(qr.get('part1_inquiry_total', 0))
                    for qr in all_quotation_results
                )
        else:
            discount_formula = 0

        _sc(ws, discount_row, 13, discount_formula, font=black_bold_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
        ws.row_dimensions[discount_row].height = 30

        unit_price_row = discount_row + 1
        ws.merge_cells(f'A{unit_price_row}:L{unit_price_row}')
        _sc(ws, unit_price_row, 1, _t('common_discount_exw_unit_price', lang).format(currency=currency_label), font=black_bold_font, align=right_a, border=thin_border, fill=BLUE_FILL)
        for ci in range(2, 13):
            ws.cell(row=unit_price_row, column=ci).border = thin_border
        ws.merge_cells(f'M{unit_price_row}:O{unit_price_row}')
        if total_base_count > 0:
            _sc(ws, unit_price_row, 13, f'=M{discount_row}/{total_base_count}', font=black_bold_font, align=center, border=thin_border, number_format='#,##0.000')
        else:
            _sc(ws, unit_price_row, 13, 0, font=black_bold_font, align=center, border=thin_border, number_format='#,##0.000')
        ws.row_dimensions[unit_price_row].height = 30

    if not skip_freight and trade_method != 'EXW':
        freight_title_row = unit_price_row + 2
        ws.merge_cells(f'A{freight_title_row}:O{freight_title_row}')
        _part2_title = f'Part 2: Quotation - {trade_method} Freight Cost'
        _sc(ws, freight_title_row, 1, _part2_title, font=section_font, align=center, border=thin_border, fill=BLUE_FILL)
        ws.row_dimensions[freight_title_row].height = 27

        container_details = kwargs.get('container_details') or []

        freight_hdr_row = freight_title_row + 1
        _hdr_cols = [
            (1, 2, _t('common_freight_hdr_no', lang)),
            (3, 4, _t('common_freight_hdr_port', lang)),
            (5, 6, _t('common_freight_hdr_container', lang)),
            (7, 8, _t('common_freight_hdr_quantity', lang)),
            (9, 12, _t('common_freight_hdr_unit_price', lang)),
            (13, 15, _t('common_freight_hdr_amount', lang).format(currency=currency_label)),
        ]
        for c_start, c_end, hdr_text in _hdr_cols:
            if c_start != c_end:
                ws.merge_cells(f'{get_column_letter(c_start)}{freight_hdr_row}:{get_column_letter(c_end)}{freight_hdr_row}')
            _sc(ws, freight_hdr_row, c_start, hdr_text, font=normal_font, align=center, border=thin_border, fill=BLUE_FILL)
            for cc in range(c_start + 1, c_end + 1):
                ws.cell(row=freight_hdr_row, column=cc).border = thin_border
        ws.row_dimensions[freight_hdr_row].height = 25

        freight_data_start = freight_hdr_row + 1
        freight_amount_cells = []

        if container_details:
            dr = freight_data_start
            seq = 1
            for cd in container_details:
                ct = str(cd.get('type', '')).upper()
                cq = _safe_float(cd.get('qty', 0))
                cpu = cd.get('freight_per_unit', 0)

                ws.merge_cells(f'A{dr}:B{dr}')
                _sc(ws, dr, 1, seq, font=normal_font, align=center, border=thin_border)
                ws.cell(row=dr, column=2).border = thin_border

                ws.merge_cells(f'C{dr}:D{dr}')
                _sc(ws, dr, 3, dest_port, font=normal_font, align=center, border=thin_border)
                ws.cell(row=dr, column=4).border = thin_border

                ws.merge_cells(f'E{dr}:F{dr}')
                _sc(ws, dr, 5, ct, font=normal_font, align=center, border=thin_border)
                ws.cell(row=dr, column=6).border = thin_border

                ws.merge_cells(f'G{dr}:H{dr}')
                _sc(ws, dr, 7, _safe_int(cq) if isinstance(cq, float) and cq % 1 == 0 else cq, font=normal_font, align=center, border=thin_border)
                ws.cell(row=dr, column=8).border = thin_border

                ws.merge_cells(f'I{dr}:L{dr}')
                _sc(ws, dr, 9, float(cpu) if cpu else 0, font=normal_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
                for cc in (10, 11, 12):
                    ws.cell(row=dr, column=cc).border = thin_border

                ws.merge_cells(f'M{dr}:O{dr}')
                _sc(ws, dr, 13, f'=I{dr}*G{dr}', font=normal_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
                for cc in (14, 15):
                    ws.cell(row=dr, column=cc).border = thin_border

                freight_amount_cells.append(f'M{dr}')
                ws.row_dimensions[dr].height = 25
                seq += 1
                dr += 1
            freight_data_end = dr - 1
        else:
            dr = freight_data_start
            ws.merge_cells(f'A{dr}:H{dr}')
            _sc(ws, dr, 1, _t('common_freight_desc', lang).format(port=dest_port), font=normal_font, align=center, border=thin_border)
            for cc in range(2, 9):
                ws.cell(row=dr, column=cc).border = thin_border

            ws.merge_cells(f'I{dr}:L{dr}')
            _sc(ws, dr, 9, ko_cif_freight, font=normal_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
            for cc in (10, 11, 12):
                ws.cell(row=dr, column=cc).border = thin_border

            ws.merge_cells(f'M{dr}:O{dr}')
            _sc(ws, dr, 13, f'=I{dr}*1', font=normal_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
            for cc in (14, 15):
                ws.cell(row=dr, column=cc).border = thin_border

            freight_amount_cells.append(f'M{dr}')
            ws.row_dimensions[dr].height = 25
            freight_data_end = dr

        subtotal_row = freight_data_end + 1
        ws.merge_cells(f'A{subtotal_row}:L{subtotal_row}')
        _sc(ws, subtotal_row, 1, f'{trade_method} {_t("common_freight_subtotal", lang)}', font=black_bold_font, align=Alignment(horizontal='right', vertical='center'), border=thin_border, fill=BLUE_FILL)
        for cc in range(2, 13):
            ws.cell(row=subtotal_row, column=cc).border = thin_border
        ws.merge_cells(f'M{subtotal_row}:O{subtotal_row}')
        if freight_amount_cells:
            _sc(ws, subtotal_row, 13, f'={"+".join(freight_amount_cells)}', font=black_bold_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
        else:
            _sc(ws, subtotal_row, 13, 0, font=black_bold_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
        for cc in (14, 15):
            ws.cell(row=subtotal_row, column=cc).border = thin_border
        ws.row_dimensions[subtotal_row].height = 27

        total_port_row = subtotal_row + 1
        ws.merge_cells(f'A{total_port_row}:L{total_port_row}')
        _sc(ws, total_port_row, 1, f'Total Amount of {trade_method} {dest_port} Port ({currency_label}):', font=black_bold_font, align=Alignment(horizontal='right', vertical='center'), border=thin_border, fill=BLUE_FILL)
        for cc in range(2, 13):
            ws.cell(row=total_port_row, column=cc).border = thin_border
        ws.merge_cells(f'M{total_port_row}:O{total_port_row}')
        _sc(ws, total_port_row, 13, f'=M{discount_row}+M{subtotal_row}', font=black_bold_font, align=center, border=thin_border, number_format=CURRENCY_FMT)
        for cc in (14, 15):
            ws.cell(row=total_port_row, column=cc).border = thin_border
        ws.row_dimensions[total_port_row].height = 30

        fob_unit_row = total_port_row + 1
        ws.merge_cells(f'A{fob_unit_row}:L{fob_unit_row}')
        _sc(ws, fob_unit_row, 1, f'{trade_method} {dest_port} Unit Price ({currency_label}):', font=black_bold_font, align=Alignment(horizontal='right', vertical='center'), border=thin_border, fill=BLUE_FILL)
        for cc in range(2, 13):
            ws.cell(row=fob_unit_row, column=cc).border = thin_border
        ws.merge_cells(f'M{fob_unit_row}:O{fob_unit_row}')
        if total_base_count > 0:
            _sc(ws, fob_unit_row, 13, f'=M{total_port_row}/{total_base_count}', font=black_bold_font, align=center, border=thin_border, number_format='#,##0.000')
        else:
            _sc(ws, fob_unit_row, 13, 0, font=black_bold_font, align=center, border=thin_border, number_format='#,##0.000')
        for cc in (14, 15):
            ws.cell(row=fob_unit_row, column=cc).border = thin_border
        ws.row_dimensions[fob_unit_row].height = 30

    if image_path and os.path.exists(image_path):
        try:
            img = XLImage(img=image_path)
            img.width = 280
            img.height = 70
            ws.add_image(img, 'A1')
        except Exception:
            pass

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(top=0.5, bottom=0.5, left=0.25, right=0.25, header=0.3, footer=0.3)

    sheet_names = workbook.sheetnames
    localized_title = _t('common_summary_title', lang)
    if localized_title in sheet_names:
        idx = sheet_names.index(localized_title)
        workbook.move_sheet(localized_title, offset=-idx)

    print(f"   [EN-COMMON] Summary sheet(Quotation) created")
    return ws.title
