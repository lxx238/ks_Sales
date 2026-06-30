"""验证无折扣汇总页改动（Calibri/H-L公司信息/红色金额/末行提示）。"""
import os
from openpyxl import Workbook
from backend.core.ap_ground_no_discount.quotation_engine import (
    create_ap_ground_no_disc_detail_sheet, create_ap_ground_no_disc_summary_sheet,
)

wb = Workbook()
wb.remove(wb.active)
site = {'rows': 4, 'cols': 12, 'tables': 2, 'kw': 6.6, 'span_info': '2400'}
items = [{'code': 'KR-001', 'name': 'Rail', 'material': 'AL6005-T5',
          'spec': 'C40*80*2400mm', 'spec_norm': '40*80*2400', 'unit_price': 12.5,
          'qty_per_table': 10, 'is_matched': True}]
detail = create_ap_ground_no_disc_detail_sheet(
    wb, site, items, matrix_data={'angle': '30', 'module_size': '2278*1134', 'module_wattage': 550},
    config={}, contact_info={'contact_name': 'Tom', 'phone': '138', 'tel': '0592-1'})

create_ap_ground_no_disc_summary_sheet(
    wb, [detail], matrix_data={'project_name': 'TEST', 'panel_orientation': 'Portrait'},
    contact_info={'contact_name': 'Tom'}, trade_method='CIF', dest_port='JEDDAH',
    container_details=[{'type': '40GP', 'qty': 3, 'freight_per_unit': 1500}])
ws = wb['Total']

# 公司信息 H-L 合并?
print('H7:L7 merged?', 'H7:L7' in [str(r) for r in ws.merged_cells.ranges])
# 字体 Calibri?
print('A3 font:', ws['A3'].font.name, '| A15 hdr font:', ws['A15'].font.name)
# 红色金额 (Total/per/w/EXW)
for r, label in [(22, 'Total'), (23, 'per/w'), (24, 'TOTAL AMOUNT(EXW)')]:
    c = ws.cell(row=r, column=10)
    print(f'R{r}({label}) J font color:', c.font.color.rgb if c.font.color else None, 'val:', c.value)
# EXW (无运费时 frame_bottom=24, note=26); 有运费时 note 更靠后。这里 CIF+柜型 -> Part2
# 找提示行
for r in range(25, 35):
    v = ws.cell(row=r, column=1).value
    if v and 'Please kindly' in str(v):
        print(f'Note row {r} font:', ws.cell(row=r, column=1).font.name,
              'size:', ws.cell(row=r, column=1).font.size,
              'color:', ws.cell(row=r, column=1).font.color.rgb if ws.cell(row=r, column=1).font.color else None)
        break
wb.save(os.path.join('output', '_nd_v2.xlsx'))
print('OK saved')
